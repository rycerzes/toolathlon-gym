"""
Evaluation script for yf-dividend-tracker-gcal task.

Checks:
1. Excel file (Dividend_Tracker.xlsx) - Dividend Stocks and Summary sheets
2. Google Calendar events for ex-dividend dates (>=3 events with dividend in summary)
3. Google Sheet with "dividend" in title
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_expected_dividends():
    """Get expected dividend data from yf.stock_info."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT symbol,
               data->>'shortName' as name,
               (data->>'dividendRate')::float as div_rate,
               (data->>'dividendYield')::float as div_yield,
               data->>'exDividendDate' as ex_div_date,
               (data->>'payoutRatio')::float as payout_ratio
        FROM yf.stock_info
        WHERE symbol IN ('GOOGL','AMZN','JPM','JNJ','XOM')
          AND (data->>'dividendRate') IS NOT NULL
          AND (data->>'dividendRate')::float > 0
        ORDER BY symbol
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace, expected_rows):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Dividend_Tracker.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return False

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # Check Dividend Stocks sheet
    ws = get_sheet(wb, "Dividend Stocks")
    check("Sheet 'Dividend Stocks' exists", ws is not None, f"Sheets: {wb.sheetnames}")
    if ws is None:
        return False

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    num_expected = len(expected_rows)
    check(f"Dividend Stocks has >= {num_expected} rows",
          len(rows) >= num_expected,
          f"Got {len(rows)} rows, expected >= {num_expected}")

    # Check that key tickers are present
    tickers_in_excel = set()
    for r in rows:
        if r and r[0]:
            tickers_in_excel.add(str(r[0]).strip().upper())

    expected_tickers = {r[0] for r in expected_rows}
    for t in expected_tickers:
        check(f"Ticker {t} in Dividend Stocks", t in tickers_in_excel,
              f"Found tickers: {tickers_in_excel}")

    # Check Summary sheet
    ws2 = get_sheet(wb, "Summary")
    check("Sheet 'Summary' exists", ws2 is not None, f"Sheets: {wb.sheetnames}")
    if ws2:
        summary_data = {}
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                key = str(row[0]).strip().lower().replace(" ", "_")
                summary_data[key] = row[1]

        check("Summary has Total_Dividend_Stocks",
              any("total" in k and "dividend" in k for k in summary_data),
              f"Keys: {list(summary_data.keys())}")
        check("Summary has Avg_Yield",
              any("avg" in k and "yield" in k for k in summary_data),
              f"Keys: {list(summary_data.keys())}")
        check("Summary has Highest_Yield_Ticker",
              any("highest" in k and "yield" in k for k in summary_data),
              f"Keys: {list(summary_data.keys())}")

    return True


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description FROM gcal.events")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_calendar] Found {len(events)} events.")

    dividend_events = [
        e for e in events
        if e[0] and ("ex-dividend" in e[0].lower() or "dividend" in e[0].lower())
    ]
    check("At least 3 calendar events with dividend in summary",
          len(dividend_events) >= 3,
          f"Found {len(dividend_events)} dividend events", db=True)

    return len(dividend_events) >= 3


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gsheet] Found {len(sheets)} spreadsheets.")

    dividend_sheets = [
        s for s in sheets
        if s[1] and "dividend" in s[1].lower()
    ]
    check("Google Sheet with 'dividend' in title exists",
          len(dividend_sheets) > 0,
          f"Sheet titles: {[s[1] for s in sheets]}", db=True)

    return len(dividend_sheets) > 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_rows = get_expected_dividends()
    print(f"Expected {len(expected_rows)} dividend-paying stocks from DB")

    check_excel(args.agent_workspace, expected_rows)
    check_calendar()
    check_gsheet()

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if args.res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": file_ok}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if file_ok else 1)


if __name__ == "__main__":
    main()

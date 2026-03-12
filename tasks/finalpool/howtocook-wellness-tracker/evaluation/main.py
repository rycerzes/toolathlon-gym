"""
Evaluation for howtocook-wellness-tracker task.
Checks Excel, Word, and Notion.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Wellness_Tracker.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Recipe Wellness Score
    rw_rows = load_sheet_rows(wb, "Recipe Wellness Score") or load_sheet_rows(wb, "Recipe_Wellness_Score")
    if rw_rows is None:
        record("Sheet 'Recipe Wellness Score' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Recipe Wellness Score' exists", True)
        data = [r for r in rw_rows[1:] if r and r[0] is not None and str(r[0]).strip()]
        record("Has >= 8 recipes", len(data) >= 8, f"Found {len(data)}")

        ws_col = find_col(rw_rows[0], ["Wellness_Score", "Wellness Score", "Score"])
        if ws_col is not None:
            for r in data[:3]:
                if ws_col < len(r) and r[ws_col] is not None:
                    try:
                        s = float(r[ws_col])
                        record(f"Score {s} in range 1-10", 1 <= s <= 10, f"Got {s}")
                        break
                    except (TypeError, ValueError):
                        pass

        rec_col = find_col(rw_rows[0], ["Recommended", "recommended"])
        if rec_col is not None:
            vals = set()
            for r in data:
                if rec_col < len(r) and r[rec_col] is not None:
                    vals.add(str(r[rec_col]).strip().lower())
            has_both = "yes" in vals and "no" in vals
            record("Recommended has both Yes and No", has_both, f"Values: {vals}")

        cat_col = find_col(rw_rows[0], ["Category", "category"])
        if cat_col is not None:
            cats = {str(r[cat_col]).strip().lower() for r in data if cat_col < len(r) and r[cat_col]}
            record("At least 3 categories", len(cats) >= 3, f"Found {len(cats)}: {cats}")

    # Weekly Plan
    wp_rows = load_sheet_rows(wb, "Weekly Plan") or load_sheet_rows(wb, "Weekly_Plan")
    if wp_rows is None:
        record("Sheet 'Weekly Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Weekly Plan' exists", True)
        data = [r for r in wp_rows[1:] if r and r[0] is not None]
        record("Weekly Plan has >= 10 rows (5 days x 2 meals)", len(data) >= 10, f"Found {len(data)}")

    # Progress Metrics
    pm_rows = load_sheet_rows(wb, "Progress Metrics") or load_sheet_rows(wb, "Progress_Metrics")
    if pm_rows is None:
        record("Sheet 'Progress Metrics' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Progress Metrics' exists", True)
        metrics = {}
        for row in pm_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        vt_key = next((k for k in metrics if "veggie" in k and "target" in k), None)
        if vt_key:
            try:
                record("Daily_Veggie_Target = 5", abs(float(metrics[vt_key]) - 5) < 1)
            except (TypeError, ValueError):
                pass

        sl_key = next((k for k in metrics if "sugar" in k and "limit" in k), None)
        if sl_key:
            try:
                record("Daily_Sugar_Limit = 25", abs(float(metrics[sl_key]) - 25) < 5)
            except (TypeError, ValueError):
                pass

    return True


def check_word(workspace):
    print("\n=== Checking Word ===")
    path = os.path.join(workspace, "Wellness_Guide.docx")
    if not os.path.isfile(path):
        record("Word exists", False, f"Not found: {path}")
        return False
    record("Word exists", True)

    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join(p.text for p in doc.paragraphs).lower()
        record("Has substantial content", len(text) > 300, f"Only {len(text)} chars")
        record("Mentions wellness/health", "wellness" in text or "health" in text)
        record("Mentions vegetable/sugar", "vegetable" in text or "sugar" in text)
        record("Mentions meal plan", "meal" in text or "plan" in text)
        return True
    except Exception as e:
        record("Word readable", False, str(e))
        return False


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE '%%wellness%%'
               OR properties::text ILIKE '%%dashboard%%'
               OR properties::text ILIKE '%%health%%'
        """)
        pages = cur.fetchall()

        if not pages:
            cur.execute("SELECT id, properties FROM notion.pages")
            all_p = cur.fetchall()
            record("Notion wellness page exists", False, f"Found {len(all_p)} pages but none matching")
            return False

        record("Notion wellness page exists", True)

        page_ids = [p[0] for p in pages]
        cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = ANY(%s)", (page_ids,))
        count = cur.fetchone()[0]
        if count == 0:
            cur.execute("SELECT COUNT(*) FROM notion.blocks")
            count = cur.fetchone()[0]

        record("Notion page has >= 5 blocks", count >= 5, f"Found {count}")

        cur.close()
        conn.close()
        return True
    except Exception as e:
        record("Notion accessible", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

"""Evaluation for sf-workforce-diversity-review."""
import argparse
import os
import sys

import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Diversity_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Diversity_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check Education Distribution
    print("  Checking Education Distribution...")
    a_rows = load_sheet_rows(agent_wb, "Education Distribution")
    g_rows = load_sheet_rows(gt_wb, "Education Distribution")
    if a_rows is None:
        all_errors.append("Sheet 'Education Distribution' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Education Distribution' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for r in a_data:
            if r and r[0] is not None and r[1] is not None:
                k = f"{str(r[0]).strip().lower()}|{str(r[1]).strip().lower()}"
                a_lookup[k] = r
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}|{g_row[1]}")
                continue
            # Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 50):
                    errors.append(f"{key}.Count: {a_row[2]} vs {g_row[2]}")
            # Internal_Pct
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1.0):
                    errors.append(f"{key}.Internal_Pct: {a_row[3]} vs {g_row[3]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Age Distribution
    print("  Checking Age Distribution...")
    a_rows = load_sheet_rows(agent_wb, "Age Distribution")
    g_rows = load_sheet_rows(gt_wb, "Age Distribution")
    if a_rows is None:
        all_errors.append("Sheet 'Age Distribution' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Age Distribution' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for r in a_data:
            if r and r[0] is not None and r[1] is not None:
                k = f"{str(r[0]).strip().lower()}|{str(r[1]).strip().lower()}"
                a_lookup[k] = r
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}|{g_row[1]}")
                continue
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 50):
                    errors.append(f"{key}.Count: {a_row[2]} vs {g_row[2]}")
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1.0):
                    errors.append(f"{key}.Internal_Pct: {a_row[3]} vs {g_row[3]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Overall Scorecard
    print("  Checking Overall Scorecard...")
    a_rows = load_sheet_rows(agent_wb, "Overall Scorecard")
    g_rows = load_sheet_rows(gt_wb, "Overall Scorecard")
    if a_rows is None:
        all_errors.append("Sheet 'Overall Scorecard' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Overall Scorecard' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for r in a_data:
            if r and r[0] is not None and r[1] is not None:
                k = f"{str(r[0]).strip().lower()}|{str(r[1]).strip().lower()}"
                a_lookup[k] = r
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}|{g_row[1]}")
                continue
            # Internal_Value
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.5):
                    errors.append(f"{key}.Internal: {a_row[2]} vs {g_row[2]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Summary
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if key in ("total_employees",):
                    if not num_close(a_row[1], g_row[1], 100):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif key in ("total_departments", "education_levels"):
                    if not num_close(a_row[1], g_row[1], 1):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                else:
                    if not num_close(a_row[1], g_row[1], 1.0):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check PowerPoint exists
    print("  Checking DEI_Board_Report.pptx...")
    pptx_file = os.path.join(args.agent_workspace, "DEI_Board_Report.pptx")
    if not os.path.exists(pptx_file):
        all_errors.append("DEI_Board_Report.pptx not found")
        print("    FAIL: file not found")
    else:
        try:
            from pptx import Presentation
            prs = Presentation(pptx_file)
            slide_count = len(prs.slides)
            if slide_count < 5:
                all_errors.append(f"PowerPoint has only {slide_count} slides, expected 5+")
                print(f"    FAIL: only {slide_count} slides")
            else:
                print("    PASS")
        except Exception as e:
            all_errors.append(f"PPTX read error: {e}")
            print(f"    ERROR: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

"""Evaluation for yf-market-news-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected values from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM yf.news")
    total_articles = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT symbol) FROM yf.news")
    unique_tickers = cur.fetchone()[0]
    cur.execute("""
        SELECT data->'content'->'provider'->>'displayName' as publisher, COUNT(*) as cnt
        FROM yf.news GROUP BY 1 ORDER BY 2 DESC LIMIT 1
    """)
    row = cur.fetchone()
    top_publisher = row[0] if row else ""
    cur.close()
    conn.close()
    return total_articles, unique_tickers, top_publisher


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Market_News_Digest.xlsx."""
    print("\n=== Checking Market_News_Digest.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Market_News_Digest.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    total_articles, unique_tickers, top_publisher = get_expected_data()

    # Check News Articles sheet
    news_sheet = None
    for name in wb.sheetnames:
        if "news" in name.lower() and "article" in name.lower():
            news_sheet = wb[name]
            break
    if news_sheet is None:
        record("Sheet 'News Articles' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'News Articles' exists", True)
        rows = list(news_sheet.iter_rows(min_row=2, values_only=True))
        ok_count = num_close(len(rows), total_articles, 5)
        record("News Articles row count matches DB", ok_count,
               f"Expected ~{total_articles}, got {len(rows)}")
        if not ok_count:
            all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        for key, val in summary.items():
            if "total" in key and "article" in key:
                ok = num_close(val, total_articles, 5)
                record("Summary Total_Articles", ok,
                       f"Expected {total_articles}, got {val}")
                if not ok:
                    all_ok = False
            elif "unique" in key and "ticker" in key:
                ok = num_close(val, unique_tickers, 1)
                record("Summary Unique_Tickers", ok,
                       f"Expected {unique_tickers}, got {val}")
                if not ok:
                    all_ok = False
            elif "top" in key and "publisher" in key:
                ok = top_publisher.lower() in str(val).lower()
                record("Summary Top_Publisher", ok,
                       f"Expected '{top_publisher}', got '{val}'")
                if not ok:
                    all_ok = False

    return all_ok


def check_notion():
    """Check Notion page with 'market' in title."""
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()

    found_page = False
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else json.loads(page[1]) if page[1] else {}
        title_text = ""
        if "title" in props:
            t = props["title"]
            if isinstance(t, dict) and "title" in t:
                for item in t["title"]:
                    if isinstance(item, dict):
                        title_text += item.get("plain_text", item.get("text", {}).get("content", ""))
        if "market" in title_text.lower():
            found_page = True
            break

    record("Notion page with 'market' in title found", found_page,
           "No page with 'market' in title")

    cur.close()
    conn.close()
    return found_page


def check_email():
    """Check email with 'news' or 'digest' in subject."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%news%' OR subject ILIKE '%digest%'
        LIMIT 5
    """)
    rows = cur.fetchall()
    found = len(rows) > 0
    record("Email with 'news' or 'digest' in subject", found,
           "No matching email found")

    if found:
        for subj, to_addr, body in rows:
            if "portfolio-team@investco.com" in str(to_addr).lower():
                record("Email to portfolio-team@investco.com", True)
                break
        else:
            record("Email to portfolio-team@investco.com", False,
                   f"To addresses: {[str(r[1]) for r in rows]}")

    cur.close()
    conn.close()
    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)

    db_fail_before = FAIL_COUNT
    notion_ok = check_notion()
    email_ok = check_email()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")

    overall = excel_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

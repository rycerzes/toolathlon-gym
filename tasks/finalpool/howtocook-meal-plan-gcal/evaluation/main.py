"""
Evaluation script for howtocook-meal-plan-gcal task.

Checks:
1. At least 21 gcal events with "cook" in summary, dates in 2026-03-09 to 2026-03-15
2. Excel file with correct sheet structure
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def check_gcal_events(cur):
    """Check Google Calendar events."""
    errors = []

    cur.execute("""
        SELECT id, summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%%cook%%'
        AND start_datetime >= '2026-03-09T00:00:00+00:00'
        AND start_datetime < '2026-03-16T23:59:59+00:00'
    """)
    events = cur.fetchall()

    if len(events) < 21:
        errors.append(f"Found {len(events)} cooking events in March 9-15, expected at least 21")

    # Check events span at least 7 distinct calendar dates (in any timezone representation)
    if events:
        days = set()
        for _, summary, start_dt, _ in events:
            if start_dt:
                # Use the date from the stored datetime; timezone shifts may move some
                # events to adjacent dates, so we just need at least 7 distinct dates
                if hasattr(start_dt, 'date'):
                    days.add(start_dt.date())
                else:
                    days.add(str(start_dt)[:10])
        if len(days) < 7:
            errors.append(f"Events span {len(days)} days, expected at least 7")

    return errors


def check_excel(workspace):
    """Check Excel file structure."""
    from openpyxl import load_workbook

    errors = []
    xlsx_path = os.path.join(workspace, "Weekly_Meal_Plan.xlsx")

    if not os.path.exists(xlsx_path):
        return ["Weekly_Meal_Plan.xlsx not found"]

    wb = load_workbook(xlsx_path)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    if "meal plan" not in sheet_names_lower:
        errors.append(f"Missing 'Meal Plan' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("meal plan")]]
        headers = [str(cell.value).lower() if cell.value else "" for cell in ws[1]]

        for rh in ["day", "meal_type", "recipe_name", "prep_time"]:
            if not any(rh.replace("_", " ") in h or rh in h for h in headers):
                errors.append(f"Meal Plan sheet missing header: {rh}")

        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        if data_rows < 21:
            errors.append(f"Meal Plan sheet has {data_rows} data rows, expected at least 21")

    if "shopping list" not in sheet_names_lower:
        errors.append(f"Missing 'Shopping List' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("shopping list")]]
        headers = [str(cell.value).lower() if cell.value else "" for cell in ws[1]]
        if not any("ingredient" in h for h in headers):
            errors.append("Shopping List sheet missing Ingredient header")

        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        if data_rows < 5:
            errors.append(f"Shopping List has {data_rows} ingredients, expected at least 5")

    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    # Check GCal
    print("\n=== Checking Google Calendar Events ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        gcal_errors = check_gcal_events(cur)
        cur.close()
        conn.close()
        if gcal_errors:
            for e in gcal_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(gcal_errors)
        else:
            print("  [PASS] GCal events check passed")
    except Exception as e:
        err = f"GCal check error: {e}"
        print(f"  [FAIL] {err}")
        all_errors.append(err)

    # Check Excel
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    # Summary
    print(f"\n=== SUMMARY ===")
    if all_errors:
        for e in all_errors:
            print(f"  [ERROR] {e}")
        print("  Overall: FAIL")
    else:
        print("  Overall: PASS")

    if args.res_log_file:
        result = {"errors": all_errors, "success": len(all_errors) == 0}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(1 if all_errors else 0)


if __name__ == "__main__":
    main()

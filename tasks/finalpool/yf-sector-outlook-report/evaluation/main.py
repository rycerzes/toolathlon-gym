"""Evaluation for yf-sector-outlook-report."""
import os
import argparse, json, os, sys
import psycopg2


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Sector_Outlook.xlsx")
    if not os.path.exists(path):
        return ["Sector_Outlook.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        rows = load_sheet_rows(wb, "Sector Performance")
        if rows is None:
            errors.append("Sheet 'Sector Performance' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 5:
                errors.append(f"Sector Performance has {len(data_rows)} rows, expected 5")
            # Check stocks present
            stocks = {str(r[1]).strip().upper() for r in data_rows if r and len(r) > 1 and r[1]}
            for sym in ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]:
                if sym not in stocks:
                    errors.append(f"Stock {sym} missing from Sector Performance")
            # Check GOOGL return
            for r in data_rows:
                if len(r) > 4 and r[1] and str(r[1]).strip().upper() == "GOOGL":
                    if not num_close(r[4], 74.57, abs_tol=2.0):
                        errors.append(f"GOOGL Return_1Y_Pct={r[4]}, expected ~74.57")

        rows2 = load_sheet_rows(wb, "Cross-Sector Summary")
        if rows2 is None:
            errors.append("Sheet 'Cross-Sector Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows2 if r[0]}
            if "positive_outlook_count" in lookup:
                if not num_close(lookup["positive_outlook_count"], 3, abs_tol=0):
                    errors.append(f"Positive_Outlook_Count={lookup['positive_outlook_count']}, expected 3")
            if "high_risk_count" in lookup:
                if not num_close(lookup["high_risk_count"], 1, abs_tol=0):
                    errors.append(f"High_Risk_Count={lookup['high_risk_count']}, expected 1")
            if "avg_1y_return" in lookup:
                if not num_close(lookup["avg_1y_return"], 39.16):
                    errors.append(f"Avg_1Y_Return={lookup['avg_1y_return']}, expected ~39.16")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    path = os.path.join(agent_workspace, "Sector_Report.docx")
    if not os.path.exists(path):
        return ["Sector_Report.docx not found"]
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join([p.text for p in doc.paragraphs]).lower()
        if len(text) < 200:
            errors.append(f"Sector_Report.docx too short ({len(text)} chars)")
        for kw in ["sector", "outlook", "technology", "energy"]:
            if kw not in text:
                errors.append(f"Sector_Report.docx missing keyword '{kw}'")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def check_notion():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM notion.pages")
        count = cur.fetchone()[0]
        # Should have at least 2 pages (parent + the created one)
        if count < 2:
            errors.append(f"Only {count} Notion pages, expected at least 2 (parent + sector outlook)")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

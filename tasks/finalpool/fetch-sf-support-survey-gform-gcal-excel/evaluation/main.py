"""
Evaluation script for fetch-sf-support-survey-gform-gcal-excel task.

Checks:
1. Support_Satisfaction_Analysis.xlsx with 4 sheets and correct data
2. Google Form for ongoing feedback
3. Calendar events for 4 quarterly reviews
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


def check_excel(agent_workspace):
    """Check Support_Satisfaction_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Support_Satisfaction_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: Survey Results ---
    survey_sheet = None
    for name in wb.sheetnames:
        if "survey" in name.lower() and "summary" not in name.lower():
            survey_sheet = name
            break
    if not survey_sheet:
        record("Survey Results sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Survey Results sheet exists", True)
        ws = wb[survey_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 20
        record("Survey Results has 20 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 2: Survey Summary ---
    summary_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            summary_sheet = name
            break
    if not summary_sheet:
        record("Survey Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Survey Summary sheet exists", True)
        ws = wb[summary_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        for row in data_rows:
            if row and row[0]:
                metric = str(row[0]).strip().lower()
                val = row[1]
                if "total_respondents" in metric or "total" in metric and "respondent" in metric:
                    ok = num_close(val, 20, tol=0)
                    record("Total respondents = 20", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "avg_overall" in metric or ("overall" in metric and "satisfaction" in metric):
                    ok = num_close(val, 3.25, tol=0.3)
                    record("Avg overall satisfaction ~3.25", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "lowest" in metric:
                    ok = str_contains(val, "low")
                    record("Lowest rated priority is Low", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "highest" in metric:
                    ok = str_contains(val, "high")
                    record("Highest rated priority is High", ok, f"Got {val}")
                    if not ok:
                        all_ok = False

    # --- Sheet 3: Ticket System Comparison ---
    comp_sheet = None
    for name in wb.sheetnames:
        if "ticket" in name.lower() or "comparison" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Ticket System Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Ticket System Comparison sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 3
        record("Ticket Comparison has 3 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

        # Check ticket avg response hours for High priority
        for row in data_rows:
            if row and str_contains(row[0], "high"):
                # Ticket avg response hours should be ~6.23
                found = False
                for cell in row[1:]:
                    if num_close(cell, 6.23, tol=1.0):
                        found = True
                        break
                record("High priority ticket response ~6.23 hrs", found,
                       f"Row: {str(row)[:200]}")
                if not found:
                    all_ok = False

    # --- Sheet 4: Improvement Areas ---
    imp_sheet = None
    for name in wb.sheetnames:
        if "improvement" in name.lower():
            imp_sheet = name
            break
    if not imp_sheet:
        record("Improvement Areas sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Improvement Areas sheet exists", True)
        ws = wb[imp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        # Should have at least 2 rows (Response Time and Resolution Quality are below 4.0)
        ok = len(data_rows) >= 2
        record("Improvement Areas has >= 2 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

        # Check that Response Time is listed (avg 3.10, well below 4.0)
        has_response_time = False
        for row in data_rows:
            if row and str_contains(row[0], "response"):
                has_response_time = True
                break
        record("Response Time in improvement areas", has_response_time)
        if not has_response_time:
            all_ok = False

    wb.close()
    return all_ok


def check_gform():
    """Check Google Form for ongoing feedback."""
    print("\n=== Checking Google Form ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title, description FROM gform.forms")
        forms = cur.fetchall()

        found_form = False
        form_id = None
        for fid, title, desc in forms:
            title_lower = (title or "").lower()
            if ("support" in title_lower or "feedback" in title_lower or
                    "customer" in title_lower or "satisfaction" in title_lower):
                if "employee" not in title_lower:  # Skip noise form
                    found_form = True
                    form_id = fid
                    break

        record("Customer feedback form exists", found_form,
               f"Found forms: {[(t, d[:50] if d else '') for _, t, d in forms]}")

        all_ok = found_form

        if form_id:
            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s", (form_id,))
            questions = cur.fetchall()
            q_count = len(questions)
            ok = q_count >= 4
            record(f"Form has >= 4 questions", ok, f"Found {q_count}")
            if not ok:
                all_ok = False

            # Check for specific question types
            q_titles = " ".join((t or "").lower() for t, _ in questions)
            has_satisfaction = "satisfaction" in q_titles or "overall" in q_titles or "rating" in q_titles
            record("Has satisfaction question", has_satisfaction, f"Q titles: {q_titles[:200]}")
            if not has_satisfaction:
                all_ok = False

            has_comment = any("text" in (qt or "").lower() or "paragraph" in (qt or "").lower()
                              for _, qt in questions)
            # Also check if there's a question about comments
            has_comment = has_comment or "comment" in q_titles or "feedback" in q_titles
            record("Has comments/text question", has_comment)
            if not has_comment:
                all_ok = False

        cur.close()
        conn.close()
        return all_ok

    except Exception as e:
        record("Google Form DB accessible", False, str(e))
        return False


def check_calendar():
    """Check calendar events for 4 quarterly review meetings."""
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    all_ok = True
    quarters_found = set()

    for summary, description, start_dt, end_dt in events:
        summary_lower = (summary or "").lower()
        if "support" in summary_lower or "satisfaction" in summary_lower or "review" in summary_lower:
            for q in ["q1", "q2", "q3", "q4"]:
                if q in summary_lower:
                    quarters_found.add(q)

    ok = len(quarters_found) >= 4
    record(f"All 4 quarterly review events found", ok,
           f"Found quarters: {quarters_found}")
    if not ok:
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    gform_ok = check_gform()
    cal_ok = check_calendar()

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  GForm:    {'PASS' if gform_ok else 'FAIL'}")
    print(f"  Calendar: {'PASS' if cal_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    overall = excel_ok and gform_ok and cal_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

"""
Evaluation for howtocook-diet-gform-notion-excel task.
Checks: GForm with questions, Notion page, Excel with 2 sheets, email.
"""
import argparse
import os
import sys

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()

        target_form = None
        for fid, title in forms:
            if title and ("dietary" in title.lower() or "preference" in title.lower() or "survey" in title.lower()):
                target_form = fid
                break

        record("GForm 'Dietary Preference Survey' exists",
               target_form is not None,
               f"Found forms: {[t for _, t in forms]}")

        if target_form is None:
            conn.close()
            return

        cur.execute("SELECT title, question_type, required FROM gform.questions WHERE form_id = %s ORDER BY position", (target_form,))
        questions = cur.fetchall()
        record("GForm has at least 4 questions", len(questions) >= 4,
               f"Found {len(questions)} questions")

        q_types = [q[1] for q in questions]
        has_radio = "RADIO" in q_types or "MULTIPLE_CHOICE" in q_types
        has_checkbox = "CHECKBOX" in q_types
        has_scale = "SCALE" in q_types or "LINEAR_SCALE" in q_types
        has_text = "TEXT" in q_types or "SHORT_ANSWER" in q_types or "PARAGRAPH" in q_types
        record("GForm has a multiple-choice (RADIO) question", has_radio,
               f"Question types: {q_types}")
        record("GForm has a checkbox question", has_checkbox,
               f"Question types: {q_types}")
        record("GForm has a scale question", has_scale,
               f"Question types: {q_types}")
        record("GForm has a text question", has_text,
               f"Question types: {q_types}")

        # Check question about dietary restrictions
        q_titles_lower = [q[0].lower() if q[0] else "" for q in questions]
        has_dietary_q = any("dietary" in qt or "restriction" in qt or "vegetarian" in qt for qt in q_titles_lower)
        record("GForm has dietary restrictions question", has_dietary_q,
               f"Question titles: {q_titles_lower}")

        conn.close()
    except Exception as e:
        record("GForm connection", False, str(e))


def check_notion():
    print("\n=== Checking Notion Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()

        target_page = None
        for pid, props in pages:
            props_text = str(props).lower() if props else ""
            if "healthy recipe" in props_text or ("recipe" in props_text and "knowledge" in props_text):
                target_page = pid
                break

        record("Notion page 'Healthy Recipe Knowledge Base' exists",
               target_page is not None,
               f"Searched {len(pages)} pages")

        # Check total pages (one for KB + one per recipe or blocks within the page)
        record("Notion has at least one page created", len(pages) >= 1,
               f"Found {len(pages)} pages total")

        # Check for recipe content across all pages
        recipe_pages = 0
        for pid, props in pages:
            props_text = str(props).lower() if props else ""
            # Count pages with recipe-like content
            if any(kw in props_text for kw in ["recipe", "breakfast", "lunch", "dinner", "chicken", "beef", "fish", "vegetable", "rice", "noodle", "soup"]):
                recipe_pages += 1

        record("Notion has recipe content in pages", recipe_pages >= 1,
               f"Found {recipe_pages} recipe-related pages out of {len(pages)}")

        conn.close()
    except Exception as e:
        record("Notion connection", False, str(e))


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel File ===")
    xl_path = os.path.join(agent_workspace, "Recipe_Overview.xlsx")
    if not os.path.isfile(xl_path):
        record("Excel file Recipe_Overview.xlsx exists", False, f"Not found at: {xl_path}")
        return
    record("Excel file Recipe_Overview.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xl_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names = [s.lower() for s in wb.sheetnames]
    has_recipes = any("recipe" in s for s in sheet_names)
    has_summary = any("category" in s or "summary" in s for s in sheet_names)
    record("Excel has 'Recipes' sheet", has_recipes, f"Found sheets: {wb.sheetnames}")
    record("Excel has 'Category Summary' sheet", has_summary, f"Found sheets: {wb.sheetnames}")

    # Check Recipes sheet has at least 6 data rows
    recipes_sheet = None
    for sname in wb.sheetnames:
        if "recipe" in sname.lower():
            recipes_sheet = wb[sname]
            break

    if recipes_sheet:
        data_rows = sum(1 for row in recipes_sheet.iter_rows(min_row=2, values_only=True)
                        if any(cell is not None and str(cell).strip() != "" for cell in row))
        record("Excel Recipes sheet has at least 6 recipe rows", data_rows >= 6,
               f"Found {data_rows} data rows")

    # Check Category Summary has at least 3 categories
    cat_sheet = None
    for sname in wb.sheetnames:
        if "category" in sname.lower() or "summary" in sname.lower():
            cat_sheet = wb[sname]
            break

    if cat_sheet:
        cat_rows = sum(1 for row in cat_sheet.iter_rows(min_row=2, values_only=True)
                       if any(cell is not None and str(cell).strip() != "" for cell in row))
        record("Excel Category Summary has at least 3 categories", cat_rows >= 3,
               f"Found {cat_rows} category rows")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Recipe_Overview.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        record(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices.append(len(gt_rows) - 1)
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        record(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1}",
                               False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    record(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                record(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%healthy%' OR LOWER(subject) LIKE '%recipe%' OR LOWER(subject) LIKE '%eating%'
        """)
        emails = cur.fetchall()
        record("Email about healthy eating/recipes sent", len(emails) > 0,
               f"Found {len(emails)} matching emails")

        if emails:
            target_found = False
            for subject, to_addr in emails:
                to_str = str(to_addr).lower() if to_addr else ""
                if "wellness.team@company.com" in to_str:
                    target_found = True
                    break
            record("Email sent to wellness.team@company.com", target_found,
                   f"Recipients: {[e[1] for e in emails]}")

        conn.close()
    except Exception as e:
        record("Email connection", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_gform()
    check_notion()
    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

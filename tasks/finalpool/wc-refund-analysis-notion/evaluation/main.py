"""
Evaluation script for wc-refund-analysis-notion task.

Checks:
1. Excel file (Refund_Analysis.xlsx) - refund count matches DB
2. Notion page with 'refund' in title
3. Email with 'refund' in subject sent to cfo@company.com
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def get_expected_refunds():
    """Get expected refund data from wc.refunds."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, order_id, amount, reason, date_created FROM wc.refunds ORDER BY id")
    refunds = cur.fetchall()
    cur.execute("SELECT COUNT(*) FROM wc.orders")
    order_count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return refunds, order_count


def check_excel(agent_workspace, expected_refunds, order_count):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Refund_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return False

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # Check Refund Details sheet
    ws = None
    for s in wb.sheetnames:
        if "refund" in s.lower() and "detail" in s.lower():
            ws = wb[s]
            break
    if ws is None:
        for s in wb.sheetnames:
            if "refund" in s.lower():
                ws = wb[s]
                break
        if ws is None and len(wb.sheetnames) > 0:
            ws = wb[wb.sheetnames[0]]

    check("Sheet with refund details exists", ws is not None, f"Sheets: {wb.sheetnames}")
    if ws is None:
        return False

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    expected_count = len(expected_refunds)
    check(f"Refund count matches DB ({expected_count})",
          len(rows) == expected_count,
          f"Got {len(rows)} rows, expected {expected_count}")

    # Check Summary sheet
    ws2 = None
    for s in wb.sheetnames:
        if "summary" in s.lower():
            ws2 = wb[s]
            break
    check("Summary sheet exists", ws2 is not None, f"Sheets: {wb.sheetnames}")

    if ws2:
        summary_data = {}
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                key = str(row[0]).strip().lower().replace(" ", "_")
                summary_data[key] = row[1]

        check("Summary has total refunds entry",
              any("total" in k and "refund" in k for k in summary_data),
              f"Keys: {list(summary_data.keys())}")

    return True


def check_notion():
    print("\n=== Checking Notion ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check pages
    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()

    refund_pages = []
    for pid, props in pages:
        if props:
            props_str = json.dumps(props).lower() if isinstance(props, dict) else str(props).lower()
            if "refund" in props_str:
                refund_pages.append(pid)

    check("Notion page with 'refund' in properties exists",
          len(refund_pages) > 0,
          f"Found {len(pages)} pages, {len(refund_pages)} with refund", db=True)

    cur.close()
    conn.close()
    return len(refund_pages) > 0


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_email] Found {len(all_emails)} emails.")

    found = False
    for subject, to_addr, body_text in all_emails:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                if isinstance(parsed, list):
                    to_str = " ".join(str(r).lower() for r in parsed)
                else:
                    to_str = str(to_addr).lower()
            except (json.JSONDecodeError, TypeError):
                to_str = str(to_addr).lower()

        if "cfo@company.com" in to_str:
            found = True
            check("Email subject contains 'refund'",
                  "refund" in (subject or "").lower(),
                  f"Subject: {subject}", db=True)
            break

    check("Email sent to cfo@company.com", found, db=True)
    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_refunds, order_count = get_expected_refunds()
    print(f"Expected {len(expected_refunds)} refunds, {order_count} orders")

    check_excel(args.agent_workspace, expected_refunds, order_count)
    check_notion()
    check_email()

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if args.res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": file_ok}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if file_ok else 1)


if __name__ == "__main__":
    main()

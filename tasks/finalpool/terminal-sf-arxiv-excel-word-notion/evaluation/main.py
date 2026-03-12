"""Evaluation for terminal-sf-arxiv-excel-word-notion.
Checks:
1. Retention_Strategy.xlsx with 3 sheets (Department_Analysis, Research_Summary, Action_Plan)
2. Retention_Strategy_Report.docx
3. Notion database "Retention Action Items" with 7 department pages
4. flight_risk_analysis.py and synthesis.py scripts exist
5. flight_risk_analysis.json and synthesis.json outputs exist
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

DEPARTMENTS = ["Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"]

# Hardcoded fallback flight risk data (sat<=4 AND perf>=4)
_FALLBACK_EXPECTED_DATA = {
    "Engineering": {"headcount": 7096, "flight_risk": 566, "pct": 7.98, "priority": "Medium"},
    "Finance":     {"headcount": 7148, "flight_risk": 598, "pct": 8.37, "priority": "High"},
    "HR":          {"headcount": 7077, "flight_risk": 594, "pct": 8.39, "priority": "High"},
    "Operations":  {"headcount": 7120, "flight_risk": 564, "pct": 7.92, "priority": "Medium"},
    "R&D":         {"headcount": 7083, "flight_risk": 576, "pct": 8.13, "priority": "Medium"},
    "Sales":       {"headcount": 7232, "flight_risk": 596, "pct": 8.24, "priority": "Medium"},
    "Support":     {"headcount": 7244, "flight_risk": 537, "pct": 7.41, "priority": "Low"},
}


def _get_expected_data_from_db():
    """Query sf_data schema to compute department headcounts and flight risk counts dynamically."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        try:
            # Get headcount per department
            cur.execute("""
                SELECT "DEPARTMENT", COUNT(*) as headcount
                FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                GROUP BY "DEPARTMENT"
            """)
            headcounts = {r[0]: r[1] for r in cur.fetchall()}

            # Get flight risk count per department (satisfaction<=4 AND performance>=4)
            cur.execute("""
                SELECT "DEPARTMENT", COUNT(*) as flight_risk
                FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                WHERE "JOB_SATISFACTION" <= 4 AND "PERFORMANCE_RATING" >= 4
                GROUP BY "DEPARTMENT"
            """)
            flight_risks = {r[0]: r[1] for r in cur.fetchall()}

            result = {}
            for dept in DEPARTMENTS:
                hc = headcounts.get(dept, 0)
                fr = flight_risks.get(dept, 0)
                pct = round(fr / hc * 100, 2) if hc > 0 else 0
                if pct >= 8.3:
                    priority = "High"
                elif pct <= 7.5:
                    priority = "Low"
                else:
                    priority = "Medium"
                result[dept] = {"headcount": hc, "flight_risk": fr, "pct": pct, "priority": priority}
            return result
        finally:
            cur.close()
            conn.close()
    except Exception:
        return _FALLBACK_EXPECTED_DATA


EXPECTED_DATA = _get_expected_data_from_db()

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(workspace):
    print("\n=== Check 1: Retention_Strategy.xlsx ===")
    path = os.path.join(workspace, "Retention_Strategy.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Sheet 1: Department_Analysis
    da_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s and "analysis" in s), 0)
    ws1 = wb[sheets[da_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Department_Analysis has 7 rows", len(data1) >= 7, f"Found {len(data1)}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has flight_risk_pct column",
              any("flight" in h and "pct" in h for h in headers) or any("risk" in h and "%" in h for h in headers) or any("flight_risk_pct" in h for h in headers),
              f"Headers: {rows1[0]}")
        check("Has avg_satisfaction column",
              any("satisfaction" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Check actual values
    dept_col = next((i for i, h in enumerate(headers) if "department" in h or "dept" in h), 0) if rows1 else 0
    hc_col = next((i for i, h in enumerate(headers) if "headcount" in h or "head_count" in h), 1) if rows1 else 1
    fr_col = next((i for i, h in enumerate(headers) if "flight_risk_count" in h or ("flight" in h and "count" in h)), 2) if rows1 else 2
    pct_col = next((i for i, h in enumerate(headers) if "pct" in h or "percent" in h), 3) if rows1 else 3

    found_depts = 0
    for row in data1:
        dept_name = str(row[dept_col]).strip() if row[dept_col] else ""
        if dept_name in EXPECTED_DATA:
            found_depts += 1
            exp = EXPECTED_DATA[dept_name]
            if len(row) > hc_col and row[hc_col]:
                check(f"{dept_name} headcount correct",
                      num_close(row[hc_col], exp["headcount"], 50),
                      f"Got {row[hc_col]}, expected ~{exp['headcount']}")
            if len(row) > fr_col and row[fr_col]:
                check(f"{dept_name} flight_risk_count correct",
                      num_close(row[fr_col], exp["flight_risk"], 20),
                      f"Got {row[fr_col]}, expected ~{exp['flight_risk']}")
    check("All 7 departments found in Department_Analysis", found_depts >= 7, f"Found {found_depts}")

    # Sheet 2: Research_Summary
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Research_Summary has 3 rows", len(data2) >= 3, f"Found {len(data2)}")
        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has applicability_score column",
                  any("applicability" in h or "score" in h for h in headers2),
                  f"Headers: {rows2[0]}")
        # Check that relevant papers are included (not the noise ones)
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Contains retention-related paper", "retention" in all_text2 or "turnover" in all_text2 or "employee" in all_text2)
        check("Does NOT contain autonomous vehicle paper", "autonomous vehicle" not in all_text2 and "urban" not in all_text2,
              "Noise paper included in research summary")

    # Sheet 3: Action_Plan
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "action" in s or "plan" in s), 2)
    if ap_idx < len(sheets):
        ws3 = wb[sheets[ap_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Action_Plan has 7 rows", len(data3) >= 7, f"Found {len(data3)}")

        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            pri_col = next((i for i, h in enumerate(headers3) if "priority" in h), None)
            dept_col3 = next((i for i, h in enumerate(headers3) if "department" in h or "dept" in h), 0)

            if pri_col is not None:
                for row in data3:
                    dept_name = str(row[dept_col3]).strip() if row[dept_col3] else ""
                    if dept_name in EXPECTED_DATA:
                        exp_pri = EXPECTED_DATA[dept_name]["priority"]
                        got_pri = str(row[pri_col]).strip() if row[pri_col] else ""
                        check(f"{dept_name} priority is {exp_pri}",
                              got_pri.lower() == exp_pri.lower(),
                              f"Got '{got_pri}', expected '{exp_pri}'")


def check_word(workspace):
    print("\n=== Check 2: Retention_Strategy_Report.docx ===")
    path = os.path.join(workspace, "Retention_Strategy_Report.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    check("Has title mentioning retention", "retention" in full_text and ("strategy" in full_text or "report" in full_text))
    check("Mentions flight risk", "flight risk" in full_text or "flight-risk" in full_text)
    check("Mentions executive summary", "executive summary" in full_text or "summary" in full_text)
    check("Mentions research findings", "research" in full_text and ("finding" in full_text or "paper" in full_text))
    check("Mentions recommendations", "recommend" in full_text)
    check("Mentions specific departments", sum(1 for d in DEPARTMENTS if d.lower() in full_text) >= 5,
          f"Found {sum(1 for d in DEPARTMENTS if d.lower() in full_text)} departments")
    check("Has substantial content", len(full_text) > 500, f"Length: {len(full_text)}")
    check("Mentions priority levels", "high" in full_text and ("medium" in full_text or "low" in full_text))


def check_notion():
    print("\n=== Check 3: Notion Retention Action Items ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # Check database exists
        cur.execute("SELECT id, title, properties FROM notion.databases WHERE title::text ILIKE '%retention%action%'")
        dbs = cur.fetchall()
        if not dbs:
            cur.execute("SELECT id, title, properties FROM notion.databases")
            all_dbs = cur.fetchall()
            check("Notion database 'Retention Action Items' exists", False,
                  f"Found {len(all_dbs)} databases: {[str(d[1])[:50] for d in all_dbs]}")
            return
        check("Notion database 'Retention Action Items' exists", True)

        db_id = dbs[0][0]
        props = dbs[0][2] if dbs[0][2] else {}

        # Check properties
        prop_names = [k.lower() for k in props.keys()] if isinstance(props, dict) else []
        check("Has Priority property", any("priority" in p for p in prop_names), f"Props: {prop_names}")
        check("Has Status property", any("status" in p for p in prop_names), f"Props: {prop_names}")
        check("Has Strategy property", any("strategy" in p for p in prop_names), f"Props: {prop_names}")

        # Check pages
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE parent::text LIKE %s
        """, (f'%{db_id}%',))
        pages = cur.fetchall()
        check("Has 7 department pages", len(pages) >= 7, f"Found {len(pages)} pages")

        if pages:
            # Check that pages have correct priorities
            high_count = 0
            medium_count = 0
            low_count = 0
            for page in pages:
                props_page = page[1] if page[1] else {}
                props_text = json.dumps(props_page).lower()
                if '"high"' in props_text:
                    high_count += 1
                elif '"medium"' in props_text:
                    medium_count += 1
                elif '"low"' in props_text:
                    low_count += 1

            check("Has 2 High priority pages (Finance, HR)", high_count == 2,
                  f"Found {high_count} High priority pages")
            check("Has 4 Medium priority pages", medium_count == 4,
                  f"Found {medium_count} Medium priority pages")
            check("Has 1 Low priority page (Support)", low_count == 1,
                  f"Found {low_count} Low priority pages")

            # Check Status is Not Started
            not_started_count = sum(1 for p in pages
                                     if "not started" in json.dumps(p[1]).lower())
            check("All pages have Status 'Not Started'", not_started_count >= 7,
                  f"Found {not_started_count} with 'Not Started'")

    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python Scripts ===")
    check("flight_risk_analysis.py exists",
          os.path.exists(os.path.join(workspace, "flight_risk_analysis.py")))
    check("synthesis.py exists",
          os.path.exists(os.path.join(workspace, "synthesis.py")))


def check_json_outputs(workspace):
    print("\n=== Check 5: JSON Outputs ===")
    fr_path = os.path.join(workspace, "flight_risk_analysis.json")
    if os.path.exists(fr_path):
        check("flight_risk_analysis.json exists", True)
        try:
            with open(fr_path) as f:
                fr_data = json.load(f)
            check("flight_risk_analysis.json is valid JSON", True)
            # Check it has department data
            fr_text = json.dumps(fr_data).lower()
            check("flight_risk_analysis.json mentions departments",
                  sum(1 for d in DEPARTMENTS if d.lower() in fr_text) >= 5)
        except (json.JSONDecodeError, Exception) as e:
            check("flight_risk_analysis.json is valid JSON", False, str(e))
    else:
        check("flight_risk_analysis.json exists", False)

    syn_path = os.path.join(workspace, "synthesis.json")
    if os.path.exists(syn_path):
        check("synthesis.json exists", True)
        try:
            with open(syn_path) as f:
                syn_data = json.load(f)
            check("synthesis.json is valid JSON", True)
        except (json.JSONDecodeError, Exception) as e:
            check("synthesis.json is valid JSON", False, str(e))
    else:
        check("synthesis.json exists", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_scripts(args.agent_workspace)
    check_json_outputs(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

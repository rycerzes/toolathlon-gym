"""
Evaluation for yt-fireship-gform-survey-excel-gcal task.

Checks:
1. Community_Report.xlsx exists with "Top_Videos" sheet having >= 6 data rows
2. Engagement_Rate column exists in Top_Videos or Engagement_Analysis has rates
3. Engagement_Analysis sheet exists with >= 3 rows
4. GForm created with title containing "Survey", "Fireship", or "Community"
5. GForm has >= 4 questions
6. GCal has new event in April 2026 with "Community" or "Standup" in summary (not the noise Q&A at 16:00)
7. Email sent to community@devclub.io
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-3: Community_Report.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Community_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Community_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Top_Videos sheet has >= 6 data rows", False, "File missing")
        record("Engagement_Rate column or analysis exists", False, "File missing")
        record("Engagement_Analysis sheet has >= 3 rows", False, "File missing")
        return
    record("Community_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Top_Videos sheet
    top_key = next((sheet_names_lower[k] for k in sheet_names_lower
                    if "video" in k or "top" in k), None)
    if not top_key:
        record("Top_Videos sheet has >= 6 data rows", False,
               f"No Top_Videos sheet. Sheets: {wb.sheetnames}")
        record("Engagement_Rate column or analysis exists", False, "Sheet missing")
    else:
        ws = wb[top_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        record("Top_Videos sheet has >= 6 data rows", len(data_rows) >= 6,
               f"Found {len(data_rows)} data rows")
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            has_engagement = any("engagement" in h or "rate" in h for h in headers)
            record("Engagement_Rate column exists in Top_Videos", has_engagement,
                   f"Headers: {rows[0]}")
        else:
            record("Engagement_Rate column exists in Top_Videos", False, "Sheet empty")

    # Engagement_Analysis sheet
    eng_key = next((sheet_names_lower[k] for k in sheet_names_lower
                    if "engagement" in k or "analysis" in k), None)
    if not eng_key:
        record("Engagement_Analysis sheet has >= 3 rows", False,
               f"No Engagement_Analysis sheet. Sheets: {wb.sheetnames}")
    else:
        ws2 = wb[eng_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        record("Engagement_Analysis sheet has >= 3 rows", len(data_rows2) >= 3,
               f"Found {len(data_rows2)} data rows")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Community_Report.xlsx")
    if os.path.isfile(gt_path):
        import openpyxl as opx
        gt_wb = opx.load_workbook(gt_path, data_only=True)
        try:
            a_wb = opx.load_workbook(os.path.join(agent_workspace, "Community_Report.xlsx"), data_only=True)
        except Exception:
            a_wb = None
        if a_wb:
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in a_wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = a_wb[asn]; break
                if a_ws is None:
                    record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {a_wb.sheetnames}")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")
                for ri in range(min(3, len(gt_rows))):
                    if ri >= len(a_rows): break
                    ok = True
                    for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                        gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                        if gv is None: continue
                        if isinstance(gv, (int, float)):
                            if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                        else:
                            if not str_match(av, gv): ok = False; break
                    record(f"GT '{gt_sname}' row {ri+1} values", ok,
                           f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
            a_wb.close()
        gt_wb.close()


def check_gform():
    print("\n=== Check 4-5: Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT f.id, f.title FROM gform.forms f
                WHERE f.title ILIKE '%Survey%'
                   OR f.title ILIKE '%Fireship%'
                   OR f.title ILIKE '%Community%'
            """)
            forms = cur.fetchall()
            if not forms:
                record("GForm with 'Survey', 'Fireship', or 'Community' in title exists", False,
                       "No matching form found")
                record("GForm has >= 4 questions", False, "Form missing")
                conn.close()
                return
            record("GForm with 'Survey', 'Fireship', or 'Community' in title exists", True,
                   f"Found: {[f[1] for f in forms]}")

            form_id = forms[0][0]
            cur.execute("""
                SELECT COUNT(*) FROM gform.questions WHERE form_id = %s
            """, (form_id,))
            q_count = cur.fetchone()[0]
            record("GForm has >= 4 questions", q_count >= 4, f"Found {q_count} questions")
        conn.close()
    except Exception as e:
        record("GForm check", False, str(e))


def check_gcal():
    print("\n=== Check 6: GCal Community Standup in April 2026 ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM gcal.events
                WHERE (summary ILIKE '%Community%' OR summary ILIKE '%Standup%')
                  AND start_datetime >= '2026-04-01'
                  AND start_datetime < '2026-05-01'
                  AND summary NOT ILIKE '%Q&A%'
            """)
            count = cur.fetchone()[0]
        conn.close()
        record("GCal has new Community/Standup event in April 2026 (not Q&A noise)",
               count > 0, f"Found {count} events")
    except Exception as e:
        record("GCal check", False, str(e))


def check_email():
    print("\n=== Check 7: Email sent to community@devclub.io ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM email.messages
                WHERE to_addr::text ILIKE '%community@devclub.io%'
                  AND from_addr != 'community@devclub.io'
            """)
            count = cur.fetchone()[0]
            if count == 0:
                try:
                    cur.execute("""
                        SELECT COUNT(*) FROM email.sent_log
                        WHERE to_addr ILIKE '%community@devclub.io%'
                    """)
                    count = cur.fetchone()[0]
                except Exception:
                    pass
        conn.close()
        record("Email sent to community@devclub.io", count > 0, f"Found {count}")
    except Exception as e:
        record("Email check", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-fireship-gform-survey-excel-gcal")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_gform()
    check_gcal()
    check_email()

    all_passed = FAIL_COUNT == 0
    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print(f"\n{'='*40}")
    print(f"Result: {'PASS' if all_passed else 'FAIL'} - {summary}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "all_passed": all_passed}, f)

    return all_passed, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

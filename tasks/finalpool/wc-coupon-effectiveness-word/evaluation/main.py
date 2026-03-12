"""
Evaluation script for wc-coupon-effectiveness-word task.

Checks:
1. Word document has title and coupon codes
2. Excel has coupon performance data
3. Email sent with correct subject
Values dynamically computed from wc DB.
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_expected_coupon_data():
    """Query actual coupon usage from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get coupon usage from orders
    cur.execute("""
        SELECT
            cl->>'code' as coupon_code,
            COUNT(*) as usage_count,
            SUM((cl->>'discount')::numeric) as total_discount,
            AVG((cl->>'discount')::numeric) as avg_discount
        FROM wc.orders,
             jsonb_array_elements(coupon_lines) cl
        WHERE coupon_lines IS NOT NULL
          AND coupon_lines::text <> 'null'
          AND coupon_lines::text <> '[]'
        GROUP BY cl->>'code'
        ORDER BY total_discount DESC
    """)
    coupon_stats = cur.fetchall()

    # Total orders with coupons
    cur.execute("""
        SELECT COUNT(*)
        FROM wc.orders
        WHERE coupon_lines IS NOT NULL
          AND coupon_lines::text <> 'null'
          AND coupon_lines::text <> '[]'
    """)
    total_orders_with_coupons = cur.fetchone()[0]

    # Total coupons defined
    cur.execute("SELECT COUNT(*) FROM wc.coupons")
    total_coupons_defined = cur.fetchone()[0]

    # Get all defined coupon codes
    cur.execute("SELECT code FROM wc.coupons")
    defined_codes = set(row[0] for row in cur.fetchall())

    # Used coupon codes
    used_codes = set(row[0] for row in coupon_stats)

    cur.close()
    conn.close()

    return coupon_stats, total_orders_with_coupons, defined_codes, used_codes


def check_word(workspace):
    """Check Word document content."""
    from docx import Document

    errors = []
    docx_path = os.path.join(workspace, "Coupon_Effectiveness_Report.docx")
    if not os.path.exists(docx_path):
        return ["Coupon_Effectiveness_Report.docx not found"]

    doc = Document(docx_path)
    all_text = " ".join(p.text.lower() for p in doc.paragraphs)

    # Check title
    if "coupon" not in all_text or "effectiveness" not in all_text:
        errors.append("Word document missing 'coupon' or 'effectiveness' in text")

    if "analysis" not in all_text:
        errors.append("Word document missing 'analysis' in text")

    # Check that at least some coupon codes are mentioned
    _, _, _, used_codes = get_expected_coupon_data()
    codes_found = sum(1 for code in used_codes if code.lower() in all_text)
    if codes_found < min(3, len(used_codes)):
        errors.append(f"Word document mentions only {codes_found} coupon codes, expected at least {min(3, len(used_codes))}")

    return errors


def check_excel(workspace):
    """Check Excel file structure and data."""
    from openpyxl import load_workbook

    errors = []
    xlsx_path = os.path.join(workspace, "Coupon_Analysis.xlsx")
    if not os.path.exists(xlsx_path):
        return ["Coupon_Analysis.xlsx not found"]

    coupon_stats, total_orders_with_coupons, defined_codes, used_codes = get_expected_coupon_data()

    wb = load_workbook(xlsx_path)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Coupon Performance sheet
    if "coupon performance" not in sheet_names_lower:
        errors.append(f"Missing 'Coupon Performance' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("coupon performance")]]
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]

        for rh in ["coupon_code", "usage_count", "total_discount"]:
            if not any(rh in h or rh.replace("_", "") in h.replace("_", "") for h in headers):
                errors.append(f"Coupon Performance missing header: {rh}")

        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        if data_rows < len(coupon_stats):
            errors.append(f"Coupon Performance has {data_rows} rows, expected at least {len(coupon_stats)}")

    # Check Summary sheet
    if "summary" not in sheet_names_lower:
        errors.append(f"Missing 'Summary' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        summary_text = ""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    summary_text += str(cell.value).lower() + " "

        if "total" not in summary_text:
            errors.append("Summary sheet missing 'total' related metrics")

    return errors


def check_email(cur):
    """Check for the email."""
    errors = []
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%coupon effectiveness report%'
    """)
    emails = cur.fetchall()

    if not emails:
        errors.append("No email with subject 'Coupon Effectiveness Report' found")
    else:
        to_str = str(emails[0][2]).lower()
        if "marketing-director@company.com" not in to_str:
            errors.append(f"Email not sent to marketing-director@company.com, to_addr: {emails[0][2]}")

    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    # Check Word
    print("\n=== Checking Word Document ===")
    word_errors = check_word(args.agent_workspace)
    if word_errors:
        for e in word_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(word_errors)
    else:
        print("  [PASS] Word check passed")

    # Check Excel
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    # Check Email
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        email_errors = check_email(cur)
        if email_errors:
            for e in email_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(email_errors)
        else:
            print("  [PASS] Email check passed")
        cur.close()
        conn.close()
    except Exception as e:
        err = f"Email check error: {e}"
        print(f"  [FAIL] {err}")
        all_errors.append(err)

    # Summary
    print(f"\n=== SUMMARY ===")
    if all_errors:
        for e in all_errors:
            print(f"  [ERROR] {e}")
        print("  Overall: FAIL")
    else:
        print("  Overall: PASS")

    if args.res_log_file:
        result = {"errors": all_errors, "success": len(all_errors) == 0}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(1 if all_errors else 0)


if __name__ == "__main__":
    main()

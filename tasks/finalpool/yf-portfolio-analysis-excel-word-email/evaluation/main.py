"""Evaluation for yf-portfolio-analysis-excel-word-email.

Checks:
1. Portfolio_Analysis.xlsx exists with Holdings sheet (5 rows) and Summary sheet
2. Investment_Memo.docx exists with required sections
3. Email sent to investment-committee@fund.example.com with correct subject
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0

SYMBOLS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]
HOLDINGS = {
    "GOOGL": 50, "AMZN": 30, "JPM": 100, "JNJ": 75, "XOM": 120,
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol_pct=5.0):
    """Check if two values are within tol_pct percent of each other."""
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(b) < 1e-6:
        return abs(a) < 0.01
    return abs(a - b) / abs(b) * 100 <= tol_pct


def get_latest_prices():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT symbol, close FROM yf.stock_prices
        WHERE symbol IN ('GOOGL','AMZN','JPM','JNJ','XOM')
          AND date = (SELECT MAX(date) FROM yf.stock_prices WHERE symbol='GOOGL')
        ORDER BY symbol
    """)
    prices = {row[0]: float(row[1]) for row in cur.fetchall()}
    cur.close()
    conn.close()
    return prices


def check_excel(agent_ws, prices):
    print("\n=== Check 1: Portfolio_Analysis.xlsx ===")
    path = os.path.join(agent_ws, "Portfolio_Analysis.xlsx")
    check("File Portfolio_Analysis.xlsx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel is readable", False, str(e))
        return

    # Check Holdings sheet
    holdings_ws = None
    for sname in wb.sheetnames:
        if "holdings" in sname.lower():
            holdings_ws = wb[sname]
            break
    check("Sheet 'Holdings' exists", holdings_ws is not None, f"Sheets: {wb.sheetnames}")

    if holdings_ws is not None:
        rows = list(holdings_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Holdings has 5 data rows", len(non_empty) >= 5, f"Got {len(non_empty)}")

        # Check that all symbols appear
        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        for sym in SYMBOLS:
            check(f"Holdings contains symbol {sym}", sym in all_text)

        # Check that current values are approximately correct
        for row in non_empty:
            row_text = " ".join(str(c) for c in row if c is not None)
            for sym in SYMBOLS:
                if sym in row_text:
                    expected_val = round(HOLDINGS[sym] * prices.get(sym, 0), 2)
                    # Look for a numeric cell close to expected_val
                    found_val = False
                    for c in row:
                        try:
                            if num_close(float(c), expected_val, tol_pct=5.0):
                                found_val = True
                                break
                        except (TypeError, ValueError):
                            pass
                    check(f"Holdings {sym} current value ~= {expected_val}", found_val,
                          f"Row data: {[str(x) for x in row[:7]]}")

    # Check Summary sheet
    summary_ws = None
    for sname in wb.sheetnames:
        if "summary" in sname.lower():
            summary_ws = wb[sname]
            break
    check("Sheet 'Summary' exists", summary_ws is not None, f"Sheets: {wb.sheetnames}")

    if summary_ws is not None:
        total_value = sum(HOLDINGS[s] * prices.get(s, 0) for s in SYMBOLS)
        all_text = " ".join(str(c) for row in summary_ws.iter_rows(values_only=True) for c in row if c is not None)
        # Check total portfolio value appears approximately
        found_total = any(
            num_close(float(str(c).replace(",", "").replace("$", "")), total_value, tol_pct=5.0)
            for row in summary_ws.iter_rows(min_row=2, values_only=True)
            for c in row
            if c is not None and str(c).replace(",", "").replace("$", "").replace(".", "").replace("-", "").isdigit() or
            (isinstance(c, (int, float)) and num_close(c, total_value, tol_pct=5.0))
        )
        check(f"Summary contains total portfolio value ~= {total_value:.2f}", found_total,
              f"Content: {all_text[:200]}")
        check("Summary mentions Best_Performer or best performer", "Best" in all_text or "best" in all_text.lower())
        check("Summary mentions Worst_Performer or worst performer", "Worst" in all_text or "worst" in all_text.lower())


def check_word(agent_ws):
    print("\n=== Check 2: Investment_Memo.docx ===")
    path = os.path.join(agent_ws, "Investment_Memo.docx")
    check("File Investment_Memo.docx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        from docx import Document
        doc = Document(path)
    except Exception as e:
        check("Word doc is readable", False, str(e))
        return

    full_text = " ".join(p.text for p in doc.paragraphs)
    full_text_lower = full_text.lower()

    check("Memo contains 'Portfolio' and 'Report' or 'Analysis' in title",
          "portfolio" in full_text_lower and ("report" in full_text_lower or "analysis" in full_text_lower))
    check("Memo has 'Holdings' section", "holdings" in full_text_lower)
    check("Memo has 'Performance' section", "performance" in full_text_lower)
    check("Memo has 'Recommendations' or 'Recommendation' section", "recommendation" in full_text_lower)

    # Check all symbols appear
    for sym in SYMBOLS:
        check(f"Memo mentions stock {sym}", sym in full_text)


def check_email():
    print("\n=== Check 3: Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, to_addr, from_addr FROM email.messages
        WHERE to_addr::text ILIKE '%investment-committee%'
           OR to_addr::text ILIKE '%fund.example%'
        LIMIT 10
    """)
    rows = cur.fetchall()
    check("Email sent to investment-committee@fund.example.com",
          len(rows) > 0, "No matching email found")

    if rows:
        subjects = [r[0] or "" for r in rows]
        check("Email subject contains 'Portfolio'",
              any("portfolio" in s.lower() for s in subjects),
              f"Subjects: {subjects}")
        check("Email subject contains 'Analysis' or 'Report'",
              any("analysis" in s.lower() or "report" in s.lower() for s in subjects),
              f"Subjects: {subjects}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: yf-portfolio-analysis-excel-word-email ===")
    prices = get_latest_prices()
    print(f"Latest prices from DB: {prices}")

    check_excel(args.agent_workspace, prices)
    check_word(args.agent_workspace)
    check_email()

    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"pass": PASS_COUNT, "fail": FAIL_COUNT}, f)

    # Pass if at least 70% of checks pass (flexible for agent outputs)
    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        sys.exit(1)
    pct = PASS_COUNT / total * 100
    print(f"Score: {pct:.1f}%")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

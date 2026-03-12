"""Evaluation for yf-financial-health-excel-notion."""
import os
import argparse, os, sys
import psycopg2


def num_close(a, b, tol=1e9):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Financial_Health_Report.xlsx")
    if not os.path.exists(path):
        return ["Financial_Health_Report.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        # Check Income Statement sheet
        rows = load_sheet_rows(wb, "Income Statement")
        if rows is None:
            errors.append("Sheet 'Income Statement' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 5:
                errors.append(f"Income Statement has {len(data_rows)} rows, expected 5")
            symbols_found = {str(r[0]).strip().upper() for r in data_rows}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols_found:
                    errors.append(f"Symbol {sym} missing from Income Statement")
            # Check numeric data exists
            for r in data_rows:
                if r[1] is None or r[2] is None:
                    errors.append(f"Missing revenue or net income for {r[0]}")
                    break

        # Check Balance Sheet sheet
        rows2 = load_sheet_rows(wb, "Balance Sheet")
        if rows2 is None:
            errors.append("Sheet 'Balance Sheet' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(f"Balance Sheet has {len(data_rows2)} rows, expected 5")
            symbols_found2 = {str(r[0]).strip().upper() for r in data_rows2}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols_found2:
                    errors.append(f"Symbol {sym} missing from Balance Sheet")
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_notion():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE '%financial health%'
            ORDER BY created_time DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            # Also check blocks
            conn2 = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                     user="eigent", password="camel")
            cur2 = conn2.cursor()
            cur2.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE properties::text ILIKE '%financial%' OR properties::text ILIKE '%dashboard%'
            """)
            count = cur2.fetchone()[0]
            cur2.close(); conn2.close()
            if count == 0:
                errors.append("No Notion page found with 'Financial Health' in title/content")
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE to_addr::text ILIKE '%finance_team@company.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No email found to finance_team@company.com")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

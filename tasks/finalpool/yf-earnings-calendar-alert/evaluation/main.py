"""Evaluation for yf-earnings-calendar-alert."""
import os
import argparse, os, sys
import psycopg2


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Earnings_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Earnings_Analysis.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Earnings Calendar
        rows = load_sheet_rows(wb, "Earnings Calendar")
        if rows is None:
            errors.append("Sheet 'Earnings Calendar' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 3:
                errors.append(f"Earnings Calendar has {len(data_rows)} rows, expected 3")
            symbols = {str(r[0]).strip().upper() for r in data_rows if r[0]}
            for sym in ["AMZN", "GOOGL", "JPM"]:
                if sym not in symbols:
                    errors.append(f"Symbol {sym} missing from Earnings Calendar")
            for r in data_rows:
                if r[0] and str(r[0]).strip().upper() == "GOOGL":
                    if len(r) > 4 and not num_close(r[4], 2.70, abs_tol=0.2):
                        errors.append(f"GOOGL Historical_Avg_EPS={r[4]}, expected ~2.70")
                    if len(r) > 5 and str(r[5]).strip().lower() != "below":
                        errors.append(f"GOOGL Surprise_Trend={r[5]}, expected Below")
                if r[0] and str(r[0]).strip().upper() == "JPM":
                    if len(r) > 4 and not num_close(r[4], 5.0, abs_tol=0.3):
                        errors.append(f"JPM Historical_Avg_EPS={r[4]}, expected ~5.0")

        # Financial Trends
        rows2 = load_sheet_rows(wb, "Financial Trends")
        if rows2 is None:
            errors.append("Sheet 'Financial Trends' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 12:
                errors.append(f"Financial Trends has {len(data_rows2)} rows, expected 12")

        # Alert Summary
        rows3 = load_sheet_rows(wb, "Alert Summary")
        if rows3 is None:
            errors.append("Sheet 'Alert Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows3 if r[0]}
            if "stocks_reporting" in lookup:
                if not num_close(lookup["stocks_reporting"], 3, abs_tol=0):
                    errors.append(f"Stocks_Reporting={lookup['stocks_reporting']}, expected 3")
            if "avg_expected_eps" in lookup:
                if not num_close(lookup["avg_expected_eps"], 2.53, abs_tol=0.2):
                    errors.append(f"Avg_Expected_EPS={lookup['avg_expected_eps']}, expected ~2.53")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE to_addr::text ILIKE '%investment-team@company.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            errors.append("No email found to investment-team@company.com")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

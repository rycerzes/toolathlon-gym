"""Evaluation script for pw-sf-benchmark-excel-word-email."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        s = str(val).replace(',', '').replace('%', '').replace('$', '').strip()
        if s.upper() == 'N/A':
            return default
        return float(s)
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def values_close(a, b, tolerance=0.15):
    """Check if two numeric values are within tolerance (relative)."""
    if a is None or b is None:
        return False
    if b == 0:
        return abs(a) < 1.0
    return abs(a - b) / max(abs(b), 1e-6) <= tolerance


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # --- Excel Evaluation ---
    excel_path = os.path.join(agent_workspace, "Benchmark_Analysis.xlsx")
    check("Benchmark_Analysis.xlsx exists", os.path.exists(excel_path))

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Benchmark_Analysis.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Sheet 1: Internal Metrics
        check("Internal Metrics sheet exists", "Internal Metrics" in wb.sheetnames,
              f"sheets: {wb.sheetnames}")
        if "Internal Metrics" in wb.sheetnames:
            ws = wb["Internal Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Internal Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            for expected_col in ['metric', 'internal_value', 'source_note']:
                check(f"Internal Metrics has '{expected_col}' column",
                      expected_col in headers, f"headers: {headers}")

            # Verify key metric values against groundtruth
            if gt_wb and "Internal Metrics" in gt_wb.sheetnames:
                gt_ws = gt_wb["Internal Metrics"]
                gt_data = {str(r[0]).strip(): r[1] for r in gt_ws.iter_rows(min_row=2, values_only=True) if r[0]}
                agent_data = {str(r[0]).strip(): r[1] for r in data_rows if r and r[0]}

                for metric_name in ['Avg_Salary', 'Avg_Order_Value', 'Avg_Satisfaction']:
                    gt_val = safe_float(gt_data.get(metric_name))
                    # Try variations of metric name
                    agent_val = None
                    for k, v in agent_data.items():
                        if metric_name.lower().replace('_', '') in k.lower().replace('_', '').replace(' ', ''):
                            agent_val = safe_float(v)
                            break
                    if agent_val is None:
                        agent_val = safe_float(agent_data.get(metric_name))
                    check(f"Internal Metrics: {metric_name} value matches",
                          values_close(agent_val, gt_val),
                          f"agent={agent_val}, gt={gt_val}")

        # Sheet 2: Gap Analysis
        check("Gap Analysis sheet exists", "Gap Analysis" in wb.sheetnames,
              f"sheets: {wb.sheetnames}")
        if "Gap Analysis" in wb.sheetnames:
            ws = wb["Gap Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Gap Analysis has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            for expected_col in ['metric', 'internal_value', 'industry_avg', 'classification', 'priority']:
                check(f"Gap Analysis has '{expected_col}' column",
                      expected_col in headers, f"headers: {headers}")

            # Check classifications
            if gt_wb and "Gap Analysis" in gt_wb.sheetnames:
                gt_ws = gt_wb["Gap Analysis"]
                gt_class = {}
                for r in gt_ws.iter_rows(min_row=2, values_only=True):
                    if r and r[0]:
                        gt_class[str(r[0]).strip()] = str(r[7]).strip() if r[7] else ""

                agent_class = {}
                class_col_idx = None
                for i, h in enumerate(headers):
                    if 'classification' in h:
                        class_col_idx = i
                        break
                if class_col_idx is not None:
                    for r in data_rows:
                        if r and r[0]:
                            agent_class[str(r[0]).strip()] = str(r[class_col_idx]).strip() if r[class_col_idx] else ""

                for metric_name in ['Avg_Order_Value', 'Avg_Salary', 'Avg_Satisfaction', 'Revenue_Per_Employee']:
                    gt_c = gt_class.get(metric_name, "")
                    agent_c = None
                    for k, v in agent_class.items():
                        if metric_name.lower().replace('_', '') in k.lower().replace('_', '').replace(' ', ''):
                            agent_c = v
                            break
                    if agent_c is None:
                        agent_c = agent_class.get(metric_name, "")
                    check(f"Gap Analysis: {metric_name} classification matches",
                          gt_c.lower() == str(agent_c).lower(),
                          f"agent='{agent_c}', gt='{gt_c}'")

        # Sheet 3: Action Plan
        check("Action Plan sheet exists", "Action Plan" in wb.sheetnames,
              f"sheets: {wb.sheetnames}")
        if "Action Plan" in wb.sheetnames:
            ws = wb["Action Plan"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Action Plan has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            for expected_col in ['metric', 'classification', 'recommended_action']:
                check(f"Action Plan has '{expected_col}' column",
                      expected_col in headers, f"headers: {headers}")

    # --- Word Document Evaluation ---
    word_path = os.path.join(agent_workspace, "Benchmark_Report.docx")
    check("Benchmark_Report.docx exists", os.path.exists(word_path))

    if os.path.exists(word_path):
        from docx import Document
        doc = Document(word_path)
        full_text = "\n".join([p.text for p in doc.paragraphs])
        full_lower = full_text.lower()

        required_sections = [
            "Executive Summary", "Methodology", "Internal Performance Overview",
            "Benchmark Comparison", "Gap Analysis", "Strategic Recommendations",
            "Implementation Timeline"
        ]
        for section in required_sections:
            check(f"Word doc has '{section}' section",
                  section.lower() in full_lower,
                  "section not found")

        # Check key values mentioned
        check("Word doc mentions avg salary value",
              "58396" in full_text or "58,396" in full_text,
              "avg salary not found")
        check("Word doc mentions avg order value",
              "152.45" in full_text or "152" in full_text,
              "avg order value not found")
        check("Word doc mentions 'Critical' classification",
              "critical" in full_lower)
        check("Word doc mentions 'Moderate' classification",
              "moderate" in full_lower)

    # --- Email Evaluation ---
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Check for operations head email
        cur.execute("""SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s""", ('%operations_head%',))
        ops_emails = cur.fetchall()
        check("Email to operations_head@company.com sent", len(ops_emails) >= 1,
              f"found {len(ops_emails)}")

        # Check for HR director email
        cur.execute("""SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s""", ('%hr_director%',))
        hr_emails = cur.fetchall()
        check("Email to hr_director@company.com sent", len(hr_emails) >= 1,
              f"found {len(hr_emails)}")

        # Check for sales VP email
        cur.execute("""SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s""", ('%sales_vp%',))
        sales_emails = cur.fetchall()
        check("Email to sales_vp@company.com sent", len(sales_emails) >= 1,
              f"found {len(sales_emails)}")

        # Check email subjects
        cur.execute("""SELECT subject FROM email.messages
            WHERE subject ILIKE %s""", ('%benchmark%',))
        benchmark_emails = cur.fetchall()
        check("At least 3 benchmark-related emails sent", len(benchmark_emails) >= 3,
              f"found {len(benchmark_emails)}")

        conn.close()
    except Exception as e:
        check("Email verification", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

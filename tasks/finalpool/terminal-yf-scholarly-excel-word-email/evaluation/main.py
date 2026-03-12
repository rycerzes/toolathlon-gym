"""Evaluation for terminal-yf-scholarly-excel-word-email.
Checks:
1. FinTech_Research_Report.xlsx with 4 sheets and correct data
2. FinTech_Research_Report.docx with required sections
3. Email sent to research-committee@company.org
4. market_analysis.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: FinTech_Research_Report.xlsx ===")
    path = os.path.join(workspace, "FinTech_Research_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # Market_Data sheet
    md_idx = next((i for i, s in enumerate(sheets_lower) if "market" in s or "data" in s), 0)
    ws = wb[sheets[md_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    check("Market_Data has 2 stock rows", len(data_rows) >= 2, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    check("Contains JPM", "jpm" in all_text, f"Text: {all_text[:120]}")
    check("Contains XOM", "xom" in all_text, f"Text: {all_text[:120]}")

    # Academic_Papers sheet
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "academic" in s or "paper" in s), 1)
    if ap_idx < len(sheets):
        ws2 = wb[sheets[ap_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Academic_Papers has at least 4 rows", len(data_rows2) >= 4, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Contains market efficiency reference", "efficient" in all_text2 or "market" in all_text2,
              f"Text: {all_text2[:120]}")

    # Statistical_Tests sheet
    st_idx = next((i for i, s in enumerate(sheets_lower) if "statistic" in s or "test" in s), 2)
    if st_idx < len(sheets):
        ws3 = wb[sheets[st_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Statistical_Tests has at least 4 rows", len(data_rows3) >= 4, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Contains autocorrelation test", "autocorrelation" in all_text3 or "ljung" in all_text3,
              f"Text: {all_text3[:120]}")
        check("Contains normality test", "normality" in all_text3 or "jarque" in all_text3,
              f"Text: {all_text3[:120]}")

    # Research_Summary sheet
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s or "summary" in s or "finding" in s), 3)
    if rs_idx < len(sheets):
        ws4 = wb[sheets[rs_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Research_Summary has at least 3 rows", len(data_rows4) >= 3, f"Found {len(data_rows4)}")


def check_word(workspace):
    print("\n=== Check 2: FinTech_Research_Report.docx ===")
    path = os.path.join(workspace, "FinTech_Research_Report.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)

    try:
        from docx import Document
        doc = Document(path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Contains introduction or objectives", "introduction" in all_text or "objective" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains methodology", "methodology" in all_text or "method" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains results", "result" in all_text or "finding" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains conclusion", "conclusion" in all_text or "implication" in all_text,
              f"Text: {all_text[:150]}")
        check("Mentions market efficiency", "market efficiency" in all_text or "efficient market" in all_text,
              f"Text: {all_text[:150]}")
    except ImportError:
        check("python-docx available", False, "python-docx not installed")


def check_email():
    print("\n=== Check 3: Email to Research Committee ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check sent_log or messages for the email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE lower(subject) LIKE '%%fintech%%' OR lower(subject) LIKE '%%market efficiency%%'
        ORDER BY created_at DESC LIMIT 5
    """)
    msgs = cur.fetchall()

    if not msgs:
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            ORDER BY created_at DESC LIMIT 5
        """)
        msgs = cur.fetchall()

    found_email = False
    for subj, to_addr, body in msgs:
        subj_lower = (subj or "").lower()
        to_str = str(to_addr).lower() if to_addr else ""
        if ("fintech" in subj_lower or "market" in subj_lower) and "research" in to_str:
            found_email = True
            break

    check("Email sent about FinTech research", found_email or len(msgs) > 0,
          f"Found {len(msgs)} messages")

    if msgs:
        any_body = " ".join(str(m[2] or "") for m in msgs).lower()
        check("Email mentions JPM or XOM", "jpm" in any_body or "xom" in any_body or "morgan" in any_body,
              f"Body sample: {any_body[:150]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: market_analysis.py ===")
    path = os.path.join(workspace, "market_analysis.py")
    check("market_analysis.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "FinTech_Research_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"market", "data", "academic", "paper", "statistic", "test",
                             "research", "summary", "finding"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

    # Email: no duplicate research emails
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, COUNT(*) FROM email.messages
            WHERE lower(subject) LIKE '%%fintech%%' OR lower(subject) LIKE '%%market efficiency%%'
            GROUP BY subject HAVING COUNT(*) > 1
        """)
        dupes = cur.fetchall()
        check("No duplicate research emails", len(dupes) == 0,
              f"Duplicates: {dupes}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

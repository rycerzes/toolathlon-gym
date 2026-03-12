#!/usr/bin/env python3
"""
Evaluation script for ecommerce-market-benchmark task.

Queries PostgreSQL directly to compute expected values, then compares
against the agent's Excel output.
"""

import sys
import os
from argparse import ArgumentParser
from pathlib import Path

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "database": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def compute_expected_monthly_revenue():
    """Compute expected monthly revenue from completed WooCommerce orders."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT to_char(date_created, 'YYYY-MM') as month,
               COUNT(*) as orders,
               SUM(total::numeric) as revenue
        FROM wc.orders
        WHERE status = 'completed'
        GROUP BY to_char(date_created, 'YYYY-MM')
        ORDER BY month
    """)
    rows = cur.fetchall()
    conn.close()
    result = []
    for month, orders, revenue in rows:
        result.append({
            "month": month,
            "orders": int(orders),
            "revenue": round(float(revenue), 2),
        })
    return result


def compute_expected_market_indicators():
    """Compute expected monthly avg closes for ^DJI and GC=F from 2025-03 to 2026-02."""
    conn = get_db_connection()
    cur = conn.cursor()

    indicators = {}
    for symbol in ["^DJI", "GC=F"]:
        cur.execute("""
            SELECT to_char(date, 'YYYY-MM') as month, AVG(close) as avg_close
            FROM yf.stock_prices
            WHERE symbol = %s AND date >= '2025-03-01' AND date < '2026-03-01'
            GROUP BY to_char(date, 'YYYY-MM')
            ORDER BY month
        """, (symbol,))
        indicators[symbol] = {row[0]: round(float(row[1]), 2) for row in cur.fetchall()}

    conn.close()
    return indicators


def read_sheet_rows(wb, sheet_name):
    """Read a sheet into a list of dicts using first row as headers."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return data


def read_summary_sheet(wb, sheet_name="Summary"):
    """Read key-value layout from Summary sheet (col A = key, col B = value)."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    kv = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            kv[str(row[0]).strip()] = row[1]
    return kv


def check_monthly_revenue(wb, expected):
    """Check the 'Monthly Revenue' sheet."""
    data = read_sheet_rows(wb, "Monthly Revenue")
    if data is None:
        return False, "Sheet 'Monthly Revenue' not found"
    if len(data) != 12:
        return False, f"Expected 12 rows in 'Monthly Revenue', found {len(data)}"

    errors = []
    for i, exp in enumerate(expected):
        if i >= len(data):
            errors.append(f"Missing row for month {exp['month']}")
            continue
        row = data[i]

        # Check month
        agent_month = str(row.get("Month", "")).strip()
        if agent_month != exp["month"]:
            errors.append(f"Row {i+1}: Expected month '{exp['month']}', found '{agent_month}'")

        # Check orders
        agent_orders = safe_float(row.get("Orders"))
        if agent_orders is None:
            errors.append(f"Row {i+1}: Orders is missing or not numeric")
        elif abs(int(agent_orders) - exp["orders"]) > 1:
            errors.append(
                f"Row {i+1} ({exp['month']}): Orders mismatch: agent={int(agent_orders)}, expected={exp['orders']}"
            )

        # Check revenue
        agent_rev = safe_float(row.get("Revenue"))
        if agent_rev is None:
            errors.append(f"Row {i+1}: Revenue is missing or not numeric")
        elif abs(agent_rev - exp["revenue"]) > 5.0:
            errors.append(
                f"Row {i+1} ({exp['month']}): Revenue mismatch: agent={agent_rev}, expected={exp['revenue']}"
            )

    if errors:
        return False, "Monthly Revenue errors:\n  " + "\n  ".join(errors)
    return True, f"Monthly Revenue: all 12 months verified"


def check_market_indicators(wb, indicators):
    """Check the 'Market Indicators' sheet."""
    data = read_sheet_rows(wb, "Market Indicators")
    if data is None:
        return False, "Sheet 'Market Indicators' not found"
    if len(data) != 12:
        return False, f"Expected 12 rows in 'Market Indicators', found {len(data)}"

    dji_data = indicators["^DJI"]
    gold_data = indicators["GC=F"]
    expected_months = sorted(dji_data.keys())

    errors = []
    for i, exp_month in enumerate(expected_months):
        if i >= len(data):
            errors.append(f"Missing row for month {exp_month}")
            continue
        row = data[i]

        agent_month = str(row.get("Month", "")).strip()
        if agent_month != exp_month:
            errors.append(f"Row {i+1}: Expected month '{exp_month}', found '{agent_month}'")

        # DJI
        agent_dji = safe_float(row.get("DJI_Avg_Close"))
        exp_dji = dji_data.get(exp_month)
        if agent_dji is None:
            errors.append(f"Row {i+1}: DJI_Avg_Close is missing or not numeric")
        elif exp_dji is not None and abs(agent_dji - exp_dji) > 50.0:
            errors.append(
                f"Row {i+1} ({exp_month}): DJI_Avg_Close mismatch: agent={agent_dji}, expected={exp_dji}"
            )

        # Gold
        agent_gold = safe_float(row.get("Gold_Avg_Close"))
        exp_gold = gold_data.get(exp_month)
        if agent_gold is None:
            errors.append(f"Row {i+1}: Gold_Avg_Close is missing or not numeric")
        elif exp_gold is not None and abs(agent_gold - exp_gold) > 5.0:
            errors.append(
                f"Row {i+1} ({exp_month}): Gold_Avg_Close mismatch: agent={agent_gold}, expected={exp_gold}"
            )

    if errors:
        return False, "Market Indicators errors:\n  " + "\n  ".join(errors)
    return True, f"Market Indicators: all 12 months verified"


def check_summary(wb, expected_revenue, indicators):
    """Check the 'Summary' sheet."""
    kv = read_summary_sheet(wb, "Summary")
    if kv is None:
        return False, "Sheet 'Summary' not found"

    errors = []

    # Compute expected summary values
    total_revenue = round(sum(m["revenue"] for m in expected_revenue), 2)
    total_orders = sum(m["orders"] for m in expected_revenue)
    best_month = max(expected_revenue, key=lambda m: m["revenue"])["month"]
    worst_month = min(expected_revenue, key=lambda m: m["revenue"])["month"]

    dji_months = sorted(indicators["^DJI"].keys())
    dji_first = indicators["^DJI"][dji_months[0]]
    dji_last = indicators["^DJI"][dji_months[-1]]
    dji_trend = "Up" if dji_last > dji_first else "Down"

    gold_first = indicators["GC=F"][dji_months[0]]
    gold_last = indicators["GC=F"][dji_months[-1]]
    gold_trend = "Up" if gold_last > gold_first else "Down"

    # Check Total_Revenue
    agent_total_rev = safe_float(kv.get("Total_Revenue"))
    if agent_total_rev is None:
        errors.append("Total_Revenue is missing or not numeric")
    elif abs(agent_total_rev - total_revenue) > 10.0:
        errors.append(f"Total_Revenue mismatch: agent={agent_total_rev}, expected={total_revenue}")

    # Check Total_Orders
    agent_total_orders = safe_float(kv.get("Total_Orders"))
    if agent_total_orders is None:
        errors.append("Total_Orders is missing or not numeric")
    elif abs(int(agent_total_orders) - total_orders) > 1:
        errors.append(f"Total_Orders mismatch: agent={int(agent_total_orders)}, expected={total_orders}")

    # Check Best_Month
    agent_best = str(kv.get("Best_Month", "")).strip()
    if agent_best.lower() != best_month.lower():
        errors.append(f"Best_Month mismatch: agent='{agent_best}', expected='{best_month}'")

    # Check Worst_Month
    agent_worst = str(kv.get("Worst_Month", "")).strip()
    if agent_worst.lower() != worst_month.lower():
        errors.append(f"Worst_Month mismatch: agent='{agent_worst}', expected='{worst_month}'")

    # Check DJI_Trend
    agent_dji_trend = str(kv.get("DJI_Trend", "")).strip()
    if agent_dji_trend.lower() != dji_trend.lower():
        errors.append(f"DJI_Trend mismatch: agent='{agent_dji_trend}', expected='{dji_trend}'")

    # Check Gold_Trend
    agent_gold_trend = str(kv.get("Gold_Trend", "")).strip()
    if agent_gold_trend.lower() != gold_trend.lower():
        errors.append(f"Gold_Trend mismatch: agent='{agent_gold_trend}', expected='{gold_trend}'")

    if errors:
        return False, "Summary errors:\n  " + "\n  ".join(errors)
    return True, f"Summary: all 6 metrics verified"


def run_evaluation(agent_workspace):
    print("=" * 70)
    print("ECOMMERCE MARKET BENCHMARK - EVALUATION")
    print("=" * 70)

    # Compute expected values from DB
    print("\nComputing expected values from database...")
    try:
        expected_revenue = compute_expected_monthly_revenue()
        print(f"  Monthly revenue: {len(expected_revenue)} months")
    except Exception as e:
        return False, f"Failed to compute expected monthly revenue: {e}"

    try:
        indicators = compute_expected_market_indicators()
        print(f"  Market indicators: {len(indicators['^DJI'])} months for DJI, {len(indicators['GC=F'])} months for Gold")
    except Exception as e:
        return False, f"Failed to compute expected market indicators: {e}"

    # Find Excel file
    ws = Path(agent_workspace)
    excel_path = ws / "Market_Benchmark_Report.xlsx"
    if not excel_path.exists():
        for f in ws.iterdir():
            if f.name.lower() == "market_benchmark_report.xlsx":
                excel_path = f
                break
    if not excel_path.exists():
        return False, f"Excel file 'Market_Benchmark_Report.xlsx' not found in {agent_workspace}"

    print(f"\nLoading Excel file: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"  Sheets found: {wb.sheetnames}")
    except Exception as e:
        return False, f"Failed to open Excel file: {e}"

    results = []

    print("\n--- Check 1: Monthly Revenue ---")
    passed, msg = check_monthly_revenue(wb, expected_revenue)
    results.append(("Monthly Revenue", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    print("\n--- Check 2: Market Indicators ---")
    passed, msg = check_market_indicators(wb, indicators)
    results.append(("Market Indicators", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    print("\n--- Check 3: Summary ---")
    passed, msg = check_summary(wb, expected_revenue, indicators)
    results.append(("Summary", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    wb.close()

    total = len(results)
    passed_count = sum(1 for _, p, _ in results if p)
    all_passed = passed_count == total

    summary_lines = ["\n" + "=" * 70, "EVALUATION SUMMARY", "=" * 70]
    for name, p, msg in results:
        status = "PASSED" if p else "FAILED"
        summary_lines.append(f"  [{status}] {name}: {msg}")
    summary_lines.append(f"\nOverall: {passed_count}/{total} checks passed")
    if all_passed:
        summary_lines.append("ALL CHECKS PASSED")
    else:
        summary_lines.append("SOME CHECKS FAILED")

    return all_passed, "\n".join(summary_lines)


def main():
    parser = ArgumentParser(description="Evaluate Ecommerce Market Benchmark Task")
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    print(f"Agent workspace: {args.agent_workspace}")

    try:
        success, message = run_evaluation(args.agent_workspace)
        print(message)

        if success:
            print("\nEVALUATION PASSED")
            sys.exit(0)
        else:
            print("\nEVALUATION FAILED")
            sys.exit(1)
    except Exception as e:
        print(f"Critical evaluation error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

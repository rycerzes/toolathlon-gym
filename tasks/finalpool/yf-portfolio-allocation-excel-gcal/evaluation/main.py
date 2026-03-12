"""Evaluation for yf-portfolio-allocation-excel-gcal."""
import os
import argparse, os, sys
import psycopg2


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_pdf(agent_workspace):
    errors = []
    pdf_path = os.path.join(agent_workspace, "Portfolio_Guidelines.pdf")
    if not os.path.exists(pdf_path):
        errors.append("Portfolio_Guidelines.pdf not found in agent workspace")
    return errors


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Portfolio_Allocation.xlsx")
    if not os.path.exists(path):
        return ["Portfolio_Allocation.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        # Check Stock Analysis sheet
        rows = load_sheet_rows(wb, "Stock Analysis")
        if rows is None:
            errors.append("Sheet 'Stock Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 5:
                errors.append(f"Stock Analysis has {len(data_rows)} rows, expected 5")
            # Check GOOGL weight
            googl_rows = [r for r in data_rows if r[0] and str(r[0]).strip().upper() == "GOOGL"]
            if not googl_rows:
                errors.append("GOOGL row not found in Stock Analysis")
            else:
                weight = googl_rows[0][4] if len(googl_rows[0]) > 4 else None
                if not num_close(weight, 25.0, 1.0):
                    errors.append(f"GOOGL Allocated_Weight_Pct={weight}, expected 25.0")
            # Check symbols
            symbols = {str(r[0]).strip().upper() for r in data_rows}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols:
                    errors.append(f"Symbol {sym} missing from Stock Analysis")

        # Check Allocation Summary sheet
        rows2 = load_sheet_rows(wb, "Allocation Summary")
        if rows2 is None:
            errors.append("Sheet 'Allocation Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows2 if r[0]}
            if "total_invested" in lookup:
                if not num_close(lookup["total_invested"], 100000, 1000):
                    errors.append(f"Total_Invested={lookup['total_invested']}, expected 100000")
            else:
                errors.append("Total_Invested not found in Allocation Summary")
            if "strong_buy_count" in lookup:
                if not num_close(lookup["strong_buy_count"], 2, 0):
                    errors.append(f"Strong_Buy_Count={lookup['strong_buy_count']}, expected 2")
            else:
                errors.append("Strong_Buy_Count not found in Allocation Summary")
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE start_datetime::date = '2026-04-30'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No GCal event found on 2026-04-30")
        else:
            summaries = [r[0].lower() if r[0] else "" for r in rows]
            if not any("portfolio" in s or "rebalancing" in s or "rebalance" in s for s in summaries):
                errors.append(f"No portfolio rebalancing event on 2026-04-30 (found: {[r[0] for r in rows]})")
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE to_addr::text ILIKE '%portfolio_manager@wealth.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No email found to portfolio_manager@wealth.com")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking PDF in workspace...")
    errs = check_pdf(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal event...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

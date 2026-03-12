"""
Evaluation for yt-ml-repos-github-notion-excel-email task.

Checks:
1. ML_Research_Tracker.xlsx exists
2. "Videos" sheet has >= 5 data rows, has Video_ID and Title columns
3. "Papers" sheet has >= 3 data rows, has ArXiv_ID and Title columns
4. "Summary" sheet has >= 3 rows
5. Notion page exists with "ML Tech" or "Research" in title
6. Email sent to research@lab.edu
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-4: ML_Research_Tracker.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "ML_Research_Tracker.xlsx")
    if not os.path.exists(xlsx_path):
        record("ML_Research_Tracker.xlsx exists", False, f"Not found at {xlsx_path}")
        for msg in ["Videos sheet has >= 5 rows with Video_ID and Title",
                    "Papers sheet has >= 3 rows with ArXiv_ID and Title",
                    "Summary sheet has >= 3 rows"]:
            record(msg, False, "File missing")
        return
    record("ML_Research_Tracker.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Videos sheet
    videos_key = next((sheet_names_lower[k] for k in sheet_names_lower if "video" in k), None)
    if not videos_key:
        record("Videos sheet has >= 5 rows with Video_ID and Title", False,
               f"No Videos sheet. Sheets: {wb.sheetnames}")
    else:
        ws = wb[videos_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        headers = [str(c).strip().lower() if c else "" for c in rows[0]] if rows else []
        has_video_id = any("video_id" in h or "videoid" in h or "video id" in h for h in headers)
        has_title = any("title" in h for h in headers)
        record("Videos sheet has >= 5 rows with Video_ID and Title",
               len(data_rows) >= 5 and has_video_id and has_title,
               f"Rows: {len(data_rows)}, Headers: {rows[0] if rows else []}")

    # Papers sheet
    papers_key = next((sheet_names_lower[k] for k in sheet_names_lower if "paper" in k), None)
    if not papers_key:
        record("Papers sheet has >= 3 rows with ArXiv_ID and Title", False,
               f"No Papers sheet. Sheets: {wb.sheetnames}")
    else:
        ws2 = wb[papers_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        headers2 = [str(c).strip().lower() if c else "" for c in rows2[0]] if rows2 else []
        has_arxiv_id = any("arxiv" in h or "id" in h for h in headers2)
        has_title2 = any("title" in h for h in headers2)
        record("Papers sheet has >= 3 rows with ArXiv_ID and Title",
               len(data_rows2) >= 3 and has_arxiv_id and has_title2,
               f"Rows: {len(data_rows2)}, Headers: {rows2[0] if rows2 else []}")

    # Summary sheet
    summary_key = next((sheet_names_lower[k] for k in sheet_names_lower if "summar" in k), None)
    if not summary_key:
        record("Summary sheet has >= 3 rows", False, f"No Summary sheet. Sheets: {wb.sheetnames}")
    else:
        ws3 = wb[summary_key]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)] if rows3 else []
        record("Summary sheet has >= 3 rows", len(data_rows3) >= 3,
               f"Found {len(data_rows3)} data rows")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "ML_Research_Tracker.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 5: Notion page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE properties->>'title' ILIKE '%ML Tech%'
                   OR properties->>'title' ILIKE '%Research Hub%'
                   OR properties->>'title' ILIKE '%ML Research%'
                   OR properties->>'title' ILIKE '%Research%'
            """)
            rows = cur.fetchall()
        conn.close()
        record("Notion page with 'ML Tech' or 'Research' in title exists",
               len(rows) > 0, f"Found {len(rows)} pages")
    except Exception as e:
        record("Notion check", False, str(e))


def check_email():
    print("\n=== Check 6: Email sent to research@lab.edu ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM email.messages
                WHERE to_addr::text ILIKE '%research@lab.edu%'
                  AND folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
            """)
            count = cur.fetchone()[0]
            if count == 0:
                # Also check any folder (agent might store in Sent or another way)
                cur.execute("""
                    SELECT COUNT(*) FROM email.messages
                    WHERE to_addr::text ILIKE '%research@lab.edu%'
                      AND from_addr != 'pi@lab.edu'
                """)
                count = cur.fetchone()[0]
            if count == 0:
                try:
                    cur.execute("""
                        SELECT COUNT(*) FROM email.sent_log
                        WHERE to_addr ILIKE '%research@lab.edu%'
                    """)
                    count = cur.fetchone()[0]
                except Exception:
                    pass
        conn.close()
        record("Email sent to research@lab.edu", count > 0, f"Found {count}")
    except Exception as e:
        record("Email check", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-ml-repos-github-notion-excel-email")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_notion()
    check_email()

    all_passed = FAIL_COUNT == 0
    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print(f"\n{'='*40}")
    print(f"Result: {'PASS' if all_passed else 'FAIL'} - {summary}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "all_passed": all_passed}, f)

    return all_passed, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

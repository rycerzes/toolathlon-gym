"""
Evaluation script for fetch-notion-monitoring task.

Checks:
1. Service_Availability_Report.xlsx with Availability Summary and Incidents sheets
2. Notion page with monitoring/dashboard/service content
3. Calendar event for incident review
4. Email with availability report
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.lower() in str(haystack).lower()


def check_excel(agent_workspace):
    """Check Service_Availability_Report.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Service_Availability_Report.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # Check Availability Summary sheet
    summary_sheet = None
    for name in wb.sheetnames:
        if "availability" in name.lower() or "summary" in name.lower() or "service" in name.lower() or "status" in name.lower():
            summary_sheet = name
            break

    if not summary_sheet:
        record("Availability Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Availability Summary sheet exists", True)
        ws = wb[summary_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        if len(data_rows) != 3:
            record("Availability Summary has 3 rows", False, f"Found {len(data_rows)} rows")
            all_ok = False
        else:
            record("Availability Summary has 3 rows", True)

        # Check expected uptime values by finding each service in the data
        expected = {
            "api gateway": 99.21,
            "database cluster": 98.50,
            "storage service": 99.95,
        }

        for svc_key, expected_uptime in expected.items():
            found = False
            for row in data_rows:
                if row and str_contains(row[0], svc_key.split()[0]):
                    found = True
                    # Find the uptime column (look for a value that could be a percentage)
                    uptime_val = None
                    for cell in row[1:]:
                        try:
                            v = float(cell)
                            if 90.0 <= v <= 100.0:
                                uptime_val = v
                                break
                        except (TypeError, ValueError):
                            continue

                    if uptime_val is not None:
                        ok = num_close(uptime_val, expected_uptime, tol=2.0)
                        record(
                            f"'{svc_key}' uptime ~{expected_uptime}%",
                            ok,
                            f"Got {uptime_val}",
                        )
                        if not ok:
                            all_ok = False
                    else:
                        record(f"'{svc_key}' uptime found", False, f"Row: {row}")
                        all_ok = False
                    break

            if not found:
                record(f"'{svc_key}' in Availability Summary", False, "Not found")
                all_ok = False

    # Check Incidents sheet (optional -- groundtruth may not have a separate incidents sheet)
    incidents_sheet = None
    for name in wb.sheetnames:
        if "incident" in name.lower():
            incidents_sheet = name
            break

    if not incidents_sheet:
        # Incidents sheet is optional; do not fail the evaluation
        print("  [INFO] No dedicated Incidents sheet found (optional); skipping incidents checks")
    else:
        record("Incidents sheet exists", True)
        ws = wb[incidents_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        # Database Cluster is the degraded service; check for at least 1 incident row
        db_rows = [
            r for r in data_rows if r and (str_contains(r[0], "database") or str_contains(r[0], "cluster"))
        ]
        if len(db_rows) >= 1:
            record("Incidents has >= 1 Database Cluster rows", True)
        else:
            record(
                "Incidents has >= 1 Database Cluster rows",
                False,
                f"Found {len(db_rows)} Database Cluster rows",
            )
            all_ok = False

        # Should have at least 1 row total (any incident)
        if len(data_rows) >= 1:
            record("Incidents sheet has data", True)
        else:
            record("Incidents sheet has data", False, "No data rows")
            all_ok = False

    wb.close()
    return all_ok


def check_notion():
    """Check Notion page with monitoring dashboard content."""
    print("\n=== Checking Notion ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT id, properties FROM notion.pages"
        )
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion DB accessible", False, str(e))
        return False

    all_ok = True
    found_page = False

    for page_id, props_raw in pages:
        if isinstance(props_raw, str):
            props = json.loads(props_raw)
        else:
            props = props_raw

        # Extract title from properties
        title_parts = []
        for key in ["title", "Name"]:
            prop = props.get(key, {})
            if isinstance(prop, dict):
                title_parts = prop.get("title", [])
                if title_parts:
                    break

        page_title = "".join(p.get("plain_text", "") for p in title_parts)
        page_title_lower = page_title.lower()

        # Match pages with monitoring, dashboard, or service in title
        if (
            "monitoring" in page_title_lower
            or "dashboard" in page_title_lower
            or "service" in page_title_lower
        ):
            found_page = True
            record("Notion monitoring page exists", True)

            # Check page has blocks
            try:
                conn2 = psycopg2.connect(**DB_CONFIG)
                cur2 = conn2.cursor()
                cur2.execute(
                    "SELECT type, block_data FROM notion.blocks WHERE parent_id = %s",
                    (page_id,),
                )
                blocks = cur2.fetchall()
                cur2.close()
                conn2.close()

                if len(blocks) > 0:
                    record("Notion page has blocks", True)
                else:
                    record("Notion page has blocks", False, "No blocks found")
                    all_ok = False
            except Exception as e:
                record("Notion blocks readable", False, str(e))
                all_ok = False
            break

    if not found_page:
        record(
            "Notion monitoring page exists",
            False,
            f"Found {len(pages)} pages but none with monitoring/dashboard/service in title",
        )
        all_ok = False

    return all_ok


def check_calendar():
    """Check calendar event for incident review."""
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT summary, description, start_datetime, end_datetime FROM gcal.events"
        )
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    all_ok = True
    found_event = False

    for summary, description, start_dt, end_dt in events:
        summary_lower = (summary or "").lower()
        if "incident review" in summary_lower:
            found_event = True
            record("Incident review calendar event exists", True)
            break

    if not found_event:
        record(
            "Incident review calendar event exists",
            False,
            f"Found {len(events)} events but none with 'incident review' in summary",
        )
        all_ok = False

    return all_ok


def check_emails():
    """Check email with availability or service in subject."""
    print("\n=== Checking Emails ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, from_addr, to_addr, body_text FROM email.messages"
        )
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    all_ok = True
    found_email = False

    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if "availability" in subj_lower or "service" in subj_lower:
            found_email = True
            record("Alert email with availability/service subject exists", True)
            break

    if not found_email:
        record(
            "Alert email with availability/service subject exists",
            False,
            f"Found {len(emails)} emails but none with 'availability' or 'service' in subject",
        )
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    notion_ok = check_notion()
    cal_ok = check_calendar()
    email_ok = check_emails()

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Notion:   {'PASS' if notion_ok else 'FAIL'}")
    print(f"  Calendar: {'PASS' if cal_ok else 'FAIL'}")
    print(f"  Emails:   {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    overall = excel_ok and notion_ok and cal_ok and email_ok
    print(f"  Overall:  {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

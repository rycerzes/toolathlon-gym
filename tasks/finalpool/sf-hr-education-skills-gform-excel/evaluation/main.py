"""Evaluation for sf-hr-education-skills-gform-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "HR_Education_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Education_Analysis.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("HR_Education_Analysis.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Education Breakdown sheet
        print("  Checking Education Breakdown sheet...")
        a_rows = load_sheet_rows(agent_wb, "Education Breakdown")
        g_rows = load_sheet_rows(gt_wb, "Education Breakdown")
        if a_rows is None:
            all_errors.append("Sheet 'Education Breakdown' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []
            if len(a_data) != 35:
                all_errors.append(f"Education Breakdown row count: {len(a_data)}, expected 35")

            # Build lookup: (dept, edu_level) -> row
            a_lookup = {}
            for row in a_data:
                if row and row[0] and row[1]:
                    key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                    a_lookup[key] = row

            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing row: {g_row[0]} / {g_row[1]}")
                    continue
                # Check Employee_Count
                if not num_close(a_row[2], g_row[2], 1):
                    all_errors.append(f"{g_row[0]}/{g_row[1]} Employee_Count: {a_row[2]} vs {g_row[2]}")
                # Check Avg_Performance
                if not num_close(a_row[3], g_row[3], 0.05):
                    all_errors.append(f"{g_row[0]}/{g_row[1]} Avg_Performance: {a_row[3]} vs {g_row[3]}")
            print("    Done.")

        # Check Department Summary sheet
        print("  Checking Department Summary sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Department Summary")
        g_rows2 = load_sheet_rows(gt_wb, "Department Summary")
        if a_rows2 is None:
            all_errors.append("Sheet 'Department Summary' not found in agent output")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            g_data2 = g_rows2[1:] if len(g_rows2) > 1 else []
            if len(a_data2) != 7:
                all_errors.append(f"Department Summary row count: {len(a_data2)}, expected 7")

            a_lookup2 = {}
            for row in a_data2:
                if row and row[0]:
                    a_lookup2[str(row[0]).strip().lower()] = row

            for g_row in g_data2:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup2.get(key)
                if a_row is None:
                    all_errors.append(f"Department Summary missing dept: {g_row[0]}")
                    continue
                # Total_Employees
                if not num_close(a_row[1], g_row[1], 1):
                    all_errors.append(f"{g_row[0]} Total_Employees: {a_row[1]} vs {g_row[1]}")
                # Higher_Ed_Pct (index 6)
                if len(a_row) > 6 and len(g_row) > 6:
                    if not num_close(a_row[6], g_row[6], 0.5):
                        all_errors.append(f"{g_row[0]} Higher_Ed_Pct: {a_row[6]} vs {g_row[6]}")
            print("    Done.")

    # --- Check 2: GForm exists ---
    print("Checking Google Form...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gform.forms WHERE LOWER(title) LIKE '%training interest%'")
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        if count == 0:
            all_errors.append("Google Form 'Training Interest Survey' not found in gform.forms")
        else:
            print(f"    GForm found ({count} matching forms)")
    except Exception as e:
        all_errors.append(f"Error checking GForm: {e}")

    # --- Check 3: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%training@company.com%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        if count == 0:
            all_errors.append("No email sent to training@company.com")
        else:
            print(f"    Email found ({count} messages)")
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

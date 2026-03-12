"""Evaluation for fetch-sf-hr-training-notion-gcal-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Training_Budget_Plan.xlsx")
    gt_file = os.path.join(gt_dir, "Training_Budget_Plan.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Training_Budget_Plan.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Department Analysis sheet
        print("  Checking Department Analysis sheet...")
        a_rows = load_sheet_rows(agent_wb, "Department Analysis")
        g_rows = load_sheet_rows(gt_wb, "Department Analysis")
        if a_rows is None:
            all_errors.append("Sheet 'Department Analysis' not found")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            if len(a_data) < 7:
                all_errors.append(f"Department Analysis: {len(a_data)} rows, expected 7")
            else:
                a_lookup = {}
                for row in a_data:
                    if row and row[0]:
                        a_lookup[str(row[0]).strip().lower()] = row
                for g_row in g_rows[1:]:
                    if not g_row or not g_row[0]:
                        continue
                    key = str(g_row[0]).strip().lower()
                    a_row = a_lookup.get(key)
                    if a_row is None:
                        all_errors.append(f"Missing department: {g_row[0]}")
                        continue
                    # Total_Employees
                    if not num_close(a_row[1], g_row[1], 5):
                        all_errors.append(f"{g_row[0]} Total: {a_row[1]} vs {g_row[1]}")
                    # Avg_Performance
                    if not num_close(a_row[2], g_row[2], 0.05):
                        all_errors.append(f"{g_row[0]} Avg_Perf: {a_row[2]} vs {g_row[2]}")
            print("    Done.")

        # Check Training Catalog sheet
        print("  Checking Training Catalog sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Training Catalog")
        if a_rows2 is None:
            all_errors.append("Sheet 'Training Catalog' not found")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            if len(a_data2) < 12:
                all_errors.append(f"Training Catalog: {len(a_data2)} rows, expected 12")
            print("    Done.")

        # Check Budget Allocation sheet
        print("  Checking Budget Allocation sheet...")
        a_rows3 = load_sheet_rows(agent_wb, "Budget Allocation")
        if a_rows3 is None:
            all_errors.append("Sheet 'Budget Allocation' not found")
        else:
            a_data3 = [r for r in a_rows3[1:] if r and r[0] and str(r[0]).strip()]
            if len(a_data3) < 3:
                all_errors.append(f"Budget Allocation: {len(a_data3)} department rows, expected 3")
            else:
                # Check that Operations and Sales are included (top 2 lowest perf)
                dept_names = [str(r[0]).strip().lower() for r in a_data3]
                if "operations" not in dept_names:
                    all_errors.append("Budget Allocation missing Operations (lowest avg perf)")
                if "sales" not in dept_names:
                    all_errors.append("Budget Allocation missing Sales (2nd lowest avg perf)")
            print("    Done.")

    # --- Check 2: Notion database ---
    print("Checking Notion database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE title::text ILIKE '%training%program%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            all_errors.append("Notion database 'Training Program 2026' not found")
        else:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages p
                JOIN notion.databases d ON p.parent::jsonb->>'database_id' = d.id
                WHERE d.title::text ILIKE '%training%program%'
                AND p.archived = false
            """)
            page_count = cur.fetchone()[0]
            if page_count < 3:
                all_errors.append(f"Notion: found {page_count}/3 department training pages")
            else:
                print(f"    Found {page_count} training pages")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking Notion: {e}")

    # --- Check 3: GCal events ---
    print("Checking GCal events...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date
            FROM gcal.events
            WHERE summary ILIKE '%training%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        training_events = [r for r in rows if "training" in r[0].lower()]
        if len(training_events) < 3:
            all_errors.append(f"GCal: found {len(training_events)} training events, expected 3")
        else:
            dates = [str(r[1]) for r in training_events]
            for expected_date in ["2026-05-04", "2026-05-11", "2026-05-18"]:
                if expected_date not in dates:
                    all_errors.append(f"GCal: no training event on {expected_date}")
    except Exception as e:
        all_errors.append(f"Error checking GCal: {e}")

    # --- Check 4: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%department-managers@company.com%'
            AND subject ILIKE '%training%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            all_errors.append("No email sent to department-managers@company.com about training")
        else:
            print(f"    Email found ({count})")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

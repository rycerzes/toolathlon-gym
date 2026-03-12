"""
Evaluation for yt-transcript-notion-song-report-gcal-email task.

Checks:
1. Song_Analysis_Report.xlsx exists
2. Tracklist sheet has >= 8 rows with Song_Title and Artist columns
3. Artist_Stats sheet has >= 4 rows
4. Publication_Plan sheet has >= 3 rows with Publish_Date column
5. Notion page exists with Afrobeat or Mix or Analysis in title
6. Notion database exists with >= 8 song entries
7. GCal has >= 3 new publication events in March 2026
8. Email sent to editorial@musicblog.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Song_Analysis_Report.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Song_Analysis_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Song_Analysis_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Song_Analysis_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # Check Tracklist sheet
    tracklist_sheet = None
    for name in wb.sheetnames:
        if "track" in name.lower():
            tracklist_sheet = wb[name]
            break
    if tracklist_sheet is None:
        record("Tracklist sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Tracklist sheet exists", True)
        rows = list(tracklist_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_title = any("song" in h or "title" in h for h in headers)
        has_artist = any("artist" in h for h in headers)
        record("Tracklist has Song_Title and Artist columns", has_title and has_artist,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Tracklist has >= 8 data rows", len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")

    # Check Artist_Stats sheet
    artist_sheet = None
    for name in wb.sheetnames:
        if "artist" in name.lower() or "stat" in name.lower():
            artist_sheet = wb[name]
            break
    if artist_sheet is None:
        record("Artist_Stats sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Artist_Stats sheet exists", True)
        rows = list(artist_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Artist_Stats has >= 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} rows")

    # Check Publication_Plan sheet
    pub_sheet = None
    for name in wb.sheetnames:
        if "pub" in name.lower() or "plan" in name.lower():
            pub_sheet = wb[name]
            break
    if pub_sheet is None:
        record("Publication_Plan sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Publication_Plan sheet exists", True)
        rows = list(pub_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_date = any("date" in h or "publish" in h for h in headers)
        record("Publication_Plan has Publish_Date column", has_date,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Publication_Plan has >= 3 data rows", len(data_rows) >= 3,
               f"Found {len(data_rows)} rows")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Song_Analysis_Report.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion Pages and Database ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, parent, properties FROM notion.pages")
    pages = cur.fetchall()
    cur.close()
    conn.close()

    # Check for analysis page
    found_page = False
    db_id = None
    for page_id, parent, props in pages:
        try:
            # workspace-level page
            parent_type = parent.get("type", "") if isinstance(parent, dict) else ""
            if parent_type == "workspace":
                title_items = []
                for key, val in props.items():
                    if isinstance(val, dict) and val.get("type") == "title":
                        title_items = val.get("title", [])
                        break
                title_text = " ".join(
                    item.get("text", {}).get("content", "") for item in title_items
                    if isinstance(item, dict)
                ).lower()
                if "afrobeat" in title_text or "mix" in title_text or "analysis" in title_text:
                    found_page = True
            # database-level page (song entry)
            elif parent_type == "database_id":
                if db_id is None:
                    db_id = parent.get("database_id")
        except Exception:
            continue

    record("Notion page exists with Afrobeat/Mix/Analysis in title", found_page,
           f"Total pages found: {len(pages)}")

    # Count song entries (pages with database parent)
    db_entries = [p for p in pages if isinstance(p[1], dict) and p[1].get("type") == "database_id"]
    record("Notion database has >= 8 song entries", len(db_entries) >= 8,
           f"Found {len(db_entries)} database-parented pages")


def check_gcal():
    print("\n=== Check 3: GCal Publication Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-15' AND start_datetime < '2026-04-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    pub_events = [e for e in events
                  if any(kw in (e[0] or "").lower()
                         for kw in ["publish", "article", "genre", "afrobeat", "spotlight", "mix", "feature"])]
    record("GCal has >= 3 publication-related events in March 2026", len(pub_events) >= 3,
           f"Found {len(pub_events)} events: {[e[0] for e in pub_events]}")


def check_email_sent():
    print("\n=== Check 4: Email Sent to editorial@musicblog.com ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # Check sent messages in Sent/SENT folders or via sent_log join
        cur.execute("""
            SELECT count(*) FROM email.messages m
            JOIN email.folders f ON m.folder_id = f.id
            WHERE UPPER(f.name) IN ('SENT')
              AND m.to_addr::text ILIKE '%editorial@musicblog.com%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            # Fallback: check sent_log joined with messages
            cur.execute("""
                SELECT count(*) FROM email.sent_log sl
                JOIN email.messages m ON sl.message_id = m.id
                WHERE m.to_addr::text ILIKE '%editorial@musicblog.com%'
            """)
            count = cur.fetchone()[0]
        record("Email sent to editorial@musicblog.com", count >= 1, f"Sent count: {count}")
    except Exception as e:
        record("Email sent_log check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()
    check_email_sent()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""Evaluation for canvas-enrollment-forecast-excel-gform-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Enrollment_Analysis.xlsx")
    gt_path = os.path.join(gt_dir, "Enrollment_Analysis.xlsx")

    if not os.path.isfile(xlsx_path):
        check("Enrollment_Analysis.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Enrollment_Analysis.xlsx exists", True)

    if not os.path.isfile(gt_path):
        check("Groundtruth Excel exists", False, f"Not found: {gt_path}")
        return

    try:
        agent_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # Sheet 1: Enrollment Trends
    print("  --- Enrollment Trends ---")
    a_rows = load_sheet_rows(agent_wb, "Enrollment Trends")
    g_rows = load_sheet_rows(gt_wb, "Enrollment Trends")
    if a_rows is None:
        check("Sheet 'Enrollment Trends' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Enrollment Trends' exists in GT", False)
    else:
        check("Sheet 'Enrollment Trends' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Enrollment Trends has 22 data rows", len(a_data) == 22, f"Found {len(a_data)}")

        # Build lookup: (base_name_lower, semester_lower, year) -> row
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower(), int(row[2]) if row[2] else 0)
                a_lookup[key] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower(), int(g_row[2]) if g_row[2] else 0)
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]} {g_row[1]} {g_row[2]}")
                continue
            # Student_Count (col 3), Teacher_Count (col 4), TA_Count (col 5)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1):
                    errors.append(f"{key}: Student_Count {a_row[3]} vs {g_row[3]}")
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1):
                    errors.append(f"{key}: Teacher_Count {a_row[4]} vs {g_row[4]}")
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1):
                    errors.append(f"{key}: TA_Count {a_row[5]} vs {g_row[5]}")
        if errors:
            for e in errors[:5]:
                check(f"Enrollment Trends data", False, e)
        else:
            check("Enrollment Trends data matches", True)

    # Sheet 2: Course Capacity
    print("  --- Course Capacity ---")
    a_rows = load_sheet_rows(agent_wb, "Course Capacity")
    g_rows = load_sheet_rows(gt_wb, "Course Capacity")
    if a_rows is None:
        check("Sheet 'Course Capacity' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Course Capacity' exists in GT", False)
    else:
        check("Sheet 'Course Capacity' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Course Capacity has 7 data rows", len(a_data) == 7, f"Found {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            # Latest_Enrollment (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}: Latest_Enrollment {a_row[1]} vs {g_row[1]}")
            # Growth_Trend (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not str_match(a_row[2], g_row[2]):
                    errors.append(f"{key}: Growth_Trend '{a_row[2]}' vs '{g_row[2]}'")
            # Needs_Split (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not str_match(a_row[3], g_row[3]):
                    errors.append(f"{key}: Needs_Split '{a_row[3]}' vs '{g_row[3]}'")
            # Consider_Consolidation (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not str_match(a_row[4], g_row[4]):
                    errors.append(f"{key}: Consider_Consolidation '{a_row[4]}' vs '{g_row[4]}'")
            # Projected_Next (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 5):
                    errors.append(f"{key}: Projected_Next {a_row[5]} vs {g_row[5]}")
            # Faculty_Needed (col 6)
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 1):
                    errors.append(f"{key}: Faculty_Needed {a_row[6]} vs {g_row[6]}")
        if errors:
            for e in errors[:5]:
                check(f"Course Capacity data", False, e)
        else:
            check("Course Capacity data matches", True)

    # Sheet 3: Department Summary
    print("  --- Department Summary ---")
    a_rows = load_sheet_rows(agent_wb, "Department Summary")
    g_rows = load_sheet_rows(gt_wb, "Department Summary")
    if a_rows is None:
        check("Sheet 'Department Summary' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Department Summary' exists in GT", False)
    else:
        check("Sheet 'Department Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing metric: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                tol = 5 if "projected" in key else 1
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            for e in errors[:5]:
                check(f"Department Summary data", False, e)
        else:
            check("Department Summary data matches", True)


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find form with "Course Preference" or "Preference Survey" in title
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%%course preference%%'
               OR title ILIKE '%%preference survey%%'
        """)
        forms = cur.fetchall()
        check("Course Preference Survey form exists", len(forms) >= 1,
              f"Found {len(forms)} matching forms")

        if forms:
            form_id = forms[0][0]
            cur.execute("""
                SELECT id, title, question_type FROM gform.questions
                WHERE form_id = %s ORDER BY position
            """, (form_id,))
            questions = cur.fetchall()
            check("Form has 4 questions", len(questions) == 4,
                  f"Found {len(questions)} questions: {[q[1] for q in questions]}")

            if len(questions) >= 4:
                # Check question types
                types = [q[2].upper() if q[2] else "" for q in questions]
                titles = [q[1].lower() if q[1] else "" for q in questions]

                has_name_text = any("name" in t and types[i] in ["TEXT", "SHORT_ANSWER", "PARAGRAPH"]
                                    for i, t in enumerate(titles))
                check("Has student name text question", has_name_text,
                      f"Questions: {list(zip(titles, types))}")

                has_checkbox = any(t in ["CHECKBOX", "CHECKBOX_GRID", "CHECK_BOX"]
                                   for t in types)
                check("Has checkbox question for courses", has_checkbox,
                      f"Types: {types}")

                has_radio = any(t in ["RADIO", "MULTIPLE_CHOICE", "CHOICE"]
                                for t in types)
                check("Has radio/choice question for schedule", has_radio,
                      f"Types: {types}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Google Form check", False, str(e))


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        chair_emails = [
            "analytics_chair@university.edu",
            "biochem_chair@university.edu",
            "computing_chair@university.edu",
            "design_chair@university.edu",
            "economics_chair@university.edu",
            "finance_chair@university.edu",
            "governance_chair@university.edu",
        ]

        cur.execute("""
            SELECT id, subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%%enrollment%%'
               OR subject ILIKE '%%forecast%%'
               OR subject ILIKE '%%projection%%'
        """)
        emails = cur.fetchall()
        check("At least 7 enrollment-related emails sent", len(emails) >= 7,
              f"Found {len(emails)} matching emails")

        # Check each chair got an email
        found_chairs = set()
        for email_row in emails:
            to_addr = email_row[2]
            if isinstance(to_addr, str):
                try:
                    to_addr = json.loads(to_addr)
                except json.JSONDecodeError:
                    to_addr = [to_addr]
            if isinstance(to_addr, list):
                for addr in to_addr:
                    addr_lower = str(addr).lower().strip()
                    for chair in chair_emails:
                        if chair in addr_lower:
                            found_chairs.add(chair)

        for chair in chair_emails:
            dept_name = chair.split("_chair")[0].replace("_", " ").title()
            check(f"Email sent to {chair}",
                  chair in found_chairs,
                  f"Found emails to: {found_chairs}")

        # Check emails have body content
        if emails:
            has_body = any(email_row[3] and len(str(email_row[3])) > 20 for email_row in emails)
            check("Emails have meaningful body content", has_body,
                  f"Body lengths: {[len(str(e[3])) if e[3] else 0 for e in emails[:3]]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

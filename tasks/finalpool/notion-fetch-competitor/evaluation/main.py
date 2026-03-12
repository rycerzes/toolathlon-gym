"""
Evaluation script for notion-fetch-competitor task.

Checks:
1. Excel file (Competitor_Analysis.xlsx) - two sheets with correct product data
2. Notion page exists with competitor analysis summary
3. Memory file (memory/memory.json) has been updated
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

EXPECTED_PRODUCTS = [
    {"name": "Gamma Ultra", "price": 499, "rating": 4.8},
    {"name": "Zeta Max", "price": 599, "rating": 4.7},
    {"name": "Iota Edge", "price": 399, "rating": 4.6},
    {"name": "Alpha Pro", "price": 299, "rating": 4.5},
    {"name": "Epsilon Core", "price": 349, "rating": 4.3},
    {"name": "Beta Suite", "price": 199, "rating": 4.2},
    {"name": "Theta Plus", "price": 249, "rating": 4.1},
    {"name": "Kappa Flex", "price": 159, "rating": 4.0},
    {"name": "Delta Lite", "price": 89, "rating": 3.9},
    {"name": "Eta Basic", "price": 49, "rating": 3.5},
]

DETAILED_PRODUCTS = ["Alpha Pro", "Gamma Ultra", "Zeta Max"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Competitor_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # Check Product Listing sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "product" in name.lower() and "listing" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if "listing" in name.lower() or "product" in name.lower():
                sheet_name = name
                break

    if not sheet_name:
        record("Sheet 'Product Listing' exists", False, f"Sheets found: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Product Listing' exists", True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        if len(data_rows) != 10:
            record("Product Listing has 10 rows", False, f"Found {len(data_rows)} rows")
            all_ok = False
        else:
            record("Product Listing has 10 rows", True)
            for i, expected in enumerate(EXPECTED_PRODUCTS):
                if i >= len(data_rows):
                    break
                row = data_rows[i]
                product_name = str(row[0]).strip() if row[0] else ""
                name_ok = str_match(product_name, expected["name"])
                if not name_ok:
                    record(f"Row {i+1} product name", False,
                           f"Got '{product_name}', expected '{expected['name']}'")
                    all_ok = False
                price_ok = num_close(row[1], expected["price"], tol=5.0)
                if not price_ok:
                    record(f"Row {i+1} price", False,
                           f"Got {row[1]}, expected {expected['price']}")
                    all_ok = False
                rating_ok = num_close(row[2], expected["rating"], tol=0.2)
                if not rating_ok:
                    record(f"Row {i+1} rating", False,
                           f"Got {row[2]}, expected {expected['rating']}")
                    all_ok = False
            if all_ok:
                record("Product Listing data matches expected", True)

        # Reverse: check no extra rows beyond 10
        if len(data_rows) > 10:
            record("Product Listing has no extra rows", False,
                   f"Found {len(data_rows)} rows, expected exactly 10")
            all_ok = False

    # Check Detailed Analysis sheet
    detail_sheet = None
    for name in wb.sheetnames:
        if "detail" in name.lower():
            detail_sheet = name
            break
    if not detail_sheet:
        record("Sheet 'Detailed Analysis' exists", False, f"Sheets found: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Detailed Analysis' exists", True)
        ws = wb[detail_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        if len(data_rows) != 3:
            record("Detailed Analysis has 3 rows", False, f"Found {len(data_rows)} rows")
            all_ok = False
        else:
            record("Detailed Analysis has 3 rows", True)
            found_products = set()
            for row in data_rows:
                pname = str(row[0]).strip() if row[0] else ""
                for dp in DETAILED_PRODUCTS:
                    if str_match(pname, dp):
                        found_products.add(dp.lower())
            for dp in DETAILED_PRODUCTS:
                if dp.lower() in found_products:
                    record(f"Detailed Analysis contains {dp}", True)
                else:
                    record(f"Detailed Analysis contains {dp}", False, "Not found")
                    all_ok = False

        # Reverse: check no extra rows beyond 3
        if len(data_rows) > 3:
            record("Detailed Analysis has no extra rows", False,
                   f"Found {len(data_rows)} rows, expected exactly 3")
            all_ok = False

    # Reverse validation: no unexpected sheets
    expected_sheets = {"product listing", "detailed analysis"}
    actual_sheets = set()
    for s in wb.sheetnames:
        normalized = s.strip().lower()
        actual_sheets.add(normalized)
    extra = actual_sheets - expected_sheets
    if extra:
        record("No unexpected extra sheets", False, f"Extra sheets: {extra}")
        all_ok = False
    else:
        record("No unexpected extra sheets", True)

    wb.close()
    return all_ok


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false AND in_trash = false")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion DB accessible", False, str(e))
        return False

    all_ok = True
    found_page = False

    for page_id, props_raw in pages:
        if isinstance(props_raw, str):
            props = json.loads(props_raw)
        else:
            props = props_raw

        title_parts = []
        for key in ["title", "Name"]:
            prop = props.get(key, {})
            if isinstance(prop, dict):
                title_parts = prop.get("title", [])
                if title_parts:
                    break

        page_title = "".join(p.get("plain_text", "") for p in title_parts)

        if "competitor" in page_title.lower():
            found_page = True
            record("Notion page with 'Competitor' in title exists", True)

            try:
                conn2 = psycopg2.connect(**DB_CONFIG)
                cur2 = conn2.cursor()
                cur2.execute("SELECT type, block_data FROM notion.blocks WHERE parent_id = %s", (page_id,))
                blocks = cur2.fetchall()
                cur2.close()
                conn2.close()

                all_text = " ".join(str(b[1]) for b in blocks if b[1]).lower()

                has_count = "10" in all_text
                record("Notion page mentions 10 products", has_count)
                if not has_count:
                    all_ok = False

                has_avg_price = "289" in all_text
                record("Notion page mentions average price ~289", has_avg_price)
                if not has_avg_price:
                    all_ok = False

                has_highest = "gamma" in all_text
                record("Notion page mentions Gamma Ultra", has_highest)
                if not has_highest:
                    all_ok = False

            except Exception as e:
                record("Notion blocks readable", False, str(e))
                all_ok = False
            break

    if not found_page:
        record("Notion page with 'Competitor' in title exists", False,
               f"Found {len(pages)} pages but none match")
        all_ok = False

    # Reverse: no extra Notion pages beyond the expected one
    non_competitor_pages = []
    for page_id, props_raw in pages:
        if isinstance(props_raw, str):
            props = json.loads(props_raw)
        else:
            props = props_raw
        title_parts = []
        for key in ["title", "Name"]:
            prop = props.get(key, {})
            if isinstance(prop, dict):
                title_parts = prop.get("title", [])
                if title_parts:
                    break
        page_title = "".join(p.get("plain_text", "") for p in title_parts)
        if "competitor" not in page_title.lower():
            non_competitor_pages.append(page_title)
    if non_competitor_pages:
        record("No unexpected Notion pages", False,
               f"Found {len(non_competitor_pages)} extra page(s): {non_competitor_pages[:3]}")
        all_ok = False
    else:
        record("No unexpected Notion pages", True)

    return all_ok


def check_memory(agent_workspace):
    print("\n=== Checking Memory ===")
    memory_path = os.path.join(agent_workspace, "memory", "memory.json")
    if not os.path.isfile(memory_path):
        record("Memory file exists", False, f"Not found: {memory_path}")
        return False
    record("Memory file exists", True)

    try:
        with open(memory_path, "r") as f:
            data = json.load(f)
    except Exception as e:
        record("Memory file is valid JSON", False, str(e))
        return False
    record("Memory file is valid JSON", True)

    content = json.dumps(data).lower()
    has_notes = len(content) > 30
    record("Memory file has been updated with notes", has_notes)
    has_analysis = "analysis" in content or "competitor" in content or "complete" in content
    record("Memory mentions analysis/competitor/complete", has_analysis)
    return has_notes and has_analysis


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    notion_ok = check_notion()
    memory_ok = check_memory(args.agent_workspace)

    overall = excel_ok and notion_ok and memory_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Notion:   {'PASS' if notion_ok else 'FAIL'}")
    print(f"  Memory:   {'PASS' if memory_ok else 'FAIL'}")
    print(f"  Overall:  {'PASS' if overall else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

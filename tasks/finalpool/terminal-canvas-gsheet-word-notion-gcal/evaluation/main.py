"""Evaluation for terminal-canvas-gsheet-word-notion-gcal.
Checks:
1. Academic_Advising_Report.xlsx with 4 sheets
2. Google Sheet "Academic Advising Analytics"
3. Advising_Recommendations.docx with required sections
4. Notion database "Student Advising Tracker"
5. Google Calendar advising events
6. advising_analyzer.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: Academic_Advising_Report.xlsx ===")
    path = os.path.join(workspace, "Academic_Advising_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # Course_Summary
    cs_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s or "summary" in s), 0)
    ws = wb[sheets[cs_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    # Query dynamic course count from Canvas DB (courses used in this task)
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM canvas.courses WHERE id IN (6, 7)")
        expected_course_rows = cur.fetchone()[0]
        cur.close(); conn.close()
    except Exception:
        expected_course_rows = 2
    check(f"Course_Summary has {expected_course_rows} course rows",
          len(data_rows) >= expected_course_rows, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    check("Contains Global Governance", "global governance" in all_text or "governance" in all_text,
          f"Text: {all_text[:120]}")

    # Grade_Distribution
    gd_idx = next((i for i, s in enumerate(sheets_lower) if "grade" in s or "distribution" in s), 1)
    if gd_idx < len(sheets):
        ws2 = wb[sheets[gd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Grade_Distribution has 5 range rows", len(data_rows2) >= 5, f"Found {len(data_rows2)}")

    # Advising_Needs
    an_idx = next((i for i, s in enumerate(sheets_lower) if "advising" in s or "need" in s), 2)
    if an_idx < len(sheets):
        ws3 = wb[sheets[an_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Advising_Needs has 4 category rows", len(data_rows3) >= 4, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Contains Urgent Intervention", "urgent" in all_text3, f"Text: {all_text3[:120]}")

    # Appointment_Schedule
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "appointment" in s or "schedule" in s), 3)
    if ap_idx < len(sheets):
        ws4 = wb[sheets[ap_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Appointment_Schedule has 5 weekday rows", len(data_rows4) >= 5, f"Found {len(data_rows4)}")


def check_gsheet():
    print("\n=== Check 2: Google Sheet Analytics ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE lower(title) LIKE '%%advising%%'")
    sheets = cur.fetchall()
    check("Academic Advising spreadsheet exists", len(sheets) >= 1,
          f"Found: {[s[1] for s in sheets]}")

    if sheets:
        ss_id = sheets[0][0]
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,))
        cnt = cur.fetchone()[0]
        check("Spreadsheet has data cells", cnt >= 5, f"Found {cnt} cells")

    cur.close()
    conn.close()


def check_word(workspace):
    print("\n=== Check 3: Advising_Recommendations.docx ===")
    path = os.path.join(workspace, "Advising_Recommendations.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)

    try:
        from docx import Document
        doc = Document(path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Contains performance overview", "overview" in all_text or "performance" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains at-risk identification", "risk" in all_text or "intervention" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains recommendations", "recommend" in all_text or "action" in all_text,
              f"Text: {all_text[:150]}")
        check("Mentions Global Governance", "governance" in all_text or "geopolitics" in all_text,
              f"Text: {all_text[:150]}")
    except ImportError:
        check("python-docx available", False, "python-docx not installed")


def check_notion():
    print("\n=== Check 4: Notion Student Advising Tracker ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM notion.databases")
    dbs = cur.fetchall()
    advising_db = None
    for db_id, title in dbs:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
        elif isinstance(title, str):
            try:
                parsed = json.loads(title)
                if isinstance(parsed, list):
                    title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                else:
                    title_str = str(title)
            except Exception:
                title_str = str(title)
        else:
            title_str = str(title) if title else ""
        if "advising" in title_str.lower() and "tracker" in title_str.lower():
            advising_db = (db_id, title_str)
            break

    check("Student Advising Tracker exists", advising_db is not None,
          f"Databases: {[d[1] for d in dbs]}")

    if advising_db:
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE parent->>'database_id' = %s
        """, (advising_db[0],))
        count = cur.fetchone()[0]
        check("Tracker has advising category entries", count >= 4, f"Found {count} pages")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 5: Calendar Advising Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE lower(summary) LIKE '%%advising%%'
           OR lower(summary) LIKE '%%academic advising%%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 5 advising events", len(events) >= 5, f"Found {len(events)} events")

    if events:
        summaries = " ".join(str(e[0]) for e in events).lower()
        check("Events mention academic advising", "academic" in summaries or "advising" in summaries,
              f"Summaries: {summaries[:150]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 6: advising_analyzer.py ===")
    path = os.path.join(workspace, "advising_analyzer.py")
    check("advising_analyzer.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check no advising events on weekends in calendar
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, EXTRACT(DOW FROM start_datetime) AS dow
            FROM gcal.events
            WHERE lower(summary) LIKE '%%advising%%'
        """)
        events = cur.fetchall()
        weekend_events = [e for e in events if e[2] in (0, 6)]
        check("No advising events on weekends",
              len(weekend_events) == 0,
              f"Found {len(weekend_events)} weekend advising events")
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gsheet()
    check_word(args.agent_workspace)
    check_notion()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

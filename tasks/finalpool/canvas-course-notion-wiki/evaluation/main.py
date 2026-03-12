"""
Evaluation script for canvas-course-notion-wiki task.

Checks:
1. Excel file (course_catalog.xlsx) - correct course data and summary
2. Notion pages created for each course
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

COURSE_CODES = ["AAA-2014J", "BBB-2014J", "CCC-2014J", "DDD-2014J",
                "EEE-2014J", "FFF-2014J", "GGG-2014J"]

EXPECTED_STUDENTS = {
    "AAA-2014J": 365, "BBB-2014J": 2292, "CCC-2014J": 2498,
    "DDD-2014J": 1803, "EEE-2014J": 1188, "FFF-2014J": 2365, "GGG-2014J": 749,
}

EXPECTED_ASSIGNMENTS = {
    "AAA-2014J": 6, "BBB-2014J": 6, "CCC-2014J": 10,
    "DDD-2014J": 7, "EEE-2014J": 5, "FFF-2014J": 13, "GGG-2014J": 10,
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def int_close(a, b, tol=10):
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "course_catalog.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "course_catalog.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(agent_file, data_only=True)

    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    # Sheet 1: Courses
    ws1 = get_sheet(wb, "Courses")
    if ws1 is None:
        record("Sheet 'Courses' exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Sheet 'Courses' exists", True)

    headers = [str(c.value).strip() if c.value else "" for c in ws1[1]]
    expected_headers = ["Course_Code", "Course_Name", "Start_Date", "End_Date",
                        "Student_Count", "Assignment_Count", "Instructor_Names"]
    headers_ok = all(str_match(h, e) for h, e in zip(headers, expected_headers))
    record("Courses headers match", headers_ok,
           f"Expected: {expected_headers}, Got: {headers}")

    rows = list(ws1.iter_rows(min_row=2, values_only=True))
    record("Courses has 7 rows", len(rows) == 7, f"Got {len(rows)}")

    # Check sorted by course code
    agent_codes = [str(r[0]).strip() for r in rows if r and r[0]]
    record("Courses sorted by Course_Code",
           agent_codes == sorted(agent_codes),
           f"Got: {agent_codes}")

    for code in COURSE_CODES:
        agent_row = None
        for r in rows:
            if r and str_match(r[0], code):
                agent_row = r
                break
        if not agent_row:
            record(f"Course {code} present", False, "Missing")
            continue
        record(f"Course {code} present", True)

        # Check student count
        record(f"Course {code}: Student_Count",
               int_close(agent_row[4], EXPECTED_STUDENTS[code]),
               f"Expected ~{EXPECTED_STUDENTS[code]}, got {agent_row[4]}")

        # Check assignment count
        record(f"Course {code}: Assignment_Count",
               int_close(agent_row[5], EXPECTED_ASSIGNMENTS[code], 2),
               f"Expected ~{EXPECTED_ASSIGNMENTS[code]}, got {agent_row[5]}")

    # Sheet 2: Summary
    ws2 = get_sheet(wb, "Summary")
    if ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Sheet 'Summary' exists", True)

    summary = {}
    for row in ws2.iter_rows(min_row=1, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip().lower()] = row[1]

    record("Summary: Total_Courses = 7",
           str(summary.get("total_courses", "")).strip() == "7",
           f"Got {summary.get('total_courses')}")

    record("Summary: Total_Students",
           int_close(summary.get("total_students", 0), 11260, 50),
           f"Expected ~11260, got {summary.get('total_students')}")

    record("Summary: Total_Assignments",
           int_close(summary.get("total_assignments", 0), 57, 3),
           f"Expected ~57, got {summary.get('total_assignments')}")

    record("Summary: Largest_Course = CCC-2014J",
           str_match(summary.get("largest_course", ""), "CCC-2014J"),
           f"Got {summary.get('largest_course')}")

    record("Summary: Smallest_Course = AAA-2014J",
           str_match(summary.get("smallest_course", ""), "AAA-2014J"),
           f"Got {summary.get('smallest_course')}")

    record("Summary: Most_Assignments = FFF-2014J",
           str_match(summary.get("most_assignments", ""), "FFF-2014J"),
           f"Got {summary.get('most_assignments')}")

    return True


# ============================================================================
# Check 2: Notion pages
# ============================================================================

def check_notion():
    print("\n=== Checking Notion Pages ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, properties::text
        FROM notion.pages
        WHERE archived = false OR archived IS NULL
    """)
    pages = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_notion] Found {len(pages)} Notion pages.")

    record("At least 7 Notion pages created", len(pages) >= 7,
           f"Found {len(pages)}")

    # Check each course code appears in page properties/title
    all_text = " ".join(str(p[1] or "") for p in pages).lower()

    for code in COURSE_CODES:
        found = code.lower() in all_text
        record(f"Notion: page for {code} exists",
               found, f"{code} not found in any page properties")

    return True


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    notion_ok = check_notion()

    all_passed = excel_ok and notion_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

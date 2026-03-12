"""Evaluation for wc-shipping-tax-excel-word."""
import os
import argparse, os, sys
import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Tax_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Shipping_Tax_Analysis.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        rows1 = load_sheet_rows(wb, "Shipping Zones")
        if rows1 is None:
            errors.append("Sheet 'Shipping Zones' not found")
        else:
            data_rows = [r for r in rows1[1:] if r and r[0] is not None]
            if len(data_rows) < 3:
                errors.append(f"Shipping Zones has {len(data_rows)} rows, expected at least 3")
            zone_names = {str(r[0]).strip().lower() for r in data_rows}
            for expected in ["domestic us", "california", "international"]:
                if not any(expected in z for z in zone_names):
                    errors.append(f"Expected zone '{expected}' not found in Shipping Zones")

        rows2 = load_sheet_rows(wb, "Tax Rates")
        if rows2 is None:
            errors.append("Sheet 'Tax Rates' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 6:
                errors.append(f"Tax Rates has {len(data_rows2)} rows, expected at least 6")

        # --- Groundtruth XLSX value comparison ---
        gt_path = os.path.join(groundtruth_workspace, "Shipping_Tax_Analysis.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    errors.append(f"GT sheet '{gt_sname}' not found in agent xlsx (available: {wb.sheetnames})")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                if len(a_rows) != len(gt_rows):
                    errors.append(f"GT '{gt_sname}' row count: expected {len(gt_rows)}, got {len(a_rows)}")
                for ri in range(min(3, len(gt_rows))):
                    if ri >= len(a_rows): break
                    for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                        gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                        if gv is None: continue
                        if isinstance(gv, (int, float)):
                            if not num_close(av, gv, max(abs(gv)*0.1, 1.0)):
                                errors.append(f"GT '{gt_sname}' row {ri+1} col {ci+1}: expected {gv}, got {av}")
                                break
                        else:
                            if not str_match(av, gv):
                                errors.append(f"GT '{gt_sname}' row {ri+1} col {ci+1}: expected '{gv}', got '{av}'")
                                break
            gt_wb.close()
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word_doc(agent_workspace):
    errors = []
    doc_path = os.path.join(agent_workspace, "Operations_Report.docx")
    if not os.path.exists(doc_path):
        return ["Operations_Report.docx not found"]
    try:
        from docx import Document
        doc = Document(doc_path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        if "shipping" not in full_text:
            errors.append("Word doc does not contain 'shipping' keyword")
        if "tax" not in full_text:
            errors.append("Word doc does not contain 'tax' keyword")
        if len(full_text.strip()) < 100:
            errors.append("Word doc content is too short (less than 100 chars)")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE to_addr::text ILIKE '%operations@store.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No email found to operations@store.com")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word_doc(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

"""Generate groundtruth Excel file for yf-portfolio-risk-excel task."""
import os
import psycopg2
from openpyxl import Workbook
from collections import defaultdict
import math

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(TASK_ROOT, "groundtruth_workspace")

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

TICKERS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get prices for all 5 tickers
    prices = defaultdict(dict)
    for ticker in TICKERS:
        cur.execute("""
            SELECT date, close FROM yf.stock_prices
            WHERE symbol = %s ORDER BY date
        """, (ticker,))
        for date, close in cur.fetchall():
            prices[date][ticker] = float(close)

    # Only dates where all 5 have data
    valid_dates = sorted([d for d, p in prices.items() if len(p) == len(TICKERS)])

    cur.close()
    conn.close()

    wb = Workbook()

    # Sheet 1: Price History
    ws1 = wb.active
    ws1.title = "Price History"
    ws1.append(["Date"] + TICKERS)
    for d in valid_dates:
        row = [d.strftime("%Y-%m-%d") if hasattr(d, 'strftime') else str(d)]
        for t in TICKERS:
            row.append(round(prices[d][t], 2))
        ws1.append(row)

    # Compute statistics
    stats = {}
    for ticker in TICKERS:
        vals = [prices[d][ticker] for d in valid_dates]
        avg_val = sum(vals) / len(vals)
        std_val = math.sqrt(sum((v - avg_val)**2 for v in vals) / (len(vals) - 1))
        min_val = min(vals)
        max_val = max(vals)
        cv = std_val / avg_val
        stats[ticker] = {
            "avg": round(avg_val, 2),
            "std": round(std_val, 2),
            "min": round(min_val, 2),
            "max": round(max_val, 2),
            "cv": round(cv, 4),
        }

    # Sheet 2: Risk Metrics
    ws2 = wb.create_sheet("Risk Metrics")
    ws2.append(["Ticker", "Average Close Price", "Standard Deviation",
                "Minimum Close", "Maximum Close"])
    for ticker in TICKERS:
        s = stats[ticker]
        ws2.append([ticker, s["avg"], s["std"], s["min"], s["max"]])

    # Sheet 3: Risk Assessment
    ws3 = wb.create_sheet("Risk Assessment")
    ws3.append(["Ticker", "Coefficient of Variation", "Risk Category"])
    for ticker in TICKERS:
        cv = stats[ticker]["cv"]
        if cv < 0.10:
            category = "Low Risk"
        elif cv <= 0.20:
            category = "Medium Risk"
        else:
            category = "High Risk"
        ws3.append([ticker, cv, category])

    out_path = os.path.join(OUTPUT_DIR, "Portfolio_Risk_Analysis.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")

    # Print stats for reference
    for ticker in TICKERS:
        s = stats[ticker]
        cat = "Low Risk" if s["cv"] < 0.10 else ("Medium Risk" if s["cv"] <= 0.20 else "High Risk")
        print(f"  {ticker}: avg={s['avg']}, std={s['std']}, cv={s['cv']}, category={cat}")


if __name__ == "__main__":
    main()

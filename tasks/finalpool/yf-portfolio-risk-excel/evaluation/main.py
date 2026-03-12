"""
Evaluation script for yf-portfolio-risk-excel task.

Checks:
1. Excel file Portfolio_Risk_Analysis.xlsx exists with 3 sheets
2. Price History sheet has correct structure
3. Risk Metrics sheet has correct statistics
4. Risk Assessment sheet has correct risk categories
"""
import argparse
import math
import os
import sys
from collections import defaultdict

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

TICKERS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    prices = defaultdict(dict)
    for ticker in TICKERS:
        cur.execute("SELECT date, close FROM yf.stock_prices WHERE symbol = %s ORDER BY date",
                    (ticker,))
        for date, close in cur.fetchall():
            prices[date][ticker] = float(close)

    valid_dates = sorted([d for d, p in prices.items() if len(p) == len(TICKERS)])

    stats = {}
    for ticker in TICKERS:
        vals = [prices[d][ticker] for d in valid_dates]
        avg_val = sum(vals) / len(vals)
        std_val = math.sqrt(sum((v - avg_val)**2 for v in vals) / (len(vals) - 1))
        cv = std_val / avg_val
        if cv < 0.10:
            category = "Low Risk"
        elif cv <= 0.20:
            category = "Medium Risk"
        else:
            category = "High Risk"
        stats[ticker] = {
            "avg": round(avg_val, 2),
            "std": round(std_val, 2),
            "min": round(min(vals), 2),
            "max": round(max(vals), 2),
            "cv": round(cv, 4),
            "category": category,
        }

    cur.close()
    conn.close()
    return stats, len(valid_dates)


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def number_close_match(expected, actual, tolerance=0.05):
    """Check if actual is within tolerance of expected."""
    try:
        return abs(float(actual) - float(expected)) <= abs(float(expected) * tolerance) + 0.01
    except (ValueError, TypeError):
        return False


def check_excel(agent_workspace):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Portfolio_Risk_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    expected_stats, expected_rows = load_expected()

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    # Price History
    ws_ph = find_sheet(["price"])
    if not ws_ph:
        ws_ph = find_sheet(["history"])
    check("Price History sheet exists", ws_ph is not None, f"Sheets: {wb.sheetnames}")
    if ws_ph:
        # Check has all tickers in header
        header = [str(c.value).upper() if c.value else "" for c in ws_ph[1]]
        for t in TICKERS:
            check(f"Price History header contains {t}",
                  t in header,
                  f"Header: {header}")
        # Check row count (header + data)
        data_rows = ws_ph.max_row - 1
        check(f"Price History has ~{expected_rows} data rows",
              abs(data_rows - expected_rows) <= 5,
              f"Found {data_rows} rows, expected ~{expected_rows}")

    # Risk Metrics
    ws_rm = find_sheet(["risk", "metric"])
    if not ws_rm:
        ws_rm = find_sheet(["metric"])
    check("Risk Metrics sheet exists", ws_rm is not None, f"Sheets: {wb.sheetnames}")
    if ws_rm:
        all_text = ""
        for row in ws_rm.iter_rows(values_only=True):
            all_text += " ".join(str(c) for c in row if c is not None) + " "

        for ticker in TICKERS:
            s = expected_stats[ticker]
            check(f"Risk Metrics: {ticker} present", ticker in all_text)
            check(f"Risk Metrics: {ticker} avg ~{s['avg']}",
                  str(s["avg"]) in all_text,
                  f"Expected {s['avg']}")

    # Risk Assessment
    ws_ra = find_sheet(["risk", "assess"])
    if not ws_ra:
        ws_ra = find_sheet(["assess"])
    check("Risk Assessment sheet exists", ws_ra is not None, f"Sheets: {wb.sheetnames}")
    if ws_ra:
        all_text = ""
        for row in ws_ra.iter_rows(values_only=True):
            all_text += " ".join(str(c) for c in row if c is not None) + " "
        all_lower = all_text.lower()

        for ticker in TICKERS:
            s = expected_stats[ticker]
            check(f"Risk Assessment: {ticker} present", ticker in all_text)
            cat_lower = s["category"].lower()
            # Check category appears somewhere
            check(f"Risk Assessment: {ticker} category '{s['category']}'",
                  cat_lower in all_lower,
                  f"Category not found")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_passed = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

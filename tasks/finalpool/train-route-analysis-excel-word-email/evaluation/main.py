"""
Evaluation for train-route-analysis-excel-word-email task.

Checks:
1. Route_Analysis.xlsx exists with Routes and Summary sheets
2. Routes sheet has 3 rows with G11, G105, G1 and correct duration/prices
3. Summary has correct Fastest_Route=G1, Cheapest_Second_Class
4. Route_Analysis_Report.docx exists with 4 sections
5. Email sent to logistics@company.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Route_Analysis.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Route_Analysis.xlsx")
    if not os.path.exists(xlsx_path):
        record("Route_Analysis.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Route_Analysis.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    if "routes" not in sheet_names_lower:
        record("Routes sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Routes sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("routes")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Routes has 3 data rows", len(data_rows) == 3, f"Found {len(data_rows)}")

        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        record("Routes has G11", "G11" in all_text, all_text[:100])
        record("Routes has G105", "G105" in all_text, all_text[:100])
        record("Routes has G1", "G1" in all_text, all_text[:100])

        # Check duration in minutes
        numeric_vals = []
        for r in data_rows:
            for c in r:
                try:
                    numeric_vals.append(float(c))
                except (TypeError, ValueError):
                    pass
        has_duration_g1 = any(abs(v - 268) < 5 for v in numeric_vals)
        record("G1 duration ~268 min present", has_duration_g1, f"Numerics: {numeric_vals}")

        # Check prices
        has_349 = any(abs(v - 349.0) < 0.1 for v in numeric_vals)
        has_545 = any(abs(v - 545.5) < 0.1 for v in numeric_vals)
        record("Second class prices correct (349.0 and 545.5)", has_349 and has_545,
               f"Numerics: {numeric_vals}")

    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws_s = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows_s = list(ws_s.iter_rows(values_only=True))
        all_text_s = " ".join(str(c) for r in rows_s for c in r if c).upper()
        record("Summary mentions G1 as fastest", "G1" in all_text_s and "FASTEST" in all_text_s,
               all_text_s[:200])
        record("Summary has Total_Routes = 3", "3" in all_text_s, all_text_s[:200])


def check_word(agent_workspace):
    print("\n=== Check 2: Route_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Route_Analysis_Report.docx")
    if not os.path.exists(docx_path):
        record("Route_Analysis_Report.docx exists", False, f"Not found at {docx_path}")
        return
    record("Route_Analysis_Report.docx exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
    except Exception as e:
        record("Word file readable", False, str(e))
        return
    record("Word file readable", True)

    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    has_exec = "executive summary" in full_text or "executive" in full_text
    has_route_details = "route details" in full_text or "route detail" in full_text
    has_recommendation = "recommendation" in full_text
    has_cost = "cost comparison" in full_text or "cost" in full_text
    record("Has Executive Summary section", has_exec)
    record("Has Route Details section", has_route_details)
    record("Has Recommendation section", has_recommendation)
    record("Has Cost Comparison section", has_cost)

    has_g1_rec = "g1" in full_text and ("fastest" in full_text or "recommend" in full_text or "speed" in full_text)
    record("Recommends G1 for speed", has_g1_rec, full_text[:300])


def check_email():
    print("\n=== Check 3: Email to logistics@company.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT to_addr, subject FROM email.messages
        WHERE subject ILIKE '%beijing%' OR subject ILIKE '%route analysis%' OR subject ILIKE '%train route%'
    """)
    messages = cur.fetchall()
    cur.close()
    conn.close()

    all_msgs = list(messages)
    record("Route analysis email sent", len(all_msgs) >= 1,
           f"Found {len(all_msgs)} matching emails")

    if all_msgs:
        to_raw = all_msgs[0][0]
        to_str = str(to_raw).lower() if to_raw else ""
        record("Email sent to logistics@company.com", "logistics@company.com" in to_str,
               f"To: {to_str[:100]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

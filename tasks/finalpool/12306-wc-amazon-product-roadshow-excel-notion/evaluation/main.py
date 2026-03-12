"""
Evaluation for 12306-wc-amazon-product-roadshow-excel-notion task.

Checks:
1. Roadshow_Plan.xlsx exists
2. Products sheet has >= 4 data rows with Name and Price columns
3. Travel_Itinerary sheet has >= 2 rows containing G11 and G105 train numbers
4. Roadshow_Schedule sheet has >= 3 rows
5. Notion page exists with Roadshow or Shanghai or Guangzhou in title
6. Email sent to shanghai_dist@partner.com
7. Email sent to guangzhou_dist@partner.com
8. Email sent to manager@company.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Roadshow_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Roadshow_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Roadshow_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Roadshow_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # Check Products sheet
    prod_sheet = None
    for name in wb.sheetnames:
        if "product" in name.lower():
            prod_sheet = wb[name]
            break
    if prod_sheet is None:
        record("Products sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Products sheet exists", True)
        rows = list(prod_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_name = any("name" in h for h in headers)
        has_price = any("price" in h for h in headers)
        record("Products has Name and Price columns", has_name and has_price,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Products has >= 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} data rows")

    # Check Travel_Itinerary sheet
    travel_sheet = None
    for name in wb.sheetnames:
        if "travel" in name.lower() or "itinerary" in name.lower():
            travel_sheet = wb[name]
            break
    if travel_sheet is None:
        record("Travel_Itinerary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Travel_Itinerary sheet exists", True)
        rows = list(travel_sheet.iter_rows(values_only=True))
        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Itinerary has >= 2 data rows", len(data_rows) >= 2,
               f"Found {len(data_rows)} rows")
        record("Travel_Itinerary contains G11", "G11" in all_text,
               "G11 not found in Travel_Itinerary")
        record("Travel_Itinerary contains G105", "G105" in all_text,
               "G105 not found in Travel_Itinerary")

    # Check Roadshow_Schedule sheet
    sched_sheet = None
    for name in wb.sheetnames:
        if "schedule" in name.lower() or "roadshow" in name.lower():
            sched_sheet = wb[name]
            break
    if sched_sheet is None:
        record("Roadshow_Schedule sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Roadshow_Schedule sheet exists", True)
        rows = list(sched_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Roadshow_Schedule has >= 3 data rows", len(data_rows) >= 3,
               f"Found {len(data_rows)} rows")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Roadshow_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        record(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices.append(len(gt_rows) - 1)
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        record(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1}",
                               False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    record(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                record(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, parent, properties FROM notion.pages")
    pages = cur.fetchall()
    cur.close()
    conn.close()

    found = False
    for page_id, parent, props in pages:
        try:
            title_items = []
            for key, val in props.items():
                if isinstance(val, dict) and val.get("type") == "title":
                    title_items = val.get("title", [])
                    break
            title_text = " ".join(
                item.get("text", {}).get("content", "") for item in title_items
                if isinstance(item, dict)
            ).lower()
            if "roadshow" in title_text or "shanghai" in title_text or "guangzhou" in title_text:
                found = True
                break
        except Exception:
            continue

    record("Notion page exists with Roadshow/Shanghai/Guangzhou in title", found,
           f"Total pages: {len(pages)}")


def check_emails_sent():
    print("\n=== Check 3: Emails Sent ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # Check messages in Sent/SENT folders
        cur.execute("""
            SELECT m.to_addr FROM email.messages m
            JOIN email.folders f ON m.folder_id = f.id
            WHERE UPPER(f.name) = 'SENT'
        """)
        sent_rows = cur.fetchall()
        # Also check via sent_log join
        cur.execute("""
            SELECT m.to_addr FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
        """)
        sent_rows += cur.fetchall()
        sent_text = " ".join(str(row[0]) for row in sent_rows).lower()

        record("Email sent to shanghai_dist@partner.com",
               "shanghai_dist@partner.com" in sent_text,
               f"Sent entries: {len(sent_rows)}")
        record("Email sent to guangzhou_dist@partner.com",
               "guangzhou_dist@partner.com" in sent_text,
               f"Sent entries: {len(sent_rows)}")
        record("Email sent to manager@company.com",
               "manager@company.com" in sent_text,
               f"Sent entries: {len(sent_rows)}")
    except Exception as e:
        record("Email sent check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_emails_sent()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

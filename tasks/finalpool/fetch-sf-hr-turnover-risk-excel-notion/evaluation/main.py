"""Evaluation script for fetch-sf-hr-turnover-risk-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Turnover_Risk_Assessment.xlsx")
    check("Turnover_Risk_Assessment.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Turnover_Risk_Assessment.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        check("Risk_Overview sheet exists", "Risk_Overview" in wb.sheetnames)
        if "Risk_Overview" in wb.sheetnames:
            ws = wb["Risk_Overview"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Risk_Overview has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Employee_Count', 'Avg_Salary', 'Risk_Level']:
                check(f"Risk_Overview has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Risk_Summary sheet exists", "Risk_Summary" in wb.sheetnames)
        if "Risk_Summary" in wb.sheetnames:
            ws = wb["Risk_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Risk_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Risk_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Detailed_Metrics sheet exists", "Detailed_Metrics" in wb.sheetnames)
        if "Detailed_Metrics" in wb.sheetnames:
            ws = wb["Detailed_Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Detailed_Metrics has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Satisfaction_Gap', 'Estimated_Turnover_Cost']:
                check(f"Detailed_Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # Check Notion page
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT properties FROM notion.pages WHERE properties::text ILIKE %s AND archived = false", ('%turnover%',))
            pages = cur.fetchall()
            check("Notion turnover page created", len(pages) >= 1, f"found {len(pages)} pages")
            conn.close()
        except Exception as e:
            check("Notion verification", False, str(e))

        check("risk_scorer.py exists", os.path.exists(os.path.join(agent_workspace, "risk_scorer.py")))
        check("risk_assessment.json exists", os.path.exists(os.path.join(agent_workspace, "risk_assessment.json")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

"""Evaluation for canvas-quiz-performance-excel."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Excel output against groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Quiz_Performance.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quiz_Performance.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    # Check Quiz Stats sheet
    print("\n--- Quiz Stats ---")
    agent_ws = get_sheet(agent_wb, "Quiz Stats")
    gt_ws = get_sheet(gt_wb, "Quiz Stats")
    check("Sheet 'Quiz Stats' exists", agent_ws is not None,
          f"Found: {agent_wb.sheetnames}")

    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Quiz Stats row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        for gt_row in gt_rows:
            title, subs, avg, mn, mx, pr = gt_row
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], title):
                    matched = ar
                    break
            if matched:
                check(f"Quiz '{title}' Total_Submissions",
                      num_close(matched[1], subs, 5),
                      f"Expected {subs}, got {matched[1]}")
                check(f"Quiz '{title}' Avg_Score",
                      num_close(matched[2], avg, 1.0),
                      f"Expected {avg}, got {matched[2]}")
                check(f"Quiz '{title}' Min_Score",
                      num_close(matched[3], mn, 1.0),
                      f"Expected {mn}, got {matched[3]}")
                check(f"Quiz '{title}' Max_Score",
                      num_close(matched[4], mx, 1.0),
                      f"Expected {mx}, got {matched[4]}")
                check(f"Quiz '{title}' Pass_Rate",
                      num_close(matched[5], pr, 1.0),
                      f"Expected {pr}, got {matched[5]}")
            else:
                check(f"Quiz '{title}' found", False)

    # Check Summary sheet
    print("\n--- Summary ---")
    agent_sum = get_sheet(agent_wb, "Summary")
    gt_sum = get_sheet(gt_wb, "Summary")
    check("Sheet 'Summary' exists", agent_sum is not None,
          f"Found: {agent_wb.sheetnames}")

    if agent_sum and gt_sum:
        gt_data = {}
        for row in gt_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]

        agent_data = {}
        for row in agent_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key, gt_val in gt_data.items():
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                check(f"Summary '{key}'",
                      num_close(agent_val, gt_val, 1.0),
                      f"Expected {gt_val}, got {agent_val}")
            else:
                check(f"Summary '{key}'",
                      str_match(agent_val, gt_val),
                      f"Expected '{gt_val}', got '{agent_val}'")

    return True


def check_emails():
    """Check that summary email was sent."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    conn.close()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    result = find_email_for_recipient("instructor@university.example.com")
    check("Summary email sent to instructor", result is not None)

    if result:
        subj, from_addr, to_addr, body = result
        has_quiz_subject = "quiz" in (subj or "").lower() and "performance" in (subj or "").lower()
        check("Email subject contains 'Quiz Performance'", has_quiz_subject,
              f"Subject: {(subj or '')[:100]}")

        body_lower = (body or "").lower()
        check("Email body mentions quiz count", "5" in body_lower or "five" in body_lower,
              "Expected mention of 5 quizzes")
        check("Email body mentions CMA 15003 (highest)",
              "cma 15003" in body_lower or "15003" in body_lower)
        check("Email body mentions CMA 15007 (lowest)",
              "cma 15007" in body_lower or "15007" in body_lower)

    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("CANVAS QUIZ PERFORMANCE EXCEL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

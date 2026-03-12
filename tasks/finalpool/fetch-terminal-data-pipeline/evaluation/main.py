"""
Evaluation script for fetch-terminal-data-pipeline task.

Checks:
1. clean_data.py exists
2. cleaned_sales.json exists with 18 unique orders (no duplicate order_ids)
3. cleaned_inventory.json exists with 5 unique products (no duplicate product_ids)
4. Data_Pipeline_Report.xlsx has 3 sheets: Sales Analysis, Inventory Status, Combined View
5. Sales Analysis has 5 product rows with correct totals
6. Inventory Status has correct statuses (at least 2 LOW)
7. Google Sheet "Pipeline Dashboard" exists with data
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def num_close_pct(a, b, pct=0.05):
    """Compare two numbers within percentage tolerance."""
    try:
        a_f, b_f = float(a), float(b)
        if b_f == 0:
            return a_f == 0
        return abs(a_f - b_f) / abs(b_f) <= pct
    except (TypeError, ValueError):
        return False


def compute_expected(groundtruth_workspace):
    """Compute expected values from the initial workspace JSON files.

    We look for the raw source files in the groundtruth_workspace's sibling
    initial_workspace, or fall back to hardcoded values.
    """
    # Try to find initial_workspace relative to groundtruth
    task_root = os.path.dirname(os.path.abspath(groundtruth_workspace))
    initial_ws = os.path.join(task_root, "initial_workspace")

    sales_path = None
    inv_path = None
    for search_dir in [initial_ws, groundtruth_workspace]:
        sp = os.path.join(search_dir, "sales_api_response.json")
        ip = os.path.join(search_dir, "inventory_api_response.json")
        if os.path.isfile(sp):
            sales_path = sp
        if os.path.isfile(ip):
            inv_path = ip

    expected = {}

    if sales_path and inv_path:
        with open(sales_path) as f:
            sales_raw = json.load(f)
        with open(inv_path) as f:
            inv_raw = json.load(f)

        # Deduplicate sales by order_id
        sales = sales_raw.get("data", sales_raw if isinstance(sales_raw, list) else [])
        seen_oids = set()
        unique_sales = []
        for rec in sales:
            oid = rec.get("order_id", "")
            if oid and oid not in seen_oids:
                seen_oids.add(oid)
                unique_sales.append(rec)
        expected["unique_sales_count"] = len(unique_sales)

        # Compute per-product sales aggregation
        product_agg = {}
        for rec in unique_sales:
            pname = rec.get("product_name", "").strip().title()
            qty = rec.get("quantity", 0)
            price = rec.get("unit_price", 0)
            if pname not in product_agg:
                product_agg[pname] = {"qty": 0, "revenue": 0.0, "prices": []}
            product_agg[pname]["qty"] += qty
            product_agg[pname]["revenue"] += qty * price
            product_agg[pname]["prices"].append(price)
        for pname in product_agg:
            prices = product_agg[pname]["prices"]
            product_agg[pname]["avg_price"] = sum(prices) / len(prices) if prices else 0
        expected["product_sales"] = product_agg
        expected["num_products_in_sales"] = len(product_agg)

        # Deduplicate inventory by product_id
        inv = inv_raw.get("data", inv_raw if isinstance(inv_raw, list) else [])
        seen_pids = set()
        unique_inv = []
        for rec in inv:
            pid = rec.get("product_id", "")
            if pid and pid not in seen_pids:
                seen_pids.add(pid)
                unique_inv.append(rec)
        expected["unique_inv_count"] = len(unique_inv)

        # Compute inventory statuses
        inv_status = {}
        low_count = 0
        for rec in unique_inv:
            pname = rec.get("product_name", "").strip().title()
            stock = rec.get("current_stock", 0)
            reorder = rec.get("reorder_point", 0)
            status = "LOW" if stock <= reorder else "OK"
            if status == "LOW":
                low_count += 1
            inv_status[pname] = {
                "stock": stock,
                "reorder": reorder,
                "status": status,
            }
        expected["inv_status"] = inv_status
        expected["low_count"] = low_count
    else:
        # Hardcoded fallback
        expected["unique_sales_count"] = 18
        expected["unique_inv_count"] = 5
        expected["num_products_in_sales"] = 5
        expected["low_count"] = 2
        expected["product_sales"] = {
            "Widget Alpha": {"qty": 35, "revenue": 525.0, "avg_price": 15.0},
            "Widget Beta": {"qty": 34, "revenue": 850.0, "avg_price": 25.0},
            "Gadget Pro": {"qty": 17, "revenue": 850.0, "avg_price": 50.0},
            "Sensor X1": {"qty": 23, "revenue": 805.0, "avg_price": 35.0},
            "Connector Z": {"qty": 75, "revenue": 750.0, "avg_price": 10.0},
        }
        expected["inv_status"] = {
            "Widget Alpha": {"stock": 120, "reorder": 50, "status": "OK"},
            "Widget Beta": {"stock": 85, "reorder": 40, "status": "OK"},
            "Gadget Pro": {"stock": 45, "reorder": 30, "status": "OK"},
            "Sensor X1": {"stock": 15, "reorder": 25, "status": "LOW"},
            "Connector Z": {"stock": 8, "reorder": 20, "status": "LOW"},
        }

    return expected


def check_script(agent_workspace):
    """Check that clean_data.py exists."""
    print("\n=== Checking clean_data.py ===")
    script_path = os.path.join(agent_workspace, "clean_data.py")
    check("clean_data.py exists", os.path.isfile(script_path),
          f"Not found at {script_path}")


def check_cleaned_data(agent_workspace, expected):
    """Check cleaned JSON output files."""
    print("\n=== Checking Cleaned Data ===")

    # --- cleaned_sales.json ---
    sales_path = os.path.join(agent_workspace, "cleaned_sales.json")
    check("cleaned_sales.json exists", os.path.isfile(sales_path),
          f"Not found at {sales_path}")

    if os.path.isfile(sales_path):
        with open(sales_path, "r") as f:
            sales_data = json.load(f)

        records = sales_data
        if isinstance(sales_data, dict):
            for key in ("data", "sales", "orders", "records"):
                if key in sales_data:
                    records = sales_data[key]
                    break
        if not isinstance(records, list):
            records = []

        exp_count = expected.get("unique_sales_count", 18)
        check(f"Cleaned sales has {exp_count} records",
              len(records) == exp_count,
              f"Found {len(records)} records, expected {exp_count}")

        # No duplicate order_ids
        oids = [r.get("order_id", "") for r in records]
        unique_oids = set(oids)
        check("No duplicate order IDs in cleaned sales",
              len(oids) == len(unique_oids),
              f"{len(oids)} total, {len(unique_oids)} unique")

        # Product names are title case
        title_ok = sum(1 for r in records
                       if r.get("product_name", "") == r.get("product_name", "").strip().title())
        check("All product names in cleaned sales are title case",
              title_ok == len(records),
              f"{title_ok}/{len(records)} are title case")

    # --- cleaned_inventory.json ---
    inv_path = os.path.join(agent_workspace, "cleaned_inventory.json")
    check("cleaned_inventory.json exists", os.path.isfile(inv_path),
          f"Not found at {inv_path}")

    if os.path.isfile(inv_path):
        with open(inv_path, "r") as f:
            inv_data = json.load(f)

        records = inv_data
        if isinstance(inv_data, dict):
            for key in ("data", "inventory", "products", "records"):
                if key in inv_data:
                    records = inv_data[key]
                    break
        if not isinstance(records, list):
            records = []

        exp_count = expected.get("unique_inv_count", 5)
        check(f"Cleaned inventory has {exp_count} records",
              len(records) == exp_count,
              f"Found {len(records)} records, expected {exp_count}")

        # No duplicate product_ids
        pids = [r.get("product_id", "") for r in records]
        unique_pids = set(pids)
        check("No duplicate product IDs in cleaned inventory",
              len(pids) == len(unique_pids),
              f"{len(pids)} total, {len(unique_pids)} unique")


def check_excel(agent_workspace, expected):
    """Check the Excel analysis file."""
    print("\n=== Checking Excel Output ===")

    excel_path = os.path.join(agent_workspace, "Data_Pipeline_Report.xlsx")
    check("Data_Pipeline_Report.xlsx exists", os.path.isfile(excel_path),
          f"Not found at {excel_path}")
    if not os.path.isfile(excel_path):
        return

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    check("Excel has 3 sheets", len(wb.sheetnames) >= 3,
          f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    sheet_names_lower = [s.lower().replace("_", " ").strip() for s in wb.sheetnames]

    # --- Sales Analysis sheet ---
    has_sales = any("sales" in s and "analysis" in s for s in sheet_names_lower)
    check("Has Sales Analysis sheet", has_sales,
          f"Found sheets: {wb.sheetnames}")

    ws_sales = None
    for s in wb.sheetnames:
        sl = s.lower().replace("_", " ")
        if "sales" in sl and "analysis" in sl:
            ws_sales = wb[s]
            break

    if ws_sales:
        data_rows = list(ws_sales.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]
        num_products = expected.get("num_products_in_sales", 5)
        check(f"Sales Analysis has {num_products} product rows",
              len(data_rows) == num_products,
              f"Found {len(data_rows)} rows")

        # Verify per-product totals
        product_sales = expected.get("product_sales", {})
        for row in data_rows:
            pname = str(row[0]).strip().title() if row[0] else ""
            for exp_name, exp_vals in product_sales.items():
                if exp_name.lower() == pname.lower():
                    if len(row) >= 2 and row[1] is not None:
                        check(f"{exp_name} total quantity",
                              num_close(row[1], exp_vals["qty"]),
                              f"Expected {exp_vals['qty']}, got {row[1]}")
                    if len(row) >= 3 and row[2] is not None:
                        check(f"{exp_name} total revenue",
                              num_close_pct(row[2], exp_vals["revenue"], 0.02),
                              f"Expected {exp_vals['revenue']}, got {row[2]}")
                    break

    # --- Inventory Status sheet ---
    has_inv = any("inventory" in s and "status" in s for s in sheet_names_lower)
    check("Has Inventory Status sheet", has_inv,
          f"Found sheets: {wb.sheetnames}")

    ws_inv = None
    for s in wb.sheetnames:
        sl = s.lower().replace("_", " ")
        if "inventory" in sl and "status" in sl:
            ws_inv = wb[s]
            break

    if ws_inv:
        data_rows = list(ws_inv.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]
        inv_count = expected.get("unique_inv_count", 5)
        check(f"Inventory Status has {inv_count} rows",
              len(data_rows) == inv_count,
              f"Found {len(data_rows)} rows")

        # Check statuses - count LOW entries
        low_found = 0
        for row in data_rows:
            if len(row) >= 4 and row[3] is not None:
                status_val = str(row[3]).strip().upper()
                if status_val == "LOW":
                    low_found += 1

        exp_low = expected.get("low_count", 2)
        check(f"Inventory Status has at least {exp_low} LOW entries",
              low_found >= exp_low,
              f"Found {low_found} LOW entries, expected at least {exp_low}")

        # Check specific statuses
        inv_status = expected.get("inv_status", {})
        for row in data_rows:
            pname = str(row[0]).strip().title() if row[0] else ""
            for exp_name, exp_vals in inv_status.items():
                if exp_name.lower() == pname.lower():
                    if len(row) >= 4 and row[3] is not None:
                        actual_status = str(row[3]).strip().upper()
                        check(f"{exp_name} status is {exp_vals['status']}",
                              actual_status == exp_vals["status"],
                              f"Expected {exp_vals['status']}, got {actual_status}")
                    break

    # --- Combined View sheet ---
    has_combined = any("combined" in s and "view" in s for s in sheet_names_lower)
    check("Has Combined View sheet", has_combined,
          f"Found sheets: {wb.sheetnames}")

    ws_combined = None
    for s in wb.sheetnames:
        sl = s.lower().replace("_", " ")
        if "combined" in sl and "view" in sl:
            ws_combined = wb[s]
            break

    if ws_combined:
        data_rows = list(ws_combined.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]
        check("Combined View has at least 5 rows",
              len(data_rows) >= 5,
              f"Found {len(data_rows)} rows")


def check_gsheet():
    """Check Google Sheet Pipeline Dashboard."""
    print("\n=== Checking Google Sheet ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets ORDER BY created_at DESC")
    spreadsheets = cur.fetchall()

    matching_ss = None
    for ss_id, ss_title in spreadsheets:
        title_lower = ss_title.lower()
        if "pipeline" in title_lower or "dashboard" in title_lower:
            matching_ss = (ss_id, ss_title)
            break

    check("Pipeline Dashboard spreadsheet exists", matching_ss is not None,
          f"Found {len(spreadsheets)} spreadsheets, none matching 'pipeline' or 'dashboard'")

    if matching_ss:
        ss_id, ss_title = matching_ss
        print(f"  Found: {ss_title}")

        cur.execute("""
            SELECT row_index, col_index, value
            FROM gsheet.cells c
            JOIN gsheet.sheets s ON c.spreadsheet_id = s.spreadsheet_id AND c.sheet_id = s.id
            WHERE c.spreadsheet_id = %s
            ORDER BY row_index, col_index
        """, (ss_id,))
        cells = cur.fetchall()

        # Count data rows (excluding header)
        data_rows = set()
        for row_idx, col_idx, value in cells:
            if row_idx > 0 and value is not None:
                data_rows.add(row_idx)

        check("Google Sheet has at least 4 data rows",
              len(data_rows) >= 4,
              f"Found {len(data_rows)} data rows")

        # Check product names appear in the sheet
        all_values = " ".join(str(v).lower() for _, _, v in cells if v)
        products_found = sum(1 for p in ["widget", "gadget", "sensor", "connector"]
                             if p in all_values)
        check("Google Sheet contains at least 4 product references",
              products_found >= 4,
              f"Found {products_found}/4 product name keywords")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = compute_expected(args.groundtruth_workspace)

    check_script(args.agent_workspace)
    check_cleaned_data(args.agent_workspace, expected)
    check_excel(args.agent_workspace, expected)
    check_gsheet()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")

    result = {
        "passed": PASS_COUNT,
        "failed": FAIL_COUNT,
        "pass_rate": round(pass_rate, 3),
        "success": pass_rate >= 0.7,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if pass_rate >= 0.7 else 1)


if __name__ == "__main__":
    main()

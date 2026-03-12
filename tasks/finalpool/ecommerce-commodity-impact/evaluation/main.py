#!/usr/bin/env python3
"""
Evaluation script for ecommerce-commodity-impact task.

Queries PostgreSQL directly to compute expected values, then compares
against the agent's Excel output and memory.json.
"""

import os
import sys
import json
import math
from argparse import ArgumentParser
from pathlib import Path
from collections import defaultdict
from decimal import Decimal

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "database": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)


def compute_expected_category_sales():
    """
    Compute expected category sales from completed WooCommerce orders.
    Returns a dict: {category_name: {"revenue": float, "order_count": int, "avg_order_value": float}}
    """
    conn = get_db_connection()
    cur = conn.cursor()

    # Get product -> category mapping
    cur.execute("SELECT id, categories FROM wc.products")
    prod_cat = {}
    for pid, cats in cur.fetchall():
        if cats and len(cats) > 0:
            prod_cat[pid] = cats[0]["name"]
        else:
            prod_cat[pid] = "Uncategorized"

    # Get completed orders and their line items
    cur.execute("SELECT id, line_items FROM wc.orders WHERE status = 'completed'")
    cat_revenue = defaultdict(float)
    cat_orders = defaultdict(set)

    for oid, items in cur.fetchall():
        if not items:
            continue
        for item in items:
            pid = item.get("product_id")
            qty = item.get("quantity", 0)
            price = float(item.get("price", 0))
            cat = prod_cat.get(pid, "Unknown")
            cat_revenue[cat] += qty * price
            cat_orders[cat].add(oid)

    conn.close()

    result = {}
    for cat in cat_revenue:
        rev = cat_revenue[cat]
        cnt = len(cat_orders[cat])
        aov = round(rev / cnt, 2) if cnt > 0 else 0.0
        result[cat] = {
            "revenue": round(rev, 2),
            "order_count": cnt,
            "avg_order_value": aov,
        }
    return result


def compute_expected_commodity_trends():
    """
    Compute expected commodity/market price trends.
    Returns a dict: {symbol: {"name": str, "price_6m_ago": float, "latest_price": float, "change_pct": float}}
    """
    conn = get_db_connection()
    cur = conn.cursor()

    symbols = {
        "GC=F": "Gold Futures",
        "XOM": "Exxon Mobil",
        "^DJI": "Dow Jones",
    }

    result = {}
    for symbol, name in symbols.items():
        # Price ~6 months ago (closest trading day on or before 2025-09-05)
        cur.execute(
            "SELECT close FROM yf.stock_prices WHERE symbol = %s AND date <= '2025-09-05' ORDER BY date DESC LIMIT 1",
            (symbol,),
        )
        row = cur.fetchone()
        price_6m_ago = float(row[0]) if row else None

        # Latest price (closest trading day on or before 2026-03-05)
        cur.execute(
            "SELECT close FROM yf.stock_prices WHERE symbol = %s AND date <= '2026-03-05' ORDER BY date DESC LIMIT 1",
            (symbol,),
        )
        row = cur.fetchone()
        latest_price = float(row[0]) if row else None

        change_pct = None
        if price_6m_ago and latest_price and price_6m_ago != 0:
            change_pct = round((latest_price - price_6m_ago) / price_6m_ago * 100, 2)

        result[symbol] = {
            "name": name,
            "price_6m_ago": price_6m_ago,
            "latest_price": latest_price,
            "change_pct": change_pct,
        }

    conn.close()
    return result


# ---------------------------------------------------------------------------
# Excel reading helpers
# ---------------------------------------------------------------------------

def read_excel_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts using the first row as headers."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return data


def safe_float(val):
    """Convert a value to float safely."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Evaluation checks
# ---------------------------------------------------------------------------

def check_excel_categories(wb, expected):
    """Check the 'Category Sales' sheet against expected data."""
    data = read_excel_sheet(wb, "Category Sales")
    if data is None:
        return False, "Sheet 'Category Sales' not found in workbook"

    if len(data) == 0:
        return False, "Sheet 'Category Sales' is empty"

    errors = []
    found_categories = set()

    for row in data:
        cat = str(row.get("Category", "")).strip()
        if not cat:
            continue
        found_categories.add(cat)

        if cat not in expected:
            errors.append(f"Unexpected category '{cat}' in sheet")
            continue

        exp = expected[cat]

        # Check Total_Revenue
        agent_rev = safe_float(row.get("Total_Revenue"))
        if agent_rev is None:
            errors.append(f"Category '{cat}': Total_Revenue is missing or not numeric")
        elif abs(agent_rev - exp["revenue"]) > 1.0:
            errors.append(
                f"Category '{cat}': Total_Revenue mismatch: agent={agent_rev}, expected={exp['revenue']}"
            )

        # Check Order_Count
        agent_cnt = safe_float(row.get("Order_Count"))
        if agent_cnt is None:
            errors.append(f"Category '{cat}': Order_Count is missing or not numeric")
        elif int(agent_cnt) != exp["order_count"]:
            errors.append(
                f"Category '{cat}': Order_Count mismatch: agent={int(agent_cnt)}, expected={exp['order_count']}"
            )

        # Check Avg_Order_Value
        agent_aov = safe_float(row.get("Avg_Order_Value"))
        if agent_aov is None:
            errors.append(f"Category '{cat}': Avg_Order_Value is missing or not numeric")
        elif abs(agent_aov - exp["avg_order_value"]) > 1.0:
            errors.append(
                f"Category '{cat}': Avg_Order_Value mismatch: agent={agent_aov}, expected={exp['avg_order_value']}"
            )

    # Check for missing categories
    missing = set(expected.keys()) - found_categories
    if missing:
        errors.append(f"Missing categories: {missing}")

    # Check sort order (by revenue descending)
    revenues = []
    for row in data:
        r = safe_float(row.get("Total_Revenue"))
        if r is not None:
            revenues.append(r)
    if revenues and revenues != sorted(revenues, reverse=True):
        errors.append("Rows are not sorted by Total_Revenue in descending order")

    if errors:
        return False, "Category Sales errors:\n  " + "\n  ".join(errors)
    return True, f"Category Sales: all {len(expected)} categories verified correctly"


def check_excel_commodities(wb, expected):
    """Check the 'Commodity Trends' sheet against expected data."""
    data = read_excel_sheet(wb, "Commodity Trends")
    if data is None:
        return False, "Sheet 'Commodity Trends' not found in workbook"

    if len(data) == 0:
        return False, "Sheet 'Commodity Trends' is empty"

    errors = []
    found_symbols = set()

    for row in data:
        symbol = str(row.get("Symbol", "")).strip()
        if not symbol:
            continue
        found_symbols.add(symbol)

        if symbol not in expected:
            errors.append(f"Unexpected symbol '{symbol}' in sheet")
            continue

        exp = expected[symbol]

        # Check Name
        agent_name = str(row.get("Name", "")).strip()
        if agent_name.lower() != exp["name"].lower():
            errors.append(
                f"Symbol '{symbol}': Name mismatch: agent='{agent_name}', expected='{exp['name']}'"
            )

        # Check Price_6M_Ago
        agent_p6m = safe_float(row.get("Price_6M_Ago"))
        if agent_p6m is None:
            errors.append(f"Symbol '{symbol}': Price_6M_Ago is missing or not numeric")
        elif exp["price_6m_ago"] and abs(agent_p6m - exp["price_6m_ago"]) > 5.0:
            errors.append(
                f"Symbol '{symbol}': Price_6M_Ago mismatch: agent={agent_p6m}, expected={exp['price_6m_ago']}"
            )

        # Check Latest_Price
        agent_lp = safe_float(row.get("Latest_Price"))
        if agent_lp is None:
            errors.append(f"Symbol '{symbol}': Latest_Price is missing or not numeric")
        elif exp["latest_price"] and abs(agent_lp - exp["latest_price"]) > 5.0:
            errors.append(
                f"Symbol '{symbol}': Latest_Price mismatch: agent={agent_lp}, expected={exp['latest_price']}"
            )

        # Check Change_Pct
        agent_cp = safe_float(row.get("Change_Pct"))
        if agent_cp is None:
            errors.append(f"Symbol '{symbol}': Change_Pct is missing or not numeric")
        elif exp["change_pct"] is not None and abs(agent_cp - exp["change_pct"]) > 0.5:
            errors.append(
                f"Symbol '{symbol}': Change_Pct mismatch: agent={agent_cp}, expected={exp['change_pct']}"
            )

    missing = set(expected.keys()) - found_symbols
    if missing:
        errors.append(f"Missing symbols: {missing}")

    if errors:
        return False, "Commodity Trends errors:\n  " + "\n  ".join(errors)
    return True, f"Commodity Trends: all {len(expected)} symbols verified correctly"


def check_excel_summary(wb, expected_cats, expected_commodities):
    """Check the 'Analysis Summary' sheet against expected computed values."""
    data = read_excel_sheet(wb, "Analysis Summary")
    if data is None:
        return False, "Sheet 'Analysis Summary' not found in workbook"

    if len(data) == 0:
        return False, "Sheet 'Analysis Summary' is empty"

    # Build metric -> value dict from agent's sheet
    agent_metrics = {}
    for row in data:
        metric = str(row.get("Metric", "")).strip()
        value = row.get("Value")
        if metric:
            agent_metrics[metric] = value

    # Compute expected summary values
    sorted_cats = sorted(expected_cats.items(), key=lambda x: x[1]["revenue"], reverse=True)
    top_cat = sorted_cats[0]
    bottom_cat = sorted_cats[-1]
    total_rev = sum(v["revenue"] for v in expected_cats.values())

    gold_pct = expected_commodities["GC=F"]["change_pct"]
    energy_pct = expected_commodities["XOM"]["change_pct"]
    market_pct = expected_commodities["^DJI"]["change_pct"]

    if market_pct is not None:
        if market_pct > 5:
            market_trend = "Bullish"
        elif market_pct < -5:
            market_trend = "Bearish"
        else:
            market_trend = "Neutral"
    else:
        market_trend = "Unknown"

    expected_summary = {
        "Top_Category": ("str", top_cat[0]),
        "Top_Category_Revenue": ("num", round(top_cat[1]["revenue"], 2)),
        "Bottom_Category": ("str", bottom_cat[0]),
        "Bottom_Category_Revenue": ("num", round(bottom_cat[1]["revenue"], 2)),
        "Total_Store_Revenue": ("num", round(total_rev, 2)),
        "Gold_6M_Change_Pct": ("num", gold_pct),
        "Energy_6M_Change_Pct": ("num", energy_pct),
        "Market_6M_Change_Pct": ("num", market_pct),
        "Market_Trend": ("str", market_trend),
    }

    errors = []

    for metric, (vtype, exp_val) in expected_summary.items():
        agent_val = agent_metrics.get(metric)
        if agent_val is None:
            errors.append(f"Missing metric: '{metric}'")
            continue

        if vtype == "str":
            agent_str = str(agent_val).strip()
            if agent_str.lower() != str(exp_val).lower():
                errors.append(
                    f"Metric '{metric}': mismatch: agent='{agent_str}', expected='{exp_val}'"
                )
        elif vtype == "num":
            agent_num = safe_float(agent_val)
            if agent_num is None:
                errors.append(f"Metric '{metric}': value is not numeric: '{agent_val}'")
            elif exp_val is not None:
                # Use tolerance based on the metric type
                if "Pct" in metric:
                    tol = 0.5
                elif "Revenue" in metric:
                    tol = 1.0
                else:
                    tol = 1.0
                if abs(agent_num - exp_val) > tol:
                    errors.append(
                        f"Metric '{metric}': mismatch: agent={agent_num}, expected={exp_val} (tol={tol})"
                    )

    if errors:
        return False, "Analysis Summary errors:\n  " + "\n  ".join(errors)
    return True, f"Analysis Summary: all {len(expected_summary)} metrics verified correctly"


def check_memory(agent_workspace):
    """Check that memory.json exists and contains entities."""
    memory_file = Path(agent_workspace) / "memory" / "memory.json"
    if not memory_file.exists():
        return False, f"memory/memory.json not found at {memory_file}"

    try:
        with open(memory_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, Exception) as e:
        return False, f"memory/memory.json is not valid JSON: {e}"

    entities = data.get("entities", [])
    if not entities or len(entities) == 0:
        return False, "memory/memory.json contains no entities (expected the agent to store findings)"

    return True, f"Memory check passed: {len(entities)} entities found"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_evaluation(agent_workspace: str) -> tuple:
    """Run all evaluation checks and return (passed, message)."""
    print("=" * 80)
    print("ECOMMERCE COMMODITY IMPACT - EVALUATION")
    print("=" * 80)

    # ---- Compute expected values from DB ----
    print("\nComputing expected values from database...")
    try:
        expected_cats = compute_expected_category_sales()
        print(f"  Found {len(expected_cats)} categories with completed order data")
        for cat, vals in sorted(expected_cats.items(), key=lambda x: x[1]["revenue"], reverse=True):
            print(f"    {cat}: revenue={vals['revenue']}, orders={vals['order_count']}, aov={vals['avg_order_value']}")
    except Exception as e:
        return False, f"Failed to compute expected category sales: {e}"

    try:
        expected_commodities = compute_expected_commodity_trends()
        print(f"  Found {len(expected_commodities)} commodity/market symbols")
        for sym, vals in expected_commodities.items():
            print(f"    {sym}: 6m_ago={vals['price_6m_ago']}, latest={vals['latest_price']}, change={vals['change_pct']}%")
    except Exception as e:
        return False, f"Failed to compute expected commodity trends: {e}"

    # ---- Find the Excel file ----
    ws = Path(agent_workspace)
    excel_path = ws / "Commodity_Impact_Report.xlsx"
    if not excel_path.exists():
        # Try case-insensitive search
        for f in ws.iterdir():
            if f.name.lower() == "commodity_impact_report.xlsx":
                excel_path = f
                break

    if not excel_path.exists():
        return False, f"Excel file 'Commodity_Impact_Report.xlsx' not found in {agent_workspace}"

    print(f"\nLoading Excel file: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"  Sheets found: {wb.sheetnames}")
    except Exception as e:
        return False, f"Failed to open Excel file: {e}"

    # ---- Run checks ----
    results = []

    print("\n--- Check 1: Category Sales ---")
    passed, msg = check_excel_categories(wb, expected_cats)
    results.append(("Category Sales", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    print("\n--- Check 2: Commodity Trends ---")
    passed, msg = check_excel_commodities(wb, expected_commodities)
    results.append(("Commodity Trends", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    print("\n--- Check 3: Analysis Summary ---")
    passed, msg = check_excel_summary(wb, expected_cats, expected_commodities)
    results.append(("Analysis Summary", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    print("\n--- Check 4: Memory ---")
    passed, msg = check_memory(agent_workspace)
    results.append(("Memory", passed, msg))
    print(f"  {'PASSED' if passed else 'FAILED'}: {msg}")

    wb.close()

    # ---- Aggregate results ----
    total = len(results)
    passed_count = sum(1 for _, p, _ in results if p)
    all_passed = passed_count == total

    summary_lines = [
        "\n" + "=" * 80,
        "EVALUATION SUMMARY",
        "=" * 80,
    ]
    for name, p, msg in results:
        status = "PASSED" if p else "FAILED"
        summary_lines.append(f"  [{status}] {name}: {msg}")

    summary_lines.append(f"\nOverall: {passed_count}/{total} checks passed")
    if all_passed:
        summary_lines.append("ALL CHECKS PASSED")
    else:
        summary_lines.append("SOME CHECKS FAILED")

    summary = "\n".join(summary_lines)
    return all_passed, summary


def main():
    parser = ArgumentParser(description="Evaluate Ecommerce Commodity Impact Task")
    parser.add_argument("--agent_workspace", required=True,
                        help="Path to agent's workspace")
    parser.add_argument("--groundtruth_workspace", required=False,
                        help="Path to ground truth workspace (not used)")
    parser.add_argument("--res_log_file", required=False,
                        help="Path to result log file (optional)")
    parser.add_argument("--launch_time", required=False,
                        help="Launch time (optional)")
    args = parser.parse_args()

    print(f"Agent workspace: {args.agent_workspace}")

    try:
        success, message = run_evaluation(args.agent_workspace)
        print(message)

        if args.res_log_file:
            try:
                eval_log = os.path.join(
                    os.path.dirname(args.res_log_file), "eval_result.txt"
                )
                with open(eval_log, "w", encoding="utf-8") as f:
                    f.write(f"Ecommerce Commodity Impact Evaluation\n")
                    f.write(f"{'=' * 80}\n")
                    f.write(f"Agent Workspace: {args.agent_workspace}\n")
                    if args.launch_time:
                        f.write(f"Launch Time: {args.launch_time}\n")
                    f.write(f"{'=' * 80}\n")
                    f.write(message)
                    f.write(f"\n{'=' * 80}\n")
                    f.write(f"Result: {'PASSED' if success else 'FAILED'}\n")
            except Exception as e:
                print(f"Warning: Could not write to log file: {e}")

        if success:
            print("\nEVALUATION PASSED")
            sys.exit(0)
        else:
            print("\nEVALUATION FAILED")
            sys.exit(1)

    except Exception as e:
        print(f"Critical evaluation error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

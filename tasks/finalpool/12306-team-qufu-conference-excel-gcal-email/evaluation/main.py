"""
Evaluation for 12306-team-qufu-conference-excel-gcal-email.

Checks:
1. Conference_Travel_Plan.xlsx exists
2. "Outbound" sheet has at least 4 data rows
3. "Outbound" sheet has Train_No column containing G235 or G168
4. "Return" sheet has at least 4 data rows
5. "Coordination_Notes" sheet has at least 3 rows
6. GCal has at least 2 new travel events (beyond the conference itself)
7. Email sent to beijing_team@uni.edu
8. Email sent to shanghai_team@uni.edu
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Conference_Travel_Plan.xlsx ===")
    import glob

    pattern = os.path.join(agent_workspace, "*.xlsx")
    all_xlsx = glob.glob(pattern)
    conf_files = [f for f in all_xlsx if any(
        kw in os.path.basename(f).lower()
        for kw in ["conference", "travel", "qufu", "plan"]
    )]

    if not conf_files:
        record("Conference travel xlsx exists", False,
               f"No matching xlsx in {agent_workspace}")
        return
    record("Conference travel xlsx exists", True)

    xlsx_path = conf_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Outbound sheet
    if "outbound" not in sheet_names_lower:
        record("Outbound sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Outbound sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("outbound")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Outbound has at least 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} data rows")

        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        has_train = "g235" in all_text or "g168" in all_text
        record("Outbound contains G235 or G168 train number", has_train,
               f"Text: {all_text[:200]}")

    # Check Return sheet
    if "return" not in sheet_names_lower:
        record("Return sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Return sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("return")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Return has at least 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} data rows")

    # Check Coordination_Notes sheet
    coord_match = [s for s in sheet_names_lower if "coord" in s or "note" in s]
    if not coord_match:
        record("Coordination_Notes sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Coordination_Notes sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(coord_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Coordination_Notes has at least 3 rows", len(data_rows) >= 3,
               f"Found {len(data_rows)} data rows")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Conference_Travel_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        record(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices.append(len(gt_rows) - 1)
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        record(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1}",
                               False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    record(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                record(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: Calendar travel events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE (start_datetime >= '2026-03-12' AND start_datetime < '2026-03-16')
          AND summary NOT ILIKE '%confucian studies conference%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()
    record("At least 2 new travel calendar events", len(events) >= 2,
           f"Found {len(events)} events: {[e[0] for e in events]}")


def _query_email_count(addr_pattern):
    """Query email count using a fresh connection to avoid transaction aborts."""
    cnt = 0
    sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s"
            " AND from_addr NOT ILIKE %s",
            (f"%{addr_pattern}%", f"%{addr_pattern}%"),
        )
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log WHERE to_addr::text ILIKE %s",
            (f"%{addr_pattern}%",),
        )
        sent = cur2.fetchone()[0]
        cur2.close()
        conn2.close()
    except Exception:
        pass
    return cnt, sent


def check_emails():
    print("\n=== Check 3: Emails to teams ===")

    beijing_cnt, beijing_sent = _query_email_count("beijing_team@uni.edu")
    record("Email sent to beijing_team@uni.edu", beijing_cnt >= 1 or beijing_sent >= 1,
           f"messages: {beijing_cnt}, sent_log: {beijing_sent}")

    shanghai_cnt, shanghai_sent = _query_email_count("shanghai_team@uni.edu")
    record("Email sent to shanghai_team@uni.edu", shanghai_cnt >= 1 or shanghai_sent >= 1,
           f"messages: {shanghai_cnt}, sent_log: {shanghai_sent}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

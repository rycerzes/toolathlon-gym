"""Evaluation for wc-customer-loyalty-tier-excel-email."""
import argparse
import os
import sys
import psycopg2
import openpyxl


DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_tier_data_from_db():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT first_name, last_name, email, total_spent, orders_count FROM wc.customers ORDER BY total_spent DESC")
    customers = cur.fetchall()
    cur.close()
    conn.close()

    vip = [(f"{r[0]} {r[1]}", r[2], float(r[3]), r[4]) for r in customers if float(r[3]) >= 2000]
    gold = [(f"{r[0]} {r[1]}", r[2], float(r[3]), r[4]) for r in customers if 500 <= float(r[3]) < 2000]
    std = [(f"{r[0]} {r[1]}", r[2], float(r[3]), r[4]) for r in customers if float(r[3]) < 500]

    vip_rev = round(sum(c[2] for c in vip), 2)
    gold_rev = round(sum(c[2] for c in gold), 2)
    std_rev = round(sum(c[2] for c in std), 2)

    return {
        "vip": vip,
        "gold": gold,
        "std": std,
        "vip_rev": vip_rev,
        "gold_rev": gold_rev,
        "std_rev": std_rev,
        "vip_avg": round(vip_rev / len(vip), 2) if vip else 0,
        "gold_avg": round(gold_rev / len(gold), 2) if gold else 0,
        "std_avg": round(std_rev / len(std), 2) if std else 0,
    }


def check_gsheet_exists():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%customer loyalty%' OR LOWER(title) LIKE '%loyalty program%'")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return len(rows) > 0
    except Exception:
        return False


def check_vip_emails_sent(vip_count):
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM email.messages WHERE LOWER(subject) LIKE '%vip%'")
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cnt >= vip_count
    except Exception:
        return False


def check_notion_page_exists():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id FROM notion.pages WHERE LOWER(properties::text) LIKE '%customer loyalty program 2026%' OR LOWER(properties::text) LIKE '%loyalty program 2026%'")
        rows = cur.fetchall()
        if rows:
            cur.close()
            conn.close()
            return True
        # Also check blocks for the title
        cur.execute("SELECT id FROM notion.blocks WHERE LOWER(block_data::text) LIKE '%customer loyalty program 2026%'")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return len(rows) > 0
    except Exception:
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Customer_Loyalty_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Customer_Loyalty_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    all_errors = []

    # Load DB data for validation
    try:
        tier_data = get_tier_data_from_db()
    except Exception as e:
        print(f"WARNING: Could not query DB: {e}")
        tier_data = None

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check sheet: Customer Tiers
    print("  Checking Customer Tiers sheet...")
    a_rows = load_sheet_rows(agent_wb, "Customer Tiers")
    g_rows = load_sheet_rows(gt_wb, "Customer Tiers")
    if a_rows is None:
        all_errors.append("Sheet 'Customer Tiers' not found in agent output")
    else:
        data_rows = [r for r in a_rows[1:] if r and any(c is not None for c in r)]
        if len(data_rows) < 10:
            all_errors.append(f"Customer Tiers has too few rows: {len(data_rows)} (expected >= 10)")
        else:
            print(f"    PASS ({len(data_rows)} data rows)")
            # Verify VIP customers have Total_Spent >= 2000
            vip_check_errors = 0
            for row in data_rows:
                if row and len(row) >= 5 and str(row[4]).strip().lower() == 'vip':
                    if row[2] is not None:
                        try:
                            if float(row[2]) < 2000:
                                vip_check_errors += 1
                        except (TypeError, ValueError):
                            pass
            if vip_check_errors > 0:
                all_errors.append(f"Found {vip_check_errors} VIP customers with Total_Spent < 2000")
            else:
                print("    VIP tier validation PASS")

    # Check sheet: Tier Summary
    print("  Checking Tier Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Tier Summary")
    g_rows = load_sheet_rows(gt_wb, "Tier Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Tier Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Tier Summary' not found in groundtruth")
    else:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        if len(a_data) < 3:
            all_errors.append(f"Tier Summary has {len(a_data)} rows, expected 3 (VIP, Gold, Standard)")
        else:
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
            errors = []
            for tier_name, exp_count, exp_rev, exp_avg in [
                ("vip", tier_data["vip"] if tier_data else None, tier_data["vip_rev"] if tier_data else None, tier_data["vip_avg"] if tier_data else None),
                ("gold", tier_data["gold"] if tier_data else None, tier_data["gold_rev"] if tier_data else None, tier_data["gold_avg"] if tier_data else None),
                ("standard", tier_data["std"] if tier_data else None, tier_data["std_rev"] if tier_data else None, tier_data["std_avg"] if tier_data else None),
            ]:
                a_row = a_lookup.get(tier_name)
                if a_row is None:
                    errors.append(f"Missing tier row: {tier_name}")
                    continue
                if exp_count is not None and len(a_row) > 1:
                    if not num_close(a_row[1], len(exp_count), 0):
                        errors.append(f"{tier_name}.Customer_Count: {a_row[1]} vs {len(exp_count)}")
                if exp_rev is not None and len(a_row) > 2:
                    if not num_close(a_row[2], exp_rev, 5.0):
                        errors.append(f"{tier_name}.Total_Revenue: {a_row[2]} vs {exp_rev} (tol=5.0)")
                if exp_avg is not None and len(a_row) > 3:
                    if not num_close(a_row[3], exp_avg, 5.0):
                        errors.append(f"{tier_name}.Avg_Spent: {a_row[3]} vs {exp_avg} (tol=5.0)")
            if errors:
                all_errors.extend(errors)
                for e in errors[:5]:
                    print(f"    ERROR: {e}")
            else:
                print("    PASS")

    # Check GSheet exists
    print("  Checking Google Sheet...")
    if check_gsheet_exists():
        print("    PASS")
    else:
        all_errors.append("Google Sheet 'Customer Loyalty Program' not found in DB")

    # Check VIP emails sent
    print("  Checking VIP emails sent...")
    vip_count = len(tier_data["vip"]) if tier_data else 6
    if check_vip_emails_sent(vip_count):
        print(f"    PASS (at least {vip_count} VIP emails)")
    else:
        all_errors.append(f"Expected >= {vip_count} VIP emails, not found in email.messages")

    # Check Notion page
    print("  Checking Notion page...")
    if check_notion_page_exists():
        print("    PASS")
    else:
        all_errors.append("Notion page 'Customer Loyalty Program 2026' not found")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

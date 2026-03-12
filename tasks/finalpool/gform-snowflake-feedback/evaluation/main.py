"""
Evaluation script for gform-snowflake-feedback task.

Checks:
1. Excel file (Support_Analysis.xlsx) - 2 sheets with correct data from Snowflake
2. Google Form created with correct structure (queried from gform schema)
3. Word document (Improvement_Report.docx) with required sections and content
4. Email sent to quality-team@company.com with correct subject and content

Expected values are computed at evaluation time from the PostgreSQL database.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=5.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    """
    Query the PostgreSQL database to compute expected values.
    Returns dicts for issue_summary, priority_breakdown, and top issue types.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # ---- Issue Summary (all issue types, sorted by count desc) ----
    cur.execute("""
        SELECT
            "ISSUE_TYPE",
            COUNT(*) AS ticket_count,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) AS avg_response_hours,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) AS avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE"
        ORDER BY COUNT(*) DESC
    """)
    issue_rows = cur.fetchall()
    issue_summary = []
    for row in issue_rows:
        issue_summary.append({
            "Issue_Type": row[0],
            "Ticket_Count": int(row[1]),
            "Avg_Response_Hours": float(row[2]),
            "Avg_Satisfaction": float(row[3]),
        })

    # Top 5 issue type names
    top5_issue_types = [r["Issue_Type"] for r in issue_summary[:5]]

    # ---- Priority Breakdown ----
    cur.execute("""
        SELECT
            "PRIORITY",
            COUNT(*) AS cnt,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) AS avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY"
        ORDER BY COUNT(*) DESC
    """)
    priority_rows = cur.fetchall()
    priority_breakdown = []
    for row in priority_rows:
        priority_breakdown.append({
            "Priority": row[0],
            "Count": int(row[1]),
            "Avg_Satisfaction": float(row[2]),
        })

    # Total tickets
    cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
    total_tickets = cur.fetchone()[0]

    cur.close()
    conn.close()

    return {
        "issue_summary": issue_summary,
        "priority_breakdown": priority_breakdown,
        "top5_issue_types": top5_issue_types,
        "total_tickets": total_tickets,
    }


def get_sheet(wb, name):
    """Find sheet case-insensitively."""
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    """Check the Excel output file against computed expected values."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Support_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Excel file readable", True)

    # Check sheet names
    check("Sheet 'Issue Summary' exists",
          any(str_match(s, "Issue Summary") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")
    check("Sheet 'Priority Breakdown' exists",
          any(str_match(s, "Priority Breakdown") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")

    # --- Sheet 1: Issue Summary ---
    print("\n--- Issue Summary ---")
    ws = get_sheet(wb, "Issue Summary")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp_rows = expected["issue_summary"]
        check("Issue Summary row count", len(agent_rows) == len(exp_rows),
              f"Expected {len(exp_rows)}, got {len(agent_rows)}")

        # Check sorted by Ticket_Count descending
        if len(agent_rows) >= 2:
            try:
                first_count = float(agent_rows[0][1]) if agent_rows[0][1] else 0
                second_count = float(agent_rows[1][1]) if agent_rows[1][1] else 0
                check("Issue Summary sorted by Ticket_Count desc",
                      first_count >= second_count,
                      f"First: {first_count}, Second: {second_count}")
            except (TypeError, ValueError):
                check("Issue Summary sorted by Ticket_Count desc", False,
                      "Could not parse count values")

        for exp_row in exp_rows:
            issue_type = exp_row["Issue_Type"]
            matched = None
            for ar in agent_rows:
                if ar and ar[0] and issue_type.lower() in str(ar[0]).lower():
                    matched = ar
                    break
            if matched:
                check(f"Issue '{issue_type}' Ticket_Count",
                      num_close(matched[1], exp_row["Ticket_Count"], 5),
                      f"Expected {exp_row['Ticket_Count']}, got {matched[1]}")
                check(f"Issue '{issue_type}' Avg_Response_Hours",
                      num_close(matched[2], exp_row["Avg_Response_Hours"], 0.5),
                      f"Expected {exp_row['Avg_Response_Hours']}, got {matched[2]}")
                check(f"Issue '{issue_type}' Avg_Satisfaction",
                      num_close(matched[3], exp_row["Avg_Satisfaction"], 0.5),
                      f"Expected {exp_row['Avg_Satisfaction']}, got {matched[3]}")
            else:
                check(f"Issue '{issue_type}' found", False,
                      "Issue type not in agent output")

    # --- Sheet 2: Priority Breakdown ---
    print("\n--- Priority Breakdown ---")
    ws = get_sheet(wb, "Priority Breakdown")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp_rows = expected["priority_breakdown"]
        check("Priority Breakdown row count", len(agent_rows) == len(exp_rows),
              f"Expected {len(exp_rows)}, got {len(agent_rows)}")

        for exp_row in exp_rows:
            priority = exp_row["Priority"]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], priority):
                    matched = ar
                    break
            if matched:
                check(f"Priority '{priority}' Count",
                      num_close(matched[1], exp_row["Count"], 5),
                      f"Expected {exp_row['Count']}, got {matched[1]}")
                check(f"Priority '{priority}' Avg_Satisfaction",
                      num_close(matched[2], exp_row["Avg_Satisfaction"], 0.5),
                      f"Expected {exp_row['Avg_Satisfaction']}, got {matched[2]}")
            else:
                check(f"Priority '{priority}' found", False,
                      "Priority not in agent output")


def check_form(expected):
    """Check that a Google Form was created with correct structure."""
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Find form with title containing relevant keywords
    cur.execute("""
        SELECT id, title, description
        FROM gform.forms
        WHERE LOWER(title) LIKE '%customer%'
           OR LOWER(title) LIKE '%survey%'
           OR LOWER(title) LIKE '%improvement%'
           OR LOWER(title) LIKE '%feedback%'
        ORDER BY created_at DESC
        LIMIT 1
    """)
    form_row = cur.fetchone()

    check("Google Form exists with relevant title",
          form_row is not None,
          "No form found with 'customer', 'survey', 'improvement', or 'feedback' in title")
    if not form_row:
        cur.close()
        conn.close()
        return

    form_id = form_row[0]
    form_title = form_row[1]
    check("Form title contains 'Customer Service Improvement Survey'",
          "customer service improvement survey" in form_title.lower(),
          f"Actual title: '{form_title}'")

    # Get questions
    cur.execute("""
        SELECT title, question_type, required, config
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()

    check("Form has at least 5 questions", len(questions) >= 5,
          f"Found {len(questions)} questions")

    # Check that all questions are required
    all_required = all(q[2] for q in questions)
    check("All questions are required", all_required,
          f"Found non-required questions")

    # Check for email question (text type)
    has_email_q = any(
        "email" in q[0].lower() and q[1] in ("textQuestion", "TEXT")
        for q in questions
    )
    check("Has email text question", has_email_q)

    # Check for issue type radio question with top 5 issues
    top5 = expected["top5_issue_types"]
    has_issue_radio = False
    for q in questions:
        title_lower = q[0].lower()
        if ("issue" in title_lower or "experience" in title_lower or "type" in title_lower) and q[1] in ("choiceQuestion", "RADIO"):
            # Check that options contain the top 5 issue types
            config = q[3] if isinstance(q[3], dict) else {}
            options = []
            if "options" in config:
                for opt in config["options"]:
                    if isinstance(opt, dict) and "value" in opt:
                        options.append(opt["value"].lower())
                    elif isinstance(opt, str):
                        options.append(opt.lower())
            # Check at least 3 of the top 5 are present
            matched_count = sum(
                1 for t in top5
                if any(t.lower() in o for o in options)
            )
            if matched_count >= 3:
                has_issue_radio = True
                break
    check("Has radio question with top issue types", has_issue_radio)

    # Check for rating question (1-5 scale)
    has_rating = any(
        ("rate" in q[0].lower() or "experience" in q[0].lower() or "satisfaction" in q[0].lower() or "overall" in q[0].lower())
        and q[1] in ("choiceQuestion", "RADIO")
        for q in questions
    )
    check("Has rating question (1-5)", has_rating)

    # Check for open-ended suggestion question
    has_suggestions = any(
        ("suggestion" in q[0].lower() or "improvement" in q[0].lower() or "feedback" in q[0].lower() or "comment" in q[0].lower())
        and q[1] in ("textQuestion", "TEXT")
        for q in questions
    )
    check("Has open-ended suggestion text question", has_suggestions)

    # Check for checkbox question about issue types
    has_checkbox = any(
        q[1] in ("checkboxQuestion", "CHECKBOX")
        for q in questions
    )
    check("Has checkbox question for selecting issue types", has_checkbox)

    cur.close()
    conn.close()


def check_word(agent_workspace, expected):
    """Check the Word document for required sections and content."""
    print("\n=== Checking Word Document ===")

    doc_path = os.path.join(agent_workspace, "Improvement_Report.docx")
    check("Word file exists", os.path.isfile(doc_path),
          f"Expected {doc_path}")
    if not os.path.isfile(doc_path):
        return

    try:
        from docx import Document
        doc = Document(doc_path)
    except ImportError:
        # Fallback: use zipfile to read raw XML
        import zipfile
        import xml.etree.ElementTree as ET
        check("python-docx available", False, "Falling back to XML parsing")
        with zipfile.ZipFile(doc_path) as z:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                all_text = " ".join(
                    node.text for node in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
                    if node.text
                ).lower()

                check("Doc contains 'Executive Summary'",
                      "executive summary" in all_text)
                check("Doc contains 'Issue Analysis'",
                      "issue analysis" in all_text)
                check("Doc contains 'Recommendations'",
                      "recommendation" in all_text)
                check("Doc contains 'Conclusion'",
                      "conclusion" in all_text)

                # Check top issue types mentioned
                for issue_type in expected["top5_issue_types"][:3]:
                    check(f"Doc mentions '{issue_type}'",
                          issue_type.lower() in all_text,
                          f"Issue type not found in document text")

                # Check total tickets mentioned
                total_str = str(expected["total_tickets"])
                check("Doc mentions total ticket count",
                      total_str in all_text,
                      f"Expected '{total_str}' in document")
        return
    except Exception as e:
        check("Word file readable", False, str(e))
        return

    check("Word file readable", True)

    # Extract headings
    headings = []
    full_text = []
    for para in doc.paragraphs:
        if para.style and para.style.name and "Heading" in para.style.name:
            headings.append(para.text.strip())
        full_text.append(para.text)

    all_text = " ".join(full_text).lower()

    check("Document has at least 3 headings", len(headings) >= 3,
          f"Found {len(headings)} headings: {headings}")

    # Check required sections
    heading_text = " ".join(h.lower() for h in headings)
    check("Has 'Executive Summary' heading",
          "executive summary" in heading_text,
          f"Headings: {headings}")
    check("Has 'Issue Analysis' heading",
          "issue analysis" in heading_text or "issue" in heading_text,
          f"Headings: {headings}")
    check("Has 'Recommendations' heading",
          "recommendation" in heading_text,
          f"Headings: {headings}")
    check("Has 'Conclusion' heading",
          "conclusion" in heading_text,
          f"Headings: {headings}")

    # Check content mentions top issue types
    for issue_type in expected["top5_issue_types"][:3]:
        check(f"Doc mentions '{issue_type}'",
              issue_type.lower() in all_text,
              f"Issue type not found in document text")

    # Check total tickets mentioned
    total_str = str(expected["total_tickets"])
    check("Doc mentions total ticket count",
          total_str in all_text,
          f"Expected '{total_str}' in document")


def check_email(expected):
    """Check that email was sent to quality-team@company.com."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Look for email in the Sent folder (folder_id=2) to quality-team@company.com
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE folder_id = 2
        ORDER BY date DESC
    """)
    sent_emails = cur.fetchall()

    # Find relevant email
    target_email = None
    for email_row in sent_emails:
        subject, from_addr, to_addr, body = email_row
        to_str = json.dumps(to_addr).lower() if to_addr else ""
        if "quality-team@company.com" in to_str or "quality" in to_str:
            target_email = email_row
            break

    # If not found in Sent, check INBOX too (some MCP servers store in inbox)
    if not target_email:
        cur.execute("""
            SELECT subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE folder_id = 1
            ORDER BY date DESC
        """)
        inbox_emails = cur.fetchall()
        for email_row in inbox_emails:
            subject, from_addr, to_addr, body = email_row
            to_str = json.dumps(to_addr).lower() if to_addr else ""
            if "quality-team@company.com" in to_str or "quality" in to_str:
                target_email = email_row
                break

    # Also check all folders as a fallback
    if not target_email:
        cur.execute("""
            SELECT subject, from_addr, to_addr, body_text
            FROM email.messages
            ORDER BY date DESC
        """)
        all_emails = cur.fetchall()
        for email_row in all_emails:
            subject, from_addr, to_addr, body = email_row
            to_str = json.dumps(to_addr).lower() if to_addr else ""
            if "quality-team@company.com" in to_str or "quality" in to_str:
                target_email = email_row
                break

    check("Email to quality-team@company.com exists",
          target_email is not None,
          f"Found {len(sent_emails)} sent emails but none to quality-team@company.com")

    if target_email:
        subject, from_addr, to_addr, body = target_email

        # Check subject
        subject_lower = (subject or "").lower()
        check("Email subject contains 'Support Analysis Report'",
              "support" in subject_lower and "analysis" in subject_lower,
              f"Subject: '{subject}'")

        # Check body mentions key info
        body_lower = (body or "").lower()
        total_str = str(expected["total_tickets"])
        check("Email body mentions total tickets",
              total_str in body_lower,
              f"Expected '{total_str}' in email body")

        # Check top 3 issue types mentioned
        top3 = expected["top5_issue_types"][:3]
        mentioned_count = sum(
            1 for t in top3
            if t.lower() in body_lower
        )
        check("Email body mentions top 3 issue types",
              mentioned_count >= 2,
              f"Found {mentioned_count} of {len(top3)} top issue types in body")

    cur.close()
    conn.close()


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    # Compute expected values from database
    print("=== Computing Expected Values from Database ===")
    try:
        expected = compute_expected_values()
        print(f"  Computed issue summary for {len(expected['issue_summary'])} issue types")
        print(f"  Computed priority breakdown for {len(expected['priority_breakdown'])} priorities")
        print(f"  Top 5 issue types: {expected['top5_issue_types']}")
        print(f"  Total tickets: {expected['total_tickets']}")
    except Exception as e:
        print(f"  ERROR computing expected values: {e}")
        import traceback
        traceback.print_exc()
        return False, f"Failed to compute expected values: {e}"

    # Run all checks
    check_excel(agent_workspace, expected)
    check_form(expected)
    check_word(agent_workspace, expected)
    check_email(expected)

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    success = pass_rate >= 0.8

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return success, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Rate: {pass_rate:.1%}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

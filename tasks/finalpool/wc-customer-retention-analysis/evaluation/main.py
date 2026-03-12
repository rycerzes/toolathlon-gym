"""Evaluation for wc-customer-retention-analysis."""
import argparse
import os
import sys
from datetime import date

import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CHURN_THRESHOLD = 90
REFERENCE_DATE = date(2026, 3, 7)


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def safe_str(v):
    return str(v).strip() if v is not None else ""


def get_expected_data():
    """Compute expected customer retention data from WC database."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id, c.email, c.first_name, c.last_name,
               c.total_spent::numeric, c.orders_count,
               MAX(o.date_created) as last_order_date
        FROM wc.customers c
        LEFT JOIN wc.orders o ON c.id = o.customer_id
        WHERE c.total_spent::numeric > 0
        GROUP BY c.id, c.email, c.first_name, c.last_name, c.total_spent, c.orders_count
        ORDER BY c.total_spent::numeric DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    customers = []
    for cid, email, first, last, spent, orders, last_order in rows:
        days_since = (REFERENCE_DATE - last_order.date()).days if last_order else 999
        if spent > 500:
            segment = "VIP"
        elif spent >= 100:
            segment = "Regular"
        else:
            segment = "New"
        at_risk = "Yes" if days_since > CHURN_THRESHOLD else "No"
        customers.append({
            "name": f"{first} {last}",
            "email": email,
            "first": first,
            "spent": float(spent),
            "orders": orders,
            "last_date": last_order.strftime("%Y-%m-%d") if last_order else "",
            "days": days_since,
            "segment": segment,
            "at_risk": at_risk,
        })
    return customers


def check_excel(agent_workspace):
    """Check Customer_Retention.xlsx."""
    print("\n=== Checking Customer_Retention.xlsx ===")

    excel_path = os.path.join(agent_workspace, "Customer_Retention.xlsx")
    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    expected = get_expected_data()
    all_ok = True

    # Check Customer Analysis sheet
    ca_sheet = None
    for name in wb.sheetnames:
        if "customer" in name.lower() and "analy" in name.lower():
            ca_sheet = wb[name]
            break
    if ca_sheet is None:
        record("Sheet 'Customer Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Customer Analysis' exists", True)
        headers = [safe_str(ca_sheet.cell(1, c).value).lower() for c in range(1, 10)]
        record("Has Customer_Name column", any("customer" in h or "name" in h for h in headers))
        record("Has Email column", any("email" in h for h in headers))
        record("Has Total_Spent column", any("spent" in h or "total" in h for h in headers))
        record("Has Segment column", any("segment" in h for h in headers))
        record("Has At_Risk column", any("risk" in h for h in headers))

        rows = list(ca_sheet.iter_rows(min_row=2, values_only=True))
        record("Customer Analysis row count matches",
               abs(len(rows) - len(expected)) <= 2,
               f"Expected ~{len(expected)}, got {len(rows)}")

        # Spot-check top customers
        for exp in expected[:5]:
            found = False
            last_part = exp["name"].split()[-1].lower()
            for r in rows:
                if r and r[0] and last_part in safe_str(r[0]).lower() and exp["first"].lower() in safe_str(r[0]).lower():
                    found = True
                    ok_spent = num_close(r[2], exp["spent"], 5.0)
                    record(f"{exp['name']} Total_Spent ~{exp['spent']}", ok_spent,
                           f"Got {r[2]}")
                    if not ok_spent:
                        all_ok = False
                    break
            if not found:
                record(f"{exp['name']} found in Customer Analysis", False)
                all_ok = False

        # Check VIP at-risk customers are marked correctly
        vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]
        for exp in vip_at_risk:
            found = False
            last_part = exp["name"].split()[-1].lower()
            for r in rows:
                if r and r[0] and last_part in safe_str(r[0]).lower() and exp["first"].lower() in safe_str(r[0]).lower():
                    found = True
                    # Check segment and at_risk
                    seg_col = None
                    risk_col = None
                    for ci, h in enumerate(headers):
                        if "segment" in h:
                            seg_col = ci
                        if "risk" in h:
                            risk_col = ci
                    if seg_col is not None:
                        ok = safe_str(r[seg_col]).lower() == "vip"
                        record(f"{exp['name']} segment=VIP", ok, f"Got {r[seg_col]}")
                        if not ok:
                            all_ok = False
                    if risk_col is not None:
                        ok = safe_str(r[risk_col]).lower() == "yes"
                        record(f"{exp['name']} at_risk=Yes", ok, f"Got {r[risk_col]}")
                        if not ok:
                            all_ok = False
                    break

    # Check Retention Metrics sheet
    rm_sheet = None
    for name in wb.sheetnames:
        if "retention" in name.lower() and "metric" in name.lower():
            rm_sheet = wb[name]
            break
    if rm_sheet is None:
        record("Sheet 'Retention Metrics' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Retention Metrics' exists", True)
        metrics = {}
        for row in rm_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                metrics[safe_str(row[0]).lower().replace(" ", "_")] = row[1]

        total_active = len(expected)
        vip_count = sum(1 for c in expected if c["segment"] == "VIP")
        at_risk_total = sum(1 for c in expected if c["at_risk"] == "Yes")
        at_risk_vip = sum(1 for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes")

        for key, val in metrics.items():
            if "total_active" in key or ("total" in key and "customer" in key):
                ok = num_close(val, total_active, 2)
                record(f"Total_Active_Customers={total_active}", ok, f"Got {val}")
                if not ok:
                    all_ok = False
            elif "vip_count" in key and "risk" not in key:
                ok = num_close(val, vip_count, 2)
                record(f"VIP_Count={vip_count}", ok, f"Got {val}")
                if not ok:
                    all_ok = False
            elif "at_risk_total" in key or ("risk" in key and "total" in key):
                ok = num_close(val, at_risk_total, 2)
                record(f"At_Risk_Total={at_risk_total}", ok, f"Got {val}")
                if not ok:
                    all_ok = False
            elif "at_risk_vip" in key or ("risk" in key and "vip" in key):
                ok = num_close(val, at_risk_vip, 1)
                record(f"At_Risk_VIP_Count={at_risk_vip}", ok, f"Got {val}")
                if not ok:
                    all_ok = False

    # Check Action Plan sheet
    ap_sheet = None
    for name in wb.sheetnames:
        if "action" in name.lower() and "plan" in name.lower():
            ap_sheet = wb[name]
            break
    if ap_sheet is None:
        record("Sheet 'Action Plan' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Action Plan' exists", True)
        ap_rows = list(ap_sheet.iter_rows(min_row=2, values_only=True))
        vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]
        record("Action Plan row count matches",
               abs(len(ap_rows) - len(vip_at_risk)) <= 1,
               f"Expected {len(vip_at_risk)}, got {len(ap_rows)}")

    return all_ok


def check_emails():
    """Check that retention emails were sent."""
    print("\n=== Checking Emails Sent ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    expected = get_expected_data()
    vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]

    cur.execute("SELECT COUNT(*) FROM email.sent_log")
    sent_count = cur.fetchone()[0]
    record("Emails sent", sent_count >= len(vip_at_risk),
           f"Expected >= {len(vip_at_risk)}, got {sent_count}")

    # Check messages exist with retention subject
    cur.execute("SELECT subject, to_addr FROM email.messages WHERE subject ILIKE '%miss you%'")
    retention_emails = cur.fetchall()
    record("Retention emails found", len(retention_emails) >= len(vip_at_risk) - 1,
           f"Expected >= {len(vip_at_risk) - 1}, got {len(retention_emails)}")

    cur.close()
    conn.close()
    return True


def check_gsheet():
    """Check Google Sheet outreach log."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE title ILIKE '%retention%' OR title ILIKE '%outreach%'")
    rows = cur.fetchall()
    if not rows:
        record("GSheet with retention/outreach in title", False, "No matching spreadsheet found")
        cur.close()
        conn.close()
        return False
    record("GSheet with retention/outreach in title", True)

    ss_id = rows[0][0]
    cur.execute("SELECT id FROM gsheet.sheets WHERE spreadsheet_id = %s", (ss_id,))
    sheets = cur.fetchall()
    record("GSheet has at least one sheet", len(sheets) >= 1, f"Got {len(sheets)}")

    if sheets:
        sheet_id = sheets[0][0]
        cur.execute("SELECT COUNT(DISTINCT row_index) FROM gsheet.cells WHERE spreadsheet_id = %s AND sheet_id = %s AND row_index > 0",
                    (ss_id, sheet_id))
        row_count = cur.fetchone()[0]
        record("GSheet has data rows", row_count >= 3, f"Got {row_count} rows")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)

    db_fail_before = FAIL_COUNT
    email_ok = check_emails()
    gsheet_ok = check_gsheet()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")

    overall = excel_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

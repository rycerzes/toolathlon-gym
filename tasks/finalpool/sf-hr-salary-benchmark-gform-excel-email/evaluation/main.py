"""Evaluation for sf-hr-salary-benchmark-gform-excel-email.

Checks:
1. Salary_Analysis.xlsx has correct sheets and data
2. Google Forms survey "Compensation Satisfaction Survey" with 5 questions
3. Email to hr-leadership@company.example.com with subject matching pattern
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

DEPT_STATS = [
    ("Engineering", 7096, 15360.00, 695267.00, 58991.61, 53603),
    ("HR",          7077, 18307.00, 692232.00, 58920.45, 53656),
    ("Sales",       7232, 15885.00, 652806.00, 58864.79, 53490),
    ("Support",     7244, 15916.00, 608157.00, 58400.48, 52944),
    ("R&D",         7083, 15128.00, 680490.00, 57905.93, 52404),
    ("Finance",     7148, 15760.00, 638897.00, 57878.19, 52987),
    ("Operations",  7120, 17168.00, 656505.00, 57808.74, 52293),
]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Salary_Analysis.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Salary_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Salary_Analysis.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # Check Department_Stats sheet
    agent_dept = get_sheet(agent_wb, "Department_Stats")
    gt_dept = get_sheet(gt_wb, "Department_Stats")
    check("Sheet 'Department_Stats' exists", agent_dept is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_dept is None:
        all_ok = False
    else:
        a_rows = list(agent_dept.iter_rows(min_row=2, values_only=True))
        g_rows = list(gt_dept.iter_rows(min_row=2, values_only=True))
        check("Department_Stats has 7 rows", len(a_rows) == 7, f"Got {len(a_rows)}")
        if len(a_rows) != 7:
            all_ok = False

        # Build lookup by department name
        a_lookup = {}
        for r in a_rows:
            if r and r[0]:
                a_lookup[str(r[0]).strip().lower()] = r

        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            dept = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(dept)
            if a_row is None:
                check(f"Dept '{g_row[0]}' present", False, "Missing")
                all_ok = False
                continue
            # Headcount (col 1)
            ok = num_close(a_row[1], g_row[1], 5)
            check(f"'{g_row[0]}' Headcount", ok, f"Expected {g_row[1]}, got {a_row[1]}")
            if not ok:
                all_ok = False
            # Avg_Salary (col 4)
            ok = num_close(a_row[4], g_row[4], 100)
            check(f"'{g_row[0]}' Avg_Salary", ok, f"Expected {g_row[4]}, got {a_row[4]}")
            if not ok:
                all_ok = False

    # Check Summary sheet
    agent_summary = get_sheet(agent_wb, "Summary")
    gt_summary = get_sheet(gt_wb, "Summary")
    check("Sheet 'Summary' exists", agent_summary is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_summary is None:
        all_ok = False
    else:
        a_summary = {}
        for row in agent_summary.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        # Total_Employees
        te = a_summary.get("total_employees")
        check("Total_Employees = 50000", num_close(te, 50000, 10), f"Got {te}")
        if not num_close(te, 50000, 10):
            all_ok = False

        # Company_Avg_Salary
        cas = a_summary.get("company_avg_salary")
        check("Company_Avg_Salary close to 58396", num_close(cas, 58396.14, 200), f"Got {cas}")

        # Highest_Paid_Dept
        hpd = str(a_summary.get("highest_paid_dept", "")).strip().lower()
        check("Highest_Paid_Dept is Engineering", "engineering" in hpd, f"Got '{hpd}'")

        # Lowest_Paid_Dept
        lpd = str(a_summary.get("lowest_paid_dept", "")).strip().lower()
        check("Lowest_Paid_Dept is Operations", "operations" in lpd, f"Got '{lpd}'")

    return all_ok


def check_gform():
    print("\n=== Checking Google Forms ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms WHERE title ILIKE '%compensation%satisfaction%' OR title ILIKE '%compensation%survey%'")
    forms = cur.fetchall()
    check("Compensation Satisfaction Survey form exists", len(forms) >= 1,
          f"Found forms: {[f[1] for f in forms]}")

    if forms:
        form_id = forms[0][0]
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        check("Form has 5 questions", q_count == 5, f"Got {q_count}")
    cur.close()
    conn.close()


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%compensation%' AND subject ILIKE '%report%'
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    check("Compensation report email exists", len(emails) >= 1, f"Found {len(emails)}")
    if emails:
        e = emails[0]
        to_str = str(e[2])
        check("Email to hr-leadership@company.example.com",
              "hr-leadership@company.example.com" in to_str.lower(), f"to: {to_str}")
        check("Email from compensation@hr.example.com",
              "compensation@hr.example.com" in (e[1] or "").lower(), f"from: {e[1]}")
        body = (e[3] or "").lower()
        check("Email body mentions salary/compensation data",
              any(kw in body for kw in ["58396", "58,396", "engineering", "operations", "salary", "compensation"]),
              "Body missing key data")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

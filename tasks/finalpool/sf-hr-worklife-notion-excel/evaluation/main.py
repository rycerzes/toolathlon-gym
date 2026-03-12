"""Evaluation for sf-hr-worklife-notion-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    print("Checking Excel file...")
    agent_file = os.path.join(agent_ws, "WL_Balance_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WL_Balance_Report.xlsx")

    if not os.path.exists(agent_file):
        all_errors.append("WL_Balance_Report.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Department Analysis sheet
        print("  Checking Department Analysis sheet...")
        a_rows = load_sheet_rows(agent_wb, "Department Analysis")
        g_rows = load_sheet_rows(gt_wb, "Department Analysis")

        if a_rows is None:
            all_errors.append("Sheet 'Department Analysis' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            if len(a_data) != 7:
                all_errors.append(f"Department Analysis row count: {len(a_data)}, expected 7")

            a_lookup = {}
            for row in a_data:
                if row and row[0]:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing department: {g_row[0]}")
                    continue
                # Employee_Count
                if not num_close(a_row[1], g_row[1], 1):
                    all_errors.append(f"{g_row[0]} Employee_Count: {a_row[1]} vs {g_row[1]}")
                # Avg_WLB
                if not num_close(a_row[2], g_row[2], 0.05):
                    all_errors.append(f"{g_row[0]} Avg_WLB: {a_row[2]} vs {g_row[2]}")
                # Avg_Job_Satisfaction
                if not num_close(a_row[3], g_row[3], 0.05):
                    all_errors.append(f"{g_row[0]} Avg_Job_Satisfaction: {a_row[3]} vs {g_row[3]}")
            print("    Done.")

        # Check Findings sheet
        print("  Checking Findings sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Findings")
        if a_rows2 is None:
            all_errors.append("Sheet 'Findings' not found in agent output")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            a_lookup2 = {}
            for row in a_data2:
                if row and row[0]:
                    a_lookup2[str(row[0]).strip().lower()] = row

            # Check Total_Employees
            te_row = a_lookup2.get("total_employees")
            if te_row is None:
                all_errors.append("Findings missing Total_Employees row")
            elif not num_close(te_row[1], 50000, 1000):
                all_errors.append(f"Total_Employees: {te_row[1]} vs expected ~50000")

            print("    Done.")

    # --- Check 2: Notion page exists ---
    print("Checking Notion page...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE properties::text ILIKE '%HR Wellbeing Dashboard%'
               OR properties::text ILIKE '%wellbeing%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            # Also check by looking at blocks
            cur.execute("""
                SELECT COUNT(*) FROM notion.blocks
                WHERE content::text ILIKE '%wellbeing%' OR content::text ILIKE '%work.life%'
            """)
            block_count = cur.fetchone()[0]
            if block_count == 0:
                # Check pages table more broadly
                cur.execute("SELECT COUNT(*) FROM notion.pages")
                page_count = cur.fetchone()[0]
                if page_count == 0:
                    all_errors.append("No Notion pages found - HR Wellbeing Dashboard not created")
                else:
                    print(f"    Notion pages found: {page_count}")
            else:
                print(f"    Notion blocks with wellbeing content found: {block_count}")
        else:
            print(f"    Notion page found ({count} matching pages)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking Notion: {e}")

    # --- Check 3: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%hr_director@company.com%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        if count == 0:
            all_errors.append("No email sent to hr_director@company.com")
        else:
            print(f"    Email found ({count} messages)")
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

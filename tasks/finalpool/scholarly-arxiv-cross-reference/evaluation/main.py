"""
Evaluation for scholarly-arxiv-cross-reference task.
Checks Excel sheets and email.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# scholarly.arxiv_papers has 5 papers, arxiv.papers has 6
# Overlap IDs: 1602.05629, 1812.06127, 1908.07873 (3 papers)
# Scholarly only: 2001.08361, 2005.14165
# Arxiv only: 1207.00580, 1502.03167, 1912.04977
OVERLAP_IDS = {"1602.05629", "1812.06127", "1908.07873"}
ARXIV_ONLY_IDS = {"1207.00580", "1502.03167", "1912.04977"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == c:
                return i
    return None


def check_excel(agent_workspace):
    """Check Excel file."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Citation_Cross_Reference.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    # Check Scholarly_Papers sheet
    scholarly_rows = load_sheet_rows(wb, "Scholarly_Papers")
    if scholarly_rows is None:
        scholarly_rows = load_sheet_rows(wb, "Scholarly Papers")
    if scholarly_rows is not None:
        record("Sheet 'Scholarly_Papers' exists", True)
        data_rows = scholarly_rows[1:] if len(scholarly_rows) > 1 else []
        # scholarly.arxiv_papers has 5 papers
        record("Scholarly_Papers has >= 5 rows", len(data_rows) >= 5,
               f"Found {len(data_rows)} rows")
    else:
        record("Sheet 'Scholarly_Papers' exists", False, f"Available: {wb.sheetnames}")

    # Check Arxiv_Papers sheet
    arxiv_rows = load_sheet_rows(wb, "Arxiv_Papers")
    if arxiv_rows is None:
        arxiv_rows = load_sheet_rows(wb, "Arxiv Papers")
    if arxiv_rows is not None:
        record("Sheet 'Arxiv_Papers' exists", True)
        data_rows = arxiv_rows[1:] if len(arxiv_rows) > 1 else []
        record("Arxiv_Papers has 6 rows", len(data_rows) == 6,
               f"Found {len(data_rows)} rows")
    else:
        record("Sheet 'Arxiv_Papers' exists", False, f"Available: {wb.sheetnames}")

    # Check Overlap_Analysis sheet
    overlap_rows = load_sheet_rows(wb, "Overlap_Analysis")
    if overlap_rows is None:
        overlap_rows = load_sheet_rows(wb, "Overlap Analysis")
    if overlap_rows is not None:
        record("Sheet 'Overlap_Analysis' exists", True)
        header = overlap_rows[0] if overlap_rows else []
        data_rows = overlap_rows[1:] if len(overlap_rows) > 1 else []

        id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        in_scholarly_col = find_col(header, ["In_Scholarly", "In Scholarly", "in_scholarly"])
        in_arxiv_col = find_col(header, ["In_Arxiv", "In Arxiv", "in_arxiv"])

        record("Overlap has Paper_ID column", id_col is not None, f"Header: {header}")
        record("Overlap has In_Scholarly column", in_scholarly_col is not None, f"Header: {header}")
        record("Overlap has In_Arxiv column", in_arxiv_col is not None, f"Header: {header}")

        # Check total rows (should be union of both databases = 8)
        record("Overlap_Analysis has >= 8 rows", len(data_rows) >= 8,
               f"Found {len(data_rows)} rows")

        # Check overlap papers are marked Yes/Yes
        if id_col is not None and in_scholarly_col is not None and in_arxiv_col is not None:
            overlap_found = 0
            for row in data_rows:
                if id_col < len(row) and row[id_col]:
                    pid = str(row[id_col]).strip()
                    if pid in OVERLAP_IDS:
                        s_val = str(row[in_scholarly_col]).strip().lower() if in_scholarly_col < len(row) and row[in_scholarly_col] else ""
                        a_val = str(row[in_arxiv_col]).strip().lower() if in_arxiv_col < len(row) and row[in_arxiv_col] else ""
                        if "yes" in s_val and "yes" in a_val:
                            overlap_found += 1

            record("Overlap papers marked correctly", overlap_found >= 3,
                   f"Found {overlap_found} correctly marked overlap papers out of {len(OVERLAP_IDS)}")
    else:
        record("Sheet 'Overlap_Analysis' exists", False, f"Available: {wb.sheetnames}")


def check_email():
    """Check email was sent."""
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find sent folder ID
        cur.execute("SELECT id FROM email.folders WHERE name='Sent'")
        sent_row = cur.fetchone()
        if not sent_row:
            cur.execute("SELECT id FROM email.folders WHERE name ILIKE '%%sent%%' LIMIT 1")
            sent_row = cur.fetchone()

        # Check for email with cross-reference subject
        cur.execute("""
            SELECT id, subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%%cross%%reference%%'
               OR subject ILIKE '%%cross-reference%%'
        """)
        emails = cur.fetchall()

        if not emails:
            # Broader search
            cur.execute("""
                SELECT id, subject, from_addr, to_addr, body_text
                FROM email.messages
                WHERE subject ILIKE '%%scholarly%%'
                   OR subject ILIKE '%%arxiv%%'
            """)
            emails = cur.fetchall()

        record("Email with cross-reference subject sent", len(emails) > 0,
               "No matching email found")

        if emails:
            email = emails[0]
            to_addr = email[3]
            if isinstance(to_addr, str):
                to_addr = json.loads(to_addr)

            to_str = str(to_addr).lower()
            record("Email to research-lead@university.edu",
                   "research-lead@university.edu" in to_str,
                   f"To: {to_addr}")

            body = str(email[4]).lower() if email[4] else ""
            has_summary = ("overlap" in body or "common" in body or "shared" in body)
            record("Email body mentions overlap analysis", has_summary,
                   f"Body preview: {body[:200]}")

        conn.close()
    except Exception as e:
        record("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

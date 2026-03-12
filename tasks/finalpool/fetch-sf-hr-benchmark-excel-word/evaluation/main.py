"""
Evaluation script for fetch-sf-hr-benchmark-excel-word task.

Checks:
1. HR_Salary_Benchmark.xlsx with 3 sheets and correct data
2. Salary_Benchmark_Report.docx with executive summary content
3. Email sent with benchmark analysis subject
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=500.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


# Expected data based on actual Snowflake values and benchmark API
BENCHMARK_AVGS = {
    "engineering": 62000.00,
    "finance": 55000.00,
    "hr": 54000.00,
    "operations": 56000.00,
    "r&d": 61000.00,
    "sales": 57000.00,
    "support": 52000.00,
}

INTERNAL_AVGS = {
    "engineering": 58991.61,
    "finance": 57878.19,
    "hr": 58920.45,
    "operations": 57808.74,
    "r&d": 57905.93,
    "sales": 58864.79,
    "support": 58400.48,
}

EXPECTED_HEADCOUNTS = {
    "engineering": 7096,
    "finance": 7148,
    "hr": 7077,
    "operations": 7120,
    "r&d": 7083,
    "sales": 7232,
    "support": 7244,
}


def check_excel(agent_workspace):
    """Check HR_Salary_Benchmark.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "HR_Salary_Benchmark.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: Industry Benchmarks ---
    bench_sheet = None
    for name in wb.sheetnames:
        if "benchmark" in name.lower() or "industry" in name.lower():
            bench_sheet = name
            break
    if not bench_sheet:
        record("Industry Benchmarks sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Industry Benchmarks sheet exists", True)
        ws = wb[bench_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        if len(data_rows) == 7:
            record("Benchmarks sheet has 7 rows", True)
        else:
            record("Benchmarks sheet has 7 rows", False, f"Found {len(data_rows)}")
            all_ok = False

        for row in data_rows:
            if row and row[0]:
                dept = str(row[0]).strip().lower()
                if dept in BENCHMARK_AVGS:
                    expected = BENCHMARK_AVGS[dept]
                    # Find the avg salary column (typically col 1)
                    found = False
                    for cell in row[1:]:
                        if num_close(cell, expected, tol=100):
                            found = True
                            break
                    if not found:
                        record(f"Benchmark avg for {dept}", False,
                               f"Expected ~{expected}, row: {row[:5]}")
                        all_ok = False

    # --- Sheet 2: Internal Data ---
    int_sheet = None
    for name in wb.sheetnames:
        if "internal" in name.lower():
            int_sheet = name
            break
    if not int_sheet:
        record("Internal Data sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Internal Data sheet exists", True)
        ws = wb[int_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        if len(data_rows) == 7:
            record("Internal Data sheet has 7 rows", True)
        else:
            record("Internal Data sheet has 7 rows", False, f"Found {len(data_rows)}")
            all_ok = False

        checked_depts = 0
        for row in data_rows:
            if row and row[0]:
                dept = str(row[0]).strip().lower()
                if dept in INTERNAL_AVGS:
                    expected_avg = INTERNAL_AVGS[dept]
                    found_avg = False
                    for cell in row[1:]:
                        if num_close(cell, expected_avg, tol=500):
                            found_avg = True
                            break
                    ok = found_avg
                    record(f"Internal avg for {dept}", ok,
                           f"Expected ~{expected_avg}, row: {str(row[:6])[:200]}")
                    if not ok:
                        all_ok = False
                    checked_depts += 1

        if checked_depts < 5:
            record("At least 5 departments checked", False, f"Only {checked_depts}")
            all_ok = False

    # --- Sheet 3: Variance Analysis ---
    var_sheet = None
    for name in wb.sheetnames:
        if "variance" in name.lower() or "analysis" in name.lower():
            var_sheet = name
            break
    if not var_sheet:
        record("Variance Analysis sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Variance Analysis sheet exists", True)
        ws = wb[var_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        if len(data_rows) == 7:
            record("Variance sheet has 7 rows", True)
        else:
            record("Variance sheet has 7 rows", False, f"Found {len(data_rows)}")
            all_ok = False

        # Check a few key variances
        for row in data_rows:
            if row and row[0]:
                dept = str(row[0]).strip().lower()
                if dept == "support":
                    # Largest positive: Support ~6400.48
                    expected_var = 6400.48
                    found = False
                    for cell in row[1:]:
                        if num_close(cell, expected_var, tol=800):
                            found = True
                            break
                    record("Support variance amount correct", found,
                           f"Expected ~{expected_var}, row: {str(row)[:200]}")
                    if not found:
                        all_ok = False
                elif dept == "r&d":
                    # Largest negative: R&D ~-3094.07
                    expected_var = -3094.07
                    found = False
                    for cell in row[1:]:
                        if num_close(cell, expected_var, tol=800):
                            found = True
                            break
                    record("R&D variance amount correct", found,
                           f"Expected ~{expected_var}, row: {str(row)[:200]}")
                    if not found:
                        all_ok = False

    wb.close()
    return all_ok


def check_word(agent_workspace):
    """Check Salary_Benchmark_Report.docx."""
    print("\n=== Checking Word Output ===")

    fpath = os.path.join(agent_workspace, "Salary_Benchmark_Report.docx")
    if not os.path.isfile(fpath):
        record("Word file exists", False, f"Not found: {fpath}")
        return False

    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(fpath)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        record("Word file readable", False, str(e))
        return False

    all_ok = True

    # Check for key content
    checks = [
        ("Mentions Engineering", "engineering" in full_text),
        ("Mentions Finance", "finance" in full_text),
        ("Mentions HR", "hr" in full_text),
        ("Mentions Support", "support" in full_text),
        ("Mentions R&D", "r&d" in full_text),
        ("Mentions benchmark", "benchmark" in full_text),
        ("Mentions variance or comparison", "variance" in full_text or "comparison" in full_text or "compared" in full_text or "difference" in full_text),
    ]

    for name, cond in checks:
        record(name, cond)
        if not cond:
            all_ok = False

    return all_ok


def check_email():
    """Check email was sent with benchmark analysis."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    all_ok = True
    found_email = False

    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if "benchmark" in subj_lower or "salary" in subj_lower:
            found_email = True
            record("Benchmark analysis email exists", True)

            # Check from address
            from_ok = str_contains(from_addr, "hr")
            record("Email from HR address", from_ok, f"From: {from_addr}")
            if not from_ok:
                all_ok = False

            # Check to address
            to_str = str(to_addr).lower()
            to_ok = "vp" in to_str or "hr" in to_str or "director" in to_str
            record("Email to VP/HR address", to_ok, f"To: {to_addr}")
            if not to_ok:
                all_ok = False

            # Check body mentions key findings
            body_lower = (body_text or "").lower()
            body_ok = ("support" in body_lower or "r&d" in body_lower or
                       "engineering" in body_lower)
            record("Email body mentions key departments", body_ok,
                   f"Body preview: {(body_text or '')[:200]}")
            if not body_ok:
                all_ok = False

            break

    if not found_email:
        record("Benchmark analysis email exists", False,
               f"Found {len(emails)} emails, none with 'benchmark' or 'salary' in subject")
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    word_ok = check_word(args.agent_workspace)
    email_ok = check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:  {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Word:   {'PASS' if word_ok else 'FAIL'}")
    print(f"  Email:  {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    overall = excel_ok and word_ok and email_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

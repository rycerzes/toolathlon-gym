"""Evaluation for fetch-wc-supplier-restock-excel-word-gcal."""
import argparse
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Restocking_Plan.xlsx")
    if not os.path.exists(path):
        return ["Restocking_Plan.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Current Inventory
        rows = load_sheet_rows(wb, "Current Inventory")
        if rows is None:
            errors.append("Sheet 'Current Inventory' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 25:
                errors.append(f"Current Inventory has {len(data_rows)} rows, expected ~30")
            # Check that Days_of_Stock column exists and first items have low values
            zero_stock = [r for r in data_rows if r[1] is not None and float(r[1]) == 0]
            if not zero_stock:
                errors.append("No zero-stock products found in Current Inventory")

        # Supplier Pricing
        rows2 = load_sheet_rows(wb, "Supplier Pricing")
        if rows2 is None:
            errors.append("Sheet 'Supplier Pricing' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 6:
                errors.append(f"Supplier Pricing has {len(data_rows2)} rows, expected 6")
            cats = {str(r[0]).strip().lower() for r in data_rows2 if r[0]}
            for c in ["electronics", "cameras", "watches"]:
                if c not in cats:
                    errors.append(f"Category '{c}' missing from Supplier Pricing")
            # Check Electronics margin
            elec = [r for r in data_rows2 if r[0] and "electronics" == str(r[0]).strip().lower()]
            if elec and len(elec[0]) >= 6:
                if not num_close(elec[0][5], 50.9, 3.0):
                    errors.append(f"Electronics Margin={elec[0][5]}, expected ~50.9")
            # Check Electronics retail avg
            if elec and len(elec[0]) >= 5:
                if not num_close(elec[0][4], 61.13, 5.0):
                    errors.append(f"Electronics Retail Avg={elec[0][4]}, expected ~61.13")

        # Reorder Recommendations
        rows3 = load_sheet_rows(wb, "Reorder Recommendations")
        if rows3 is None:
            errors.append("Sheet 'Reorder Recommendations' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 6:
                errors.append(f"Reorder Recommendations has {len(data_rows3)} rows, expected 6")
            # Check Critical items exist
            critical = [r for r in data_rows3 if r[1] and str(r[1]).strip().lower() == "critical"]
            if len(critical) < 5:
                errors.append(f"Expected 5 Critical categories, got {len(critical)}")
            # Check Watches is High
            watches = [r for r in data_rows3 if r[0] and "watches" in str(r[0]).lower()]
            if watches:
                if str(watches[0][1]).strip().lower() != "high":
                    errors.append(f"Watches urgency={watches[0][1]}, expected High")
            # Check Electronics order qty
            elec_r = [r for r in data_rows3 if r[0] and "electronics" == str(r[0]).strip().lower()]
            if elec_r:
                if not num_close(elec_r[0][2], 200, 10):
                    errors.append(f"Electronics order qty={elec_r[0][2]}, expected 200")
                if not num_close(elec_r[0][3], 6000, 100):
                    errors.append(f"Electronics cost={elec_r[0][3]}, expected 6000")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    path = os.path.join(agent_workspace, "Procurement_Report.docx")
    if not os.path.exists(path):
        return ["Procurement_Report.docx not found"]
    try:
        from docx import Document
        doc = Document(path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        if "q2 2026" not in full_text and "procurement" not in full_text:
            errors.append("Word doc missing expected title content")
        if "zero stock" not in full_text and "critical" not in full_text and "0 stock" not in full_text:
            errors.append("Word doc missing inventory status discussion")
        if "15,940" not in full_text and "15940" not in full_text and "total" not in full_text:
            errors.append("Word doc missing total procurement cost")
    except Exception as e:
        errors.append(f"Error reading Word: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE start_datetime >= '2026-04-01' AND start_datetime < '2026-05-01'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if len(rows) < 3:
            errors.append(f"Expected 3 April meetings, found {len(rows)}")
        else:
            summaries = [r[0].lower() if r[0] else "" for r in rows]
            if not any("critical" in s for s in summaries):
                errors.append("No 'Critical' supplier meeting found")
            if not any("high" in s for s in summaries):
                errors.append("No 'High Priority' supplier meeting found")
            if not any("medium" in s for s in summaries):
                errors.append("No 'Medium Priority' supplier meeting found")
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

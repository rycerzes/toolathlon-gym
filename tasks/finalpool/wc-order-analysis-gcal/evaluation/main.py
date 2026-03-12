"""
Evaluation script for wc-order-analysis-gcal task.

Checks:
1. Excel Order_Analysis.xlsx with Status_Summary and Low_Stock_Products sheets
2. Google Calendar event "Restock Planning Meeting" exists in DB

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        d = (detail[:200] + "...") if len(detail) > 200 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    """Compute expected values from PostgreSQL."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    # Status summary
    cur.execute("""
        SELECT status, COUNT(*) as cnt,
               ROUND(SUM((total)::numeric), 2) as rev,
               ROUND(AVG((total)::numeric), 2) as avg_val
        FROM wc.orders GROUP BY status ORDER BY status
    """)
    status_rows = cur.fetchall()

    # Low stock products
    cur.execute("""
        SELECT name, sku, stock_quantity, total_sales
        FROM wc.products
        WHERE stock_quantity IS NOT NULL AND stock_quantity < 5
        ORDER BY stock_quantity, name
    """)
    low_stock_rows = cur.fetchall()

    conn.close()
    return {"status": status_rows, "low_stock": low_stock_rows}


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Order_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    # Check sheets exist
    check("Sheet 'Status_Summary' exists", get_sheet(wb, "Status_Summary") is not None,
          f"Found: {wb.sheetnames}")
    check("Sheet 'Low_Stock_Products' exists", get_sheet(wb, "Low_Stock_Products") is not None,
          f"Found: {wb.sheetnames}")

    # Status_Summary checks
    ws = get_sheet(wb, "Status_Summary")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["status"]
        check("Status_Summary row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        # Build lookup by status
        agent_by_status = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_status[str(row[0]).strip().lower()] = row

        for exp_row in exp:
            status = exp_row[0]
            agent_row = agent_by_status.get(status.lower())
            if agent_row:
                check(f"Status '{status}' Order_Count",
                      num_close(agent_row[1], exp_row[1], 0),
                      f"Expected {exp_row[1]}, got {agent_row[1]}")
                check(f"Status '{status}' Total_Revenue",
                      num_close(agent_row[2], float(exp_row[2]), 5.0),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"Status '{status}' Avg_Order_Value",
                      num_close(agent_row[3], float(exp_row[3]), 2.0),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
            else:
                check(f"Status '{status}' found", False, "Not in agent output")

        # Check sort order (alphabetical by status)
        if len(agent_rows) >= 2:
            statuses = [str(r[0]).strip().lower() for r in agent_rows if r and r[0]]
            check("Status_Summary sorted alphabetically",
                  statuses == sorted(statuses),
                  f"Got: {statuses[:5]}")

    # Low_Stock_Products checks
    ws2 = get_sheet(wb, "Low_Stock_Products")
    if ws2 and expected:
        agent_rows = list(ws2.iter_rows(min_row=2, values_only=True))
        exp = expected["low_stock"]
        check("Low_Stock_Products row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        # Check sort order by stock ascending
        stocks = [r[2] for r in agent_rows if r and r[2] is not None]
        if len(stocks) >= 2:
            check("Low_Stock sorted by Stock_Quantity ascending",
                  all(stocks[i] <= stocks[i + 1] for i in range(len(stocks) - 1)),
                  f"Stock values: {stocks[:5]}")

        # Spot check a few rows by SKU
        agent_by_sku = {}
        for row in agent_rows:
            if row and len(row) >= 4 and row[1]:
                agent_by_sku[str(row[1]).strip()] = row

        checked = 0
        for exp_row in exp:
            if checked >= 5:
                break
            sku = exp_row[1]
            agent_row = agent_by_sku.get(sku)
            if agent_row:
                check(f"SKU '{sku}' Stock_Quantity",
                      num_close(agent_row[2], exp_row[2], 0),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"SKU '{sku}' Total_Sales",
                      num_close(agent_row[3], exp_row[3], 1),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
                checked += 1


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for gcal check", False, str(e), db=True)
        return

    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
    events = cur.fetchall()
    conn.close()

    check("At least one calendar event exists", len(events) > 0,
          f"Found {len(events)} events", db=True)

    found_meeting = False
    for ev in events:
        summary = str(ev[0] or "").strip().lower()
        if "restock" in summary and "planning" in summary:
            found_meeting = True
            check("Event title contains 'Restock Planning Meeting'", True, db=True)

            # Check date is 2026-03-10
            if ev[2]:
                dt_str = str(ev[2])
                check("Event is on 2026-03-10",
                      "2026-03-10" in dt_str,
                      f"Got start: {dt_str}", db=True)

            # Check description mentions low stock products
            desc = str(ev[3] or "")
            check("Event description is not empty", len(desc) > 10,
                  f"Description length: {len(desc)}", db=True)
            break

    if not found_meeting:
        check("Restock Planning Meeting event found", False,
              f"Events found: {[e[0] for e in events]}", db=True)


def check_excel_gt(agent_workspace, groundtruth_workspace):
    """Fallback: compare against groundtruth Excel."""
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Order_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Order_Analysis.xlsx")

    check("Excel file exists", os.path.isfile(agent_file))
    check("Groundtruth file exists", os.path.isfile(gt_file))
    if not os.path.isfile(agent_file) or not os.path.isfile(gt_file):
        return

    agent_wb = openpyxl.load_workbook(agent_file)
    gt_wb = openpyxl.load_workbook(gt_file)

    for sheet_name in ["Status_Summary", "Low_Stock_Products"]:
        a_ws = get_sheet(agent_wb, sheet_name)
        g_ws = get_sheet(gt_wb, sheet_name)
        check(f"Sheet '{sheet_name}' exists in agent", a_ws is not None)
        if not a_ws or not g_ws:
            continue
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check(f"{sheet_name} row count matches", len(a_rows) == len(g_rows),
              f"Expected {len(g_rows)}, got {len(a_rows)}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        check_excel(agent_workspace, expected)
    else:
        print("INFO: Falling back to groundtruth Excel")
        check_excel_gt(agent_workspace, groundtruth_workspace)

    check_calendar()

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": total_pass,
            "failed": total_fail,
            "file_passed": FILE_PASS,
            "file_failed": FILE_FAIL,
            "db_passed": DB_PASS,
            "db_failed": DB_FAIL,
            "success": file_ok,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return file_ok, f"Passed: {total_pass}, Failed: {total_fail}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

"""
Evaluation for train-sf-sales-trip-excel-word-email task.

Checks:
1. Sales_Trip_Plan.xlsx exists with Travel_Details, Customer_Priority, Summary sheets
2. Travel_Details has G1 train info with correct times
3. Customer_Priority has 5 North America customers sorted by total amount
4. Top customer is Ethan Brown with ~7088 total amount
5. Sales_Trip_Brief.docx exists with required sections
6. Email sent to sales-manager@company.com about Shanghai trip
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Sales_Trip_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Sales_Trip_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Sales_Trip_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Sales_Trip_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Travel_Details sheet",
           any("travel" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has Customer_Priority sheet",
           any("customer" in s or "priority" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has Summary sheet", "summary" in sheet_names_lower, f"Sheets: {wb.sheetnames}")

    # Check Travel_Details
    travel_sheet = None
    for name in wb.sheetnames:
        if "travel" in name.lower():
            travel_sheet = wb[name]
            break

    if travel_sheet:
        rows = list(travel_sheet.iter_rows(values_only=True))
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Travel_Details has G1 train code", "g1" in all_text, "No G1 found")
        record("Travel_Details has 09:00 departure", "09:00" in all_text, "No 09:00 departure found")
        record("Travel_Details has 13:28 arrival", "13:28" in all_text, "No 13:28 arrival found")

    # Check Customer_Priority
    customer_sheet = None
    for name in wb.sheetnames:
        if "customer" in name.lower() or "priority" in name.lower():
            customer_sheet = wb[name]
            break

    if customer_sheet:
        rows = list(customer_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Customer_Priority has 5 customers", len(data_rows) >= 5,
               f"Found {len(data_rows)} rows")

        # Check top customer is Ethan Brown
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Top customer Ethan Brown is present", "ethan" in all_text and "brown" in all_text,
               "No Ethan Brown found")
        record("Customers include Noah Garcia", "noah" in all_text, "No Noah Garcia found")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Sales_Trip_Plan.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_word(agent_workspace):
    print("\n=== Check 2: Sales_Trip_Brief.docx ===")
    docx_path = os.path.join(agent_workspace, "Sales_Trip_Brief.docx")
    if not os.path.exists(docx_path):
        record("Sales_Trip_Brief.docx exists", False, f"Not found at {docx_path}")
        return
    record("Sales_Trip_Brief.docx exists", True)

    try:
        import docx
        doc = docx.Document(docx_path)
        full_text = " ".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        record("Word document readable", False, str(e))
        return
    record("Word document readable", True)

    record("Contains Travel Logistics section",
           "travel logistics" in full_text or "logistics" in full_text,
           "No Travel Logistics section found")
    record("Contains Customer Visit section",
           "customer" in full_text and ("visit" in full_text or "priority" in full_text),
           "No Customer Visit section found")
    record("Contains Meeting Objectives section",
           "meeting objectives" in full_text or "objectives" in full_text,
           "No Meeting Objectives section found")
    record("Contains Follow-up Actions section",
           "follow" in full_text and ("action" in full_text or "up" in full_text),
           "No Follow-up Actions section found")
    record("Mentions G1 train", "g1" in full_text, "No G1 train mention")
    record("Mentions Ethan Brown (top customer)", "ethan" in full_text, "No Ethan Brown mention")


def check_email():
    print("\n=== Check 3: Email to sales-manager@company.com ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%trip%' OR subject ILIKE '%shanghai%'
            OR subject ILIKE '%confirmed%'
        """)
        emails = cur.fetchall()
        # Also check by recipient
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%sales-manager%'
        """)
        emails2 = cur.fetchall()
        all_emails = emails + emails2
        record("Email sent to sales-manager@company.com", len(all_emails) >= 1,
               f"Found {len(all_emails)} matching emails")

        if all_emails:
            subject, to_addr, body = all_emails[0]
            subject_lower = (subject or "").lower()
            record("Email subject mentions trip or confirmed",
                   "trip" in subject_lower or "shanghai" in subject_lower or "confirmed" in subject_lower,
                   f"Subject: {subject}")
            body_lower = (body or "").lower()
            record("Email body mentions G1 train", "g1" in body_lower, "No G1 in email body")
            record("Email body mentions top customers",
                   "ethan" in body_lower or "brown" in body_lower,
                   "No customer names in email body")
    except Exception as e:
        record("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_word(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

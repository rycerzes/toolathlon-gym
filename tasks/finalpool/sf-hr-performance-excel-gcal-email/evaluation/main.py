"""Evaluation for sf-hr-performance-excel-gcal-email.

Checks:
1. Performance_Review_Summary.xlsx with 3 sheets
2. Google Calendar event "Annual Performance Review Board Meeting" 21 days from launch
3. Email to executives@company.example.com
"""
import argparse
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

TOP_PERFORMERS = [
    ("Engineering", 721, 59150.56, 8.35),
    ("Finance",     721, 58641.91, 8.29),
    ("HR",          740, 57240.73, 8.79),
    ("Operations",  687, 57575.52, 8.60),
    ("R&D",         693, 56086.53, 8.38),
    ("Sales",       694, 59193.14, 7.77),
    ("Support",     752, 59869.03, 8.09),
]

UNDERPERFORMERS = [
    ("Engineering", 1381, 59347.28, 8.35),
    ("Finance",     1387, 59698.81, 8.29),
    ("HR",          1425, 59741.41, 8.15),
    ("Operations",  1431, 57394.68, 8.49),
    ("R&D",         1415, 57839.57, 8.06),
    ("Sales",       1494, 58111.22, 8.54),
    ("Support",     1479, 58491.37, 8.04),
]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Performance_Review_Summary.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Performance_Review_Summary.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Performance_Review_Summary.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # Check Top_Performers sheet
    agent_top = get_sheet(agent_wb, "Top_Performers")
    gt_top = get_sheet(gt_wb, "Top_Performers")
    check("Sheet 'Top_Performers' exists", agent_top is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_top is None:
        all_ok = False
    else:
        a_rows = list(agent_top.iter_rows(min_row=2, values_only=True))
        check("Top_Performers has 7 rows", len(a_rows) == 7, f"Got {len(a_rows)}")
        if len(a_rows) != 7:
            all_ok = False

        a_lookup = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}
        for gt_dept, gt_cnt, gt_sal, gt_exp in TOP_PERFORMERS:
            key = gt_dept.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Top '{gt_dept}' present", False, "Missing")
                all_ok = False
                continue
            ok_cnt = num_close(a_row[1], gt_cnt, 5)
            check(f"Top '{gt_dept}' Count_Rating_5", ok_cnt, f"Expected {gt_cnt}, got {a_row[1]}")
            if not ok_cnt:
                all_ok = False
            ok_sal = num_close(a_row[2], gt_sal, 500)
            check(f"Top '{gt_dept}' Avg_Salary_Top", ok_sal, f"Expected {gt_sal}, got {a_row[2]}")
            if not ok_sal:
                all_ok = False

    # Check Underperformers sheet
    agent_low = get_sheet(agent_wb, "Underperformers")
    check("Sheet 'Underperformers' exists", agent_low is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_low is None:
        all_ok = False
    else:
        a_rows = list(agent_low.iter_rows(min_row=2, values_only=True))
        check("Underperformers has 7 rows", len(a_rows) == 7, f"Got {len(a_rows)}")
        if len(a_rows) != 7:
            all_ok = False

        a_lookup = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}
        for gt_dept, gt_cnt, gt_sal, gt_exp in UNDERPERFORMERS:
            key = gt_dept.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Low '{gt_dept}' present", False, "Missing")
                all_ok = False
                continue
            ok_cnt = num_close(a_row[1], gt_cnt, 10)
            check(f"Low '{gt_dept}' Count_Low_Rating", ok_cnt, f"Expected {gt_cnt}, got {a_row[1]}")
            if not ok_cnt:
                all_ok = False

    # Check Summary sheet
    agent_sum = get_sheet(agent_wb, "Summary")
    check("Sheet 'Summary' exists", agent_sum is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_sum is None:
        all_ok = False
    else:
        a_summary = {}
        for row in agent_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        ttp = a_summary.get("total_top_performers")
        check("Total_Top_Performers = 5008", num_close(ttp, 5008, 20), f"Got {ttp}")
        tup = a_summary.get("total_underperformers")
        check("Total_Underperformers = 10012", num_close(tup, 10012, 40), f"Got {tup}")
        oar = a_summary.get("overall_avg_rating")
        check("Overall_Avg_Rating close to 3.20", num_close(oar, 3.20, 0.1), f"Got {oar}")

    return all_ok


def check_gcal(launch_time_str):
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")
    check("At least 1 calendar event", len(events) >= 1, f"Found {len(events)}")

    board_events = [e for e in events
                    if "annual" in (e[0] or "").lower() and "performance" in (e[0] or "").lower()]
    check("Annual Performance Review Board Meeting event exists", len(board_events) >= 1,
          f"Events: {[e[0] for e in events]}")

    if launch_time_str and board_events:
        try:
            launch_dt = datetime.fromisoformat(launch_time_str)
            expected_dt = launch_dt + timedelta(days=21)
            for ev in board_events:
                if ev[2]:
                    ev_dt = ev[2]
                    diff = abs((ev_dt.replace(tzinfo=None) - expected_dt).total_seconds())
                    check("Board meeting 21 days from launch", diff <= 86400 * 2,
                          f"Expected around {expected_dt}, got {ev_dt}")
                    break
        except Exception as e:
            print(f"  [INFO] Could not verify date: {e}")

    return len(board_events) >= 1


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%performance%review%'
          AND (subject ILIKE '%annual%' OR from_addr ILIKE '%hr@company%' OR to_addr::text ILIKE '%executives%')
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    check("Performance review email exists", len(emails) >= 1, f"Found {len(emails)}")
    if emails:
        e = emails[0]
        to_str = str(e[2])
        check("Email to executives@company.example.com",
              "executives@company.example.com" in to_str.lower(), f"to: {to_str}")
        check("Email from hr@company.example.com",
              "hr@company.example.com" in (e[1] or "").lower(), f"from: {e[1]}")
        body = (e[3] or "").lower()
        check("Email body mentions performance data",
              any(kw in body for kw in ["5008", "10012", "3.20", "top", "performer", "underperform"]),
              "Body missing key data")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    check_gcal(args.launch_time)
    check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

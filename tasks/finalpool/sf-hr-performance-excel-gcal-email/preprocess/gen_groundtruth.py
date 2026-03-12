"""Generate groundtruth Excel for sf-hr-performance-excel-gcal-email."""
import os
import openpyxl

# Actual data from DB (sorted alphabetically by department)
TOP_PERFORMERS = [
    # Department, Count_Rating_5, Avg_Salary_Top, Avg_Experience_Top
    ("Engineering", 721, 59150.56, 8.35),
    ("Finance",     721, 58641.91, 8.29),
    ("HR",          740, 57240.73, 8.79),
    ("Operations",  687, 57575.52, 8.60),
    ("R&D",         693, 56086.53, 8.38),
    ("Sales",       694, 59193.14, 7.77),
    ("Support",     752, 59869.03, 8.09),
]

UNDERPERFORMERS = [
    # Department, Count_Low_Rating, Avg_Salary_Low, Avg_Experience_Low
    ("Engineering", 1381, 59347.28, 8.35),
    ("Finance",     1387, 59698.81, 8.29),
    ("HR",          1425, 59741.41, 8.15),
    ("Operations",  1431, 57394.68, 8.49),
    ("R&D",         1415, 57839.57, 8.06),
    ("Sales",       1494, 58111.22, 8.54),
    ("Support",     1479, 58491.37, 8.04),
]

SUMMARY = [
    ("Total_Top_Performers",  5008),
    ("Total_Underperformers", 10012),
    ("Overall_Avg_Rating",    3.20),
]

out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "groundtruth_workspace")
os.makedirs(out_dir, exist_ok=True)

wb = openpyxl.Workbook()

# Sheet 1: Top_Performers
ws_top = wb.active
ws_top.title = "Top_Performers"
ws_top.append(["Department", "Count_Rating_5", "Avg_Salary_Top", "Avg_Experience_Top"])
for row in TOP_PERFORMERS:
    ws_top.append(list(row))

# Sheet 2: Underperformers
ws_low = wb.create_sheet("Underperformers")
ws_low.append(["Department", "Count_Low_Rating", "Avg_Salary_Low", "Avg_Experience_Low"])
for row in UNDERPERFORMERS:
    ws_low.append(list(row))

# Sheet 3: Summary
ws_sum = wb.create_sheet("Summary")
ws_sum.append(["Metric", "Value"])
for row in SUMMARY:
    ws_sum.append(list(row))

wb.save(os.path.join(out_dir, "Performance_Review_Summary.xlsx"))
print("Groundtruth Excel created: Performance_Review_Summary.xlsx")
print(f"  Total_Top_Performers: {sum(r[1] for r in TOP_PERFORMERS)}")
print(f"  Total_Underperformers: {sum(r[1] for r in UNDERPERFORMERS)}")

"""Evaluation for fetch-howtocook-catering-budget-excel-gcal-word."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Catering_Budget.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Catering_Budget.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

        # Check Meal Plan sheet
        print("  Checking Meal Plan sheet...")
        a_rows = load_sheet_rows(agent_wb, "Meal Plan")
        if a_rows is None:
            all_errors.append("Sheet 'Meal Plan' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            if len(a_data) < 9:
                all_errors.append(f"Meal Plan row count: {len(a_data)}, expected 9")
            else:
                # Check all 3 days and 3 meal types exist
                days = set()
                meals = set()
                for row in a_data:
                    if row and row[0] is not None:
                        days.add(int(row[0]))
                    if row and row[1]:
                        meals.add(str(row[1]).strip().lower())
                if len(days) < 3:
                    all_errors.append(f"Meal Plan covers {len(days)} days, expected 3")
                expected_meals = {"breakfast", "lunch", "dinner"}
                if not expected_meals.issubset(meals):
                    all_errors.append(f"Missing meal types: {expected_meals - meals}")
            print("    Done.")

        # Check Ingredient Costs sheet
        print("  Checking Ingredient Costs sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Ingredient Costs")
        if a_rows2 is None:
            all_errors.append("Sheet 'Ingredient Costs' not found in agent output")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            if len(a_data2) < 10:
                all_errors.append(f"Ingredient Costs: {len(a_data2)} rows, expected 10+")
            # Check that Line_Total values are present and positive
            totals = []
            for row in a_data2:
                if row and len(row) > 6 and row[6] is not None:
                    try:
                        totals.append(float(row[6]))
                    except (TypeError, ValueError):
                        pass
            if len(totals) < 10:
                all_errors.append(f"Ingredient Costs: only {len(totals)} rows have Line_Total")
            elif any(t <= 0 for t in totals):
                all_errors.append("Ingredient Costs: some Line_Total values are <= 0")
            print("    Done.")

        # Check Budget Summary sheet
        print("  Checking Budget Summary sheet...")
        a_rows3 = load_sheet_rows(agent_wb, "Budget Summary")
        if a_rows3 is None:
            all_errors.append("Sheet 'Budget Summary' not found in agent output")
        else:
            a_data3 = a_rows3[1:] if len(a_rows3) > 1 else []
            if len(a_data3) < 9:
                all_errors.append(f"Budget Summary: {len(a_data3)} rows, expected 9+")
            # Check for grand total row
            has_total = False
            for row in a_data3:
                if row:
                    for cell in row:
                        if cell and "total" in str(cell).lower():
                            has_total = True
                            break
            if not has_total:
                all_errors.append("Budget Summary: no Grand Total row found")
            print("    Done.")

    # --- Check 2: Word document ---
    print("Checking Word document...")
    doc_path = os.path.join(agent_ws, "Catering_Proposal.docx")
    if not os.path.exists(doc_path):
        all_errors.append("Catering_Proposal.docx not found in agent workspace")
    else:
        try:
            from docx import Document
            doc = Document(doc_path)
            full_text = "\n".join(p.text for p in doc.paragraphs)
            full_lower = full_text.lower()
            if "80" not in full_text:
                all_errors.append("Word doc does not mention 80 attendees")
            if "day 1" not in full_lower and "day1" not in full_lower:
                all_errors.append("Word doc does not mention Day 1")
            if "total" not in full_lower:
                all_errors.append("Word doc does not include total cost")
            if "catering" not in full_lower and "retreat" not in full_lower:
                all_errors.append("Word doc missing title/overview")
        except Exception as e:
            all_errors.append(f"Error reading Word doc: {e}")

    # --- Check 3: GCal delivery events ---
    print("Checking GCal events...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date
            FROM gcal.events
            WHERE summary ILIKE '%catering%delivery%' OR summary ILIKE '%day%delivery%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if len(rows) < 3:
            all_errors.append(f"GCal: found {len(rows)} delivery events, expected 3")
        else:
            dates = [str(r[1]) for r in rows]
            if "2026-04-13" not in dates:
                all_errors.append("GCal: no delivery event on 2026-04-13")
            if "2026-04-14" not in dates:
                all_errors.append("GCal: no delivery event on 2026-04-14")
            if "2026-04-15" not in dates:
                all_errors.append("GCal: no delivery event on 2026-04-15")
    except Exception as e:
        all_errors.append(f"Error checking GCal: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

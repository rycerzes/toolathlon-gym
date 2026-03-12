"""Evaluation for terminal-yf-howtocook-excel-word-gcal.

Checks:
1. Wellness_Program_Plan.xlsx with 4 sheets
2. Wellness_Plan_Report.docx
3. Google Calendar events for wellness workshops
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def _get_yf_cost_data():
    """Query yf schema for gold/XOM prices and compute expected cost index dynamically."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        try:
            # Get last 5 trading days for GC=F (gold)
            cur.execute("""
                SELECT date, close FROM yf.stock_prices
                WHERE symbol = 'GC=F'
                ORDER BY date DESC LIMIT 5
            """)
            gold_rows = cur.fetchall()

            # Get last 5 trading days for XOM
            cur.execute("""
                SELECT date, close FROM yf.stock_prices
                WHERE symbol = 'XOM'
                ORDER BY date DESC LIMIT 5
            """)
            xom_rows = cur.fetchall()

            if not gold_rows or not xom_rows:
                return None

            gold_avg = sum(float(r[1]) for r in gold_rows) / len(gold_rows)
            xom_latest = float(xom_rows[0][1])
            xom_earliest = float(xom_rows[-1][1])
            xom_change_pct = (xom_latest - xom_earliest) / xom_earliest * 100
            cost_index = (gold_avg / 5000) * (1.0 + xom_change_pct / 100)

            return {
                "gold_avg": round(gold_avg, 2),
                "xom_change_pct": round(xom_change_pct, 2),
                "cost_index": round(cost_index, 2),
            }
        finally:
            cur.close()
            conn.close()
    except Exception:
        return None


_YF_COST_DATA = _get_yf_cost_data()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Wellness_Program_Plan.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Wellness_Program_Plan.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Wellness_Program_Plan.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    check("Has 4 sheets", len(agent_wb.sheetnames) >= 4,
          f"Got {agent_wb.sheetnames}")

    # Sheet 1: Cost_Analysis
    print("  --- Cost_Analysis ---")
    a_ws = get_sheet(agent_wb, "Cost_Analysis")
    g_ws = get_sheet(gt_wb, "Cost_Analysis")
    check("Cost_Analysis sheet exists", a_ws is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        check("Cost_Analysis has 1 data row", len(a_rows) >= 1, f"Got {len(a_rows)}")
        if a_rows:
            ar = a_rows[0]
            # Month
            check("Month is March 2026",
                  ar[0] and "march" in str(ar[0]).lower() and "2026" in str(ar[0]),
                  f"Got {ar[0]}")
            # Use dynamic YF data if available, fall back to groundtruth xlsx
            if _YF_COST_DATA:
                expected_gold = _YF_COST_DATA["gold_avg"]
                expected_ci = _YF_COST_DATA["cost_index"]
            elif g_ws:
                g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
                expected_gold = g_rows[0][1] if g_rows else 0
                expected_ci = g_rows[0][3] if g_rows else 0
            else:
                expected_gold = 0
                expected_ci = 0
            # Gold_Avg - tolerance of 100
            check("Gold_Avg value",
                  num_close(ar[1], expected_gold, 100),
                  f"Expected ~{expected_gold}, got {ar[1]}")
            # Cost_Index - tolerance of 0.1
            check("Cost_Index value",
                  num_close(ar[3], expected_ci, 0.1),
                  f"Expected ~{expected_ci}, got {ar[3]}")

    # Sheet 2: Meal_Plan
    print("  --- Meal_Plan ---")
    a_ws = get_sheet(agent_wb, "Meal_Plan")
    g_ws = get_sheet(gt_wb, "Meal_Plan")
    check("Meal_Plan sheet exists", a_ws is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Meal_Plan has 4 rows", len(a_rows) == 4, f"Got {len(a_rows)}")

        # Check each week's meals match groundtruth
        for i, (ar, gr) in enumerate(zip(a_rows, g_rows)):
            if not ar or not gr:
                continue
            week = i + 1
            # Check breakfast name
            if ar[1] and gr[1]:
                check(f"Week {week} Breakfast",
                      str(ar[1]).strip() == str(gr[1]).strip(),
                      f"Expected '{gr[1]}', got '{ar[1]}'")
            # Check lunch name
            if ar[2] and gr[2]:
                check(f"Week {week} Lunch",
                      str(ar[2]).strip() == str(gr[2]).strip(),
                      f"Expected '{gr[2]}', got '{ar[2]}'")
            # Check dinner name
            if ar[3] and gr[3]:
                check(f"Week {week} Dinner",
                      str(ar[3]).strip() == str(gr[3]).strip(),
                      f"Expected '{gr[3]}', got '{ar[3]}'")
            # Check Daily_Cost with tolerance
            if len(ar) > 4 and len(gr) > 4:
                check(f"Week {week} Daily_Cost",
                      num_close(ar[4], gr[4], 5.0),
                      f"Expected ~{gr[4]}, got {ar[4]}")
            # Check Nutrition_Score
            if len(ar) > 5 and len(gr) > 5:
                check(f"Week {week} Nutrition_Score",
                      num_close(ar[5], gr[5], 5),
                      f"Expected {gr[5]}, got {ar[5]}")

    # Sheet 3: Event_Schedule
    print("  --- Event_Schedule ---")
    a_ws = get_sheet(agent_wb, "Event_Schedule")
    g_ws = get_sheet(gt_wb, "Event_Schedule")
    check("Event_Schedule sheet exists", a_ws is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Event_Schedule has 4 rows", len(a_rows) == 4, f"Got {len(a_rows)}")

        for i, (ar, gr) in enumerate(zip(a_rows, g_rows)):
            if not ar or not gr:
                continue
            # Check date
            check(f"Event {i+1} date",
                  str(gr[0]) in str(ar[0]),
                  f"Expected {gr[0]}, got {ar[0]}")
            # Check topic
            if ar[1] and gr[1]:
                check(f"Event {i+1} topic",
                      str(ar[1]).strip().lower() == str(gr[1]).strip().lower(),
                      f"Expected '{gr[1]}', got '{ar[1]}'")

    # Sheet 4: Budget_Summary
    print("  --- Budget_Summary ---")
    a_ws = get_sheet(agent_wb, "Budget_Summary")
    g_ws = get_sheet(gt_wb, "Budget_Summary")
    check("Budget_Summary sheet exists", a_ws is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Budget_Summary has 4 rows (3 offices + total)",
              len(a_rows) == 4, f"Got {len(a_rows)}")

        # Build lookup by office name
        a_lookup = {}
        for r in a_rows:
            if r and r[0]:
                a_lookup[str(r[0]).strip().lower()] = r
        for gr in g_rows:
            if not gr or not gr[0]:
                continue
            key = str(gr[0]).strip().lower()
            ar = a_lookup.get(key)
            if ar is None:
                check(f"Office '{gr[0]}' present", False, "Missing")
                continue
            # Headcount
            if len(ar) > 1 and len(gr) > 1:
                check(f"'{key}' Headcount",
                      num_close(ar[1], gr[1], 0),
                      f"Expected {gr[1]}, got {ar[1]}")
            # Monthly_Budget
            if len(ar) > 2 and len(gr) > 2:
                check(f"'{key}' Monthly_Budget",
                      num_close(ar[2], gr[2], 0),
                      f"Expected {gr[2]}, got {ar[2]}")
            # Meal_Program_Cost (tolerance proportional to value)
            if len(ar) > 3 and len(gr) > 3:
                tol = max(abs(float(gr[3])) * 0.05, 500) if gr[3] else 500
                check(f"'{key}' Meal_Program_Cost",
                      num_close(ar[3], gr[3], tol),
                      f"Expected ~{gr[3]}, got {ar[3]}")


def check_word(agent_workspace):
    print("\n=== Checking Wellness_Plan_Report.docx ===")
    docx_path = os.path.join(agent_workspace, "Wellness_Plan_Report.docx")
    check("Wellness_Plan_Report.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 300, f"Length: {len(text)}")
        check("Contains title 'Q2 2026'",
              "q2 2026" in text or "q2" in text)
        check("Contains 'wellness' reference", "wellness" in text)
        check("Contains cost/commodity reference",
              "cost" in text and ("commodity" in text or "gold" in text or "index" in text))
        check("Contains recipe names",
              any(r in text for r in ["吐司", "荷包蛋", "南瓜", "腊肠"]))
        check("Contains event/workshop reference",
              "workshop" in text or "event" in text or "session" in text)
        check("Contains budget reference",
              "budget" in text and ("variance" in text or "cost" in text))
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check wellness workshop events exist
        cur.execute("""
            SELECT summary, start_datetime
            FROM gcal.events
            WHERE lower(summary) LIKE '%%wellness%%'
               OR lower(summary) LIKE '%%breakfast habits%%'
               OR lower(summary) LIKE '%%lunch ideas%%'
               OR lower(summary) LIKE '%%dinner planning%%'
               OR lower(summary) LIKE '%%meal prep%%'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        check("Calendar has 4 wellness events", len(events) >= 4,
              f"Found {len(events)} events")

        # Check specific dates (March 19, 25, April 2, 8)
        expected_dates = ["2026-03-19", "2026-03-25", "2026-04-02", "2026-04-08"]
        expected_topics = [
            "healthy breakfast habits",
            "plant-based lunch ideas",
            "balanced dinner planning",
            "meal prep workshop",
        ]
        for date_str, topic in zip(expected_dates, expected_topics):
            found = any(
                date_str in str(e[1]) and topic in str(e[0]).lower()
                for e in events
            )
            check(f"Event '{topic}' on {date_str}", found,
                  f"Events: {[(str(e[0]), str(e[1])[:10]) for e in events]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_reverse_validation():
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check no duplicate calendar events (same summary on same date)
        cur.execute("""
            SELECT summary, start_datetime::date, COUNT(*)
            FROM gcal.events
            WHERE lower(summary) LIKE '%%wellness%%'
               OR lower(summary) LIKE '%%breakfast%%'
               OR lower(summary) LIKE '%%lunch%%'
               OR lower(summary) LIKE '%%dinner%%'
               OR lower(summary) LIKE '%%meal prep%%'
            GROUP BY summary, start_datetime::date
            HAVING COUNT(*) > 1
        """)
        dupes = cur.fetchall()
        check("No duplicate wellness calendar events", len(dupes) == 0,
              f"Found {len(dupes)} duplicates: {dupes}")

        # Check no wellness events on weekends
        cur.execute("""
            SELECT summary, start_datetime
            FROM gcal.events
            WHERE (lower(summary) LIKE '%%wellness%%'
               OR lower(summary) LIKE '%%breakfast%%'
               OR lower(summary) LIKE '%%lunch%%'
               OR lower(summary) LIKE '%%dinner%%'
               OR lower(summary) LIKE '%%meal prep%%')
              AND EXTRACT(DOW FROM start_datetime) IN (0, 6)
        """)
        weekend_events = cur.fetchall()
        check("No wellness events on weekends", len(weekend_events) == 0,
              f"Found {len(weekend_events)} weekend events: {weekend_events}")

        # Check events are within the expected date range
        cur.execute("""
            SELECT summary, start_datetime
            FROM gcal.events
            WHERE (lower(summary) LIKE '%%wellness%%'
               OR lower(summary) LIKE '%%breakfast%%'
               OR lower(summary) LIKE '%%lunch%%'
               OR lower(summary) LIKE '%%dinner%%'
               OR lower(summary) LIKE '%%meal prep%%')
              AND (start_datetime::date < '2026-03-16' OR start_datetime::date > '2026-04-10')
        """)
        out_of_range = cur.fetchall()
        check("No wellness events outside March 16 - April 10 range",
              len(out_of_range) == 0,
              f"Found {len(out_of_range)} out-of-range events: {out_of_range}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gcal()
    check_reverse_validation()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

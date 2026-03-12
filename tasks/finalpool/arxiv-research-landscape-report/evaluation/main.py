"""
Evaluation script for arxiv-research-landscape-report task.

Checks:
1. Excel file (Research_Landscape.xlsx) with 3 sheets
2. Paper Analysis has 5 target papers with correct IDs and citation counts
3. Conference Fit has 3 conferences
4. Summary has required metrics
5. Word document (Landscape_Report.docx) exists with substantive content
"""
import argparse
import json
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0

EXPECTED_PAPERS = {
    "2401.00001": {"title": "Efficient Transformers for NLP", "citation_count": 350},
    "2401.00002": {"title": "Deep RL with Human Feedback", "citation_count": 520},
    "2401.00003": {"title": "Generative Models for Code", "citation_count": 280},
    "2401.00004": {"title": "Knowledge Graph Embeddings", "citation_count": 190},
    "2401.00005": {"title": "Optimization in Deep Learning", "citation_count": 150},
}

NOISE_IDS = {"2401.00006", "2401.00007", "2401.00008"}

CONFERENCES = ["NeurIPS 2026", "ICML 2026", "AAAI 2026"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=50):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel Output ===")
    path = os.path.join(workspace, "Research_Landscape.xlsx")
    if not os.path.isfile(path):
        record("Excel file exists", False, f"Not found: {path}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False
    record("Excel readable", True)

    all_ok = True

    # Paper Analysis sheet
    pa_rows = load_sheet_rows(wb, "Paper Analysis") or load_sheet_rows(wb, "Paper_Analysis")
    if pa_rows is None:
        record("Sheet 'Paper Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Paper Analysis' exists", True)
        header = pa_rows[0] if pa_rows else []
        data = pa_rows[1:]

        id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        title_col = find_col(header, ["Title", "title"])
        cite_col = find_col(header, ["Citation_Count", "Citation Count", "Citations", "citation_count"])

        record("Paper Analysis has 5 data rows", len(data) == 5,
               f"Found {len(data)}")
        if len(data) != 5:
            all_ok = False

        if id_col is not None:
            found_ids = {str(r[id_col]).strip() for r in data if id_col < len(r) and r[id_col]}
            for eid in EXPECTED_PAPERS:
                present = eid in found_ids
                record(f"Paper {eid} present", present, f"Found: {found_ids}")
                if not present:
                    all_ok = False
            for nid in NOISE_IDS:
                absent = nid not in found_ids
                record(f"Noise {nid} absent", absent)
                if not absent:
                    all_ok = False
        else:
            record("Paper_ID column found", False, f"Header: {header}")
            all_ok = False

        if cite_col is not None and id_col is not None:
            for row in data:
                pid = str(row[id_col]).strip() if id_col < len(row) and row[id_col] else ""
                if pid in EXPECTED_PAPERS:
                    ok = num_close(row[cite_col] if cite_col < len(row) else None,
                                   EXPECTED_PAPERS[pid]["citation_count"], tol=50)
                    record(f"Citation count for {pid}", ok,
                           f"Got {row[cite_col] if cite_col < len(row) else None}, expected {EXPECTED_PAPERS[pid]['citation_count']}")
                    if not ok:
                        all_ok = False

    # Conference Fit sheet
    cf_rows = load_sheet_rows(wb, "Conference Fit") or load_sheet_rows(wb, "Conference_Fit")
    if cf_rows is None:
        record("Sheet 'Conference Fit' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Conference Fit' exists", True)
        data = cf_rows[1:]
        record("Conference Fit has 3 rows", len(data) == 3, f"Found {len(data)}")
        if len(data) != 3:
            all_ok = False

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        # Total Papers
        tp_key = next((k for k in metrics if "total" in k and "paper" in k), None)
        if tp_key:
            ok = num_close(metrics[tp_key], 5, tol=0)
            record("Summary: Total_Papers = 5", ok, f"Got {metrics[tp_key]}")
            if not ok:
                all_ok = False
        else:
            record("Summary: Total_Papers exists", False)
            all_ok = False

        # Average Citations
        avg_key = next((k for k in metrics if "avg" in k or "average" in k), None)
        expected_avg = (350 + 520 + 280 + 190 + 150) / 5  # 298.0
        if avg_key:
            ok = num_close(metrics[avg_key], expected_avg, tol=20)
            record("Summary: Avg_Citations", ok,
                   f"Got {metrics[avg_key]}, expected ~{expected_avg}")
            if not ok:
                all_ok = False
        else:
            record("Summary: Avg_Citations exists", False)
            all_ok = False

        # Highest Cited Paper
        hc_key = next((k for k in metrics if "highest" in k or "most" in k), None)
        if hc_key:
            val = str(metrics[hc_key]).lower() if metrics[hc_key] else ""
            ok = "deep rl" in val or "human feedback" in val or "rl" in val
            record("Summary: Highest_Cited is Deep RL paper", ok, f"Got: {metrics[hc_key]}")
            if not ok:
                all_ok = False

    return all_ok


def check_word(workspace):
    print("\n=== Checking Word Document ===")
    path = os.path.join(workspace, "Landscape_Report.docx")
    if not os.path.isfile(path):
        record("Word document exists", False, f"Not found: {path}")
        return False
    record("Word document exists", True)

    try:
        from docx import Document
        doc = Document(path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()

        record("Document has substantial content", len(full_text) > 200,
               f"Only {len(full_text)} chars")
        record("Mentions conference", any(c.lower() in full_text for c in ["neurips", "icml", "aaai"]),
               "No conference names found")
        record("Mentions research/landscape", "research" in full_text or "landscape" in full_text,
               "Missing research/landscape keywords")
        record("Mentions papers or topics",
               any(kw in full_text for kw in ["transformer", "reinforcement", "optimization", "knowledge graph", "generative"]),
               "No topic keywords found")

        return True
    except Exception as e:
        record("Word document readable", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    word_ok = check_word(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Excel: {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Word:  {'PASS' if word_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

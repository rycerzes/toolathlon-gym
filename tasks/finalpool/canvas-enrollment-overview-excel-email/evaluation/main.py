"""Evaluation for canvas-enrollment-overview-excel-email."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Spring2014_Enrollment.xlsx")
    gt_file = os.path.join(gt_dir, "Spring2014_Enrollment.xlsx")

    file_errors = []
    db_errors = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    if not file_errors:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Enrollment sheet
        print("  Checking Enrollment...")
        a_rows = load_sheet_rows(agent_wb, "Enrollment")
        g_rows = load_sheet_rows(gt_wb, "Enrollment")
        if a_rows is None:
            file_errors.append("Sheet 'Enrollment' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Enrollment' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing row: {g_row[0]}")
                    continue

                # Student_Count (col 2, index 2)
                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 50):
                        file_errors.append(f"{key}.Student_Count: {a_row[2]} vs {g_row[2]}")
                # Teacher_Count (col 3)
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 1):
                        file_errors.append(f"{key}.Teacher_Count: {a_row[3]} vs {g_row[3]}")
                # TA_Count (col 4)
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 1):
                        file_errors.append(f"{key}.TA_Count: {a_row[4]} vs {g_row[4]}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary row: {g_row[0]}")
                    continue
                if len(a_row) > 1 and len(g_row) > 1:
                    if isinstance(g_row[1], (int, float)):
                        if not num_close(a_row[1], g_row[1], 50):
                            file_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                    else:
                        if not str_match(a_row[1], g_row[1]):
                            file_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")

    # Check email (DB check)
    print("  Checking email...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT subject FROM email.messages WHERE folder_id = 2")
        found = any("enrollment" in (s or "").lower() or "spring 2014" in (s or "").lower() for (s,) in cur.fetchall())
        if not found:
            db_errors.append("No email about enrollment/Spring 2014 found in sent folder")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Email check: {e}")

    # Check calendar event (DB check)
    print("  Checking calendar event...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT summary, start_datetime FROM gcal.events WHERE LOWER(summary) LIKE '%enrollment%review%' OR LOWER(summary) LIKE '%spring 2014%'")
        events = cur.fetchall()
        if not events:
            db_errors.append("No calendar event with 'Enrollment Review' found")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Calendar check: {e}")

    # Final result
    print(f"\n=== SUMMARY ===")
    print(f"  File errors: {len(file_errors)}")
    print(f"  DB errors:   {len(db_errors)} (not blocking)")
    if db_errors:
        for e in db_errors[:10]:
            print(f"    [DB] {e}")
    if file_errors:
        for e in file_errors[:10]:
            print(f"    [FILE] {e}")
        print(f"  Overall: FAIL")
        sys.exit(1)
    else:
        print(f"  Overall: PASS")
        sys.exit(0)


if __name__ == "__main__":
    main()

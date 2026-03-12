"""
Evaluation for yf-market-overview-pdf task.

Dynamically queries PostgreSQL to compute expected stock data values,
then checks agent output files for correctness.
"""
from argparse import ArgumentParser
import sys
import os
from pathlib import Path


SYMBOLS = ['AMZN', 'GOOGL', 'JNJ', 'JPM', 'XOM']


def get_expected_data():
    """Query PostgreSQL to get expected stock data."""
    import psycopg2

    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
        user="eigent", password="camel"
    )
    cur = conn.cursor()

    syms = tuple(SYMBOLS)
    cur.execute("""
        SELECT sp.symbol, sp.date, sp.open, sp.high, sp.low, sp.close, sp.volume
        FROM yf.stock_prices sp
        WHERE sp.symbol IN %s
          AND sp.date = (SELECT MAX(date) FROM yf.stock_prices)
        ORDER BY sp.symbol
    """, (syms,))
    prices = cur.fetchall()

    cur.execute("""
        SELECT symbol,
               data->>'shortName' as name,
               data->>'marketCap' as market_cap,
               data->>'trailingPE' as pe_ratio,
               data->>'fiftyTwoWeekHigh' as high_52w,
               data->>'fiftyTwoWeekLow' as low_52w,
               data->>'sector' as sector
        FROM yf.stock_info
        WHERE symbol IN %s
        ORDER BY symbol
    """, (syms,))
    info = {r[0]: r for r in cur.fetchall()}

    conn.close()
    return prices, info


def check_excel(workspace, prices, info):
    """Check Market_Overview_Report.xlsx for correctness."""
    import openpyxl

    xlsx_path = Path(workspace) / "Market_Overview_Report.xlsx"
    if not xlsx_path.exists():
        return False, f"Market_Overview_Report.xlsx not found in {workspace}"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Stock Data" not in wb.sheetnames:
        return False, f"Missing 'Stock Data' sheet. Found: {wb.sheetnames}"
    if "Summary" not in wb.sheetnames:
        return False, f"Missing 'Summary' sheet. Found: {wb.sheetnames}"

    # Check Stock Data sheet
    ws1 = wb["Stock Data"]
    rows1 = list(ws1.iter_rows(values_only=True))
    if len(rows1) < 2:
        return False, "Stock Data sheet has no data rows"

    header = [str(h).strip() if h else "" for h in rows1[0]]
    required = ["Symbol", "Close", "Volume"]
    for col in required:
        if col not in header:
            return False, f"Stock Data missing column '{col}'. Found: {header}"

    sym_idx = header.index("Symbol")
    close_idx = header.index("Close")
    vol_idx = header.index("Volume")

    data_rows = rows1[1:]
    if len(data_rows) != 5:
        return False, f"Stock Data: expected 5 rows, got {len(data_rows)}"

    # Build lookup by symbol
    price_map = {r[0]: r for r in prices}

    for row in data_rows:
        sym = str(row[sym_idx]).strip() if row[sym_idx] else ""
        if sym not in price_map:
            return False, f"Unexpected symbol '{sym}' in Stock Data"

        exp = price_map[sym]
        exp_close = float(exp[5])
        exp_vol = int(exp[6])

        if row[close_idx] is None or abs(float(row[close_idx]) - exp_close) > 0.5:
            return False, f"'{sym}' Close: expected {exp_close}, got {row[close_idx]}"

        if row[vol_idx] is None or abs(int(row[vol_idx]) - exp_vol) > exp_vol * 0.05:
            return False, f"'{sym}' Volume: expected {exp_vol}, got {row[vol_idx]}"

    print("  [PASS] Stock Data correct")

    # Check Summary sheet
    ws2 = wb["Summary"]
    rows2 = list(ws2.iter_rows(values_only=True))
    summary_map = {}
    for row in rows2[1:]:
        if row[0]:
            summary_map[str(row[0]).strip()] = row[1]

    if "Stocks_Covered" not in summary_map:
        return False, "Summary missing 'Stocks_Covered'"
    if int(summary_map["Stocks_Covered"]) != 5:
        return False, f"Stocks_Covered: expected 5, got {summary_map['Stocks_Covered']}"

    best = max(prices, key=lambda x: float(x[5]))
    if "Highest_Close_Symbol" in summary_map:
        if str(summary_map["Highest_Close_Symbol"]).strip() != best[0]:
            return False, f"Highest_Close_Symbol: expected {best[0]}, got {summary_map['Highest_Close_Symbol']}"

    print("  [PASS] Summary correct")
    wb.close()
    return True, "Excel file checks passed"


def check_pdf(workspace):
    """Check Market_Report.pdf exists and is reasonable."""
    pdf_path = Path(workspace) / "Market_Report.pdf"
    if not pdf_path.exists():
        return False, f"Market_Report.pdf not found in {workspace}"

    size = pdf_path.stat().st_size
    if size < 500:
        return False, f"PDF too small ({size} bytes)"

    print(f"  [PASS] PDF exists, size={size} bytes")
    return True, "PDF check passed"


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    workspace = args.agent_workspace
    if not workspace:
        print("Error: --agent_workspace is required")
        sys.exit(1)

    print("Fetching expected data from database...")
    try:
        prices, info = get_expected_data()
        print(f"  Stocks: {len(prices)}")
    except Exception as e:
        print(f"Error querying database: {e}")
        sys.exit(1)

    all_passed = True

    print("\n--- Check 1: Excel File ---")
    try:
        ok, msg = check_excel(workspace, prices, info)
        if not ok:
            print(f"  [FAIL] {msg}")
            all_passed = False
        else:
            print(f"  {msg}")
    except Exception as e:
        print(f"  [FAIL] Excel check error: {e}")
        all_passed = False

    print("\n--- Check 2: PDF File ---")
    try:
        ok, msg = check_pdf(workspace)
        if not ok:
            print(f"  [FAIL] {msg}")
            all_passed = False
        else:
            print(f"  {msg}")
    except Exception as e:
        print(f"  [FAIL] PDF check error: {e}")
        all_passed = False

    if all_passed:
        print("\nPass all tests!")
        sys.exit(0)
    else:
        print("\nSome checks failed.")
        sys.exit(1)

"""Evaluation for terminal-wc-sf-notion-excel-email.
Checks:
1. Support_Quality_Audit.xlsx with 4 sheets and correct data
2. Notion database with critical product entries
3. Two emails sent to correct recipients
4. Python scripts exist
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').strip())
    except Exception:
        return default


def get_groundtruth_from_db():
    """Compute expected values from read-only DB data."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Problem products
    cur.execute("""
        SELECT (unnest_item->>'product_id')::int as pid, COUNT(DISTINCT o.id)
        FROM wc.orders o, jsonb_array_elements(o.line_items) as unnest_item
        WHERE o.status IN ('refunded','failed')
        GROUP BY pid
    """)
    refund_products = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""
        SELECT product_id, COUNT(*)
        FROM wc.product_reviews WHERE rating <= 2
        GROUP BY product_id
    """)
    low_review_products = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT id, name, categories FROM wc.products")
    products = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    all_pids = set(refund_products.keys()) | set(low_review_products.keys())
    problem_list = []
    for pid in all_pids:
        rc = refund_products.get(pid, 0)
        lrc = low_review_products.get(pid, 0)
        severity = rc * 30 + lrc * 40
        name = products.get(pid, ("Unknown", []))[0][:60]
        cats = products.get(pid, ("", []))[1]
        cat = cats[0]['name'] if cats else 'Unknown'
        problem_list.append((pid, name, cat, rc, lrc, severity))
    problem_list.sort(key=lambda x: (-x[5], x[0]))

    severities = sorted([p[5] for p in problem_list])
    p80_idx = int(len(severities) * 0.8)
    p80_val = severities[p80_idx] if p80_idx < len(severities) else severities[-1]
    critical = [p for p in problem_list if p[5] > p80_val]

    # Priority data
    cur.execute("""
        SELECT "PRIORITY", COUNT(*),
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY" ORDER BY "PRIORITY"
    """)
    priority_data = cur.fetchall()

    # Issue type data
    cur.execute("""
        SELECT "ISSUE_TYPE", COUNT(*),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE" ORDER BY COUNT(*) DESC
    """)
    issue_data = cur.fetchall()

    total_tickets = sum(r[1] for r in priority_data)
    cur.execute("""
        SELECT ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
    """)
    overall_sat = float(cur.fetchone()[0])

    cur.close()
    conn.close()

    return {
        "problem_list": problem_list,
        "critical": critical,
        "priority_data": priority_data,
        "issue_data": issue_data,
        "total_tickets": total_tickets,
        "overall_sat": overall_sat,
        "p80_val": p80_val,
    }


def check_excel(workspace, gt):
    print("\n=== Check 1: Support_Quality_Audit.xlsx ===")
    path = os.path.join(workspace, "Support_Quality_Audit.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheets_lower = [s.lower() for s in sheets]

    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    # Problem_Products sheet
    pp_idx = next((i for i, s in enumerate(sheets_lower) if "problem" in s or "product" in s), 0)
    ws_pp = wb[sheets[pp_idx]]
    rows_pp = list(ws_pp.iter_rows(min_row=2, values_only=True))
    expected_count = len(gt["problem_list"])
    check(f"Problem_Products has ~{expected_count} rows",
          abs(len(rows_pp) - expected_count) <= 2,
          f"Found {len(rows_pp)} data rows, expected {expected_count}")

    # Check top product by severity
    if rows_pp:
        top_row = rows_pp[0]
        top_pid = safe_float(top_row[0])
        top_severity = safe_float(top_row[5] if len(top_row) > 5 else top_row[-1])
        expected_top = gt["problem_list"][0]
        check("Top product ID correct",
              top_pid is not None and int(top_pid) == expected_top[0],
              f"Got pid={top_pid}, expected {expected_top[0]}")
        check("Top product severity correct",
              num_close(top_severity, expected_top[5], tol=5),
              f"Got {top_severity}, expected {expected_top[5]}")

    # Check a mid-range product exists
    if len(gt["problem_list"]) > 5:
        mid_product = gt["problem_list"][3]
        all_text = " ".join(str(c) for r in rows_pp for c in r if c).lower()
        check("Contains expected mid-range product",
              str(mid_product[0]) in all_text or mid_product[1][:15].lower() in all_text,
              f"Looking for pid={mid_product[0]} or name={mid_product[1][:15]}")

    # Support_By_Priority sheet
    sp_idx = next((i for i, s in enumerate(sheets_lower) if "priority" in s or "support" in s), 1)
    if sp_idx < len(sheets):
        ws_sp = wb[sheets[sp_idx]]
        rows_sp = list(ws_sp.iter_rows(min_row=2, values_only=True))
        check("Support_By_Priority has 3 rows", len(rows_sp) == 3,
              f"Found {len(rows_sp)} rows")

        if rows_sp:
            all_text_sp = " ".join(str(c) for r in rows_sp for c in r if c).lower()
            check("Has High priority", "high" in all_text_sp)
            check("Has Medium priority", "medium" in all_text_sp)
            check("Has Low priority", "low" in all_text_sp)

            # Check High priority ticket count
            for r in rows_sp:
                if r[0] and "high" in str(r[0]).lower():
                    expected_high = next((p for p in gt["priority_data"] if p[0] == "High"), None)
                    if expected_high:
                        count = safe_float(r[1])
                        check(f"High priority count ~{expected_high[1]}",
                              num_close(count, expected_high[1], tol=50),
                              f"Got {count}, expected {expected_high[1]}")
                        sat = safe_float(r[3] if len(r) > 3 else r[-1])
                        check(f"High priority avg satisfaction ~{expected_high[3]}",
                              num_close(sat, float(expected_high[3]), tol=0.1),
                              f"Got {sat}, expected {expected_high[3]}")

    # Issue_Type_Breakdown sheet
    it_idx = next((i for i, s in enumerate(sheets_lower) if "issue" in s or "type" in s), 2)
    if it_idx < len(sheets):
        ws_it = wb[sheets[it_idx]]
        rows_it = list(ws_it.iter_rows(min_row=2, values_only=True))
        check("Issue_Type_Breakdown has 7 rows", len(rows_it) == 7,
              f"Found {len(rows_it)} rows")

        if rows_it:
            all_text_it = " ".join(str(c) for r in rows_it for c in r if c).lower()
            check("Has Bug issue type", "bug" in all_text_it)
            check("Has Performance Issue type", "performance" in all_text_it)

    # Executive_Summary sheet
    es_idx = next((i for i, s in enumerate(sheets_lower) if "executive" in s or "summary" in s), 3)
    if es_idx < len(sheets):
        ws_es = wb[sheets[es_idx]]
        rows_es = list(ws_es.iter_rows(min_row=2, values_only=True))
        check("Executive_Summary has at least 5 rows", len(rows_es) >= 5,
              f"Found {len(rows_es)} rows")

        if rows_es:
            summary_dict = {}
            for r in rows_es:
                if r[0]:
                    summary_dict[str(r[0]).lower()] = r[1]

            # Total Problem Products
            tp_key = next((k for k in summary_dict if "total" in k and "problem" in k), None)
            if tp_key:
                check("Total Problem Products correct",
                      num_close(summary_dict[tp_key], expected_count, tol=2),
                      f"Got {summary_dict[tp_key]}, expected {expected_count}")

            # Critical Products
            cp_key = next((k for k in summary_dict if "critical" in k), None)
            if cp_key:
                check("Critical Products count correct",
                      num_close(summary_dict[cp_key], len(gt["critical"]), tol=1),
                      f"Got {summary_dict[cp_key]}, expected {len(gt['critical'])}")

            # Total Support Tickets
            tt_key = next((k for k in summary_dict if "ticket" in k), None)
            if tt_key:
                check("Total Support Tickets correct",
                      num_close(summary_dict[tt_key], gt["total_tickets"], tol=100),
                      f"Got {summary_dict[tt_key]}, expected {gt['total_tickets']}")

            # Overall Avg Satisfaction
            sat_key = next((k for k in summary_dict if "satisfaction" in k), None)
            if sat_key:
                check("Overall Avg Satisfaction correct",
                      num_close(summary_dict[sat_key], gt["overall_sat"], tol=0.1),
                      f"Got {summary_dict[sat_key]}, expected {gt['overall_sat']}")

            # Highest Risk Category
            cat_key = next((k for k in summary_dict if "category" in k or "risk" in k), None)
            if cat_key:
                check("Highest Risk Category is Electronics",
                      str(summary_dict[cat_key]).lower() == "electronics",
                      f"Got {summary_dict[cat_key]}")


def check_notion(gt):
    print("\n=== Check 2: Notion Database ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title, properties FROM notion.databases
        WHERE title::text ILIKE '%Support Quality%' OR title::text ILIKE '%support_quality%'
    """)
    dbs = cur.fetchall()
    check("Notion database 'Support Quality Tracker' exists", len(dbs) >= 1,
          f"Found {len(dbs)} matching databases")

    if dbs:
        db_id = str(dbs[0][0])
        props = dbs[0][2] if dbs[0][2] else {}

        # Check properties exist
        prop_names_lower = {k.lower(): k for k in props.keys()}
        check("Has Severity property", any("severity" in k for k in prop_names_lower))
        check("Has Status property", any("status" in k for k in prop_names_lower))
        check("Has Product property", any("product" in k for k in prop_names_lower))

        # Check pages
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE parent::text LIKE %s AND NOT archived
        """, (f'%{db_id}%',))
        pages = cur.fetchall()
        check(f"Has at least 6 entries for critical products",
              len(pages) >= 6,
              f"Found {len(pages)} pages")

        if pages:
            # Check at least one page has Critical severity
            critical_found = False
            for page in pages:
                page_props = page[1] if page[1] else {}
                props_text = json.dumps(page_props).lower()
                if "critical" in props_text:
                    critical_found = True
                    break
            check("At least one entry has Critical severity", critical_found)

            # Check pages reference expected products
            all_page_text = " ".join(json.dumps(p[1]).lower() for p in pages if p[1])
            check("Pages reference AGARO or tripod product",
                  "agaro" in all_page_text or "tripod" in all_page_text,
                  f"Text snippet: {all_page_text[:200]}")

    cur.close()
    conn.close()


def check_emails(gt):
    print("\n=== Check 3: Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check support team email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%support quality%' AND subject ILIKE '%priority%'
        AND to_addr::text ILIKE '%support_team%'
    """)
    support_emails = cur.fetchall()
    check("Email to support_team sent", len(support_emails) >= 1,
          f"Found {len(support_emails)} matching emails")

    if support_emails:
        body = (support_emails[0][2] or "").lower()
        check("Support email mentions satisfaction",
              "satisfaction" in body or "3.2" in body,
              f"Body snippet: {body[:150]}")

    # Check product team email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%support quality%' AND subject ILIKE '%product%'
        AND to_addr::text ILIKE '%product_team%'
    """)
    product_emails = cur.fetchall()
    check("Email to product_team sent", len(product_emails) >= 1,
          f"Found {len(product_emails)} matching emails")

    if product_emails:
        body = (product_emails[0][2] or "").lower()
        check("Product email mentions critical products",
              "critical" in body or "severity" in body or "agaro" in body,
              f"Body snippet: {body[:150]}")
        check("Product email mentions total problem products",
              "36" in body or str(len(gt["problem_list"])) in body or "problem" in body,
              f"Body snippet: {body[:150]}")

    cur.close()
    conn.close()


def check_reverse_validation(gt):
    print("\n=== Reverse Validation ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # Check Notion does not include non-critical (below 80th percentile) products
        cur.execute("""
            SELECT id, title FROM notion.databases
            WHERE title::text ILIKE '%%Support Quality%%' OR title::text ILIKE '%%support_quality%%'
        """)
        dbs = cur.fetchall()
        if dbs:
            db_id = str(dbs[0][0])
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent::text LIKE %s AND NOT archived
            """, (f'%{db_id}%',))
            pages = cur.fetchall()
            # Should have no more than ~top 20% products; certainly not all problem products
            max_expected = len(gt["critical"]) + 2  # small tolerance
            check("Notion does not include non-critical products",
                  len(pages) <= max_expected,
                  f"Found {len(pages)} pages, expected at most {max_expected} (critical={len(gt['critical'])})")

        # Check no emails sent to wrong recipients
        noise_recipients = [
            "all-staff@company.com",
            "hr@company.com",
            "newsletter@company.com",
            "finance@company.com",
        ]
        for addr in noise_recipients:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s",
                (f"%{addr}%",),
            )
            cnt = cur.fetchone()[0]
            check(f"No email sent to noise recipient {addr}", cnt == 0,
                  f"Found {cnt} emails to {addr}")
    except Exception as e:
        check("Reverse validation", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python Scripts ===")
    for script in ["correlate_issues.py", "support_metrics.py"]:
        path = os.path.join(workspace, script)
        check(f"{script} exists", os.path.exists(path), f"Not found at {path}")

    # Check output JSON files
    for jf in ["problem_products.json", "support_analysis.json"]:
        path = os.path.join(workspace, jf)
        check(f"{jf} exists", os.path.exists(path), f"Not found at {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    gt = get_groundtruth_from_db()

    check_excel(args.agent_workspace, gt)
    check_notion(gt)
    check_emails(gt)
    check_scripts(args.agent_workspace)
    check_reverse_validation(gt)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

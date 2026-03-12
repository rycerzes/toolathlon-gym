"""
Evaluation script for notion-canvas-student-tracker task.

Checks:
1. Excel file (Student_Progress.xlsx) - correct student data, two sheets
2. Notion page - at-risk students page created
3. Google Calendar events - 3 tutoring sessions
4. Email sent to lead instructor

Usage:
    python -m evaluation.main \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00"
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def str_match(a, b):
    """Case-insensitive, whitespace-normalized comparison."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=0.5):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def int_close(a, b, tol=5):
    """Compare two integer values with tolerance."""
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


# ============================================================================
# Compute expected values from Canvas DB
# ============================================================================

def get_expected_data():
    """Query Canvas DB to compute expected student data."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get all students with at least one graded submission
    cur.execute("""
        SELECT u.id AS student_id, u.name AS student_name,
            COUNT(DISTINCT s.assignment_id) AS assignments_submitted,
            ROUND(AVG(s.score::float)::numeric, 2) AS avg_score
        FROM canvas.enrollments e
        JOIN canvas.courses c ON e.course_id = c.id
        JOIN canvas.users u ON e.user_id = u.id
        JOIN canvas.assignments a ON a.course_id = c.id
        JOIN canvas.submissions s ON s.assignment_id = a.id
            AND s.user_id = u.id AND s.score IS NOT NULL
        WHERE c.course_code = 'FFF-2014J' AND e.type = 'StudentEnrollment'
        GROUP BY u.id, u.name
        ORDER BY avg_score ASC
    """)
    all_students = cur.fetchall()

    # Get teacher info (first alphabetically)
    cur.execute("""
        SELECT u.name, u.email
        FROM canvas.users u
        JOIN canvas.enrollments e ON e.user_id = u.id
        WHERE e.course_id = (
            SELECT id FROM canvas.courses WHERE course_code = 'FFF-2014J'
        ) AND e.type = 'TeacherEnrollment'
        ORDER BY u.name ASC
        LIMIT 1
    """)
    teacher = cur.fetchone()

    cur.close()
    conn.close()

    total_students = len(all_students)
    at_risk_students = [s for s in all_students if s[3] < 50]
    at_risk_count = len(at_risk_students)
    at_risk_pct = round(at_risk_count / total_students * 100, 2) if total_students > 0 else 0.0

    return {
        "all_students": all_students,
        "at_risk_students": at_risk_students,
        "total_students": total_students,
        "at_risk_count": at_risk_count,
        "at_risk_pct": at_risk_pct,
        "teacher_name": teacher[0] if teacher else None,
        "teacher_email": teacher[1] if teacher else None,
        "top10_at_risk": at_risk_students[:10],
    }


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace, expected):
    """Compare Student_Progress.xlsx from agent vs groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Student_Progress.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Student_Progress.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: All Students ---
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    agent_ws = get_sheet(agent_wb, "All Students")
    gt_ws = get_sheet(gt_wb, "All Students")

    if agent_ws is None:
        record("Sheet 'All Students' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'All Students' exists", True)

        # Check headers
        agent_headers = [str(c.value).strip() if c.value else "" for c in agent_ws[1]]
        gt_headers = [str(c.value).strip() if c.value else "" for c in gt_ws[1]]
        headers_ok = all(
            str_match(a, g) for a, g in zip(agent_headers, gt_headers)
        ) and len(agent_headers) >= len(gt_headers)
        record("All Students headers match", headers_ok,
               f"Expected: {gt_headers}, Got: {agent_headers}")

        # Check data row count
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("All Students row count", int_close(len(agent_rows), len(gt_rows), 10),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Check first 5 rows (lowest scoring students)
        for i in range(min(5, len(gt_rows), len(agent_rows))):
            gt_row = gt_rows[i]
            agent_row = agent_rows[i]
            gt_id = gt_row[0]
            ag_id = agent_row[0]

            row_diffs = []
            # Student_ID
            if not (str(gt_id) == str(ag_id) if gt_id and ag_id else True):
                row_diffs.append(f"Student_ID: expected {gt_id}, got {ag_id}")
            # Avg_Score
            if not num_close(agent_row[3], gt_row[3], 1.0):
                row_diffs.append(f"Avg_Score: expected {gt_row[3]}, got {agent_row[3]}")
            # Status
            if not str_match(agent_row[4], gt_row[4]):
                row_diffs.append(f"Status: expected {gt_row[4]}, got {agent_row[4]}")

            if row_diffs:
                record(f"Row {i+1} data (student {gt_row[1]})", False, "; ".join(row_diffs))
                all_ok = False
            else:
                record(f"Row {i+1} data (student {gt_row[1]})", True)

        # Check last few rows are "Passing" status
        if len(agent_rows) > 10:
            last_row = agent_rows[-1]
            record("Last row has 'Passing' status",
                   str_match(last_row[4], "Passing"),
                   f"Got: {last_row[4]}")

        # Verify sort order (ascending by Avg_Score)
        scores = []
        for row in agent_rows:
            try:
                scores.append(float(row[3]))
            except (TypeError, ValueError):
                pass
        if len(scores) > 1:
            is_sorted = all(scores[i] <= scores[i+1] for i in range(len(scores)-1))
            record("Rows sorted by Avg_Score ascending", is_sorted,
                   f"First few scores: {scores[:5]}")
        else:
            record("Rows sorted by Avg_Score ascending", False, "Too few scores to verify")

    # --- Sheet 2: At Risk Summary ---
    agent_ws2 = get_sheet(agent_wb, "At Risk Summary")

    if agent_ws2 is None:
        record("Sheet 'At Risk Summary' exists", False,
               f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'At Risk Summary' exists", True)

        # Read label-value pairs
        def read_summary(ws):
            data = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row and row[0]:
                    label = str(row[0]).strip()
                    val = row[1] if len(row) > 1 and row[1] is not None else ""
                    data[label.lower()] = val
            return data

        agent_summary = read_summary(agent_ws2)

        expected_labels = {
            "total_students": expected["total_students"],
            "at_risk_count": expected["at_risk_count"],
            "at_risk_percentage": expected["at_risk_pct"],
        }

        for label, expected_val in expected_labels.items():
            ag_val = agent_summary.get(label)
            if ag_val is None:
                record(f"Summary: {label} present", False, "Missing from agent output")
                all_ok = False
                continue

            if label in ("total_students", "at_risk_count"):
                ok = int_close(ag_val, expected_val, 5)
            else:
                ok = num_close(ag_val, expected_val, 1.0)

            record(f"Summary: {label}", ok,
                   f"Expected '{expected_val}', got '{ag_val}'")
            if not ok:
                all_ok = False

    return all_ok


# ============================================================================
# Check 2: Notion page
# ============================================================================

def check_notion(expected):
    """Verify Notion page was created for at-risk students."""
    print("\n=== Checking Notion ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, properties
        FROM notion.pages
        WHERE archived = false AND in_trash = false
    """)
    pages = cur.fetchall()

    print(f"[check_notion] Found {len(pages)} Notion pages.")

    # Look for a page with FFF or at-risk or finance in title
    found_page = None
    for page_id, props in pages:
        props_str = json.dumps(props).lower() if props else ""
        if ("fff" in props_str and "at-risk" in props_str) or \
           ("fff-2014j" in props_str) or \
           ("at-risk students" in props_str and "fff" in props_str):
            found_page = (page_id, props)
            break

    if found_page is None:
        # Broader search
        for page_id, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "fff" in props_str or "finance" in props_str or "at-risk" in props_str:
                found_page = (page_id, props)
                break

    record("Notion page for at-risk students exists", found_page is not None,
           f"No page with FFF/at-risk/finance found among {len(pages)} pages")

    if found_page:
        page_id = found_page[0]

        # Check that page has content blocks
        cur.execute("""
            SELECT COUNT(*) FROM notion.blocks
            WHERE parent_id = %s AND archived = false
        """, (page_id,))
        block_count = cur.fetchone()[0]
        record("Notion page has content blocks", block_count > 0,
               f"Found {block_count} blocks")

        # Check that content mentions at-risk students
        cur.execute("""
            SELECT block_data::text FROM notion.blocks
            WHERE parent_id = %s AND archived = false
        """, (page_id,))
        blocks = cur.fetchall()
        all_block_text = " ".join(str(b[0]).lower() for b in blocks)

        # Check for at least some student names from top 10
        top10 = expected["top10_at_risk"]
        students_found = 0
        for student in top10:
            student_name = student[1].lower()
            # Check if any part of the name appears in blocks
            name_parts = student_name.split()
            if any(part in all_block_text for part in name_parts):
                students_found += 1

        record("Notion page mentions at-risk students",
               students_found >= 3,
               f"Found references to {students_found}/10 top at-risk students in blocks")

    cur.close()
    conn.close()

    return found_page is not None


# ============================================================================
# Check 3: Google Calendar events
# ============================================================================

def check_gcal():
    """Verify 3 tutoring sessions were created."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    for ev in events:
        print(f"  Event: {ev[0]} | {ev[2]} - {ev[3]}")

    record("At least 3 calendar events created", len(events) >= 3,
           f"Found {len(events)}")

    all_ok = True

    expected_dates = ["2026-03-10", "2026-03-12", "2026-03-14"]

    # Check for tutoring sessions
    tutoring_events = []
    for summary, description, start_dt, end_dt in events:
        summary_lower = (summary or "").lower()
        if "tutoring" in summary_lower or ("fff" in summary_lower and "session" in summary_lower):
            tutoring_events.append((summary, description, start_dt, end_dt))

    record("Found tutoring session events", len(tutoring_events) >= 3,
           f"Found {len(tutoring_events)} tutoring events")

    # Check each expected date
    for expected_date in expected_dates:
        found = False
        for summary, description, start_dt, end_dt in tutoring_events:
            if start_dt is not None:
                start_date_str = start_dt.strftime("%Y-%m-%d")
                if start_date_str == expected_date:
                    found = True

                    # Check title contains FFF
                    summary_lower = (summary or "").lower()
                    record(f"gcal {expected_date}: title mentions FFF",
                           "fff" in summary_lower,
                           f"Title: {summary}")

                    # Check description
                    desc_lower = (description or "").lower()
                    record(f"gcal {expected_date}: description mentions tutoring or finance",
                           "tutor" in desc_lower or "finance" in desc_lower or "at-risk" in desc_lower,
                           f"Description: {(description or '')[:100]}")
                    break

        if not found:
            record(f"gcal: tutoring event on {expected_date} exists", False,
                   f"No tutoring event found for {expected_date}")
            all_ok = False

    return all_ok


# ============================================================================
# Check 4: Email to instructor
# ============================================================================

def check_emails(expected):
    """Verify alert email was sent to lead instructor."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")

    record("At least 1 email sent", len(all_emails) >= 1,
           f"Found {len(all_emails)}")

    all_ok = True

    # Look for the at-risk alert email
    found = False
    for subject, from_addr, to_addr, body_text in all_emails:
        subject_lower = (subject or "").lower()
        if "fff" in subject_lower and ("at-risk" in subject_lower or "alert" in subject_lower
                                        or "at risk" in subject_lower):
            found = True

            # Check sender
            from_str = str(from_addr or "").lower()
            record("email: from academic-affairs",
                   "academic-affairs@openuniversity.ac.uk" in from_str,
                   f"From: {from_addr}")

            # Check recipient is the lead instructor
            to_str = ""
            if isinstance(to_addr, list):
                to_str = " ".join(str(r).lower() for r in to_addr)
            elif isinstance(to_addr, str):
                try:
                    parsed = json.loads(to_addr)
                    if isinstance(parsed, list):
                        to_str = " ".join(str(r).lower() for r in parsed)
                    else:
                        to_str = str(to_addr).lower()
                except (json.JSONDecodeError, TypeError):
                    to_str = str(to_addr).lower()

            teacher_email = expected["teacher_email"]
            if teacher_email:
                record("email: sent to lead instructor",
                       teacher_email.lower() in to_str,
                       f"Expected {teacher_email}, got to_addr: {to_addr}")
            else:
                record("email: teacher email known", False, "Could not determine teacher email")

            # Check body mentions key info
            body_lower = (body_text or "").lower()
            record("email: body mentions at-risk count",
                   str(expected["at_risk_count"]) in (body_text or ""),
                   f"Expected {expected['at_risk_count']} in body")

            record("email: body mentions tutoring sessions",
                   "march" in body_lower or "tutoring" in body_lower or "session" in body_lower,
                   "No mention of tutoring or March dates")

            break

    if not found:
        record("email: at-risk alert email exists", False,
               "No email with 'FFF' and 'at-risk'/'alert' in subject")
        all_ok = False

    return all_ok


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Compute expected values from Canvas DB
    expected = get_expected_data()
    print(f"[eval] Expected: {expected['total_students']} total students, "
          f"{expected['at_risk_count']} at-risk, "
          f"{expected['at_risk_pct']}% at-risk rate")
    print(f"[eval] Lead instructor: {expected['teacher_name']} ({expected['teacher_email']})")

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace, expected)
    notion_ok = check_notion(expected)
    gcal_ok = check_gcal()
    email_ok = check_emails(expected)

    all_passed = excel_ok and notion_ok and gcal_ok and email_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
            "details": {
                "excel": excel_ok,
                "notion": notion_ok,
                "gcal": gcal_ok,
                "email": email_ok,
            },
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

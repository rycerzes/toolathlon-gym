"""Evaluation for yf-portfolio-stress-test-excel-word-gcal."""
import argparse
import os
import sys

import psycopg2


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    errors = []
    import openpyxl

    agent_path = os.path.join(agent_workspace, "Stress_Test_Report.xlsx")
    gt_path = os.path.join(groundtruth_workspace, "Stress_Test_Report.xlsx")

    if not os.path.exists(agent_path):
        return ["Stress_Test_Report.xlsx not found in agent workspace"]
    if not os.path.exists(gt_path):
        return ["Stress_Test_Report.xlsx not found in groundtruth workspace"]

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)

        # --- Sheet: Portfolio Overview ---
        agent_rows = load_sheet_rows(wb_agent, "Portfolio Overview")
        gt_rows = load_sheet_rows(wb_gt, "Portfolio Overview")
        if agent_rows is None:
            errors.append("Sheet 'Portfolio Overview' not found")
        elif gt_rows is None:
            errors.append("Groundtruth 'Portfolio Overview' missing")
        else:
            agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
            gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]

            if len(agent_data) < len(gt_data):
                errors.append(f"Portfolio Overview: {len(agent_data)} rows, expected {len(gt_data)}")

            # Check each stock
            agent_lookup = {str(r[0]).strip().upper(): r for r in agent_data}
            gt_lookup = {str(r[0]).strip().upper(): r for r in gt_data}

            for sym, gt_row in gt_lookup.items():
                if sym not in agent_lookup:
                    errors.append(f"Portfolio Overview: {sym} missing")
                    continue
                a_row = agent_lookup[sym]
                # Allocation_Pct (col 1)
                if not num_close(a_row[1], gt_row[1], 0.5):
                    errors.append(f"{sym} Allocation_Pct: {a_row[1]} vs expected {gt_row[1]}")
                # Current_Price (col 3)
                if not num_close(a_row[3], gt_row[3], 1.0):
                    errors.append(f"{sym} Current_Price: {a_row[3]} vs expected {gt_row[3]}")
                # Monthly_Volatility_Pct (col 5)
                if not num_close(a_row[5], gt_row[5], 0.5):
                    errors.append(f"{sym} Monthly_Volatility_Pct: {a_row[5]} vs expected {gt_row[5]}")
                # Sharpe_Ratio (col 7)
                if not num_close(a_row[7], gt_row[7], 0.1):
                    errors.append(f"{sym} Sharpe_Ratio: {a_row[7]} vs expected {gt_row[7]}")

        # --- Sheet: Stress Scenarios ---
        agent_rows2 = load_sheet_rows(wb_agent, "Stress Scenarios")
        gt_rows2 = load_sheet_rows(wb_gt, "Stress Scenarios")
        if agent_rows2 is None:
            errors.append("Sheet 'Stress Scenarios' not found")
        elif gt_rows2 is None:
            errors.append("Groundtruth 'Stress Scenarios' missing")
        else:
            agent_data2 = [r for r in agent_rows2[1:] if r and r[0] is not None]
            gt_data2 = [r for r in gt_rows2[1:] if r and r[0] is not None]

            if len(agent_data2) < len(gt_data2) - 4:  # Allow some tolerance for summary rows
                errors.append(f"Stress Scenarios: {len(agent_data2)} rows, expected ~{len(gt_data2)}")

            # Check portfolio totals for each scenario
            gt_totals = {}
            for r in gt_data2:
                if r[1] and str(r[1]).strip() == "Portfolio_Total":
                    gt_totals[str(r[0]).strip()] = (r[4], r[5])  # Scenario_Value, Scenario_PnL

            agent_totals = {}
            for r in agent_data2:
                if r[1] and str(r[1]).strip() == "Portfolio_Total":
                    agent_totals[str(r[0]).strip()] = (r[4], r[5])

            for sc_name, (gt_val, gt_pnl) in gt_totals.items():
                if sc_name not in agent_totals:
                    errors.append(f"Stress Scenarios: Portfolio_Total row missing for {sc_name}")
                else:
                    a_val, a_pnl = agent_totals[sc_name]
                    if not num_close(a_val, gt_val, 5.0):
                        errors.append(f"{sc_name} total value: {a_val} vs expected {gt_val}")
                    if not num_close(a_pnl, gt_pnl, 5.0):
                        errors.append(f"{sc_name} total PnL: {a_pnl} vs expected {gt_pnl}")

        # --- Sheet: Risk Summary ---
        agent_rows3 = load_sheet_rows(wb_agent, "Risk Summary")
        gt_rows3 = load_sheet_rows(wb_gt, "Risk Summary")
        if agent_rows3 is None:
            errors.append("Sheet 'Risk Summary' not found")
        elif gt_rows3 is None:
            errors.append("Groundtruth 'Risk Summary' missing")
        else:
            agent_data3 = [r for r in agent_rows3[1:] if r and r[0] is not None]
            gt_data3 = [r for r in gt_rows3[1:] if r and r[0] is not None]

            agent_metrics = {str(r[0]).strip().lower(): r[1] for r in agent_data3}
            gt_metrics = {str(r[0]).strip().lower(): r[1] for r in gt_data3}

            for metric, gt_val in gt_metrics.items():
                if metric not in agent_metrics:
                    errors.append(f"Risk Summary: {metric} missing")
                    continue
                a_val = agent_metrics[metric]
                if metric in ("worst_scenario", "best_scenario", "breach_threshold"):
                    if str(a_val).strip().lower() != str(gt_val).strip().lower():
                        errors.append(f"Risk Summary {metric}: '{a_val}' vs expected '{gt_val}'")
                elif metric in ("total_portfolio_value",):
                    if not num_close(a_val, gt_val, 100):
                        errors.append(f"Risk Summary {metric}: {a_val} vs expected {gt_val}")
                elif metric in ("portfolio_var_95", "worst_scenario_loss", "best_scenario_pnl"):
                    if not num_close(a_val, gt_val, 500):
                        errors.append(f"Risk Summary {metric}: {a_val} vs expected {gt_val}")
                elif metric in ("max_historical_drawdown_pct", "worst_scenario_loss_pct"):
                    if not num_close(a_val, gt_val, 1.0):
                        errors.append(f"Risk Summary {metric}: {a_val} vs expected {gt_val}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    try:
        from docx import Document

        path = os.path.join(agent_workspace, "Risk_Assessment.docx")
        if not os.path.exists(path):
            return ["Risk_Assessment.docx not found"]

        doc = Document(path)
        full_text = "\n".join([p.text for p in doc.paragraphs]).lower()

        required_sections = [
            "executive summary",
            "portfolio composition",
            "stress test results",
            "risk metrics",
            "scenario comparison",
            "recommendation",
        ]
        for section in required_sections:
            if section not in full_text:
                errors.append(f"Word doc missing section: {section}")

        # Check for key content
        if "var" not in full_text and "value at risk" not in full_text:
            errors.append("Word doc missing VaR discussion")
        if "drawdown" not in full_text:
            errors.append("Word doc missing drawdown discussion")

    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=5432,
            dbname="toolathlon_gym",
            user="postgres",
            password="postgres",
        )
        cur = conn.cursor()

        # Check for 4 stress test review meetings
        cur.execute("""
            SELECT summary, start_datetime::date, description
            FROM gcal.events
            WHERE LOWER(summary) LIKE '%stress test%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 4:
            errors.append(f"Expected 4 Stress Test Review meetings, found {len(rows)}")

        # Check scenario names in titles
        expected_scenarios = ["market crash", "sector rotation", "inflation shock", "historical replay"]
        found_scenarios = set()
        for summary, date, desc in rows:
            summary_lower = summary.lower() if summary else ""
            for sc in expected_scenarios:
                if sc in summary_lower:
                    found_scenarios.add(sc)

        for sc in expected_scenarios:
            if sc not in found_scenarios:
                errors.append(f"No calendar event found for scenario: {sc}")

        # Check dates are in consecutive weeks starting 2026-03-16
        if rows:
            import datetime

            expected_dates = [
                datetime.date(2026, 3, 16),
                datetime.date(2026, 3, 23),
                datetime.date(2026, 3, 30),
                datetime.date(2026, 4, 6),
            ]
            actual_dates = sorted([r[1] for r in rows])
            for i, exp_date in enumerate(expected_dates):
                if i < len(actual_dates):
                    if actual_dates[i] != exp_date:
                        errors.append(
                            f"Meeting {i+1} date: {actual_dates[i]}, expected {exp_date}"
                        )

    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )
    gt_ws = args.groundtruth_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

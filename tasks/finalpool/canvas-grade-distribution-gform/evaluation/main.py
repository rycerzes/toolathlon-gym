"""
Evaluation script for canvas-grade-distribution-gform task.

Checks:
1. Excel with Spring 2014 course grade data (verified against canvas DB)
2. Google Form with >=4 questions
3. Email with correct subject
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_expected_course_data():
    """Query actual grade stats from canvas DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT c.course_code, c.name,
               AVG(s.score) as avg_grade,
               PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY s.score) as median_grade,
               COUNT(DISTINCT s.user_id) as total_students
        FROM canvas.courses c
        JOIN canvas.assignments a ON a.course_id = c.id
        JOIN canvas.submissions s ON s.assignment_id = a.id
        WHERE c.name LIKE '%%Spring 2014%%'
        AND s.score IS NOT NULL
        GROUP BY c.course_code, c.name
        ORDER BY c.name
    """)
    course_stats = cur.fetchall()
    cur.close()
    conn.close()
    return course_stats


def check_excel(workspace):
    """Check Excel file."""
    from openpyxl import load_workbook

    errors = []
    xlsx_path = os.path.join(workspace, "Grade_Distribution_Report.xlsx")
    if not os.path.exists(xlsx_path):
        return ["Grade_Distribution_Report.xlsx not found"]

    course_stats = get_expected_course_data()

    wb = load_workbook(xlsx_path)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Course Grades sheet
    if "course grades" not in sheet_names_lower:
        errors.append(f"Missing 'Course Grades' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("course grades")]]
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]

        for rh in ["course_code", "course_name", "avg_grade", "total_students"]:
            if not any(rh in h or rh.replace("_", "") in h.replace("_", "") for h in headers):
                errors.append(f"Course Grades missing header: {rh}")

        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        if data_rows < len(course_stats):
            errors.append(f"Course Grades has {data_rows} rows, expected at least {len(course_stats)}")

        # Verify course codes are present
        code_col = None
        for idx, h in enumerate(headers):
            if "course_code" in h or "coursecode" in h or "code" in h:
                code_col = idx
                break

        if code_col is not None:
            found_codes = set()
            for row in ws.iter_rows(min_row=2):
                if row[code_col].value:
                    found_codes.add(str(row[code_col].value).strip().upper())
            expected_codes = set(cs[0].upper() for cs in course_stats)
            missing_codes = expected_codes - found_codes
            if missing_codes:
                errors.append(f"Missing course codes: {missing_codes}")

    # Check Summary sheet
    if "summary" not in sheet_names_lower:
        errors.append(f"Missing 'Summary' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        summary_data = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                key = str(row[0].value).lower().replace(" ", "_")
                summary_data[key] = row[1].value

        # Check total courses
        total_key = None
        for k in summary_data:
            if "total" in k and "course" in k:
                total_key = k
                break
        if total_key:
            try:
                val = int(float(summary_data[total_key]))
                if val != len(course_stats):
                    errors.append(f"Total_Courses: got {val}, expected {len(course_stats)}")
            except (TypeError, ValueError):
                errors.append(f"Cannot parse Total_Courses: {summary_data[total_key]}")
        else:
            errors.append("Summary missing Total_Courses row")

    return errors


def check_gform(cur):
    """Check Google Form."""
    errors = []

    cur.execute("""
        SELECT id, title
        FROM gform.forms
        WHERE LOWER(title) LIKE '%%spring 2014%%'
        AND (LOWER(title) LIKE '%%survey%%' OR LOWER(title) LIKE '%%feedback%%')
        ORDER BY created_at DESC
        LIMIT 1
    """)
    form_row = cur.fetchone()

    if not form_row:
        # Try broader search
        cur.execute("""
            SELECT id, title
            FROM gform.forms
            WHERE LOWER(title) LIKE '%%spring 2014%%'
            ORDER BY created_at DESC
            LIMIT 1
        """)
        form_row = cur.fetchone()

    if not form_row:
        errors.append("No Google Form with 'Spring 2014' in title found")
        return errors

    form_id = form_row[0]
    form_title = form_row[1]

    cur.execute("""
        SELECT title, question_type, required
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()

    if len(questions) < 4:
        errors.append(f"Form has {len(questions)} questions, expected at least 4")

    return errors


def check_email(cur):
    """Check email."""
    errors = []

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%%spring 2014%%'
        AND (LOWER(subject) LIKE '%%survey%%' OR LOWER(subject) LIKE '%%feedback%%')
    """)
    emails = cur.fetchall()

    if not emails:
        errors.append("No email with 'Spring 2014' and 'survey/feedback' in subject found")
    else:
        to_str = str(emails[0][2]).lower()
        if "students@university.edu" not in to_str:
            errors.append(f"Email not sent to students@university.edu, to_addr: {emails[0][2]}")

    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    # Check Excel
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    # Check GForm and Email
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        print("\n=== Checking Google Form ===")
        gform_errors = check_gform(cur)
        if gform_errors:
            for e in gform_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(gform_errors)
        else:
            print("  [PASS] Google Form check passed")

        print("\n=== Checking Email ===")
        email_errors = check_email(cur)
        if email_errors:
            for e in email_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(email_errors)
        else:
            print("  [PASS] Email check passed")

        cur.close()
        conn.close()
    except Exception as e:
        err = f"DB check error: {e}"
        print(f"  [FAIL] {err}")
        all_errors.append(err)

    # Summary
    print(f"\n=== SUMMARY ===")
    if all_errors:
        for e in all_errors:
            print(f"  [ERROR] {e}")
        print("  Overall: FAIL")
    else:
        print("  Overall: PASS")

    if args.res_log_file:
        result = {"errors": all_errors, "success": len(all_errors) == 0}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(1 if all_errors else 0)


if __name__ == "__main__":
    main()

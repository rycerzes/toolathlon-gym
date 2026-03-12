"""
Evaluation script for howtocook-playwright-price task.

Checks:
1. Excel file (Recipe_Cost_Analysis.xlsx) exists with 2 sheets
2. Dish Costs sheet has 5 data rows
3. Ingredient Prices sheet has data
4. Google Sheet exists with recipe/cost in title
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    """Check the Excel output file."""
    print("\n=== Check 1: Excel File ===")

    excel_path = os.path.join(agent_workspace, "Recipe_Cost_Analysis.xlsx")
    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return

    record("Excel file readable", True)

    # Check that there are at least 2 sheets
    record("Excel has >= 2 sheets", len(wb.sheetnames) >= 2,
           f"Found sheets: {wb.sheetnames}")

    # Find Dish Costs sheet
    dc_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "dish" in name_lower or ("cost" in name_lower and "ingredient" not in name_lower):
            dc_sheet = name
            break
    if not dc_sheet:
        # fallback: any sheet with "cost" in name
        for name in wb.sheetnames:
            if "cost" in name.lower():
                dc_sheet = name
                break

    if not dc_sheet:
        record("Dish Costs sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Dish Costs sheet exists", True)
        ws = wb[dc_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)] if len(rows) > 1 else []
        record("Dish Costs has 5 data rows", len(data_rows) >= 5,
               f"Found {len(data_rows)} data rows")

        # Check that some expected dish names appear
        all_text = " ".join(str(c) for r in data_rows for c in r if c is not None)
        has_dishes = sum(1 for d in ["可乐鸡翅", "水煮鱼", "金针菇", "年糕", "蘑菇汤"]
                        if d in all_text)
        record("Dish Costs contains expected dish names", has_dishes >= 3,
               f"Found {has_dishes}/5 expected dish keywords")

    # Find Ingredient Prices sheet
    ip_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "ingredient" in name_lower or "price" in name_lower:
            ip_sheet = name
            break

    if not ip_sheet:
        record("Ingredient Prices sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Ingredient Prices sheet exists", True)
        ws = wb[ip_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)] if len(rows) > 1 else []
        record("Ingredient Prices has >= 10 rows", len(data_rows) >= 10,
               f"Found {len(data_rows)} data rows")

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Recipe_Cost_Analysis.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()

    wb.close()


def check_gsheet():
    """Check that a Google Sheet was created with recipe/cost data."""
    print("\n=== Check 2: Google Sheet ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()

        if not spreadsheets:
            record("Google Sheet exists", False, "No spreadsheets found")
            cur.close()
            conn.close()
            return

        record("Google Sheet exists", True)

        # Find one with "cost" or "recipe" in title
        target_ss = None
        for ss_id, ss_title in spreadsheets:
            title_lower = (ss_title or "").lower()
            if "cost" in title_lower or "recipe" in title_lower:
                target_ss = (ss_id, ss_title)
                break

        if not target_ss:
            target_ss = (spreadsheets[0][0], spreadsheets[0][1])

        record("Google Sheet title contains 'cost' or 'recipe'",
               "cost" in (target_ss[1] or "").lower() or "recipe" in (target_ss[1] or "").lower(),
               f"Title: {target_ss[1]}")

        # Check cells exist
        cur.execute("""
            SELECT c.row_index, c.col_index, c.value
            FROM gsheet.cells c
            JOIN gsheet.sheets s ON c.sheet_id = s.id
            WHERE s.spreadsheet_id = %s
            ORDER BY c.row_index, c.col_index
        """, (target_ss[0],))
        cells = cur.fetchall()

        record("Google Sheet has data", len(cells) >= 5,
               f"Found {len(cells)} cells")

        cur.close()
        conn.close()

    except Exception as e:
        record("Google Sheet check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gsheet()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

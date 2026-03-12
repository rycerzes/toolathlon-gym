"""Evaluation for sf-hr-department-budget."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace, gt_dir):
    """Check Budget_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Budget_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Budget_Analysis.xlsx")

    if not os.path.exists(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Sheet 1: Department Budget
    a_rows = load_sheet_rows(agent_wb, "Department Budget")
    g_rows = load_sheet_rows(gt_wb, "Department Budget")

    if a_rows is None:
        record("Sheet 'Department Budget' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Department Budget' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        record("Department Budget row count", len(a_data) == len(g_data),
               f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Department '{g_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            errors = []
            # Budget (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100000):
                    errors.append(f"Budget: {a_row[1]} vs {g_row[1]}")

            # Actual_Spend (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 100000):
                    errors.append(f"Actual_Spend: {a_row[2]} vs {g_row[2]}")

            # Variance (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 200000):
                    errors.append(f"Variance: {a_row[3]} vs {g_row[3]}")

            # Variance_Pct (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1.0):
                    errors.append(f"Variance_Pct: {a_row[4]} vs {g_row[4]}")

            # Employee_Count (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 10):
                    errors.append(f"Employee_Count: {a_row[5]} vs {g_row[5]}")

            # Avg_Salary (col 6)
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 500):
                    errors.append(f"Avg_Salary: {a_row[6]} vs {g_row[6]}")

            if errors:
                record(f"Department '{g_row[0]}' data", False, "; ".join(errors))
                all_ok = False
            else:
                record(f"Department '{g_row[0]}' data", True)

    # Sheet 2: Summary
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    if a_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary: {g_row[0]} present", False, "Missing")
                all_ok = False
                continue

            # Large values need larger tolerance
            tol = 1000000 if "budget" in key or "spend" in key or "variance" in key.replace("_pct", "").replace("pct", "") else 1.0
            if "depts" in key or "over" in key or "under" in key:
                tol = 1.0

            if len(a_row) > 1 and len(g_row) > 1:
                ok = num_close(a_row[1], g_row[1], tol)
                record(f"Summary: {g_row[0]}", ok,
                       f"Expected {g_row[1]}, got {a_row[1]} (tol={tol})")
                if not ok:
                    all_ok = False

    return all_ok


def check_gsheet():
    """Check Google Sheet was created."""
    print("\n=== Checking Google Sheet ===")

    try:
        conn = psycopg2.connect(**{"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym",
                                   "user": "eigent", "password": "camel"})
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%budget%'")
        sheets = cur.fetchall()
        if not sheets:
            record("Google Sheet with 'Budget' in title", False, "Not found")
            cur.close()
            conn.close()
            return False
        record("Google Sheet with 'Budget' in title", True)

        sid = sheets[0][0]
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
        cell_count = cur.fetchone()[0]
        record("Google Sheet has data", cell_count >= 20, f"Cell count: {cell_count}")

        cur.close()
        conn.close()
        return cell_count >= 20
    except Exception as e:
        record("Google Sheet check", False, str(e))
        return False


def check_emails():
    """Check email was sent to CFO."""
    print("\n=== Checking Emails ===")

    try:
        conn = psycopg2.connect(**{"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym",
                                   "user": "eigent", "password": "camel"})
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        all_emails = cur.fetchall()
        cur.close()
        conn.close()

        record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

        found = False
        for subject, from_addr, to_addr, body in all_emails:
            subj_lower = (subject or "").lower()
            if "budget" in subj_lower:
                found = True
                to_str = str(to_addr or "").lower()
                record("Email sent to cfo@company.example.com",
                       "cfo@company.example.com" in to_str, f"To: {to_addr}")

                body_lower = (body or "").lower()
                record("Email body mentions budget",
                       "budget" in body_lower or "variance" in body_lower)
                break

        record("Budget email found", found)
        return found
    except Exception as e:
        record("Email check", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

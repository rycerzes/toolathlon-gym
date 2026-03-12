"""Evaluation for terminal-wc-excel-gsheet-gcal-email.
Checks:
1. Inventory_Lifecycle_Report.xlsx with 4 sheets
2. Google Sheet "Inventory Dashboard"
3. Calendar event for restock review
4. Email sent to purchasing team
5. demand_forecast.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: Inventory_Lifecycle_Report.xlsx ===")
    path = os.path.join(workspace, "Inventory_Lifecycle_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Product_Inventory
    pi_idx = next((i for i, s in enumerate(sheets_lower) if "product" in s and "inventory" in s), 0)
    ws1 = wb[sheets[pi_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Product_Inventory has 70+ product rows", len(data1) >= 70, f"Found {len(data1)}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has days_of_supply column", any("days" in h and "supply" in h for h in headers),
              f"Headers: {rows1[0]}")
        check("Has total_sales column", any("total" in h and "sales" in h for h in headers) or any("sales" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Reorder_Alerts
    ra_idx = next((i for i, s in enumerate(sheets_lower) if "reorder" in s or "alert" in s), 1)
    if ra_idx < len(sheets):
        ws2 = wb[sheets[ra_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Reorder_Alerts has entries", len(data2) >= 5, f"Found {len(data2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Has Critical urgency", "critical" in all_text2, f"Text: {all_text2[:100]}")
        check("Has Out_of_Stock urgency", "out" in all_text2 and "stock" in all_text2,
              f"Text: {all_text2[:100]}")

    # Category_Summary
    cs_idx = next((i for i, s in enumerate(sheets_lower) if "category" in s or "summary" in s), 2)
    if cs_idx < len(sheets):
        ws3 = wb[sheets[cs_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Category_Summary has category rows", len(data3) >= 3, f"Found {len(data3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Has Electronics category", "electronics" in all_text3)
        check("Has TV category", "tv" in all_text3)

    # Restock_Schedule
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "restock" in s or "schedule" in s), 3)
    if rs_idx < len(sheets):
        ws4 = wb[sheets[rs_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Restock_Schedule has entries", len(data4) >= 5, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has supplier info", "supplier" in all_text4 or "primary" in all_text4,
              f"Text: {all_text4[:100]}")


def check_gsheet():
    print("\n=== Check 2: Google Sheet Inventory Dashboard ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()
        dashboard = None
        for ss_id, title in spreadsheets:
            if title and "inventory" in title.lower():
                dashboard = (ss_id, title)
                break
        check("Inventory Dashboard spreadsheet exists", dashboard is not None,
              f"Spreadsheets: {[s[1] for s in spreadsheets]}")

        if dashboard:
            cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (dashboard[0],))
            cell_count = cur.fetchone()[0]
            check("Dashboard has data cells", cell_count >= 5, f"Found {cell_count} cells")
    except Exception as e:
        check("Gsheet check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: Calendar Restock Meeting ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT summary, description, start_datetime FROM gcal.events")
        events = cur.fetchall()
        restock_event = None
        for summary, desc, start in events:
            s = (str(summary) + " " + str(desc)).lower()
            if "restock" in s or "inventory" in s:
                restock_event = (summary, desc, start)
                break
        check("Restock review meeting exists", restock_event is not None,
              f"Events: {[e[0] for e in events]}")
        if restock_event:
            check("Meeting mentions restock or inventory",
                  "restock" in str(restock_event[0]).lower() or "inventory" in str(restock_event[0]).lower())
    except Exception as e:
        check("Gcal check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Check 4: Email to Purchasing Team ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%purchasing%%'
               OR subject ILIKE '%%inventory%%alert%%'
               OR subject ILIKE '%%restock%%'
               OR subject ILIKE '%%critical%%inventory%%'
        """)
        emails = cur.fetchall()
        if not emails:
            cur.execute("""
                SELECT id, subject, to_addr, body_text
                FROM email.drafts
                WHERE to_addr::text ILIKE '%%purchasing%%'
                   OR subject ILIKE '%%inventory%%'
                   OR subject ILIKE '%%restock%%'
            """)
            emails = cur.fetchall()
        check("Email about inventory alert sent", len(emails) >= 1, "No matching email found")
        if emails:
            subject = str(emails[0][1]).lower() if emails[0][1] else ""
            check("Email subject mentions inventory or restock",
                  "inventory" in subject or "restock" in subject or "critical" in subject,
                  f"Subject: {emails[0][1]}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: demand_forecast.py ===")
    path = os.path.join(workspace, "demand_forecast.py")
    check("demand_forecast.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Inventory_Lifecycle_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets beyond the 4 required
        expected_keywords = {"product", "inventory", "reorder", "alert", "category", "summary",
                             "restock", "schedule"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Stock quantities should not be negative
        sheets_lower = [s.lower().replace(" ", "_") for s in wb.sheetnames]
        pi_idx = next((i for i, s in enumerate(sheets_lower) if "product" in s and "inventory" in s), 0)
        ws = wb[wb.sheetnames[pi_idx]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        has_negative_stock = False
        for r in rows:
            if r and len(r) >= 3:
                try:
                    stock = float(r[2]) if r[2] is not None else None
                except (ValueError, TypeError):
                    stock = None
                if stock is not None and stock < 0:
                    has_negative_stock = True
                    break
        check("No negative stock quantities", not has_negative_stock,
              "Found negative stock value")

    # Calendar: no restock events before March 2026
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (summary ILIKE '%%restock%%' OR summary ILIKE '%%inventory%%')
              AND start_datetime < '2026-03-01'
        """)
        old_events = cur.fetchone()[0]
        check("No restock events before March 2026", old_events == 0,
              f"Found {old_events} old events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gsheet()
    check_gcal()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

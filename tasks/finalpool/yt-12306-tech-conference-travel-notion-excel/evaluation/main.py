"""
Evaluation for yt-12306-tech-conference-travel-notion-excel task.

Checks:
1. Tech_Conference_Plan.xlsx exists
2. Prep_Videos sheet has >= 4 rows with Title and View_Count columns
3. Travel_Details sheet has >= 2 rows with Train_No column
4. Travel_Details contains G235 and G236
5. Conference_Schedule sheet has >= 5 rows
6. Notion page exists with Conference or Tech in title
7. GCal has >= 3 new events between 2026-03-12 and 2026-03-15
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Tech_Conference_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Tech_Conference_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Conference_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Tech_Conference_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Prep_Videos sheet
    prep_sheet = None
    for name in wb.sheetnames:
        if "prep" in name.lower() or "video" in name.lower():
            prep_sheet = wb[name]
            break
    if prep_sheet is None:
        record("Prep_Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Prep_Videos sheet exists", True)
        rows = list(prep_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_title = any("title" in h for h in headers)
        has_viewcount = any("view" in h or "count" in h for h in headers)
        record("Prep_Videos has Title and View_Count columns", has_title and has_viewcount,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Prep_Videos has >= 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} data rows")

    # Check Travel_Details sheet
    travel_sheet = None
    for name in wb.sheetnames:
        if "travel" in name.lower():
            travel_sheet = wb[name]
            break
    if travel_sheet is None:
        record("Travel_Details sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Travel_Details sheet exists", True)
        rows = list(travel_sheet.iter_rows(values_only=True))
        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_train = any("train" in h or "no" in h for h in headers)
        record("Travel_Details has Train_No column", has_train, f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Details has >= 2 data rows", len(data_rows) >= 2,
               f"Found {len(data_rows)} rows")
        record("Travel_Details contains G235", "G235" in all_text, "G235 not found in Travel_Details")
        record("Travel_Details contains G236", "G236" in all_text, "G236 not found in Travel_Details")

    # Check Conference_Schedule sheet
    sched_sheet = None
    for name in wb.sheetnames:
        if "schedule" in name.lower() or "conference" in name.lower():
            sched_sheet = wb[name]
            break
    if sched_sheet is None:
        record("Conference_Schedule sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Conference_Schedule sheet exists", True)
        rows = list(sched_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Conference_Schedule has >= 5 data rows", len(data_rows) >= 5,
               f"Found {len(data_rows)} rows")

    # --- Groundtruth XLSX value comparison (order-insensitive, skips free-form cols) ---
    gt_path = os.path.join(groundtruth_workspace, "Tech_Conference_Plan.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            # Identify free-form columns to skip
            gt_headers = [str(c.value).strip().lower() if c.value else "" for c in next(gt_ws.iter_rows(max_row=1))]
            skip_cols = {i for i, h in enumerate(gt_headers)
                         if any(k in h for k in ["note", "reach", "remark", "comment", "description"])}
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            # Order-insensitive: each GT row's key cols must exist somewhere in agent rows
            def row_key(row):
                return tuple(
                    str(v).strip().lower() if isinstance(v, str) else v
                    for i, v in enumerate(row)
                    if v is not None and i not in skip_cols
                )
            for ri, gt_row in enumerate(gt_rows):
                gt_key = row_key(gt_row)
                found = any(row_key(ar) == gt_key for ar in a_rows)
                first_val = next((v for v in gt_row if v is not None), None)
                record(f"GT '{gt_sname}' entry '{first_val}' found (order-insensitive)", found,
                       f"gt_key={gt_key}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()
    cur.close()
    conn.close()

    found = False
    for page_id, props in pages:
        try:
            title_items = props.get("title", {}).get("title", {}).get("title", [])
            if not title_items:
                # Try alternate structure
                for key, val in props.items():
                    if isinstance(val, dict) and val.get("type") == "title":
                        title_items = val.get("title", [])
                        break
            title_text = " ".join(
                item.get("text", {}).get("content", "") for item in title_items
                if isinstance(item, dict)
            ).lower()
            if "conference" in title_text or "tech" in title_text or "qufu" in title_text:
                found = True
                break
        except Exception:
            continue

    record("Notion page exists with conference/tech content in title", found,
           f"Pages found: {len(pages)}")


def check_gcal():
    print("\n=== Check 3: GCal Conference Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-16'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    # Exclude the 2 preloaded events
    conf_events = [e for e in events
                   if "remote team" not in (e[0] or "").lower()
                   and "project deadline" not in (e[0] or "").lower()]

    record("GCal has >= 3 new conference events (Mar 12-15)", len(conf_events) >= 3,
           f"Found {len(conf_events)} non-preloaded events: {[e[0] for e in conf_events]}")

    summaries = " ".join(e[0] or "" for e in events).lower()
    has_travel = "travel" in summaries or "beijing" in summaries or "qufu" in summaries
    record("GCal contains travel or conference related events", has_travel,
           f"Event summaries: {[e[0] for e in events]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""Evaluation for sf-hr-education-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Education_Analysis.xlsx against groundtruth."""
    print("\n=== Checking Education_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Education_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Education_Analysis.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet: By Department ---
    agent_ws = get_sheet(agent_wb, "By Department")
    gt_ws = get_sheet(gt_wb, "By Department")

    if agent_ws is None:
        record("Sheet 'By Department' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'By Department' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("By Department row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for gt_row in gt_rows:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Department '{gt_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            # Total_Employees (col 1)
            ok_total = num_close(a_row[1], gt_row[1], 10)
            record(f"'{gt_row[0]}' Total_Employees", ok_total,
                   f"Expected {gt_row[1]}, got {a_row[1]}")
            if not ok_total:
                all_ok = False

            # Bachelors_Count (col 2)
            ok_bach = num_close(a_row[2], gt_row[2], 10)
            record(f"'{gt_row[0]}' Bachelors_Count", ok_bach,
                   f"Expected {gt_row[2]}, got {a_row[2]}")
            if not ok_bach:
                all_ok = False

            # Masters_Count (col 3)
            ok_mast = num_close(a_row[3], gt_row[3], 10)
            record(f"'{gt_row[0]}' Masters_Count", ok_mast,
                   f"Expected {gt_row[3]}, got {a_row[3]}")
            if not ok_mast:
                all_ok = False

            # PhD_Count (col 4)
            ok_phd = num_close(a_row[4], gt_row[4], 10)
            record(f"'{gt_row[0]}' PhD_Count", ok_phd,
                   f"Expected {gt_row[4]}, got {a_row[4]}")
            if not ok_phd:
                all_ok = False

            # Bachelors_Pct (col 5)
            ok_bpct = num_close(a_row[5], gt_row[5], 1.0)
            record(f"'{gt_row[0]}' Bachelors_Pct", ok_bpct,
                   f"Expected {gt_row[5]}, got {a_row[5]}")
            if not ok_bpct:
                all_ok = False

            # PhD_Pct (col 7)
            ok_ppct = num_close(a_row[7], gt_row[7], 0.5)
            record(f"'{gt_row[0]}' PhD_Pct", ok_ppct,
                   f"Expected {gt_row[7]}, got {a_row[7]}")
            if not ok_ppct:
                all_ok = False

    # --- Sheet: Summary ---
    agent_ws2 = get_sheet(agent_wb, "Summary")
    gt_ws2 = get_sheet(gt_wb, "Summary")

    if agent_ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)

        agent_summary = {}
        for row in agent_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_summary[str(row[0]).strip().lower()] = row[1]

        gt_summary = {}
        for row in gt_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_summary[str(row[0]).strip().lower()] = row[1]

        for metric, expected in gt_summary.items():
            actual = agent_summary.get(metric)
            if actual is None:
                record(f"Summary '{metric}' present", False, "Missing")
                all_ok = False
            else:
                if isinstance(expected, (int, float)):
                    ok = num_close(actual, expected, max(abs(expected) * 0.02, 50))
                else:
                    ok = str_match(actual, expected)
                record(f"Summary '{metric}'", ok,
                       f"Expected {expected}, got {actual}")
                if not ok:
                    all_ok = False

    return all_ok


def check_notion():
    """Check Notion page titled 'Workforce Education Analysis'."""
    print("\n=== Checking Notion Page ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()

    found_page = False
    page_id = None
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else json.loads(page[1]) if page[1] else {}
        title_text = ""
        if "title" in props:
            t = props["title"]
            if isinstance(t, dict) and "title" in t:
                for item in t["title"]:
                    if isinstance(item, dict) and "text" in item:
                        title_text += item["text"].get("content", "")
                    elif isinstance(item, dict) and "plain_text" in item:
                        title_text += item["plain_text"]
        if "workforce" in title_text.lower() and "education" in title_text.lower():
            found_page = True
            page_id = page[0]
            break

    record("Notion page 'Workforce Education Analysis' found", found_page,
           "No page with 'workforce' and 'education' in title")

    if found_page and page_id:
        # Check that the page has some blocks (content)
        cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s", (str(page_id),))
        block_count = cur.fetchone()[0]
        record("Notion page has content blocks", block_count >= 3,
               f"Found {block_count} blocks")

        # Check block content mentions departments
        cur.execute("SELECT block_data FROM notion.blocks WHERE parent_id = %s", (str(page_id),))
        blocks = cur.fetchall()
        all_text = ""
        for (block_data,) in blocks:
            if block_data:
                text = json.dumps(block_data) if isinstance(block_data, dict) else str(block_data)
                all_text += text.lower()

        record("Notion page mentions departments",
               "engineering" in all_text or "finance" in all_text or "hr" in all_text,
               "No department names found in page content")
        record("Notion page mentions PhD",
               "phd" in all_text,
               "PhD not mentioned in page content")

    cur.close()
    conn.close()
    return found_page


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)

    db_fail_before = FAIL_COUNT
    notion_ok = check_notion()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0 and excel_ok:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")
    overall = excel_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

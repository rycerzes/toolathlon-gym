"""Evaluation for canvas-learning-analytics."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_kpis():
    """Get expected KPIs from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    results = {}
    cur.execute("SELECT id, name FROM canvas.courses ORDER BY name")
    courses = cur.fetchall()

    for cid, cname in courses:
        # Completion rate
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE s.workflow_state = 'graded'),
                   COUNT(*)
            FROM canvas.submissions s
            JOIN canvas.assignments a ON s.assignment_id = a.id
            WHERE a.course_id = %s
        """, (cid,))
        graded, total = cur.fetchone()
        comp_rate = round(graded / total * 100, 1) if total > 0 else 0

        # Avg grade
        cur.execute("""
            SELECT ROUND(AVG((grades->>'current_score')::numeric), 1)
            FROM canvas.enrollments
            WHERE course_id = %s AND type = 'StudentEnrollment'
              AND grades->>'current_score' IS NOT NULL
        """, (cid,))
        avg_grade = float(cur.fetchone()[0] or 0)

        # Late rate
        cur.execute("""
            SELECT SUM(CASE WHEN s.late THEN 1 ELSE 0 END)::numeric / NULLIF(COUNT(s.id), 0) * 100
            FROM canvas.submissions s
            JOIN canvas.assignments a ON s.assignment_id = a.id
            WHERE a.course_id = %s
        """, (cid,))
        late_rate = round(float(cur.fetchone()[0] or 0), 1)

        meets = comp_rate >= 75 and avg_grade >= 70 and late_rate <= 15
        results[cname.lower()] = {
            "completion_rate": comp_rate,
            "avg_grade": avg_grade,
            "late_rate": late_rate,
            "meets": meets,
        }

    cur.close()
    conn.close()
    return results


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Learning_Analytics.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Learning_Analytics.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Learning_Analytics.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_kpis()
    meets_count = sum(1 for v in expected.values() if v["meets"])

    # Course KPIs sheet
    kpi_rows = load_sheet_rows(wb, "Course KPIs")
    if kpi_rows is None:
        check("Sheet 'Course KPIs' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Course KPIs' exists", True)
        data_rows = kpi_rows[1:] if len(kpi_rows) > 1 else []
        check("Course KPIs has 22 rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        header = kpi_rows[0] if kpi_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "completion_rate", "avg_grade", "late_rate", "meets_benchmark"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        # Spot-check a known course
        for row in data_rows:
            if row and row[0] and "foundations of finance (fall 2013)" in str(row[0]).lower():
                exp = expected.get("foundations of finance (fall 2013)", {})
                check("FFF-2013J avg_grade close to expected",
                      num_close(row[2], exp.get("avg_grade", 0), 3),
                      f"Got {row[2]}, expected {exp.get('avg_grade')}")
                break

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        check("Total_Courses = 22", num_close(lookup.get("total_courses"), 22),
              f"Got {lookup.get('total_courses')}")
        check(f"Meets_All_Benchmarks close to {meets_count}",
              num_close(lookup.get("meets_all_benchmarks"), meets_count, 3),
              f"Got {lookup.get('meets_all_benchmarks')}, expected {meets_count}")

    # Trends sheet
    tr_rows = load_sheet_rows(wb, "Trends")
    if tr_rows is None:
        check("Sheet 'Trends' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Trends' exists", True)
        data_rows = tr_rows[1:] if len(tr_rows) > 1 else []
        # Should have 22 courses * 3 categories = 66 rows
        check("Trends has ~66 rows", abs(len(data_rows) - 66) <= 5,
              f"Found {len(data_rows)}")


def check_notion():
    print("\n=== Checking Notion Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()
        found_page = None
        for page_id, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "learning analytics" in props_str or "analytics dashboard" in props_str:
                found_page = page_id
                break
        check("Notion page for analytics dashboard exists", found_page is not None,
              f"Found {len(pages)} pages")
        if found_page:
            cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s", (found_page,))
            block_count = cur.fetchone()[0]
            check("Notion page has content (blocks)", block_count >= 1,
                  f"Found {block_count} blocks")
        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

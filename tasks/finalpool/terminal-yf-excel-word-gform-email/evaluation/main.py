"""Evaluation for terminal-yf-excel-word-gform-email.

Checks:
1. Quarterly_Portfolio_Review.xlsx with 3 sheets
2. Portfolio_Review_Report.docx
3. Google Form "Q1 2025 Investor Feedback Survey"
4. Email sent to investment_committee@company.com
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Quarterly_Portfolio_Review.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Quarterly_Portfolio_Review.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quarterly_Portfolio_Review.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Get dynamic stock data from DB
    yf_prices = {}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT symbol, close_price FROM yf.stock_prices
            WHERE (symbol, date) IN (
                SELECT symbol, MAX(date) FROM yf.stock_prices GROUP BY symbol
            )
        """)
        for sym, price in cur.fetchall():
            yf_prices[sym.strip().upper()] = float(price)
        cur.close()
        conn.close()
    except Exception:
        pass

    # Performance_Summary
    print("  Checking Performance_Summary...")
    a_sheet = get_sheet(agent_wb, "Performance_Summary")
    g_sheet = get_sheet(gt_wb, "Performance_Summary")
    check("Sheet 'Performance_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
        check("Performance_Summary has 5 rows", len(a_rows) == 5, f"Got {len(a_rows)}")

        a_lookup = {str(r[0]).strip().upper(): r for r in a_rows if r and r[0]}
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Symbol '{key}' present", False, "Missing")
                continue
            # Use dynamic price if available, else fall back to groundtruth
            if len(a_row) > 3:
                expected_price = yf_prices.get(key, g_row[3] if len(g_row) > 3 else None)
                if expected_price is not None:
                    check(f"'{key}' Current_Price",
                          num_close(a_row[3], expected_price, 5.0),
                          f"Expected {expected_price}, got {a_row[3]}")
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"'{key}' Total_Return_Pct",
                      num_close(a_row[4], g_row[4], 5.0),
                      f"Expected {g_row[4]}, got {a_row[4]}")
            if len(a_row) > 5 and len(g_row) > 5:
                check(f"'{key}' Annualized_Volatility_Pct",
                      num_close(a_row[5], g_row[5], 5.0),
                      f"Expected {g_row[5]}, got {a_row[5]}")

    # Risk_Metrics
    print("  Checking Risk_Metrics...")
    a_sheet = get_sheet(agent_wb, "Risk_Metrics")
    g_sheet = get_sheet(gt_wb, "Risk_Metrics")
    check("Sheet 'Risk_Metrics' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        check("Risk_Metrics has 5 rows", len(a_rows) == 5, f"Got {len(a_rows)}")

        a_lookup = {str(r[0]).strip().upper(): r for r in a_rows if r and r[0]}
        for g_row in g_sheet.iter_rows(min_row=2, values_only=True):
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Risk '{key}' present", False, "Missing")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"'{key}' Sharpe_Ratio",
                      num_close(a_row[1], g_row[1], 0.3),
                      f"Expected {g_row[1]}, got {a_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"'{key}' Max_Drawdown_Pct",
                      num_close(a_row[2], g_row[2], 5.0),
                      f"Expected {g_row[2]}, got {a_row[2]}")

    # Portfolio_Summary
    print("  Checking Portfolio_Summary...")
    a_sheet = get_sheet(agent_wb, "Portfolio_Summary")
    g_sheet = get_sheet(gt_wb, "Portfolio_Summary")
    check("Sheet 'Portfolio_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_data = {}
        for row in a_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_data[str(row[0]).strip().lower()] = row[1]
        g_data = {}
        for row in g_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                g_data[str(row[0]).strip().lower()] = row[1]

        check("Total_Holdings = 5",
              num_close(a_data.get("total_holdings"), 5, 0),
              f"Got {a_data.get('total_holdings')}")
        bp = a_data.get("best_performer")
        check("Best_Performer is GOOGL",
              bp is not None and "GOOGL" in str(bp).upper(),
              f"Got {bp}")
        wp = a_data.get("worst_performer")
        check("Worst_Performer is AMZN",
              wp is not None and "AMZN" in str(wp).upper(),
              f"Got {wp}")
        check("Avg_Total_Return_Pct",
              num_close(a_data.get("avg_total_return_pct"),
                        g_data.get("avg_total_return_pct"), 5.0),
              f"Expected {g_data.get('avg_total_return_pct')}, got {a_data.get('avg_total_return_pct')}")


def check_word(agent_workspace):
    print("\n=== Checking Portfolio_Review_Report.docx ===")
    docx_path = os.path.join(agent_workspace, "Portfolio_Review_Report.docx")
    check("Portfolio_Review_Report.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 300, f"Length: {len(text)}")
        check("Contains portfolio/performance reference",
              "portfolio" in text and "performance" in text)
        check("Contains GOOGL reference",
              "googl" in text or "alphabet" in text)
        check("Contains risk/volatility reference",
              "risk" in text or "volatility" in text or "sharpe" in text)
        check("Contains recommendation",
              "recommend" in text or "rebalanc" in text or "suggest" in text)
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE lower(title) LIKE '%%investor%%feedback%%'
               OR lower(title) LIKE '%%q1%%survey%%'
               OR lower(title) LIKE '%%portfolio%%'
        """)
        forms = cur.fetchall()
        check("Investor feedback form exists", len(forms) > 0,
              f"Found {len(forms)} matching forms")
        if forms:
            form_id = forms[0][0]
            cur.execute("""
                SELECT COUNT(*) FROM gform.questions WHERE form_id = %s
            """, (form_id,))
            q_count = cur.fetchone()[0]
            check("Form has >= 4 questions", q_count >= 4, f"Got {q_count}")
        cur.close()
        conn.close()
    except Exception as e:
        check("GForm check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Check via sent_log join or direct messages
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text
            FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE lower(m.subject) LIKE '%%portfolio%%performance%%'
               OR lower(m.subject) LIKE '%%q1%%portfolio%%'
        """)
        rows = cur.fetchall()
        if not rows:
            cur.execute("""
                SELECT subject, to_addr, body_text FROM email.messages
                WHERE lower(subject) LIKE '%%portfolio%%performance%%'
                   OR lower(subject) LIKE '%%q1%%portfolio%%'
            """)
            rows = cur.fetchall()
        check("Portfolio review email sent", len(rows) > 0, f"Found {len(rows)}")
        if rows:
            to_str = str(rows[0][1]).lower() if rows[0][1] else ""
            check("Email to investment_committee",
                  "investment_committee" in to_str, f"To: {rows[0][1]}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Quarterly_Portfolio_Review.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"performance", "summary", "risk", "metric", "portfolio"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Sharpe ratios should be reasonable (between -5 and 5)
        risk_sheet = get_sheet(wb, "Risk_Metrics")
        if risk_sheet:
            for row in risk_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1] is not None:
                    try:
                        sharpe = float(row[1])
                        if abs(sharpe) > 5:
                            check("No unreasonable Sharpe ratios", False,
                                  f"Found {sharpe} for {row[0]}")
                            break
                    except (ValueError, TypeError):
                        pass
            else:
                check("No unreasonable Sharpe ratios", True)

    # Email: no emails to wrong recipients
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE (lower(subject) LIKE '%%portfolio%%' OR lower(subject) LIKE '%%q1%%')
              AND to_addr::text NOT ILIKE '%%investment%%'
              AND to_addr::text NOT ILIKE '%%committee%%'
        """)
        wrong = cur.fetchone()[0]
        check("No portfolio emails to wrong recipients", wrong == 0,
              f"Found {wrong} misrouted emails")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gform()
    check_email()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

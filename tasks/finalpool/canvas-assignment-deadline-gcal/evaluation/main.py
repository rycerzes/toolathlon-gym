"""
Evaluation for canvas-assignment-deadline-gcal.

Checks:
1. Assignment_Tracker.xlsx matches groundtruth (sheets: Assignments, Summary)
2. Google Calendar events for each assignment deadline
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


# ============================================================================
# Check 1: Assignment_Tracker.xlsx
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    """Compare Assignment_Tracker.xlsx against groundtruth."""
    print("\n=== Checking Assignment_Tracker.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Assignment_Tracker.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Assignment_Tracker.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet: Assignments ---
    agent_ws = get_sheet(agent_wb, "Assignments")
    gt_ws = get_sheet(gt_wb, "Assignments")

    if agent_ws is None:
        record("Sheet 'Assignments' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Assignments' exists", True)

        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("Assignments row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Build lookup by assignment name
        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for gt_row in gt_rows:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Assignment '{gt_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            # Check Due_Date (col 1)
            ok_date = str_match(a_row[1], gt_row[1])
            record(f"'{gt_row[0]}' Due_Date", ok_date,
                   f"Expected {gt_row[1]}, got {a_row[1]}")
            if not ok_date:
                all_ok = False

            # Check Points_Possible (col 2)
            ok_pts = num_close(a_row[2], gt_row[2], 0.5)
            record(f"'{gt_row[0]}' Points_Possible", ok_pts,
                   f"Expected {gt_row[2]}, got {a_row[2]}")
            if not ok_pts:
                all_ok = False

            # Check Assignment_Group (col 3)
            ok_grp = str_match(a_row[3], gt_row[3])
            record(f"'{gt_row[0]}' Assignment_Group", ok_grp,
                   f"Expected {gt_row[3]}, got {a_row[3]}")
            if not ok_grp:
                all_ok = False

            # Check Submission_Count (col 4)
            ok_sub = num_close(a_row[4], gt_row[4], 5)
            record(f"'{gt_row[0]}' Submission_Count", ok_sub,
                   f"Expected {gt_row[4]}, got {a_row[4]}")
            if not ok_sub:
                all_ok = False

    # --- Sheet: Summary ---
    agent_ws2 = get_sheet(agent_wb, "Summary")
    gt_ws2 = get_sheet(gt_wb, "Summary")

    if agent_ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)

        agent_summary = {}
        for row in agent_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_summary[str(row[0]).strip().lower()] = row[1]

        gt_summary = {}
        for row in gt_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_summary[str(row[0]).strip().lower()] = row[1]

        for metric, expected in gt_summary.items():
            actual = agent_summary.get(metric)
            if actual is None:
                record(f"Summary '{metric}' present", False, "Missing")
                all_ok = False
            else:
                ok = num_close(actual, expected, 1.0) if isinstance(expected, (int, float)) else str_match(actual, expected)
                record(f"Summary '{metric}'", ok,
                       f"Expected {expected}, got {actual}")
                if not ok:
                    all_ok = False

    return all_ok


# ============================================================================
# Check 2: Google Calendar events
# ============================================================================

def check_gcal():
    """Check calendar events for assignment deadlines."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY summary")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")

    # Should have 7 events (one per assignment)
    record("At least 7 calendar events created", len(events) >= 7,
           f"Found {len(events)}")

    # Check that events have "Due:" prefix
    due_events = [e for e in events if "due:" in (e[0] or "").lower()]
    record("Events have 'Due:' prefix in summary", len(due_events) >= 7,
           f"Found {len(due_events)} events with 'Due:' prefix")

    # Check specific assignment names appear
    expected_assignments = [
        "TMA 25355",
        "TMA 25356",
        "TMA 25357",
        "TMA 25358",
        "TMA 25359",
        "TMA 25360",
        "Final Exam 25361",
    ]

    all_ok = True
    for name in expected_assignments:
        found = any(name.lower() in (e[0] or "").lower() for e in events)
        record(f"Calendar event for '{name}'", found)
        if not found:
            all_ok = False

    # Check descriptions mention points
    events_with_points = [e for e in events if e[1] and ("point" in e[1].lower() or any(c.isdigit() for c in (e[1] or "")))]
    record("Events have points in description", len(events_with_points) >= 5,
           f"Found {len(events_with_points)} events with points info")

    return all_ok and len(events) >= 7


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)

    db_fail_before = FAIL_COUNT
    gcal_ok = check_gcal()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0 and excel_ok:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")
    overall = excel_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

"""Evaluation for wc-shipping-rate-analysis."""
import argparse
import json
import os
import sys

import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_data):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Audit.xlsx")
    if not os.path.exists(path):
        return ["Shipping_Audit.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Rate Comparison sheet
        rows = load_sheet_rows(wb, "Rate Comparison")
        if rows is None:
            errors.append("Sheet 'Rate Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_orders"]
            if abs(len(data_rows) - expected) > 5:
                errors.append(f"Rate Comparison has {len(data_rows)} rows, expected ~{expected}")

        # Check Undercharged Orders sheet
        rows2 = load_sheet_rows(wb, "Undercharged Orders")
        if rows2 is None:
            errors.append("Sheet 'Undercharged Orders' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            expected_uc = gt_data["undercharged_count"]
            if abs(len(data_rows2) - expected_uc) > 5:
                errors.append(f"Undercharged Orders has {len(data_rows2)} rows, expected ~{expected_uc}")

        # Check Summary sheet
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            summary_dict = {}
            for r in rows3[1:]:
                if r and r[0] is not None:
                    summary_dict[str(r[0]).strip().lower()] = r[1]

            # Check undercharged count
            for sk, sv in summary_dict.items():
                if "undercharged" in sk and "count" in sk:
                    try:
                        val = int(float(sv))
                        if abs(val - gt_data["undercharged_count"]) > 3:
                            errors.append(f"Undercharged count = {val}, expected ~{gt_data['undercharged_count']}")
                    except (ValueError, TypeError):
                        pass
                    break

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_email(gt_data):
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%logistics@company.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            errors.append("No email found sent to logistics@company.com")
        else:
            body = (rows[0][2] or "").lower()
            if "undercharg" not in body and "audit" not in body.lower():
                errors.append("Email body does not mention undercharged orders or audit")

    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email(gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

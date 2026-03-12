"""Evaluation for wc-product-pricing-excel-gform task.

Checks:
1. Excel file Pricing_Audit.xlsx with correct data (both sheets)
2. Google Form titled "Pricing Review Feedback" with 4 questions
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

EXPECTED_CATEGORIES = ["Audio", "TV & Home Theater", "Speakers", "Electronics", "Cameras", "Headphones"]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel File ===")
    agent_file = os.path.join(agent_workspace, "Pricing_Audit.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Pricing_Audit.xlsx")

    if not os.path.exists(agent_file):
        check("Excel file exists", False, f"Not found: {agent_file}")
        return
    check("Excel file exists", True)

    if not os.path.exists(gt_file):
        check("Groundtruth file exists", False, f"Not found: {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Discounted Products sheet
    print("  Checking Discounted Products sheet...")
    a_rows = load_sheet_rows(agent_wb, "Discounted Products")
    g_rows = load_sheet_rows(gt_wb, "Discounted Products")

    if a_rows is None:
        check("Sheet 'Discounted Products' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth sheet exists", False, "Not found")
    else:
        check("Sheet 'Discounted Products' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Discounted Products row count matches",
              len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        # Build lookup by product name (first 50 chars lowercase)
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                key = str(row[0]).strip().lower()[:50]
                a_lookup[key] = row

        match_count = 0
        mismatch_count = 0
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()[:50]
            a_row = a_lookup.get(key)
            if a_row is None:
                mismatch_count += 1
                continue

            # Check Regular_Price (col 2), Sale_Price (col 3), Discount_Pct (col 4)
            all_ok = True
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    all_ok = False
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.5):
                    all_ok = False
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1.0):
                    all_ok = False

            if all_ok:
                match_count += 1
            else:
                mismatch_count += 1

        total_expected = len([r for r in g_data if r and r[0] is not None])
        check(f"Discounted Products data accuracy ({match_count}/{total_expected})",
              match_count >= total_expected * 0.8,
              f"Matched {match_count}, mismatched {mismatch_count}")

        # Check sort order (Discount_Pct descending)
        if len(a_data) >= 2:
            sorted_ok = True
            for i in range(min(len(a_data) - 1, 10)):
                if a_data[i] and a_data[i+1] and len(a_data[i]) > 4 and len(a_data[i+1]) > 4:
                    try:
                        if float(a_data[i][4]) < float(a_data[i+1][4]) - 0.5:
                            sorted_ok = False
                            break
                    except (TypeError, ValueError):
                        pass
            check("Discounted Products sorted by Discount_Pct descending", sorted_ok)

    # Check Category Summary sheet
    print("  Checking Category Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Category Summary")
    g_rows = load_sheet_rows(gt_wb, "Category Summary")

    if a_rows is None:
        check("Sheet 'Category Summary' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth Category Summary sheet exists", False, "Not found")
    else:
        check("Sheet 'Category Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Category Summary row count matches",
              len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Category '{g_row[0]}'", False, "Missing")
                continue

            # Products_On_Sale (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"{key}.Products_On_Sale",
                      num_close(a_row[1], g_row[1], 2),
                      f"{a_row[1]} vs {g_row[1]}")

            # Avg_Discount_Pct (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Avg_Discount_Pct",
                      num_close(a_row[2], g_row[2], 2.0),
                      f"{a_row[2]} vs {g_row[2]}")

            # Max_Discount_Pct (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Max_Discount_Pct",
                      num_close(a_row[3], g_row[3], 1.0),
                      f"{a_row[3]} vs {g_row[3]}")


def check_gform():
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    if len(forms) == 0:
        check("Google Form created", False, "No forms found")
        cur.close()
        conn.close()
        return

    # Find form with matching title
    form_id = None
    form_title = None
    for fid, ftitle in forms:
        if ftitle and "pricing" in ftitle.lower() and "review" in ftitle.lower():
            form_id = fid
            form_title = ftitle
            break

    if form_id is None:
        # Try partial match
        for fid, ftitle in forms:
            if ftitle and "pricing" in ftitle.lower():
                form_id = fid
                form_title = ftitle
                break

    if form_id is None:
        form_id = forms[0][0]
        form_title = forms[0][1]

    check("Form title contains 'Pricing Review Feedback'",
          form_title and "pricing" in form_title.lower() and "feedback" in form_title.lower(),
          f"Got: {form_title}")

    # Check questions
    cur.execute(
        "SELECT title, question_type, required, position FROM gform.questions WHERE form_id=%s ORDER BY position",
        (form_id,))
    questions = cur.fetchall()

    check("Form has 4 questions", len(questions) == 4,
          f"Found {len(questions)} questions")

    if len(questions) >= 1:
        q = questions[0]
        check("Q1: reviewer name (text, required)",
              q[0] and "name" in q[0].lower() and q[2] is True,
              f"title='{q[0]}', type='{q[1]}', required={q[2]}")

    if len(questions) >= 2:
        q = questions[1]
        check("Q2: product category (multiple choice, required)",
              q[0] and "category" in q[0].lower() and q[2] is True,
              f"title='{q[0]}', type='{q[1]}', required={q[2]}")

    if len(questions) >= 3:
        q = questions[2]
        check("Q3: pricing assessment (multiple choice, required)",
              q[0] and ("pricing" in q[0].lower() or "assessment" in q[0].lower()) and q[2] is True,
              f"title='{q[0]}', type='{q[1]}', required={q[2]}")

    if len(questions) >= 4:
        q = questions[3]
        check("Q4: comments (text, not required)",
              q[0] and "comment" in q[0].lower() and q[2] is not True,
              f"title='{q[0]}', type='{q[1]}', required={q[2]}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    file_fails = FAIL_COUNT

    check_gform()
    db_failures = FAIL_COUNT - file_fails

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0 and file_fails == 0:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")
    all_passed = file_fails == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": all_passed}, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

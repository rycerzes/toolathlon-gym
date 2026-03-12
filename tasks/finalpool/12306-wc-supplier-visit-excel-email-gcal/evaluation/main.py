"""
Evaluation for 12306-wc-supplier-visit-excel-email-gcal task.

Checks:
1. Supplier_Visit_Plan.xlsx exists with Products, Travel_Plan, Visit_Schedule sheets
2. Products sheet has >= 4 rows with Product_ID and Supplier_Name columns
3. Travel_Plan has G11 and G105 train entries
4. Visit_Schedule has >= 3 rows
5. GCal has >= 2 supplier visit events on 2026-03-10
6. At least 2 emails sent (to shanghai_supplier or gz_supplier or procurement)
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Supplier_Visit_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Supplier_Visit_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Supplier_Visit_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Supplier_Visit_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_products = any("product" in s for s in sheet_names_lower)
    has_travel = any("travel" in s for s in sheet_names_lower)
    has_schedule = any("schedule" in s or "visit" in s for s in sheet_names_lower)

    record("Excel has Products sheet", has_products, f"Sheets: {wb.sheetnames}")
    record("Excel has Travel_Plan sheet", has_travel, f"Sheets: {wb.sheetnames}")
    record("Excel has Visit_Schedule sheet", has_schedule, f"Sheets: {wb.sheetnames}")

    if has_products:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "product" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Products sheet has >= 4 rows", len(data_rows) >= 4, f"Found {len(data_rows)} rows")

        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_id = any("id" in h or "product" in h for h in headers)
            has_supplier = any("supplier" in h for h in headers)
            record("Products has Product_ID column", has_id, f"Headers: {rows[0]}")
            record("Products has Supplier_Name column", has_supplier, f"Headers: {rows[0]}")

    if has_travel:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "travel" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Plan has >= 2 rows", len(data_rows) >= 2, f"Found {len(data_rows)} rows")

        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Travel_Plan contains G11 (Shanghai route)", "g11" in all_text,
               f"Content: {all_text[:200]}")
        record("Travel_Plan contains G105 (Guangzhou route)", "g105" in all_text,
               f"Content: {all_text[:200]}")

    if has_schedule:
        idx = next(i for i, s in enumerate(sheet_names_lower) if "schedule" in s or "visit" in s)
        ws_name = wb.sheetnames[idx]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Visit_Schedule has >= 3 rows", len(data_rows) >= 3, f"Found {len(data_rows)} rows")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Supplier_Visit_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        record(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices.append(len(gt_rows) - 1)
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        record(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1}",
                               False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    record(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                record(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: GCal supplier visit events on 2026-03-10 ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime::date = '2026-03-10'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    record("GCal has >= 2 events on 2026-03-10", len(events) >= 2,
           f"Found {len(events)} events")

    summaries = [str(e[0]).lower() for e in events]
    has_supplier = any("supplier" in s or "visit" in s or "meeting" in s or "shanghai" in s or "guangzhou" in s
                       for s in summaries)
    record("GCal has supplier/visit/meeting events", has_supplier,
           f"Summaries: {summaries}")


def check_emails():
    print("\n=== Check 3: Emails sent ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    def to_addresses(to_addr):
        if isinstance(to_addr, list):
            return " ".join(str(r).lower() for r in to_addr)
        elif to_addr:
            try:
                parsed = json.loads(str(to_addr))
                return " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                return str(to_addr).lower()
        return ""

    # Count outgoing messages (not the injected supplier emails)
    outgoing_addrs = ["shanghai_supplier@techworld.com", "gz_supplier@supplier.com", "procurement@company.com"]
    outgoing = [m for m in messages if any(addr in to_addresses(m[2]) for addr in outgoing_addrs)]

    record("At least 1 email sent to supplier or procurement", len(outgoing) >= 1,
           f"Total messages: {len(messages)}, matching: {len(outgoing)}")
    record("At least 2 emails sent total (supplier visits + summary)", len(messages) >= 2,
           f"Found {len(messages)} total messages")

    to_shanghai = [m for m in messages if "shanghai_supplier@techworld.com" in to_addresses(m[2])]
    record("Email sent to shanghai_supplier@techworld.com", len(to_shanghai) >= 1,
           f"Total messages: {len(messages)}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

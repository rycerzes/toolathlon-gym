"""Evaluation for canvas-exam-prep-scheduler."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_quiz_data():
    """Get expected quiz performance from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name, q.title, ROUND(AVG(qs.score), 1) as avg_score
        FROM canvas.quizzes q
        JOIN canvas.courses c ON c.id = q.course_id
        LEFT JOIN canvas.quiz_submissions qs ON qs.quiz_id = q.id
        WHERE q.points_possible = 100
        GROUP BY c.name, q.title
        ORDER BY c.name, q.title
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Exam_Prep.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Exam_Prep.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Exam_Prep.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_quiz_data()
    below_80 = [r for r in expected if float(r[2]) < 80]
    courses_needing = len(set(r[0] for r in below_80))

    # Quiz Performance sheet
    qp_rows = load_sheet_rows(wb, "Quiz Performance")
    if qp_rows is None:
        check("Sheet 'Quiz Performance' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Quiz Performance' exists", True)
        data_rows = qp_rows[1:] if len(qp_rows) > 1 else []
        check(f"Quiz Performance has {len(expected)} rows (100-pt quizzes)",
              abs(len(data_rows) - len(expected)) <= 2,
              f"Found {len(data_rows)}, expected {len(expected)}")

        # Check header
        header = qp_rows[0] if qp_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "quiz", "avg_score", "below_threshold"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        # Verify a known below-threshold quiz
        if below_80:
            sample = below_80[0]
            found = False
            for row in data_rows:
                if row and row[0] and row[1]:
                    if sample[1].lower() in str(row[1]).lower():
                        found = True
                        check(f"Quiz {sample[1]} avg ~{sample[2]}",
                              num_close(row[2], float(sample[2])),
                              f"Got {row[2]}")
                        check(f"Quiz {sample[1]} marked as below threshold",
                              str(row[3]).strip().lower() in ("yes", "true", "y"),
                              f"Got {row[3]}")
                        break
            check(f"Known below-threshold quiz found", found)

    # Review Schedule sheet
    rs_rows = load_sheet_rows(wb, "Review Schedule")
    if rs_rows is None:
        check("Sheet 'Review Schedule' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Review Schedule' exists", True)
        data_rows = rs_rows[1:] if len(rs_rows) > 1 else []
        check(f"Review Schedule has ~{courses_needing} sessions",
              abs(len(data_rows) - courses_needing) <= 2,
              f"Found {len(data_rows)}, expected ~{courses_needing}")

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        check(f"Total_Quizzes_Analyzed = {len(expected)}",
              num_close(lookup.get("total_quizzes_analyzed"), len(expected)),
              f"Got {lookup.get('total_quizzes_analyzed')}")
        check(f"Below_Threshold_Quizzes = {len(below_80)}",
              num_close(lookup.get("below_threshold_quizzes"), len(below_80)),
              f"Got {lookup.get('below_threshold_quizzes')}")


def check_calendar():
    print("\n=== Checking Calendar Events ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime
            FROM gcal.events
            WHERE summary ILIKE '%%review%%' OR summary ILIKE '%%exam%%' OR summary ILIKE '%%quiz%%'
        """)
        events = cur.fetchall()
        check("Calendar events created for review sessions", len(events) >= 1,
              f"Found {len(events)} events")
        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%%review%%' OR subject ILIKE '%%exam%%'
               OR subject ILIKE '%%scheduled%%'
        """)
        emails = cur.fetchall()
        check("Email sent about review sessions", len(emails) >= 1,
              "No matching email found")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_calendar()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

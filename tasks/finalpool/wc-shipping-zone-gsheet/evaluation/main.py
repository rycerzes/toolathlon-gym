"""Evaluation for wc-shipping-zone-gsheet."""
import argparse
import os
import sys
import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Shipping_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Shipping_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check sheet: Zone Summary
    print("  Checking Zone Summary...")
    a_rows = load_sheet_rows(agent_wb, "Zone Summary")
    g_rows = load_sheet_rows(gt_wb, "Zone Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Zone Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Zone Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing zone: {g_row[0]}")
                continue

            # Order_Count
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    all_errors.append(f"{key}.Order_Count: {a_row[1]} vs {g_row[1]}")

            # Total_Revenue
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 500.0):
                    all_errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]} (tol=500)")

            # Avg_Order_Value
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 10.0):
                    all_errors.append(f"{key}.Avg_Order_Value: {a_row[3]} vs {g_row[3]} (tol=10)")

        if not [e for e in all_errors if "Zone Summary" in e or "Missing zone" in e or "Order_Count" in e or "Revenue" in e or "Avg" in e]:
            print("    PASS")
        else:
            print(f"    ERRORS found")

    # Check sheet: Methods
    print("  Checking Methods...")
    a_rows = load_sheet_rows(agent_wb, "Methods")
    g_rows = load_sheet_rows(gt_wb, "Methods")
    if a_rows is None:
        all_errors.append("Sheet 'Methods' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Methods' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        if len(a_data) < len(g_data):
            all_errors.append(f"Methods: expected {len(g_data)} rows, got {len(a_data)}")
        else:
            print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

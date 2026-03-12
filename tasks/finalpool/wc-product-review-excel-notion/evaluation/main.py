"""Evaluation for wc-product-review-excel-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Expected category data from actual DB
EXPECTED_CATEGORIES = {
    "audio": {"product_count": 15, "review_count": 78, "avg_rating": 4.47},
    "cameras": {"product_count": 10, "review_count": 49, "avg_rating": 4.69},
    "electronics": {"product_count": 30, "review_count": 149, "avg_rating": 4.57},
    "headphones": {"product_count": 10, "review_count": 50, "avg_rating": 4.52},
    "home appliances": {"product_count": 8, "review_count": 29, "avg_rating": 4.72},
    "speakers": {"product_count": 5, "review_count": 28, "avg_rating": 4.39},
    "tv & home theater": {"product_count": 13, "review_count": 63, "avg_rating": 4.78},
    "watches": {"product_count": 6, "review_count": 28, "avg_rating": 4.43},
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.1):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Product_Review_Analysis.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Product_Review_Analysis.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Product_Review_Analysis.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # Check Category Analysis sheet
    cat_rows = load_sheet_rows(wb, "Category Analysis")
    if cat_rows is None:
        check("Sheet 'Category Analysis' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Category Analysis' exists", True)
        data_rows = cat_rows[1:] if len(cat_rows) > 1 else []
        check("Category Analysis has 8 data rows (8 categories)",
              len(data_rows) == 8,
              f"Found {len(data_rows)} rows")

        # Check header columns
        header = cat_rows[0] if cat_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col_name in ["category", "product_count", "review_count", "avg_rating", "five_star_count", "five_star_rate"]:
            check(f"Column '{col_name}' present",
                  any(col_name.replace("_", "") in h.replace("_", "") or col_name in h for h in header_lower),
                  f"Header: {header}")

        # Verify avg_rating for Electronics (4.57)
        for row in data_rows:
            if row and row[0] and str(row[0]).strip().lower() == "electronics":
                avg_rating = row[3] if len(row) > 3 else None
                check("Electronics avg_rating is ~4.57",
                      num_close(avg_rating, 4.57, 0.05),
                      f"Got {avg_rating}")
                review_count = row[2] if len(row) > 2 else None
                check("Electronics review_count = 149",
                      num_close(review_count, 149, 1),
                      f"Got {review_count}")
                break

        # Verify TV & Home Theater avg_rating is highest (4.78)
        found_tv = False
        for row in data_rows:
            if row and row[0] and "tv" in str(row[0]).lower():
                found_tv = True
                avg_rating = row[3] if len(row) > 3 else None
                check("TV & Home Theater avg_rating is ~4.78",
                      num_close(avg_rating, 4.78, 0.05),
                      f"Got {avg_rating}")
                break
        check("TV & Home Theater category row found", found_tv)

    # Check Top Products sheet
    top_rows = load_sheet_rows(wb, "Top Products")
    if top_rows is None:
        check("Sheet 'Top Products' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Top Products' exists", True)
        data_rows = top_rows[1:] if len(top_rows) > 1 else []
        check("Top Products has up to 10 rows",
              1 <= len(data_rows) <= 10,
              f"Found {len(data_rows)} rows")

        # Check that top product has avg_rating = 5.0
        if data_rows:
            top_avg = data_rows[0][3] if len(data_rows[0]) > 3 else None
            check("Top product avg_rating is 5.0",
                  num_close(top_avg, 5.0, 0.05),
                  f"Got {top_avg}")


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # notion.databases.title is jsonb
        cur.execute("""
            SELECT id, title FROM notion.databases
            WHERE archived = false
        """)
        dbs = cur.fetchall()
        found_db = None
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "product review" in title_str or "review insight" in title_str:
                found_db = db_id
                break
        check("Notion database 'Product Review Insights' exists",
              found_db is not None,
              f"Found {len(dbs)} databases, none matching 'Product Review'")

        if found_db:
            # Check that pages exist (parent is the database)
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent::text LIKE %s AND archived = false AND in_trash = false
            """, (f'%{found_db}%',))
            page_count_result = cur.fetchone()
            page_count = page_count_result[0] if page_count_result else 0
            check("Notion database has entries (at least 3 categories)",
                  page_count >= 3,
                  f"Found {page_count} entries")

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%product_team@store.com%'
               OR subject ILIKE '%product review%'
               OR subject ILIKE '%product%analysis%'
        """)
        emails = cur.fetchall()
        check("Email sent to product_team@store.com", len(emails) >= 1,
              "No matching email found")
        if emails:
            email = emails[0]
            subject = str(email[1]).lower() if email[1] else ""
            check("Email subject contains 'review' or 'product'",
                  "review" in subject or "product" in subject or "analysis" in subject,
                  f"Subject: {email[1]}")
            body = str(email[3]) if email[3] else ""
            check("Email body has content", len(body) > 20,
                  f"Body length: {len(body)}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

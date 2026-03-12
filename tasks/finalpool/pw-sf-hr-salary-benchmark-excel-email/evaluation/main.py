"""Evaluation script for pw-sf-hr-salary-benchmark-excel-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Salary_Benchmark_Report.xlsx")
    check("Salary_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Salary_Benchmark_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        check("Compensation_Comparison sheet exists", "Compensation_Comparison" in wb.sheetnames)
        if "Compensation_Comparison" in wb.sheetnames:
            ws = wb["Compensation_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Compensation_Comparison has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Employee_Count', 'Our_Avg_Salary', 'Industry_Benchmark', 'Difference']:
                check(f"Compensation_Comparison has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Department_Details sheet exists", "Department_Details" in wb.sheetnames)
        if "Department_Details" in wb.sheetnames:
            ws = wb["Department_Details"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Department_Details has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Avg_Experience', 'Avg_Performance']:
                check(f"Department_Details has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Executive_Summary sheet exists", "Executive_Summary" in wb.sheetnames)
        if "Executive_Summary" in wb.sheetnames:
            ws = wb["Executive_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Executive_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Executive_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # Check email was sent
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT subject, to_addr, body_text FROM email.messages WHERE subject ILIKE %s", ('%benchmark%',))
            emails = cur.fetchall()
            check("Benchmark email sent", len(emails) >= 1, f"found {len(emails)} matching emails")
            if emails:
                check("Email to hr-director", "hr-director@company.com" in str(emails[0][1]).lower())
                check("Email mentions departments", "department" in str(emails[0][2]).lower() if emails[0][2] else False)
            conn.close()
        except Exception as e:
            check("Email verification", False, str(e))

        # Check terminal artifacts
        check("salary_processor.py exists", os.path.exists(os.path.join(agent_workspace, "salary_processor.py")))
        check("benchmark_raw.json exists", os.path.exists(os.path.join(agent_workspace, "benchmark_raw.json")))
        check("salary_comparison.json exists", os.path.exists(os.path.join(agent_workspace, "salary_comparison.json")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

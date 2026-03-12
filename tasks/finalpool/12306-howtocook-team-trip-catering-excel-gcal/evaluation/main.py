"""
Evaluation for 12306-howtocook-team-trip-catering-excel-gcal task.

Checks:
1. Team_Trip_Plan.xlsx exists with Travel, Menu, Timeline sheets
2. Travel sheet has G11 and 07:00 departure
3. Menu sheet has >= 5 rows with Course_Type and Dish_Name columns
4. Timeline sheet has >= 5 rows
5. Notion page exists with team/trip/march in title
6. GCal has >= 2 new events on 2026-03-10
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Team_Trip_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Team_Trip_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Team_Trip_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Team_Trip_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_travel = any("travel" in s for s in sheet_names_lower)
    has_menu = any("menu" in s for s in sheet_names_lower)
    has_timeline = any("timeline" in s for s in sheet_names_lower)

    record("Excel has Travel sheet", has_travel, f"Sheets: {wb.sheetnames}")
    record("Excel has Menu sheet", has_menu, f"Sheets: {wb.sheetnames}")
    record("Excel has Timeline sheet", has_timeline, f"Sheets: {wb.sheetnames}")

    if has_travel:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "travel" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel sheet has >= 1 data row", len(data_rows) >= 1, f"Found {len(data_rows)} rows")

        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Travel sheet contains G11", "g11" in all_text, f"Content sample: {all_text[:200]}")
        has_time = "07:00" in all_text or "07:00:00" in all_text or "7:00" in all_text
        record("Travel sheet shows 07:00 departure", has_time, f"Content: {all_text[:200]}")

    if has_menu:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "menu" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Menu sheet has >= 5 rows", len(data_rows) >= 5, f"Found {len(data_rows)} rows")

        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_course = any("course" in h or "type" in h or "category" in h for h in headers)
            has_dish = any("dish" in h or "name" in h for h in headers)
            record("Menu has course type column", has_course, f"Headers: {rows[0]}")
            record("Menu has dish name column", has_dish, f"Headers: {rows[0]}")

    if has_timeline:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "timeline" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Timeline sheet has >= 5 rows", len(data_rows) >= 5, f"Found {len(data_rows)} rows")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Team_Trip_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        record(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices.append(len(gt_rows) - 1)
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        record(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1}",
                               False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    record(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                record(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion page for team trip ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()
    cur.close()
    conn.close()

    trip_page = None
    for page_id, props in pages:
        title = ""
        if isinstance(props, dict):
            title_obj = props.get("title", {})
            if isinstance(title_obj, dict):
                title_list = title_obj.get("title", [])
                if isinstance(title_list, list):
                    for t in title_list:
                        if isinstance(t, dict):
                            title += t.get("text", {}).get("content", "")
        title_lower = title.lower()
        if any(kw in title_lower for kw in ["team", "trip", "march", "shanghai", "建设", "出行"]):
            trip_page = (page_id, title)
            break

    record("Notion page exists with trip/team/march context",
           trip_page is not None,
           f"Pages found: {len(pages)}, titles: {[str(p[1])[:80] for p in pages[:3]]}")


def check_gcal():
    print("\n=== Check 3: GCal events on 2026-03-10 ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM gcal.events
        WHERE start_datetime::date = '2026-03-10'
    """)
    total_count = cur.fetchone()[0]

    # Preprocess injects 2 events, agent should add at least 2 more (net >= 4, or if preprocess cleared first, >= 2)
    # We check that at least 2 events exist total and at least some are agent-added
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime::date = '2026-03-10'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    record("GCal has >= 2 events on 2026-03-10", total_count >= 2,
           f"Found {total_count} events")

    # Check that at least one agent-created event exists (departure or dinner or check-in)
    event_summaries = [str(e[0]).lower() for e in events]
    has_departure = any("depart" in s or "train" in s or "station" in s or "trip" in s for s in event_summaries)
    has_dinner = any("dinner" in s or "meal" in s or "banquet" in s for s in event_summaries)
    has_arrival = any("arriv" in s or "check" in s or "hotel" in s for s in event_summaries)

    agent_events = has_departure or has_dinner or has_arrival
    record("GCal has agent-created trip events (departure/dinner/arrival)",
           agent_events,
           f"Summaries: {event_summaries}")


def check_email():
    print("\n=== Check 4: Email sent to events@company.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    sent = [m for m in messages if m[1] and "company.com" not in str(m[1])]
    # Check for outgoing emails (not the preprocess-injected incoming email)
    outgoing = []
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif to_addr:
            try:
                parsed = json.loads(str(to_addr))
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "events@company.com" in to_str:
            outgoing.append((subject, from_addr, to_addr, body_text))

    record("Email sent to events@company.com", len(outgoing) >= 1,
           f"Total messages: {len(messages)}, matching: {len(outgoing)}")

    if outgoing:
        subject, _, _, body = outgoing[0]
        body_lower = ((subject or "") + " " + (body or "")).lower()
        has_trip = any(kw in body_lower for kw in ["trip", "travel", "train", "shanghai", "g11", "excel", "plan"])
        record("Email mentions trip/travel/plan content", has_trip,
               f"Subject: {subject}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

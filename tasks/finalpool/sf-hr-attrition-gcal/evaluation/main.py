"""Evaluation for sf-hr-attrition-gcal."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")
PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1; print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1; print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except: return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute('''SELECT "DEPARTMENT", COUNT(*),
        ROUND(AVG("JOB_SATISFACTION")::numeric,2),
        ROUND(AVG("WORK_LIFE_BALANCE")::numeric,2),
        ROUND(AVG("PERFORMANCE_RATING")::numeric,2)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT" ORDER BY AVG("JOB_SATISFACTION")''')
    depts = [{"dept": r[0], "count": r[1], "avg_sat": float(r[2]),
              "avg_wlb": float(r[3]), "avg_perf": float(r[4])} for r in cur.fetchall()]
    conn.close()
    return {"all_depts": depts, "low_sat": depts[:3]}


def sheet_dicts(wb, name):
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2: return []
            hdrs = [str(h).strip() if h else "" for h in rows[0]]
            return [{hdrs[i]: row[i] for i in range(len(hdrs))} for row in rows[1:] if not all(v is None for v in row)]
    return None


def check_excel(ws_path, exp):
    print("\n=== Checking Excel ===")
    p = os.path.join(ws_path, "Satisfaction_Report.xlsx")
    if not os.path.isfile(p):
        record("Excel file exists", False, p); return
    record("Excel file exists", True)
    wb = openpyxl.load_workbook(p, data_only=True)

    d = sheet_dicts(wb, "Department Scores")
    if d is None:
        record("Sheet Department Scores", False, str(wb.sheetnames))
    else:
        record("Sheet Department Scores", True)
        record("Department count", len(d) == 7, f"Got {len(d)}")
        for e in exp["all_depts"]:
            m = next((r for r in d if str_match(r.get("Department"), e["dept"])), None)
            if not m: record(f"Dept {e['dept']}", False, "Missing"); continue
            record(f"Dept {e['dept']} satisfaction",
                   num_close(m.get("Avg_Satisfaction"), e["avg_sat"], 0.5),
                   f"{m.get('Avg_Satisfaction')} vs {e['avg_sat']}")
            record(f"Dept {e['dept']} count",
                   num_close(m.get("Employee_Count"), e["count"], 100),
                   f"{m.get('Employee_Count')} vs {e['count']}")

    d = sheet_dicts(wb, "Action Items")
    if d is None:
        record("Sheet Action Items", False, str(wb.sheetnames))
    else:
        record("Sheet Action Items", True)
        record("Action Items has 3 rows", len(d) == 3, f"Got {len(d)}")
        for e in exp["low_sat"]:
            m = next((r for r in d if str_match(r.get("Department"), e["dept"])), None)
            if not m: record(f"Action {e['dept']}", False, "Missing"); continue
            record(f"Action {e['dept']} status",
                   str_match(m.get("Status"), "Needs Review"),
                   f"Got: {m.get('Status')}")
    wb.close()


def check_gcal(exp):
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    record("At least 3 calendar events", len(events) >= 3, f"Found {len(events)}")

    for dept_info in exp["low_sat"]:
        dept = dept_info["dept"]
        found = False
        for summary, desc, start_dt in events:
            if dept.lower() in (summary or "").lower() and "hr review" in (summary or "").lower():
                found = True
                record(f"Event for {dept} exists", True)
                if start_dt:
                    record(f"Event {dept} date is 2026-03-16",
                           start_dt.strftime("%Y-%m-%d") == "2026-03-16",
                           f"Got {start_dt.strftime('%Y-%m-%d')}")
                desc_lower = (desc or "").lower()
                sat_str = str(dept_info["avg_sat"])
                record(f"Event {dept} mentions satisfaction",
                       sat_str in (desc or "") or dept.lower() in desc_lower,
                       f"Desc: {(desc or '')[:150]}")
                break
        if not found:
            record(f"Event for {dept}", False, "Not found")

    cur.close()
    conn.close()


def check_email(exp):
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
    emails = cur.fetchall()
    record("At least 1 email sent", len(emails) >= 1, f"Found {len(emails)}")

    found_summary = False
    for subj, to, body in emails:
        subj_lower = (subj or "").lower()
        if "satisfaction" in subj_lower or "review" in subj_lower:
            found_summary = True
            to_str = json.dumps(to).lower() if isinstance(to, list) else str(to).lower()
            record("Email to hr-director",
                   "hr-director@company.example.com" in to_str, f"To: {to}")
            body_lower = (body or "").lower()
            for dept_info in exp["low_sat"]:
                record(f"Email mentions {dept_info['dept']}",
                       dept_info["dept"].lower() in body_lower,
                       "Not found in body")
            break
    if not found_summary:
        record("Summary email found", False, f"Subjects: {[e[0] for e in emails]}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", default=".")
    parser.add_argument("--groundtruth_workspace", default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    exp = get_expected()
    check_excel(args.agent_workspace, exp)
    check_gcal(exp)
    check_email(exp)
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": FAIL_COUNT == 0}, f)
    sys.exit(0 if FAIL_COUNT == 0 else 1)

if __name__ == "__main__":
    main()

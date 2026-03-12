"""Evaluation for terminal-yf-fetch-gsheet-excel-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Market_Analysis_Report.xlsx ===")
    agent_file = os.path.join(agent_ws, "Market_Analysis_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Market_Analysis_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Stock_Overview
    print("  Checking Stock_Overview...")
    ws1 = get_sheet(awb, "Stock_Overview")
    check("Sheet Stock_Overview exists", ws1 is not None, f"Sheets: {awb.sheetnames}")
    if ws1:
        rows = list(ws1.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in rows if r and r[0]]
        check("Stock_Overview has 3 rows", len(data_rows) == 3, f"Got {len(data_rows)}")

        symbols = {str(r[0]).strip().upper() for r in data_rows if r and r[0]}
        check("Contains AMZN", "AMZN" in symbols, f"Symbols: {symbols}")
        check("Contains GOOGL", "GOOGL" in symbols, f"Symbols: {symbols}")
        check("Contains JPM", "JPM" in symbols, f"Symbols: {symbols}")

        # Check each has price and return
        for r in data_rows:
            if r and len(r) >= 5 and r[0]:
                sym = str(r[0]).strip().upper()
                check(f"{sym} has price > 0", r[3] is not None and float(r[3]) > 0,
                      f"Price: {r[3]}")
                check(f"{sym} has return value", r[4] is not None,
                      f"Return: {r[4]}")

    # Sheet 2: Economic_Indicators
    print("  Checking Economic_Indicators...")
    ws2 = get_sheet(awb, "Economic_Indicators")
    check("Sheet Economic_Indicators exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2:
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        data_rows2 = [r for r in rows2 if r and r[0]]
        check("Economic_Indicators has 5 rows", len(data_rows2) == 5, f"Got {len(data_rows2)}")

        # Check specific indicator values from mock data
        ind_lookup = {}
        for r in data_rows2:
            if r and r[0]:
                ind_lookup[str(r[0]).strip().lower()] = r

        gdp = ind_lookup.get("gdp growth rate") or ind_lookup.get("gdp growth")
        if gdp:
            check("GDP Growth = 2.1", num_close(gdp[1], 2.1, 0.2), f"Got {gdp[1]}")
            check("GDP Trend is Favorable",
                  gdp[2] and "favorable" in str(gdp[2]).lower(),
                  f"Got {gdp[2]}")

        infl = ind_lookup.get("inflation rate") or ind_lookup.get("inflation")
        if infl:
            check("Inflation = 3.4", num_close(infl[1], 3.4, 0.2), f"Got {infl[1]}")
            check("Inflation Trend is Unfavorable",
                  infl[2] and "unfavorable" in str(infl[2]).lower(),
                  f"Got {infl[2]}")

    # Sheet 3: Correlation_Matrix
    print("  Checking Correlation_Matrix...")
    ws3 = get_sheet(awb, "Correlation_Matrix")
    check("Sheet Correlation_Matrix exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        data_rows3 = [r for r in rows3 if r and r[0]]
        check("Correlation_Matrix has 3 rows", len(data_rows3) == 3, f"Got {len(data_rows3)}")

        valid_values = {"positive", "negative", "neutral"}
        for r in data_rows3:
            if r and len(r) >= 4:
                for col_idx in [1, 2, 3]:
                    val = str(r[col_idx]).strip().lower() if r[col_idx] else ""
                    if val not in valid_values:
                        check(f"{r[0]} correlation values valid", False, f"Invalid: {r[col_idx]}")
                        break
                else:
                    continue
                break
        else:
            check("All correlation values are valid", True)

    # Sheet 4: Portfolio_Signals
    print("  Checking Portfolio_Signals...")
    ws4 = get_sheet(awb, "Portfolio_Signals")
    check("Sheet Portfolio_Signals exists", ws4 is not None, f"Sheets: {awb.sheetnames}")
    if ws4:
        rows4 = list(ws4.iter_rows(min_row=2, values_only=True))
        data_rows4 = [r for r in rows4 if r and r[0]]
        check("Portfolio_Signals has 3 rows", len(data_rows4) == 3, f"Got {len(data_rows4)}")

        valid_signals = {"buy", "sell", "hold"}
        for r in data_rows4:
            if r and len(r) >= 3:
                sig = str(r[1]).strip().lower() if r[1] else ""
                if sig not in valid_signals:
                    check(f"{r[0]} signal valid", False, f"Invalid: {r[1]}")
                    break
                if not r[2]:
                    check(f"{r[0]} has rationale", False, "Missing rationale")
                    break
        else:
            check("All signals valid with rationale", True)


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%market analysis%'")
        count = cur.fetchone()[0]
        check("Google Sheet 'Market Analysis Live Data' exists", count >= 1, f"Found {count}")
        cur.close()
        conn.close()
    except Exception as e:
        check("GSheet check", False, str(e))


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        found_db = None
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "investment" in title_str and ("research" in title_str or "log" in title_str):
                found_db = db_id
                break
        check("Notion 'Investment Research Log' exists",
              found_db is not None, f"Found {len(dbs)} dbs")

        if found_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent::text LIKE %s AND archived = false AND in_trash = false
            """, (f'%{found_db}%',))
            page_count = cur.fetchone()[0]
            check("Notion DB has 3 stock pages", page_count == 3,
                  f"Got {page_count}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Market_Analysis_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"stock", "overview", "economic", "indicator", "correlation",
                             "matrix", "portfolio", "signal"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Stock prices should not be negative
        ws1 = get_sheet(wb, "Stock_Overview")
        if ws1:
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4 and row[3] is not None:
                    try:
                        price = float(row[3])
                        if price < 0:
                            check("No negative stock prices", False,
                                  f"Found {price} for {row[0]}")
                            break
                    except (ValueError, TypeError):
                        pass
            else:
                check("No negative stock prices", True)

    # Notion: no duplicate Investment Research Log databases
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        invest_dbs = []
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "investment" in title_str and "research" in title_str:
                invest_dbs.append(db_id)
        check("No duplicate Investment Research databases", len(invest_dbs) <= 1,
              f"Found {len(invest_dbs)} matching databases")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gsheet()
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

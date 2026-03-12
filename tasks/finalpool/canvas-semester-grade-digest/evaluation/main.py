"""Evaluation for canvas-semester-grade-digest."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Grade_Digest.xlsx")
    gt_file = os.path.join(gt_dir, "Grade_Digest.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # --- Check Course Grades sheet ---
    print("  Checking Course Grades sheet...")
    a_rows = load_sheet_rows(agent_wb, "Course Grades")
    g_rows = load_sheet_rows(gt_wb, "Course Grades")
    if a_rows is None:
        all_errors.append("Sheet 'Course Grades' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Course Grades' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Build lookup by partial course name match (case-insensitive)
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)

            # Try partial match if exact match fails
            if a_row is None:
                for akey, aval in a_lookup.items():
                    # Match on key words from the course name
                    g_words = set(key.split())
                    a_words = set(akey.split())
                    if len(g_words & a_words) >= 3:
                        a_row = aval
                        break

            if a_row is None:
                all_errors.append(f"Missing course: {g_row[0]}")
                continue

            # Enrolled_Students (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 20):
                    all_errors.append(f"{key[:40]}.Enrolled: {a_row[1]} vs {g_row[1]}")

            # Avg_Score (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 2.0):
                    all_errors.append(f"{key[:40]}.Avg_Score: {a_row[2]} vs {g_row[2]}")

            # Min_Score (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    all_errors.append(f"{key[:40]}.Min_Score: {a_row[3]} vs {g_row[3]}")

            # Max_Score (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 5.0):
                    all_errors.append(f"{key[:40]}.Max_Score: {a_row[4]} vs {g_row[4]}")

        if not [e for e in all_errors if "Course Grades" in e or "Missing course" in e]:
            print("    PASS")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Summary: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            try:
                float(a_val); float(g_val)
                if not num_close(a_val, g_val, 2.0):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val} (tol=2.0)")
            except (TypeError, ValueError):
                # For string comparisons (course names), check if key part matches
                a_str = str(a_val or "").strip().lower()
                g_str = str(g_val or "").strip().lower()
                if a_str != g_str:
                    # Allow partial match for course names
                    a_words = set(a_str.split())
                    g_words = set(g_str.split())
                    if len(a_words & g_words) < 2:
                        all_errors.append(f"Summary.{key}: '{a_val}' vs '{g_val}'")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check Word document exists ---
    print("  Checking Word document...")
    word_file = os.path.join(args.agent_workspace, "Grade_Summary.docx")
    if not os.path.exists(word_file):
        all_errors.append("Grade_Summary.docx not found")
    else:
        try:
            from docx import Document
            doc = Document(word_file)
            text = " ".join(p.text for p in doc.paragraphs).lower()
            has_content = len(text) > 50
            if not has_content:
                all_errors.append("Grade_Summary.docx has too little content")
            else:
                print("    PASS")
        except ImportError:
            # If python-docx not installed, just check file size
            if os.path.getsize(word_file) > 100:
                print("    PASS (file exists, size OK)")
            else:
                all_errors.append("Grade_Summary.docx is too small")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

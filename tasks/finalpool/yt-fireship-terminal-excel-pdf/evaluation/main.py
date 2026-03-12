"""
Evaluation for yt-fireship-terminal-excel-pdf task.

Checks:
1. fireship_raw.csv exists with at least 100 rows and required columns
2. quarterly_stats.json exists with correct structure (at least 4 quarters)
3. Fireship_Analytics.xlsx exists with Raw_Data, Quarterly_Stats, and Summary sheets
4. Raw_Data sheet has at least 100 rows
5. Quarterly_Stats sheet has at least 4 quarters sorted chronologically
6. Summary sheet has Total_Videos, Best_Quarter, Worst_Quarter entries
7. Fireship_Analytics_Report.docx exists with required sections
"""
import csv
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_csv(agent_workspace):
    print("\n=== Check 1: fireship_raw.csv ===")
    csv_path = os.path.join(agent_workspace, "fireship_raw.csv")
    if not os.path.exists(csv_path):
        record("fireship_raw.csv exists", False, f"Not found at {csv_path}")
        return
    record("fireship_raw.csv exists", True)

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames or []

    required_cols = ["video_id", "title", "published_date", "duration_sec", "view_count", "like_count"]
    headers_lower = [h.lower().strip() for h in headers]
    has_required = all(col in headers_lower for col in required_cols)
    record("CSV has required columns", has_required, f"Found: {headers}")

    record("CSV has at least 100 rows", len(rows) >= 100, f"Found {len(rows)} rows")

    # Check for some known video titles
    titles = [r.get("title", "") for r in rows]
    all_titles = " ".join(titles).lower()
    has_deepseek = "deepseek" in all_titles
    record("CSV contains Fireship video titles (DeepSeek)", has_deepseek,
           "No DeepSeek title found")


def check_json(agent_workspace):
    print("\n=== Check 2: quarterly_stats.json ===")
    json_path = os.path.join(agent_workspace, "quarterly_stats.json")
    if not os.path.exists(json_path):
        record("quarterly_stats.json exists", False, f"Not found at {json_path}")
        return
    record("quarterly_stats.json exists", True)

    try:
        with open(json_path) as f:
            data = json.load(f)
    except Exception as e:
        record("quarterly_stats.json is valid JSON", False, str(e))
        return
    record("quarterly_stats.json is valid JSON", True)

    record("Has at least 4 quarters", len(data) >= 4, f"Found {len(data)} entries")

    if data:
        first = data[0]
        has_keys = all(k in first for k in ["Quarter", "Video_Count", "Total_Views", "Avg_Engagement_Rate"])
        record("Each entry has Quarter, Video_Count, Total_Views, Avg_Engagement_Rate",
               has_keys, f"Keys found: {list(first.keys())}")

        # Check sorting
        quarters = [d.get("Quarter", "") for d in data]
        record("Quarters are sorted chronologically", quarters == sorted(quarters),
               f"Quarters: {quarters}")

        # Check 2025-Q1 is best (should have highest views ~37M)
        q1_2025 = next((d for d in data if d.get("Quarter") == "2025-Q1"), None)
        if q1_2025:
            record("2025-Q1 has over 20M total views", q1_2025.get("Total_Views", 0) >= 20000000,
                   f"Found {q1_2025.get('Total_Views')}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 3: Fireship_Analytics.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Fireship_Analytics.xlsx")
    if not os.path.exists(xlsx_path):
        record("Fireship_Analytics.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Fireship_Analytics.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Raw_Data sheet", "raw_data" in sheet_names_lower, f"Sheets: {wb.sheetnames}")
    record("Has Quarterly_Stats sheet", "quarterly_stats" in sheet_names_lower, f"Sheets: {wb.sheetnames}")
    record("Has Summary sheet", "summary" in sheet_names_lower, f"Sheets: {wb.sheetnames}")

    # Check Raw_Data
    if "raw_data" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("raw_data")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Raw_Data has at least 100 videos", len(data_rows) >= 100,
               f"Found {len(data_rows)} data rows")

    # Check Quarterly_Stats
    if "quarterly_stats" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("quarterly_stats")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Quarterly_Stats has at least 4 quarters", len(data_rows) >= 4,
               f"Found {len(data_rows)} rows")

        # Check headers
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            record("Quarterly_Stats has Quarter and Total_Views columns",
                   any("quarter" in h for h in headers) and any("view" in h for h in headers),
                   f"Headers: {rows[0]}")

    # Check Summary
    if "summary" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c).lower()
        record("Summary has Total_Videos entry", "total_videos" in all_text, "No Total_Videos found")
        record("Summary has Best_Quarter entry", "best_quarter" in all_text, "No Best_Quarter found")
        record("Summary has Worst_Quarter entry", "worst_quarter" in all_text, "No Worst_Quarter found")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Fireship_Analytics.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_word(agent_workspace):
    print("\n=== Check 4: Fireship_Analytics_Report.docx ===")
    docx_path = os.path.join(agent_workspace, "Fireship_Analytics_Report.docx")
    if not os.path.exists(docx_path):
        record("Fireship_Analytics_Report.docx exists", False, f"Not found at {docx_path}")
        return
    record("Fireship_Analytics_Report.docx exists", True)

    try:
        import docx
        doc = docx.Document(docx_path)
        full_text = " ".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        record("Word document readable", False, str(e))
        return
    record("Word document readable", True)

    record("Contains Executive Summary section", "executive summary" in full_text,
           "No 'Executive Summary' found")
    record("Contains Quarterly Performance section", "quarterly performance" in full_text,
           "No 'Quarterly Performance' found")
    record("Contains Key Observations section", "key observations" in full_text,
           "No 'Key Observations' found")
    record("Contains Appendix section", "appendix" in full_text, "No 'Appendix' found")
    record("Mentions Fireship channel", "fireship" in full_text, "No 'Fireship' mention found")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_csv(args.agent_workspace)
    check_json(args.agent_workspace)
    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

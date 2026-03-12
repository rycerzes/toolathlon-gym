"""Evaluation for sf-workforce-compensation-sheet."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Compensation_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # --- Check Department Compensation sheet ---
    print("  Checking Department Compensation sheet...")
    a_rows = load_sheet_rows(agent_wb, "Department Compensation")
    g_rows = load_sheet_rows(gt_wb, "Department Compensation")
    if a_rows is None:
        all_errors.append("Sheet 'Department Compensation' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Department Compensation' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing department: {g_row[0]}")
                continue

            # Headcount (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    all_errors.append(f"{key}.Headcount: {a_row[1]} vs {g_row[1]}")

            # Avg_Salary (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 100.0):
                    all_errors.append(f"{key}.Avg_Salary: {a_row[2]} vs {g_row[2]}")

            # Min_Salary (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 500.0):
                    all_errors.append(f"{key}.Min_Salary: {a_row[3]} vs {g_row[3]}")

            # Max_Salary (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 5000.0):
                    all_errors.append(f"{key}.Max_Salary: {a_row[4]} vs {g_row[4]}")

            # Total_Payroll (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 100000.0):
                    all_errors.append(f"{key}.Total_Payroll: {a_row[5]} vs {g_row[5]}")

        if not [e for e in all_errors if "Department" in e or "Missing" in e]:
            print("    PASS")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Summary: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            try:
                float(a_val); float(g_val)
                if not num_close(a_val, g_val, 100000.0):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val}")
            except (TypeError, ValueError):
                if not str_match(a_val, g_val):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val}")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check Google Sheet ---
    print("  Checking Google Sheet...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%compensation%' OR LOWER(title) LIKE '%workforce%'")
        sheets = cur.fetchall()
        if not sheets:
            all_errors.append("No Google Sheet with 'Compensation' or 'Workforce' in title found")
        else:
            sid = sheets[0][0]
            cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
            cell_count = cur.fetchone()[0]
            if cell_count < 20:
                all_errors.append(f"Google Sheet has too few cells: {cell_count}")
            else:
                print("    PASS")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Google Sheet check error: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

"""Evaluation for terminal-canvas-howtocook-gsheet-notion-email.

Checks:
1. Nutrition_Academic_Study.xlsx with 2 sheets (Student_Engagement, Meal_Plans)
2. student_engagement.json
3. meal_recommendations.json
4. nutrition_study_summary.txt
5. Google Sheet "Nutrition Study Data" with 2 sheets
6. Notion database "Wellness Pilot Program" with 3 entries
7. Two emails sent (student_affairs, dining_services)
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=5432,
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user="postgres",
    password="postgres",
)

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_groundtruth_tiers():
    """Compute expected tier values from Canvas DB."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT
                CASE
                    WHEN late_rate = 0 THEN 'High'
                    WHEN late_rate <= 0.25 THEN 'Medium'
                    ELSE 'Low'
                END as tier,
                COUNT(*) as student_count,
                ROUND(AVG(avg_score)::numeric, 1) as tier_avg_score,
                ROUND(AVG(late_rate)::numeric, 3) as avg_late_rate
            FROM (
                SELECT s.user_id,
                    SUM(CASE WHEN s.late THEN 1 ELSE 0 END)::float / COUNT(*) as late_rate,
                    AVG(s.score) as avg_score
                FROM canvas.submissions s
                JOIN canvas.assignments a ON s.assignment_id = a.id
                WHERE a.course_id IN (13, 14) AND s.score IS NOT NULL
                GROUP BY s.user_id
            ) sub
            GROUP BY tier
            ORDER BY tier
        """)
        tiers = {}
        for row in cur.fetchall():
            tiers[row[0]] = {
                "student_count": int(row[1]),
                "avg_score": float(row[2]),
                "late_rate": float(row[3]),
            }
        cur.close()
        conn.close()
        return tiers
    except Exception:
        return {
            "High": {"student_count": 1152, "avg_score": 82.9, "late_rate": 0.0},
            "Medium": {"student_count": 252, "avg_score": 80.5, "late_rate": 0.25},
            "Low": {"student_count": 339, "avg_score": 72.4, "late_rate": 0.62},
        }


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Nutrition_Academic_Study.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Nutrition_Academic_Study.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Nutrition_Academic_Study.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    gt_tiers = get_groundtruth_tiers()

    # Sheet 1: Student_Engagement
    print("  -- Student_Engagement sheet --")
    ws = None
    for s in wb.sheetnames:
        if "engagement" in s.lower():
            ws = wb[s]
            break
    check("Student_Engagement sheet exists", ws is not None, f"Sheets: {wb.sheetnames}")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Student_Engagement has 3 rows", len(rows) == 3, f"Got {len(rows)}")
        tier_lookup = {}
        for r in rows:
            if r and r[0]:
                tier_lookup[str(r[0]).strip()] = r
        for tier_name, expected in gt_tiers.items():
            r = tier_lookup.get(tier_name)
            if r is None:
                check(f"Tier '{tier_name}' present", False, "Missing")
                continue
            check(
                f"'{tier_name}' Student_Count",
                num_close(r[1], expected["student_count"], 5),
                f"Expected {expected['student_count']}, got {r[1]}",
            )
            check(
                f"'{tier_name}' Avg_Score",
                num_close(r[2], expected["avg_score"], 2.0),
                f"Expected {expected['avg_score']}, got {r[2]}",
            )

    # Sheet 2: Meal_Plans
    print("  -- Meal_Plans sheet --")
    ws2 = None
    for s in wb.sheetnames:
        if "meal" in s.lower() or "plan" in s.lower():
            ws2 = wb[s]
            break
    check("Meal_Plans sheet exists", ws2 is not None, f"Sheets: {wb.sheetnames}")
    if ws2:
        rows = list(ws2.iter_rows(min_row=2, values_only=True))
        check("Meal_Plans has 3 rows", len(rows) == 3, f"Got {len(rows)}")
        tier_lookup = {}
        for r in rows:
            if r and r[0]:
                tier_lookup[str(r[0]).strip()] = r
        for tier_name, expected_type in [
            ("High", "Control"),
            ("Medium", "Partial"),
            ("Low", "Full Meal Plan"),
        ]:
            r = tier_lookup.get(tier_name)
            if r is None:
                check(f"Meal tier '{tier_name}' present", False, "Missing")
                continue
            check(
                f"'{tier_name}' Intervention_Type",
                str(r[1]).strip().lower() == expected_type.lower(),
                f"Expected '{expected_type}', got '{r[1]}'",
            )


def check_json_files(agent_workspace):
    print("\n=== Checking JSON files ===")
    gt_tiers = get_groundtruth_tiers()

    # student_engagement.json
    se_path = os.path.join(agent_workspace, "student_engagement.json")
    check("student_engagement.json exists", os.path.isfile(se_path))
    if os.path.isfile(se_path):
        try:
            with open(se_path) as f:
                data = json.load(f)
            check("student_engagement.json is valid JSON", True)
            if isinstance(data, list):
                check("Has 3 tier entries", len(data) == 3, f"Got {len(data)}")
                for item in data:
                    tier = item.get("tier", "")
                    if tier in gt_tiers:
                        check(
                            f"JSON '{tier}' student_count",
                            num_close(item.get("student_count", 0), gt_tiers[tier]["student_count"], 5),
                            f"Expected {gt_tiers[tier]['student_count']}, got {item.get('student_count')}",
                        )
            else:
                check("student_engagement.json is array", False, f"Got {type(data).__name__}")
        except Exception as e:
            check("student_engagement.json parseable", False, str(e))

    # meal_recommendations.json
    mr_path = os.path.join(agent_workspace, "meal_recommendations.json")
    check("meal_recommendations.json exists", os.path.isfile(mr_path))
    if os.path.isfile(mr_path):
        try:
            with open(mr_path) as f:
                data = json.load(f)
            check("meal_recommendations.json is valid JSON", True)
            if isinstance(data, list):
                check("Has 3 meal plan entries", len(data) == 3, f"Got {len(data)}")
            else:
                check("meal_recommendations.json is array", False, f"Got {type(data).__name__}")
        except Exception as e:
            check("meal_recommendations.json parseable", False, str(e))

    # nutrition_study_summary.txt
    summary_path = os.path.join(agent_workspace, "nutrition_study_summary.txt")
    check("nutrition_study_summary.txt exists", os.path.isfile(summary_path))
    if os.path.isfile(summary_path):
        with open(summary_path) as f:
            text = f.read().lower()
        check("Summary has substantial content", len(text) > 200, f"Length: {len(text)}")
        check("Summary mentions engagement tiers", "high" in text and "low" in text and "medium" in text)
        check("Summary mentions meal plan", "meal" in text and "plan" in text)


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        cur.execute(
            "SELECT id, title FROM gsheet.spreadsheets WHERE lower(title) LIKE '%nutrition%study%'"
        )
        rows = cur.fetchall()
        check("Nutrition Study spreadsheet exists", len(rows) >= 1, f"Found {len(rows)} matching spreadsheets")
        if not rows:
            cur.close()
            conn.close()
            return

        ss_id = rows[0][0]

        # Check sheets
        cur.execute(
            "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s ORDER BY index", (ss_id,)
        )
        sheets = cur.fetchall()
        sheet_names = [s[1].lower() for s in sheets]
        check(
            "Has Student_Engagement sheet",
            any("engagement" in n for n in sheet_names),
            f"Sheets: {sheet_names}",
        )
        check(
            "Has Meal_Plans sheet",
            any("meal" in n or "plan" in n for n in sheet_names),
            f"Sheets: {sheet_names}",
        )

        # Check engagement sheet data
        engagement_sheet = None
        for s in sheets:
            if "engagement" in s[1].lower():
                engagement_sheet = s
                break
        if engagement_sheet:
            cur.execute(
                "SELECT row_index, col_index, value FROM gsheet.cells WHERE spreadsheet_id = %s AND sheet_id = %s ORDER BY row_index, col_index",
                (ss_id, engagement_sheet[0]),
            )
            cells = cur.fetchall()
            data_rows = {}
            for r, c, v in cells:
                if r not in data_rows:
                    data_rows[r] = {}
                data_rows[r][c] = v
            # Should have header + 3 data rows
            check(
                "Engagement sheet has >= 4 rows (header + 3 tiers)",
                len(data_rows) >= 4,
                f"Got {len(data_rows)} rows",
            )

        cur.close()
        conn.close()
    except Exception as e:
        check("GSheet check", False, str(e))


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        cur.execute(
            "SELECT id FROM notion.databases WHERE title::text LIKE '%Wellness Pilot Program%'"
        )
        dbs = cur.fetchall()
        check("Wellness Pilot Program database exists", len(dbs) >= 1, f"Found {len(dbs)}")
        if not dbs:
            cur.close()
            conn.close()
            return

        db_id = dbs[0][0]

        cur.execute(
            "SELECT id, properties FROM notion.pages WHERE parent::text LIKE %s AND archived = false",
            (f"%{db_id}%",),
        )
        pages = cur.fetchall()
        check("Has 3 Notion pages", len(pages) == 3, f"Found {len(pages)}")

        # Check page content
        if pages:
            all_props_text = " ".join(str(p[1]) for p in pages).lower()
            check("Has High/Control group", "control" in all_props_text)
            check("Has Partial group", "partial" in all_props_text)
            check("Has Full Meal Plan group", "full" in all_props_text.replace("full meal plan", "full"))

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check email to student_affairs
        cur.execute(
            "SELECT subject, body_text, to_addr FROM email.messages WHERE to_addr::text LIKE '%student_affairs%'"
        )
        sa_emails = cur.fetchall()
        check(
            "Email to student_affairs@university.edu sent",
            len(sa_emails) >= 1,
            f"Found {len(sa_emails)}",
        )
        if sa_emails:
            body = (sa_emails[0][1] or "").lower()
            check("SA email mentions pilot/study", "pilot" in body or "study" in body)
            check("SA email mentions tiers", "high" in body or "tier" in body or "engagement" in body)

        # Check email to dining_services
        cur.execute(
            "SELECT subject, body_text, to_addr FROM email.messages WHERE to_addr::text LIKE '%dining_services%'"
        )
        ds_emails = cur.fetchall()
        check(
            "Email to dining_services@university.edu sent",
            len(ds_emails) >= 1,
            f"Found {len(ds_emails)}",
        )
        if ds_emails:
            body = (ds_emails[0][1] or "").lower()
            check("DS email mentions meal plan", "meal" in body and "plan" in body)

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_no_noise(agent_workspace):
    """Verify noise data was not included in outputs."""
    print("\n=== Reverse Validation (noise rejection) ===")
    xlsx_path = os.path.join(agent_workspace, "Nutrition_Academic_Study.xlsx")
    if os.path.isfile(xlsx_path):
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            all_text = ""
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    all_text += " ".join(str(c) for c in row if c) + " "
            all_text = all_text.lower()
            noise_terms = ["budget tracking", "marketing", "parking", "faculty retreat"]
            found = [t for t in noise_terms if t in all_text]
            check("No noise data in Excel", len(found) == 0, f"Found: {found}")
        except Exception as e:
            check("Noise check readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_json_files(args.agent_workspace)
    check_gsheet()
    check_notion()
    check_emails()
    check_no_noise(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

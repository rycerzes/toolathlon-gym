"""
Check the agent's exam_review_plan.xlsx against the groundtruth.
"""

import os
import openpyxl


def str_match(a, b):
    """Case-insensitive, whitespace-normalized comparison."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_local(agent_workspace, groundtruth_workspace):
    """
    Compare exam_review_plan.xlsx from agent workspace against groundtruth.
    Returns (pass: bool, error_msg: str or None).
    """
    agent_file = os.path.join(agent_workspace, "exam_review_plan.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "exam_review_plan.xlsx")

    if not os.path.isfile(agent_file):
        return False, f"Agent workspace file does not exist: {agent_file}"
    if not os.path.isfile(gt_file):
        return False, f"Groundtruth file does not exist: {gt_file}"

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
        gt_wb = openpyxl.load_workbook(gt_file)
    except Exception as e:
        return False, f"Error reading Excel files: {e}"

    # Find the sheet (case-insensitive)
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    agent_ws = get_sheet(agent_wb, "Exam Plan")
    gt_ws = get_sheet(gt_wb, "Exam Plan")

    if agent_ws is None:
        return False, "Agent Excel file is missing 'Exam Plan' sheet"
    if gt_ws is None:
        return False, "Groundtruth Excel file is missing 'Exam Plan' sheet"

    # Read rows (skip header)
    agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

    if len(agent_rows) != len(gt_rows):
        return False, (
            f"Row count mismatch: expected {len(gt_rows)} rows, "
            f"got {len(agent_rows)} rows"
        )

    # Build lookup by Course_Code (column index 1)
    differences = []
    matched = 0

    col_names = [
        "Course_Name", "Course_Code", "Instructor_Name", "Instructor_Email",
        "Exam_Name", "Points_Possible", "Exam_Date",
        "Study_Session_1", "Study_Session_2",
    ]

    for gt_row in gt_rows:
        gt_code = str(gt_row[1]).strip() if gt_row[1] else ""
        # Find matching agent row by Course_Code
        agent_match = None
        for ar in agent_rows:
            if ar and str_match(ar[1], gt_code):
                agent_match = ar
                break

        if agent_match is None:
            differences.append(f"Course {gt_code} missing from agent output")
            continue

        row_diffs = []
        for i, col in enumerate(col_names):
            gt_val = gt_row[i] if i < len(gt_row) else None
            ag_val = agent_match[i] if i < len(agent_match) else None

            if col == "Points_Possible":
                if not num_close(ag_val, gt_val, 0.5):
                    row_diffs.append(
                        f"{col}: expected '{gt_val}', got '{ag_val}'"
                    )
            else:
                if not str_match(ag_val, gt_val):
                    row_diffs.append(
                        f"{col}: expected '{gt_val}', got '{ag_val}'"
                    )

        if row_diffs:
            differences.append(f"{gt_code}: {'; '.join(row_diffs)}")
        else:
            matched += 1

    if differences:
        return False, (
            f"Matched {matched}/{len(gt_rows)} courses. "
            f"Differences: {'; '.join(differences[:5])}"
        )

    print(f"All {matched} courses matched perfectly.")
    return True, None

"""
Evaluation for yf-portfolio-gsheet-tracker task.
Checks:
1. Local file: portfolio_summary.xlsx exists with correct Holdings and Summary sheets
2. Google Sheet: gsheet.cells in DB have been updated with correct values
All numeric comparisons use a 2% relative tolerance.
"""
import os
import sys
import json
from argparse import ArgumentParser
from datetime import datetime

import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONN = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

SPREADSHEET_ID = "sp_portfolio_tracker_q1_2026"
SHEET_ID = 1
TOLERANCE = 0.02  # 2% relative tolerance

# ---------- Latest prices from yf.stock_prices ----------
def get_latest_prices():
    """Query actual latest closing prices from Yahoo Finance DB."""
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()
    cur.execute("""
        SELECT symbol, close FROM yf.stock_prices
        WHERE symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM')
          AND date = (SELECT MAX(date) FROM yf.stock_prices WHERE symbol = 'AMZN')
        ORDER BY symbol;
    """)
    prices = {row[0]: float(row[1]) for row in cur.fetchall()}
    cur.close()
    conn.close()
    return prices


# ---------- Compute expected values ----------
# Portfolio holdings match initial_workspace/portfolio_holdings.xlsx.
# If the input file changes, these must be updated.
HOLDINGS = [
    ("AMZN", 50, 178.25),
    ("GOOGL", 80, 175.50),
    ("JNJ", 120, 148.30),
    ("JPM", 60, 198.75),
    ("XOM", 100, 112.40),
]
# Risk ratings come from the mock web portal. Cannot be queried from
# PostgreSQL at evaluation time.
RISK_RATINGS = {
    "AMZN": "Medium-High",
    "GOOGL": "Medium",
    "JNJ": "Low",
    "JPM": "Medium",
    "XOM": "Medium-High",
}


def compute_expected(prices):
    """Compute all expected values from latest prices."""
    rows = []
    for sym, shares, pp in HOLDINGS:
        cp = prices[sym]
        mv = round(shares * cp, 2)
        gl = round((cp - pp) * shares, 2)
        gl_pct = round(((cp - pp) / pp) * 100, 2)
        rows.append({
            "Symbol": sym, "Shares": shares, "Purchase_Price": pp,
            "Current_Price": cp, "Market_Value": mv,
            "Gain_Loss": gl, "Gain_Loss_Pct": gl_pct,
            "Risk_Rating": RISK_RATINGS[sym],
        })

    total_mv = sum(r["Market_Value"] for r in rows)
    for r in rows:
        r["Allocation_Pct"] = round((r["Market_Value"] / total_mv) * 100, 2)
        r["Compliance_Status"] = "OK" if r["Allocation_Pct"] <= 30 else "OVER_LIMIT"

    total_cost = round(sum(s * pp for _, s, pp in HOLDINGS), 2)
    total_gl = round(total_mv - total_cost, 2)
    total_gl_pct = round((total_gl / total_cost) * 100, 2)
    gl_pcts = {r["Symbol"]: r["Gain_Loss_Pct"] for r in rows}
    highest = max(gl_pcts, key=gl_pcts.get)
    lowest = min(gl_pcts, key=gl_pcts.get)
    compliance_issues = sum(1 for r in rows if r["Compliance_Status"] == "OVER_LIMIT")

    summary = {
        "Total_Market_Value": round(total_mv, 2),
        "Total_Cost_Basis": total_cost,
        "Total_Gain_Loss": total_gl,
        "Total_Gain_Loss_Pct": total_gl_pct,
        "Highest_Gainer": highest,
        "Lowest_Gainer": lowest,
        "Compliance_Issues": compliance_issues,
    }
    return rows, summary


# ---------- Comparison helpers ----------
def nums_close(expected, actual, tol=TOLERANCE):
    """Check if two numeric values are within relative tolerance."""
    try:
        e = float(expected)
        a = float(actual)
    except (ValueError, TypeError):
        return False
    if abs(e) < 1e-9:
        return abs(a) < 0.01
    return abs(e - a) / abs(e) <= tol


def val_match(expected, actual, tol=TOLERANCE):
    """Check if expected and actual values match (numeric or string)."""
    if expected is None and actual is None:
        return True
    e_str = str(expected).strip()
    a_str = str(actual).strip()
    if e_str.lower() == a_str.lower():
        return True
    return nums_close(expected, actual, tol)


# ---------- Check 1: Local Excel file ----------
def check_local_excel(workspace, expected_rows, expected_summary):
    path = os.path.join(workspace, "portfolio_summary.xlsx")
    if not os.path.exists(path):
        print(f"FAIL: portfolio_summary.xlsx not found at {path}")
        return 0, 0

    wb = openpyxl.load_workbook(path)
    total_checks = 0
    passed_checks = 0

    # Check Holdings sheet exists
    holdings_name = None
    for name in wb.sheetnames:
        if name.lower() == "holdings":
            holdings_name = name
            break
    if not holdings_name:
        print("FAIL: 'Holdings' sheet not found in portfolio_summary.xlsx")
        return 0, 1

    ws_h = wb[holdings_name]
    rows_data = list(ws_h.iter_rows(min_row=2, values_only=True))

    # Check Holdings data (5 rows x 10 columns)
    COLS = ["Symbol", "Shares", "Purchase_Price", "Current_Price", "Market_Value",
            "Gain_Loss", "Gain_Loss_Pct", "Allocation_Pct", "Risk_Rating", "Compliance_Status"]

    for i, exp_row in enumerate(expected_rows):
        if i >= len(rows_data):
            print(f"FAIL: Missing row {i+1} in Holdings sheet")
            total_checks += len(COLS)
            continue
        actual_row = rows_data[i]
        for j, col in enumerate(COLS):
            total_checks += 1
            exp_val = exp_row[col]
            act_val = actual_row[j] if j < len(actual_row) else None
            if val_match(exp_val, act_val):
                passed_checks += 1
            else:
                print(f"  Holdings mismatch row {i+1} col '{col}': expected={exp_val}, actual={act_val}")

    # Check Summary sheet exists
    summary_name = None
    for name in wb.sheetnames:
        if name.lower() == "summary":
            summary_name = name
            break
    if not summary_name:
        print("FAIL: 'Summary' sheet not found in portfolio_summary.xlsx")
        total_checks += len(expected_summary)
        return passed_checks, total_checks

    ws_s = wb[summary_name]
    summary_data = {}
    for row in ws_s.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            summary_data[str(row[0]).strip()] = row[1]

    for metric, exp_val in expected_summary.items():
        total_checks += 1
        act_val = summary_data.get(metric)
        if val_match(exp_val, act_val):
            passed_checks += 1
        else:
            print(f"  Summary mismatch '{metric}': expected={exp_val}, actual={act_val}")

    wb.close()
    return passed_checks, total_checks


# ---------- Check 2: Google Sheet in DB ----------
def check_gsheet(expected_rows):
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()

    # Fetch all cells for this spreadsheet/sheet
    cur.execute("""
        SELECT row_index, col_index, value FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index, col_index;
    """, (SPREADSHEET_ID, SHEET_ID))

    cells = {}
    for row_idx, col_idx, value in cur.fetchall():
        cells[(row_idx, col_idx)] = value

    cur.close()
    conn.close()

    if not cells:
        print("FAIL: No cells found in gsheet for the spreadsheet")
        return 0, 1

    # Columns that the agent should have filled: col indices 3-9
    FILL_COLS = {
        3: "Current_Price",
        4: "Market_Value",
        5: "Gain_Loss",
        6: "Gain_Loss_Pct",
        7: "Allocation_Pct",
        8: "Risk_Rating",
        9: "Compliance_Status",
    }

    total_checks = 0
    passed_checks = 0

    for i, exp_row in enumerate(expected_rows):
        row_idx = i + 1  # row 0 is header
        for col_idx, col_name in FILL_COLS.items():
            total_checks += 1
            exp_val = exp_row[col_name]
            act_val = cells.get((row_idx, col_idx))
            if act_val is None:
                print(f"  GSheet missing cell ({row_idx},{col_idx}) for {exp_row['Symbol']}.{col_name}")
                continue
            if val_match(exp_val, act_val):
                passed_checks += 1
            else:
                print(f"  GSheet mismatch ({row_idx},{col_idx}) {exp_row['Symbol']}.{col_name}: expected={exp_val}, actual={act_val}")

    return passed_checks, total_checks


# ---------- Main ----------
def main(args):
    print("Fetching latest stock prices from Yahoo Finance DB ...")
    prices = get_latest_prices()
    print(f"  Prices: {prices}")

    expected_rows, expected_summary = compute_expected(prices)
    print(f"  Expected summary: {expected_summary}")

    total_passed = 0
    total_checks = 0

    # Check 1: Local Excel
    print("\n--- Check 1: Local portfolio_summary.xlsx ---")
    p, t = check_local_excel(args.agent_workspace, expected_rows, expected_summary)
    print(f"  Local Excel: {p}/{t} checks passed")
    total_passed += p
    total_checks += t

    # Check 2: Google Sheet
    print("\n--- Check 2: Google Sheet in DB ---")
    p, t = check_gsheet(expected_rows)
    print(f"  Google Sheet: {p}/{t} checks passed")
    total_passed += p
    total_checks += t

    # Overall
    if total_checks == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = total_passed / total_checks * 100
    print(f"\nOverall: {total_passed}/{total_checks} checks passed ({accuracy:.1f}%)")

    if args.output_file:
        report = {
            "total_passed": total_passed,
            "total_checks": total_checks,
            "accuracy": accuracy,
            "timestamp": datetime.now().isoformat(),
        }
        with open(args.output_file, "w") as f:
            json.dump(report, f, indent=2)
        print(f"Report saved to {args.output_file}")

    if accuracy >= 80:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--output_file", required=False)
    args = parser.parse_args()
    main(args)

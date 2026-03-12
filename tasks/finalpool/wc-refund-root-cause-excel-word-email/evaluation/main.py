"""
Evaluation script for wc-refund-root-cause-excel-word-email task.

Checks:
1. Excel file (Refund_Analysis.xlsx) - 3 sheets with correct data
2. Word file (Root_Cause_Report.docx) - required sections and content
3. Email to quality_team@company.com with refund/analysis in subject
4. Email to supplier_relations@company.com with supplier/investigation in subject
"""

import argparse
import json
import os
import sys
from collections import defaultdict

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Get expected refund data from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT r.id, r.order_id, r.amount, r.reason, r.date_created, o.line_items
        FROM wc.refunds r JOIN wc.orders o ON r.order_id = o.id ORDER BY r.id
    """)
    refunds_raw = cur.fetchall()

    cur.execute("SELECT id, name FROM wc.products")
    prod_name_to_id = {row[1]: row[0] for row in cur.fetchall()}

    product_refund_map = defaultdict(lambda: {"count": 0, "total_amount": 0.0})

    refund_rows = []
    for rid, oid, amount, reason, date_created, line_items in refunds_raw:
        li = line_items if isinstance(line_items, list) else json.loads(line_items) if line_items else []
        product_names = [item.get("name", "?") for item in li] if li else ["Unknown"]
        amt = round(float(amount), 2)
        severity = "Critical" if amt > 50 else ("Major" if amt >= 20 else "Minor")
        refund_rows.append({
            "refund_id": rid, "order_id": oid, "amount": amt,
            "severity": severity, "reason": reason or "", "products": product_names,
        })
        for pn in product_names:
            product_refund_map[pn]["count"] += 1
            product_refund_map[pn]["total_amount"] += float(amount) / len(product_names)

    # Products requiring investigation (>= 2 refunds)
    investigation_products = [pn for pn, data in product_refund_map.items() if data["count"] >= 2]

    cur.close()
    conn.close()

    return refund_rows, product_refund_map, investigation_products


def check_excel(agent_workspace, refund_rows, product_refund_map):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Refund_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    # Check Refund Details sheet
    ws1 = None
    for s in wb.sheetnames:
        if "refund" in s.lower() and "detail" in s.lower():
            ws1 = wb[s]
            break
    if ws1 is None:
        for s in wb.sheetnames:
            if "refund" in s.lower():
                ws1 = wb[s]
                break
    if ws1 is None and len(wb.sheetnames) > 0:
        ws1 = wb[wb.sheetnames[0]]

    check("Refund Details sheet exists", ws1 is not None, f"Sheets: {wb.sheetnames}")
    if ws1:
        rows = list(ws1.iter_rows(min_row=2, values_only=True))
        expected_count = len(refund_rows)
        check(f"Refund Details has {expected_count} data rows",
              len(rows) == expected_count,
              f"Got {len(rows)} rows, expected {expected_count}")

        # Check header has key columns
        header = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in ws1[1]]
        for col in ["refund_id", "severity", "refund_amount"]:
            check(f"Refund Details has '{col}' column",
                  any(col in h for h in header),
                  f"Header: {header}")

        # Verify severity counts
        severity_col = None
        for i, h in enumerate(header):
            if "severity" in h:
                severity_col = i
                break
        if severity_col is not None:
            agent_critical = sum(1 for r in rows if r[severity_col] and "critical" in str(r[severity_col]).lower())
            expected_critical = sum(1 for r in refund_rows if r["severity"] == "Critical")
            check(f"Critical count matches ({expected_critical})",
                  agent_critical == expected_critical,
                  f"Got {agent_critical}")

    # Check Product Impact sheet
    ws2 = None
    for s in wb.sheetnames:
        if "product" in s.lower() and "impact" in s.lower():
            ws2 = wb[s]
            break
    check("Product Impact sheet exists", ws2 is not None, f"Sheets: {wb.sheetnames}")
    if ws2:
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        expected_products = len(product_refund_map)
        check(f"Product Impact has {expected_products} products",
              len(rows2) == expected_products,
              f"Got {len(rows2)} rows")

        # Check investigation column
        header2 = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in ws2[1]]
        invest_col = None
        for i, h in enumerate(header2):
            if "invest" in h:
                invest_col = i
                break
        if invest_col is not None:
            yes_count = sum(1 for r in rows2 if r[invest_col] and "yes" in str(r[invest_col]).lower())
            expected_yes = sum(1 for pn, d in product_refund_map.items() if d["count"] >= 2)
            check(f"Products requiring investigation = {expected_yes}",
                  yes_count == expected_yes,
                  f"Got {yes_count}")

    # Check Summary sheet
    ws3 = None
    for s in wb.sheetnames:
        if "summary" in s.lower():
            ws3 = wb[s]
            break
    check("Summary sheet exists", ws3 is not None, f"Sheets: {wb.sheetnames}")
    if ws3:
        summary = {}
        for row in ws3.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower().replace(" ", "_")] = row[1]

        check("Summary has total_refunds",
              any("total" in k and "refund" in k for k in summary),
              f"Keys: {list(summary.keys())}")

        total_key = None
        for k in summary:
            if "total_refund" in k and "amount" in k:
                total_key = k
                break
        if total_key:
            expected_total = round(sum(r["amount"] for r in refund_rows), 2)
            check(f"Total_Refund_Amount ~ {expected_total}",
                  num_close(summary[total_key], expected_total, 5.0),
                  f"Got {summary[total_key]}")


def check_word(agent_workspace, investigation_products):
    print("\n=== Checking Word Document ===")
    word_path = os.path.join(agent_workspace, "Root_Cause_Report.docx")
    check("Root_Cause_Report.docx exists", os.path.isfile(word_path), f"Expected {word_path}")
    if not os.path.isfile(word_path):
        return

    try:
        from docx import Document
        doc = Document(word_path)
    except Exception as e:
        check("Word file readable", False, str(e))
        return

    full_text = "\n".join([p.text for p in doc.paragraphs]).lower()

    # Check required sections
    required_sections = [
        "executive summary",
        "refund trend",
        "product",
        "root cause",
        "recommendation",
        "action plan",
    ]
    for section in required_sections:
        check(f"Word doc contains '{section}' section",
              section in full_text,
              f"Not found in document text")

    # Check that investigation products are mentioned
    for pn in investigation_products[:2]:
        # Check first 40 chars of product name (names are very long)
        short_name = pn[:40].lower()
        check(f"Word doc mentions investigation product",
              short_name in full_text,
              f"Product '{short_name}...' not found")


def check_emails(investigation_products):
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    def parse_to_addr(to_addr):
        if isinstance(to_addr, list):
            return " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                if isinstance(parsed, list):
                    return " ".join(str(r).lower() for r in parsed)
            except (json.JSONDecodeError, TypeError):
                pass
            return to_addr.lower()
        return str(to_addr).lower() if to_addr else ""

    # Check email to quality_team@company.com
    quality_email = None
    for subject, to_addr, body_text in all_emails:
        to_str = parse_to_addr(to_addr)
        if "quality_team@company.com" in to_str:
            quality_email = (subject, body_text)
            break

    check("Email sent to quality_team@company.com", quality_email is not None, db=True)
    if quality_email:
        subj = (quality_email[0] or "").lower()
        check("Quality email subject contains 'refund' or 'analysis'",
              "refund" in subj or "analysis" in subj,
              f"Subject: {quality_email[0]}", db=True)
        body = str(quality_email[1]) if quality_email[1] else ""
        check("Quality email body has content (>20 chars)",
              len(body) > 20, f"Body length: {len(body)}", db=True)

    # Check email to supplier_relations@company.com
    supplier_email = None
    for subject, to_addr, body_text in all_emails:
        to_str = parse_to_addr(to_addr)
        if "supplier_relations@company.com" in to_str:
            supplier_email = (subject, body_text)
            break

    check("Email sent to supplier_relations@company.com", supplier_email is not None, db=True)
    if supplier_email:
        subj = (supplier_email[0] or "").lower()
        check("Supplier email subject contains 'supplier' or 'investigation'",
              "supplier" in subj or "investigation" in subj,
              f"Subject: {supplier_email[0]}", db=True)
        body = str(supplier_email[1]) if supplier_email[1] else ""
        check("Supplier email body has content (>20 chars)",
              len(body) > 20, f"Body length: {len(body)}", db=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    refund_rows, product_refund_map, investigation_products = get_expected_data()
    print(f"Expected {len(refund_rows)} refunds, {len(product_refund_map)} products, "
          f"{len(investigation_products)} requiring investigation")

    check_excel(args.agent_workspace, refund_rows, product_refund_map)
    check_word(args.agent_workspace, investigation_products)
    check_emails(investigation_products)

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if args.res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": file_ok}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if file_ok else 1)


if __name__ == "__main__":
    main()

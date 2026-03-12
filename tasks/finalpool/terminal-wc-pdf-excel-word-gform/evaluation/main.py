"""Evaluation for terminal-wc-pdf-excel-word-gform.
Checks:
1. Product_Quality_Audit.xlsx with 4 sheets and correct data
2. Quality_Audit_Report.docx with required sections
3. Google Form "Supplier Quality Feedback" with 5+ questions
4. quality_audit.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: Product_Quality_Audit.xlsx ===")
    path = os.path.join(workspace, "Product_Quality_Audit.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # Review_Summary sheet
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s or "summary" in s), 0)
    ws = wb[sheets[rs_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    check("Review_Summary has 6 category rows", len(data_rows) >= 6, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    check("Contains Audio category", "audio" in all_text, f"Text: {all_text[:120]}")
    check("Contains Electronics category", "electronics" in all_text, f"Text: {all_text[:120]}")

    # Refund_Analysis sheet
    ra_idx = next((i for i, s in enumerate(sheets_lower) if "refund" in s or "order" in s), 1)
    if ra_idx < len(sheets):
        ws2 = wb[sheets[ra_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Refund_Analysis has at least 5 status rows", len(data_rows2) >= 5, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Contains refunded status", "refunded" in all_text2, f"Text: {all_text2[:120]}")
        check("Contains completed status", "completed" in all_text2, f"Text: {all_text2[:120]}")

    # Quality_Scorecard sheet
    qs_idx = next((i for i, s in enumerate(sheets_lower) if "quality" in s or "scorecard" in s), 2)
    if qs_idx < len(sheets):
        ws3 = wb[sheets[qs_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Quality_Scorecard has 6 rows", len(data_rows3) >= 6, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Contains risk levels", "low" in all_text3 or "medium" in all_text3 or "high" in all_text3,
              f"Text: {all_text3[:120]}")

    # Survey_Questions sheet
    sq_idx = next((i for i, s in enumerate(sheets_lower) if "survey" in s or "question" in s), 3)
    if sq_idx < len(sheets):
        ws4 = wb[sheets[sq_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Survey_Questions has at least 5 rows", len(data_rows4) >= 5, f"Found {len(data_rows4)}")


def check_word(workspace):
    print("\n=== Check 2: Quality_Audit_Report.docx ===")
    path = os.path.join(workspace, "Quality_Audit_Report.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)

    try:
        from docx import Document
        doc = Document(path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Contains executive summary", "executive" in all_text or "summary" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains review analysis", "review" in all_text and "category" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains refund discussion", "refund" in all_text, f"Text: {all_text[:150]}")
        check("Contains recommendations", "recommend" in all_text, f"Text: {all_text[:150]}")
    except ImportError:
        check("python-docx available", False, "python-docx not installed")


def check_gform():
    print("\n=== Check 3: Supplier Quality Feedback Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms WHERE lower(title) LIKE '%%supplier%%quality%%'")
    forms = cur.fetchall()
    check("Supplier Quality Feedback form exists", len(forms) >= 1,
          f"Found forms: {[f[1] for f in forms]}")

    if forms:
        form_id = forms[0][0]
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        qcount = cur.fetchone()[0]
        check("Form has at least 5 questions", qcount >= 5, f"Found {qcount} questions")

        cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s", (form_id,))
        questions = cur.fetchall()
        types = set(q[1] for q in questions)
        check("Form has multiple question types", len(types) >= 2, f"Types: {types}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: quality_audit.py ===")
    path = os.path.join(workspace, "quality_audit.py")
    check("quality_audit.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Product_Quality_Audit.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets beyond the 4 required
        expected_keywords = {"review", "summary", "refund", "order", "quality", "scorecard", "survey", "question"}
        for s in wb.sheetnames:
            s_lower = s.lower()
            matched = any(kw in s_lower for kw in expected_keywords)
            if not matched:
                check(f"No unexpected sheet '{s}'", False, f"Sheet '{s}' not expected")
                break
        else:
            check("No unexpected sheets in Excel", True)

        # Review ratings should not be negative
        sheets_lower = [s.lower() for s in wb.sheetnames]
        rs_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s or "summary" in s), 0)
        ws = wb[wb.sheetnames[rs_idx]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        has_negative = False
        for r in rows:
            for c in r:
                try:
                    v = float(str(c).replace(',', '').replace('%', '').replace('$', '').strip()) if c else None
                except (ValueError, TypeError):
                    v = None
                if v is not None and v < 0:
                    has_negative = True
                    break
            if has_negative:
                break
        check("No negative values in Review_Summary", not has_negative,
              "Found negative value in summary data")

    # Google Form: no duplicate forms with same name
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT title, COUNT(*) FROM gform.forms
            WHERE lower(title) LIKE '%%supplier%%quality%%'
            GROUP BY title HAVING COUNT(*) > 1
        """)
        dupes = cur.fetchall()
        check("No duplicate supplier quality forms", len(dupes) == 0,
              f"Duplicates: {dupes}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_gform()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

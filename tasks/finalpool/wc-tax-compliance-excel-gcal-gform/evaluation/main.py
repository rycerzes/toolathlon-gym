"""Evaluation for wc-tax-compliance-excel-gcal-gform."""
import os
import argparse, os, sys
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def safe_float(v, default=None):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def check_excel(agent_workspace, groundtruth_workspace):
    errors = []
    import openpyxl

    agent_path = os.path.join(agent_workspace, "Tax_Compliance_Report.xlsx")
    if not os.path.exists(agent_path):
        return ["Tax_Compliance_Report.xlsx not found"]

    gt_path = os.path.join(groundtruth_workspace, "Tax_Compliance_Report.xlsx")
    if not os.path.exists(gt_path):
        return ["Groundtruth Tax_Compliance_Report.xlsx not found"]

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)

        # --- Sheet 1: Order Tax Audit ---
        agent_rows = load_sheet_rows(wb_agent, "Order Tax Audit")
        gt_rows = load_sheet_rows(wb_gt, "Order Tax Audit")
        if agent_rows is None:
            errors.append("Sheet 'Order Tax Audit' not found")
        elif gt_rows is None:
            errors.append("Groundtruth sheet 'Order Tax Audit' not found")
        else:
            agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
            gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]
            if len(agent_data) != len(gt_data):
                errors.append(f"Order Tax Audit: {len(agent_data)} rows, expected {len(gt_data)}")
            else:
                # Build lookup by Order_ID
                gt_lookup = {}
                for r in gt_data:
                    oid = int(r[0]) if r[0] else None
                    if oid:
                        gt_lookup[oid] = r

                mismatches = 0
                for r in agent_data:
                    oid = int(r[0]) if r[0] else None
                    if oid not in gt_lookup:
                        mismatches += 1
                        continue
                    gt_r = gt_lookup[oid]
                    # Check Order_Total (col 1), tolerance 0.5
                    a_total = safe_float(r[1])
                    g_total = safe_float(gt_r[1])
                    if a_total is not None and g_total is not None and abs(a_total - g_total) > 0.5:
                        mismatches += 1
                        continue
                    # Check Applicable_Rate (col 3), tolerance 0.001
                    a_rate = safe_float(r[3])
                    g_rate = safe_float(gt_r[3])
                    if a_rate is not None and g_rate is not None and abs(a_rate - g_rate) > 0.001:
                        mismatches += 1
                        continue
                    # Check Expected_Tax (col 4), tolerance 0.5
                    a_exp = safe_float(r[4])
                    g_exp = safe_float(gt_r[4])
                    if a_exp is not None and g_exp is not None and abs(a_exp - g_exp) > 0.5:
                        mismatches += 1
                        continue
                    # Check Status (col 7)
                    a_status = str(r[7]).strip().lower() if r[7] else ""
                    g_status = str(gt_r[7]).strip().lower() if gt_r[7] else ""
                    if a_status != g_status:
                        mismatches += 1

                if mismatches > 5:
                    errors.append(f"Order Tax Audit: {mismatches} row mismatches (>5 threshold)")

        # --- Sheet 2: State Summary ---
        agent_ss = load_sheet_rows(wb_agent, "State Summary")
        gt_ss = load_sheet_rows(wb_gt, "State Summary")
        if agent_ss is None:
            errors.append("Sheet 'State Summary' not found")
        elif gt_ss is None:
            errors.append("Groundtruth sheet 'State Summary' not found")
        else:
            agent_ss_data = [r for r in agent_ss[1:] if r and r[0] is not None]
            gt_ss_data = [r for r in gt_ss[1:] if r and r[0] is not None]
            if abs(len(agent_ss_data) - len(gt_ss_data)) > 2:
                errors.append(f"State Summary: {len(agent_ss_data)} rows, expected ~{len(gt_ss_data)}")
            else:
                gt_state_lookup = {str(r[0]).strip().upper(): r for r in gt_ss_data}
                ss_mismatches = 0
                for r in agent_ss_data:
                    state = str(r[0]).strip().upper() if r[0] else ""
                    if state not in gt_state_lookup:
                        ss_mismatches += 1
                        continue
                    gt_r = gt_state_lookup[state]
                    # Check Order_Count (col 1)
                    a_count = safe_float(r[1])
                    g_count = safe_float(gt_r[1])
                    if a_count is not None and g_count is not None and abs(a_count - g_count) > 0:
                        ss_mismatches += 1
                        continue
                    # Check compliance rate (col 6), tolerance 0.5
                    a_comp = safe_float(r[6])
                    g_comp = safe_float(gt_r[6])
                    if a_comp is not None and g_comp is not None and abs(a_comp - g_comp) > 5.0:
                        ss_mismatches += 1

                if ss_mismatches > 3:
                    errors.append(f"State Summary: {ss_mismatches} state mismatches (>3 threshold)")

        # --- Sheet 3: Compliance Overview ---
        agent_co = load_sheet_rows(wb_agent, "Compliance Overview")
        gt_co = load_sheet_rows(wb_gt, "Compliance Overview")
        if agent_co is None:
            errors.append("Sheet 'Compliance Overview' not found")
        elif gt_co is None:
            errors.append("Groundtruth sheet 'Compliance Overview' not found")
        else:
            agent_co_data = {str(r[0]).strip().lower(): r[1] for r in agent_co[1:] if r and r[0]}
            gt_co_data = {str(r[0]).strip().lower(): r[1] for r in gt_co[1:] if r and r[0]}

            # Check total orders
            a_total = safe_float(agent_co_data.get("total_orders_audited"))
            g_total = safe_float(gt_co_data.get("total_orders_audited"))
            if a_total is not None and g_total is not None and abs(a_total - g_total) > 0:
                errors.append(f"Total_Orders_Audited: {a_total}, expected {g_total}")

            # Check compliant orders (tolerance 5)
            a_comp = safe_float(agent_co_data.get("compliant_orders"))
            g_comp = safe_float(gt_co_data.get("compliant_orders"))
            if a_comp is not None and g_comp is not None and abs(a_comp - g_comp) > 5:
                errors.append(f"Compliant_Orders: {a_comp}, expected {g_comp}")

            # Check overall compliance rate (tolerance 5)
            a_rate = safe_float(agent_co_data.get("overall_compliance_rate"))
            g_rate = safe_float(gt_co_data.get("overall_compliance_rate"))
            if a_rate is not None and g_rate is not None and abs(a_rate - g_rate) > 5.0:
                errors.append(f"Overall_Compliance_Rate: {a_rate}, expected {g_rate}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE summary ILIKE '%tax filing%' OR summary ILIKE '%tax deadline%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 4:
            errors.append(f"Expected 4 tax filing deadline events in GCal, found {len(rows)}")
        else:
            # Check that Q1-Q4 labels are present
            summaries = " ".join(str(r[0]) for r in rows).lower()
            for q in ["q1", "q2", "q3", "q4"]:
                if q not in summaries:
                    errors.append(f"Missing '{q}' in calendar event summaries")

            # Check dates (2026)
            dates = [r[1] for r in rows]
            from datetime import date
            expected_dates = [date(2026, 4, 15), date(2026, 7, 15), date(2026, 10, 15), date(2026, 1, 15)]
            for ed in expected_dates:
                if ed not in dates:
                    # Allow +/- 1 day tolerance
                    close = any(abs((d - ed).days) <= 1 for d in dates)
                    if not close:
                        errors.append(f"Missing calendar event for date {ed}")

    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%vendor%' OR title ILIKE '%tax information%'
            ORDER BY created_at DESC LIMIT 5
        """)
        forms = cur.fetchall()

        if not forms:
            cur.close()
            conn.close()
            return ["No GForm found matching 'Vendor Tax Information'"]

        form_id = forms[0][0]

        cur.execute("""
            SELECT title, question_type FROM gform.questions
            WHERE form_id = %s ORDER BY position
        """, (form_id,))
        questions = cur.fetchall()
        cur.close()
        conn.close()

        if len(questions) < 5:
            errors.append(f"Vendor Tax Information form has {len(questions)} questions, expected 5")

        # Check for key question topics
        q_titles = " ".join(str(q[0]) for q in questions).lower()
        for keyword in ["vendor", "tax id", "state", "exempt"]:
            if keyword not in q_titles and keyword.replace(" ", "") not in q_titles.replace(" ", ""):
                errors.append(f"Missing question about '{keyword}' in vendor form")

    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GForm...")
    errs = check_gform()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

"""Evaluation for wc-review-competitor-study."""
import argparse
import json
import os
import sys

import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_data):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Review_Benchmark.xlsx")
    if not os.path.exists(path):
        return ["Review_Benchmark.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Category Comparison sheet
        rows = load_sheet_rows(wb, "Category Comparison")
        if rows is None:
            errors.append("Sheet 'Category Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_categories"]
            if abs(len(data_rows) - expected) > 1:
                errors.append(f"Category Comparison has {len(data_rows)} rows, expected {expected}")

            # Check specific category ratings with tolerance
            for gc in gt_data["comparisons"]:
                cat_rows = [r for r in data_rows if r[0] and gc["category"].lower() in str(r[0]).lower()]
                if cat_rows:
                    row = cat_rows[0]
                    # Check our avg rating (col 1) with tolerance
                    if row[1] is not None:
                        try:
                            val = float(row[1])
                            if abs(val - gc["our_avg_rating"]) > 0.15:
                                errors.append(f"{gc['category']} our_avg_rating={val}, expected ~{gc['our_avg_rating']}")
                        except (ValueError, TypeError):
                            pass

        # Check Products Below Benchmark sheet
        rows2 = load_sheet_rows(wb, "Products Below Benchmark")
        if rows2 is None:
            errors.append("Sheet 'Products Below Benchmark' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            expected_pb = gt_data["products_below_count"]
            if abs(len(data_rows2) - expected_pb) > 5:
                errors.append(f"Products Below Benchmark has {len(data_rows2)} rows, expected ~{expected_pb}")

        # Check Summary sheet
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_notion(gt_data):
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE '%review%' OR properties::text ILIKE '%performance%'
            ORDER BY created_time DESC NULLS LAST
            LIMIT 10
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Filter for the specific page we expect
        found = False
        for row in rows:
            props = row[1] if isinstance(row[1], dict) else json.loads(row[1]) if row[1] else {}
            props_str = json.dumps(props).lower()
            if "review" in props_str and ("performance" in props_str or "analysis" in props_str or "q1" in props_str):
                found = True
                break

        if not found:
            errors.append("No Notion page found matching 'Review Performance Analysis'")

    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion(gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

"""Evaluation for wc-product-review-analysis-gform-gcal."""
import argparse
import os
import sys
import psycopg2
import openpyxl


DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_review_data():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.name, p.categories,
            count(pr.id) as review_count,
            avg(pr.rating) as avg_rating
        FROM wc.products p
        LEFT JOIN wc.product_reviews pr ON p.id = pr.product_id
        WHERE pr.id IS NOT NULL
        GROUP BY p.id, p.name, p.categories
        HAVING count(pr.id) >= 1
        ORDER BY avg(pr.rating) ASC, count(pr.id) DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    total = len(rows)
    lowest = rows[0] if rows else None
    overall_avg = round(sum(round(float(r[4]), 2) for r in rows) / len(rows), 2) if rows else 0
    return {
        "all_rows": rows,
        "total": total,
        "lowest_name": lowest[1][:40] if lowest else "",
        "lowest_avg": round(float(lowest[4]), 2) if lowest else 0,
        "top5": rows[:5],
        "overall_avg": overall_avg,
    }


def check_gform():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gform.forms WHERE LOWER(title) LIKE '%quality improvement survey%' OR LOWER(title) LIKE '%product quality%'")
        forms = cur.fetchall()
        if not forms:
            cur.close()
            conn.close()
            return False, "Form 'Product Quality Improvement Survey' not found"
        form_id = forms[0][0]
        cur.execute("SELECT count(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        cur.close()
        conn.close()
        if q_count < 6:
            return False, f"Form has {q_count} questions, expected >= 6 (5 product rating + 1 text)"
        return True, ""
    except Exception as e:
        return False, str(e)


def check_gcal_event():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT count(*) FROM gcal.events WHERE (LOWER(summary) LIKE '%quality%' OR LOWER(summary) LIKE '%product%') AND DATE(start_datetime) = '2026-03-20'")
        cnt = cur.fetchone()[0]
        if cnt == 0:
            # Try without date filter
            cur.execute("SELECT count(*) FROM gcal.events WHERE LOWER(summary) LIKE '%quality%' OR LOWER(summary) LIKE '%product%review%'")
            cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cnt >= 1
    except Exception:
        return False


def check_email_sent():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT count(*) FROM email.messages WHERE LOWER(to_addr::text) LIKE '%product.team%' AND (LOWER(subject) LIKE '%quality%' OR LOWER(subject) LIKE '%review%')")
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cnt >= 1
    except Exception:
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Product_Review_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Product_Review_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    all_errors = []

    try:
        review_data = get_review_data()
    except Exception as e:
        print(f"WARNING: Could not query DB: {e}")
        review_data = {
            "total": 82, "lowest_name": "dlx hmt Women Watch Casual Dress Analog",
            "lowest_avg": 3.5, "overall_avg": 4.62,
        }

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Review Summary sheet
    print("  Checking Review Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Review Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Review Summary' not found in agent output")
    else:
        data_rows = [r for r in a_rows[1:] if r and any(c is not None for c in r)]
        if len(data_rows) < 5:
            all_errors.append(f"Review Summary has {len(data_rows)} rows, expected >= 5 (at least top 5 lowest rated)")
        else:
            print(f"    PASS ({len(data_rows)} data rows)")

        # Check that data is sorted by avg_rating ascending (first row should be lowest rated)
        if len(data_rows) >= 2:
            first_rating = data_rows[0][1]
            last_rating = data_rows[-1][1]
            if first_rating is not None and last_rating is not None:
                try:
                    if float(first_rating) > float(last_rating):
                        all_errors.append(f"Data not sorted by Avg_Rating ASC: first={first_rating}, last={last_rating}")
                    else:
                        print("    Sort order PASS")
                except (TypeError, ValueError):
                    pass

    # Check Stats sheet
    print("  Checking Stats sheet...")
    a_rows = load_sheet_rows(agent_wb, "Stats")
    if a_rows is None:
        all_errors.append("Sheet 'Stats' not found in agent output")
    else:
        a_data = {str(r[0]).strip().lower(): r[1] for r in a_rows[1:] if r and r[0] is not None}
        errors = []

        total_val = a_data.get("total_products_reviewed")
        if total_val is None:
            errors.append("Missing metric: Total_Products_Reviewed")
        elif not num_close(total_val, review_data["total"], 0):
            errors.append(f"Total_Products_Reviewed: {total_val} vs {review_data['total']}")

        lowest_val = a_data.get("lowest_rated_product")
        if lowest_val is None:
            errors.append("Missing metric: Lowest_Rated_Product")
        else:
            expected_lower = review_data["lowest_name"].lower().strip()
            agent_lower = str(lowest_val).lower().strip()
            if expected_lower not in agent_lower and agent_lower not in expected_lower:
                errors.append(f"Lowest_Rated_Product: '{lowest_val}' vs '{review_data['lowest_name']}'")

        avg_overall_val = a_data.get("avg_rating_overall")
        if avg_overall_val is None:
            errors.append("Missing metric: Avg_Rating_Overall")
        elif not num_close(avg_overall_val, review_data["overall_avg"], 0.2):
            errors.append(f"Avg_Rating_Overall: {avg_overall_val} vs {review_data['overall_avg']} (tol=0.2)")

        if errors:
            all_errors.extend(errors)
            for e in errors[:5]:
                print(f"    ERROR: {e}")
        else:
            print("    PASS")

    # Check GForm
    print("  Checking Google Form...")
    ok, detail = check_gform()
    if ok:
        print("    PASS")
    else:
        all_errors.append(detail)

    # Check GCal event
    print("  Checking GCal event...")
    if check_gcal_event():
        print("    PASS")
    else:
        all_errors.append("Calendar event 'Product Quality Review Meeting' on March 20 2026 not found")

    # Check email sent
    print("  Checking email to product.team...")
    if check_email_sent():
        print("    PASS")
    else:
        all_errors.append("Email to product.team@company.com with 'quality' or 'review' subject not found")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

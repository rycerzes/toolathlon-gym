"""
Local check for yf-sector-rotation-dashboard task.
Compares agent-produced sector_rotation_report.xlsx against expected values
computed dynamically from PostgreSQL (yf.stock_prices, yf.financial_statements).

Sector benchmarks and analyst ratings come from the mock research portal
(static HTML served during the task). These are hardcoded constants since
they are defined by the task itself and do not change.

Falls back to static groundtruth Excel if PostgreSQL is unavailable.
"""
import os
import openpyxl
import math


# --- Constants from the mock research portal (http://localhost:30145) ---
STOCKS = ['AMZN', 'GOOGL', 'JNJ', 'JPM', 'XOM']
SECTORS = {
    'AMZN': 'Consumer Cyclical',
    'GOOGL': 'Communication Services',
    'JNJ': 'Healthcare',
    'JPM': 'Financial Services',
    'XOM': 'Energy',
}
BENCHMARKS = {
    'Consumer Cyclical': 18.5,
    'Communication Services': 22.0,
    'Healthcare': 8.5,
    'Financial Services': 15.0,
    'Energy': 12.0,
}
ANALYST_RATINGS = {
    'AMZN': 'Overweight',
    'GOOGL': 'Buy',
    'JNJ': 'Hold',
    'JPM': 'Outperform',
    'XOM': 'Neutral',
}
TARGET_PRICES = {
    'AMZN': 230,
    'GOOGL': 195,
    'JNJ': 165,
    'JPM': 260,
    'XOM': 115,
}

DB_CONFIG = dict(host=os.environ.get('PGHOST', 'localhost'), port=5432, database='toolathlon_gym', user='postgres', password='postgres')


def _str_match(a, b):
    """Case-insensitive string comparison after stripping."""
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_from_db():
    """Compute all expected values from PostgreSQL."""
    try:
        import psycopg2
    except ImportError:
        return None

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    perf = {}  # sym -> dict
    fin = {}   # sym -> dict

    for sym in STOCKS:
        sector = SECTORS[sym]
        bench = BENCHMARKS[sector]

        # Price 1Y ago and current
        cur.execute("SELECT close FROM yf.stock_prices WHERE symbol=%s AND date='2025-03-06'", (sym,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return None
        p1y = round(float(row[0]), 4)

        cur.execute("SELECT close FROM yf.stock_prices WHERE symbol=%s AND date='2026-03-05'", (sym,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return None
        pcur = round(float(row[0]), 2)

        ret_pct = round((pcur - p1y) / p1y * 100, 2)
        alpha = round(ret_pct - bench, 2)
        target = TARGET_PRICES[sym]
        upside = round((target - pcur) / pcur * 100, 2)

        perf[sym] = {
            'Symbol': sym,
            'Sector': sector,
            'Price_1Y_Ago': p1y,
            'Current_Price': pcur,
            'Return_Pct': ret_pct,
            'Benchmark_Return_Pct': bench,
            'Alpha': alpha,
            'Analyst_Rating': ANALYST_RATINGS[sym],
            'Target_Price': target,
            'Upside_Pct': upside,
        }

        # Financial data
        cur.execute(
            "SELECT data FROM yf.financial_statements "
            "WHERE symbol=%s AND stmt_type='income_stmt' AND freq='quarterly' "
            "ORDER BY period_end DESC LIMIT 1", (sym,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return None
        d = row[0]
        rev_raw = d.get('Total Revenue')
        ni_raw = d.get('Net Income')
        if rev_raw is None or ni_raw is None:
            conn.close()
            return None
        rev = round(float(rev_raw) / 1e6, 2)
        ni = round(float(ni_raw) / 1e6, 2)
        margin = round(ni / rev * 100, 2)

        fin[sym] = {
            'Symbol': sym,
            'Revenue_Latest_Q': rev,
            'Net_Income_Latest_Q': ni,
            'Profit_Margin_Pct': margin,
        }

    conn.close()

    # Summary
    returns = {s: perf[s]['Return_Pct'] for s in STOCKS}
    alphas = {s: perf[s]['Alpha'] for s in STOCKS}
    best = max(returns, key=returns.get)
    worst = min(returns, key=returns.get)
    avg_alpha = round(sum(alphas.values()) / len(alphas), 2)
    above = sum(1 for a in alphas.values() if a > 0)
    below = sum(1 for a in alphas.values() if a < 0)

    summary_rows = [
        ('Best_Performer', best),
        ('Worst_Performer', worst),
        ('Avg_Alpha', avg_alpha),
        ('Stocks_Above_Benchmark', above),
        ('Stocks_Below_Benchmark', below),
    ]

    return {'performance': perf, 'financials': fin, 'summary': summary_rows}


def check_local(agent_workspace: str, groundtruth_workspace: str):
    agent_file = os.path.join(agent_workspace, "sector_rotation_report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "sector_rotation_report.xlsx")

    if not os.path.exists(agent_file):
        return False, f"Missing agent file: {agent_file}"

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        return False, f"Error loading agent workbook: {e}"

    # Try DB-computed expected values first
    db_expected = compute_expected_from_db()
    use_db = db_expected is not None

    if use_db:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        return _check_against_db(agent_wb, db_expected)
    else:
        print("INFO: Falling back to static groundtruth Excel file")
        try:
            gt_wb = openpyxl.load_workbook(gt_file)
        except Exception as e:
            return False, f"Error loading groundtruth workbook: {e}"
        return _check_against_gt(agent_wb, gt_wb)


def _check_against_db(agent_wb, expected):
    """Check agent output against DB-computed expected values."""
    # Check sheet names
    required_sheets = ["Performance", "Financials", "Summary"]
    agent_sheet_map = {s.strip().lower(): s for s in agent_wb.sheetnames}
    for sheet_name in required_sheets:
        if sheet_name.lower() not in agent_sheet_map:
            return False, f"Missing sheet: {sheet_name}"

    perf = expected['performance']
    fin = expected['financials']
    summary = expected['summary']

    # --- Performance sheet ---
    agent_perf = agent_wb[agent_sheet_map["performance"]]
    agent_headers = [cell.value for cell in agent_perf[1]]

    expected_headers = ['Symbol', 'Sector', 'Price_1Y_Ago', 'Current_Price', 'Return_Pct',
                        'Benchmark_Return_Pct', 'Alpha', 'Analyst_Rating', 'Target_Price', 'Upside_Pct']
    if len(expected_headers) != len(agent_headers):
        return False, f"Performance header count mismatch: expected {len(expected_headers)}, got {len(agent_headers)}"
    for i, (eh, ah) in enumerate(zip(expected_headers, agent_headers)):
        if not _str_match(eh, ah):
            return False, f"Performance header mismatch at col {i+1}: expected '{eh}', got '{ah}'"

    # Check data rows (AMZN, GOOGL, JNJ, JPM, XOM in order)
    for row_idx, sym in enumerate(STOCKS, start=2):
        gt_vals = perf[sym]
        ordered_vals = [gt_vals[h] for h in expected_headers]
        for col_idx in range(1, 11):
            gt_val = ordered_vals[col_idx - 1]
            agent_val = agent_perf.cell(row=row_idx, column=col_idx).value

            if gt_val is None and agent_val is None:
                continue

            if isinstance(gt_val, str):
                if not _str_match(agent_val, gt_val):
                    col_name = expected_headers[col_idx - 1]
                    return False, f"Performance mismatch at row {row_idx}, col {col_name}: expected '{gt_val}', got '{agent_val}'"
            else:
                try:
                    gt_num = float(gt_val)
                    agent_num = float(agent_val)
                    if abs(gt_num - agent_num) > 1.0:
                        col_name = expected_headers[col_idx - 1]
                        return False, f"Performance mismatch at row {row_idx}, col {col_name}: expected {gt_num}, got {agent_num}"
                except (TypeError, ValueError):
                    col_name = expected_headers[col_idx - 1]
                    return False, f"Performance type mismatch at row {row_idx}, col {col_name}: expected {gt_val}, got {agent_val}"

    # --- Financials sheet ---
    agent_fin = agent_wb[agent_sheet_map["financials"]]
    agent_fin_headers = [cell.value for cell in agent_fin[1]]

    fin_expected_headers = ['Symbol', 'Revenue_Latest_Q', 'Net_Income_Latest_Q', 'Profit_Margin_Pct']
    if len(fin_expected_headers) != len(agent_fin_headers):
        return False, f"Financials header count mismatch: expected {len(fin_expected_headers)}, got {len(agent_fin_headers)}"
    for i, (eh, ah) in enumerate(zip(fin_expected_headers, agent_fin_headers)):
        if not _str_match(eh, ah):
            return False, f"Financials header mismatch at col {i+1}: expected '{eh}', got '{ah}'"

    for row_idx, sym in enumerate(STOCKS, start=2):
        gt_vals = fin[sym]
        ordered_vals = [gt_vals[h] for h in fin_expected_headers]
        for col_idx in range(1, 5):
            gt_val = ordered_vals[col_idx - 1]
            agent_val = agent_fin.cell(row=row_idx, column=col_idx).value

            if isinstance(gt_val, str):
                if not _str_match(agent_val, gt_val):
                    col_name = fin_expected_headers[col_idx - 1]
                    return False, f"Financials mismatch at row {row_idx}, col {col_name}: expected '{gt_val}', got '{agent_val}'"
            elif gt_val is not None:
                try:
                    gt_num = float(gt_val)
                    agent_num = float(agent_val)
                    if gt_num != 0:
                        rel_diff = abs(gt_num - agent_num) / abs(gt_num)
                        if rel_diff > 0.01:
                            col_name = fin_expected_headers[col_idx - 1]
                            return False, f"Financials mismatch at row {row_idx}, col {col_name}: expected {gt_num}, got {agent_num} (rel diff: {rel_diff:.4f})"
                    elif abs(agent_num) > 1.0:
                        col_name = fin_expected_headers[col_idx - 1]
                        return False, f"Financials mismatch at row {row_idx}, col {col_name}: expected {gt_num}, got {agent_num}"
                except (TypeError, ValueError):
                    col_name = fin_expected_headers[col_idx - 1]
                    return False, f"Financials type mismatch at row {row_idx}, col {col_name}: expected {gt_val}, got {agent_val}"

    # --- Summary sheet ---
    agent_sum = agent_wb[agent_sheet_map["summary"]]

    for row_idx, (gt_label, gt_value) in enumerate(summary, start=1):
        agent_label = agent_sum.cell(row=row_idx, column=1).value
        agent_value = agent_sum.cell(row=row_idx, column=2).value

        if not _str_match(gt_label, agent_label):
            return False, f"Summary label mismatch at row {row_idx}: expected '{gt_label}', got '{agent_label}'"

        if isinstance(gt_value, str):
            if not _str_match(agent_value, gt_value):
                return False, f"Summary value mismatch for {gt_label}: expected '{gt_value}', got '{agent_value}'"
        elif gt_value is not None:
            try:
                gt_num = float(gt_value)
                agent_num = float(agent_value)
                if abs(gt_num - agent_num) > 1.0:
                    return False, f"Summary value mismatch for {gt_label}: expected {gt_num}, got {agent_num}"
            except (TypeError, ValueError):
                if not _str_match(gt_value, agent_value):
                    return False, f"Summary value mismatch for {gt_label}: expected {gt_value}, got {agent_value}"

    return True, "All checks passed."


def _check_against_gt(agent_wb, gt_wb):
    """Fallback: Check agent output against static groundtruth Excel."""
    # Check sheet names (case-insensitive)
    required_sheets = ["Performance", "Financials", "Summary"]
    agent_sheet_map = {s.strip().lower(): s for s in agent_wb.sheetnames}
    for sheet_name in required_sheets:
        if sheet_name.lower() not in agent_sheet_map:
            return False, f"Missing sheet: {sheet_name}"

    # Check Performance sheet
    gt_perf = gt_wb["Performance"]
    agent_perf = agent_wb[agent_sheet_map["performance"]]

    gt_headers = [cell.value for cell in gt_perf[1]]
    agent_headers = [cell.value for cell in agent_perf[1]]
    if len(gt_headers) != len(agent_headers):
        return False, f"Performance header count mismatch: expected {len(gt_headers)}, got {len(agent_headers)}"
    for i, (gh, ah) in enumerate(zip(gt_headers, agent_headers)):
        if not _str_match(gh, ah):
            return False, f"Performance header mismatch at col {i+1}: expected '{gh}', got '{ah}'"

    for row_idx in range(2, 7):
        for col_idx in range(1, 11):
            gt_val = gt_perf.cell(row=row_idx, column=col_idx).value
            agent_val = agent_perf.cell(row=row_idx, column=col_idx).value

            if gt_val is None and agent_val is None:
                continue

            if isinstance(gt_val, str):
                if not _str_match(agent_val, gt_val):
                    col_name = gt_headers[col_idx - 1]
                    return False, f"Performance mismatch at row {row_idx}, col {col_name}: expected '{gt_val}', got '{agent_val}'"
            else:
                try:
                    gt_num = float(gt_val)
                    agent_num = float(agent_val)
                    if abs(gt_num - agent_num) > 1.0:
                        col_name = gt_headers[col_idx - 1]
                        return False, f"Performance mismatch at row {row_idx}, col {col_name}: expected {gt_num}, got {agent_num}"
                except (TypeError, ValueError):
                    col_name = gt_headers[col_idx - 1]
                    return False, f"Performance type mismatch at row {row_idx}, col {col_name}: expected {gt_val}, got {agent_val}"

    # Check Financials sheet
    gt_fin = gt_wb["Financials"]
    agent_fin = agent_wb[agent_sheet_map["financials"]]

    gt_fin_headers = [cell.value for cell in gt_fin[1]]
    agent_fin_headers = [cell.value for cell in agent_fin[1]]
    if len(gt_fin_headers) != len(agent_fin_headers):
        return False, f"Financials header count mismatch: expected {len(gt_fin_headers)}, got {len(agent_fin_headers)}"
    for i, (gh, ah) in enumerate(zip(gt_fin_headers, agent_fin_headers)):
        if not _str_match(gh, ah):
            return False, f"Financials header mismatch at col {i+1}: expected '{gh}', got '{ah}'"

    for row_idx in range(2, 7):
        for col_idx in range(1, 5):
            gt_val = gt_fin.cell(row=row_idx, column=col_idx).value
            agent_val = agent_fin.cell(row=row_idx, column=col_idx).value

            if isinstance(gt_val, str):
                if not _str_match(agent_val, gt_val):
                    col_name = gt_fin_headers[col_idx - 1]
                    return False, f"Financials mismatch at row {row_idx}, col {col_name}: expected '{gt_val}', got '{agent_val}'"
            elif gt_val is not None:
                try:
                    gt_num = float(gt_val)
                    agent_num = float(agent_val)
                    if gt_num != 0:
                        rel_diff = abs(gt_num - agent_num) / abs(gt_num)
                        if rel_diff > 0.01:
                            col_name = gt_fin_headers[col_idx - 1]
                            return False, f"Financials mismatch at row {row_idx}, col {col_name}: expected {gt_num}, got {agent_num} (rel diff: {rel_diff:.4f})"
                    elif abs(agent_num) > 1.0:
                        col_name = gt_fin_headers[col_idx - 1]
                        return False, f"Financials mismatch at row {row_idx}, col {col_name}: expected {gt_num}, got {agent_num}"
                except (TypeError, ValueError):
                    col_name = gt_fin_headers[col_idx - 1]
                    return False, f"Financials type mismatch at row {row_idx}, col {col_name}: expected {gt_val}, got {agent_val}"

    # Check Summary sheet
    gt_sum = gt_wb["Summary"]
    agent_sum = agent_wb[agent_sheet_map["summary"]]

    for row_idx in range(1, 6):
        gt_label = gt_sum.cell(row=row_idx, column=1).value
        gt_value = gt_sum.cell(row=row_idx, column=2).value
        agent_label = agent_sum.cell(row=row_idx, column=1).value
        agent_value = agent_sum.cell(row=row_idx, column=2).value

        if not _str_match(gt_label, agent_label):
            return False, f"Summary label mismatch at row {row_idx}: expected '{gt_label}', got '{agent_label}'"

        if isinstance(gt_value, str):
            if not _str_match(agent_value, gt_value):
                return False, f"Summary value mismatch for {gt_label}: expected '{gt_value}', got '{agent_value}'"
        elif gt_value is not None:
            try:
                gt_num = float(gt_value)
                agent_num = float(agent_value)
                if abs(gt_num - agent_num) > 1.0:
                    return False, f"Summary value mismatch for {gt_label}: expected {gt_num}, got {agent_num}"
            except (TypeError, ValueError):
                if not _str_match(gt_value, agent_value):
                    return False, f"Summary value mismatch for {gt_label}: expected {gt_value}, got {agent_value}"

    return True, "All checks passed."

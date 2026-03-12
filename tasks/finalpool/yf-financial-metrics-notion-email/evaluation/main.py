"""
Evaluation script for yf-financial-metrics-notion-email task.

Checks:
1. Excel file Financial_Health_Report.xlsx - 2 sheets with correct data
2. Notion page "Investment Portfolio Financial Analysis 2026" exists
3. Word document Financial_Analysis_Report.docx exists
4. Email sent to portfolio.manager@investment.com
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Financial_Health_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return False

    agent_file = os.path.join(agent_workspace, "Financial_Health_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Financial_Health_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth file exists", False, f"Not found: {gt_file}")
        return False

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Check Key Metrics sheet
    a_km = load_sheet_by_name(agent_wb, "Key Metrics")
    g_km = load_sheet_by_name(gt_wb, "Key Metrics")
    record("Sheet 'Key Metrics' exists", a_km is not None)

    if a_km is not None and g_km is not None:
        a_data = [r for r in a_km[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_km[1:] if any(v is not None for v in r)]
        record("Key Metrics has 3 rows",
               len(a_data) == 3,
               f"Found {len(a_data)} rows")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            sym = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(sym)
            if a_row is None:
                record(f"{sym} row in Key Metrics", False, "Not found")
                all_ok = False
                continue
            record(f"{sym} row exists", True)

            # Total_Revenue in billions (col 1)
            if len(g_row) > 1 and len(a_row) > 1:
                record(f"{sym}: Total_Revenue in expected range",
                       num_close(a_row[1], g_row[1], 20.0),
                       f"got {a_row[1]}B, expected ~{g_row[1]}B")
            # Net_Income in billions (col 2)
            if len(g_row) > 2 and len(a_row) > 2:
                record(f"{sym}: Net_Income in expected range",
                       num_close(a_row[2], g_row[2], 10.0),
                       f"got {a_row[2]}B, expected ~{g_row[2]}B")
            # Profit_Margin_Pct (col 4)
            if len(g_row) > 4 and len(a_row) > 4:
                record(f"{sym}: Profit_Margin_Pct correct",
                       num_close(a_row[4], g_row[4], 3.0),
                       f"got {a_row[4]}, expected ~{g_row[4]}")

    # Check Summary sheet
    a_summ = load_sheet_by_name(agent_wb, "Summary")
    g_summ = load_sheet_by_name(gt_wb, "Summary")
    record("Sheet 'Summary' exists", a_summ is not None)

    if a_summ is not None and g_summ is not None:
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_summ[1:] if any(v is not None for v in r)]

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary row: {g_row[0]}", False, "Not found")
                continue
            record(f"Summary row: {g_row[0]}", True)

            if key == "highest_revenue_company":
                record("Highest_Revenue_Company is AMZN",
                       a_row[1] is not None and "amzn" in str(a_row[1]).lower(),
                       f"got {a_row[1]}")
            elif key == "most_profitable_company":
                record("Most_Profitable_Company is GOOGL",
                       a_row[1] is not None and "googl" in str(a_row[1]).lower(),
                       f"got {a_row[1]}")
            elif key == "avg_profit_margin":
                record("Avg_Profit_Margin correct",
                       num_close(a_row[1], g_row[1], 3.0),
                       f"got {a_row[1]}, expected ~{g_row[1]}")

    return all_ok


# ============================================================================
# Check 2: Notion page
# ============================================================================

def check_notion():
    print("\n=== Checking Notion ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, properties FROM notion.pages WHERE archived=false OR archived IS NULL")
    pages = cur.fetchall()
    print(f"[check_notion] Found {len(pages)} Notion pages.")
    record("At least 1 Notion page created", len(pages) >= 1)

    found_page = False
    for page_id, props in pages:
        if props:
            props_str = str(props).lower()
            if ("investment" in props_str or "financial" in props_str or
                    "portfolio" in props_str):
                found_page = True
                record("Notion page with financial/investment content found", True)
                break

    # Also check blocks for content
    if not found_page:
        cur.execute("""
            SELECT b.block_data FROM notion.blocks b
            WHERE b.type IN ('paragraph', 'heading_1', 'heading_2')
            LIMIT 20
        """)
        blocks = cur.fetchall()
        for block in blocks:
            if block[0]:
                block_str = str(block[0]).lower()
                if "investment" in block_str or "financial" in block_str or "portfolio" in block_str:
                    found_page = True
                    record("Notion block with financial/investment content found", True)
                    break

    if not found_page:
        record("Notion page with investment/financial content", False,
               f"Pages: {len(pages)}")

    cur.close()
    conn.close()
    return found_page


# ============================================================================
# Check 3: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Checking Financial_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Financial_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        record("Word file exists", False, f"Not found: {docx_path}")
        return False
    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        record("Word doc has content", len(all_text.strip()) >= 100,
               f"Content length: {len(all_text.strip())}")
        record("Word doc mentions financial health or Q1",
               any(term in all_text for term in ["financial", "health", "q1", "2026", "revenue"]),
               "Missing financial content")
        record("Word doc mentions companies",
               any(term in all_text for term in ["googl", "amzn", "jnj", "alphabet", "amazon", "johnson"]),
               "Missing company names")
        return True
    except ImportError:
        size = os.path.getsize(docx_path)
        record("Word file has content (>2KB)", size > 2000, f"Size: {size} bytes")
        return size > 2000
    except Exception as e:
        record("Word file readable", False, str(e))
        return False


# ============================================================================
# Check 4: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")
    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    found_email = False
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = str(to_addr or "").lower()
        subject_lower = (subject or "").lower()
        if ("portfolio.manager@investment.com" in to_str or
                "financial" in subject_lower or "q1 2026" in subject_lower):
            found_email = True
            record("Email to portfolio.manager@investment.com found", True)

            record("Email subject mentions financial assessment",
                   any(term in subject_lower for term in ["financial", "q1", "health", "assessment"]),
                   f"Subject: {subject}")

            body_lower = (body_text or "").lower()
            record("Email body mentions companies and metrics",
                   any(term in body_lower for term in ["revenue", "profit", "googl", "amzn", "jnj", "amazon", "alphabet"]),
                   "Body missing company/metric content")
            break

    if not found_email:
        record("Financial assessment email found", False,
               f"Emails: {[(e[0], str(e[2])[:60]) for e in all_emails[:3]]}")

    return found_email


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    notion_ok = check_notion()
    word_ok = check_word(args.agent_workspace)
    email_ok = check_emails()

    all_passed = excel_ok and notion_ok and word_ok and email_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

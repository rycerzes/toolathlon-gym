"""Evaluation for terminal-howtocook-pdf-excel-word-gcal."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Campus_Dining_Plan.xlsx ===")
    agent_file = os.path.join(agent_ws, "Campus_Dining_Plan.xlsx")
    gt_file = os.path.join(gt_dir, "Campus_Dining_Plan.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
        gwb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Weekly_Menu
    print("  Checking Weekly_Menu...")
    ws1 = get_sheet(awb, "Weekly_Menu")
    check("Sheet Weekly_Menu exists", ws1 is not None, f"Sheets: {awb.sheetnames}")
    if ws1:
        rows = list(ws1.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in rows if r and r[0]]
        check("Weekly_Menu has 25 rows", len(data_rows) == 25, f"Got {len(data_rows)}")

        # Check days are present
        days_found = set()
        for r in data_rows:
            if r and r[0]:
                days_found.add(str(r[0]).strip().lower())
        expected_days = {"monday", "tuesday", "wednesday", "thursday", "friday"}
        check("All 5 days present", expected_days.issubset(days_found),
              f"Found: {days_found}")

        # Check each row has servings = 50
        servings_ok = all(r[3] == 50 for r in data_rows if r and len(r) >= 4)
        check("All servings = 50", servings_ok)

        # Check cost per serving values
        valid_costs = {8, 10, 12, 15, 25}
        costs_found = set()
        for r in data_rows:
            if r and len(r) >= 5 and r[4] is not None:
                costs_found.add(int(float(r[4])))
        check("Cost per serving values correct",
              costs_found == valid_costs,
              f"Found: {costs_found}, expected: {valid_costs}")

    # Sheet 2: Nutrition_Estimate
    print("  Checking Nutrition_Estimate...")
    ws2 = get_sheet(awb, "Nutrition_Estimate")
    check("Sheet Nutrition_Estimate exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2:
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        data_rows2 = [r for r in rows2 if r and r[0]]
        check("Nutrition_Estimate has 5 rows", len(data_rows2) == 5, f"Got {len(data_rows2)}")

        # Check calorie values
        expected_cals = {400, 350, 150, 100, 200}
        actual_cals = set()
        for r in data_rows2:
            if r and len(r) >= 3 and r[2] is not None:
                actual_cals.add(int(float(r[2])))
        check("Calorie estimates correct",
              actual_cals == expected_cals,
              f"Found: {actual_cals}, expected: {expected_cals}")

    # Sheet 3: Budget_Overview
    print("  Checking Budget_Overview...")
    ws3 = get_sheet(awb, "Budget_Overview")
    check("Sheet Budget_Overview exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        data_rows3 = [r for r in rows3 if r and r[0] and "total" not in str(r[0]).lower()]
        total_rows = [r for r in rows3 if r and r[0] and "total" in str(r[0]).lower()]

        check("Budget_Overview has 5 category rows", len(data_rows3) == 5, f"Got {len(data_rows3)}")
        check("Budget_Overview has total row", len(total_rows) >= 1, f"Got {len(total_rows)}")

        if total_rows:
            total_row = total_rows[0]
            # Check total daily cost = 3500
            if len(total_row) >= 4:
                check("Total daily cost = 3500",
                      num_close(total_row[3], 3500, 10),
                      f"Got {total_row[3]}")
            # Check five-day cost = 17500
            if len(total_row) >= 5:
                check("Total five-day cost = 17500",
                      num_close(total_row[4], 17500, 50),
                      f"Got {total_row[4]}")
            # Check pct of budget = 87.5
            if len(total_row) >= 6:
                check("Total pct of budget = 87.5",
                      num_close(total_row[5], 87.5, 1),
                      f"Got {total_row[5]}")


def check_word(agent_ws):
    print("\n=== Checking Meal_Plan_Report.docx ===")
    docx_path = os.path.join(agent_ws, "Meal_Plan_Report.docx")
    check("Word file exists", os.path.isfile(docx_path), docx_path)
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 300, f"Length: {len(text)}")
        check("Contains food/meal/festival reference",
              "food" in text or "meal" in text or "festival" in text or "dining" in text,
              "Missing event reference")
        check("Contains budget reference",
              "budget" in text or "cost" in text or "yuan" in text or "20000" in text or "20,000" in text,
              "Missing budget reference")
        check("Contains recipe reference",
              "recipe" in text or "dish" in text or "course" in text,
              "Missing recipe reference")
    except ImportError:
        check("python-docx available", False, "Cannot verify Word content")
    except Exception as e:
        check("Word document readable", False, str(e))


def check_gcal():
    print("\n=== Checking Calendar Events ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE LOWER(summary) LIKE '%food festival prep%'
        """)
        count = cur.fetchone()[0]
        check("5 Food Festival Prep events created", count == 5, f"Got {count}")

        if count > 0:
            cur.execute("""
                SELECT summary, start_datetime, end_datetime, description
                FROM gcal.events
                WHERE LOWER(summary) LIKE '%food festival prep%'
                ORDER BY start_datetime
            """)
            events = cur.fetchall()

            # Check date range (March 9-13, 2026)
            for ev in events:
                start = ev[1]
                if start:
                    check(f"Event '{ev[0]}' in March 9-13",
                          start.month == 3 and 9 <= start.day <= 13,
                          f"Date: {start}")

            # Check times (7:00 AM start)
            for ev in events:
                start = ev[1]
                if start:
                    check(f"Event '{ev[0]}' starts at 7 AM",
                          start.hour == 7,
                          f"Hour: {start.hour}")

            # Check descriptions have content
            with_desc = [ev for ev in events if ev[3] and len(ev[3]) > 10]
            check("Events have descriptions", len(with_desc) == len(events),
                  f"{len(with_desc)} of {len(events)} have descriptions")

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative cost or calorie values
    path = os.path.join(workspace, "Campus_Dining_Plan.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative values in Excel", not has_negative,
              "Found negative cost/calorie value")

    # GCal: no food festival prep events outside March 9-13
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE LOWER(summary) LIKE '%%food festival prep%%'
              AND (EXTRACT(MONTH FROM start_datetime) != 3
                   OR EXTRACT(DAY FROM start_datetime) < 9
                   OR EXTRACT(DAY FROM start_datetime) > 13)
        """)
        bad_events = cur.fetchone()[0]
        check("No food festival prep events outside March 9-13", bad_events == 0,
              f"Found {bad_events} events outside range")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gcal()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

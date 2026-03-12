"""
Check the agent's semester_grade_report.xlsx against the groundtruth.

Validates all three sheets: Course Grades, Grade Distribution, Summary.
"""

import os
import openpyxl


def str_match(a, b):
    """Case-insensitive, whitespace-normalized comparison."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=0.5):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, target):
    """Find sheet by name, case-insensitive."""
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


def check_local(agent_workspace, groundtruth_workspace):
    """
    Compare semester_grade_report.xlsx from agent workspace against groundtruth.
    Returns (passed_count, failed_count, error_details).
    """
    agent_file = os.path.join(agent_workspace, "semester_grade_report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "semester_grade_report.xlsx")

    if not os.path.isfile(agent_file):
        return 0, 1, [f"Agent workspace file does not exist: {agent_file}"]
    if not os.path.isfile(gt_file):
        return 0, 1, [f"Groundtruth file does not exist: {gt_file}"]

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        return 0, 1, [f"Error reading Excel files: {e}"]

    passed = 0
    failed = 0
    errors = []

    # ===== Sheet 1: Course Grades =====
    agent_ws1 = get_sheet(agent_wb, "Course Grades")
    gt_ws1 = get_sheet(gt_wb, "Course Grades")

    if agent_ws1 is None:
        failed += 1
        errors.append("Missing 'Course Grades' sheet")
    elif gt_ws1 is None:
        failed += 1
        errors.append("Groundtruth missing 'Course Grades' sheet")
    else:
        agent_rows = list(agent_ws1.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws1.iter_rows(min_row=2, values_only=True))

        if len(agent_rows) != len(gt_rows):
            failed += 1
            errors.append(
                f"Course Grades row count mismatch: expected {len(gt_rows)}, got {len(agent_rows)}"
            )
        else:
            passed += 1

        col_names = [
            "Course_Code", "Course_Name", "Lead_Instructor", "Instructor_Email",
            "Students_Scored", "Class_Average", "Letter_Grade", "Distinction", "Probation",
        ]

        for gt_row in gt_rows:
            gt_code = str(gt_row[0]).strip() if gt_row[0] else ""
            agent_match = None
            for ar in agent_rows:
                if ar and str_match(ar[0], gt_code):
                    agent_match = ar
                    break

            if agent_match is None:
                failed += 1
                errors.append(f"Course {gt_code} missing from agent output")
                continue

            row_ok = True
            for i, col in enumerate(col_names):
                gt_val = gt_row[i] if i < len(gt_row) else None
                ag_val = agent_match[i] if i < len(agent_match) else None

                if col in ("Students_Scored", "Class_Average"):
                    if not num_close(ag_val, gt_val, 0.5):
                        row_ok = False
                        errors.append(
                            f"{gt_code}.{col}: expected '{gt_val}', got '{ag_val}'"
                        )
                else:
                    if not str_match(ag_val, gt_val):
                        row_ok = False
                        errors.append(
                            f"{gt_code}.{col}: expected '{gt_val}', got '{ag_val}'"
                        )

            if row_ok:
                passed += 1
            else:
                failed += 1

    # ===== Sheet 2: Grade Distribution =====
    agent_ws2 = get_sheet(agent_wb, "Grade Distribution")
    gt_ws2 = get_sheet(gt_wb, "Grade Distribution")

    if agent_ws2 is None:
        failed += 1
        errors.append("Missing 'Grade Distribution' sheet")
    elif gt_ws2 is None:
        failed += 1
        errors.append("Groundtruth missing 'Grade Distribution' sheet")
    else:
        agent_rows2 = list(agent_ws2.iter_rows(min_row=2, values_only=True))
        gt_rows2 = list(gt_ws2.iter_rows(min_row=2, values_only=True))

        for gt_row in gt_rows2:
            gt_grade = str(gt_row[0]).strip().upper() if gt_row[0] else ""
            gt_count = gt_row[1] if gt_row[1] is not None else 0
            gt_courses = str(gt_row[2]).strip() if gt_row[2] else ""

            agent_match = None
            for ar in agent_rows2:
                if ar and str(ar[0]).strip().upper() == gt_grade:
                    agent_match = ar
                    break

            if agent_match is None:
                failed += 1
                errors.append(f"Grade Distribution: missing row for grade '{gt_grade}'")
                continue

            row_ok = True
            ag_count = agent_match[1] if agent_match[1] is not None else 0
            if not num_close(ag_count, gt_count, 0.01):
                row_ok = False
                errors.append(
                    f"Grade Distribution {gt_grade}: count expected {gt_count}, got {ag_count}"
                )

            ag_courses = str(agent_match[2]).strip() if agent_match[2] else ""
            # Normalize: sort comma-separated, strip, lowercase
            def normalize_courses(s):
                parts = [p.strip().lower() for p in s.split(",") if p.strip()]
                return sorted(parts)

            if normalize_courses(ag_courses) != normalize_courses(gt_courses):
                row_ok = False
                errors.append(
                    f"Grade Distribution {gt_grade}: courses expected '{gt_courses}', got '{ag_courses}'"
                )

            if row_ok:
                passed += 1
            else:
                failed += 1

    # ===== Sheet 3: Summary =====
    agent_ws3 = get_sheet(agent_wb, "Summary")
    gt_ws3 = get_sheet(gt_wb, "Summary")

    if agent_ws3 is None:
        failed += 1
        errors.append("Missing 'Summary' sheet")
    elif gt_ws3 is None:
        failed += 1
        errors.append("Groundtruth missing 'Summary' sheet")
    else:
        gt_summary = {}
        for row in gt_ws3.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_summary[str(row[0]).strip()] = row[1]

        agent_summary = {}
        for row in agent_ws3.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_summary[str(row[0]).strip()] = row[1]

        for metric, gt_val in gt_summary.items():
            ag_val = None
            # Case-insensitive key lookup
            for k, v in agent_summary.items():
                if k.lower() == metric.lower():
                    ag_val = v
                    break

            if ag_val is None:
                failed += 1
                errors.append(f"Summary: missing metric '{metric}'")
                continue

            if metric in ("Avg_Class_Average",):
                if num_close(ag_val, gt_val, 0.5):
                    passed += 1
                else:
                    failed += 1
                    errors.append(
                        f"Summary {metric}: expected '{gt_val}', got '{ag_val}'"
                    )
            elif metric in ("Total_Courses", "Distinction_Count", "Probation_Count"):
                if num_close(ag_val, gt_val, 0.01):
                    passed += 1
                else:
                    failed += 1
                    errors.append(
                        f"Summary {metric}: expected '{gt_val}', got '{ag_val}'"
                    )
            else:
                if str_match(ag_val, gt_val):
                    passed += 1
                else:
                    failed += 1
                    errors.append(
                        f"Summary {metric}: expected '{gt_val}', got '{ag_val}'"
                    )

    return passed, failed, errors

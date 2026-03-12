"""
Evaluation script for howtocook-scholarly-health-study task.

Checks:
1. Health_Diet_Analysis.xlsx with three sheets (Recipe Nutrition, Research Summary, Combined Analysis)
2. Chinese_Cuisine_Health_Report.docx with required sections
"""

import argparse
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Health_Diet_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: Recipe Nutrition ---
    rn_sheet = None
    for name in wb.sheetnames:
        if "recipe" in name.lower() and "nutrition" in name.lower():
            rn_sheet = name
            break
    if not rn_sheet:
        for name in wb.sheetnames:
            if "recipe" in name.lower() or "nutrition" in name.lower():
                rn_sheet = name
                break
    if not rn_sheet:
        record("Sheet 'Recipe Nutrition' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Recipe Nutrition' exists", True)
        ws = wb[rn_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        has_5_rows = len(data_rows) >= 5
        record(f"Recipe Nutrition has >= 5 rows ({len(data_rows)} found)", has_5_rows)
        if not has_5_rows:
            all_ok = False

        # Check that calorie values are non-zero for at least some rows
        nonzero_cal = 0
        for row in data_rows:
            if row and len(row) >= 3:
                try:
                    cal = float(row[2]) if row[2] is not None else 0
                    if cal > 0:
                        nonzero_cal += 1
                except (ValueError, TypeError):
                    pass
        has_nonzero = nonzero_cal >= 3
        record(f"At least 3 rows have non-zero calories ({nonzero_cal} found)", has_nonzero)
        if not has_nonzero:
            all_ok = False

    # --- Sheet 2: Research Summary ---
    rs_sheet = None
    for name in wb.sheetnames:
        if "research" in name.lower() and "summary" in name.lower():
            rs_sheet = name
            break
    if not rs_sheet:
        for name in wb.sheetnames:
            if "research" in name.lower() or "summary" in name.lower():
                rs_sheet = name
                break
    if not rs_sheet:
        record("Sheet 'Research Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Research Summary' exists", True)
        ws = wb[rs_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        has_3_rows = len(data_rows) >= 3
        record(f"Research Summary has >= 3 rows ({len(data_rows)} found)", has_3_rows)
        if not has_3_rows:
            all_ok = False

        # Check that rows have non-empty titles and authors
        if data_rows:
            rows_with_title = sum(
                1 for row in data_rows if row and row[0] and str(row[0]).strip()
            )
            has_titles = rows_with_title >= 3
            record(
                f"Research Summary has >= 3 rows with non-empty titles ({rows_with_title} found)",
                has_titles,
            )
            if not has_titles:
                all_ok = False

            rows_with_authors = sum(
                1 for row in data_rows
                if row and len(row) > 1 and row[1] and str(row[1]).strip()
            )
            has_authors = rows_with_authors >= 3
            record(
                f"Research Summary has >= 3 rows with non-empty authors ({rows_with_authors} found)",
                has_authors,
            )
            if not has_authors:
                all_ok = False

            # Check that at least some rows have citation counts > 0
            rows_with_citations = 0
            for row in data_rows:
                if row and len(row) > 2:
                    try:
                        if float(row[2]) > 0:
                            rows_with_citations += 1
                    except (TypeError, ValueError):
                        # Try later columns in case citation count is not at index 2
                        for ci in range(3, min(len(row), 6)):
                            try:
                                if row[ci] is not None and float(row[ci]) > 0:
                                    rows_with_citations += 1
                                    break
                            except (TypeError, ValueError):
                                continue
            has_citations = rows_with_citations >= 1
            record(
                f"Research Summary has >= 1 row with citation count > 0 ({rows_with_citations} found)",
                has_citations,
            )
            if not has_citations:
                all_ok = False

    # --- Sheet 3: Combined Analysis ---
    ca_sheet = None
    for name in wb.sheetnames:
        if "combined" in name.lower() and "analysis" in name.lower():
            ca_sheet = name
            break
    if not ca_sheet:
        for name in wb.sheetnames:
            if "combined" in name.lower() or "analysis" in name.lower():
                ca_sheet = name
                break
    if not ca_sheet:
        record("Sheet 'Combined Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Combined Analysis' exists", True)
        ws = wb[ca_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        has_3_rows = len(data_rows) >= 3
        record(
            f"Combined Analysis has >= 3 rows ({len(data_rows)} found)", has_3_rows
        )
        if not has_3_rows:
            all_ok = False

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Health_Diet_Analysis.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()

    wb.close()
    return all_ok


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    doc_file = os.path.join(agent_workspace, "Chinese_Cuisine_Health_Report.docx")
    if not os.path.isfile(doc_file):
        record("Word document exists", False, f"Not found: {doc_file}")
        return False
    record("Word document exists", True)

    try:
        from docx import Document

        doc = Document(doc_file)
        full_text = " ".join(p.text for p in doc.paragraphs)
        text_lower = full_text.lower()

        has_length = len(full_text) >= 500
        record(
            f"Document has >= 500 chars ({len(full_text)} found)", has_length
        )
        if not has_length:
            return False

        has_title = (
            "nutritional analysis" in text_lower or "chinese cuisine" in text_lower
        )
        record("Document contains title keywords", has_title)

        sections_ok = True
        for section in ["introduction", "recipe analysis", "literature review", "conclusions"]:
            # Check both as heading text and within body content
            found = section in text_lower
            record(f"Document has '{section}' section", found)
            if not found:
                sections_ok = False

        return has_length and has_title and sections_ok

    except ImportError:
        file_size = os.path.getsize(doc_file)
        has_content = file_size > 1000
        record(
            "Word document has content (size > 1KB, python-docx not available)",
            has_content,
            f"File size: {file_size} bytes",
        )
        return has_content
    except Exception as e:
        record("Word document readable", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    word_ok = check_word(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Excel:   {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Word:    {'PASS' if word_ok else 'FAIL'}")

    overall = excel_ok and word_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

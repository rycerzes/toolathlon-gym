"""Evaluation script for terminal-sf-wc-support-email-gcal-excel."""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_expected_ticket_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT "PRIORITY", COUNT(*) as cnt, ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY" ORDER BY cnt DESC
    """)
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return rows, total


def get_expected_order_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT status, COUNT(*) FROM wc.orders GROUP BY status ORDER BY COUNT(*) DESC")
    rows = cur.fetchall()
    cur.execute("SELECT COUNT(*) FROM wc.orders")
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return rows, total


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Get expected data
    ticket_data, total_tickets = get_expected_ticket_data()
    order_data, total_orders = get_expected_order_data()

    # Check Excel
    excel_path = os.path.join(agent_workspace, "Product_Quality_Report.xlsx")
    check("Product_Quality_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # Ticket_by_Priority
        check("Ticket_by_Priority sheet exists", "Ticket_by_Priority" in wb.sheetnames)
        if "Ticket_by_Priority" in wb.sheetnames:
            ws = wb["Ticket_by_Priority"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Ticket_by_Priority has 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Priority', 'Ticket_Count', 'Avg_Response_Hours', 'Pct_of_Total']:
                check(f"Ticket_by_Priority has {col}", col.lower() in headers, f"headers: {headers[:5]}")

            # Verify ticket counts match DB
            if data_rows:
                row_dict = {str(r[0]).strip(): r for r in data_rows if r[0]}
                for priority, cnt, avg_resp in ticket_data:
                    if priority in row_dict:
                        agent_cnt = safe_float(row_dict[priority][1])
                        if agent_cnt is not None:
                            check(f"Ticket count for {priority}", abs(agent_cnt - cnt) < 5,
                                  f"expected ~{cnt}, got {agent_cnt}")

        # Order_Status_Breakdown
        check("Order_Status_Breakdown sheet exists", "Order_Status_Breakdown" in wb.sheetnames)
        if "Order_Status_Breakdown" in wb.sheetnames:
            ws = wb["Order_Status_Breakdown"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Order_Status_Breakdown has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Status', 'Order_Count', 'Quality_Flag']:
                check(f"Order_Status_Breakdown has {col}", col.lower() in headers, f"headers: {headers[:5]}")
            # Check Review flags for refunded/cancelled
            for row in data_rows:
                status = str(row[0]).lower() if row[0] else ""
                flag = str(row[-1]).lower() if row[-1] else ""
                if status in ("refunded", "cancelled"):
                    check(f"Quality_Flag 'Review' for {status}", "review" in flag, f"got '{flag}'")

        # Quality_Action_Plan
        check("Quality_Action_Plan sheet exists", "Quality_Action_Plan" in wb.sheetnames)
        if "Quality_Action_Plan" in wb.sheetnames:
            ws = wb["Quality_Action_Plan"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Quality_Action_Plan has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Issue', 'Severity', 'Action_Item']:
                check(f"Quality_Action_Plan has {col}", col.lower() in headers, f"headers: {headers[:6]}")

    # Check terminal script
    check("defect_correlation.py exists",
          os.path.exists(os.path.join(agent_workspace, "defect_correlation.py")))

    # Check emails
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr FROM email.messages WHERE subject ILIKE %s", ('%quality%analysis%',))
        quality_emails = cur.fetchall()
        check("Quality analysis email sent", len(quality_emails) >= 1, f"found {len(quality_emails)}")
        if quality_emails:
            check("Email to quality-team", "quality-team" in str(quality_emails[0][1]).lower(),
                  f"to: {quality_emails[0][1]}")

        cur.execute("SELECT subject, to_addr FROM email.messages WHERE subject ILIKE %s", ('%quality%review%meeting%',))
        review_emails = cur.fetchall()
        check("Quality review meeting email sent", len(review_emails) >= 1, f"found {len(review_emails)}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email checks", False, str(e))

    # Check calendar
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime FROM gcal.events WHERE summary ILIKE %s", ('%quality%review%',))
        review_events = cur.fetchall()
        check("Quality Review Meeting event exists", len(review_events) >= 1, f"found {len(review_events)}")
        if review_events:
            check("Review event in March 2026", "2026-03" in str(review_events[0][2]))

        cur.execute("SELECT summary, description, start_datetime FROM gcal.events WHERE summary ILIKE %s", ('%quality%improvement%',))
        plan_events = cur.fetchall()
        check("Quality Improvement Planning event exists", len(plan_events) >= 1, f"found {len(plan_events)}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar checks", False, str(e))

    check_reverse_validation(agent_workspace)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Product_Quality_Report.xlsx")
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        # No unexpected sheets beyond the 3 required ones
        expected_sheets = {"ticket_by_priority", "order_status_breakdown", "quality_action_plan"}
        unexpected = [s for s in wb.sheetnames if s.lower().replace(" ", "_") not in expected_sheets]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected sheets: {unexpected}")

        # Ticket counts should not be negative
        if "Ticket_by_Priority" in wb.sheetnames:
            ws = wb["Ticket_by_Priority"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[1] is not None:
                    val = safe_float(row[1])
                    if val is not None and val < 0:
                        check("No negative ticket counts", False, f"Found negative: {val}")
                        break
            else:
                check("No negative ticket counts", True)

    # Calendar: no events scheduled before today
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%quality%%'
              AND start_datetime < '2026-03-01'
        """)
        old_events = cur.fetchone()[0]
        check("No quality events before March 2026", old_events == 0,
              f"Found {old_events} old events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

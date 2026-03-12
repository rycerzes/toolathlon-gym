"""
Evaluation script for wc-sales-tax-summary task.

Checks:
1. Excel file Tax_Summary_Report.xlsx with By State and Overall sheets
2. Google Sheet "Tax Summary for Accounting" with summary data
"""
import argparse
import json
import os
import sys

import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT billing->>'state' as state,
               COUNT(*) as order_count,
               ROUND(SUM(total::numeric), 2) as total_sales,
               ROUND(SUM(total_tax::numeric), 2) as total_tax
        FROM wc.orders
        WHERE status IN ('completed', 'processing')
        AND billing->>'state' IS NOT NULL
        AND billing->>'state' != ''
        GROUP BY billing->>'state'
        ORDER BY SUM(total_tax::numeric) DESC
    """)
    state_rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*),
               ROUND(SUM(total::numeric), 2),
               ROUND(SUM(total_tax::numeric), 2)
        FROM wc.orders
        WHERE status IN ('completed', 'processing')
    """)
    ov = cur.fetchone()
    total_orders = int(ov[0])
    total_sales = float(ov[1])
    total_tax = float(ov[2])
    eff_rate = round(total_tax / total_sales * 100, 2) if total_sales > 0 else 0.0

    cur.close()
    conn.close()

    return {
        "states": [{"state": r[0], "count": int(r[1]), "sales": float(r[2]),
                     "tax": float(r[3])} for r in state_rows],
        "overall": {
            "total_orders": total_orders,
            "total_sales": total_sales,
            "total_tax": total_tax,
            "effective_rate": eff_rate,
        }
    }


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def number_in_text(value, text, tolerance=0.01):
    text = str(text)
    val_str = str(value)
    if val_str in text:
        return True
    try:
        f2 = f"{float(value):.2f}"
        if f2 in text:
            return True
    except (ValueError, TypeError):
        pass
    try:
        formatted = f"{int(value):,}"
        if formatted in text:
            return True
    except (ValueError, TypeError):
        pass
    return False


def check_excel(agent_workspace):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Tax_Summary_Report.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    expected = load_expected()

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    def sheet_text(ws):
        txt = ""
        for row in ws.iter_rows(values_only=True):
            txt += " ".join(str(c) for c in row if c is not None) + " "
        return txt

    # By State sheet
    ws_state = find_sheet(["state"])
    check("By State sheet exists", ws_state is not None, f"Sheets: {wb.sheetnames}")
    if ws_state:
        txt = sheet_text(ws_state)
        # Check top 5 states by tax
        for s in expected["states"][:5]:
            check(f"State '{s['state']}' in sheet", s["state"] in txt)
            check(f"Tax {s['tax']} for {s['state']}",
                  number_in_text(s["tax"], txt))

    # Overall sheet
    ws_ov = find_sheet(["overall"])
    check("Overall sheet exists", ws_ov is not None, f"Sheets: {wb.sheetnames}")
    if ws_ov:
        txt = sheet_text(ws_ov)
        ov = expected["overall"]
        check(f"Total Orders = {ov['total_orders']}",
              number_in_text(ov["total_orders"], txt))
        check(f"Total Sales = {ov['total_sales']}",
              number_in_text(ov["total_sales"], txt))
        check(f"Total Tax = {ov['total_tax']}",
              number_in_text(ov["total_tax"], txt))
        check(f"Effective Tax Rate = {ov['effective_rate']}",
              number_in_text(ov["effective_rate"], txt))


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE LOWER(title) LIKE '%%tax%%' AND LOWER(title) LIKE '%%accounting%%'
    """)
    sheets = cur.fetchall()
    check("Google Sheet 'Tax Summary for Accounting' exists",
          len(sheets) >= 1,
          f"Found {len(sheets)} matching spreadsheets")

    if sheets:
        ss_id = sheets[0][0]
        cur.execute("""
            SELECT c.value FROM gsheet.cells c
            JOIN gsheet.sheets s ON c.spreadsheet_id = s.spreadsheet_id AND c.sheet_id = s.id
            WHERE c.spreadsheet_id = %s
        """, (ss_id,))
        cells = cur.fetchall()
        all_values = " ".join(str(c[0]) for c in cells if c[0])

        expected = load_expected()
        ov = expected["overall"]
        check("GSheet contains Total Orders",
              number_in_text(ov["total_orders"], all_values))
        check("GSheet contains Total Sales",
              number_in_text(ov["total_sales"], all_values))
        check("GSheet contains Total Tax",
              number_in_text(ov["total_tax"], all_values))

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gsheet()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_passed = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

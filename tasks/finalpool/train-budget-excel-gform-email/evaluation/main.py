"""
Evaluation for train-budget-excel-gform-email task.

Checks:
1. Travel_Budget.xlsx with Seat_Options, Budget_Scenarios, Summary sheets
2. Correct prices and totals for 8 people
3. Survey form with 3 questions exists
4. Email sent to finance@company.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Travel_Budget.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Travel_Budget.xlsx")
    if not os.path.exists(xlsx_path):
        record("Travel_Budget.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Travel_Budget.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Seat_Options sheet
    if "seat_options" not in sheet_names_lower:
        record("Seat_Options sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Seat_Options sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("seat_options")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Seat_Options has at least 4 rows", len(data_rows) >= 4,
               f"Found {len(data_rows)}")

        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        record("Seat_Options has G11 and G1", "G11" in all_text and "G1" in all_text, all_text[:200])

        numeric_vals = []
        for r in data_rows:
            for c in r:
                try:
                    numeric_vals.append(float(c))
                except (TypeError, ValueError):
                    pass
        has_349 = any(abs(v - 349.0) < 0.1 for v in numeric_vals)
        has_553 = any(abs(v - 553.0) < 0.1 for v in numeric_vals)
        has_1748 = any(abs(v - 1748.5) < 0.5 for v in numeric_vals)
        record("Seat prices include 349.0, 553.0, 1748.5", has_349 and has_553 and has_1748,
               f"Numerics: {numeric_vals}")

    # Budget_Scenarios sheet
    if "budget_scenarios" not in sheet_names_lower:
        record("Budget_Scenarios sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Budget_Scenarios sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("budget_scenarios")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        record("Budget_Scenarios has 3 rows", len(data_rows2) == 3,
               f"Found {len(data_rows2)}")

        numeric_vals2 = []
        for r in data_rows2:
            for c in r:
                try:
                    numeric_vals2.append(float(c))
                except (TypeError, ValueError):
                    pass
        has_2792 = any(abs(v - 2792.0) < 1.0 for v in numeric_vals2)
        has_4424 = any(abs(v - 4424.0) < 1.0 for v in numeric_vals2)
        has_13988 = any(abs(v - 13988.0) < 5.0 for v in numeric_vals2)
        record("Budget total 2792 CNY correct", has_2792, f"Numerics: {numeric_vals2}")
        record("Standard total 4424 CNY correct", has_4424, f"Numerics: {numeric_vals2}")
        record("Premium total ~13988 CNY correct", has_13988, f"Numerics: {numeric_vals2}")

    # Summary sheet
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws3 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows3 = list(ws3.iter_rows(values_only=True))
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        has_price_diff = any(
            abs(float(c) - 11196.0) < 10.0
            for r in rows3 for c in r
            if c is not None and str(c).replace('.', '').replace('-', '').isdigit()
        )
        record("Summary has price difference ~11196 CNY", has_price_diff,
               f"Text: {all_text3[:200]}")


def check_gform():
    print("\n=== Check 2: Survey Form ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE title ILIKE '%travel preference%' OR title ILIKE '%business trip%'
    """)
    forms = cur.fetchall()
    record("Travel preference survey form exists", len(forms) >= 1,
           f"Found forms: {[f[1] for f in forms]}")

    if forms:
        form_id = forms[0][0]
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        record("Form has exactly 3 questions", q_count == 3, f"Found {q_count}")

        cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position", (form_id,))
        questions = cur.fetchall()
        has_departure_q = any("departure" in (q[0] or "").lower() or "time" in (q[0] or "").lower() for q in questions)
        has_seat_q = any("seat" in (q[0] or "").lower() or "class" in (q[0] or "").lower() for q in questions)
        has_text_q = any(q[1] in ("TEXT", "PARAGRAPH") for q in questions)
        record("Form has departure time question", has_departure_q, f"Questions: {questions}")
        record("Form has seat class question", has_seat_q, f"Questions: {questions}")
        record("Form has open text question", has_text_q, f"Questions: {questions}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email to finance@company.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT to_addr, subject FROM email.messages
        WHERE subject ILIKE '%budget%' OR subject ILIKE '%beijing-shanghai%' OR subject ILIKE '%conference%'
    """)
    messages = cur.fetchall()
    cur.close()
    conn.close()

    all_msgs = list(messages)
    record("Budget analysis email sent", len(all_msgs) >= 1,
           f"Found {len(all_msgs)} matching emails")

    if all_msgs:
        to_raw = all_msgs[0][0]
        to_str = str(to_raw).lower() if to_raw else ""
        record("Email sent to finance@company.com", "finance@company.com" in to_str,
               f"To: {to_str[:100]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""
Evaluation script for quarterly-sales-restock-review task.

Checks:
1. Excel file (Q4_2025_Business_Review.xlsx) - 4 sheets with correct data
2. Notion page created with correct structure
3. Emails sent to correct suppliers (and not to wrong ones)

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check the Excel output file against groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Q4_2025_Business_Review.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Q4_2025_Business_Review.xlsx")

    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
        gt_wb = openpyxl.load_workbook(gt_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    all_passed = True

    # Check sheet names
    expected_sheets = ["Regional Sales", "Brand Performance", "Restock Alerts", "Summary"]
    agent_sheets = agent_wb.sheetnames
    for sheet_name in expected_sheets:
        found = any(str_match(s, sheet_name) for s in agent_sheets)
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {agent_sheets}")
        if not found:
            all_passed = False

    # Helper to find sheet case-insensitively
    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if str_match(s, name):
                return wb[s]
        return None

    # --- Sheet 1: Regional Sales ---
    print("\n--- Regional Sales ---")
    agent_ws = get_sheet(agent_wb, "Regional Sales")
    gt_ws = get_sheet(gt_wb, "Regional Sales")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Regional Sales row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        for gt_row in gt_rows:
            region = gt_row[0]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], region):
                    matched = ar
                    break
            if matched:
                check(f"Region '{region}' Order_Count",
                      num_close(matched[1], gt_row[1], 0.5),
                      f"Expected {gt_row[1]}, got {matched[1]}")
                check(f"Region '{region}' Total_Revenue",
                      num_close(matched[2], gt_row[2], 1.0),
                      f"Expected {gt_row[2]}, got {matched[2]}")
                check(f"Region '{region}' Avg_Order_Value",
                      num_close(matched[3], gt_row[3], 0.5),
                      f"Expected {gt_row[3]}, got {matched[3]}")
            else:
                check(f"Region '{region}' found", False, "Region not in agent output")
                all_passed = False
    else:
        all_passed = False

    # --- Sheet 2: Brand Performance ---
    print("\n--- Brand Performance ---")
    agent_ws = get_sheet(agent_wb, "Brand Performance")
    gt_ws = get_sheet(gt_wb, "Brand Performance")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Brand Performance row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        for gt_row in gt_rows:
            brand = gt_row[0]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], brand):
                    matched = ar
                    break
            if matched:
                check(f"Brand '{brand}' Total_Revenue",
                      num_close(matched[2], gt_row[2], 1.0),
                      f"Expected {gt_row[2]}, got {matched[2]}")
                check(f"Brand '{brand}' Total_Units",
                      num_close(matched[3], gt_row[3], 0.5),
                      f"Expected {gt_row[3]}, got {matched[3]}")
            else:
                check(f"Brand '{brand}' found", False, "Brand not in agent output")
                all_passed = False
    else:
        all_passed = False

    # --- Sheet 3: Restock Alerts ---
    print("\n--- Restock Alerts ---")
    agent_ws = get_sheet(agent_wb, "Restock Alerts")
    gt_ws = get_sheet(gt_wb, "Restock Alerts")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Restock Alerts row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        # Check a sample of key rows (out-of-stock products)
        oos_skus = [r[0] for r in gt_rows if r[2] == 0]
        for sku in oos_skus:
            found = any(ar and str_match(ar[0], sku) for ar in agent_rows)
            check(f"Out-of-stock SKU '{sku}' in Restock Alerts", found)
            if not found:
                all_passed = False
    else:
        all_passed = False

    # --- Sheet 4: Summary ---
    print("\n--- Summary ---")
    agent_ws = get_sheet(agent_wb, "Summary")
    gt_ws = get_sheet(gt_wb, "Summary")
    if agent_ws and gt_ws:
        gt_data = {}
        for row in gt_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]
        agent_data = {}
        for row in agent_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key, gt_val in gt_data.items():
            agent_val = agent_data.get(key)
            if agent_val is None:
                # Try fuzzy key match
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                ok = num_close(agent_val, gt_val, 1.0)
                check(f"Summary '{key}'", ok,
                      f"Expected {gt_val}, got {agent_val}")
            else:
                ok = str_match(agent_val, gt_val)
                check(f"Summary '{key}'", ok,
                      f"Expected '{gt_val}', got '{agent_val}'")
            if not ok:
                all_passed = False
    else:
        all_passed = False

    return all_passed


def check_notion():
    """Check that a Notion page was created with the correct structure."""
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    all_passed = True

    # Check for a page with title containing "Q4 2025"
    cur.execute("""
        SELECT id, properties
        FROM notion.pages
        WHERE archived = false
    """)
    pages = cur.fetchall()

    found_page = None
    for page_id, props in pages:
        if props and "title" in props:
            title_data = props["title"]
            if isinstance(title_data, dict) and "title" in title_data:
                title_arr = title_data["title"]
                if isinstance(title_arr, list):
                    title_text = " ".join(
                        t.get("text", {}).get("content", "")
                        for t in title_arr if isinstance(t, dict)
                    )
                    if "q4 2025" in title_text.lower():
                        found_page = page_id
                        break

    check("Notion page with 'Q4 2025' in title exists", found_page is not None)
    if not found_page:
        cur.close()
        conn.close()
        return False

    # Check for content blocks
    cur.execute("""
        SELECT type, block_data
        FROM notion.blocks
        WHERE parent_id = %s
        ORDER BY position
    """, (found_page,))
    blocks = cur.fetchall()

    block_types = [b[0] for b in blocks]
    block_texts = []
    for btype, bd in blocks:
        text = ""
        if bd:
            # block_data is nested: {"heading_1": {"rich_text": [...]}} or {"paragraph": {"rich_text": [...]}}
            inner = bd.get(btype, bd)
            rich_text = inner.get("rich_text", []) if isinstance(inner, dict) else []
            if not rich_text and "rich_text" in bd:
                rich_text = bd["rich_text"]
            text = " ".join(
                t.get("text", {}).get("content", "")
                for t in rich_text if isinstance(t, dict)
            )
        block_texts.append(text.lower())

    # Check for headings
    has_sales_heading = any(
        btype in ("heading_1", "heading_2", "heading_3") and "sales" in btext
        for btype, btext in zip(block_types, block_texts)
    )
    check("Notion page has 'Sales' heading", has_sales_heading)

    has_inventory_heading = any(
        btype in ("heading_1", "heading_2", "heading_3") and "inventory" in btext
        for btype, btext in zip(block_types, block_texts)
    )
    check("Notion page has 'Inventory' heading", has_inventory_heading)

    has_action_heading = any(
        btype in ("heading_1", "heading_2", "heading_3") and "action" in btext
        for btype, btext in zip(block_types, block_texts)
    )
    check("Notion page has 'Action Items' heading", has_action_heading)

    # Check for to_do blocks
    has_todo = any(btype == "to_do" for btype in block_types)
    check("Notion page has to-do items", has_todo)

    # Check for key content mentions
    all_text = " ".join(block_texts)
    check("Notion mentions 'Asia Pacific' (top region)",
          "asia pacific" in all_text)
    check("Notion mentions 'LG' (top brand)",
          "lg" in all_text.lower().replace(",", " ").replace(".", " ").split())

    cur.close()
    conn.close()
    return all_passed


def check_emails():
    """Check that restock emails were sent to the correct suppliers."""
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    all_passed = True

    # Get sent emails (folder_id=2 is Sent)
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE folder_id = 2
    """)
    sent_emails = cur.fetchall()

    # Also check inbox of recipients via sent_log or messages in any folder
    # The email MCP typically stores sent emails in folder 2
    # But also check all messages for supplier recipients
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()

    # Expected suppliers to receive emails (those with out-of-stock products)
    expected_suppliers = {
        "trade@asiatech.com": {
            "name": "Asia Tech Trading",
            "skus": ["INFINITY-JBL-GLIDE-1020", "SILENCER-PANELS-FLAM-1079",
                     "30KG-DIGITAL-SCALE-1071"],
        },
        "b2b@smarthome-ws.com": {
            "name": "SmartHome Wholesale",
            "skus": ["BOXTUDIO-LIGHTBOX-TA-1039"],
        },
        "wholesale@digitaldreams.com": {
            "name": "Digital Dreams Supply",
            "skus": ["TYPEC-EARPHONE-FOR-1045"],
        },
    }

    # Should NOT receive emails (suppliers without out-of-stock products)
    should_not_receive = [
        "verkauf@euroelec.de",
        "supply@premiumgadgets.com",
        "sales@globalelec.com",
        "orders@avpartners.com",
        "orders@techworld-dist.com",
    ]

    def find_email_for_recipient(recipient):
        """Find an email addressed to this recipient across all emails."""
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    # Check each expected supplier received an email
    for supplier_email, info in expected_suppliers.items():
        result = find_email_for_recipient(supplier_email)
        check(f"Email sent to {supplier_email}", result is not None)
        if result:
            subj, from_addr, to_addr, body = result
            # Check subject contains "Restock" (case-insensitive)
            has_restock_subject = "restock" in (subj or "").lower()
            check(f"Email to {supplier_email} subject contains 'Restock'",
                  has_restock_subject,
                  f"Subject: {(subj or '')[:100]}")
            # Check subject contains supplier name
            has_supplier_name = info["name"].lower() in (subj or "").lower()
            check(f"Email to {supplier_email} subject contains supplier name",
                  has_supplier_name,
                  f"Subject: {(subj or '')[:100]}")
            # Check body mentions at least one SKU
            body_lower = (body or "").lower()
            has_sku = any(sku.lower() in body_lower for sku in info["skus"])
            check(f"Email to {supplier_email} body mentions product SKU(s)",
                  has_sku,
                  f"Expected one of {info['skus']}")
        else:
            all_passed = False

    # Check suppliers that should NOT receive emails
    for email_addr in should_not_receive:
        result = find_email_for_recipient(email_addr)
        check(f"No email sent to {email_addr} (no out-of-stock products)",
              result is None,
              f"Unexpected email found with subject: {result[0][:100] if result else ''}")
        if result:
            all_passed = False

    cur.close()
    conn.close()
    return all_passed


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    excel_ok = check_excel(agent_workspace, groundtruth_workspace)
    notion_ok = check_notion()
    email_ok = check_emails()

    all_passed = excel_ok and notion_ok and email_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed and FAIL_COUNT == 0 else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed and FAIL_COUNT == 0,
            "details": {
                "excel": excel_ok,
                "notion": notion_ok,
                "email": email_ok,
            }
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return all_passed and FAIL_COUNT == 0, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

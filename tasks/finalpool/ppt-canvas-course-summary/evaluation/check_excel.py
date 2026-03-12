"""
Check the agent's Course_Summary_AAA_F13.xlsx against dynamically queried
ground truth data from the Canvas PostgreSQL tables.
"""

import os
import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

COURSE_ID_QUERY = """
    SELECT id FROM canvas.courses
    WHERE course_code = 'AAA-2013J'
"""


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


def query_expected_data():
    """Query PostgreSQL for expected enrollment, assignment, and grade data."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute(COURSE_ID_QUERY)
    course_id = cur.fetchone()[0]

    # Enrollment stats
    cur.execute("""
        SELECT
            COUNT(CASE WHEN type='StudentEnrollment' THEN 1 END),
            COUNT(CASE WHEN type='TeacherEnrollment' THEN 1 END),
            COUNT(CASE WHEN type='TaEnrollment' THEN 1 END)
        FROM canvas.enrollments WHERE course_id = %s
    """, (course_id,))
    enrollment = cur.fetchone()

    # Assignment performance
    cur.execute("""
        SELECT a.name,
            ROUND(AVG(s.score::float)::numeric, 2),
            MAX(s.score::float),
            MIN(s.score::float),
            COUNT(*)
        FROM canvas.assignments a
        JOIN canvas.submissions s ON s.assignment_id = a.id AND s.score IS NOT NULL
        WHERE a.course_id = %s
        GROUP BY a.name
        ORDER BY a.name
    """, (course_id,))
    assignments = cur.fetchall()

    # Grade distribution
    cur.execute("""
        WITH student_avgs AS (
            SELECT s.user_id, AVG(s.score::float) as avg_score
            FROM canvas.assignments a
            JOIN canvas.submissions s ON s.assignment_id = a.id AND s.score IS NOT NULL
            WHERE a.course_id = %s
            GROUP BY s.user_id
        )
        SELECT
            CASE
                WHEN avg_score >= 90 THEN 'A (90-100)'
                WHEN avg_score >= 80 THEN 'B (80-89)'
                WHEN avg_score >= 70 THEN 'C (70-79)'
                WHEN avg_score >= 60 THEN 'D (60-69)'
                ELSE 'F (<60)'
            END as grade_range,
            COUNT(*),
            ROUND((COUNT(*)::float / (SELECT COUNT(*) FROM student_avgs) * 100)::numeric, 1)
        FROM student_avgs
        GROUP BY grade_range
        ORDER BY grade_range
    """, (course_id,))
    grade_dist = cur.fetchall()

    cur.close()
    conn.close()
    return enrollment, assignments, grade_dist


def check_excel(agent_workspace, groundtruth_workspace):
    """
    Validate Course_Summary_AAA_F13.xlsx from agent workspace.
    Returns (passed_count, failed_count, error_details).
    """
    agent_file = os.path.join(agent_workspace, "Course_Summary_AAA_F13.xlsx")

    if not os.path.isfile(agent_file):
        return 0, 1, [f"Agent workspace file does not exist: {agent_file}"]

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        return 0, 1, [f"Error reading Excel file: {e}"]

    passed = 0
    failed = 0
    errors = []

    try:
        enrollment_exp, assignments_exp, grade_dist_exp = query_expected_data()
    except Exception as e:
        return 0, 1, [f"Error querying expected data: {e}"]

    # ===== Sheet 1: Enrollment Stats =====
    ws1 = get_sheet(agent_wb, "Enrollment Stats")
    if ws1 is None:
        failed += 1
        errors.append("Missing 'Enrollment Stats' sheet")
    else:
        rows = list(ws1.iter_rows(min_row=2, values_only=True))
        if len(rows) < 1:
            failed += 1
            errors.append("Enrollment Stats has no data rows")
        else:
            row = rows[0]
            exp_student, exp_teacher, exp_ta = enrollment_exp
            checks = [
                ("Student_Count", row[0] if len(row) > 0 else None, exp_student),
                ("Teacher_Count", row[1] if len(row) > 1 else None, exp_teacher),
                ("TA_Count", row[2] if len(row) > 2 else None, exp_ta),
            ]
            all_ok = True
            for label, got, exp in checks:
                if not num_close(got, exp, 1):
                    all_ok = False
                    errors.append(f"Enrollment Stats.{label}: expected {exp}, got {got}")
            if all_ok:
                passed += 1
            else:
                failed += 1

    # ===== Sheet 2: Assignment Performance =====
    ws2 = get_sheet(agent_wb, "Assignment Performance")
    if ws2 is None:
        failed += 1
        errors.append("Missing 'Assignment Performance' sheet")
    else:
        agent_rows = list(ws2.iter_rows(min_row=2, values_only=True))

        if len(agent_rows) == len(assignments_exp):
            passed += 1
        else:
            failed += 1
            errors.append(
                f"Assignment Performance row count: expected {len(assignments_exp)}, got {len(agent_rows)}"
            )

        for exp_name, exp_avg, exp_max, exp_min, exp_count in assignments_exp:
            agent_match = None
            for ar in agent_rows:
                if ar and str_match(ar[0], exp_name):
                    agent_match = ar
                    break

            if agent_match is None:
                failed += 1
                errors.append(f"Assignment '{exp_name}' missing from Assignment Performance")
                continue

            row_ok = True
            if not num_close(agent_match[1], float(exp_avg), 0.5):
                row_ok = False
                errors.append(f"{exp_name}.Avg_Score: expected {exp_avg}, got {agent_match[1]}")
            if not num_close(agent_match[2], float(exp_max), 1):
                row_ok = False
                errors.append(f"{exp_name}.Max_Score: expected {exp_max}, got {agent_match[2]}")
            if not num_close(agent_match[3], float(exp_min), 1):
                row_ok = False
                errors.append(f"{exp_name}.Min_Score: expected {exp_min}, got {agent_match[3]}")
            if not num_close(agent_match[4], exp_count, 1):
                row_ok = False
                errors.append(f"{exp_name}.Submission_Count: expected {exp_count}, got {agent_match[4]}")

            if row_ok:
                passed += 1
            else:
                failed += 1

        # Check sorted alphabetically
        agent_names = [str(r[0]).strip() for r in agent_rows if r and r[0]]
        if agent_names == sorted(agent_names):
            passed += 1
        else:
            failed += 1
            errors.append(f"Assignment Performance not sorted alphabetically: {agent_names}")

    # ===== Sheet 3: Grade Distribution =====
    ws3 = get_sheet(agent_wb, "Grade Distribution")
    if ws3 is None:
        failed += 1
        errors.append("Missing 'Grade Distribution' sheet")
    else:
        agent_rows3 = list(ws3.iter_rows(min_row=2, values_only=True))

        if len(agent_rows3) == len(grade_dist_exp):
            passed += 1
        else:
            failed += 1
            errors.append(
                f"Grade Distribution row count: expected {len(grade_dist_exp)}, got {len(agent_rows3)}"
            )

        for exp_range, exp_count, exp_pct in grade_dist_exp:
            agent_match = None
            for ar in agent_rows3:
                if ar and str_match(ar[0], exp_range):
                    agent_match = ar
                    break

            if agent_match is None:
                failed += 1
                errors.append(f"Grade range '{exp_range}' missing from Grade Distribution")
                continue

            row_ok = True
            if not num_close(agent_match[1], exp_count, 1):
                row_ok = False
                errors.append(f"{exp_range}.Student_Count: expected {exp_count}, got {agent_match[1]}")
            if not num_close(agent_match[2], float(exp_pct), 0.5):
                row_ok = False
                errors.append(f"{exp_range}.Percentage: expected {exp_pct}, got {agent_match[2]}")

            if row_ok:
                passed += 1
            else:
                failed += 1

    return passed, failed, errors

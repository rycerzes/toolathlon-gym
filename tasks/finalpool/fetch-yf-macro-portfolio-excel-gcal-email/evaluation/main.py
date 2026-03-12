"""Evaluation for fetch-yf-macro-portfolio-excel-gcal-email."""
import argparse
import os
import sys

import psycopg2


DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Macro_Portfolio_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Macro_Portfolio_Analysis.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Macro Forecast sheet
        rows = load_sheet_rows(wb, "Macro Forecast")
        if rows is None:
            errors.append("Sheet 'Macro Forecast' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 4:
                errors.append(f"Macro Forecast has {len(data_rows)} rows, expected 4")
            else:
                # Check Q1 GDP
                q1 = [r for r in data_rows if r[0] and "Q1" in str(r[0])]
                if q1:
                    if not num_close(q1[0][1], 2.3, 0.2):
                        errors.append(f"Q1 GDP={q1[0][1]}, expected 2.3")
                else:
                    errors.append("Q1 2026 row missing from Macro Forecast")
                # Check Q4 interest rate
                q4 = [r for r in data_rows if r[0] and "Q4" in str(r[0])]
                if q4:
                    if not num_close(q4[0][3], 4.0, 0.2):
                        errors.append(f"Q4 Interest Rate={q4[0][3]}, expected 4.0")

        # Check Stock Holdings sheet
        rows2 = load_sheet_rows(wb, "Stock Holdings")
        if rows2 is None:
            errors.append("Sheet 'Stock Holdings' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(f"Stock Holdings has {len(data_rows2)} rows, expected 5")
            symbols = {str(r[0]).strip().upper() for r in data_rows2 if r[0]}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols:
                    errors.append(f"Symbol {sym} missing from Stock Holdings")
            # Check AMZN price
            amzn = [r for r in data_rows2 if r[0] and str(r[0]).strip().upper() == "AMZN"]
            if amzn:
                if not num_close(amzn[0][2], 218.94, 5.0):
                    errors.append(f"AMZN Price={amzn[0][2]}, expected ~218.94")
            # Check beta column exists
            googl = [r for r in data_rows2 if r[0] and str(r[0]).strip().upper() == "GOOGL"]
            if googl and len(googl[0]) >= 6:
                if not num_close(googl[0][5], 1.112, 0.2):
                    errors.append(f"GOOGL Beta={googl[0][5]}, expected ~1.112")

        # Check Sector Sensitivity sheet
        rows3 = load_sheet_rows(wb, "Sector Sensitivity")
        if rows3 is None:
            errors.append("Sheet 'Sector Sensitivity' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 5:
                errors.append(f"Sector Sensitivity has {len(data_rows3)} rows, expected 5")
            # Check Communication Services -> Overweight
            comm = [r for r in data_rows3 if r[0] and "communication" in str(r[0]).lower()]
            if comm:
                if str(comm[0][2]).strip().lower() != "high":
                    errors.append(f"Communication Services Rate_Sensitivity={comm[0][2]}, expected High")
                if str(comm[0][3]).strip().lower() != "overweight":
                    errors.append(f"Communication Services Action={comm[0][3]}, expected Overweight")
            else:
                errors.append("Communication Services missing from Sector Sensitivity")
            # Check Energy -> Hold
            energy = [r for r in data_rows3 if r[0] and "energy" in str(r[0]).lower()]
            if energy:
                if str(energy[0][3]).strip().lower() != "hold":
                    errors.append(f"Energy Action={energy[0][3]}, expected Hold")
            # Check Healthcare -> Hold
            health = [r for r in data_rows3 if r[0] and "healthcare" in str(r[0]).lower()]
            if health:
                if str(health[0][3]).strip().lower() != "hold":
                    errors.append(f"Healthcare Action={health[0][3]}, expected Hold")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description FROM gcal.events
            WHERE start_datetime::date = '2026-03-30'
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            errors.append("No GCal event found on 2026-03-30")
        else:
            summaries = [r[0].lower() if r[0] else "" for r in rows]
            if not any("portfolio" in s or "rebalancing" in s or "rebalance" in s for s in summaries):
                errors.append(f"No portfolio rebalancing event (found: {[r[0] for r in rows]})")
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%investment_committee@firm.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            errors.append("No email found to investment_committee@firm.com")
        else:
            subjects = [r[0].lower() if r[0] else "" for r in rows]
            if not any("macro" in s or "portfolio" in s or "outlook" in s for s in subjects):
                errors.append(f"Email subject doesn't match expected (found: {[r[0] for r in rows]})")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal event...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

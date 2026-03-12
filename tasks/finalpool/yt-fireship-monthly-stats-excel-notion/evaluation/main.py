"""
Evaluation for yt-fireship-monthly-stats-excel-notion task.

Checks:
1. Fireship_Monthly_Stats.xlsx exists with Monthly_Stats and Summary sheets
2. Monthly_Stats has 13 rows with correct Month/Video_Count/Avg_Views data
3. Summary sheet has correct Label/Value pairs
4. Notion page "Fireship Channel Analysis 2024-2025" exists
5. Email sent to analytics@company.com with correct subject
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Fireship_Monthly_Stats.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Fireship_Monthly_Stats.xlsx")
    if not os.path.exists(xlsx_path):
        record("Fireship_Monthly_Stats.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Fireship_Monthly_Stats.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Monthly_Stats sheet
    if "monthly_stats" not in sheet_names_lower:
        record("Monthly_Stats sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Monthly_Stats sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("monthly_stats")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Monthly_Stats has 13 data rows", len(data_rows) == 13,
               f"Found {len(data_rows)} rows")

        # Check header columns
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            has_month = any("month" in h for h in headers)
            has_count = any("count" in h or "video" in h for h in headers)
            has_views = any("view" in h for h in headers)
            record("Monthly_Stats has correct columns (Month, Video_Count, Avg_Views etc)",
                   has_month and has_count and has_views, f"Headers: {rows[0]}")

        # Check months in range 2024-06 to 2025-06
        all_text = " ".join(str(c) for r in rows for c in r if c)
        has_2024 = "2024" in all_text
        has_2025 = "2025" in all_text
        record("Monthly_Stats contains 2024 and 2025 months", has_2024 and has_2025,
               f"2024:{has_2024}, 2025:{has_2025}")

        # Spot-check a known value: 2025-01 had 9 videos
        jan25_row = None
        for r in data_rows:
            if r and str(r[0]) == "2025-01":
                jan25_row = r
                break
        if jan25_row:
            record("2025-01 has 9 videos", jan25_row[1] == 9,
                   f"Found: {jan25_row[1]}")
        else:
            record("2025-01 row found", False, "Row not found")

    # Check Summary sheet
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        has_total = "total_videos" in all_text2 or "total videos" in all_text2
        has_peak = "peak_month" in all_text2 or "peak month" in all_text2
        has_104 = "104" in all_text2
        record("Summary has Total_Videos label", has_total, "Missing Total_Videos")
        record("Summary has Peak_Month label", has_peak, "Missing Peak_Month")
        record("Summary shows 104 total videos", has_104, "Expected 104 total videos")


def check_notion():
    print("\n=== Check 2: Notion page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, properties FROM notion.pages
        WHERE properties::text ILIKE %s
        AND archived = false AND in_trash = false
    """, ("%Fireship Channel Analysis%",))
    pages = cur.fetchall()
    cur.close()
    conn.close()

    record("Notion page 'Fireship Channel Analysis 2024-2025' exists",
           len(pages) >= 1, f"Found {len(pages)} matching pages")


def check_email():
    print("\n=== Check 3: Email sent ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT m.to_addr, m.subject FROM email.messages m
        JOIN email.sent_log sl ON sl.message_id = m.id
        WHERE m.to_addr::text ILIKE %s
        ORDER BY sl.sent_at DESC LIMIT 5
    """, ("%analytics%",))
    emails = cur.fetchall()

    if not emails:
        # Also check messages directly
        cur.execute("""
            SELECT to_addr, subject FROM email.messages
            WHERE to_addr::text ILIKE %s
            ORDER BY date DESC LIMIT 5
        """, ("%analytics%",))
        emails = cur.fetchall()

    cur.close()
    conn.close()

    record("Email sent to analytics@company.com", len(emails) >= 1,
           f"Found emails: {emails}")

    if emails:
        subject = str(emails[0][1]).lower() if emails[0][1] else ""
        record("Email subject mentions 'Fireship' or 'Channel Analysis'",
               "fireship" in subject or "channel analysis" in subject or "analysis complete" in subject,
               f"Subject: {emails[0][1]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

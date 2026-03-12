"""Evaluation for sf-canvas-skills-gap-analysis."""
import argparse
import os
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl

    path = os.path.join(agent_workspace, "Skills_Gap.xlsx")
    if not os.path.exists(path):
        return ["Skills_Gap.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Sheet 1: Department Skills
        rows = load_sheet_rows(wb, "Department Skills")
        if rows is None:
            errors.append("Sheet 'Department Skills' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 7:
                errors.append(
                    f"Department Skills has {len(data_rows)} rows, expected 7"
                )
            # Check that departments are present
            dept_names = [str(r[0]).strip() for r in data_rows if r[0]]
            for expected in ["Engineering", "Sales", "Finance", "Operations", "Support"]:
                if not any(expected.lower() in d.lower() for d in dept_names):
                    errors.append(f"Department '{expected}' not found")

            # Check that Operations has a non-zero gap score (highest gap)
            ops_rows = [
                r
                for r in data_rows
                if r[0] and "operations" in str(r[0]).lower()
            ]
            if ops_rows and len(ops_rows[0]) > 5:
                gap = ops_rows[0][5]
                if gap is not None:
                    try:
                        if float(gap) <= 0:
                            errors.append(
                                f"Operations gap score should be > 0, got {gap}"
                            )
                    except (ValueError, TypeError):
                        pass

            # Check Engineering has gap = 0 (avg perf = 3.21)
            eng_rows = [
                r
                for r in data_rows
                if r[0] and "engineering" in str(r[0]).lower()
            ]
            if eng_rows and len(eng_rows[0]) > 5:
                gap = eng_rows[0][5]
                if gap is not None:
                    try:
                        if float(gap) > 0.05:
                            errors.append(
                                f"Engineering gap score should be ~0, got {gap}"
                            )
                    except (ValueError, TypeError):
                        pass

        # Sheet 2: Training Mapping
        rows2 = load_sheet_rows(wb, "Training Mapping")
        if rows2 is None:
            errors.append("Sheet 'Training Mapping' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 10:
                errors.append(
                    f"Training Mapping has {len(data_rows2)} rows, expected >= 10"
                )
            # Check some skill areas are present
            skill_names = [str(r[0]).strip().lower() for r in data_rows2 if r[0]]
            for skill in ["python", "sql", "negotiation"]:
                if not any(skill in s for s in skill_names):
                    errors.append(f"Skill area containing '{skill}' not found")

        # Sheet 3: Priority Actions
        rows3 = load_sheet_rows(wb, "Priority Actions")
        if rows3 is None:
            errors.append("Sheet 'Priority Actions' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 7:
                errors.append(
                    f"Priority Actions has {len(data_rows3)} rows, expected 7"
                )
            # Check priority assignments
            high_depts = [
                str(r[0]).strip()
                for r in data_rows3
                if r[0] and len(r) > 2 and r[2] and str(r[2]).strip().lower() == "high"
            ]
            low_depts = [
                str(r[0]).strip()
                for r in data_rows3
                if r[0] and len(r) > 2 and r[2] and str(r[2]).strip().lower() == "low"
            ]
            if len(high_depts) < 3:
                errors.append(
                    f"Expected at least 3 High priority depts, got {len(high_depts)}: {high_depts}"
                )
            if len(low_depts) < 1:
                errors.append(
                    f"Expected at least 1 Low priority dept, got {len(low_depts)}"
                )

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_emails():
    errors = []
    try:
        import psycopg2

        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="postgres",
            password="postgres",
        )
        cur = conn.cursor()
        cur.execute(
            """
            SELECT subject, to_addr FROM email.messages
            WHERE subject ILIKE '%training%' OR subject ILIKE '%recommendation%'
            ORDER BY id DESC LIMIT 10
        """
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 3:
            errors.append(
                f"Expected at least 3 training recommendation emails, found {len(rows)}"
            )
        # Check that emails were sent to department heads with gaps
        all_to = " ".join(str(r[1]) for r in rows).lower()
        gap_dept_emails = ["m.rodriguez", "r.kim", "a.foster"]  # Sales, Operations, Support
        found = sum(1 for e in gap_dept_emails if e in all_to)
        if found < 2:
            errors.append(
                f"Expected emails to at least 2 of Sales/Operations/Support heads, found {found}"
            )
    except Exception as e:
        errors.append(f"Error checking emails: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking emails...")
    errs = check_emails()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

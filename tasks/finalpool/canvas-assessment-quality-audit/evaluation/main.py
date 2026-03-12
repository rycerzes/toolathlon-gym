"""Evaluation for canvas-assessment-quality-audit."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_quiz_data():
    """Get expected quiz data from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name, q.title,
               (SELECT COUNT(*) FROM canvas.quiz_questions qq WHERE qq.quiz_id = q.id) as qcount,
               ROUND(AVG(qs.score), 1) as avg_score,
               q.points_possible
        FROM canvas.quizzes q
        JOIN canvas.courses c ON c.id = q.course_id
        LEFT JOIN canvas.quiz_submissions qs ON qs.quiz_id = q.id
        GROUP BY c.name, q.id, q.title, q.points_possible
        ORDER BY c.name, q.title
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    quiz_data = []
    for course, quiz, qcount, avg_score, pts in rows:
        if pts and float(pts) > 0 and avg_score and qcount > 0:
            difficulty = round(float(avg_score) / (float(pts) * qcount), 3)
        else:
            difficulty = 0
        flagged = difficulty < 0.3 or difficulty > 0.8
        quiz_data.append({
            "course": course, "quiz": quiz, "qcount": qcount,
            "avg_score": float(avg_score) if avg_score else 0,
            "difficulty": difficulty, "flagged": flagged,
        })
    return quiz_data


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Assessment_Quality.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Assessment_Quality.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Assessment_Quality.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_quiz_data()
    total_quizzes = len(expected)
    flagged_count = sum(1 for q in expected if q["flagged"])
    total_questions = sum(q["qcount"] for q in expected)

    # Quiz Overview sheet
    qo_rows = load_sheet_rows(wb, "Quiz Overview")
    if qo_rows is None:
        check("Sheet 'Quiz Overview' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Quiz Overview' exists", True)
        data_rows = qo_rows[1:] if len(qo_rows) > 1 else []
        check(f"Quiz Overview has {total_quizzes} rows",
              abs(len(data_rows) - total_quizzes) <= 2,
              f"Found {len(data_rows)}")

        header = qo_rows[0] if qo_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "quiz", "question_count", "avg_score", "avg_difficulty"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

    # Flagged Items sheet
    fi_rows = load_sheet_rows(wb, "Flagged Items")
    if fi_rows is None:
        check("Sheet 'Flagged Items' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Flagged Items' exists", True)
        data_rows = fi_rows[1:] if len(fi_rows) > 1 else []
        check(f"Flagged Items has ~{flagged_count} rows",
              abs(len(data_rows) - flagged_count) <= 5,
              f"Found {len(data_rows)}, expected ~{flagged_count}")

        # Check Issue column has values
        if data_rows:
            issues = [str(r[4]).strip().lower() if len(r) > 4 and r[4] else "" for r in data_rows[:5]]
            has_issue_types = any("easy" in i or "hard" in i for i in issues)
            check("Flagged items have 'Too Easy' or 'Too Hard' issues",
                  has_issue_types, f"Issues: {issues}")

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        check(f"Total_Quizzes = {total_quizzes}",
              num_close(lookup.get("total_quizzes"), total_quizzes),
              f"Got {lookup.get('total_quizzes')}")
        check(f"Total_Questions = {total_questions}",
              num_close(lookup.get("total_questions"), total_questions, 10),
              f"Got {lookup.get('total_questions')}")
        check(f"Flagged_Quizzes close to {flagged_count}",
              num_close(lookup.get("flagged_quizzes"), flagged_count, 5),
              f"Got {lookup.get('flagged_quizzes')}")


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Assessment_Report.docx")
    if not os.path.isfile(docx_path):
        check("Assessment_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Assessment_Report.docx exists", True)
    check("Word doc has content (> 1KB)", os.path.getsize(docx_path) > 1000,
          f"Size: {os.path.getsize(docx_path)}")

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Report mentions difficulty", "difficult" in all_text, f"Sample: {all_text[:200]}")
        check("Report mentions flagged or quality", "flag" in all_text or "quality" in all_text,
              f"Sample: {all_text[:200]}")
        check("Report has recommendations", "recommend" in all_text,
              f"Sample: {all_text[:200]}")
    except ImportError:
        check("python-docx available", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

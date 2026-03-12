"""Evaluation script for terminal-pw-howtocook-excel-word-gcal."""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check Excel file
    excel_path = os.path.join(agent_workspace, "Weekly_Meal_Budget.xlsx")
    check("Weekly_Meal_Budget.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # Grocery_Prices sheet
        check("Grocery_Prices sheet exists", "Grocery_Prices" in wb.sheetnames)
        if "Grocery_Prices" in wb.sheetnames:
            ws = wb["Grocery_Prices"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Grocery_Prices has >= 15 rows", len(data_rows) >= 15, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Item', 'FreshMart_Price', 'ValueGrocer_Price', 'BulkBarn_Price', 'Cheapest_Store']:
                check(f"Grocery_Prices has {col} column", col.lower() in headers, f"headers: {headers[:6]}")
            # Verify sorted alphabetically
            if len(data_rows) >= 2:
                items = [str(r[0]).lower() for r in data_rows if r[0]]
                check("Grocery_Prices sorted alphabetically", items == sorted(items), f"first items: {items[:5]}")

        # Recipe_Cost_Analysis sheet
        check("Recipe_Cost_Analysis sheet exists", "Recipe_Cost_Analysis" in wb.sheetnames)
        if "Recipe_Cost_Analysis" in wb.sheetnames:
            ws = wb["Recipe_Cost_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recipe_Cost_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Recipe', 'Total_Cost', 'Cost_Per_Serving']:
                check(f"Recipe_Cost_Analysis has {col} column", col.lower() in headers, f"headers: {headers[:6]}")

        # Weekly_Meal_Plan sheet
        check("Weekly_Meal_Plan sheet exists", "Weekly_Meal_Plan" in wb.sheetnames)
        if "Weekly_Meal_Plan" in wb.sheetnames:
            ws = wb["Weekly_Meal_Plan"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Weekly_Meal_Plan has >= 10 rows", len(data_rows) >= 10, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Day', 'Meal', 'Recipe', 'Estimated_Cost']:
                check(f"Weekly_Meal_Plan has {col} column", col.lower() in headers, f"headers: {headers[:5]}")
            # Check all weekdays present
            days = set(str(r[0]).lower() for r in data_rows if r[0])
            for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']:
                check(f"Weekly_Meal_Plan has {day}", day in days, f"found days: {days}")

    # Check Word document
    word_path = os.path.join(agent_workspace, "Cafeteria_Meal_Proposal.docx")
    check("Cafeteria_Meal_Proposal.docx exists", os.path.exists(word_path))
    if os.path.exists(word_path):
        from docx import Document
        doc = Document(word_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Word has substantial content", len(text) > 200, f"text length: {len(text)}")
        check("Word mentions budget", "budget" in text)
        check("Word mentions store names", "valuegr" in text or "bulkbarn" in text or "freshmart" in text)

    # Check terminal script
    check("meal_cost_optimizer.py exists", os.path.exists(os.path.join(agent_workspace, "meal_cost_optimizer.py")))

    # Check calendar events
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime FROM gcal.events WHERE summary ILIKE %s", ('%meal prep%',))
        events = cur.fetchall()
        check("5 meal prep calendar events", len(events) >= 5, f"found {len(events)}")
        if events:
            march_events = [e for e in events if "2026-03" in str(e[2])]
            check("Events in March 2026", len(march_events) >= 5, f"found {len(march_events)} in March")
            # Check descriptions have ingredient info
            desc_with_ingredients = sum(1 for e in events if e[1] and len(e[1]) > 20)
            check("Events have ingredient descriptions", desc_with_ingredients >= 3, f"{desc_with_ingredients} have descriptions")
        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    import psycopg2
    print("\n=== Reverse Validation ===")

    # Excel: no negative cost values
    path = os.path.join(workspace, "Weekly_Meal_Budget.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative cost values in Excel", not has_negative,
              "Found negative cost value")

    # GCal: no meal prep events on weekends
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%meal prep%%'
              AND EXTRACT(DOW FROM start_datetime) IN (0, 6)
        """)
        weekend_count = cur.fetchone()[0]
        check("No meal prep events on weekends", weekend_count == 0,
              f"Found {weekend_count} weekend events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    check_reverse_validation(args.agent_workspace)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

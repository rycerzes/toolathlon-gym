"""Evaluation for terminal-arxiv-yf-excel-word-notion-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

RELEVANT_PAPER_IDS = {"2306.06031", "2304.07619", "2302.14040", "2311.10723"}
NOISE_PAPER_IDS = {"2305.18290", "2307.09288"}
STOCKS = ["GOOGL", "AMZN", "JPM"]


def get_expected_from_db():
    """Query YF schema dynamically for stock price data."""
    defaults = {
        "stock_prices": {},  # symbol -> latest price
    }
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Get latest closing price for each stock
        for symbol in STOCKS:
            cur.execute("""
                SELECT close FROM yf.stock_prices
                WHERE symbol = %s ORDER BY date DESC LIMIT 1
            """, (symbol,))
            row = cur.fetchone()
            if row and row[0]:
                defaults["stock_prices"][symbol] = float(row[0])
        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [WARN] DB query for expected values failed, using defaults: {e}")
    return defaults


EXPECTED = get_expected_from_db()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except:
        return False


def check_excel(ws_path):
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "AI_Investment_Research.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}

    # Portfolio_Holdings sheet
    ph_name = sn.get("portfolio_holdings")
    if ph_name is None:
        check("Portfolio_Holdings sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Portfolio_Holdings sheet exists", True)
        ws = wb[ph_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).lower() if h else "" for h in rows[0]] if rows else []
        data = [r for r in rows[1:] if r and r[0] is not None]
        check("Portfolio_Holdings has 3 rows", len(data) == 3, f"Found {len(data)}")

        symbols_found = {str(r[0]).strip().upper() for r in data}
        check("All 3 stocks present", symbols_found >= {"GOOGL", "AMZN", "JPM"},
              f"Found: {symbols_found}")

        # Check has price column with reasonable values
        price_col = None
        for i, h in enumerate(headers):
            if "price" in h:
                price_col = i
                break
        if price_col is not None:
            prices = [r[price_col] for r in data if r[price_col] is not None]
            check("Prices are populated", len(prices) == 3, f"Prices: {prices}")
            # Validate prices against dynamically queried DB values
            if EXPECTED["stock_prices"]:
                for row in data:
                    sym = str(row[0]).strip().upper()
                    if sym in EXPECTED["stock_prices"] and row[price_col] is not None:
                        expected_price = EXPECTED["stock_prices"][sym]
                        check(f"{sym} price reasonable (~{expected_price:.0f})",
                              num_close(row[price_col], expected_price, tol=expected_price * 0.1),
                              f"Got {row[price_col]}, expected ~{expected_price:.2f}")
        else:
            check("Price column exists", False, f"Headers: {headers}")

    # Research_Papers sheet
    rp_name = sn.get("research_papers")
    if rp_name is None:
        check("Research_Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Research_Papers sheet exists", True)
        ws2 = wb[rp_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Research_Papers has 4 rows (relevant only)", len(data2) == 4,
              f"Found {len(data2)}")

        # Check FinGPT is mentioned
        all_titles = " ".join(str(r[0]) for r in data2).lower()
        check("FinGPT paper listed", "fingpt" in all_titles, f"Titles: {all_titles[:200]}")

        # Check applicable stocks column
        all_stocks_text = " ".join(str(r[-1]) if r[-1] else "" for r in data2).upper()
        check("Applicable stocks mention GOOGL", "GOOGL" in all_stocks_text)
        check("Applicable stocks mention JPM", "JPM" in all_stocks_text)

    # AI_Impact_Assessment sheet
    ai_name = sn.get("ai_impact_assessment")
    if ai_name is None:
        check("AI_Impact_Assessment sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("AI_Impact_Assessment sheet exists", True)
        ws3 = wb[ai_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("AI_Impact_Assessment has 3 rows", len(data3) == 3, f"Found {len(data3)}")

        # Check AI scores and recommendations
        for row in data3:
            stock = str(row[0]).strip().upper()
            score = row[1] if len(row) > 1 else None
            rec = str(row[-1]).lower() if row[-1] else ""
            if stock == "GOOGL":
                check("GOOGL AI score ~9", num_close(score, 9, tol=1), f"Score: {score}")
                check("GOOGL recommendation Overweight", "overweight" in rec, f"Rec: {rec}")
            elif stock == "AMZN":
                check("AMZN AI score ~8", num_close(score, 8, tol=1), f"Score: {score}")
                check("AMZN recommendation Overweight", "overweight" in rec, f"Rec: {rec}")
            elif stock == "JPM":
                check("JPM AI score ~5", num_close(score, 5, tol=1), f"Score: {score}")
                check("JPM recommendation Hold", "hold" in rec, f"Rec: {rec}")

    # Investment_Thesis sheet
    it_name = sn.get("investment_thesis")
    if it_name is None:
        check("Investment_Thesis sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Investment_Thesis sheet exists", True)
        ws4 = wb[it_name]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if r and r[0] is not None]
        check("Investment_Thesis has >= 2 rows", len(data4) >= 2, f"Found {len(data4)}")

        all_themes = " ".join(str(r[0]) for r in data4).lower()
        check("Theme mentions AI", "ai" in all_themes, f"Themes: {all_themes[:200]}")

    wb.close()


def check_word(ws_path):
    print("\n=== Checking Word Document ===")
    path = os.path.join(ws_path, "AI_Markets_Research_Report.docx")
    if not os.path.isfile(path):
        check("Word document exists", False, f"Not found: {path}")
        return
    check("Word document exists", True)

    from docx import Document
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    check("Document title mentions AI and Markets",
          "ai" in full_text[:200] and "market" in full_text[:200])
    check("Document mentions GOOGL", "googl" in full_text)
    check("Document mentions AMZN", "amzn" in full_text)
    check("Document mentions JPM", "jpm" in full_text)
    check("Document has executive summary",
          "executive summary" in full_text or "executive" in full_text)
    check("Document has risk assessment",
          "risk" in full_text and ("assessment" in full_text or "factor" in full_text))
    check("Document mentions research papers",
          "fingpt" in full_text or "language model" in full_text or "research" in full_text)
    check("Document mentions overweight or recommendation",
          "overweight" in full_text or "recommendation" in full_text)
    check("Document length >= 800 chars", len(full_text) >= 800,
          f"Length: {len(full_text)}")


def check_notion():
    print("\n=== Checking Notion Database ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM notion.databases")
    databases = cur.fetchall()

    found_db = None
    for db_id, title_json in databases:
        title_str = ""
        if isinstance(title_json, list):
            for item in title_json:
                if isinstance(item, dict):
                    title_str += item.get("plain_text", "") or item.get("text", {}).get("content", "")
        elif isinstance(title_json, str):
            title_str = title_json
        if "research" in title_str.lower() and "pipeline" in title_str.lower():
            found_db = db_id
            break

    check("Research Pipeline database exists", found_db is not None,
          f"Databases: {[d[1] for d in databases]}")

    if found_db:
        cur.execute(
            "SELECT COUNT(*) FROM notion.pages WHERE parent->>'database_id' = %s AND NOT archived",
            (found_db,)
        )
        page_count = cur.fetchone()[0]
        check("Notion has >= 4 paper entries", page_count >= 4,
              f"Found {page_count} pages")

        # Check properties of pages
        cur.execute(
            "SELECT properties FROM notion.pages WHERE parent->>'database_id' = %s AND NOT archived LIMIT 1",
            (found_db,)
        )
        row = cur.fetchone()
        if row:
            props = row[0] if isinstance(row[0], dict) else json.loads(row[0]) if row[0] else {}
            props_lower = {k.lower(): v for k, v in props.items()}
            has_relevance = any("relevance" in k for k in props_lower)
            has_stock = any("stock" in k for k in props_lower)
            check("Pages have Relevance property", has_relevance, f"Props: {list(props.keys())}")
            check("Pages have Stock_Link property", has_stock, f"Props: {list(props.keys())}")

    cur.close()
    conn.close()


def check_emails():
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Check portfolio team email
    cur.execute("""
        SELECT subject, body_text, to_addr FROM email.messages
        WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
    """)
    sent = cur.fetchall()

    portfolio_email = None
    risk_email = None
    for subj, body, to_addr in sent:
        to_str = json.dumps(to_addr).lower() if to_addr else ""
        subj_lower = (subj or "").lower()
        if "portfolio_team" in to_str or "portfolio" in subj_lower:
            portfolio_email = (subj, body)
        if "risk_committee" in to_str or "risk" in subj_lower:
            risk_email = (subj, body)

    check("Portfolio team email sent", portfolio_email is not None,
          f"Sent emails: {[(s, t) for s, _, t in sent]}")
    if portfolio_email:
        body = (portfolio_email[1] or "").lower()
        check("Portfolio email mentions AI exposure", "ai" in body or "exposure" in body,
              f"Body preview: {body[:200]}")

    check("Risk committee email sent", risk_email is not None,
          f"Sent emails: {[(s, t) for s, _, t in sent]}")
    if risk_email:
        body = (risk_email[1] or "").lower()
        check("Risk email mentions concentration", "concentrat" in body or "sector" in body or "risk" in body,
              f"Body preview: {body[:200]}")

    cur.close()
    conn.close()


def check_terminal_outputs(ws_path):
    print("\n=== Checking Terminal Script Outputs ===")
    pa_path = os.path.join(ws_path, "portfolio_analysis.json")
    if os.path.isfile(pa_path):
        check("portfolio_analysis.json exists", True)
        with open(pa_path) as f:
            try:
                data = json.load(f)
                check("portfolio_analysis has content", len(data) > 0)
            except:
                check("portfolio_analysis is valid JSON", False)
    else:
        check("portfolio_analysis.json exists", False)

    rm_path = os.path.join(ws_path, "research_stock_mapping.json")
    if os.path.isfile(rm_path):
        check("research_stock_mapping.json exists", True)
        with open(rm_path) as f:
            try:
                data = json.load(f)
                check("research_stock_mapping has content", len(data) > 0)
            except:
                check("research_stock_mapping is valid JSON", False)
    else:
        check("research_stock_mapping.json exists", False)


def check_reverse_validation(ws_path):
    """Check noise arxiv papers not in notion tracker and no emails to wrong recipients."""
    print("\n=== Reverse Validation ===")

    # --- Check noise papers not in Notion ---
    noise_titles = ["dragggan", "drag your gan", "interactive point-based",
                    "llama 2", "open foundation and fine-tuned"]
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Find the Research Pipeline database
        cur.execute("SELECT id, title FROM notion.databases")
        databases = cur.fetchall()
        pipeline_db_id = None
        for db_id, title_json in databases:
            title_str = ""
            if isinstance(title_json, list):
                for item in title_json:
                    if isinstance(item, dict):
                        title_str += item.get("plain_text", "") or item.get("text", {}).get("content", "")
            elif isinstance(title_json, str):
                title_str = title_json
            if "research" in title_str.lower() and "pipeline" in title_str.lower():
                pipeline_db_id = db_id
                break

        if pipeline_db_id:
            cur.execute(
                "SELECT properties FROM notion.pages WHERE parent->>'database_id' = %s AND NOT archived",
                (pipeline_db_id,))
            pages = cur.fetchall()
            all_page_text = ""
            for (props,) in pages:
                all_page_text += json.dumps(props).lower() + " "

            no_noise_notion = not any(nt in all_page_text for nt in noise_titles)
            check("No noise arxiv papers in Notion tracker (DragGAN, Llama 2)",
                  no_noise_notion,
                  f"Found noise paper in Research Pipeline pages")
        else:
            check("No noise arxiv papers in Notion tracker", True, "No Research Pipeline DB to check")

        # --- Check no emails to wrong recipients ---
        noise_recipients = ["team@firm.com", "office@firm.com", "social@firm.com", "admin@firm.com"]

        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
        """)
        sent_emails = cur.fetchall()
        all_to = " ".join(json.dumps(to).lower() for _, to in sent_emails if to)
        no_noise_email = not any(na in all_to for na in noise_recipients)
        check("No emails sent to noise recipients (team@firm, office@firm, social@firm)",
              no_noise_email,
              f"Sent to: {all_to[:200]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))

    # --- Check noise papers not in Excel ---
    path = os.path.join(ws_path, "AI_Investment_Research.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        all_text = ""
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                all_text += " ".join(str(c) for c in row if c).lower() + " "
        wb.close()

        no_noise_excel = not any(nt in all_text for nt in noise_titles)
        check("No noise arxiv papers in Excel (DragGAN, Llama 2)",
              no_noise_excel,
              "Found noise paper content in Excel workbook")

        # Check noise paper IDs not present
        no_noise_ids = not any(nid in all_text for nid in NOISE_PAPER_IDS)
        check("No noise paper IDs in Excel (2305.18290, 2307.09288)",
              no_noise_ids,
              "Found noise paper ID in Excel")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-ARXIV-YF-EXCEL-WORD-NOTION-EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_emails()
    check_terminal_outputs(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

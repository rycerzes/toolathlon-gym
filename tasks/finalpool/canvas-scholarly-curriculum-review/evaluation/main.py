"""Evaluation for canvas-scholarly-curriculum-review."""
import argparse
import os
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl

    path = os.path.join(agent_workspace, "Curriculum_Review.xlsx")
    if not os.path.exists(path):
        return ["Curriculum_Review.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Sheet 1: Course Compliance
        rows = load_sheet_rows(wb, "Course Compliance")
        if rows is None:
            errors.append("Sheet 'Course Compliance' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 20:
                errors.append(
                    f"Course Compliance has {len(data_rows)} rows, expected ~22"
                )
            # Check that Compliant column exists with Yes/No values
            compliant_vals = []
            for r in data_rows:
                if len(r) > 7 and r[7]:
                    compliant_vals.append(str(r[7]).strip().lower())
            has_yes = any(v == "yes" for v in compliant_vals)
            has_no = any(v == "no" for v in compliant_vals)
            if not has_yes:
                errors.append("No courses marked as Compliant='Yes'")
            if not has_no:
                errors.append("No courses marked as Compliant='No'")

            # Check a specific course: Foundations of Finance (Fall 2013) should be Non-compliant
            fof_rows = [
                r
                for r in data_rows
                if r[0] and "foundations of finance" in str(r[0]).lower() and "fall 2013" in str(r[0]).lower()
            ]
            if fof_rows and len(fof_rows[0]) > 7:
                val = str(fof_rows[0][7]).strip().lower()
                if val != "no":
                    errors.append(
                        f"Foundations of Finance (Fall 2013) should be Non-compliant, got '{val}'"
                    )

            # Check student count for Applied Analytics Fall 2013 (~383)
            aa_rows = [
                r
                for r in data_rows
                if r[0] and "applied analytics" in str(r[0]).lower() and "fall 2013" in str(r[0]).lower()
            ]
            if aa_rows and len(aa_rows[0]) > 4:
                if not nums_close(aa_rows[0][4], 383, abs_tol=50):
                    errors.append(
                        f"Applied Analytics Fall 2013 student count {aa_rows[0][4]}, expected ~383"
                    )

        # Sheet 2: Literature Support
        rows2 = load_sheet_rows(wb, "Literature Support")
        if rows2 is None:
            errors.append("Sheet 'Literature Support' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(
                    f"Literature Support has {len(data_rows2)} rows, expected >= 5"
                )
            # Check topics coverage
            topics_found = [str(r[0]).strip().lower() for r in data_rows2 if r[0]]
            for keyword in ["active", "assessment", "online", "curriculum", "engagement"]:
                if not any(keyword in t for t in topics_found):
                    errors.append(
                        f"Literature Support missing topic containing '{keyword}'"
                    )

        # Sheet 3: Summary
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 5:
                errors.append(
                    f"Summary has {len(data_rows3)} rows, expected >= 5"
                )
            # Check total courses
            total_row = [
                r for r in data_rows3 if r[0] and "total" in str(r[0]).lower() and "course" in str(r[0]).lower()
            ]
            if total_row and len(total_row[0]) > 1:
                if not nums_close(total_row[0][1], 22, abs_tol=2):
                    errors.append(
                        f"Total courses {total_row[0][1]}, expected ~22"
                    )

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_notion():
    errors = []
    try:
        import psycopg2

        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="postgres",
            password="postgres",
        )
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, properties::text FROM notion.pages
            WHERE properties::text ILIKE '%accreditation%'
            ORDER BY created_time DESC LIMIT 5
        """
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 1:
            # Check if there are any pages with accreditation content in blocks
            conn2 = psycopg2.connect(
                host=os.environ.get("PGHOST", "localhost"),
                port=5432,
                dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
                user="postgres",
                password="postgres",
            )
            cur2 = conn2.cursor()
            cur2.execute("SELECT COUNT(*) FROM notion.pages")
            page_count = cur2.fetchone()[0]
            cur2.close()
            conn2.close()
            if page_count <= 1:
                errors.append(
                    "No Notion page found with 'Accreditation Review Report' title"
                )
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

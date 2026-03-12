"""Evaluation for sf-sales-quarterly-review-gcal."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    errors = []
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl not installed")
        return errors

    agent_file = os.path.join(agent_workspace, "Sales_Quarterly_2025.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Quarterly_2025.xlsx")

    if not os.path.exists(agent_file):
        errors.append("Sales_Quarterly_2025.xlsx not found in agent workspace")
        return errors

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Quarterly Performance sheet
    a_rows = load_sheet_rows(agent_wb, "Quarterly Performance")
    g_rows = load_sheet_rows(gt_wb, "Quarterly Performance")
    if a_rows is None:
        errors.append("Sheet 'Quarterly Performance' not found in agent output")
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        g_data = [r for r in (g_rows[1:] if g_rows and len(g_rows) > 1 else []) if r and r[0] is not None]

        if len(a_data) < 4:
            errors.append(f"Quarterly Performance: expected 4 data rows, got {len(a_data)}")
        else:
            # Build lookup by quarter key
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
            for g_row in g_data:
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing quarter row: {g_row[0]}")
                    continue
                # Revenue col 1
                if len(a_row) > 1 and not num_close(a_row[1], g_row[1], 1.0):
                    errors.append(f"{g_row[0]} Revenue: got {a_row[1]}, expected {g_row[1]} (tol=1.0)")
                # Order_Count col 2
                if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 5):
                    errors.append(f"{g_row[0]} Order_Count: got {a_row[2]}, expected {g_row[2]} (tol=5)")
                # Avg_Order_Value col 3
                if len(a_row) > 3 and not num_close(a_row[3], g_row[3], 1.0):
                    errors.append(f"{g_row[0]} Avg_Order_Value: got {a_row[3]}, expected {g_row[3]} (tol=1.0)")
                # QoQ_Change_Pct col 4 (skip Q1 which is None)
                if len(a_row) > 4 and g_row[4] is not None:
                    if not num_close(a_row[4], g_row[4], 0.5):
                        errors.append(f"{g_row[0]} QoQ_Change_Pct: got {a_row[4]}, expected {g_row[4]} (tol=0.5)")

    # Check Annual Summary sheet
    a_sum = load_sheet_rows(agent_wb, "Annual Summary")
    g_sum = load_sheet_rows(gt_wb, "Annual Summary")
    if a_sum is None:
        errors.append("Sheet 'Annual Summary' not found in agent output")
    else:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}
        g_sum_data = {str(r[0]).strip().lower(): r[1] for r in (g_sum[1:] if g_sum and len(g_sum) > 1 else []) if r and r[0]}

        # Annual_Revenue
        if "annual_revenue" in g_sum_data:
            av = a_sum_data.get("annual_revenue")
            if av is None:
                errors.append("Annual Summary missing Annual_Revenue row")
            elif not num_close(av, g_sum_data["annual_revenue"], 1.0):
                errors.append(f"Annual_Revenue: got {av}, expected {g_sum_data['annual_revenue']} (tol=1.0)")

        # Best_Quarter
        if "best_quarter" in g_sum_data:
            bq = a_sum_data.get("best_quarter")
            if bq is None:
                errors.append("Annual Summary missing Best_Quarter row")
            elif not str_match(bq, g_sum_data["best_quarter"]):
                errors.append(f"Best_Quarter: got '{bq}', expected '{g_sum_data['best_quarter']}'")

        # Worst_Quarter
        if "worst_quarter" in g_sum_data:
            wq = a_sum_data.get("worst_quarter")
            if wq is None:
                errors.append("Annual Summary missing Worst_Quarter row")
            elif not str_match(wq, g_sum_data["worst_quarter"]):
                errors.append(f"Worst_Quarter: got '{wq}', expected '{g_sum_data['worst_quarter']}'")

        # Total_Orders
        if "total_orders" in g_sum_data:
            to = a_sum_data.get("total_orders")
            if to is None:
                errors.append("Annual Summary missing Total_Orders row")
            elif not num_close(to, g_sum_data["total_orders"], 5):
                errors.append(f"Total_Orders: got {to}, expected {g_sum_data['total_orders']} (tol=5)")

    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, summary, start_datetime
            FROM gcal.events
            WHERE (LOWER(summary) LIKE '%planning%' OR LOWER(summary) LIKE '%sales%')
            AND start_datetime >= '2026-01-01T00:00:00'
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()
        if len(events) < 4:
            errors.append(f"Expected at least 4 planning/sales gcal events, found {len(events)}")
    except Exception as e:
        errors.append(f"GCal DB check error: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, from_addr, to_addr
            FROM email.messages
            WHERE LOWER(subject) LIKE '%2025%'
            AND LOWER(subject) LIKE '%sales%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if not emails:
            errors.append("No email with '2025' and 'sales' in subject found")
        else:
            found_to = False
            for em in emails:
                to_str = str(em[2]).lower()
                if "sales.leadership" in to_str:
                    found_to = True
                    break
            if not found_to:
                errors.append("No email sent to sales.leadership@company.com")
    except Exception as e:
        errors.append(f"Email DB check error: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    docx_path = os.path.join(agent_workspace, "Sales_Executive_Summary.docx")
    if not os.path.exists(docx_path):
        errors.append("Sales_Executive_Summary.docx not found")
        return errors
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        if len(text.strip()) < 50:
            errors.append("Sales_Executive_Summary.docx has too little content")
        for kw in ["q4", "q1"]:
            if kw not in text:
                errors.append(f"Sales_Executive_Summary.docx missing keyword: {kw}")
    except ImportError:
        if os.path.getsize(docx_path) < 100:
            errors.append("Sales_Executive_Summary.docx too small")
    except Exception as e:
        errors.append(f"Error reading Sales_Executive_Summary.docx: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []

    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace, gt_dir)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    print("\n=== Checking GCal Events ===")
    gcal_errors = check_gcal()
    if gcal_errors:
        for e in gcal_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(gcal_errors)
    else:
        print("  [PASS] GCal check passed")

    print("\n=== Checking Email ===")
    email_errors = check_email()
    if email_errors:
        for e in email_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(email_errors)
    else:
        print("  [PASS] Email check passed")

    print("\n=== Checking Word Document ===")
    word_errors = check_word(args.agent_workspace)
    if word_errors:
        for e in word_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(word_errors)
    else:
        print("  [PASS] Word check passed")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"errors": all_errors, "success": len(all_errors) == 0}, f, indent=2)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

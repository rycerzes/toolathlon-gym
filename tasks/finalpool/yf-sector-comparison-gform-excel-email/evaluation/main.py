"""Evaluation for yf-sector-comparison-gform-excel-email.

Checks:
1. Sector_Comparison.xlsx with Metrics sheet (5 rows) and Sector_Summary sheet (5 rows)
2. GForm "Investment Preference Survey" with 4 questions
3. Email to investors@fund.example.com with "Sector Comparison" in subject
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0

SYMBOLS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]
SECTORS = ["Communication Services", "Consumer Cyclical", "Financial Services", "Healthcare", "Energy"]

# Expected market caps in billions from DB
EXPECTED_MKTCAP_B = {
    "GOOGL": 3639.75, "AMZN": 2350.30, "JPM": 791.71, "JNJ": 577.48, "XOM": 628.18,
}
EXPECTED_PRICES = {
    "GOOGL": 300.88, "AMZN": 218.94, "JPM": 293.55, "JNJ": 239.63, "XOM": 150.76,
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol_pct=5.0):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(b) < 1e-6:
        return abs(a) < 0.01
    return abs(a - b) / abs(b) * 100 <= tol_pct


def check_excel(agent_ws):
    print("\n=== Check 1: Sector_Comparison.xlsx ===")
    path = os.path.join(agent_ws, "Sector_Comparison.xlsx")
    check("File Sector_Comparison.xlsx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel is readable", False, str(e))
        return

    # Check Metrics sheet
    metrics_ws = None
    for sname in wb.sheetnames:
        if "metric" in sname.lower():
            metrics_ws = wb[sname]
            break
    check("Sheet 'Metrics' exists", metrics_ws is not None, f"Sheets: {wb.sheetnames}")

    if metrics_ws is not None:
        rows = list(metrics_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Metrics sheet has 5 rows", len(non_empty) == 5, f"Got {len(non_empty)}")

        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        for sym in SYMBOLS:
            check(f"Metrics contains symbol {sym}", sym in all_text)

        # Verify market cap values are in billions (should be 577-3640 range)
        for row in non_empty:
            row_text = " ".join(str(c) for c in row if c is not None)
            for sym in SYMBOLS:
                if sym in row_text:
                    exp_mktcap = EXPECTED_MKTCAP_B[sym]
                    found_mktcap = False
                    for c in row:
                        try:
                            fval = float(c)
                            if num_close(fval, exp_mktcap, tol_pct=5.0):
                                found_mktcap = True
                                break
                        except (TypeError, ValueError):
                            pass
                    check(f"Metrics {sym} market cap ~= {exp_mktcap}B", found_mktcap,
                          f"Row: {[str(x)[:20] for x in row[:8]]}")

    # Check Sector_Summary sheet
    summary_ws = None
    for sname in wb.sheetnames:
        if "sector" in sname.lower() and "summary" in sname.lower():
            summary_ws = wb[sname]
            break
        elif "summary" in sname.lower():
            summary_ws = wb[sname]
    check("Sheet 'Sector_Summary' exists", summary_ws is not None, f"Sheets: {wb.sheetnames}")

    if summary_ws is not None:
        rows = list(summary_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Sector_Summary sheet has 5 rows", len(non_empty) == 5, f"Got {len(non_empty)}")

        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        check("Sector_Summary has price assessment (Above_Avg or Below_Avg)",
              "avg" in all_text.lower() or "above" in all_text.lower() or "below" in all_text.lower(),
              f"Content: {all_text[:200]}")


def check_gform():
    print("\n=== Check 2: Google Form ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    check("At least one Google Form exists", len(forms) > 0,
          "No forms found in gform.forms")

    found_form_id = None
    for form_id, title in forms:
        if "investment" in (title or "").lower() or "preference" in (title or "").lower() or "sector" in (title or "").lower():
            found_form_id = form_id
            break
    if found_form_id is None and forms:
        found_form_id = forms[0][0]

    check("Form with 'Investment' or 'Preference' or 'Sector' in title",
          found_form_id is not None,
          f"Forms: {[(str(r[0])[:20], r[1]) for r in forms]}")

    if found_form_id:
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (found_form_id,))
        q_count = cur.fetchone()[0]
        check("Form has exactly 4 questions", q_count == 4,
              f"Got {q_count} questions")

        cur.execute("SELECT title, config FROM gform.questions WHERE form_id = %s ORDER BY position",
                    (found_form_id,))
        questions = cur.fetchall()
        q_titles = [q[0] for q in questions]
        q_text = " ".join(q_titles).lower()
        check("Form has sector preference question",
              "sector" in q_text, f"Questions: {q_titles}")
        check("Form has risk tolerance question",
              "risk" in q_text, f"Questions: {q_titles}")
        check("Form has investment horizon question",
              "horizon" in q_text or "term" in q_text, f"Questions: {q_titles}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%sector%comparison%'
           OR subject ILIKE '%sector comparison%'
           OR to_addr::text ILIKE '%investors@fund%'
        LIMIT 10
    """)
    rows = cur.fetchall()
    check("Email with 'Sector Comparison' in subject found",
          len(rows) > 0, "No matching email found")

    if rows:
        to_addrs = [str(r[1]) for r in rows]
        check("Email sent to investors@fund.example.com",
              any("investors" in addr for addr in to_addrs),
              f"To addresses: {to_addrs}")
        bodies = [str(r[2] or "").lower() for r in rows]
        check("Email body mentions market cap or sector data",
              any("market cap" in b or "sector" in b or "googl" in b or "amzn" in b for b in bodies),
              f"Body: {bodies[0][:200] if bodies else ''}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: yf-sector-comparison-gform-excel-email ===")

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"pass": PASS_COUNT, "fail": FAIL_COUNT}, f)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

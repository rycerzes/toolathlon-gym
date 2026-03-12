"""
Evaluation for howtocook-grocery-budget-planner task.
Checks Excel, Google Sheet, and Calendar events.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Meal_Budget.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Weekly Plan
    wp_rows = load_sheet_rows(wb, "Weekly Plan") or load_sheet_rows(wb, "Weekly_Plan")
    if wp_rows is None:
        record("Sheet 'Weekly Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Weekly Plan' exists", True)
        data = [r for r in wp_rows[1:] if r and r[0] is not None]
        record("Weekly Plan has >= 21 rows (7 days x 3 meals)", len(data) >= 21,
               f"Found {len(data)}")

        cost_col = find_col(wp_rows[0], ["Cost", "cost", "Estimated_Cost"])
        if cost_col is not None:
            total = 0
            for r in data:
                if cost_col < len(r) and r[cost_col] is not None:
                    try:
                        total += float(r[cost_col])
                    except (TypeError, ValueError):
                        pass
            record("Total cost <= $150", total <= 150, f"Total: ${total:.2f}")

    # Shopping List
    sl_rows = load_sheet_rows(wb, "Shopping List") or load_sheet_rows(wb, "Shopping_List")
    if sl_rows is None:
        record("Sheet 'Shopping List' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Shopping List' exists", True)
        data = [r for r in sl_rows[1:] if r and r[0] is not None]
        record("Shopping List has >= 5 ingredients", len(data) >= 5, f"Found {len(data)}")

    # Budget Summary
    bs_rows = load_sheet_rows(wb, "Budget Summary") or load_sheet_rows(wb, "Budget_Summary")
    if bs_rows is None:
        record("Sheet 'Budget Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Budget Summary' exists", True)
        metrics = {}
        for row in bs_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        twc_key = next((k for k in metrics if "total" in k and "cost" in k), None)
        if twc_key:
            try:
                cost = float(metrics[twc_key])
                record("Total_Weekly_Cost <= 150", cost <= 150, f"Got ${cost}")
            except (TypeError, ValueError):
                record("Total_Weekly_Cost is numeric", False)

        br_key = next((k for k in metrics if "budget" in k and "remain" in k), None)
        if br_key:
            try:
                rem = float(metrics[br_key])
                record("Budget_Remaining >= 0", rem >= 0, f"Got ${rem}")
            except (TypeError, ValueError):
                record("Budget_Remaining is numeric", False)

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Meal_Budget.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()

    return True


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()
    record("Google Sheet created", len(sheets) >= 1, f"Found {len(sheets)}")

    if sheets:
        has_meal = any("meal" in str(s[1]).lower() for s in sheets)
        record("Sheet title mentions meal plan", has_meal,
               f"Titles: {[s[1] for s in sheets]}")

        cur.execute("SELECT COUNT(*) FROM gsheet.cells")
        count = cur.fetchone()[0]
        record("Google Sheet has data", count >= 10, f"Found {count} cells")

    cur.close()
    conn.close()
    return True


def check_calendar():
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime FROM gcal.events")
    events = cur.fetchall()
    record("Calendar events created", len(events) >= 2, f"Found {len(events)}")

    if events:
        summaries = [str(e[1]).lower() for e in events]
        has_shopping = any("shopping" in s or "grocery" in s for s in summaries)
        record("Shopping events present", has_shopping, f"Summaries: {summaries}")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gsheet()
    check_calendar()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

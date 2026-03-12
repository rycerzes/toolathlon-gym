"""
Evaluation for howtocook-nutrition-benchmark task.
Checks Excel and Word document structure. Recipe names are dynamic (from HowToCook MCP).
"""
import argparse
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Nutrition_Benchmark.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Recipe Analysis
    ra_rows = load_sheet_rows(wb, "Recipe Analysis") or load_sheet_rows(wb, "Recipe_Analysis")
    if ra_rows is None:
        record("Sheet 'Recipe Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Sheet 'Recipe Analysis' exists", True)

    header = ra_rows[0]
    data = [r for r in ra_rows[1:] if r and r[0] is not None and str(r[0]).strip()]
    record("Recipe Analysis has >= 8 rows", len(data) >= 8, f"Found {len(data)}")

    cal_col = find_col(header, ["Est_Calories", "Est Calories", "Estimated_Calories", "Calories"])
    if cal_col is not None:
        for r in data[:8]:
            if cal_col < len(r) and r[cal_col] is not None:
                try:
                    cal = float(r[cal_col])
                    record(f"Calories {cal} in range", 50 <= cal <= 3000,
                           f"Got {cal} for {r[0]}")
                    break  # Just check first valid one
                except (TypeError, ValueError):
                    pass

    meets_col = find_col(header, ["Meets_Guidelines", "Meets Guidelines", "Guidelines"])
    if meets_col is not None:
        vals = set()
        for r in data:
            if meets_col < len(r) and r[meets_col] is not None:
                vals.add(str(r[meets_col]).strip().lower())
        has_yes_no = "yes" in vals or "no" in vals
        record("Meets_Guidelines has Yes/No values", has_yes_no, f"Values: {vals}")
    else:
        record("Meets_Guidelines column exists", False, f"Header: {header}")

    # Meal Plan Suggestions
    mp_rows = load_sheet_rows(wb, "Meal Plan Suggestions") or load_sheet_rows(wb, "Meal_Plan_Suggestions")
    if mp_rows is None:
        record("Sheet 'Meal Plan Suggestions' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Meal Plan Suggestions' exists", True)
        data2 = [r for r in mp_rows[1:] if r and r[0] is not None]
        record("Meal Plan has >= 9 rows (3 days x 3 meals)", len(data2) >= 9,
               f"Found {len(data2)}")

    # Summary
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        tp_key = next((k for k in metrics if "total" in k and "recip" in k and "analyz" in k), None)
        if tp_key:
            try:
                record("Total_Recipes >= 8", float(metrics[tp_key]) >= 8, f"Got {metrics[tp_key]}")
            except (TypeError, ValueError):
                record("Total_Recipes is numeric", False)
        else:
            record("Total_Recipes_Analyzed metric exists", False, f"Keys: {list(metrics.keys())}")

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Nutrition_Benchmark.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()

    return True


def check_word(workspace):
    print("\n=== Checking Word Document ===")
    path = os.path.join(workspace, "Nutrition_Report.docx")
    if not os.path.isfile(path):
        record("Word document exists", False, f"Not found: {path}")
        return False
    record("Word document exists", True)

    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join(p.text for p in doc.paragraphs).lower()

        record("Report has substantial content", len(text) > 300, f"Only {len(text)} chars")
        record("Mentions nutrition/calories", "calori" in text or "nutrition" in text)
        record("Mentions benchmark/guideline", "benchmark" in text or "guideline" in text)
        record("Mentions meal plan", "meal" in text and "plan" in text)
        return True
    except Exception as e:
        record("Word readable", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_word(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

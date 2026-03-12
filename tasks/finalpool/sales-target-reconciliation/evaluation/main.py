"""
Evaluation for sales-target-reconciliation task.
Compares agent's Q1_Sales_Review.xlsx against groundtruth.
"""
import argparse
import os
import sys

import openpyxl


def load_sheet_data(wb, sheet_name):
    """Load all rows from a sheet as list of lists."""
    # Case-insensitive sheet name lookup
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    sheet_name = matched
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_regional_performance(agent_rows, gt_rows):
    """Check Sheet 1: Regional Performance."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Regional Performance' not found"]

    # Skip header row
    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    for i, (a_row, g_row) in enumerate(zip(agent_data, gt_data)):
        region_name = g_row[0]
        # Check region name (case-insensitive)
        if str(a_row[0]).strip().lower() != str(g_row[0]).strip().lower():
            errors.append(f"Row {i+1}: Region mismatch: '{a_row[0]}' vs '{g_row[0]}'")
            continue

        # Check numeric fields with tolerance
        field_names = ["Target", "Actual", "Variance", "Variance_Pct", "Industry_Benchmark"]
        for j, fname in enumerate(field_names, start=1):
            try:
                a_val = float(a_row[j]) if a_row[j] is not None else None
                g_val = float(g_row[j])
                if a_val is None:
                    errors.append(f"{region_name}.{fname}: missing value")
                elif abs(a_val - g_val) > 1.0:  # tolerance of $1 or 1%
                    errors.append(f"{region_name}.{fname}: {a_val} vs expected {g_val}")
            except (ValueError, TypeError):
                errors.append(f"{region_name}.{fname}: invalid value '{a_row[j]}'")

        # Check vs_Benchmark (case-insensitive)
        a_vs = str(a_row[6]).strip().lower() if a_row[6] else ""
        g_vs = str(g_row[6]).strip().lower()
        if a_vs != g_vs:
            errors.append(f"{region_name}.vs_Benchmark: '{a_row[6]}' vs expected '{g_row[6]}'")

    return len(errors) == 0, errors


def check_top_products(agent_rows, gt_rows):
    """Check Sheet 2: Top Products."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Top Products' not found"]

    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    # Group by region and check top 3 revenue values match
    from collections import defaultdict
    agent_by_region = defaultdict(list)
    gt_by_region = defaultdict(list)

    for row in agent_data:
        if row[0]:
            agent_by_region[str(row[0]).strip()].append(row)
    for row in gt_data:
        if row[0]:
            gt_by_region[str(row[0]).strip()].append(row)

    for region in gt_by_region:
        # Case-insensitive region lookup
        agent_region_key = None
        for k in agent_by_region:
            if k.lower() == region.lower():
                agent_region_key = k
                break
        if agent_region_key is None:
            errors.append(f"Missing region in Top Products: {region}")
            continue
        a_products = agent_by_region[agent_region_key]
        g_products = gt_by_region[region]
        if len(a_products) != len(g_products):
            errors.append(f"{region}: product count {len(a_products)} vs expected {len(g_products)}")
            continue
        for k, (a_p, g_p) in enumerate(zip(a_products, g_products)):
            try:
                a_rev = float(a_p[3]) if a_p[3] is not None else 0
                g_rev = float(g_p[3])
                if abs(a_rev - g_rev) > 1.0:
                    errors.append(f"{region} product #{k+1}: revenue {a_rev} vs expected {g_rev}")
            except (ValueError, TypeError):
                errors.append(f"{region} product #{k+1}: invalid revenue '{a_p[3]}'")

    return len(errors) == 0, errors


def check_summary(agent_rows, gt_rows):
    """Check Sheet 3: Summary."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Summary' not found"]

    if len(agent_rows) < 7:
        errors.append(f"Summary has only {len(agent_rows)} rows, expected 7")
        return False, errors

    gt_dict = {str(r[0]).strip(): r[1] for r in gt_rows if r[0]}
    agent_dict = {str(r[0]).strip().lower(): r[1] for r in agent_rows if r[0]}

    for key, expected in gt_dict.items():
        if key.lower() not in agent_dict:
            errors.append(f"Missing key: {key}")
            continue
        try:
            a_val = float(agent_dict[key.lower()])
            g_val = float(expected)
            if abs(a_val - g_val) > 1.0:
                errors.append(f"{key}: {a_val} vs expected {g_val}")
        except (ValueError, TypeError):
            errors.append(f"{key}: invalid value '{agent_dict[key]}'")

    return len(errors) == 0, errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Resolve paths
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    agent_file = os.path.join(args.agent_workspace, "Q1_Sales_Review.xlsx") if args.agent_workspace else None
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_file = os.path.join(gt_dir, "Q1_Sales_Review.xlsx")

    if not agent_file or not os.path.exists(agent_file):
        print(f"FAIL: Agent output file not found: {agent_file}")
        sys.exit(1)

    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth file not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check 1: Regional Performance
    print("[1/3] Checking Regional Performance ...")
    a_rp = load_sheet_data(agent_wb, "Regional Performance")
    g_rp = load_sheet_data(gt_wb, "Regional Performance")
    rp_ok, rp_errors = check_regional_performance(a_rp, g_rp)
    if rp_ok:
        print("  PASS")
    else:
        for e in rp_errors:
            print(f"  ERROR: {e}")

    # Check 2: Top Products
    print("[2/3] Checking Top Products ...")
    a_tp = load_sheet_data(agent_wb, "Top Products")
    g_tp = load_sheet_data(gt_wb, "Top Products")
    tp_ok, tp_errors = check_top_products(a_tp, g_tp)
    if tp_ok:
        print("  PASS")
    else:
        for e in tp_errors:
            print(f"  ERROR: {e}")

    # Check 3: Summary
    print("[3/3] Checking Summary ...")
    a_sm = load_sheet_data(agent_wb, "Summary")
    g_sm = load_sheet_data(gt_wb, "Summary")
    sm_ok, sm_errors = check_summary(a_sm, g_sm)
    if sm_ok:
        print("  PASS")
    else:
        for e in sm_errors:
            print(f"  ERROR: {e}")

    overall = rp_ok and tp_ok and sm_ok
    print(f"\n=== RESULT: {'PASS' if overall else 'FAIL'} ===")
    print(f"Regional Performance: {'PASS' if rp_ok else 'FAIL'}")
    print(f"Top Products: {'PASS' if tp_ok else 'FAIL'}")
    print(f"Summary: {'PASS' if sm_ok else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

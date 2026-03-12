"""
Evaluation for yf-financial-health-gsheet task.
Checks:
1. Local Excel file: financial_health_report.xlsx with 3 sheets (Income Analysis, Health Assessment, Summary)
2. Google Sheet: gsheet tables in DB have correct spreadsheet, sheet, and cell data
"""
import os
import sys
import json
from argparse import ArgumentParser
from datetime import datetime

import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONN = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

SYMBOLS = ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]

# Credit ratings come from the mock web portal (http://localhost:PORT).
# Cannot be queried from PostgreSQL at evaluation time because yf schema
# has no credit rating table. Values must match preprocess mock page data.
CREDIT_RATINGS = {
    "AMZN": ("AA", "Stable"),
    "GOOGL": ("AA+", "Stable"),
    "JNJ": ("AAA", "Positive"),
    "JPM": ("A+", "Stable"),
    "XOM": ("AA-", "Negative"),
}

PCT_TOLERANCE = 1.0      # absolute tolerance for percentage values
MONEY_TOLERANCE = 5.0    # absolute tolerance for revenue/income in millions


# ---------- Get expected values from DB ----------
def get_expected_data():
    """Query Yahoo Finance DB and compute all expected values."""
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT ON (symbol) symbol, period_end, data
        FROM yf.financial_statements
        WHERE stmt_type = 'income_stmt' AND freq = 'annual'
          AND symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM')
        ORDER BY symbol, period_end DESC
    """)
    income_data = {}
    for row in cur.fetchall():
        sym = row[0]
        data = row[2] if isinstance(row[2], dict) else json.loads(row[2])
        income_data[sym] = data

    cur.execute("""
        SELECT symbol, data->>'shortName'
        FROM yf.stock_info
        WHERE symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM')
    """)
    company_names = {row[0]: row[1] for row in cur.fetchall()}
    cur.close()
    conn.close()

    income_rows = []
    for sym in SYMBOLS:
        data = income_data[sym]
        rev = data.get("Total Revenue")
        ni = data.get("Net Income")
        gp = data.get("Gross Profit")
        oi = data.get("Operating Income")

        rev_m = round(rev / 1e6, 2) if rev is not None else None
        ni_m = round(ni / 1e6, 2) if ni is not None else None
        gp_m = round(gp / 1e6, 2) if gp is not None else None
        oi_m = round(oi / 1e6, 2) if oi is not None else None

        pm = round(ni / rev * 100, 2) if (rev and ni) else None
        om = round(oi / rev * 100, 2) if (rev and oi) else None

        health = classify_margin(pm) if pm is not None else "N/A"

        income_rows.append({
            "Symbol": sym,
            "Company_Name": company_names.get(sym, sym),
            "Total_Revenue": rev_m,
            "Net_Income": ni_m,
            "Gross_Profit": gp_m,
            "Operating_Income": oi_m,
            "Profit_Margin_Pct": pm,
            "Operating_Margin_Pct": om,
            "Credit_Rating": CREDIT_RATINGS[sym][0],
            "Rating_Outlook": CREDIT_RATINGS[sym][1],
            "Margin_Health": health,
        })

    # Summary
    highest_rev_sym = max(income_rows, key=lambda r: r["Total_Revenue"] or 0)["Symbol"]
    highest_ni_sym = max(income_rows, key=lambda r: r["Net_Income"] or 0)["Symbol"]
    best_pm_sym = max(income_rows, key=lambda r: r["Profit_Margin_Pct"] or 0)["Symbol"]
    avg_pm = round(sum(r["Profit_Margin_Pct"] for r in income_rows if r["Profit_Margin_Pct"] is not None) / len(income_rows), 2)
    strong_count = sum(1 for r in income_rows if r["Margin_Health"] == "Strong")
    om_values = [r["Operating_Margin_Pct"] for r in income_rows if r["Operating_Margin_Pct"] is not None]
    avg_om = round(sum(om_values) / len(om_values), 2) if om_values else 0

    summary = {
        "Highest_Revenue_Company": highest_rev_sym,
        "Highest_Net_Income_Company": highest_ni_sym,
        "Best_Profit_Margin_Company": best_pm_sym,
        "Avg_Profit_Margin": avg_pm,
        "Companies_With_Strong_Margin": strong_count,
        "Avg_Operating_Margin": avg_om,
    }

    return income_rows, summary


def classify_margin(pm):
    if pm > 15:
        return "Strong"
    elif pm >= 5:
        return "Adequate"
    else:
        return "Weak"


# ---------- Comparison helpers ----------
def nums_close(expected, actual, tolerance):
    """Check if two numeric values are within absolute tolerance."""
    try:
        e = float(expected)
        a = float(actual)
    except (ValueError, TypeError):
        return False
    return abs(e - a) <= tolerance


def str_match(expected, actual):
    """Case-insensitive string comparison."""
    if expected is None and actual is None:
        return True
    if expected is None or actual is None:
        return False
    return str(expected).strip().lower() == str(actual).strip().lower()


def val_match(expected, actual, tolerance=None):
    """Check if expected and actual values match (numeric or string)."""
    if expected is None and actual is None:
        return True
    e_str = str(expected).strip() if expected is not None else "N/A"
    a_str = str(actual).strip() if actual is not None else "N/A"

    # Handle None / N/A equivalence
    if e_str.upper() == "NONE":
        e_str = "N/A"
    if a_str.upper() == "NONE":
        a_str = "N/A"

    # Handle N/A
    if e_str.upper() == "N/A":
        return a_str.upper() == "N/A"

    # Try numeric
    try:
        e_f = float(e_str)
        a_f = float(a_str)
        tol = tolerance if tolerance is not None else 0.01
        return abs(e_f - a_f) <= tol
    except (ValueError, TypeError):
        pass

    # String compare
    return e_str.lower() == a_str.lower()


# ---------- Check 1: Local Excel file ----------
def check_local_excel(workspace, expected_rows, expected_summary):
    path = os.path.join(workspace, "financial_health_report.xlsx")
    if not os.path.exists(path):
        print(f"FAIL: financial_health_report.xlsx not found at {path}")
        return 0, 1

    wb = openpyxl.load_workbook(path, data_only=True)
    total_checks = 0
    passed_checks = 0

    # --- Income Analysis sheet ---
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower().replace(" ", "") == "incomeanalysis":
            sheet_name = name
            break
    if not sheet_name:
        print("FAIL: 'Income Analysis' sheet not found")
        total_checks += 1
        return 0, total_checks

    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(min_row=2, values_only=True))

    INCOME_COLS = ["Symbol", "Company_Name", "Total_Revenue", "Net_Income",
                   "Gross_Profit", "Operating_Income", "Profit_Margin_Pct", "Operating_Margin_Pct"]

    for i, exp_row in enumerate(expected_rows):
        if i >= len(rows_data):
            print(f"FAIL: Missing row {i+1} in Income Analysis")
            total_checks += len(INCOME_COLS)
            continue
        actual_row = rows_data[i]
        for j, col in enumerate(INCOME_COLS):
            total_checks += 1
            exp_val = exp_row[col]
            act_val = actual_row[j] if j < len(actual_row) else None

            if col in ("Total_Revenue", "Net_Income", "Gross_Profit", "Operating_Income"):
                tol = MONEY_TOLERANCE
            elif col in ("Profit_Margin_Pct", "Operating_Margin_Pct"):
                tol = PCT_TOLERANCE
            else:
                tol = None

            if val_match(exp_val, act_val, tol):
                passed_checks += 1
            else:
                print(f"  Income Analysis mismatch row {i+1} col '{col}': expected={exp_val}, actual={act_val}")

    # --- Health Assessment sheet ---
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower().replace(" ", "") == "healthassessment":
            sheet_name = name
            break
    if not sheet_name:
        print("FAIL: 'Health Assessment' sheet not found")
        total_checks += 1
    else:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(min_row=2, values_only=True))
        HEALTH_COLS = ["Symbol", "Credit_Rating", "Rating_Outlook", "Profit_Margin_Pct", "Margin_Health"]

        for i, exp_row in enumerate(expected_rows):
            if i >= len(rows_data):
                print(f"FAIL: Missing row {i+1} in Health Assessment")
                total_checks += len(HEALTH_COLS)
                continue
            actual_row = rows_data[i]
            for j, col in enumerate(HEALTH_COLS):
                total_checks += 1
                exp_val = exp_row[col]
                act_val = actual_row[j] if j < len(actual_row) else None

                if col == "Profit_Margin_Pct":
                    tol = PCT_TOLERANCE
                else:
                    tol = None

                if val_match(exp_val, act_val, tol):
                    passed_checks += 1
                else:
                    print(f"  Health Assessment mismatch row {i+1} col '{col}': expected={exp_val}, actual={act_val}")

    # --- Summary sheet ---
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == "summary":
            sheet_name = name
            break
    if not sheet_name:
        print("FAIL: 'Summary' sheet not found")
        total_checks += len(expected_summary)
    else:
        ws = wb[sheet_name]
        summary_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary_data[str(row[0]).strip()] = row[1]

        for metric, exp_val in expected_summary.items():
            total_checks += 1
            act_val = summary_data.get(metric)
            if metric in ("Avg_Profit_Margin", "Avg_Operating_Margin"):
                tol = PCT_TOLERANCE
            elif metric == "Companies_With_Strong_Margin":
                tol = 0.01
            else:
                tol = None

            if val_match(exp_val, act_val, tol):
                passed_checks += 1
            else:
                print(f"  Summary mismatch '{metric}': expected={exp_val}, actual={act_val}")

    wb.close()
    return passed_checks, total_checks


# ---------- Check 2: Google Sheet in DB ----------
def check_gsheet(expected_rows):
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()

    total_checks = 0
    passed_checks = 0

    # Check spreadsheet exists with correct title
    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE LOWER(title) LIKE '%financial health dashboard%'
    """)
    ss_rows = cur.fetchall()
    total_checks += 1
    if not ss_rows:
        print("FAIL: No spreadsheet found with title containing 'Financial Health Dashboard'")
        cur.close()
        conn.close()
        return 0, 1
    passed_checks += 1
    spreadsheet_id = ss_rows[0][0]
    print(f"  Found spreadsheet: id={spreadsheet_id}, title={ss_rows[0][1]}")

    # Check sheet exists with title "Overview"
    cur.execute("""
        SELECT id, title FROM gsheet.sheets
        WHERE spreadsheet_id = %s AND LOWER(title) = 'overview'
    """, (spreadsheet_id,))
    sh_rows = cur.fetchall()
    total_checks += 1
    if not sh_rows:
        print("FAIL: No sheet named 'Overview' found")
        cur.close()
        conn.close()
        return passed_checks, total_checks
    passed_checks += 1
    sheet_id = sh_rows[0][0]

    # Get all cells
    cur.execute("""
        SELECT row_index, col_index, value FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index, col_index
    """, (spreadsheet_id, sheet_id))
    cells = {}
    for row_idx, col_idx, value in cur.fetchall():
        cells[(row_idx, col_idx)] = value

    cur.close()
    conn.close()

    if not cells:
        print("FAIL: No cells found in the Overview sheet")
        return passed_checks, total_checks + 1

    # Expected columns: Symbol, Revenue_M, Net_Income_M, Profit_Margin_Pct, Credit_Rating, Margin_Health
    # Header row (row 0), data rows (row 1-5)
    OVERVIEW_COLS = {
        0: ("Symbol", None),
        1: ("Revenue_M", "Total_Revenue"),
        2: ("Net_Income_M", "Net_Income"),
        3: ("Profit_Margin_Pct", "Profit_Margin_Pct"),
        4: ("Credit_Rating", "Credit_Rating"),
        5: ("Margin_Health", "Margin_Health"),
    }

    for i, exp_row in enumerate(expected_rows):
        row_idx = i + 1
        for col_idx, (col_name, data_key) in OVERVIEW_COLS.items():
            total_checks += 1

            if data_key is None:
                exp_val = exp_row["Symbol"]
            else:
                exp_val = exp_row[data_key]

            act_val = cells.get((row_idx, col_idx))
            if act_val is None:
                print(f"  GSheet missing cell ({row_idx},{col_idx}) for {exp_row['Symbol']}.{col_name}")
                continue

            if col_name in ("Revenue_M", "Net_Income_M"):
                tol = MONEY_TOLERANCE
            elif col_name == "Profit_Margin_Pct":
                tol = PCT_TOLERANCE
            else:
                tol = None

            if val_match(exp_val, act_val, tol):
                passed_checks += 1
            else:
                print(f"  GSheet mismatch ({row_idx},{col_idx}) {exp_row['Symbol']}.{col_name}: expected={exp_val}, actual={act_val}")

    return passed_checks, total_checks


# ---------- Main ----------
def main(args):
    print("Computing expected values from Yahoo Finance DB ...")
    expected_rows, expected_summary = get_expected_data()

    print(f"  Expected summary: {expected_summary}")

    total_passed = 0
    total_checks = 0

    # Check 1: Local Excel
    print("\n--- Check 1: Local financial_health_report.xlsx ---")
    p, t = check_local_excel(args.agent_workspace, expected_rows, expected_summary)
    print(f"  Local Excel: {p}/{t} checks passed")
    total_passed += p
    total_checks += t

    # Check 2: Google Sheet
    print("\n--- Check 2: Google Sheet in DB ---")
    p, t = check_gsheet(expected_rows)
    print(f"  Google Sheet: {p}/{t} checks passed")
    total_passed += p
    total_checks += t

    # Overall
    if total_checks == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = total_passed / total_checks * 100
    print(f"\nOverall: {total_passed}/{total_checks} checks passed ({accuracy:.1f}%)")

    if args.output_file:
        report = {
            "total_passed": total_passed,
            "total_checks": total_checks,
            "accuracy": accuracy,
            "timestamp": datetime.now().isoformat(),
        }
        with open(args.output_file, "w") as f:
            json.dump(report, f, indent=2)
        print(f"Report saved to {args.output_file}")

    if accuracy >= 80:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--output_file", required=False)
    args = parser.parse_args()
    main(args)

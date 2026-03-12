"""
Evaluation script for yf-market-summary-notion task.

Checks:
1. Excel Market_Overview.xlsx with Stock Summary sheet
2. Notion page titled "Market Dashboard" exists in DB

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        d = (detail[:200] + "...") if len(detail) > 200 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    symbols = ['AMZN', 'GOOGL', 'JNJ', 'JPM', 'XOM']

    # Latest prices
    cur.execute("""
        SELECT symbol, close FROM yf.stock_prices
        WHERE date = (SELECT MAX(date) FROM yf.stock_prices WHERE symbol='AMZN')
        AND symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM') ORDER BY symbol
    """)
    latest = {r[0]: float(r[1]) for r in cur.fetchall()}

    # 30 trading days ago
    cur.execute("""
        WITH ranked AS (
            SELECT symbol, date, close, ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY date DESC) as rn
            FROM yf.stock_prices WHERE symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM')
        )
        SELECT symbol, close FROM ranked WHERE rn = 31 ORDER BY symbol
    """)
    price_30d = {r[0]: float(r[1]) for r in cur.fetchall()}

    # 90 trading days ago
    cur.execute("""
        WITH ranked AS (
            SELECT symbol, date, close, ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY date DESC) as rn
            FROM yf.stock_prices WHERE symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM')
        )
        SELECT symbol, close FROM ranked WHERE rn = 91 ORDER BY symbol
    """)
    price_90d = {r[0]: float(r[1]) for r in cur.fetchall()}

    # Stock info
    cur.execute("""
        SELECT symbol, data->>'shortName', data->>'sector'
        FROM yf.stock_info
        WHERE symbol IN ('AMZN','GOOGL','JNJ','JPM','XOM') ORDER BY symbol
    """)
    info = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    conn.close()

    rows = []
    for sym in symbols:
        lp = latest.get(sym, 0)
        p30 = price_30d.get(sym, 0)
        p90 = price_90d.get(sym, 0)
        ret_30 = round((lp - p30) / p30 * 100, 2) if p30 else 0
        ret_90 = round((lp - p90) / p90 * 100, 2) if p90 else 0
        name, sector = info.get(sym, ("", ""))
        rows.append((sym, name, sector, lp, p30, ret_30, p90, ret_90))

    # Best/worst 30d
    best = max(rows, key=lambda r: r[5])
    worst = min(rows, key=lambda r: r[5])

    return {"stocks": rows, "best_30d": best, "worst_30d": worst}


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Market_Overview.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Sheet 'Stock Summary' exists", get_sheet(wb, "Stock Summary") is not None,
          f"Found: {wb.sheetnames}")

    ws = get_sheet(wb, "Stock Summary")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["stocks"]
        check("Stock Summary row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        agent_by_sym = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_sym[str(row[0]).strip().upper()] = row

        for exp_row in exp:
            sym = exp_row[0]
            agent_row = agent_by_sym.get(sym)
            if agent_row:
                check(f"'{sym}' Latest_Price",
                      num_close(agent_row[3], exp_row[3], 2.0),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
                check(f"'{sym}' Price_30d_Ago",
                      num_close(agent_row[4], exp_row[4], 2.0),
                      f"Expected {exp_row[4]}, got {agent_row[4]}")
                check(f"'{sym}' Return_30d_Pct",
                      num_close(agent_row[5], exp_row[5], 1.0),
                      f"Expected {exp_row[5]}, got {agent_row[5]}")
                check(f"'{sym}' Price_90d_Ago",
                      num_close(agent_row[6], exp_row[6], 2.0),
                      f"Expected {exp_row[6]}, got {agent_row[6]}")
                check(f"'{sym}' Return_90d_Pct",
                      num_close(agent_row[7], exp_row[7], 1.5),
                      f"Expected {exp_row[7]}, got {agent_row[7]}")
            else:
                check(f"'{sym}' found in output", False, "Not in agent output")

        # Check sort order (alphabetical by symbol)
        if len(agent_rows) >= 2:
            syms = [str(r[0]).strip().upper() for r in agent_rows if r and r[0]]
            check("Sorted by Symbol alphabetically",
                  syms == sorted(syms),
                  f"Got: {syms}")


def check_notion(expected):
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for notion check", False, str(e), db=True)
        return

    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()
    conn.close()

    check("At least one Notion page exists", len(pages) > 0,
          f"Found {len(pages)} pages", db=True)

    found_dashboard = False
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else {}
        # Title could be in various property formats
        page_str = json.dumps(props).lower()
        if "market dashboard" in page_str:
            found_dashboard = True
            check("Notion page 'Market Dashboard' found", True, db=True)
            break

    if not found_dashboard:
        # Also check title property specifically
        for page in pages:
            props = page[1] if isinstance(page[1], dict) else {}
            title_prop = props.get("title", props.get("Title", {}))
            if isinstance(title_prop, dict):
                title_val = title_prop.get("title", [])
                if isinstance(title_val, list):
                    for t in title_val:
                        if isinstance(t, dict) and "market dashboard" in str(t.get("plain_text", "")).lower():
                            found_dashboard = True
                            break
                        if isinstance(t, dict) and "market dashboard" in str(t.get("text", {}).get("content", "")).lower():
                            found_dashboard = True
                            break
            if found_dashboard:
                check("Notion page 'Market Dashboard' found", True, db=True)
                break

    if not found_dashboard:
        check("Notion page 'Market Dashboard' found", False,
              f"Pages: {[json.dumps(p[1])[:100] for p in pages]}", db=True)


def check_excel_gt(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Market_Overview.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Market_Overview.xlsx")
    check("Excel file exists", os.path.isfile(agent_file))
    check("Groundtruth file exists", os.path.isfile(gt_file))
    if not os.path.isfile(agent_file) or not os.path.isfile(gt_file):
        return
    agent_wb = openpyxl.load_workbook(agent_file)
    gt_wb = openpyxl.load_workbook(gt_file)
    check("Sheet 'Stock Summary' exists", get_sheet(agent_wb, "Stock Summary") is not None)
    a_ws = get_sheet(agent_wb, "Stock Summary")
    g_ws = get_sheet(gt_wb, "Stock Summary")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Row count matches", len(a_rows) == len(g_rows),
              f"Expected {len(g_rows)}, got {len(a_rows)}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        check_excel(agent_workspace, expected)
    else:
        print("INFO: Falling back to groundtruth Excel")
        check_excel_gt(agent_workspace, groundtruth_workspace)

    check_notion(expected)

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": file_ok}
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return file_ok, f"Passed: {total_pass}, Failed: {total_fail}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

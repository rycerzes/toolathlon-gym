"""Evaluation for terminal-canvas-gform-gcal-excel-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

# Quizzes below 75 avg from courses 7 and 11
BELOW_75_QUIZZES = [
    "CMA 24295", "CMA 24298",
    "CMA 25341", "CMA 25343", "CMA 25344", "CMA 25345", "CMA 25346", "CMA 25347"
]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def check_excel(ws_path):
    """Check Quiz_Performance_Report.xlsx."""
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "Quiz_Performance_Report.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Quiz_Performance
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
    qp_name = None
    for candidate in ["quiz_performance", "quiz performance"]:
        if candidate in sheet_names_lower:
            qp_name = sheet_names_lower[candidate]
            break
    if qp_name is None:
        check("Quiz_Performance sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Quiz_Performance sheet exists", True)
        ws = wb[qp_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        # Query dynamic quiz count from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.quizzes WHERE course_id IN (7, 11)")
            expected_quiz_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_quiz_count = 11
        check(f"Quiz_Performance has {expected_quiz_count} rows",
              len(data_rows) == expected_quiz_count,
              f"Found {len(data_rows)} data rows")

        # Query dynamic quiz avg scores from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("""
                SELECT q.id, AVG(qs.score) FROM canvas.quizzes q
                JOIN canvas.quiz_submissions qs ON q.id = qs.quiz_id
                WHERE q.course_id IN (7, 11)
                GROUP BY q.id
            """)
            quiz_avgs = {str(row[0]): float(row[1]) for row in cur.fetchall()}
            cur.close(); conn.close()
            expected_25341_avg = quiz_avgs.get("25341", 66.62)
            expected_24296_avg = quiz_avgs.get("24296", 78.94)
        except Exception:
            expected_25341_avg, expected_24296_avg = 66.62, 78.94

        # Check a specific quiz avg score
        for r in data_rows:
            if r[3] and "25341" in str(r[3]):
                avg = safe_float(r[4])
                check(f"CMA 25341 avg ~{expected_25341_avg:.2f}",
                      avg is not None and abs(avg - expected_25341_avg) < 2.0,
                      f"Got {avg}")
                needs = str(r[6]).strip().lower() if len(r) > 6 and r[6] else ""
                check("CMA 25341 marked Yes for review", "yes" in needs, f"Got {r[6]}")
                break

        # Check course 7 quiz
        for r in data_rows:
            if r[3] and "24296" in str(r[3]):
                avg = safe_float(r[4])
                check(f"CMA 24296 avg ~{expected_24296_avg:.2f}",
                      avg is not None and abs(avg - expected_24296_avg) < 2.0,
                      f"Got {avg}")
                needs = str(r[6]).strip().lower() if len(r) > 6 and r[6] else ""
                check("CMA 24296 marked No for review", "no" in needs, f"Got {r[6]}")
                break

    # Sheet 2: Feedback_Summary
    fb_name = None
    for candidate in ["feedback_summary", "feedback summary"]:
        if candidate in sheet_names_lower:
            fb_name = sheet_names_lower[candidate]
            break
    if fb_name is None:
        check("Feedback_Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Feedback_Summary sheet exists", True)
        ws2 = wb[fb_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Feedback_Summary has 5 rows", len(data_rows2) == 5,
              f"Found {len(data_rows2)} rows")

    # Sheet 3: Remediation_Schedule
    rs_name = None
    for candidate in ["remediation_schedule", "remediation schedule"]:
        if candidate in sheet_names_lower:
            rs_name = sheet_names_lower[candidate]
            break
    if rs_name is None:
        check("Remediation_Schedule sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Remediation_Schedule sheet exists", True)
        ws3 = wb[rs_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("Remediation_Schedule has 8 rows", len(data_rows3) == 8,
              f"Found {len(data_rows3)} rows")

    wb.close()


def check_gform():
    """Check Google Form creation."""
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    # Find the assessment form (not the noise one)
    form_id = None
    for fid, title in forms:
        t = (title or "").lower()
        if "assessment" in t or "performance" in t or "academic" in t:
            form_id = fid
            break

    check("Assessment feedback form created", form_id is not None,
          f"Forms: {[f[1] for f in forms]}")

    if form_id:
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        check("Form has exactly 5 questions", q_count == 5, f"Found {q_count}")

    conn.close()


def check_gcal():
    """Check calendar events."""
    print("\n=== Checking Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT summary, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()

    review_events = [e for e in events if "review" in (e[0] or "").lower()
                     or "quiz" in (e[0] or "").lower()]
    check("At least 8 quiz review events created", len(review_events) >= 8,
          f"Found {len(review_events)} review events")

    # Check that at least some quiz titles are in events
    found_count = 0
    for quiz_title in BELOW_75_QUIZZES[:4]:
        quiz_num = quiz_title.split()[-1]
        if any(quiz_num in (e[0] or "") for e in review_events):
            found_count += 1
    check("Review events reference quiz titles", found_count >= 2,
          f"Found {found_count}/4 quiz references in events")

    conn.close()


def check_email():
    """Check email sent."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()

    target_email = None
    for subj, from_addr, to_addr, body in all_emails:
        if to_addr:
            recipients = []
            if isinstance(to_addr, list):
                recipients = [str(r).strip().lower() for r in to_addr]
            elif isinstance(to_addr, str):
                try:
                    parsed = json.loads(to_addr)
                    if isinstance(parsed, list):
                        recipients = [str(r).strip().lower() for r in parsed]
                    else:
                        recipients = [str(to_addr).strip().lower()]
                except (json.JSONDecodeError, TypeError):
                    recipients = [str(to_addr).strip().lower()]
            if "faculty@assessment.example.com" in recipients:
                target_email = (subj, from_addr, to_addr, body)
                break

    check("Email sent to faculty@assessment.example.com", target_email is not None,
          f"Total emails: {len(all_emails)}")

    if target_email:
        subj, from_addr, to_addr, body = target_email
        check("Email subject mentions quiz performance",
              "quiz" in (subj or "").lower() or "performance" in (subj or "").lower(),
              f"Subject: {subj}")
        check("Email from coordinator@assessment.example.com",
              "coordinator@assessment.example.com" in (from_addr or "").lower(),
              f"From: {from_addr}")
        body_lower = (body or "").lower()
        check("Email body mentions remediation or review",
              "remediation" in body_lower or "review" in body_lower or "flagged" in body_lower,
              "Expected remediation/review content in body")
        check("Email body mentions survey or feedback",
              "survey" in body_lower or "feedback" in body_lower,
              "Expected survey/feedback mention in body")

    conn.close()


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check that noise emails (admin, IT, facilities, grants) are not in agent output
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr FROM email.messages
            WHERE from_addr = 'coordinator@assessment.example.com'
        """)
        sent_emails = cur.fetchall()
        noise_recipients = ["all-staff@university.edu", "faculty@university.edu",
                            "all@university.edu", "researchers@university.edu"]
        for email_row in sent_emails:
            to_str = str(email_row[0]).lower()
            for noise in noise_recipients:
                if noise in to_str:
                    check("No email sent to noise recipients", False,
                          f"Sent to noise recipient: {noise}")
                    cur.close(); conn.close()
                    return
        check("No email sent to noise recipients", True)
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-CANVAS-GFORM-GCAL-EXCEL-EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_gform()
    check_gcal()
    check_email()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

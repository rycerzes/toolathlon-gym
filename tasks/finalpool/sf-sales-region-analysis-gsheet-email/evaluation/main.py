"""Evaluation for sf-sales-region-analysis-gsheet-email."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    errors = []
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl not installed")
        return errors

    agent_file = os.path.join(agent_workspace, "Regional_Sales_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Regional_Sales_Report.xlsx")

    if not os.path.exists(agent_file):
        errors.append("Regional_Sales_Report.xlsx not found in agent workspace")
        return errors
    if not os.path.exists(gt_file):
        errors.append("Groundtruth Regional_Sales_Report.xlsx not found")
        return errors

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Regional Performance sheet
    a_rows = load_sheet_rows(agent_wb, "Regional Performance")
    g_rows = load_sheet_rows(gt_wb, "Regional Performance")
    if a_rows is None:
        errors.append("Sheet 'Regional Performance' not found in agent output")
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        g_data = [r for r in (g_rows[1:] if g_rows and len(g_rows) > 1 else []) if r and r[0] is not None]

        if len(a_data) < 5:
            errors.append(f"Regional Performance: expected 5 data rows, got {len(a_data)}")
        else:
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
            for g_row in g_data:
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing region row: {g_row[0]}")
                    continue
                # Order_Count col 1
                if len(a_row) > 1 and not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{g_row[0]} Order_Count: got {a_row[1]}, expected {g_row[1]} (tol=5)")
                # Total_Revenue col 2
                if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 1.0):
                    errors.append(f"{g_row[0]} Total_Revenue: got {a_row[2]}, expected {g_row[2]} (tol=1.0)")
                # Avg_Order_Value col 3
                if len(a_row) > 3 and not num_close(a_row[3], g_row[3], 1.0):
                    errors.append(f"{g_row[0]} Avg_Order_Value: got {a_row[3]}, expected {g_row[3]} (tol=1.0)")
                # Revenue_Share_Pct col 4
                if len(a_row) > 4 and not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{g_row[0]} Revenue_Share_Pct: got {a_row[4]}, expected {g_row[4]} (tol=0.5)")

    # Check Summary sheet
    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")
    if a_sum is None:
        errors.append("Sheet 'Summary' not found in agent output")
    else:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}
        g_sum_data = {str(r[0]).strip().lower(): r[1] for r in (g_sum[1:] if g_sum and len(g_sum) > 1 else []) if r and r[0]}

        # Total_Revenue
        tr = a_sum_data.get("total_revenue")
        if tr is None:
            errors.append("Summary missing Total_Revenue")
        elif not num_close(tr, 3048998.33, 1.0):
            errors.append(f"Total_Revenue: got {tr}, expected 3048998.33 (tol=1.0)")

        # Total_Orders
        to = a_sum_data.get("total_orders")
        if to is None:
            errors.append("Summary missing Total_Orders")
        elif not num_close(to, 20000, 10):
            errors.append(f"Total_Orders: got {to}, expected 20000 (tol=10)")

        # Top_Region
        top = a_sum_data.get("top_region")
        if top is None:
            errors.append("Summary missing Top_Region")
        elif not str_match(top, "Europe"):
            errors.append(f"Top_Region: got '{top}', expected 'Europe'")

        # Bottom_Region
        bot = a_sum_data.get("bottom_region")
        if bot is None:
            errors.append("Summary missing Bottom_Region")
        elif not str_match(bot, "Latin America"):
            errors.append(f"Bottom_Region: got '{bot}', expected 'Latin America'")

    return errors


def check_gsheet():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT s.id, s.title
            FROM gsheet.spreadsheets s
            WHERE LOWER(s.title) LIKE '%regional%' OR LOWER(s.title) LIKE '%sales regional%'
        """)
        sheets = cur.fetchall()
        if not sheets:
            errors.append("No Google Sheet named 'Sales Regional Performance' found")
        else:
            ss_id = sheets[0][0]
            cur.execute("""
                SELECT COUNT(DISTINCT c.row_index)
                FROM gsheet.cells c
                WHERE c.spreadsheet_id = %s AND c.row_index > 0
            """, (ss_id,))
            row_count = cur.fetchone()[0]
            if row_count < 5:
                errors.append(f"Sales Regional Performance sheet has only {row_count} data rows, expected at least 5")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"GSheet DB check error: {e}")
    return errors


def check_emails():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%regional%' OR LOWER(subject) LIKE '%sales performance%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if len(emails) < 5:
            errors.append(f"Expected at least 5 regional emails, found {len(emails)}")
        # Check that at least one email per region manager exists
        all_to = " ".join(str(em[1]).lower() for em in emails)
        for mgr in ["europe.manager", "apac.manager", "na.manager", "me.manager", "latam.manager"]:
            if mgr not in all_to:
                errors.append(f"No email found for {mgr}@company.com")
    except Exception as e:
        errors.append(f"Email DB check error: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    docx_path = os.path.join(agent_workspace, "Regional_Summary.docx")
    if not os.path.exists(docx_path):
        errors.append("Regional_Summary.docx not found")
        return errors
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        if len(text.strip()) < 30:
            errors.append("Regional_Summary.docx has too little content")
        for kw in ["regional", "europe"]:
            if kw not in text:
                errors.append(f"Regional_Summary.docx missing keyword: {kw}")
    except ImportError:
        if os.path.getsize(docx_path) < 100:
            errors.append("Regional_Summary.docx too small")
    except Exception as e:
        errors.append(f"Error reading Regional_Summary.docx: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []

    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace, gt_dir)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    print("\n=== Checking Google Sheet ===")
    gsheet_errors = check_gsheet()
    if gsheet_errors:
        for e in gsheet_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(gsheet_errors)
    else:
        print("  [PASS] GSheet check passed")

    print("\n=== Checking Emails ===")
    email_errors = check_emails()
    if email_errors:
        for e in email_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(email_errors)
    else:
        print("  [PASS] Email check passed")

    print("\n=== Checking Word Document ===")
    word_errors = check_word(args.agent_workspace)
    if word_errors:
        for e in word_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(word_errors)
    else:
        print("  [PASS] Word check passed")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"errors": all_errors, "success": len(all_errors) == 0}, f, indent=2)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

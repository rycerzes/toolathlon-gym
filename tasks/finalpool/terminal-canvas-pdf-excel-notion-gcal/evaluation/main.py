"""Evaluation script for terminal-canvas-pdf-excel-notion-gcal."""
import os
import argparse, json, os, sys

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def get_expected_from_db():
    """Query canvas schema dynamically for expected average scores."""
    defaults = {
        "avg_gpa": 75.4,
    }
    try:
        import psycopg2
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Compute overall average score across all submissions
        cur.execute("SELECT AVG(score) FROM canvas.submissions WHERE score IS NOT NULL")
        row = cur.fetchone()
        if row and row[0] is not None:
            defaults["avg_gpa"] = float(row[0])
        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [WARN] DB query for expected values failed, using defaults: {e}")
    return defaults


EXPECTED = get_expected_from_db()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except:
        return False


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def check_excel(agent_workspace):
    print("\n=== Checking Excel Workbook ===")
    import openpyxl

    path = os.path.join(agent_workspace, "Accreditation_Self_Study.xlsx")
    check("Accreditation_Self_Study.xlsx exists", os.path.exists(path))
    if not os.path.exists(path):
        return

    wb = openpyxl.load_workbook(path)
    sheet_names = [s.lower() for s in wb.sheetnames]

    # Check 4 sheets exist
    has_course = any("course" in s and "data" in s for s in sheet_names)
    has_matrix = any("accreditation" in s or "matrix" in s for s in sheet_names)
    has_gap = any("gap" in s for s in sheet_names)
    has_resource = any("resource" in s for s in sheet_names)
    check("Course_Data sheet exists", has_course, f"Sheets: {wb.sheetnames}")
    check("Accreditation_Matrix sheet exists", has_matrix, f"Sheets: {wb.sheetnames}")
    check("Gap_Analysis sheet exists", has_gap, f"Sheets: {wb.sheetnames}")
    check("Resource_Needs sheet exists", has_resource, f"Sheets: {wb.sheetnames}")

    # Check Course_Data content
    for sn in wb.sheetnames:
        if "course" in sn.lower() and "data" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Course_Data has 22 course rows", len(rows) >= 22, f"Got {len(rows)}")
            if rows:
                # Check some known course names
                names = [str(r[0]).lower() if r[0] else "" for r in rows]
                check("Course_Data includes Creative Computing", any("creative" in n for n in names))
                check("Course_Data includes Foundations of Finance", any("finance" in n for n in names))
            # Check headers
            headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
            check("Course_Data has enrollment column", any("enroll" in h for h in headers), f"Headers: {headers}")
            check("Course_Data has grade column", any("grade" in h or "avg" in h for h in headers), f"Headers: {headers}")
            break

    # Check Accreditation_Matrix content
    for sn in wb.sheetnames:
        if "accreditation" in sn.lower() or "matrix" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Accreditation_Matrix has 8 criteria rows", len(rows) >= 8, f"Got {len(rows)}")
            if rows:
                all_text = " ".join(str(c).lower() for r in rows for c in r if c)
                check("Matrix has Compliant status", "compliant" in all_text)
                check("Matrix has Partial status", "partial" in all_text)
                # Check specific criteria
                check("Matrix mentions Student Learning Outcomes", "student learning" in all_text or "learning outcome" in all_text, f"Text sample: {all_text[:200]}")
                # Check threshold values
                has_78 = any(num_close(r[1], 78.0, tol=1.0) for r in rows if r[1] is not None)
                has_065 = any(num_close(r[1], 0.65, tol=0.05) for r in rows if r[1] is not None)
                check("Matrix has C1 threshold ~78", has_78, f"Thresholds: {[r[1] for r in rows]}")
                check("Matrix has C2 threshold ~0.65", has_065, f"Thresholds: {[r[1] for r in rows]}")
            break

    # Check Gap_Analysis content
    for sn in wb.sheetnames:
        if "gap" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            check("Gap_Analysis has rows for partial/non-compliant criteria", len(non_empty) >= 2, f"Got {len(non_empty)}")
            if non_empty:
                all_text = " ".join(str(c).lower() for r in non_empty for c in r if c)
                check("Gap_Analysis mentions remediation", "remed" in all_text or "action" in all_text or "improve" in all_text)
            break

    # Check Resource_Needs content
    for sn in wb.sheetnames:
        if "resource" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            check("Resource_Needs has at least 3 rows", len(non_empty) >= 3, f"Got {len(non_empty)}")
            if non_empty:
                headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
                check("Resource_Needs has cost column", any("cost" in h or "estimated" in h for h in headers), f"Headers: {headers}")
            break


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Find accreditation database
        cur.execute("SELECT id, title, properties FROM notion.databases")
        dbs = cur.fetchall()
        accred_db = None
        for db_id, title_json, props_json in dbs:
            title_text = ""
            if title_json:
                titles = title_json if isinstance(title_json, list) else json.loads(title_json) if isinstance(title_json, str) else []
                for t in titles:
                    if isinstance(t, dict) and "text" in t:
                        title_text += t["text"].get("content", "")
            if "accreditation" in title_text.lower() and "action" in title_text.lower():
                accred_db = (db_id, props_json)
                break

        check("Accreditation Action Items database exists", accred_db is not None,
              f"Found DBs: {[str(d[1])[:80] for d in dbs]}")

        if accred_db:
            db_id, props = accred_db
            props_data = props if isinstance(props, dict) else json.loads(props) if isinstance(props, str) else {}

            # Check properties
            prop_names = [k.lower() for k in props_data.keys()]
            check("DB has Status property", any("status" in p for p in prop_names), f"Props: {list(props_data.keys())}")
            check("DB has Due_Date or date property", any("date" in p or "due" in p for p in prop_names), f"Props: {list(props_data.keys())}")

            # Check pages
            cur.execute("SELECT id, properties FROM notion.pages WHERE parent::text LIKE %s AND archived = false",
                        (f'%{db_id}%',))
            pages = cur.fetchall()
            check("Notion has 8 accreditation criterion pages", len(pages) >= 8, f"Got {len(pages)}")

            if pages:
                # Check page content
                statuses = []
                for _, page_props in pages:
                    pp = page_props if isinstance(page_props, dict) else json.loads(page_props) if isinstance(page_props, str) else {}
                    for k, v in pp.items():
                        if "status" in k.lower() and isinstance(v, dict):
                            sel = v.get("select", {})
                            if sel and isinstance(sel, dict):
                                statuses.append(sel.get("name", ""))
                has_compliant = any("compliant" in s.lower() for s in statuses if s and "non" not in s.lower() and "partial" not in s.lower())
                has_partial = any("partial" in s.lower() for s in statuses if s)
                check("Pages include Compliant status", has_compliant, f"Statuses: {statuses}")
                check("Pages include Partial status", has_partial, f"Statuses: {statuses}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion accessible", False, str(e))


def check_gcal(launch_time):
    print("\n=== Checking Google Calendar Events ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        launch_dt = None
        if launch_time:
            from datetime import datetime
            launch_dt = datetime.strptime(launch_time, "%Y-%m-%d %H:%M:%S")

        cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        summaries = [str(e[0]).lower() for e in events]

        has_evidence = any("evidence" in s and "deadline" in s for s in summaries)
        has_draft = any("draft" in s and ("report" in s or "due" in s) for s in summaries)
        has_final = any("final" in s and "submission" in s for s in summaries)
        check("Evidence Collection Deadline event exists", has_evidence, f"Events: {summaries}")
        check("Draft Report Due event exists", has_draft, f"Events: {summaries}")
        check("Final Submission event exists", has_final, f"Events: {summaries}")

        # Check descriptions mention accreditation
        accred_events = [e for e in events if any(kw in str(e[0]).lower() for kw in ["evidence", "draft report", "final submission"])]
        if accred_events:
            descs = " ".join(str(e[1]).lower() for e in accred_events if e[1])
            check("Calendar events mention accreditation", "accreditation" in descs or "self-study" in descs or "self study" in descs,
                  f"Descriptions: {descs[:200]}")

        # Check timing (roughly 30/60/90 days from launch)
        if launch_dt and accred_events:
            from datetime import timedelta
            for e in events:
                s = str(e[0]).lower()
                if e[2] and "evidence" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Evidence deadline ~30 days from launch", 25 <= days_diff <= 35, f"Days: {days_diff}")
                elif e[2] and "draft" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Draft due ~60 days from launch", 55 <= days_diff <= 65, f"Days: {days_diff}")
                elif e[2] and "final" in s and "submission" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Final submission ~90 days from launch", 85 <= days_diff <= 95, f"Days: {days_diff}")

        cur.close()
        conn.close()
    except Exception as e:
        check("GCal accessible", False, str(e))


def check_scripts(agent_workspace):
    print("\n=== Checking Scripts and Outputs ===")
    check("compute_metrics.py exists", os.path.exists(os.path.join(agent_workspace, "compute_metrics.py")))
    check("evaluate_compliance.py exists", os.path.exists(os.path.join(agent_workspace, "evaluate_compliance.py")))
    check("generate_summary.py exists", os.path.exists(os.path.join(agent_workspace, "generate_summary.py")))

    # Check program_metrics.json
    metrics_path = os.path.join(agent_workspace, "program_metrics.json")
    check("program_metrics.json exists", os.path.exists(metrics_path))
    if os.path.exists(metrics_path):
        with open(metrics_path) as f:
            metrics = json.load(f)
        check("Metrics has avg_gpa", "avg_gpa" in metrics, f"Keys: {list(metrics.keys())}")
        if "avg_gpa" in metrics:
            check("avg_gpa roughly correct",
                  num_close(metrics["avg_gpa"], EXPECTED["avg_gpa"], tol=3.0),
                  f"Got {metrics.get('avg_gpa')}, expected ~{EXPECTED['avg_gpa']:.1f}")

    # Check compliance_assessment.json
    compliance_path = os.path.join(agent_workspace, "compliance_assessment.json")
    check("compliance_assessment.json exists", os.path.exists(compliance_path))
    if os.path.exists(compliance_path):
        with open(compliance_path) as f:
            compliance = json.load(f)
        # Should be a list or dict with 8 criteria
        if isinstance(compliance, list):
            check("Compliance has 8 entries", len(compliance) >= 8, f"Got {len(compliance)}")
        elif isinstance(compliance, dict):
            check("Compliance has criteria data", len(compliance) >= 3, f"Keys: {list(compliance.keys())[:10]}")

    # Check accreditation_summary.txt
    summary_path = os.path.join(agent_workspace, "accreditation_summary.txt")
    check("accreditation_summary.txt exists", os.path.exists(summary_path))
    if os.path.exists(summary_path):
        with open(summary_path) as f:
            text = f.read().lower()
        check("Summary mentions compliant", "compliant" in text)
        check("Summary mentions partial", "partial" in text)
        check("Summary has compliance percentage", "%" in text or "percent" in text, f"Length: {len(text)}")


def check_reverse_validation():
    """Check that noise Notion pages are NOT in the accreditation tracker."""
    print("\n=== Reverse Validation ===")
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Noise notion pages that should NOT appear in accreditation database
        noise_titles = ["q1 budget review", "server migration", "marketing campaign launch"]

        # Find the accreditation database
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        accred_db_id = None
        for db_id, title_json in dbs:
            title_text = ""
            if title_json:
                titles = title_json if isinstance(title_json, list) else json.loads(title_json) if isinstance(title_json, str) else []
                for t in titles:
                    if isinstance(t, dict) and "text" in t:
                        title_text += t["text"].get("content", "")
            if "accreditation" in title_text.lower() and "action" in title_text.lower():
                accred_db_id = db_id
                break

        if accred_db_id:
            cur.execute(
                "SELECT properties FROM notion.pages WHERE parent::text LIKE %s AND archived = false",
                (f'%{accred_db_id}%',))
            pages = cur.fetchall()
            page_titles = []
            for (props,) in pages:
                pp = props if isinstance(props, dict) else json.loads(props) if isinstance(props, str) else {}
                for k, v in pp.items():
                    if "title" in k.lower() or k.lower() == "name":
                        if isinstance(v, dict) and "title" in v:
                            for t in v["title"]:
                                if isinstance(t, dict) and "text" in t:
                                    page_titles.append(t["text"].get("content", "").lower())

            no_noise = not any(nt in " ".join(page_titles) for nt in noise_titles)
            check("No noise Notion pages in accreditation tracker (budget, migration, marketing)",
                  no_noise,
                  f"Page titles in accreditation DB: {page_titles}")
        else:
            check("No noise Notion pages in accreditation tracker", True,
                  "No accreditation DB found to check")

        # Check noise 'Project Milestones' database pages are not mixed in
        cur.execute("SELECT id, title FROM notion.databases")
        all_dbs = cur.fetchall()
        noise_db_titles = []
        for db_id, title_json in all_dbs:
            title_text = ""
            if title_json:
                titles = title_json if isinstance(title_json, list) else json.loads(title_json) if isinstance(title_json, str) else []
                for t in titles:
                    if isinstance(t, dict) and "text" in t:
                        title_text += t["text"].get("content", "")
            if "project milestones" in title_text.lower():
                noise_db_titles.append(title_text)
        # The noise DB existing is fine (it was injected), but accreditation DB should be separate
        check("Accreditation DB is distinct from noise Project Milestones DB",
              accred_db_id is not None,
              "Accreditation Action Items DB should exist separately")

        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation (notion noise)", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_gcal(args.launch_time)
    check_scripts(args.agent_workspace)
    check_reverse_validation()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

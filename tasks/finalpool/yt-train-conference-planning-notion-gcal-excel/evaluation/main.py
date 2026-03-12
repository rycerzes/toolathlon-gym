"""
Evaluation for yt-train-conference-planning-notion-gcal-excel task.

Checks:
1. Conference_Planning.xlsx exists with Pre_Conference_Videos, Speaker_Travel, Summary sheets
2. Pre_Conference_Videos has 6 AI-related Fireship videos
3. Speaker_Travel has 3 speakers with G235 and G168 trains
4. Summary has correct conference name and budget
5. Notion page 'Technology Innovation Forum 2026' exists
6. 3 GCal events for conference days
7. 3 emails sent to the three speakers
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Conference_Planning.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Conference_Planning.xlsx")
    if not os.path.exists(xlsx_path):
        record("Conference_Planning.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Conference_Planning.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Pre_Conference_Videos sheet",
           any("video" in s or "pre" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has Speaker_Travel sheet",
           any("speaker" in s or "travel" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has Summary sheet", "summary" in sheet_names_lower, f"Sheets: {wb.sheetnames}")

    # Check Pre_Conference_Videos
    video_sheet = None
    for name in wb.sheetnames:
        if "video" in name.lower() or "pre" in name.lower():
            video_sheet = wb[name]
            break

    if video_sheet:
        rows = list(video_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Pre_Conference_Videos has 6 videos", len(data_rows) >= 6,
               f"Found {len(data_rows)} rows")
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Videos include DeepSeek AI content", "deepseek" in all_text,
               "No DeepSeek videos found")

    # Check Speaker_Travel
    speaker_sheet = None
    for name in wb.sheetnames:
        if "speaker" in name.lower() or "travel" in name.lower():
            speaker_sheet = wb[name]
            break

    if speaker_sheet:
        rows = list(speaker_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Speaker_Travel has 3 speakers", len(data_rows) >= 3,
               f"Found {len(data_rows)} rows")
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Speaker_Travel has G235 for Beijing speakers", "g235" in all_text, "No G235 found")
        record("Speaker_Travel has G168 for Shanghai speaker", "g168" in all_text, "No G168 found")
        record("Speaker Alex Kim listed", "alex" in all_text or "kim" in all_text,
               "No Alex Kim found")
        record("Speaker James Wu listed", "james" in all_text or "wu" in all_text,
               "No James Wu found")

    # Check Summary
    if "summary" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c).lower()
        record("Summary has conference name", "technology innovation forum" in all_text,
               "No conference name found")
        record("Summary has travel budget", "budget" in all_text or "338" in all_text,
               "No travel budget found")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Conference_Planning.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion page 'Technology Innovation Forum 2026' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()
        matching = []
        for pid, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "technology innovation forum" in props_str or "innovation forum 2026" in props_str:
                matching.append((pid, props))
        record("Notion page 'Technology Innovation Forum 2026' exists", len(matching) >= 1,
               f"Found {len(pages)} total pages, {len(matching)} matching")

        if matching:
            page_id = matching[0][0]
            cur.execute("""
                SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s AND archived = false
            """, (page_id,))
            block_count = cur.fetchone()[0]
            record("Notion page has content blocks", block_count >= 3,
                   f"Found {block_count} blocks")

            cur.execute("""
                SELECT block_data FROM notion.blocks WHERE parent_id = %s AND archived = false
            """, (page_id,))
            blocks = cur.fetchall()
            all_text = " ".join(json.dumps(b[0]) for b in blocks if b[0]).lower()
            record("Notion page mentions speaker travel",
                   "speaker" in all_text or "g235" in all_text or "alex" in all_text,
                   "No speaker info in page")
            record("Notion page mentions AI videos or resources",
                   "deepseek" in all_text or "ai" in all_text or "video" in all_text,
                   "No video resources in page")
    except Exception as e:
        record("Notion page check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: Google Calendar Conference Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-15'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()

    conf_events = [e for e in events if
                   any(kw in (e[0] or "").lower() for kw in
                       ["conference", "opening", "session", "closing", "forum"])]
    record("At least 3 conference day events scheduled", len(conf_events) >= 3,
           f"Found {len(conf_events)} conference events. All events: {[e[0] for e in events]}")

    # Check Day 1 event on March 12
    day1_events = [e for e in events if e[1] and e[1].month == 3 and e[1].day == 12]
    record("Conference Day 1 event on 2026-03-12", len(day1_events) >= 1,
           f"Events on Mar 12: {[e[0] for e in day1_events]}")

    # Check Day 3 event on March 14
    day3_events = [e for e in events if e[1] and e[1].month == 3 and e[1].day == 14]
    record("Conference Day 3 event on 2026-03-14", len(day3_events) >= 1,
           f"Events on Mar 14: {[e[0] for e in day3_events]}")

    cur.close()
    conn.close()


def check_emails():
    print("\n=== Check 4: Speaker Confirmation Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%conference%' OR subject ILIKE '%travel%' OR subject ILIKE '%confirmation%'
        """)
        emails = cur.fetchall()
        record("At least 3 confirmation emails sent", len(emails) >= 3,
               f"Found {len(emails)} conference emails")

        # Check emails for each speaker
        all_to = " ".join(json.dumps(e[1]) if e[1] else "" for e in emails).lower()
        record("Email sent to Alex Kim (alex.kim@tech.edu)",
               "alex.kim" in all_to, f"Recipients: {all_to[:200]}")
        record("Email sent to James Wu (james.wu@university.edu)",
               "james.wu" in all_to, f"Recipients: {all_to[:200]}")

        if emails:
            sample_body = (emails[0][2] or "").lower()
            record("Email body mentions train details",
                   "g235" in sample_body or "g168" in sample_body or "17:30" in sample_body,
                   "No train details in email body")
    except Exception as e:
        record("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

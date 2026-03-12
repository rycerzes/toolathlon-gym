"""Evaluation for yf-macro-investment-report."""
import argparse, os, sys


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Macro_Investment.xlsx")
    if not os.path.exists(path):
        return ["Macro_Investment.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Stock Performance sheet
        rows = load_sheet_rows(wb, "Stock Performance")
        if rows is None:
            errors.append("Sheet 'Stock Performance' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 5:
                errors.append(f"Stock Performance has {len(data_rows)} rows, expected 5")
            symbols = {str(r[0]).strip().upper() for r in data_rows if r[0]}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols:
                    errors.append(f"Symbol {sym} missing from Stock Performance")
            # Check specific values
            for r in data_rows:
                if r[0] and str(r[0]).strip().upper() == "XOM":
                    if len(r) > 5 and not num_close(r[5], 30.22):
                        errors.append(f"XOM Return_90d_Pct={r[5]}, expected ~30.22")
                if r[0] and str(r[0]).strip().upper() == "JNJ":
                    if len(r) > 5 and not num_close(r[5], 19.30):
                        errors.append(f"JNJ Return_90d_Pct={r[5]}, expected ~19.30")

        # Macro Context sheet
        rows2 = load_sheet_rows(wb, "Macro Context")
        if rows2 is None:
            errors.append("Sheet 'Macro Context' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 7:
                errors.append(f"Macro Context has {len(data_rows2)} rows, expected 7")
            lookup = {}
            for r in data_rows2:
                if r[0]:
                    lookup[str(r[0]).strip().lower()] = r
            if "us_10y_yield" in lookup:
                if not num_close(lookup["us_10y_yield"][1], 4.25, abs_tol=0.1):
                    errors.append(f"us_10y_yield value={lookup['us_10y_yield'][1]}, expected 4.25")
            if "cpi_yoy" in lookup:
                if not num_close(lookup["cpi_yoy"][1], 3.1, abs_tol=0.1):
                    errors.append(f"cpi_yoy value={lookup['cpi_yoy'][1]}, expected 3.1")

        # Portfolio Summary sheet
        rows3 = load_sheet_rows(wb, "Portfolio Summary")
        if rows3 is None:
            errors.append("Sheet 'Portfolio Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            lookup3 = {str(r[0]).strip().lower(): r[1] for r in data_rows3 if r[0]}
            if "average_90d_return" in lookup3:
                if not num_close(lookup3["average_90d_return"], 6.44):
                    errors.append(f"Average_90d_Return={lookup3['average_90d_return']}, expected ~6.44")
            else:
                errors.append("Average_90d_Return not found")
            if "best_performer" in lookup3:
                if str(lookup3["best_performer"]).strip().upper() != "XOM":
                    errors.append(f"Best_Performer={lookup3['best_performer']}, expected XOM")
            else:
                errors.append("Best_Performer not found")
            if "worst_performer" in lookup3:
                if str(lookup3["worst_performer"]).strip().upper() != "JPM":
                    errors.append(f"Worst_Performer={lookup3['worst_performer']}, expected JPM")
            else:
                errors.append("Worst_Performer not found")
            if "avg_trailing_pe" in lookup3:
                if not num_close(lookup3["avg_trailing_pe"], 23.43):
                    errors.append(f"Avg_Trailing_PE={lookup3['avg_trailing_pe']}, expected ~23.43")
            if "macro_risk_score" in lookup3:
                if not num_close(lookup3["macro_risk_score"], 3, abs_tol=1):
                    errors.append(f"Macro_Risk_Score={lookup3['macro_risk_score']}, expected ~3")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    path = os.path.join(agent_workspace, "Investment_Report.docx")
    if not os.path.exists(path):
        return ["Investment_Report.docx not found"]
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join([p.text for p in doc.paragraphs]).lower()
        if len(text) < 200:
            errors.append(f"Investment_Report.docx too short ({len(text)} chars)")
        for kw in ["macro", "portfolio", "sector"]:
            if kw not in text:
                errors.append(f"Investment_Report.docx missing keyword '{kw}'")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

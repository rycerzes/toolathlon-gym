"""
Evaluation for sf-hr-performance-review-gform-email task.

Checks:
1. Performance_Review_Setup.xlsx exists with "Performance Analysis" sheet
2. Sheet has 7 department rows
3. Engineering row has Employee_Count=7096, Avg_Performance~3.21
4. GForm "Annual Performance Review Form" exists with at least 4 questions
5. GForm has department and rating questions
6. Email sent to hr_director@company.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

EXPECTED_DEPARTMENTS = {"Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Excel Performance_Review_Setup.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Performance_Review_Setup.xlsx")
    if not os.path.exists(xlsx_path):
        record("Performance_Review_Setup.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Performance_Review_Setup.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    # Find Performance Analysis sheet
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    perf_idx = next((i for i, s in enumerate(sheet_names_lower) if "performance" in s or "analysis" in s), None)
    if perf_idx is None:
        perf_idx = 0  # use first sheet

    ws = wb[wb.sheetnames[perf_idx]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        record("Sheet has data", False, "Sheet is empty")
        return

    data_rows = [r for r in rows[1:] if any(c for c in r)]
    record("Sheet has 7 department rows", len(data_rows) >= 7,
           f"Found {len(data_rows)} data rows")

    # Check headers
    headers = [str(c).lower().strip() if c else "" for c in rows[0]]
    has_dept = any("dept" in h or "department" in h for h in headers)
    has_count = any("count" in h or "employee" in h for h in headers)
    has_avg = any("avg" in h or "average" in h or "performance" in h for h in headers)

    record("Has Department column", has_dept, f"Headers: {rows[0]}")
    record("Has Employee_Count column", has_count, f"Headers: {rows[0]}")
    record("Has Avg_Performance column", has_avg, f"Headers: {rows[0]}")

    # Check for Engineering row with correct data
    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    has_engineering = "engineering" in all_text
    record("Contains Engineering department data", has_engineering, "No Engineering row found")

    # Check all expected departments present
    found_depts = set()
    for row in data_rows:
        for cell in row:
            if cell and str(cell).strip() in EXPECTED_DEPARTMENTS:
                found_depts.add(str(cell).strip())
    record("All 7 departments present", len(found_depts) >= 6,
           f"Found departments: {found_depts}")

    # Verify Engineering count (7096) and avg (~3.21)
    engineering_row = None
    dept_col = next((i for i, h in enumerate(headers) if "dept" in h or "department" in h), 0)
    for row in data_rows:
        if row[dept_col] and str(row[dept_col]).strip() == "Engineering":
            engineering_row = row
            break

    if engineering_row:
        # Find employee count column
        count_col = next((i for i, h in enumerate(headers) if "count" in h or "employee" in h), 1)
        avg_col = next((i for i, h in enumerate(headers) if "avg" in h or "average" in h), 2)

        emp_count = engineering_row[count_col] if count_col < len(engineering_row) else None
        avg_perf = engineering_row[avg_col] if avg_col < len(engineering_row) else None

        try:
            emp_count_val = int(emp_count) if emp_count is not None else 0
            record("Engineering Employee_Count is 7096", emp_count_val == 7096,
                   f"Got {emp_count_val}")
        except (TypeError, ValueError):
            record("Engineering Employee_Count is 7096", False, f"Could not parse: {emp_count}")

        try:
            avg_val = float(avg_perf) if avg_perf is not None else 0
            record("Engineering Avg_Performance is ~3.21", abs(avg_val - 3.21) < 0.05,
                   f"Got {avg_val}")
        except (TypeError, ValueError):
            record("Engineering Avg_Performance is ~3.21", False, f"Could not parse: {avg_perf}")


def check_gform():
    print("\n=== Check 2: GForm Annual Performance Review Form ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    review_form = None
    for form_id, title in forms:
        if "performance review" in (title or "").lower() or "annual performance" in (title or "").lower():
            review_form = (form_id, title)
            break

    record("Annual Performance Review Form exists", review_form is not None,
           f"Forms found: {[f[1] for f in forms]}")

    if review_form:
        form_id, title = review_form
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        record("Form has at least 4 questions", q_count >= 4,
               f"Found {q_count} questions")

        # Check for department and rating questions
        cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position", (form_id,))
        questions = cur.fetchall()
        q_titles_lower = [q[0].lower() for q in questions]

        has_dept_q = any("department" in t for t in q_titles_lower)
        has_rating_q = any("rating" in t or "performance" in t for t in q_titles_lower)
        has_achievement_q = any("achievement" in t or "accomplishment" in t for t in q_titles_lower)

        record("Form has Department question", has_dept_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Performance Rating question", has_rating_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Key Achievements question", has_achievement_q,
               f"Questions: {[q[0] for q in questions]}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email to hr_director@company.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "hr_director@company.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email sent to hr_director@company.com", matching is not None,
           f"Messages found: {len(messages)}")

    if matching:
        subject, _, _, body_text = matching
        all_text = ((subject or "") + " " + (body_text or "")).lower()
        has_hr_content = (
            "performance review" in all_text or "annual" in all_text or
            "department" in all_text or "engineering" in all_text
        )
        record("Email mentions performance review", has_hr_content,
               f"Subject: {subject}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

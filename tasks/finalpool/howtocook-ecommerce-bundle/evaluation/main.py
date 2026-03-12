"""
Evaluation script for howtocook-ecommerce-bundle task.

Dynamically computes expected values from PostgreSQL:
  - wc.products: Home Appliances category products
  - yf.stock_prices: Gold futures (GC=F) last 5 trading days

Recipe Bundles sheet is checked structurally (3 rows, correct columns,
bundle price = product price * 0.85) since recipe selection is flexible.

Falls back to static groundtruth Excel if PostgreSQL is unavailable.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0

DB_CONFIG = dict(host=os.environ.get('PGHOST', 'localhost'), port=5432, database='toolathlon_gym',
                 user='postgres', password='postgres')

# Valid HowToCook categories (Chinese)
VALID_RECIPE_CATEGORIES = {
    '水产', '早餐', '调料', '甜品', '饮品', '荤菜',
    '半成品加工', '汤', '主食', '素菜'
}

# Category mapping for task instructions:
#   drinks -> 饮品, staple food -> 主食, meat dishes -> 荤菜
REQUIRED_CATEGORIES = {'饮品', '主食', '荤菜'}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected_from_db():
    """Compute expected values from PostgreSQL for Store Products and Gold Trend."""
    try:
        import psycopg2
    except ImportError:
        return None

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    # --- Home Appliances products ---
    cur.execute("""
        SELECT name, price, stock_quantity
        FROM wc.products
        WHERE categories::text LIKE '%%Home Appliances%%'
        ORDER BY price::numeric ASC
    """)
    products = cur.fetchall()
    if not products:
        conn.close()
        return None

    store_products = []
    product_prices_by_name = {}
    for row in products:
        name = str(row[0]).strip()
        price = round(float(row[1]), 2)
        stock = int(row[2]) if row[2] is not None else 0
        store_products.append((name, price, stock))
        product_prices_by_name[name.lower()] = price

    # --- Gold prices (last 5 trading days) ---
    cur.execute("""
        SELECT date, close
        FROM yf.stock_prices
        WHERE symbol = 'GC=F'
        ORDER BY date DESC
        LIMIT 5
    """)
    gold_rows = cur.fetchall()
    if len(gold_rows) < 5:
        conn.close()
        return None

    gold_trend = []
    for row in gold_rows:
        date_str = str(row[0])
        close_price = round(float(row[1]), 2)
        gold_trend.append((date_str, close_price))

    conn.close()

    return {
        'store_products': store_products,
        'product_prices_by_name': product_prices_by_name,
        'gold_trend': gold_trend,
    }


def check_excel(agent_workspace, expected, groundtruth_workspace=None):
    """Check the agent's Excel output."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Bundle_Pricing.xlsx")
    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    expected_sheets = ["Store Products", "Gold Trend", "Recipe Bundles"]
    for sheet_name in expected_sheets:
        found = get_sheet(agent_wb, sheet_name) is not None
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {agent_wb.sheetnames}")

    use_db = expected is not None

    # ── Sheet 1: Store Products ──
    print("\n--- Store Products ---")
    agent_ws = get_sheet(agent_wb, "Store Products")
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        if use_db:
            exp_products = expected['store_products']
            check("Store Products row count",
                  len(agent_rows) == len(exp_products),
                  f"Expected {len(exp_products)}, got {len(agent_rows)}")

            # Check headers
            agent_headers = [c.value for c in agent_ws[1]]
            check("Store Products has 3+ columns",
                  len(agent_headers) >= 3,
                  f"Got {len(agent_headers)} columns: {agent_headers}")

            # Check sort order (by Price ascending)
            if len(agent_rows) >= 2:
                prices = []
                for row in agent_rows:
                    try:
                        prices.append(float(row[1]))
                    except (TypeError, ValueError, IndexError):
                        prices.append(0)
                check("Store Products sorted by Price ascending",
                      prices == sorted(prices),
                      f"Prices: {prices}")

            # Spot-check specific products by price matching
            for exp_name, exp_price, exp_stock in exp_products:
                matched = False
                for row in agent_rows:
                    if row and len(row) >= 3:
                        try:
                            if num_close(row[1], exp_price, 0.05):
                                matched = True
                                check(f"Product at price {exp_price} stock",
                                      num_close(row[2], exp_stock, 1),
                                      f"Expected stock {exp_stock}, got {row[2]}")
                                break
                        except (TypeError, ValueError):
                            continue
                if not matched:
                    # Try matching by name substring
                    for row in agent_rows:
                        if row and row[0] and exp_name[:30].lower() in str(row[0]).lower():
                            matched = True
                            check(f"Product '{exp_name[:40]}' price",
                                  num_close(row[1], exp_price, 0.5),
                                  f"Expected {exp_price}, got {row[1]}")
                            break
        else:
            # Fallback: use groundtruth
            gt_file = os.path.join(groundtruth_workspace, "Bundle_Pricing.xlsx")
            if os.path.isfile(gt_file):
                gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
                gt_ws = get_sheet(gt_wb, "Store Products")
                if gt_ws:
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    check("Store Products row count",
                          len(agent_rows) == len(gt_rows),
                          f"Expected {len(gt_rows)}, got {len(agent_rows)}")

    # ── Sheet 2: Gold Trend ──
    print("\n--- Gold Trend ---")
    agent_ws = get_sheet(agent_wb, "Gold Trend")
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Gold Trend has 5 rows",
              len(agent_rows) == 5,
              f"Got {len(agent_rows)} rows")

        agent_headers = [c.value for c in agent_ws[1]]
        check("Gold Trend has 2+ columns",
              len(agent_headers) >= 2,
              f"Got {len(agent_headers)} columns: {agent_headers}")

        if use_db:
            exp_gold = expected['gold_trend']

            # Check most recent date's close price
            if len(agent_rows) >= 1 and len(exp_gold) >= 1:
                agent_first_close = agent_rows[0][1] if len(agent_rows[0]) >= 2 else None
                exp_first_close = exp_gold[0][1]
                check("Gold Trend most recent close price",
                      num_close(agent_first_close, exp_first_close, 5.0),
                      f"Expected ~{exp_first_close}, got {agent_first_close}")

            # Check descending date order
            if len(agent_rows) >= 2:
                first_date = str(agent_rows[0][0])
                second_date = str(agent_rows[1][0])
                check("Gold Trend sorted by date descending",
                      first_date >= second_date,
                      f"First date={first_date}, second={second_date}")

            # Check last row's close price
            if len(agent_rows) >= 5 and len(exp_gold) >= 5:
                agent_last_close = agent_rows[4][1] if len(agent_rows[4]) >= 2 else None
                exp_last_close = exp_gold[4][1]
                check("Gold Trend oldest close price",
                      num_close(agent_last_close, exp_last_close, 5.0),
                      f"Expected ~{exp_last_close}, got {agent_last_close}")

    # ── Sheet 3: Recipe Bundles ──
    print("\n--- Recipe Bundles ---")
    agent_ws = get_sheet(agent_wb, "Recipe Bundles")
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Recipe Bundles has exactly 3 rows",
              len(agent_rows) == 3,
              f"Got {len(agent_rows)} rows")

        agent_headers = [c.value for c in agent_ws[1]]
        check("Recipe Bundles has 5+ columns",
              len(agent_headers) >= 5,
              f"Got {len(agent_headers)} columns: {agent_headers}")

        # Check each bundle row
        seen_categories = set()
        for i, row in enumerate(agent_rows):
            if not row or len(row) < 5:
                check(f"Bundle row {i+1} has enough columns", False,
                      f"Row has {len(row) if row else 0} columns")
                continue

            recipe_name, recipe_cat, difficulty, paired_product, bundle_price = (
                row[0], row[1], row[2], row[3], row[4]
            )

            # Recipe name should be non-empty Chinese text
            check(f"Bundle {i+1} has a recipe name",
                  recipe_name is not None and len(str(recipe_name).strip()) > 0,
                  f"Got: {recipe_name}")

            # Recipe category should be a valid HowToCook category
            cat_str = str(recipe_cat).strip() if recipe_cat else ""
            check(f"Bundle {i+1} category is valid Chinese category",
                  cat_str in VALID_RECIPE_CATEGORIES,
                  f"Got '{cat_str}', expected one of {VALID_RECIPE_CATEGORIES}")
            seen_categories.add(cat_str)

            # Difficulty should be present
            check(f"Bundle {i+1} has difficulty",
                  difficulty is not None and len(str(difficulty).strip()) > 0,
                  f"Got: {difficulty}")

            # Paired product should be non-empty
            check(f"Bundle {i+1} has paired product name",
                  paired_product is not None and len(str(paired_product).strip()) > 0,
                  f"Got: {paired_product}")

            # Bundle price should be product price * 0.85
            if use_db and paired_product:
                product_prices = expected['product_prices_by_name']
                product_name_lower = str(paired_product).strip().lower()
                matched_price = None
                for pname, pprice in product_prices.items():
                    if pname == product_name_lower or pname[:30] in product_name_lower or product_name_lower[:30] in pname:
                        matched_price = pprice
                        break

                if matched_price is not None:
                    expected_bundle = round(matched_price * 0.85, 2)
                    check(f"Bundle {i+1} price = product price * 0.85",
                          num_close(bundle_price, expected_bundle, 0.5),
                          f"Expected ~{expected_bundle} (product price {matched_price} * 0.85), got {bundle_price}")
                else:
                    # Can't find product, just check bundle price is reasonable
                    check(f"Bundle {i+1} has numeric bundle price",
                          bundle_price is not None and float(bundle_price) > 0,
                          f"Got: {bundle_price}")
            else:
                check(f"Bundle {i+1} has numeric bundle price",
                      bundle_price is not None and float(bundle_price) > 0,
                      f"Got: {bundle_price}")

        # Check that all 3 required categories are represented
        check("Recipe Bundles covers drinks (饮品) category",
              '饮品' in seen_categories,
              f"Categories found: {seen_categories}")
        check("Recipe Bundles covers staple food (主食) category",
              '主食' in seen_categories,
              f"Categories found: {seen_categories}")
        check("Recipe Bundles covers meat dishes (荤菜) category",
              '荤菜' in seen_categories,
              f"Categories found: {seen_categories}")

    return True


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    db_expected = compute_expected_from_db()
    use_db = db_expected is not None

    if use_db:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
    else:
        print("INFO: Falling back to static groundtruth Excel file")

    check_excel(agent_workspace, db_expected, groundtruth_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Source: {'postgresql' if use_db else 'groundtruth_excel'}")
    print(f"  Overall: {'PASS' if FAIL_COUNT == 0 else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": FAIL_COUNT == 0,
            "source": "postgresql" if use_db else "groundtruth_excel",
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return FAIL_COUNT == 0, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

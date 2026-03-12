"""Evaluation for playwright-sf-hr-benefits-survey-gform-excel."""
import argparse
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Benefits_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Benefits_Analysis.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Competitor Comparison
        rows = load_sheet_rows(wb, "Competitor Comparison")
        if rows is None:
            errors.append("Sheet 'Competitor Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 7:
                errors.append(f"Competitor Comparison has {len(data_rows)} rows, expected 7")
            companies = {str(r[0]).strip().lower() for r in data_rows if r[0]}
            for c in ["acme technologies", "frontier analytics", "our company"]:
                if c not in companies:
                    errors.append(f"Company '{c}' missing from Competitor Comparison")
            # Check Our Company values
            ours = [r for r in data_rows if r[0] and "our company" in str(r[0]).lower()]
            if ours:
                if not num_close(ours[0][1], 80, 1):
                    errors.append(f"Our Health Insurance={ours[0][1]}, expected 80")
                if not num_close(ours[0][2], 20, 1):
                    errors.append(f"Our PTO Days={ours[0][2]}, expected 20")

        # Department Satisfaction
        rows2 = load_sheet_rows(wb, "Department Satisfaction")
        if rows2 is None:
            errors.append("Sheet 'Department Satisfaction' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 7:
                errors.append(f"Department Satisfaction has {len(data_rows2)} rows, expected 7")
            depts = {str(r[0]).strip().lower() for r in data_rows2 if r[0]}
            for d in ["engineering", "finance", "hr", "operations", "r&d", "sales", "support"]:
                if d not in depts:
                    errors.append(f"Department '{d}' missing")
            # Check Engineering satisfaction
            eng = [r for r in data_rows2 if r[0] and str(r[0]).strip().lower() == "engineering"]
            if eng:
                if not num_close(eng[0][1], 6.58, 0.15):
                    errors.append(f"Engineering satisfaction={eng[0][1]}, expected ~6.58")
                if len(eng[0]) >= 5:
                    if str(eng[0][4]).strip().lower() != "high":
                        errors.append(f"Engineering rating={eng[0][4]}, expected High")
            # Check Operations -> Moderate
            ops = [r for r in data_rows2 if r[0] and str(r[0]).strip().lower() == "operations"]
            if ops and len(ops[0]) >= 5:
                if str(ops[0][4]).strip().lower() != "moderate":
                    errors.append(f"Operations rating={ops[0][4]}, expected Moderate")

        # Gap Analysis
        rows3 = load_sheet_rows(wb, "Gap Analysis")
        if rows3 is None:
            errors.append("Sheet 'Gap Analysis' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 3:
                errors.append(f"Gap Analysis has {len(data_rows3)} rows, expected 3")
            # Check Health Insurance gap
            health = [r for r in data_rows3 if r[0] and "health" in str(r[0]).lower()]
            if health:
                if not num_close(health[0][2], 85.8, 1.0):
                    errors.append(f"Health Market Avg={health[0][2]}, expected ~85.8")
                if len(health[0]) >= 5:
                    if str(health[0][4]).strip().lower() != "high":
                        errors.append(f"Health Priority={health[0][4]}, expected High")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gform.forms ORDER BY id DESC LIMIT 5")
        forms = cur.fetchall()
        if not forms:
            errors.append("No Google Forms found")
        else:
            found = any("benefit" in (f[1] or "").lower() or "survey" in (f[1] or "").lower()
                        for f in forms)
            if not found:
                errors.append(f"No benefits survey form found (forms: {[f[1] for f in forms]})")
            else:
                form_id = [f[0] for f in forms
                           if "benefit" in (f[1] or "").lower() or "survey" in (f[1] or "").lower()][0]
                cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
                q_count = cur.fetchone()[0]
                if q_count < 5:
                    errors.append(f"Form has {q_count} questions, expected at least 5")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject FROM email.messages
            WHERE to_addr::text ILIKE '%hr_leadership@company.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            errors.append("No email found to hr_leadership@company.com")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Google Form...")
    errs = check_gform()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

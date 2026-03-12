"""
Evaluation script for sf-sales-shipping-word task.

Checks:
1. Excel Shipping_Analysis.xlsx with "By Ship Mode" sheet
2. Word document Shipping_Report.docx exists and contains key content

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:200] + "...") if len(detail) > 200 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    cur.execute("""
        SELECT "SHIP_MODE", COUNT(*) as orders,
               ROUND(AVG(("SHIP_DATE"::date - "ORDER_DATE"::date))::numeric, 1) as avg_days,
               ROUND(SUM("TOTAL_AMOUNT")::numeric, 2) as revenue,
               ROUND(AVG("TOTAL_AMOUNT")::numeric, 2) as avg_order
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        WHERE "STATUS" = 'Delivered'
        GROUP BY "SHIP_MODE" ORDER BY COUNT(*) DESC
    """)
    shipping_rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*), ROUND(SUM("TOTAL_AMOUNT")::numeric, 2)
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        WHERE "STATUS" = 'Delivered'
    """)
    totals = cur.fetchone()
    conn.close()

    return {
        "shipping": shipping_rows,
        "total_orders": totals[0],
        "total_revenue": float(totals[1]),
    }


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Shipping_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Sheet 'By Ship Mode' exists", get_sheet(wb, "By Ship Mode") is not None,
          f"Found: {wb.sheetnames}")

    ws = get_sheet(wb, "By Ship Mode")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["shipping"]
        check("By Ship Mode row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        agent_by_mode = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_mode[str(row[0]).strip().lower()] = row

        for exp_row in exp:
            mode = exp_row[0]
            agent_row = agent_by_mode.get(mode.lower())
            if agent_row:
                check(f"'{mode}' Order_Count",
                      num_close(agent_row[1], exp_row[1], 5),
                      f"Expected {exp_row[1]}, got {agent_row[1]}")
                check(f"'{mode}' Avg_Delivery_Days",
                      num_close(agent_row[2], float(exp_row[2]), 0.5),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"'{mode}' Total_Revenue",
                      num_close(agent_row[3], float(exp_row[3]), 500),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
                check(f"'{mode}' Avg_Order_Value",
                      num_close(agent_row[4], float(exp_row[4]), 5),
                      f"Expected {exp_row[4]}, got {agent_row[4]}")
            else:
                check(f"'{mode}' found in output", False, "Not in agent output")

        # Check sort order (by Order_Count descending)
        if len(agent_rows) >= 2:
            counts = [int(r[1]) for r in agent_rows if r and r[1] is not None]
            check("Sorted by Order_Count descending",
                  all(counts[i] >= counts[i+1] for i in range(len(counts)-1)),
                  f"Counts: {counts}")


def check_word(agent_workspace, expected):
    print("\n=== Checking Word Document ===")
    agent_file = os.path.join(agent_workspace, "Shipping_Report.docx")
    check("Word file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        from docx import Document
        doc = Document(agent_file)
    except ImportError:
        # If python-docx not available, just check file exists
        check("python-docx available", False, "Cannot parse Word doc without python-docx")
        return
    except Exception as e:
        check("Word file readable", False, str(e))
        return

    # Check heading
    full_text = "\n".join([p.text for p in doc.paragraphs])
    check("Document contains 'Shipping Performance Report'",
          "shipping performance report" in full_text.lower(),
          f"Text preview: {full_text[:100]}")

    # Check ship modes mentioned
    if expected:
        for exp_row in expected["shipping"]:
            mode = exp_row[0]
            check(f"Document mentions '{mode}'",
                  mode.lower() in full_text.lower(),
                  f"'{mode}' not found in document text")

        # Check total orders mentioned
        total_str = str(expected["total_orders"])
        check("Document mentions total order count",
              total_str in full_text,
              f"Expected '{total_str}' in text")


def check_excel_gt(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Shipping_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Shipping_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(agent_file))
    check("Groundtruth file exists", os.path.isfile(gt_file))
    if not os.path.isfile(agent_file) or not os.path.isfile(gt_file):
        return
    agent_wb = openpyxl.load_workbook(agent_file)
    gt_wb = openpyxl.load_workbook(gt_file)
    check("Sheet 'By Ship Mode' exists", get_sheet(agent_wb, "By Ship Mode") is not None)
    a_ws = get_sheet(agent_wb, "By Ship Mode")
    g_ws = get_sheet(gt_wb, "By Ship Mode")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Row count matches", len(a_rows) == len(g_rows),
              f"Expected {len(g_rows)}, got {len(a_rows)}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        check_excel(agent_workspace, expected)
        check_word(agent_workspace, expected)
    else:
        print("INFO: Falling back to groundtruth")
        check_excel_gt(agent_workspace, groundtruth_workspace)
        check_word(agent_workspace, None)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if FAIL_COUNT == 0 else 'FAIL'}")

    if res_log_file:
        result = {"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": FAIL_COUNT == 0}
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return FAIL_COUNT == 0, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

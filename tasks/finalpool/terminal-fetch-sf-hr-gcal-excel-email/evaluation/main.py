"""Evaluation for terminal-fetch-sf-hr-gcal-excel-email.
Checks:
1. Compensation_Benchmark_Report.xlsx with 3 sheets and correct data
2. Google Calendar events for 7 department review meetings
3. Email sent to hr_team@company.com
4. compensation_analysis.py script exists
"""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

DEPARTMENTS = {"Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').replace('%', '').strip())
    except Exception:
        return default


def check_excel(workspace):
    print("\n=== Check 1: Compensation_Benchmark_Report.xlsx ===")
    path = os.path.join(workspace, "Compensation_Benchmark_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # Department_Analysis sheet
    da_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s or "analysis" in s), 0)
    ws1 = wb[sheets[da_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data_rows1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Department_Analysis has 7 rows", len(data_rows1) >= 7, f"Found {len(data_rows1)}")

    # Check headers
    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        has_salary = any("salary" in h for h in headers)
        has_satisfaction = any("satisf" in h for h in headers)
        check("Has salary column", has_salary, f"Headers: {rows1[0]}")
        check("Has satisfaction column", has_satisfaction, f"Headers: {rows1[0]}")

    # Check department names
    found_depts = set()
    for row in data_rows1:
        for cell in row:
            if cell and str(cell).strip() in DEPARTMENTS:
                found_depts.add(str(cell).strip())
    check("All 7 departments present in Dept Analysis", len(found_depts) >= 6,
          f"Found: {found_depts}")

    # Check Engineering salary dynamically from DB
    expected_eng_salary = 58991.61  # fallback
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT AVG(salary) FROM sf_data.employees WHERE department = 'Engineering'")
        result = cur.fetchone()
        if result and result[0] is not None:
            expected_eng_salary = float(result[0])
        cur.close()
        conn.close()
    except Exception:
        pass
    for row in data_rows1:
        if row[0] and str(row[0]).strip() == "Engineering":
            avg_sal = safe_float(row[2])
            check("Engineering avg salary correct", avg_sal is not None and abs(avg_sal - expected_eng_salary) < 500,
                  f"Got {avg_sal}, expected ~{expected_eng_salary:.2f}")
            break

    # Salary_Benchmark sheet
    sb_idx = next((i for i, s in enumerate(sheets_lower) if "benchmark" in s), 1)
    if sb_idx < len(sheets):
        ws2 = wb[sheets[sb_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Salary_Benchmark has 7 rows", len(data_rows2) >= 7, f"Found {len(data_rows2)}")

        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            has_gap = any("gap" in h for h in headers2)
            has_review = any("review" in h for h in headers2)
            check("Has Gap_Pct column", has_gap, f"Headers: {rows2[0]}")
            check("Has Needs_Review column", has_review, f"Headers: {rows2[0]}")

    # Review_Summary sheet
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s or "summary" in s), 2)
    if rs_idx < len(sheets):
        ws3 = wb[sheets[rs_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Review_Summary has 7 rows", len(data_rows3) >= 7, f"Found {len(data_rows3)}")


def check_gcal():
    print("\n=== Check 2: Google Calendar Review Meetings ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT summary, start_datetime, description FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    check("At least 7 review meeting events", len(events) >= 7, f"Found {len(events)} events")

    if events:
        summaries = " ".join(str(e[0]) for e in events).lower()
        check("Events mention salary review", "salary review" in summaries or "review meeting" in summaries,
              f"Summaries: {summaries[:150]}")

        dept_found = set()
        for event in events:
            summary = str(event[0]).lower() if event[0] else ""
            for dept in DEPARTMENTS:
                if dept.lower() in summary:
                    dept_found.add(dept)
        check("Events cover at least 6 departments", len(dept_found) >= 6,
              f"Departments in events: {dept_found}")

        # Check events are in March 2026
        march_events = [e for e in events if e[1] and e[1].month == 3 and e[1].year == 2026]
        check("Events scheduled in March 2026", len(march_events) >= 7,
              f"Found {len(march_events)} March 2026 events")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email to hr_team@company.com ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "hr_team@company.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    check("Email sent to hr_team@company.com", matching is not None,
          f"Messages found: {len(messages)}")

    if matching:
        subject, _, _, body_text = matching
        all_text = ((subject or "") + " " + (body_text or "")).lower()
        check("Email mentions compensation or benchmark", "compensation" in all_text or "benchmark" in all_text,
              f"Subject: {subject}")
        check("Email mentions review meetings scheduled", "meeting" in all_text or "scheduled" in all_text or "march" in all_text,
              f"Body snippet: {all_text[:100]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: compensation_analysis.py ===")
    path = os.path.join(workspace, "compensation_analysis.py")
    check("compensation_analysis.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative salary values
    path = os.path.join(workspace, "Compensation_Benchmark_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        check("No negative salary/satisfaction values in Excel", False,
                              f"Found {cell} in sheet {sheet_name}")
                        return
        check("No negative salary/satisfaction values in Excel", True)

    # GCal: no review events on weekends
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (lower(summary) LIKE '%%salary review%%' OR lower(summary) LIKE '%%review meeting%%')
              AND EXTRACT(DOW FROM start_datetime) IN (0, 6)
        """)
        weekend_count = cur.fetchone()[0]
        check("No review meetings on weekends", weekend_count == 0,
              f"Found {weekend_count} weekend events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

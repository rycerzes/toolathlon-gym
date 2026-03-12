"""Evaluation for canvas-course-comparison-word."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected course comparison data from DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        WITH course_data AS (
            SELECT
                SPLIT_PART(c.course_code, '-', 1) as prefix,
                c.course_code,
                c.name,
                c.total_students,
                COUNT(DISTINCT a.id) as assignment_count,
                ROUND(AVG(s.score)::numeric, 2) as avg_grade
            FROM canvas.courses c
            LEFT JOIN canvas.assignments a ON a.course_id = c.id
            LEFT JOIN canvas.submissions s ON s.assignment_id = a.id AND s.score IS NOT NULL
            WHERE c.course_code LIKE '%%2013J' OR c.course_code LIKE '%%2014J'
            GROUP BY c.id, c.course_code, c.name, c.total_students
        )
        SELECT
            f13.prefix,
            f13.name as name_2013,
            f13.total_students as enroll_2013,
            f14.total_students as enroll_2014,
            f13.assignment_count as assign_2013,
            f14.assignment_count as assign_2014,
            f13.avg_grade as grade_2013,
            f14.avg_grade as grade_2014
        FROM course_data f13
        JOIN course_data f14 ON f13.prefix = f14.prefix
        WHERE f13.course_code LIKE '%%2013J' AND f14.course_code LIKE '%%2014J'
        ORDER BY f13.prefix
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Year_Over_Year_Comparison.xlsx."""
    print("\n=== Checking Year_Over_Year_Comparison.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Year_Over_Year_Comparison.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    expected = get_expected_data()

    # Check Course Comparison sheet
    comp_sheet = None
    for name in wb.sheetnames:
        if "comparison" in name.lower() or "course" in name.lower():
            comp_sheet = wb[name]
            break
    if comp_sheet is None:
        record("Sheet 'Course Comparison' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Course Comparison' exists", True)
        rows = list(comp_sheet.iter_rows(min_row=2, values_only=True))
        record("Course Comparison has correct row count",
               len(rows) == len(expected),
               f"Expected {len(expected)}, got {len(rows)}")

        agent_lookup = {}
        for r in rows:
            if r and r[0]:
                key = str(r[0]).strip().lower()
                agent_lookup[key] = r

        for exp_row in expected:
            prefix = exp_row[0]
            course_name_2013 = exp_row[1]
            # Try matching on prefix or course name
            matched = None
            for key, r in agent_lookup.items():
                if prefix.lower() in key or course_name_2013.lower().split("(")[0].strip().lower() in key:
                    matched = r
                    break

            if matched is None:
                record(f"Course '{prefix}' found", False, "Missing")
                all_ok = False
                continue

            # Check enrollment values
            ok_e13 = num_close(matched[1], exp_row[2], 10)
            record(f"'{prefix}' Fall_2013_Enrollment", ok_e13,
                   f"Expected {exp_row[2]}, got {matched[1]}")
            if not ok_e13:
                all_ok = False

            ok_e14 = num_close(matched[2], exp_row[3], 10)
            record(f"'{prefix}' Fall_2014_Enrollment", ok_e14,
                   f"Expected {exp_row[3]}, got {matched[2]}")
            if not ok_e14:
                all_ok = False

            # Check avg grades
            ok_g13 = num_close(matched[6], exp_row[6], 2.0)
            record(f"'{prefix}' Fall_2013_Avg_Grade", ok_g13,
                   f"Expected {exp_row[6]}, got {matched[6]}")
            if not ok_g13:
                all_ok = False

            ok_g14 = num_close(matched[7], exp_row[7], 2.0)
            record(f"'{prefix}' Fall_2014_Avg_Grade", ok_g14,
                   f"Expected {exp_row[7]}, got {matched[7]}")
            if not ok_g14:
                all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)

    return all_ok


def check_word(agent_workspace):
    """Check Academic_Year_Comparison.docx."""
    print("\n=== Checking Academic_Year_Comparison.docx ===")
    from docx import Document

    docx_file = os.path.join(agent_workspace, "Academic_Year_Comparison.docx")
    if not os.path.isfile(docx_file):
        record("Word file exists", False, f"Not found: {docx_file}")
        return False
    record("Word file exists", True)

    try:
        doc = Document(docx_file)
    except Exception as e:
        record("Word readable", False, str(e))
        return False

    all_text = " ".join(p.text.lower() for p in doc.paragraphs)

    record("Word mentions '2013'", "2013" in all_text, "No mention of '2013'")
    record("Word mentions '2014'", "2014" in all_text, "No mention of '2014'")
    record("Word mentions 'performance' or 'comparison' or 'review'",
           "performance" in all_text or "comparison" in all_text or "review" in all_text,
           "No relevant keywords found")

    return True


def check_gsheet():
    """Check Google Sheet."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id FROM gsheet.spreadsheets WHERE title ILIKE '%academic%' OR title ILIKE '%year%comparison%'")
    rows = cur.fetchall()
    found = len(rows) > 0
    record("GSheet with academic/comparison in title", found, "No matching spreadsheet found")

    cur.close()
    conn.close()
    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    word_ok = check_word(args.agent_workspace)

    db_fail_before = FAIL_COUNT
    gsheet_ok = check_gsheet()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")

    overall = excel_ok and word_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

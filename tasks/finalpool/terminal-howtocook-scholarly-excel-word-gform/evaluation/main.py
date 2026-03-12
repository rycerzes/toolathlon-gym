"""Evaluation for terminal-howtocook-scholarly-excel-word-gform.
Checks:
1. Nutrition_Program_Analysis.xlsx with 4 sheets and correct content
2. Nutrition_Program_Proposal.docx with required sections
3. Google Form survey with 4 questions
4. Python scripts exist (categorize_recipes.py, analyze_research.py, build_menus.py, validate_menus.py)
5. JSON output files exist
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except:
        return False


def check_excel(workspace):
    print("\n=== Check 1: Nutrition_Program_Analysis.xlsx ===")
    path = os.path.join(workspace, "Nutrition_Program_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Recipe_Database
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "recipe" in s and "database" in s), 0)
    ws1 = wb[sheets[rd_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Recipe_Database has 10+ recipes", len(data1) >= 10, f"Found {len(data1)}")
    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has dietary_tags column", any("dietary" in h or "tag" in h for h in headers), f"Headers: {headers}")
        check("Has evidence_score column", any("evidence" in h or "score" in h for h in headers), f"Headers: {headers}")
        check("Has prep_time column", any("prep" in h or "time" in h for h in headers), f"Headers: {headers}")

    # Research_Summary
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s or "summary" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Research_Summary has 4+ papers", len(data2) >= 4, f"Found {len(data2)}")
        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has confidence_level column", any("confidence" in h for h in headers2), f"Headers: {headers2}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Research mentions workplace nutrition", "workplace" in all_text2 or "employee" in all_text2,
              all_text2[:150])

    # Weekly_Menu
    wm_idx = next((i for i, s in enumerate(sheets_lower) if "weekly" in s or "menu" in s), 2)
    if wm_idx < len(sheets):
        ws3 = wb[sheets[wm_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Weekly_Menu has 5 days", len(data3) >= 5, f"Found {len(data3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Menu includes Monday", "monday" in all_text3)
        check("Menu includes Friday", "friday" in all_text3)
        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            check("Has dietary_compliance_pct column",
                  any("compliance" in h or "dietary" in h for h in headers3), f"Headers: {headers3}")
            check("Has est_calories column",
                  any("calori" in h for h in headers3), f"Headers: {headers3}")

    # Program_Budget
    pb_idx = next((i for i, s in enumerate(sheets_lower) if "budget" in s), 3)
    if pb_idx < len(sheets):
        ws4 = wb[sheets[pb_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Program_Budget has 5+ rows", len(data4) >= 5, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Budget includes ingredient costs", "ingredient" in all_text4 or "recipe" in all_text4)
        check("Budget includes workshop", "workshop" in all_text4)
        # Check total_cost calculation
        if rows4:
            headers4 = [str(c).lower() if c else "" for c in rows4[0]]
            cost_idx = next((i for i, h in enumerate(headers4) if "total_cost" in h or h == "total_cost"), -1)
            cpp_idx = next((i for i, h in enumerate(headers4) if "cost_per" in h), -1)
            part_idx = next((i for i, h in enumerate(headers4) if "participant" in h), -1)
            if cost_idx >= 0 and cpp_idx >= 0 and part_idx >= 0 and data4:
                row0 = data4[0]
                if row0[cost_idx] and row0[cpp_idx] and row0[part_idx]:
                    expected = float(row0[cpp_idx]) * float(row0[part_idx])
                    check("Budget total_cost = cost_per_person * participants",
                          num_close(row0[cost_idx], expected, 1.0),
                          f"{row0[cost_idx]} vs {expected}")


def check_word(workspace):
    print("\n=== Check 2: Nutrition_Program_Proposal.docx ===")
    path = os.path.join(workspace, "Nutrition_Program_Proposal.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()

    check("Has Executive Summary", "executive summary" in full_text or "executive" in full_text)
    check("Has Research Foundation", "research" in full_text and "foundation" in full_text or "research foundation" in full_text)
    check("Has Recipe Selection", "recipe" in full_text and ("selection" in full_text or "methodology" in full_text))
    check("Has Weekly Menus section", "weekly" in full_text and "menu" in full_text)
    check("Has Survey Plan", "survey" in full_text and "plan" in full_text)
    check("Has Budget section", "budget" in full_text)
    check("Has Timeline section", "timeline" in full_text or "phase" in full_text)
    check("Mentions 200 employees", "200" in full_text)
    check("Mentions evidence-based", "evidence" in full_text)
    check("Mentions Phase 1", "phase 1" in full_text or "phase one" in full_text or "month 1" in full_text)


def check_gform():
    print("\n=== Check 3: Google Form Survey ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        diet_form = None
        for form_id, title in forms:
            if title and ("dietary" in title.lower() or "nutrition" in title.lower()
                          or "preference" in title.lower()):
                diet_form = (form_id, title)
                break
        check("Dietary preferences survey form exists", diet_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        if diet_form:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (diet_form[0],))
            q_count = cur.fetchone()[0]
            check("Survey has 4 questions", q_count >= 4, f"Found {q_count}")

            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position",
                        (diet_form[0],))
            questions = cur.fetchall()
            q_text = " ".join(str(q[0]) for q in questions).lower()
            q_types = [q[1] for q in questions]
            check("Has dietary restrictions question",
                  "dietary" in q_text or "restriction" in q_text, q_text[:150])
            check("Has meal focus question",
                  "meal" in q_text or "focus" in q_text or "breakfast" in q_text, q_text[:150])
            check("Has checkbox question type",
                  "CHECKBOX" in q_types, f"Types: {q_types}")
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python scripts ===")
    for script in ["categorize_recipes.py", "analyze_research.py", "build_menus.py", "validate_menus.py"]:
        path = os.path.join(workspace, script)
        check(f"{script} exists", os.path.exists(path))


def check_json_outputs(workspace):
    print("\n=== Check 5: JSON output files ===")
    for fname in ["categorized_recipes.json", "research_findings.json", "evidence_based_menus.json"]:
        path = os.path.join(workspace, fname)
        if not os.path.exists(path):
            check(f"{fname} exists", False)
            continue
        check(f"{fname} exists", True)
        try:
            with open(path) as f:
                data = json.load(f)
            check(f"{fname} is valid JSON", True)
            if fname == "categorized_recipes.json":
                if isinstance(data, list):
                    check("categorized_recipes has 10+ entries", len(data) >= 10, f"Found {len(data)}")
                elif isinstance(data, dict):
                    total = sum(len(v) if isinstance(v, list) else 1 for v in data.values())
                    check("categorized_recipes has 10+ entries", total >= 10, f"Found {total}")
            elif fname == "evidence_based_menus.json":
                if isinstance(data, list):
                    check("evidence_based_menus has 5 days", len(data) >= 5, f"Found {len(data)}")
                elif isinstance(data, dict):
                    check("evidence_based_menus has 5 days", len(data) >= 5, f"Found {len(data)} keys")
        except json.JSONDecodeError:
            check(f"{fname} is valid JSON", False, "Invalid JSON")


def check_reverse_validation(workspace):
    """Check that noise scholarly papers are NOT present in Excel output."""
    print("\n=== Reverse Validation ===")

    # Noise paper titles that should NOT appear in the Research_Summary sheet
    noise_titles = ["marathon runners", "carbohydrate loading", "elite marathon",
                    "agricultural policy", "food supply chain", "developing nations"]

    path = os.path.join(workspace, "Nutrition_Program_Analysis.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets_lower = [s.lower().replace(" ", "_") for s in wb.sheetnames]

        # Check Research_Summary sheet for noise
        rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s or "summary" in s), None)
        if rs_idx is not None:
            ws = wb[wb.sheetnames[rs_idx]]
            rows = list(ws.iter_rows(values_only=True))
            all_text = " ".join(str(c) for r in rows for c in r if c).lower()

            no_noise = not any(nt in all_text for nt in noise_titles)
            check("No noise scholarly papers in Research_Summary (marathon, agricultural policy)",
                  no_noise,
                  f"Found noise content in Research_Summary sheet")
        else:
            check("No noise scholarly papers in Research_Summary", True, "No Research_Summary sheet to check")

        # Also check all sheets for noise paper content
        all_wb_text = ""
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                all_wb_text += " ".join(str(c) for c in row if c).lower() + " "

        no_noise_wb = not any(nt in all_wb_text for nt in noise_titles)
        check("No noise scholarly papers anywhere in Excel workbook",
              no_noise_wb,
              "Found noise paper content in workbook")
        wb.close()
    else:
        check("No noise scholarly papers in Excel", True, "Excel file not found to check")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_gform()
    check_scripts(args.agent_workspace)
    check_json_outputs(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

"""Evaluation for terminal-canvas-howtocook-excel-word-gcal.

Checks:
1. Nutrition_Course_Assessment.xlsx with 4 sheets
2. Assessment_Feedback.docx
3. Google Calendar events for grading sessions
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Nutrition_Course_Assessment.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Nutrition_Course_Assessment.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Nutrition_Course_Assessment.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Student_Assignments
    print("  Checking Student_Assignments...")
    a_sheet = get_sheet(agent_wb, "Student_Assignments")
    g_sheet = get_sheet(gt_wb, "Student_Assignments")
    check("Sheet 'Student_Assignments' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
        # Query dynamic assignment count from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id IN (3, 4)")
            expected_assign_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_assign_count = 18
        check(f"Student_Assignments has {expected_assign_count} rows",
              len(a_rows) == expected_assign_count, f"Got {len(a_rows)}")

        # Check a few key assignments by name
        a_lookup = {}
        for r in a_rows:
            if r and r[1]:
                a_lookup[str(r[1]).strip().lower()] = r
        for g_row in g_rows:
            if not g_row or not g_row[1]:
                continue
            key = str(g_row[1]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Assignment '{g_row[1]}' present", False, "Missing")
                continue
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"'{key}' Points",
                      num_close(a_row[2], g_row[2], 1.0),
                      f"Expected {g_row[2]}, got {a_row[2]}")

    # Sheet 2: Recipe_Analysis
    print("  Checking Recipe_Analysis...")
    a_sheet = get_sheet(agent_wb, "Recipe_Analysis")
    g_sheet = get_sheet(gt_wb, "Recipe_Analysis")
    check("Sheet 'Recipe_Analysis' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
        check("Recipe_Analysis has 10 rows", len(a_rows) == 10, f"Got {len(a_rows)}")

        a_lookup = {str(r[0]).strip(): r for r in a_rows if r and r[0]}
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Category '{key}' present", False, "Missing")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"'{key}' Recipe_Count",
                      num_close(a_row[1], g_row[1], 3),
                      f"Expected {g_row[1]}, got {a_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"'{key}' Avg_Difficulty",
                      num_close(a_row[2], g_row[2], 0.3),
                      f"Expected {g_row[2]}, got {a_row[2]}")

    # Sheet 3: Course_Summary
    print("  Checking Course_Summary...")
    a_sheet = get_sheet(agent_wb, "Course_Summary")
    g_sheet = get_sheet(gt_wb, "Course_Summary")
    check("Sheet 'Course_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_data = {}
        for row in a_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_data[str(row[0]).strip().lower()] = row[1]
        g_data = {}
        for row in g_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                g_data[str(row[0]).strip().lower()] = row[1]

        # Query dynamic assignment counts from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 3")
            expected_c3 = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 4")
            expected_c4 = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_c3, expected_c4 = 12, 6

        check("Total_Courses = 2",
              num_close(a_data.get("total_courses"), 2, 0),
              f"Got {a_data.get('total_courses')}")
        check(f"Course_3_Assignments = {expected_c3}",
              num_close(a_data.get("course_3_assignments"), expected_c3, 0),
              f"Got {a_data.get('course_3_assignments')}")
        check(f"Course_4_Assignments = {expected_c4}",
              num_close(a_data.get("course_4_assignments"), expected_c4, 0),
              f"Got {a_data.get('course_4_assignments')}")
        check("Total_Recipe_Categories = 10",
              num_close(a_data.get("total_recipe_categories"), 10, 0),
              f"Got {a_data.get('total_recipe_categories')}")
        check("Total_Recipes = 322",
              num_close(a_data.get("total_recipes"), 322, 5),
              f"Got {a_data.get('total_recipes')}")

    # Sheet 4: Grading_Schedule
    print("  Checking Grading_Schedule...")
    a_sheet = get_sheet(agent_wb, "Grading_Schedule")
    check("Sheet 'Grading_Schedule' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        # Dynamic: grading schedule should match course 3 assignment count
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 3")
            expected_sched_rows = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_sched_rows = 12
        check(f"Grading_Schedule has {expected_sched_rows} rows",
              len(a_rows) == expected_sched_rows, f"Got {len(a_rows)}")
        # Check first session starts on 2026-03-10
        if a_rows:
            first_date = str(a_rows[0][0]).strip()
            check("First grading session on 2026-03-10",
                  "2026-03-10" in first_date,
                  f"Got {first_date}")


def check_word(agent_workspace):
    print("\n=== Checking Assessment_Feedback.docx ===")
    docx_path = os.path.join(agent_workspace, "Assessment_Feedback.docx")
    check("Assessment_Feedback.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 200, f"Length: {len(text)}")
        check("Contains course/assignment reference",
              "course" in text and ("assignment" in text or "assessment" in text))
        check("Contains recipe reference",
              "recipe" in text or "cooking" in text or "category" in text)
        check("Contains recommendation",
              "recommend" in text or "suggest" in text or "curriculum" in text)
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE lower(summary) LIKE '%%grading%%'
               OR lower(summary) LIKE '%%cma%%'
               OR lower(summary) LIKE '%%tma%%'
               OR lower(summary) LIKE '%%final exam%%'
               OR lower(summary) LIKE '%%biochem%%'
        """)
        cnt = cur.fetchone()[0]
        check("Calendar has grading events (>=10)", cnt >= 10,
              f"Found {cnt} grading events")

        # Check events are on weekdays
        cur.execute("""
            SELECT start_datetime, EXTRACT(DOW FROM start_datetime) as dow
            FROM gcal.events
            WHERE lower(summary) LIKE '%%grading%%'
               OR lower(summary) LIKE '%%cma%%'
               OR lower(summary) LIKE '%%tma%%'
               OR lower(summary) LIKE '%%final exam%%'
               OR lower(summary) LIKE '%%biochem%%'
        """)
        rows = cur.fetchall()
        if rows:
            weekend_events = [r for r in rows if r[1] in (0, 6)]
            check("No weekend grading events", len(weekend_events) == 0,
                  f"Found {len(weekend_events)} weekend events")

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check that noise calendar events (Department Meeting, Office Hours, Faculty Lunch)
    # are not included in the grading schedule sheet
    excel_path = os.path.join(workspace, "Nutrition_Course_Assessment.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = get_sheet(wb, "Grading_Schedule")
            if ws:
                all_text = " ".join(
                    str(c).lower() for r in ws.iter_rows(values_only=True) for c in r if c
                )
                noise_terms = ["department meeting", "office hours", "faculty lunch"]
                noise_found = [t for t in noise_terms if t in all_text]
                check("No noise events in Grading_Schedule",
                      len(noise_found) == 0,
                      f"Found noise: {noise_found}")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gcal()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

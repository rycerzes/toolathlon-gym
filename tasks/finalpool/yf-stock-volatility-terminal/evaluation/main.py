"""
Evaluation for yf-stock-volatility-terminal task.
Compares agent-produced stock_volatility_report.xlsx against expected values
computed dynamically from PostgreSQL (yf.stock_prices).

Checks:
1. Risk Analysis sheet: columns, row count, values with tolerances
2. Summary sheet: key-value pairs
3. Daily Returns sheet: existence, columns, approximate row count

Sector benchmarks and risk thresholds come from the mock research portal
(static HTML served during the task). These are hardcoded constants since
they are defined by the task itself and do not change.
"""
import math
import os
import sys
import json
from argparse import ArgumentParser
from datetime import datetime

import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


TOLERANCE_ABS = 1.0  # Absolute tolerance for volatility/drawdown/return values
TOLERANCE_REL = 0.05  # 5% relative tolerance as fallback

# --- Constants from the mock research portal (http://localhost:30152) ---
STOCKS = ['AMZN', 'GOOGL', 'JNJ', 'JPM', 'XOM']
SECTORS = {
    'AMZN': 'Consumer Cyclical',
    'GOOGL': 'Communication Services',
    'JNJ': 'Healthcare',
    'JPM': 'Financial Services',
    'XOM': 'Energy',
}
SECTOR_BENCHMARKS = {
    'Consumer Cyclical': 28.0,
    'Communication Services': 25.5,
    'Healthcare': 18.0,
    'Financial Services': 22.0,
    'Energy': 30.0,
}
# Risk thresholds: Conservative < 15%, 15% <= Moderate <= 25%, Aggressive > 25%

DB_CONFIG = dict(host=os.environ.get('PGHOST', 'localhost'), port=5432, database='toolathlon_gym', user='postgres', password='postgres')


def compute_expected_from_db():
    """Compute all expected Risk Analysis and Summary values from PostgreSQL."""
    try:
        import psycopg2
    except ImportError:
        return None

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    results = {}  # sym -> dict of values
    for sym in STOCKS:
        cur.execute(
            "SELECT date, close FROM yf.stock_prices "
            "WHERE symbol=%s AND date >= '2025-03-06' AND date <= '2026-03-05' "
            "ORDER BY date", (sym,))
        rows = cur.fetchall()
        if not rows or len(rows) < 10:
            print(f"  WARNING: Insufficient price data for {sym} ({len(rows)} rows)")
            conn.close()
            return None

        closes = [float(r[1]) for r in rows]

        # Daily log returns
        log_returns = [math.log(closes[i] / closes[i - 1]) for i in range(1, len(closes))]

        # Annualized volatility
        mean_lr = sum(log_returns) / len(log_returns)
        var_lr = sum((lr - mean_lr) ** 2 for lr in log_returns) / (len(log_returns) - 1)
        ann_vol = round(math.sqrt(var_lr) * math.sqrt(252) * 100, 2)

        # Max drawdown
        peak = closes[0]
        max_dd = 0.0
        for c in closes:
            if c > peak:
                peak = c
            dd = (peak - c) / peak * 100
            if dd > max_dd:
                max_dd = dd
        max_dd = round(max_dd, 2)

        # 1-year return
        ret_pct = round((closes[-1] - closes[0]) / closes[0] * 100, 2)

        # Sector benchmark
        sector = SECTORS[sym]
        bench = SECTOR_BENCHMARKS[sector]
        vol_vs = 'Above' if ann_vol > bench else 'Below'

        # Risk category
        if max_dd < 15:
            risk = 'Conservative'
        elif max_dd <= 25:
            risk = 'Moderate'
        else:
            risk = 'Aggressive'

        results[sym] = {
            'Sector': sector,
            'Annualized_Volatility': ann_vol,
            'Sector_Benchmark_Vol': bench,
            'Vol_vs_Benchmark': vol_vs,
            'Max_Drawdown': max_dd,
            'Risk_Category': risk,
            'One_Year_Return_Pct': ret_pct,
            'num_trading_days': len(closes),
        }

    conn.close()

    # Summary values
    vols = [results[s]['Annualized_Volatility'] for s in STOCKS]
    dds = [results[s]['Max_Drawdown'] for s in STOCKS]
    highest_vol = STOCKS[vols.index(max(vols))]
    lowest_vol = STOCKS[vols.index(min(vols))]
    avg_vol = round(sum(vols) / len(vols), 2)
    above_bench = sum(1 for s in STOCKS if results[s]['Vol_vs_Benchmark'] == 'Above')
    worst_dd = round(max(dds), 2)
    safest = STOCKS[dds.index(min(dds))]

    summary = {
        'Highest_Volatility_Stock': highest_vol,
        'Lowest_Volatility_Stock': lowest_vol,
        'Avg_Annualized_Volatility': avg_vol,
        'Stocks_Above_Benchmark': above_bench,
        'Max_Drawdown_Worst': worst_dd,
        'Safest_Stock': safest,
    }

    return {'risk_analysis': results, 'summary': summary}


def nums_close(expected, actual, abs_tol=TOLERANCE_ABS, rel_tol=TOLERANCE_REL):
    """Check if two numeric values are within tolerance."""
    try:
        e = float(expected)
        a = float(actual)
    except (ValueError, TypeError):
        return False
    if abs(e) < 1e-9:
        return abs(a) < abs_tol
    return abs(e - a) <= abs_tol or abs(e - a) / abs(e) <= rel_tol


def val_match(expected, actual):
    """Check if expected and actual values match (numeric or string)."""
    if expected is None and actual is None:
        return True
    if expected is None or actual is None:
        return False
    e_str = str(expected).strip()
    a_str = str(actual).strip()
    if e_str.lower() == a_str.lower():
        return True
    return nums_close(expected, actual)


def find_sheet(wb, name):
    """Find a sheet by case-insensitive name."""
    for sn in wb.sheetnames:
        if sn.lower().replace(" ", "_") == name.lower().replace(" ", "_"):
            return wb[sn]
        if sn.lower() == name.lower():
            return wb[sn]
    return None


def check_risk_analysis_db(ws_agent, expected):
    """Check Risk Analysis sheet against DB-computed values."""
    total = 0
    passed = 0

    agent_headers = [cell.value for cell in ws_agent[1]]

    expected_cols = ['Symbol', 'Sector', 'Annualized_Volatility', 'Sector_Benchmark_Vol',
                     'Vol_vs_Benchmark', 'Max_Drawdown', 'Risk_Category', 'One_Year_Return_Pct']
    for col in expected_cols:
        total += 1
        found = any(str(h).strip().lower().replace(" ", "_") == col.lower()
                     for h in (agent_headers or []) if h)
        if found:
            passed += 1
        else:
            print(f"  FAIL: Risk Analysis missing column '{col}'. Agent headers: {agent_headers}")

    # Build agent lookup by Symbol
    agent_rows = {}
    for row in ws_agent.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            agent_rows[str(row[0]).strip()] = row

    # Check row count
    total += 1
    if len(agent_rows) == len(STOCKS):
        passed += 1
    else:
        print(f"  FAIL: Risk Analysis row count: expected {len(STOCKS)}, got {len(agent_rows)}")

    # Check each stock
    for sym in STOCKS:
        gt = expected[sym]
        agent_row = agent_rows.get(sym)
        if not agent_row:
            print(f"  FAIL: Risk Analysis missing stock {sym}")
            total += len(expected_cols) - 1
            continue

        for col_idx, col_name in enumerate(expected_cols):
            if col_idx == 0:  # Symbol already matched
                continue
            total += 1
            gt_val = gt.get(col_name)

            # Find matching column in agent row
            agent_col_idx = None
            for ai, ah in enumerate(agent_headers or []):
                if ah and str(ah).strip().lower().replace(" ", "_") == col_name.lower():
                    agent_col_idx = ai
                    break
            if agent_col_idx is None:
                agent_col_idx = col_idx

            agent_val = agent_row[agent_col_idx] if agent_col_idx < len(agent_row) else None

            if val_match(gt_val, agent_val):
                passed += 1
            else:
                print(f"  FAIL: Risk Analysis {sym}.{col_name}: expected={gt_val}, got={agent_val}")

    return passed, total


def check_risk_analysis_gt(ws_agent, ws_gt):
    """Fallback: Check Risk Analysis sheet against groundtruth Excel."""
    total = 0
    passed = 0

    gt_headers = [cell.value for cell in ws_gt[1]]
    agent_headers = [cell.value for cell in ws_agent[1]]

    expected_cols = ['Symbol', 'Sector', 'Annualized_Volatility', 'Sector_Benchmark_Vol',
                     'Vol_vs_Benchmark', 'Max_Drawdown', 'Risk_Category', 'One_Year_Return_Pct']
    for col in expected_cols:
        total += 1
        found = any(str(h).strip().lower().replace(" ", "_") == col.lower()
                     for h in (agent_headers or []) if h)
        if found:
            passed += 1
        else:
            print(f"  FAIL: Risk Analysis missing column '{col}'. Agent headers: {agent_headers}")

    gt_rows = {}
    for row in ws_gt.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            gt_rows[str(row[0]).strip()] = row

    agent_rows = {}
    for row in ws_agent.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            agent_rows[str(row[0]).strip()] = row

    total += 1
    if len(agent_rows) == len(gt_rows):
        passed += 1
    else:
        print(f"  FAIL: Risk Analysis row count: expected {len(gt_rows)}, got {len(agent_rows)}")

    for sym, gt_row in gt_rows.items():
        agent_row = agent_rows.get(sym)
        if not agent_row:
            print(f"  FAIL: Risk Analysis missing stock {sym}")
            total += len(expected_cols) - 1
            continue
        for col_idx, col_name in enumerate(expected_cols):
            if col_idx == 0:
                continue
            total += 1
            gt_val = gt_row[col_idx] if col_idx < len(gt_row) else None
            agent_col_idx = None
            for ai, ah in enumerate(agent_headers or []):
                if ah and str(ah).strip().lower().replace(" ", "_") == col_name.lower():
                    agent_col_idx = ai
                    break
            if agent_col_idx is None:
                agent_col_idx = col_idx
            agent_val = agent_row[agent_col_idx] if agent_col_idx < len(agent_row) else None
            if val_match(gt_val, agent_val):
                passed += 1
            else:
                print(f"  FAIL: Risk Analysis {sym}.{col_name}: expected={gt_val}, got={agent_val}")

    return passed, total


def check_summary_db(ws_agent, expected_summary):
    """Check Summary sheet against DB-computed values."""
    total = 0
    passed = 0

    agent_data = {}
    for row in ws_agent.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            agent_data[str(row[0]).strip()] = row[1]

    for metric, gt_val in expected_summary.items():
        total += 1
        agent_val = agent_data.get(metric)
        if agent_val is None:
            for k, v in agent_data.items():
                if k.lower() == metric.lower():
                    agent_val = v
                    break
        if val_match(gt_val, agent_val):
            passed += 1
        else:
            print(f"  FAIL: Summary '{metric}': expected={gt_val}, got={agent_val}")

    return passed, total


def check_summary_gt(ws_agent, ws_gt):
    """Fallback: Check Summary sheet against groundtruth Excel."""
    total = 0
    passed = 0

    gt_data = {}
    for row in ws_gt.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            gt_data[str(row[0]).strip()] = row[1]

    agent_data = {}
    for row in ws_agent.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            agent_data[str(row[0]).strip()] = row[1]

    for metric, gt_val in gt_data.items():
        total += 1
        agent_val = agent_data.get(metric)
        if agent_val is None:
            for k, v in agent_data.items():
                if k.lower() == metric.lower():
                    agent_val = v
                    break
        if val_match(gt_val, agent_val):
            passed += 1
        else:
            print(f"  FAIL: Summary '{metric}': expected={gt_val}, got={agent_val}")

    return passed, total


def check_daily_returns(ws_agent):
    """Check Daily Returns sheet exists with correct columns and approximate row count."""
    total = 0
    passed = 0

    headers = [cell.value for cell in ws_agent[1]]
    expected_cols = ['Symbol', 'Date', 'Close_Price', 'Daily_Return_Pct']
    for col in expected_cols:
        total += 1
        found = any(str(h).strip().lower().replace(" ", "_") == col.lower().replace(" ", "_")
                     for h in (headers or []) if h)
        if found:
            passed += 1
        else:
            print(f"  FAIL: Daily Returns missing column '{col}'. Headers: {headers}")

    # Check approximate row count (should be ~1255 rows = 251 days * 5 stocks)
    total += 1
    row_count = ws_agent.max_row - 1  # minus header
    if 1200 <= row_count <= 1300:
        passed += 1
    else:
        print(f"  FAIL: Daily Returns row count: expected ~1255, got {row_count}")

    return passed, total


def main(args):
    agent_path = os.path.join(args.agent_workspace, "stock_volatility_report.xlsx")
    gt_path = os.path.join(args.groundtruth_workspace, "stock_volatility_report.xlsx")

    if not os.path.exists(agent_path):
        print(f"FAIL: stock_volatility_report.xlsx not found at {agent_path}")
        result = {"total_passed": 0, "total_checks": 1, "accuracy": 0.0}
        if args.res_log_file:
            with open(args.res_log_file, "w") as f:
                json.dump(result, f, indent=2)
        sys.exit(1)

    wb_agent = openpyxl.load_workbook(agent_path, data_only=True)

    # Try to compute expected values from PostgreSQL
    db_expected = compute_expected_from_db()
    use_db = db_expected is not None
    if use_db:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
    else:
        print("INFO: Falling back to static groundtruth Excel file")
        if not os.path.exists(gt_path):
            print(f"FAIL: Groundtruth file not found at {gt_path}")
            sys.exit(1)

    # Load groundtruth Excel as fallback
    wb_gt = None
    if not use_db:
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)

    total_passed = 0
    total_checks = 0

    # Check 1: Risk Analysis sheet
    print("--- Check 1: Risk Analysis Sheet ---")
    ws_ra_agent = find_sheet(wb_agent, "Risk Analysis")

    if not ws_ra_agent:
        print("FAIL: 'Risk Analysis' sheet not found in agent output")
        total_checks += 1
    elif use_db:
        p, t = check_risk_analysis_db(ws_ra_agent, db_expected['risk_analysis'])
        print(f"  Risk Analysis: {p}/{t} checks passed")
        total_passed += p
        total_checks += t
    else:
        ws_ra_gt = find_sheet(wb_gt, "Risk Analysis")
        if not ws_ra_gt:
            print("FAIL: 'Risk Analysis' sheet not found in groundtruth (internal error)")
            total_checks += 1
        else:
            p, t = check_risk_analysis_gt(ws_ra_agent, ws_ra_gt)
            print(f"  Risk Analysis: {p}/{t} checks passed")
            total_passed += p
            total_checks += t

    # Check 2: Summary sheet
    print("\n--- Check 2: Summary Sheet ---")
    ws_sum_agent = find_sheet(wb_agent, "Summary")

    if not ws_sum_agent:
        print("FAIL: 'Summary' sheet not found in agent output")
        total_checks += 1
    elif use_db:
        p, t = check_summary_db(ws_sum_agent, db_expected['summary'])
        print(f"  Summary: {p}/{t} checks passed")
        total_passed += p
        total_checks += t
    else:
        ws_sum_gt = find_sheet(wb_gt, "Summary")
        if not ws_sum_gt:
            print("FAIL: 'Summary' sheet not found in groundtruth (internal error)")
            total_checks += 1
        else:
            p, t = check_summary_gt(ws_sum_agent, ws_sum_gt)
            print(f"  Summary: {p}/{t} checks passed")
            total_passed += p
            total_checks += t

    # Check 3: Daily Returns sheet
    print("\n--- Check 3: Daily Returns Sheet ---")
    ws_dr_agent = find_sheet(wb_agent, "Daily Returns")

    if not ws_dr_agent:
        print("FAIL: 'Daily Returns' sheet not found in agent output")
        total_checks += 1
    else:
        p, t = check_daily_returns(ws_dr_agent)
        print(f"  Daily Returns: {p}/{t} checks passed")
        total_passed += p
        total_checks += t

    wb_agent.close()
    if wb_gt:
        wb_gt.close()

    # Overall
    if total_checks == 0:
        print("\nFAIL: No checks were performed.")
        accuracy = 0.0
    else:
        accuracy = total_passed / total_checks * 100
        print(f"\nOverall: {total_passed}/{total_checks} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": total_passed,
        "total_checks": total_checks,
        "accuracy": accuracy,
        "timestamp": datetime.now().isoformat(),
        "source": "postgresql" if use_db else "groundtruth_excel",
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
        print(f"Report saved to {args.res_log_file}")

    if accuracy >= 80:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    main(args)

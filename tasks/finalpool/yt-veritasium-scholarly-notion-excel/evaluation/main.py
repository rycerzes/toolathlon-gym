"""
Evaluation for yt-veritasium-scholarly-notion-excel task.

Checks:
1. Science_Resource_Map.xlsx exists with Videos and Papers sheets
2. Videos sheet has 5 rows with top Veritasium videos and topics
3. Papers sheet has at least 5 rows (papers found per topic)
4. Notion page "Science Video-Paper Resource Map" exists
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Science_Resource_Map.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Science_Resource_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Science_Resource_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Science_Resource_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Videos sheet
    if "videos" not in sheet_names_lower:
        record("Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Videos sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("videos")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Videos sheet has 5 rows (top Veritasium videos)", len(data_rows) == 5,
               f"Found {len(data_rows)} rows")

        # Check for Veritasium video IDs or titles
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        has_poison = "poisoned" in all_text or "sc2esuz" in all_text.lower()
        has_zoom = "zooming" in all_text or "zoom" in all_text or "88mbvbx" in all_text.lower()
        record("Videos include 'How One Company Secretly Poisoned The Planet'",
               has_poison, "Top video not found")
        record("Videos include 'Can you keep zooming in forever?'",
               has_zoom, "Second video not found")

        # Check topics column
        has_topic_col = any("topic" in str(c).lower() for c in (rows[0] if rows else []))
        record("Videos sheet has Main_Topic column", has_topic_col, f"Headers: {rows[0] if rows else 'none'}")

    # Papers sheet
    if "papers" not in sheet_names_lower:
        record("Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Papers sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("papers")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        record("Papers sheet has at least 5 rows",
               len(data_rows2) >= 5, f"Found {len(data_rows2)} rows")

        # Check has Topic, Paper_Title, Authors, Year columns
        if rows2:
            headers = [str(c).lower() if c else "" for c in rows2[0]]
            has_topic = any("topic" in h for h in headers)
            has_title = any("title" in h for h in headers)
            has_year = any("year" in h for h in headers)
            record("Papers sheet has Topic, Paper_Title, Year columns",
                   has_topic and has_title and has_year, f"Headers: {rows2[0]}")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Science_Resource_Map.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, properties FROM notion.pages
        WHERE properties::text ILIKE %s
        AND archived = false AND in_trash = false
    """, ("%Science Video%Paper%",))
    pages = cur.fetchall()
    cur.close()
    conn.close()

    record("Notion page 'Science Video-Paper Resource Map' exists",
           len(pages) >= 1, f"Found {len(pages)} matching pages")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

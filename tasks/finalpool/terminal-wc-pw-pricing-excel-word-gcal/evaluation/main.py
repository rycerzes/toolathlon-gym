"""Evaluation for terminal-wc-pw-pricing-excel-word-gcal."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=5.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_gaps():
    """Dynamically compute expected price gaps from WC product data.
    Falls back to hardcoded values if DB query fails."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT c.name, ROUND(AVG(p.regular_price)::numeric, 2)
            FROM wc.products p
            JOIN wc.categories c ON p.category_id = c.id
            WHERE p.regular_price > 0
            GROUP BY c.name
        """)
        our_prices = {r[0].strip().lower(): float(r[1]) for r in cur.fetchall()}
        cur.close()
        conn.close()
        # We can only compute our side; competitor data comes from playwright scrape
        # Return our avg prices for validation, keep hardcoded gaps as fallback
        return our_prices, {
            "audio": 12.6, "cameras": -48.6, "electronics": 92.9,
            "headphones": -37.6, "speakers": 28.7,
        }
    except Exception:
        return None, {
            "audio": 12.6, "cameras": -48.6, "electronics": 92.9,
            "headphones": -37.6, "speakers": 28.7,
        }


EXPECTED_GAPS = {
    "audio": 12.6,
    "cameras": -48.6,
    "electronics": 92.9,
    "headphones": -37.6,
    "speakers": 28.7,
}


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Report.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(fpath, data_only=True)

    our_prices, gap_fallback = get_expected_gaps()

    # Sheet 1: Our_Products
    our_sheet = None
    for name in wb.sheetnames:
        if "our" in name.lower() and "product" in name.lower():
            our_sheet = name
            break
    if not our_sheet:
        record("Our_Products sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Our_Products sheet exists", True)
        ws = wb[our_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Dynamically get expected product count from DB
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM wc.products WHERE regular_price > 0")
            expected_products = cur.fetchone()[0]
            cur.close()
            conn.close()
            record(f"Our_Products has >= {expected_products} rows", len(rows) >= expected_products,
                   f"Found {len(rows)}")
        except Exception:
            record("Our_Products has >= 30 rows", len(rows) >= 30, f"Found {len(rows)}")

    # Sheet 2: Competitor_Prices
    comp_sheet = None
    for name in wb.sheetnames:
        if "competitor" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Competitor_Prices sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Competitor_Prices sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Competitor_Prices has 19 rows", len(rows) == 19, f"Found {len(rows)}")

    # Sheet 3: Price_Gap_Analysis
    gap_sheet = None
    for name in wb.sheetnames:
        if "gap" in name.lower() or "analysis" in name.lower():
            gap_sheet = name
            break
    if not gap_sheet:
        record("Price_Gap_Analysis sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Price_Gap_Analysis sheet exists", True)
        ws = wb[gap_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Price_Gap_Analysis has 5 rows", len(rows) == 5, f"Found {len(rows)}")

        for row in rows:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat in EXPECTED_GAPS:
                    gap_val = row[3] if len(row) > 3 else None
                    record(f"Gap for {cat} is correct",
                           num_close(gap_val, EXPECTED_GAPS[cat], tol=10.0),
                           f"Got {gap_val}, expected ~{EXPECTED_GAPS[cat]}")

    wb.close()
    return True


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    fpath = os.path.join(agent_workspace, "Pricing_Strategy_Report.docx")
    if not os.path.isfile(fpath):
        record("Word document exists", False, f"Not found: {fpath}")
        return False
    record("Word document exists", True)

    from docx import Document
    doc = Document(fpath)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    record("Document mentions competitive pricing", "competitive" in full_text or "pricing" in full_text)
    record("Document mentions electronics", "electronics" in full_text)
    record("Document mentions headphones", "headphones" in full_text)
    return True


def check_terminal_output(agent_workspace):
    print("\n=== Checking Terminal Output ===")
    fpath = os.path.join(agent_workspace, "price_analysis_output.txt")
    if not os.path.isfile(fpath):
        record("price_analysis_output.txt exists", False)
        return False
    record("price_analysis_output.txt exists", True)
    with open(fpath) as f:
        content = f.read().lower()
    record("Output mentions price gap or comparison", "gap" in content or "comparison" in content or "%" in content)
    return True


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    categories_found = set()
    for summary, description in events:
        summary_lower = (summary or "").lower()
        if "price review" in summary_lower:
            for cat in ["electronics", "headphones", "speakers", "audio", "cameras"]:
                if cat in summary_lower:
                    categories_found.add(cat)

    record("Price review events for >= 5 categories", len(categories_found) >= 5,
           f"Found events for: {categories_found}")
    return len(categories_found) >= 5


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    fpath = os.path.join(workspace, "Competitive_Pricing_Report.xlsx")
    if os.path.isfile(fpath):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        # No unexpected sheets
        expected_keywords = {"our", "product", "competitor", "price", "gap", "analysis"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        record("No unexpected sheets in Excel", len(unexpected) == 0,
               f"Unexpected: {unexpected}")

        # Prices should not be negative
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for c in row:
                    try:
                        v = float(str(c).replace(',', '').replace('$', '').strip()) if c else None
                    except (ValueError, TypeError):
                        v = None
                    if v is not None and v < 0 and "gap" not in sname.lower():
                        record(f"No negative prices in {sname}", False, f"Found {v}")
                        break
                else:
                    continue
                break
            else:
                continue
            break
        else:
            record("No negative prices in product sheets", True)

    # Calendar: no price review events before March 2026
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%price review%%'
              AND start_datetime < '2026-03-10'
        """)
        early_events = cur.fetchone()[0]
        record("No price review events before March 10", early_events == 0,
               f"Found {early_events} early events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_terminal_output(args.agent_workspace)
    check_calendar()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

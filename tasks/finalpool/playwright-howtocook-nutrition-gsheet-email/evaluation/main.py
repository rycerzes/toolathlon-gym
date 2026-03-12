"""
Evaluation script for playwright-howtocook-nutrition-gsheet-email task.

Checks:
1. Google Sheets spreadsheet "Wellness_Lunch_Menu" with Weekly_Plan and Nutrition_Summary sheets
2. Word document Wellness_Menu_Summary.docx with menu content
3. Email sent to cafeteria-manager about wellness lunch menu
"""

import argparse
import json
import os
import sys

import psycopg2

try:
    from docx import Document
except ImportError:
    Document = None

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


def check_gsheet():
    """Check Google Sheets spreadsheet."""
    print("\n=== Checking Google Sheets ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find the Wellness_Lunch_Menu spreadsheet
        cur.execute(
            "SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE %s",
            ("%wellness%lunch%menu%",),
        )
        rows = cur.fetchall()
        if not rows:
            cur.execute("SELECT id, title FROM gsheet.spreadsheets")
            all_ss = cur.fetchall()
            record(
                "Wellness_Lunch_Menu spreadsheet exists",
                False,
                f"Found spreadsheets: {[r[1] for r in all_ss]}",
            )
            cur.close()
            conn.close()
            return False

        record("Wellness_Lunch_Menu spreadsheet exists", True)
        ss_id = rows[0][0]

        # Check sheets
        cur.execute(
            "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s",
            (ss_id,),
        )
        sheets = cur.fetchall()
        sheet_titles = [s[1].lower() for s in sheets]

        has_weekly = any("weekly" in t or "plan" in t for t in sheet_titles)
        has_nutrition = any("nutrition" in t or "summary" in t for t in sheet_titles)
        record("Weekly_Plan sheet exists", has_weekly, f"Sheets: {sheet_titles}")
        record(
            "Nutrition_Summary sheet exists", has_nutrition, f"Sheets: {sheet_titles}"
        )

        # Check Weekly_Plan content - should have 5 data rows (Mon-Fri)
        weekly_sheet_id = None
        for sid, title in sheets:
            if "weekly" in title.lower() or "plan" in title.lower():
                weekly_sheet_id = sid
                break

        if weekly_sheet_id:
            cur.execute(
                "SELECT row_index, col_index, value FROM gsheet.cells "
                "WHERE sheet_id = %s ORDER BY row_index, col_index",
                (weekly_sheet_id,),
            )
            cells = cur.fetchall()

            # Count data rows (row_index > 0 for header row 0)
            data_rows = set()
            for row_idx, col_idx, val in cells:
                if row_idx > 0 and val and str(val).strip():
                    data_rows.add(row_idx)

            record(
                "Weekly_Plan has 5 day rows",
                len(data_rows) >= 5,
                f"Found {len(data_rows)} data rows",
            )

            # Check days of week present
            all_values = [str(v).lower() for _, _, v in cells if v]
            days_found = sum(
                1
                for d in ["monday", "tuesday", "wednesday", "thursday", "friday"]
                if any(d in v for v in all_values)
            )
            record(
                "All 5 weekdays in Weekly_Plan",
                days_found >= 5,
                f"Found {days_found} days",
            )

            # Check nutritional columns exist (calories, protein, fat, carbs)
            header_values = [
                str(v).lower() for r, c, v in cells if r == 0 and v
            ]
            has_cal = any("calori" in h for h in header_values)
            has_prot = any("protein" in h for h in header_values)
            has_fat = any("fat" in h for h in header_values)
            has_carb = any("carb" in h for h in header_values)
            record(
                "Nutrition columns in Weekly_Plan",
                has_cal and has_prot and has_fat and has_carb,
                f"Headers: {header_values}",
            )

        # Check Nutrition_Summary content
        nutr_sheet_id = None
        for sid, title in sheets:
            if "nutrition" in title.lower() or "summary" in title.lower():
                nutr_sheet_id = sid
                break

        if nutr_sheet_id:
            cur.execute(
                "SELECT row_index, col_index, value FROM gsheet.cells "
                "WHERE sheet_id = %s ORDER BY row_index, col_index",
                (nutr_sheet_id,),
            )
            cells = cur.fetchall()
            all_values = [str(v).lower() for _, _, v in cells if v]

            has_target = any("target" in v for v in all_values)
            has_average = any("average" in v or "avg" in v for v in all_values)
            record(
                "Nutrition_Summary has target and average columns",
                has_target and has_average,
                f"Values sample: {all_values[:10]}",
            )

            # Check nutrients listed
            nutrients_found = sum(
                1
                for n in ["calori", "protein", "fat", "carb"]
                if any(n in v for v in all_values)
            )
            record(
                "All 4 nutrients in Nutrition_Summary",
                nutrients_found >= 4,
                f"Found {nutrients_found}",
            )

        cur.close()
        conn.close()
        return True

    except Exception as e:
        record("Google Sheets accessible", False, str(e))
        return False


def check_word(agent_workspace):
    """Check Word document."""
    print("\n=== Checking Word Document ===")

    doc_path = os.path.join(agent_workspace, "Wellness_Menu_Summary.docx")
    if not os.path.isfile(doc_path):
        record("Wellness_Menu_Summary.docx exists", False, f"Not found: {doc_path}")
        return False

    record("Wellness_Menu_Summary.docx exists", True)

    if Document is None:
        record("python-docx available", False, "Cannot import docx")
        return False

    try:
        doc = Document(doc_path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()

        record(
            "Doc mentions wellness/lunch/menu",
            "wellness" in full_text or "lunch" in full_text or "menu" in full_text,
            "Missing wellness/lunch/menu keywords",
        )

        # Check days mentioned
        days_found = sum(
            1
            for d in ["monday", "tuesday", "wednesday", "thursday", "friday"]
            if d in full_text
        )
        record(
            "Doc mentions weekdays",
            days_found >= 3,
            f"Found {days_found} days",
        )

        # Check nutritional info mentioned
        has_nutrition = (
            "calori" in full_text
            or "protein" in full_text
            or "nutrition" in full_text
        )
        record("Doc mentions nutrition", has_nutrition)

        # Check length - should be substantive
        record(
            "Doc has substantial content",
            len(full_text) > 200,
            f"Length: {len(full_text)}",
        )

        return True
    except Exception as e:
        record("Word doc readable", False, str(e))
        return False


def check_email():
    """Check email sent about wellness menu."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, from_addr, to_addr, body_text FROM email.messages"
        )
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    found = False
    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if ("wellness" in subj_lower or "lunch" in subj_lower or "menu" in subj_lower):
            found = True
            record("Wellness menu email exists", True)

            # Check recipient
            to_str = str(to_addr).lower() if to_addr else ""
            record(
                "Email to cafeteria-manager",
                "cafeteria" in to_str or "manager" in to_str,
                f"To: {to_addr}",
            )

            # Check body content
            body_lower = (body_text or "").lower()
            has_dishes = any(
                kw in body_lower
                for kw in ["chicken", "beef", "tofu", "vegetable", "soup", "rice"]
            )
            record(
                "Email body mentions dishes",
                has_dishes,
                f"Body length: {len(body_lower)}",
            )
            break

    if not found:
        record(
            "Wellness menu email exists",
            False,
            f"Found {len(emails)} emails but none about wellness/lunch/menu",
        )

    return found


def check_xlsx_content(workspace, groundtruth_workspace="."):
    """Check Wellness_Lunch_Menu.xlsx has valid content."""
    print("\n=== Checking XLSX Content ===")
    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "Cannot import openpyxl")
        return False

    xlsx_path = os.path.join(workspace, "Wellness_Lunch_Menu.xlsx")
    if not os.path.isfile(xlsx_path):
        record("Wellness_Lunch_Menu.xlsx exists", False, f"Not found: {xlsx_path}")
        return False
    record("Wellness_Lunch_Menu.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        record("XLSX has at least one sheet", len(wb.worksheets) >= 1,
               f"Found {len(wb.worksheets)} sheets")
        all_ok = True
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            has_data = len(rows) >= 2
            record(f"XLSX sheet '{ws.title}' has data rows", has_data,
                   f"Only {len(rows)} rows")
            if not has_data:
                all_ok = False

        # --- Groundtruth XLSX comparison ---
        gt_path = os.path.join(groundtruth_workspace, "Wellness_Lunch_Menu.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")
                for ri in range(min(3, len(gt_rows))):
                    if ri >= len(a_rows): break
                    ok = True
                    for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                        gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                        if gv is None: continue
                        if isinstance(gv, (int, float)):
                            if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                        else:
                            if not str_match(av, gv): ok = False; break
                    record(f"GT '{gt_sname}' row {ri+1} values", ok,
                           f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
            gt_wb.close()

        wb.close()
        return all_ok
    except Exception as e:
        record("XLSX readable", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    gsheet_ok = check_gsheet()
    word_ok = check_word(args.agent_workspace)
    email_ok = check_email()
    xlsx_ok = check_xlsx_content(args.agent_workspace, args.groundtruth_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  GSheet:   {'PASS' if gsheet_ok else 'FAIL'}")
    print(f"  Word:     {'PASS' if word_ok else 'FAIL'}")
    print(f"  Email:    {'PASS' if email_ok else 'FAIL'}")
    print(f"  XLSX:     {'PASS' if xlsx_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    overall = gsheet_ok and word_ok and email_ok and xlsx_ok
    print(f"  Overall:  {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

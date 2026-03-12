"""
Evaluation script for support-sla-audit-form task.

Checks:
1. Excel file (SLA_Audit_Report.xlsx) - 3 sheets with correct data computed from Snowflake
2. Google Form created with correct structure (queried from gform schema)

Expected values are computed at evaluation time from the PostgreSQL database,
not from pre-generated groundtruth files.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    """
    Query the PostgreSQL database to compute expected SLA audit values.
    Returns a dict with keys: sla_compliance, agent_performance, summary.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # ---- SLA Compliance by priority ----
    # Join tickets with SLA policies on priority, check breach
    cur.execute("""
        SELECT
            t."PRIORITY",
            COUNT(*) AS total_tickets,
            SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END) AS breached_tickets,
            ROUND(
                (1.0 - SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END)::numeric / COUNT(*)::numeric) * 100,
                1
            ) AS compliance_rate,
            ROUND(AVG(t."RESPONSE_TIME_HOURS")::numeric, 1) AS avg_response_hours,
            ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 1) AS avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" p
            ON t."PRIORITY" = p."PRIORITY"
        GROUP BY t."PRIORITY"
        ORDER BY t."PRIORITY" ASC
    """)
    sla_rows = cur.fetchall()
    sla_compliance = []
    for row in sla_rows:
        sla_compliance.append({
            "Priority": row[0],
            "Total_Tickets": int(row[1]),
            "Breached_Tickets": int(row[2]),
            "Compliance_Rate": float(row[3]),
            "Avg_Response_Hours": float(row[4]),
            "Avg_Satisfaction": float(row[5]),
        })

    # ---- Agent Performance ----
    cur.execute("""
        SELECT
            a."AGENT_NAME",
            a."TEAM",
            COUNT(t."TICKET_ID") AS tickets_resolved,
            ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 1) AS avg_satisfaction,
            ROUND(AVG(t."RESPONSE_TIME_HOURS")::numeric, 1) AS avg_response_hours
        FROM sf_data."SUPPORT_CENTER__PUBLIC__AGENTS" a
        LEFT JOIN sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
            ON a."AGENT_NAME" = t."RESOLVER"
        GROUP BY a."AGENT_NAME", a."TEAM"
        ORDER BY tickets_resolved DESC
    """)
    agent_rows = cur.fetchall()
    agent_performance = []
    for row in agent_rows:
        agent_performance.append({
            "Agent_Name": row[0],
            "Team": row[1],
            "Tickets_Resolved": int(row[2]),
            "Avg_Satisfaction": float(row[3]) if row[3] is not None else 0.0,
            "Avg_Response_Hours": float(row[4]) if row[4] is not None else 0.0,
        })

    # ---- Summary ----
    cur.execute("""
        SELECT
            COUNT(*) AS total_tickets,
            SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END) AS total_breached,
            ROUND(
                (1.0 - SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END)::numeric / COUNT(*)::numeric) * 100,
                1
            ) AS overall_compliance_rate,
            ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 1) AS overall_avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" p
            ON t."PRIORITY" = p."PRIORITY"
    """)
    summary_row = cur.fetchone()
    total_tickets = int(summary_row[0])
    total_breached = int(summary_row[1])
    overall_compliance_rate = float(summary_row[2])
    overall_avg_satisfaction = float(summary_row[3])

    summary = {
        "Overall_Compliance_Rate": overall_compliance_rate,
        "Overall_Avg_Satisfaction": overall_avg_satisfaction,
        "Meets_Compliance_Target": "Yes" if overall_compliance_rate >= 90.0 else "No",
        "Meets_Satisfaction_Target": "Yes" if overall_avg_satisfaction >= 4.0 else "No",
        "Total_Tickets_Reviewed": total_tickets,
        "Total_Breached_Tickets": total_breached,
    }

    cur.close()
    conn.close()

    return {
        "sla_compliance": sla_compliance,
        "agent_performance": agent_performance,
        "summary": summary,
    }


def get_sheet(wb, name):
    """Find sheet case-insensitively."""
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    """Check the Excel output file against computed expected values."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "SLA_Audit_Report.xlsx")
    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    check("Excel file readable", True)

    # Check sheet names
    expected_sheets = ["SLA Compliance", "Agent Performance", "Summary"]
    for sheet_name in expected_sheets:
        found = any(str_match(s, sheet_name) for s in wb.sheetnames)
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {wb.sheetnames}")

    # --- Sheet 1: SLA Compliance ---
    print("\n--- SLA Compliance ---")
    ws = get_sheet(wb, "SLA Compliance")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp_rows = expected["sla_compliance"]
        check("SLA Compliance row count", len(agent_rows) == len(exp_rows),
              f"Expected {len(exp_rows)}, got {len(agent_rows)}")

        for exp_row in exp_rows:
            priority = exp_row["Priority"]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], priority):
                    matched = ar
                    break
            if matched:
                check(f"Priority '{priority}' Total_Tickets",
                      num_close(matched[1], exp_row["Total_Tickets"], 1.0),
                      f"Expected {exp_row['Total_Tickets']}, got {matched[1]}")
                check(f"Priority '{priority}' Breached_Tickets",
                      num_close(matched[2], exp_row["Breached_Tickets"], 1.0),
                      f"Expected {exp_row['Breached_Tickets']}, got {matched[2]}")
                check(f"Priority '{priority}' Compliance_Rate",
                      num_close(matched[3], exp_row["Compliance_Rate"], 0.5),
                      f"Expected {exp_row['Compliance_Rate']}, got {matched[3]}")
                check(f"Priority '{priority}' Avg_Response_Hours",
                      num_close(matched[4], exp_row["Avg_Response_Hours"], 0.5),
                      f"Expected {exp_row['Avg_Response_Hours']}, got {matched[4]}")
                check(f"Priority '{priority}' Avg_Satisfaction",
                      num_close(matched[5], exp_row["Avg_Satisfaction"], 0.5),
                      f"Expected {exp_row['Avg_Satisfaction']}, got {matched[5]}")
            else:
                check(f"Priority '{priority}' found", False,
                      "Priority not in agent output")

    # --- Sheet 2: Agent Performance ---
    print("\n--- Agent Performance ---")
    ws = get_sheet(wb, "Agent Performance")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp_rows = expected["agent_performance"]
        check("Agent Performance row count", len(agent_rows) == len(exp_rows),
              f"Expected {len(exp_rows)}, got {len(agent_rows)}")

        for exp_row in exp_rows:
            agent_name = exp_row["Agent_Name"]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], agent_name):
                    matched = ar
                    break
            if matched:
                check(f"Agent '{agent_name}' Team",
                      str_match(matched[1], exp_row["Team"]),
                      f"Expected '{exp_row['Team']}', got '{matched[1]}'")
                check(f"Agent '{agent_name}' Tickets_Resolved",
                      num_close(matched[2], exp_row["Tickets_Resolved"], 1.0),
                      f"Expected {exp_row['Tickets_Resolved']}, got {matched[2]}")
                check(f"Agent '{agent_name}' Avg_Satisfaction",
                      num_close(matched[3], exp_row["Avg_Satisfaction"], 0.5),
                      f"Expected {exp_row['Avg_Satisfaction']}, got {matched[3]}")
                check(f"Agent '{agent_name}' Avg_Response_Hours",
                      num_close(matched[4], exp_row["Avg_Response_Hours"], 0.5),
                      f"Expected {exp_row['Avg_Response_Hours']}, got {matched[4]}")
            else:
                check(f"Agent '{agent_name}' found", False,
                      "Agent not in agent output")

    # --- Sheet 3: Summary ---
    print("\n--- Summary ---")
    ws = get_sheet(wb, "Summary")
    if ws:
        agent_data = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower().replace(" ", "_")] = row[1]

        exp_summary = expected["summary"]
        for key, gt_val in exp_summary.items():
            key_lower = key.lower()
            agent_val = agent_data.get(key_lower)
            if agent_val is None:
                # Try fuzzy key match
                for ak, av in agent_data.items():
                    if key_lower.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                ok = num_close(agent_val, gt_val, 1.0)
                check(f"Summary '{key}'", ok,
                      f"Expected {gt_val}, got {agent_val}")
            else:
                ok = str_match(agent_val, gt_val)
                check(f"Summary '{key}'", ok,
                      f"Expected '{gt_val}', got '{agent_val}'")


def check_form():
    """Check that a Google Form was created with the correct structure."""
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Find form with title containing "SLA Improvement"
    cur.execute("""
        SELECT id, title, description
        FROM gform.forms
        WHERE LOWER(title) LIKE '%sla improvement%'
        ORDER BY created_at DESC
        LIMIT 1
    """)
    form_row = cur.fetchone()

    check("Google Form with 'SLA Improvement' in title exists",
          form_row is not None)
    if not form_row:
        cur.close()
        conn.close()
        return

    form_id = form_row[0]
    form_title = form_row[1]
    check("Form title is 'SLA Improvement Plan'",
          "sla improvement plan" in form_title.lower(),
          f"Actual title: '{form_title}'")

    # Get questions
    cur.execute("""
        SELECT title, question_type, required, config
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()

    check("Form has 5 questions", len(questions) == 5,
          f"Found {len(questions)} questions")

    # Expected questions in order
    expected_questions = [
        {
            "title": "Your Name",
            "type": "textQuestion",
            "required": True,
            "options": None,
        },
        {
            "title": "Your Team",
            "type": "choiceQuestion",
            "required": True,
            "options": ["Tier 1", "Tier 2", "Tier 3", "Specialist"],
        },
        {
            "title": "Which priority level needs most improvement?",
            "type": "choiceQuestion",
            "required": True,
            "options": ["Critical", "High", "Medium", "Low"],
        },
        {
            "title": "Proposed improvement actions",
            "type": "textQuestion",
            "required": True,
            "options": None,
        },
        {
            "title": "Target completion date",
            "type": "textQuestion",
            "required": True,
            "options": None,
        },
    ]

    for i, exp_q in enumerate(expected_questions):
        if i < len(questions):
            actual_title, actual_type, actual_required, actual_config = questions[i]

            # Check title (fuzzy match)
            title_match = exp_q["title"].lower() in actual_title.lower()
            check(f"Q{i+1} title contains '{exp_q['title']}'",
                  title_match,
                  f"Actual: '{actual_title}'")

            # Check type
            check(f"Q{i+1} type is '{exp_q['type']}'",
                  actual_type == exp_q["type"],
                  f"Actual: '{actual_type}'")

            # Check required
            check(f"Q{i+1} required is {exp_q['required']}",
                  actual_required == exp_q["required"],
                  f"Actual: {actual_required}")

            # Check options for choice questions
            if exp_q["options"] is not None and actual_config:
                config = actual_config if isinstance(actual_config, dict) else {}
                actual_options = []
                if "options" in config:
                    for opt in config["options"]:
                        if isinstance(opt, dict) and "value" in opt:
                            actual_options.append(opt["value"])
                        elif isinstance(opt, str):
                            actual_options.append(opt)

                for exp_opt in exp_q["options"]:
                    found = any(
                        exp_opt.lower() == ao.lower()
                        for ao in actual_options
                    )
                    check(f"Q{i+1} has option '{exp_opt}'", found,
                          f"Actual options: {actual_options}")
        else:
            check(f"Q{i+1} exists", False, "Question missing")

    cur.close()
    conn.close()


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    # Compute expected values from database
    print("=== Computing Expected Values from Database ===")
    try:
        expected = compute_expected_values()
        print(f"  Computed SLA compliance for {len(expected['sla_compliance'])} priorities")
        print(f"  Computed performance for {len(expected['agent_performance'])} agents")
        print(f"  Summary: compliance={expected['summary']['Overall_Compliance_Rate']}%, "
              f"satisfaction={expected['summary']['Overall_Avg_Satisfaction']}")
    except Exception as e:
        print(f"  ERROR computing expected values: {e}")
        import traceback
        traceback.print_exc()
        return False, f"Failed to compute expected values: {e}"

    # Run checks
    check_excel(agent_workspace, expected)
    check_form()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    success = pass_rate >= 0.8

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return success, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Rate: {pass_rate:.1%}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

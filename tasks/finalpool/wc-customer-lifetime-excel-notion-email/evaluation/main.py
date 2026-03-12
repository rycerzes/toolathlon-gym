"""
Evaluation script for wc-customer-lifetime-excel-notion-email task.

Checks:
1. Excel file (Customer_CLV_Report.xlsx) - 3 sheets with correct data
2. Notion database "Customer CRM" with >= 50 customer pages
3. Emails sent to at-risk customers with retention content
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_at_risk_emails():
    """Get expected at-risk customer emails from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.email, c.first_name, c.last_name, c.orders_count, c.total_spent
        FROM wc.customers c
        WHERE c.orders_count <= 1
        ORDER BY c.email
    """)
    single_order = {row[0].lower() for row in cur.fetchall()}

    # Also get customers whose last order > 60 days ago
    cur.execute("""
        SELECT c.email
        FROM wc.customers c
        LEFT JOIN (
            SELECT customer_id, MAX(date_created) as last_order
            FROM wc.orders
            WHERE status IN ('completed', 'processing', 'on-hold', 'pending')
            GROUP BY customer_id
        ) o ON c.id = o.customer_id
        WHERE c.orders_count > 1
          AND (o.last_order IS NULL OR o.last_order < NOW() - INTERVAL '60 days')
    """)
    stale_order = {row[0].lower() for row in cur.fetchall()}

    cur.close()
    conn.close()
    return single_order | stale_order


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Customer_CLV_Report.xlsx")
    gt_path = os.path.join(gt_workspace, "Customer_CLV_Report.xlsx")

    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return False

    if not os.path.isfile(gt_path):
        print(f"  WARNING: Groundtruth not found at {gt_path}, skipping comparison")
        return False

    try:
        agent_wb = openpyxl.load_workbook(excel_path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # --- Sheet 1: CLV Analysis ---
    print("  Checking CLV Analysis sheet...")
    a_rows = load_sheet_rows(agent_wb, "CLV Analysis")
    g_rows = load_sheet_rows(gt_wb, "CLV Analysis")

    if a_rows is None:
        check("CLV Analysis sheet exists", False, f"Sheets: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("CLV Analysis sheet in groundtruth", False)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("CLV Analysis row count", abs(len(a_data) - len(g_data)) <= 2,
              f"Agent {len(a_data)} vs GT {len(g_data)}")

        # Build lookup by customer name (col 0)
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing customer: {g_row[0]}")
                continue

            # Col 2: Orders_Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1):
                    errors.append(f"{key}.Orders_Count: {a_row[2]} vs {g_row[2]}")

            # Col 3: Total_Spent
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    errors.append(f"{key}.Total_Spent: {a_row[3]} vs {g_row[3]}")

            # Col 4: Avg_Order_Value
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 5.0):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[4]} vs {g_row[4]}")

            # Col 5: CLV_Tier
            if len(a_row) > 5 and len(g_row) > 5:
                if not str_match(a_row[5], g_row[5]):
                    errors.append(f"{key}.CLV_Tier: {a_row[5]} vs {g_row[5]}")

            # Col 6: Is_At_Risk
            if len(a_row) > 6 and len(g_row) > 6:
                if not str_match(a_row[6], g_row[6]):
                    errors.append(f"{key}.Is_At_Risk: {a_row[6]} vs {g_row[6]}")

        if errors:
            check("CLV Analysis data accuracy", False,
                  f"{len(errors)} errors: {'; '.join(errors[:5])}")
        else:
            check("CLV Analysis data accuracy", True)

    # --- Sheet 2: Tier Summary ---
    print("  Checking Tier Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Tier Summary")
    g_rows = load_sheet_rows(gt_wb, "Tier Summary")

    if a_rows is None:
        check("Tier Summary sheet exists", False, f"Sheets: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Tier Summary sheet in groundtruth", False)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing tier: {g_row[0]}")
                continue

            # Col 1: Customer_Count
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Customer_Count: {a_row[1]} vs {g_row[1]}")

            # Col 2: Total_Revenue
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 10.0):
                    errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]}")

            # Col 3: Avg_CLV
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 10.0):
                    errors.append(f"{key}.Avg_CLV: {a_row[3]} vs {g_row[3]}")

            # Col 4: At_Risk_Count
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 2):
                    errors.append(f"{key}.At_Risk_Count: {a_row[4]} vs {g_row[4]}")

        if errors:
            check("Tier Summary data accuracy", False,
                  f"{len(errors)} errors: {'; '.join(errors[:5])}")
        else:
            check("Tier Summary data accuracy", True)

    # --- Sheet 3: At Risk Customers ---
    print("  Checking At Risk Customers sheet...")
    a_rows = load_sheet_rows(agent_wb, "At Risk Customers")
    g_rows = load_sheet_rows(gt_wb, "At Risk Customers")

    if a_rows is None:
        check("At Risk Customers sheet exists", False, f"Sheets: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("At Risk Customers sheet in groundtruth", False)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("At Risk row count", abs(len(a_data) - len(g_data)) <= 3,
              f"Agent {len(a_data)} vs GT {len(g_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing at-risk customer: {g_row[0]}")
                continue

            # Col 3: Total_Spent
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    errors.append(f"{key}.Total_Spent: {a_row[3]} vs {g_row[3]}")

        if errors:
            check("At Risk data accuracy", False,
                  f"{len(errors)} errors: {'; '.join(errors[:5])}")
        else:
            check("At Risk data accuracy", True)

    return True


def check_notion():
    print("\n=== Checking Notion ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check for "Customer CRM" database
    cur.execute("SELECT id, title, properties FROM notion.databases")
    databases = cur.fetchall()

    crm_db_id = None
    for db_id, title, props in databases:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(t.get("plain_text", "") for t in title if isinstance(t, dict))
        elif isinstance(title, str):
            title_str = title
        if "customer" in title_str.lower() and "crm" in title_str.lower():
            crm_db_id = db_id
            break

    check("Notion 'Customer CRM' database exists", crm_db_id is not None,
          f"Found {len(databases)} databases", db=True)

    # Check pages in the CRM database
    if crm_db_id:
        cur.execute(
            "SELECT id, properties FROM notion.pages WHERE parent->>'database_id' = %s AND archived = false",
            (crm_db_id,),
        )
        pages = cur.fetchall()
        check("Notion CRM has >= 45 customer pages", len(pages) >= 45,
              f"Found {len(pages)} pages", db=True)
    else:
        # Fallback: check all pages for customer data
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false OR archived IS NULL")
        pages = cur.fetchall()
        customer_pages = 0
        for pid, props in pages:
            if props:
                props_str = json.dumps(props).lower() if isinstance(props, dict) else str(props).lower()
                if any(tier in props_str for tier in ["platinum", "gold", "silver", "bronze"]):
                    customer_pages += 1
        check("Notion has >= 45 customer pages (fallback)", customer_pages >= 45,
              f"Found {customer_pages} customer pages out of {len(pages)} total", db=True)

    cur.close()
    conn.close()


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get all sent emails
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()

    # Get expected at-risk emails
    at_risk_emails = get_at_risk_emails()

    # Filter for retention emails (not noise)
    retention_emails = []
    for subject, from_addr, to_addr, body_text in all_emails:
        subj_lower = (subject or "").lower()
        if "retention" in subj_lower or "customer" in subj_lower:
            # Parse to_addr
            recipients = set()
            if isinstance(to_addr, list):
                for r in to_addr:
                    recipients.add(str(r).strip().lower())
            elif isinstance(to_addr, str):
                try:
                    parsed = json.loads(to_addr)
                    if isinstance(parsed, list):
                        for r in parsed:
                            recipients.add(str(r).strip().lower())
                    else:
                        recipients.add(to_addr.strip().lower())
                except (json.JSONDecodeError, TypeError):
                    recipients.add(to_addr.strip().lower())
            retention_emails.append({
                "subject": subject,
                "from": from_addr,
                "to": recipients,
                "body": body_text or "",
            })

    print(f"  Found {len(retention_emails)} retention emails out of {len(all_emails)} total")
    print(f"  Expected at-risk customers: {len(at_risk_emails)}")

    # Check that we have a reasonable number of retention emails
    check("Retention emails sent (>= 50% of at-risk)",
          len(retention_emails) >= len(at_risk_emails) * 0.5,
          f"Got {len(retention_emails)}, expected >= {int(len(at_risk_emails) * 0.5)}", db=True)

    # Check that at-risk email addresses received emails
    matched = 0
    all_recipients = set()
    for em in retention_emails:
        all_recipients.update(em["to"])
    for ar_email in at_risk_emails:
        if ar_email in all_recipients:
            matched += 1

    check("At-risk customers received emails (>= 50%)",
          matched >= len(at_risk_emails) * 0.5,
          f"Matched {matched} out of {len(at_risk_emails)} at-risk customers", db=True)

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_notion()
    check_email()

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if args.res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": file_ok}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if file_ok else 1)


if __name__ == "__main__":
    main()

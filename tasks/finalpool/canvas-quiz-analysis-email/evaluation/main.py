"""
Evaluation script for canvas-quiz-analysis-email task.

Checks:
1. Excel Quiz_Performance.xlsx with Quiz Analysis sheet
2. Email sent to course instructor

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        d = (detail[:200] + "...") if len(detail) > 200 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    cur.execute("""
        SELECT q.title, q.points_possible,
               ROUND(AVG(qs.score)::numeric, 2) as avg_score,
               ROUND(100.0 * COUNT(CASE WHEN qs.score >= q.points_possible * 0.6 THEN 1 END) / COUNT(*)::numeric, 1) as pass_rate,
               COUNT(*) as student_count
        FROM canvas.quizzes q
        JOIN canvas.quiz_submissions qs ON q.id = qs.quiz_id
        WHERE q.course_id = 3
        GROUP BY q.id, q.title, q.points_possible
        ORDER BY pass_rate
    """)
    quiz_rows = cur.fetchall()

    # Get instructor email
    cur.execute("""
        SELECT u.email FROM canvas.enrollments e
        JOIN canvas.users u ON e.user_id = u.id
        WHERE e.course_id = 3 AND e.type = 'TeacherEnrollment'
        ORDER BY e.id LIMIT 1
    """)
    row = cur.fetchone()
    instructor_email = row[0] if row else None

    conn.close()
    return {"quizzes": quiz_rows, "instructor_email": instructor_email}


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Quiz_Performance.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Sheet 'Quiz Analysis' exists", get_sheet(wb, "Quiz Analysis") is not None,
          f"Found: {wb.sheetnames}")

    ws = get_sheet(wb, "Quiz Analysis")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["quizzes"]
        check("Quiz Analysis row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        # Build lookup by title
        agent_by_title = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_title[str(row[0]).strip().lower()] = row

        for exp_row in exp:
            title = exp_row[0]
            agent_row = agent_by_title.get(title.strip().lower())
            if agent_row:
                check(f"Quiz '{title}' Points_Possible",
                      num_close(agent_row[1], float(exp_row[1]), 0.01),
                      f"Expected {exp_row[1]}, got {agent_row[1]}")
                check(f"Quiz '{title}' Avg_Score",
                      num_close(agent_row[2], float(exp_row[2]), 1.0),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"Quiz '{title}' Pass_Rate_Pct",
                      num_close(agent_row[3], float(exp_row[3]), 1.0),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
                check(f"Quiz '{title}' Student_Count",
                      num_close(agent_row[4], exp_row[4], 5),
                      f"Expected {exp_row[4]}, got {agent_row[4]}")
            else:
                check(f"Quiz '{title}' found in output", False, "Not in agent output")

        # Check sort order (by pass rate ascending)
        if len(agent_rows) >= 2:
            pass_rates = [float(r[3]) for r in agent_rows if r and r[3] is not None]
            check("Sorted by Pass_Rate_Pct ascending",
                  all(pass_rates[i] <= pass_rates[i + 1] for i in range(len(pass_rates) - 1)),
                  f"Pass rates: {pass_rates}")


def check_email(expected):
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for email check", False, str(e), db=True)
        return

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()

    # Also check sent_log
    cur.execute("SELECT COUNT(*) FROM email.sent_log")
    sent_count = cur.fetchone()[0]

    # Also check drafts
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.drafts")
    drafts = cur.fetchall()

    conn.close()

    all_items = list(messages) + list(drafts)
    check("At least one email message or draft exists", len(all_items) > 0,
          f"Found {len(messages)} messages, {len(drafts)} drafts", db=True)

    found_email = False
    for item in all_items:
        subj = str(item[0] or "").lower()
        if "quiz" in subj and "performance" in subj:
            found_email = True
            check("Email subject contains 'Quiz Performance'", True, db=True)

            # Check recipient
            to_addr = item[2]
            if expected and expected.get("instructor_email"):
                exp_email = expected["instructor_email"].lower()
                to_str = json.dumps(to_addr).lower() if isinstance(to_addr, (list, dict)) else str(to_addr or "").lower()
                check("Email sent to instructor",
                      exp_email in to_str,
                      f"Expected to contain '{exp_email}', got '{to_str}'", db=True)

            # Check body has content
            body = str(item[3] or "")
            check("Email body is not empty", len(body) > 20,
                  f"Body length: {len(body)}", db=True)
            break

    if not found_email:
        check("Quiz Performance email found", False,
              f"Subjects: {[str(i[0]) for i in all_items]}", db=True)


def check_excel_gt(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Quiz_Performance.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quiz_Performance.xlsx")
    check("Excel file exists", os.path.isfile(agent_file))
    check("Groundtruth file exists", os.path.isfile(gt_file))
    if not os.path.isfile(agent_file) or not os.path.isfile(gt_file):
        return
    agent_wb = openpyxl.load_workbook(agent_file)
    gt_wb = openpyxl.load_workbook(gt_file)
    check("Sheet 'Quiz Analysis' exists", get_sheet(agent_wb, "Quiz Analysis") is not None)
    a_ws = get_sheet(agent_wb, "Quiz Analysis")
    g_ws = get_sheet(gt_wb, "Quiz Analysis")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Row count matches", len(a_rows) == len(g_rows),
              f"Expected {len(g_rows)}, got {len(a_rows)}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        check_excel(agent_workspace, expected)
    else:
        print("INFO: Falling back to groundtruth Excel")
        check_excel_gt(agent_workspace, groundtruth_workspace)

    check_email(expected)

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": file_ok}
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return file_ok, f"Passed: {total_pass}, Failed: {total_fail}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

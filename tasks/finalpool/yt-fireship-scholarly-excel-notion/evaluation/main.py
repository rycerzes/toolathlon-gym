"""
Evaluation for yt-fireship-scholarly-excel-notion task.

Checks:
1. Tech_Research_Map.xlsx exists with Video_Paper_Mapping and All_Papers sheets
2. Video_Paper_Mapping has 8 rows with correct structure
3. Top ranked video is the most-viewed (DeepSeek R1 bubble video)
4. All_Papers sheet has at least 8 papers
5. Notion page titled 'Tech Content Research Bridge' exists
6. Notion page has content mentioning top videos and papers
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Tech_Research_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Tech_Research_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Research_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Tech_Research_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Video_Paper_Mapping sheet",
           any("video" in s and "paper" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has All_Papers sheet",
           any("all" in s and "paper" in s for s in sheet_names_lower) or
           any("papers" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")

    # Check Video_Paper_Mapping content
    mapping_sheet = None
    for name in wb.sheetnames:
        if "video" in name.lower() and "paper" in name.lower():
            mapping_sheet = wb[name]
            break
        elif "mapping" in name.lower():
            mapping_sheet = wb[name]
            break

    if mapping_sheet:
        rows = list(mapping_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Video_Paper_Mapping has 8 rows", len(data_rows) >= 8,
               f"Found {len(data_rows)} rows")

        # Check headers
        if rows:
            headers_text = " ".join(str(c) for c in rows[0] if c).lower()
            record("Has Video_Rank, Video_Title, Video_Views columns",
                   "rank" in headers_text and "title" in headers_text and "view" in headers_text,
                   f"Headers: {rows[0]}")

        # Check top video is DeepSeek R1 bubble
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Contains DeepSeek R1 bubble video", "deepseek" in all_text and "bubble" in all_text,
               "No DeepSeek R1 bubble video found")
        record("Contains tech topics", "ai" in all_text and ("security" in all_text or "linux" in all_text),
               "Missing expected tech topics")

    # Check All_Papers
    papers_sheet = None
    for name in wb.sheetnames:
        if "all" in name.lower() and "paper" in name.lower():
            papers_sheet = wb[name]
            break
        elif "paper" in name.lower() and "mapping" not in name.lower():
            papers_sheet = wb[name]
            break

    if papers_sheet:
        rows = list(papers_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("All_Papers has at least 8 papers", len(data_rows) >= 8,
               f"Found {len(data_rows)} rows")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Tech_Research_Map.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion page 'Tech Content Research Bridge' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check notion pages table
    try:
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()
        matching = []
        for pid, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "tech content" in props_str or "research bridge" in props_str:
                matching.append((pid, props))
        record("Notion page 'Tech Content Research Bridge' exists", len(matching) >= 1,
               f"Found {len(pages)} total pages, {len(matching)} matching")

        if matching:
            page_id = matching[0][0]
            cur.execute("""
                SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s AND archived = false
            """, (page_id,))
            block_count = cur.fetchone()[0]
            record("Notion page has content blocks", block_count >= 3,
                   f"Found {block_count} blocks")

            cur.execute("""
                SELECT block_data FROM notion.blocks WHERE parent_id = %s AND archived = false
            """, (page_id,))
            blocks = cur.fetchall()
            all_text = " ".join(json.dumps(b[0]) for b in blocks if b[0]).lower()
            record("Notion page mentions DeepSeek or AI topics",
                   "deepseek" in all_text or "ai" in all_text or "fireship" in all_text,
                   "No relevant content found")
    except Exception as e:
        record("Notion page check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

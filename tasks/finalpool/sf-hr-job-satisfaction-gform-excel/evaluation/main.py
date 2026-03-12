"""Evaluation for sf-hr-job-satisfaction-gform-excel."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_satisfaction():
    """Query actual satisfaction data from DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               COUNT(*) as emp_count,
               AVG("JOB_SATISFACTION") as avg_js,
               AVG("WORK_LIFE_BALANCE") as avg_wlb
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*), AVG("JOB_SATISFACTION"), AVG("WORK_LIFE_BALANCE") FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    totals = cur.fetchone()
    cur.close()
    conn.close()
    return rows, totals


def check_excel(agent_workspace):
    errors = []
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl not installed")
        return errors

    agent_file = os.path.join(agent_workspace, "Employee_Satisfaction_Analysis.xlsx")
    if not os.path.exists(agent_file):
        errors.append("Employee_Satisfaction_Analysis.xlsx not found in agent workspace")
        return errors

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

    try:
        dept_rows, totals = get_expected_satisfaction()
    except Exception as e:
        errors.append(f"Cannot query expected data: {e}")
        return errors

    total_emp = int(totals[0])
    overall_avg_js = round(float(totals[1]), 2)
    overall_avg_wlb = round(float(totals[2]), 2)

    dept_data = {}
    for r in dept_rows:
        dept = str(r[0]).strip().lower()
        dept_data[dept] = {
            "count": int(r[1]),
            "avg_js": round(float(r[2]), 2),
            "avg_wlb": round(float(r[3]), 2),
            "combined": round((float(r[2]) + float(r[3])) / 2, 2),
        }

    # Find highest/lowest by combined score
    sorted_depts = sorted(dept_data.items(), key=lambda x: x[1]["combined"], reverse=True)
    highest_dept = sorted_depts[0][0]
    lowest_dept = sorted_depts[-1][0]

    # Check Department Scores sheet
    a_rows = load_sheet_rows(agent_wb, "Department Scores")
    if a_rows is None:
        errors.append("Sheet 'Department Scores' not found in agent output")
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        if len(a_data) < 7:
            errors.append(f"Department Scores: expected 7 data rows, got {len(a_data)}")
        else:
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
            for dept_key, expected in dept_data.items():
                a_row = a_lookup.get(dept_key)
                if a_row is None:
                    errors.append(f"Missing department: {dept_key}")
                    continue
                # Employee_Count col 1
                if len(a_row) > 1 and not num_close(a_row[1], expected["count"], 10):
                    errors.append(f"{dept_key} Employee_Count: got {a_row[1]}, expected {expected['count']} (tol=10)")
                # Avg_Job_Satisfaction col 2
                if len(a_row) > 2 and not num_close(a_row[2], expected["avg_js"], 0.1):
                    errors.append(f"{dept_key} Avg_Job_Satisfaction: got {a_row[2]}, expected {expected['avg_js']} (tol=0.1)")
                # Avg_Work_Life_Balance col 3
                if len(a_row) > 3 and not num_close(a_row[3], expected["avg_wlb"], 0.1):
                    errors.append(f"{dept_key} Avg_Work_Life_Balance: got {a_row[3]}, expected {expected['avg_wlb']} (tol=0.1)")
                # Combined_Score col 4
                if len(a_row) > 4 and not num_close(a_row[4], expected["combined"], 0.1):
                    errors.append(f"{dept_key} Combined_Score: got {a_row[4]}, expected {expected['combined']} (tol=0.1)")

    # Check Summary sheet
    a_sum = load_sheet_rows(agent_wb, "Summary")
    if a_sum is None:
        errors.append("Sheet 'Summary' not found in agent output")
    else:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}

        # Total_Employees
        te = a_sum_data.get("total_employees")
        if te is None:
            errors.append("Summary missing Total_Employees")
        elif not num_close(te, total_emp, 100):
            errors.append(f"Total_Employees: got {te}, expected {total_emp} (tol=100)")

        # Highest_Satisfaction_Dept
        hsd = a_sum_data.get("highest_satisfaction_dept")
        if hsd is None:
            errors.append("Summary missing Highest_Satisfaction_Dept")
        elif str(hsd).strip().lower() != highest_dept:
            errors.append(f"Highest_Satisfaction_Dept: got '{hsd}', expected '{highest_dept}'")

        # Lowest_Satisfaction_Dept
        lsd = a_sum_data.get("lowest_satisfaction_dept")
        if lsd is None:
            errors.append("Summary missing Lowest_Satisfaction_Dept")
        elif str(lsd).strip().lower() != lowest_dept:
            errors.append(f"Lowest_Satisfaction_Dept: got '{lsd}', expected '{lowest_dept}'")

        # Overall_Avg_Satisfaction
        oas = a_sum_data.get("overall_avg_satisfaction")
        if oas is None:
            errors.append("Summary missing Overall_Avg_Satisfaction")
        elif not num_close(oas, overall_avg_js, 0.1):
            errors.append(f"Overall_Avg_Satisfaction: got {oas}, expected {overall_avg_js} (tol=0.1)")

    return errors


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT f.id, f.title, COUNT(q.id) as q_count
            FROM gform.forms f
            LEFT JOIN gform.questions q ON q.form_id = f.id
            WHERE LOWER(f.title) LIKE '%wellbeing%' OR LOWER(f.title) LIKE '%employee%survey%'
            GROUP BY f.id, f.title
        """)
        forms = cur.fetchall()
        cur.close()
        conn.close()
        if not forms:
            errors.append("No Google Form named 'Employee Wellbeing Survey 2026' found")
        else:
            form = forms[0]
            if form[2] < 4:
                errors.append(f"Form has {form[2]} questions, expected at least 4")
    except Exception as e:
        errors.append(f"GForm DB check error: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%satisfaction%' OR LOWER(subject) LIKE '%wellbeing%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if not emails:
            errors.append("No email related to satisfaction or wellbeing found")
        else:
            found_to = False
            for em in emails:
                if "hr.team" in str(em[1]).lower():
                    found_to = True
                    break
            if not found_to:
                errors.append("No email sent to hr.team@company.com")
    except Exception as e:
        errors.append(f"Email DB check error: {e}")
    return errors


def check_notion():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties
            FROM notion.pages
            WHERE LOWER(properties::text) LIKE '%employee satisfaction%'
               OR LOWER(properties::text) LIKE '%satisfaction q1%'
        """)
        pages = cur.fetchall()
        if not pages:
            # Also check blocks for title content
            cur.execute("""
                SELECT id FROM notion.blocks
                WHERE LOWER(block_data::text) LIKE '%employee satisfaction%'
                   OR LOWER(block_data::text) LIKE '%satisfaction q1%'
            """)
            blocks = cur.fetchall()
            if not blocks:
                errors.append("No Notion page titled 'Employee Satisfaction Q1 2026' found")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Notion DB check error: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    print("\n=== Checking Google Form ===")
    gform_errors = check_gform()
    if gform_errors:
        for e in gform_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(gform_errors)
    else:
        print("  [PASS] GForm check passed")

    print("\n=== Checking Email ===")
    email_errors = check_email()
    if email_errors:
        for e in email_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(email_errors)
    else:
        print("  [PASS] Email check passed")

    print("\n=== Checking Notion Page ===")
    notion_errors = check_notion()
    if notion_errors:
        for e in notion_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(notion_errors)
    else:
        print("  [PASS] Notion check passed")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"errors": all_errors, "success": len(all_errors) == 0}, f, indent=2)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

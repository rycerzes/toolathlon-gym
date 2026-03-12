"""Evaluation for terminal-arxiv-scholarly-notion-word-excel."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

TRANSFORMER_IDS = {"1706.03762", "1810.04805", "2005.14165",
                   "1409.0473", "1910.10683", "2009.06732"}
NOISE_IDS = {"1207.00580", "1502.03167", "1312.06199"}
CATEGORIES = ["Architecture Design", "Training Methods", "Applications", "Survey"]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def check_excel(ws_path):
    """Check Research_Paper_Analysis.xlsx."""
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "Research_Paper_Analysis.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}

    # Paper_Catalog
    pc_name = sn.get("paper_catalog")
    if pc_name is None:
        check("Paper_Catalog sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Paper_Catalog sheet exists", True)
        ws = wb[pc_name]
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows[1:] if r and r[0] is not None]
        check("Paper_Catalog has 9 rows", len(data) == 9, f"Found {len(data)}")

        # Check that transformer papers are present
        ids_found = {str(r[0]).strip() for r in data}
        transformer_found = len(TRANSFORMER_IDS & ids_found)
        check("All 6 transformer papers listed", transformer_found == 6,
              f"Found {transformer_found}/6")

        # Check categories assigned
        cats_found = {str(r[5]).strip() if len(r) > 5 and r[5] else "" for r in data}
        valid_cats = sum(1 for c in cats_found if c in CATEGORIES)
        check("Valid categories assigned", valid_cats >= 2,
              f"Categories found: {cats_found}")

    # Method_Comparison
    mc_name = sn.get("method_comparison")
    if mc_name is None:
        check("Method_Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Method_Comparison sheet exists", True)
        ws2 = wb[mc_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Method_Comparison has 4 category rows", len(data2) == 4,
              f"Found {len(data2)}")

    # Citation_Matrix
    cm_name = sn.get("citation_matrix")
    if cm_name is None:
        check("Citation_Matrix sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Citation_Matrix sheet exists", True)
        ws3 = wb[cm_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("Citation_Matrix has 9 rows", len(data3) == 9, f"Found {len(data3)}")

        # Check overlap marking
        overlap_count = 0
        for r in data3:
            pid = str(r[0]).strip()
            if pid in TRANSFORMER_IDS:
                overlap_val = str(r[4]).strip().lower() if len(r) > 4 and r[4] else ""
                if "yes" in overlap_val:
                    overlap_count += 1
        check("6 transformer papers marked as overlap", overlap_count >= 5,
              f"Found {overlap_count}")

    wb.close()


def check_word(ws_path):
    """Check Transformer_Literature_Review.docx."""
    print("\n=== Checking Word Document ===")
    path = os.path.join(ws_path, "Transformer_Literature_Review.docx")
    if not os.path.isfile(path):
        check("Word document exists", False, f"Not found: {path}")
        return
    check("Word document exists", True)

    from docx import Document
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    check("Document mentions transformer", "transformer" in full_text)
    check("Document mentions attention", "attention" in full_text)
    check("Document mentions BERT or pre-training", "bert" in full_text or "pre-train" in full_text)
    check("Document has conclusion section",
          "conclusion" in full_text or "summary" in full_text)
    check("Document length >= 500 chars", len(full_text) >= 500,
          f"Length: {len(full_text)}")


def check_notion():
    """Check Notion database creation."""
    print("\n=== Checking Notion Database ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM notion.databases")
    databases = cur.fetchall()

    found_db = None
    for db_id, title_json in databases:
        title_str = ""
        if isinstance(title_json, list):
            for item in title_json:
                if isinstance(item, dict):
                    title_str += item.get("plain_text", "") or item.get("text", {}).get("content", "")
        elif isinstance(title_json, str):
            title_str = title_json
        if "research" in title_str.lower() and "paper" in title_str.lower():
            found_db = db_id
            break
        if "tracker" in title_str.lower():
            found_db = db_id
            break

    check("Research Paper Tracker database exists", found_db is not None,
          f"Databases: {[d[1] for d in databases]}")

    if found_db:
        cur.execute(
            "SELECT COUNT(*) FROM notion.pages WHERE parent->>'database_id' = %s",
            (found_db,)
        )
        page_count = cur.fetchone()[0]
        check("Notion database has >= 6 paper entries", page_count >= 6,
              f"Found {page_count} pages")

    conn.close()


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Research_Paper_Analysis.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            # Check no unexpected sheets beyond the 3 required
            valid_keywords = ["paper", "catalog", "method", "comparison", "citation", "matrix"]
            unexpected = [s for s in wb.sheetnames
                          if not any(k in s.lower() for k in valid_keywords) and s.lower() != "sheet1"]
            check("No unexpected sheets in Excel", len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")
            # Check no duplicate paper IDs
            sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}
            pc_name = sn.get("paper_catalog")
            if pc_name:
                ws = wb[pc_name]
                rows = list(ws.iter_rows(values_only=True))
                ids = [str(r[0]).strip() for r in rows[1:] if r and r[0]]
                check("No duplicate paper IDs in Paper_Catalog",
                      len(ids) == len(set(ids)),
                      f"Found {len(ids)} IDs but {len(set(ids))} unique")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-ARXIV-SCHOLARLY-NOTION-WORD-EXCEL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

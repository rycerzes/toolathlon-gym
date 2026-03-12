"""
Evaluation for howtocook-nutrition-excel-email task.
Checks Excel file structure (not specific recipe names) and email.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == c:
                return i
    return None


def check_excel(agent_workspace, groundtruth_workspace="."):
    """Check Excel file structure."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Breakfast_Nutrition_Report.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    # Check Recipes sheet
    recipes_rows = load_sheet_rows(wb, "Recipes")
    if recipes_rows is not None:
        record("Sheet 'Recipes' exists", True)
        header = recipes_rows[0] if recipes_rows else []
        data_rows = recipes_rows[1:] if len(recipes_rows) > 1 else []

        # Check columns exist
        name_col = find_col(header, ["Name", "name", "Recipe_Name", "recipe_name"])
        cat_col = find_col(header, ["Category", "category"])
        ing_col = find_col(header, ["Ingredients_Count", "ingredients_count", "Ingredients Count"])
        diff_col = find_col(header, ["Difficulty", "difficulty"])

        record("Name column exists", name_col is not None, f"Header: {header}")
        record("Category column exists", cat_col is not None, f"Header: {header}")
        record("Ingredients_Count column exists", ing_col is not None, f"Header: {header}")
        record("Difficulty column exists", diff_col is not None, f"Header: {header}")

        # Check has at least 1 recipe
        record("Recipes sheet has data rows (> 0)", len(data_rows) > 0,
               f"Found {len(data_rows)} rows")

        # Check difficulty values are valid
        if diff_col is not None and len(data_rows) > 0:
            valid_diffs = {"easy", "medium", "hard"}
            all_valid = True
            for row in data_rows:
                if diff_col < len(row) and row[diff_col]:
                    val = str(row[diff_col]).strip().lower()
                    if val not in valid_diffs:
                        all_valid = False
                        break
            record("Difficulty values are valid (easy/medium/hard)", all_valid)
    else:
        record("Sheet 'Recipes' exists", False, f"Available: {wb.sheetnames}")

    # Check Summary sheet
    summary_rows = load_sheet_rows(wb, "Summary")
    if summary_rows is not None:
        record("Sheet 'Summary' exists", True)

        metrics = {}
        for row in summary_rows:
            if row and row[0] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                val = row[1] if len(row) > 1 else None
                metrics[key] = val

        # Total_Recipes
        total_key = None
        for k in metrics:
            if "total" in k and "recipe" in k:
                total_key = k
                break
        if total_key:
            val = metrics[total_key]
            ok = val is not None and int(float(val)) > 0
            record("Summary: Total_Recipes > 0", ok, f"Got {val}")
        else:
            record("Summary: Total_Recipes exists", False, f"Keys: {list(metrics.keys())}")

        # Avg_Ingredients
        avg_key = None
        for k in metrics:
            if "avg" in k and "ingredient" in k:
                avg_key = k
                break
        if avg_key:
            val = metrics[avg_key]
            ok = val is not None and float(val) > 0
            record("Summary: Avg_Ingredients > 0", ok, f"Got {val}")
        else:
            record("Summary: Avg_Ingredients exists", False, f"Keys: {list(metrics.keys())}")

        # Easiest_Recipe
        easiest_key = None
        for k in metrics:
            if "easiest" in k or ("easy" in k and "recipe" in k):
                easiest_key = k
                break
        if easiest_key:
            val = metrics[easiest_key]
            ok = val is not None and len(str(val).strip()) > 0
            record("Summary: Easiest_Recipe has value", ok, f"Got {val}")
        else:
            record("Summary: Easiest_Recipe exists", False, f"Keys: {list(metrics.keys())}")
    else:
        record("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Breakfast_Nutrition_Report.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_email():
    """Check email was sent."""
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT id, subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%%breakfast%%'
               OR subject ILIKE '%%nutrition%%report%%'
        """)
        emails = cur.fetchall()

        record("Email with breakfast/nutrition subject sent", len(emails) > 0,
               "No matching email found")

        if emails:
            email = emails[0]

            # Check from
            from_addr = str(email[2]).lower() if email[2] else ""
            record("Email from nutrition-team@company.com",
                   "nutrition" in from_addr or "team" in from_addr,
                   f"From: {email[2]}")

            # Check to
            to_addr = email[3]
            if isinstance(to_addr, str):
                to_addr = json.loads(to_addr)
            to_str = str(to_addr).lower()
            record("Email to dietician@company.com",
                   "dietician@company.com" in to_str,
                   f"To: {to_addr}")

            # Check subject
            subject = str(email[1]).lower() if email[1] else ""
            record("Subject contains 'Breakfast Recipe Nutrition Report'",
                   "breakfast" in subject and "nutrition" in subject,
                   f"Subject: {email[1]}")

            # Check body has content
            body = str(email[4]) if email[4] else ""
            record("Email body has content", len(body) > 50,
                   f"Body length: {len(body)}")

        conn.close()
    except Exception as e:
        record("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

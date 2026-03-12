"""
Evaluation script for sf-training-needs-assessment task.

Checks:
1. Excel file Training_Needs.xlsx with 3 sheets
2. Emails sent to department heads
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Training catalog - cheapest department-specific course per department
COURSE_MAP = {
    "Engineering": ("Technical Deep Dive", 1800),
    "Finance": ("Data Analysis Fundamentals", 900),
    "HR": ("People Management", 600),
    "Operations": ("Process Optimization", 1100),
    "R&D": ("Innovation Workshop", 1400),
    "Sales": ("Negotiation Skills", 950),
    "Support": ("Customer Service Mastery", 700),
}

DEPT_HEADS = {
    "Engineering": "eng-head@company.com",
    "Sales": "sales-head@company.com",
    "HR": "hr-head@company.com",
    "Finance": "finance-head@company.com",
    "R&D": "rd-head@company.com",
    "Operations": "ops-head@company.com",
    "Support": "support-head@company.com",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Count low performers by department
    cur.execute("""
        SELECT "DEPARTMENT", COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        WHERE "PERFORMANCE_RATING" = 1
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    dept_counts = {}
    for dept, cnt in cur.fetchall():
        dept_counts[dept] = cnt

    total_employees = sum(dept_counts.values())

    # Recommended courses
    recommended = []
    total_est_cost = 0
    for dept in sorted(COURSE_MAP.keys()):
        title, cost = COURSE_MAP[dept]
        count = dept_counts.get(dept, 0)
        total = cost * count
        total_est_cost += total
        recommended.append({
            "Department": dept,
            "Course_Title": title,
            "Course_Cost": cost,
            "Employee_Count": count,
            "Total_Cost": total,
        })

    summary = {
        "Total_Eligible_Employees": total_employees,
        "Total_Departments": len(dept_counts),
        "Estimated_Total_Cost": total_est_cost,
        "Available_Budget": 500000,
        "Budget_Remaining": 500000 - total_est_cost,
        "Within_Budget": "Yes" if total_est_cost <= 500000 else "No",
    }

    cur.close()
    conn.close()

    return {
        "dept_counts": dept_counts,
        "recommended": recommended,
        "summary": summary,
        "total_employees": total_employees,
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Training_Needs.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    for sn in ["Low Performers", "Recommended Courses", "Budget Summary"]:
        found = any(str_match(s, sn) for s in wb.sheetnames)
        check(f"Sheet '{sn}' exists", found, f"Found: {wb.sheetnames}")

    # --- Low Performers ---
    print("\n--- Low Performers ---")
    ws = get_sheet(wb, "Low Performers")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Low Performers has data", len(rows) > 0)
        check("Low Performers count",
              num_close(len(rows), expected["total_employees"], 10),
              f"Expected ~{expected['total_employees']}, got {len(rows)}")

        # Check department distribution
        dept_dist = {}
        for r in rows:
            if r and r[1]:
                d = str(r[1]).strip()
                dept_dist[d] = dept_dist.get(d, 0) + 1
        for dept, exp_cnt in expected["dept_counts"].items():
            actual = dept_dist.get(dept, 0)
            check(f"Dept '{dept}' count",
                  num_close(actual, exp_cnt, 5),
                  f"Expected {exp_cnt}, got {actual}")

    # --- Recommended Courses ---
    print("\n--- Recommended Courses ---")
    ws = get_sheet(wb, "Recommended Courses")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Recommended Courses row count", len(rows) == 7,
              f"Expected 7, got {len(rows)}")

        for exp_row in expected["recommended"]:
            dept = exp_row["Department"]
            matched = None
            for r in rows:
                if r and str_match(r[0], dept):
                    matched = r
                    break
            if matched:
                check(f"{dept} Course_Cost",
                      num_close(matched[2], exp_row["Course_Cost"], 1),
                      f"Expected {exp_row['Course_Cost']}, got {matched[2]}")
                check(f"{dept} Employee_Count",
                      num_close(matched[3], exp_row["Employee_Count"], 5),
                      f"Expected {exp_row['Employee_Count']}, got {matched[3]}")
                check(f"{dept} Total_Cost",
                      num_close(matched[4], exp_row["Total_Cost"], exp_row["Total_Cost"] * 0.05),
                      f"Expected {exp_row['Total_Cost']}, got {matched[4]}")
            else:
                check(f"Dept '{dept}' found", False, "Not in output")

    # --- Budget Summary ---
    print("\n--- Budget Summary ---")
    ws = get_sheet(wb, "Budget Summary")
    if ws:
        data = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                data[str(row[0]).strip().lower().replace(" ", "_")] = row[1]

        for key, gt_val in expected["summary"].items():
            key_lower = key.lower()
            agent_val = data.get(key_lower)
            if agent_val is None:
                for ak, av in data.items():
                    if key_lower.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                tol = max(1.0, abs(gt_val) * 0.05)
                check(f"Summary '{key}'",
                      num_close(agent_val, gt_val, tol),
                      f"Expected {gt_val}, got {agent_val}")
            else:
                check(f"Summary '{key}'",
                      str_match(agent_val, gt_val),
                      f"Expected '{gt_val}', got '{agent_val}'")


def check_emails(expected):
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check messages table for sent training emails
    cur.execute("""
        SELECT subject, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%training%'
        ORDER BY subject
    """)
    all_emails = cur.fetchall()

    # Also check sent_log by joining back to messages
    try:
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text
            FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE LOWER(m.subject) LIKE '%training%'
            ORDER BY m.subject
        """)
        sent = cur.fetchall()
        all_emails = all_emails + sent
    except Exception:
        pass

    check("Training emails sent", len(all_emails) >= 7,
          f"Found {len(all_emails)} training emails")

    # Check each department head received an email
    for dept, email_addr in DEPT_HEADS.items():
        found = False
        for subj, to_addr, body in all_emails:
            to_str = str(to_addr).lower() if to_addr else ""
            if email_addr.lower() in to_str:
                found = True
                break
        check(f"Email to {dept} head ({email_addr})", found)

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Computing Expected Values ===")
    try:
        expected = compute_expected_values()
        print(f"  Total eligible: {expected['total_employees']}")
        print(f"  Departments: {len(expected['dept_counts'])}")
        print(f"  Estimated cost: {expected['summary']['Estimated_Total_Cost']}")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    check_excel(args.agent_workspace, expected)
    check_emails(expected)

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    success = pass_rate >= 0.8

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

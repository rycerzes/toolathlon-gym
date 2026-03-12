"""Evaluation for sf-support-coaching-plan-excel-notion-gcal.

Checks:
1. Excel Agent_Scorecard.xlsx with Performance Metrics, Coaching Plan, and Summary sheets
2. Notion database "Agent Coaching Tracker" with 5 agent pages
3. Google Calendar with 5 coaching session events in March 16-20, 2026
"""
import argparse
import json
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

AGENT_NAMES = ["Alice", "Bob", "Charlie", "Emily", "John"]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    print("\n=== Checking Excel ===")
    try:
        import openpyxl
    except ImportError:
        check("openpyxl installed", False, "openpyxl not available")
        return

    agent_file = os.path.join(agent_workspace, "Agent_Scorecard.xlsx")
    gt_file = os.path.join(gt_dir, "Agent_Scorecard.xlsx")

    check("Agent_Scorecard.xlsx exists", os.path.exists(agent_file),
          f"Not found at {agent_file}")
    if not os.path.exists(agent_file) or not os.path.exists(gt_file):
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Performance Metrics sheet ---
    a_rows = load_sheet_rows(agent_wb, "Performance Metrics")
    g_rows = load_sheet_rows(gt_wb, "Performance Metrics")
    check("Performance Metrics sheet exists", a_rows is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_rows and g_rows:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        check("Performance Metrics has 5 data rows", len(a_data) >= 5,
              f"Got {len(a_data)}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Agent {g_row[0]} in Performance Metrics", False, "Missing")
                continue
            # Total_Tickets (col 1)
            check(f"{g_row[0]} Total_Tickets",
                  num_close(a_row[1], g_row[1], 10),
                  f"got {a_row[1]}, expected {g_row[1]}")
            # SLA_Compliance_Pct (col 3)
            check(f"{g_row[0]} SLA_Compliance_Pct",
                  num_close(a_row[3], g_row[3], 2.0),
                  f"got {a_row[3]}, expected {g_row[3]}")
            # Avg_CSAT (col 4)
            check(f"{g_row[0]} Avg_CSAT",
                  num_close(a_row[4], g_row[4], 0.1),
                  f"got {a_row[4]}, expected {g_row[4]}")

    # --- Coaching Plan sheet ---
    a_cp = load_sheet_rows(agent_wb, "Coaching Plan")
    g_cp = load_sheet_rows(gt_wb, "Coaching Plan")
    check("Coaching Plan sheet exists", a_cp is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_cp and g_cp:
        a_cp_data = [r for r in a_cp[1:] if r and r[0] is not None]
        g_cp_data = [r for r in g_cp[1:] if r and r[0] is not None]
        check("Coaching Plan has 5 data rows", len(a_cp_data) >= 5,
              f"Got {len(a_cp_data)}")

        a_cp_lookup = {str(r[0]).strip().lower(): r for r in a_cp_data}
        for g_row in g_cp_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_cp_lookup.get(key)
            if a_row is None:
                check(f"Agent {g_row[0]} in Coaching Plan", False, "Missing")
                continue
            # Performance_Tier (col 1)
            check(f"{g_row[0]} Performance_Tier",
                  str_match(a_row[1], g_row[1]),
                  f"got '{a_row[1]}', expected '{g_row[1]}'")
            # Coaching_Frequency (col 2)
            check(f"{g_row[0]} Coaching_Frequency",
                  str_match(a_row[2], g_row[2]),
                  f"got '{a_row[2]}', expected '{g_row[2]}'")

    # --- Summary sheet ---
    a_sum = load_sheet_rows(agent_wb, "Summary")
    check("Summary sheet exists", a_sum is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_sum:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in a_sum[1:] if r and r[0]}
        tt = a_sum_data.get("total_tickets")
        check("Summary Total_Tickets", num_close(tt, 31588, 50),
              f"got {tt}, expected 31588")
        ta = a_sum_data.get("total_agents")
        check("Summary Total_Agents", num_close(ta, 5, 0),
              f"got {ta}, expected 5")


def check_notion():
    print("\n=== Checking Notion ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Check database exists
    cur.execute("""
        SELECT id, title::text, properties FROM notion.databases
        WHERE LOWER(title::text) LIKE '%agent%coaching%' OR LOWER(title::text) LIKE '%coaching%tracker%'
    """)
    dbs = cur.fetchall()
    check("Agent Coaching Tracker database exists", len(dbs) >= 1,
          f"Found {len(dbs)} databases: {[d[1] for d in dbs]}")

    if not dbs:
        cur.close()
        conn.close()
        return

    db_id = dbs[0][0]

    # Check pages in the database
    cur.execute("""
        SELECT id, properties FROM notion.pages
        WHERE parent::text LIKE %s AND archived = false
    """, (f"%{db_id}%",))
    pages = cur.fetchall()
    check("At least 5 agent pages in database", len(pages) >= 5,
          f"Found {len(pages)} pages")

    # Check that agent names appear in page properties
    all_props_text = " ".join(str(p[1]) for p in pages).lower()
    found_agents = sum(1 for a in AGENT_NAMES if a.lower() in all_props_text)
    check("Agent names found in Notion pages", found_agents >= 4,
          f"Found {found_agents}/5 agents")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%coaching%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("Coaching session events exist", len(events) >= 5,
          f"Found {len(events)} coaching events")

    if events:
        # Check they are in March 16-20, 2026
        march_events = [e for e in events
                        if e[2] and e[2].year == 2026 and e[2].month == 3
                        and 16 <= e[2].day <= 20]
        check("Coaching events in March 16-20 2026", len(march_events) >= 5,
              f"Found {len(march_events)} events in target week, total={len(events)}")

        # Check agent names appear in event summaries
        all_summaries = " ".join((e[0] or "") for e in events).lower()
        found = sum(1 for a in AGENT_NAMES if a.lower() in all_summaries)
        check("Agent names in coaching event titles", found >= 4,
              f"Found {found}/5 agent names in summaries")

        # Check 30-minute duration
        valid_duration = 0
        for e in events:
            if e[2] and e[3]:
                delta = (e[3] - e[2]).total_seconds()
                if 25 * 60 <= delta <= 35 * 60:
                    valid_duration += 1
        check("Coaching sessions are ~30 minutes", valid_duration >= 4,
              f"{valid_duration} events have 30-min duration")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_notion()
    check_gcal()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": overall}, f, indent=2)

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

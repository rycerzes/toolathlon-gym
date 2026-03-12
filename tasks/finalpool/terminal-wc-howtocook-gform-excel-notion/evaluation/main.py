"""Evaluation for terminal-wc-howtocook-gform-excel-notion.
Checks:
1. Meal_Kit_Analysis.xlsx with 4 sheets (Appliance_Catalog, Recipe_Matches, Survey_Results, Product_Roadmap)
2. Google Form "Meal Kit Interest Survey" with 6 questions
3. Notion "Meal Kit Development Tracker" with 5 kit pages
4. appliance_recipe_matcher.py script exists
5. appliance_recipe_matches.json exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except:
        return False


def check_excel(workspace):
    print("\n=== Check 1: Meal_Kit_Analysis.xlsx ===")
    path = os.path.join(workspace, "Meal_Kit_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Appliance_Catalog
    ac_idx = next((i for i, s in enumerate(sheets_lower) if "appliance" in s or "catalog" in s), 0)
    ws1 = wb[sheets[ac_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Appliance_Catalog has 8 products", len(data1) >= 8, f"Found {len(data1)}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has price column", any("price" in h for h in headers), f"Headers: {rows1[0]}")
        check("Has avg_rating column", any("rating" in h for h in headers), f"Headers: {rows1[0]}")
        check("Has compatible_recipes_count column",
              any("recipe" in h and "count" in h for h in headers) or any("compatible" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Check known appliance prices
    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains Blender product", "blender" in all_text1, f"Text snippet: {all_text1[:150]}")
    check("Contains Vacuum Sealer product", "vacuum" in all_text1 or "sealing" in all_text1,
          f"Text snippet: {all_text1[:150]}")
    check("Contains Cooking Pot product", "cooking pot" in all_text1 or "cooker" in all_text1 or "steamer" in all_text1,
          f"Text snippet: {all_text1[:150]}")

    # Verify a known price
    price_col = next((i for i, h in enumerate(headers) if "price" in h and "regular" not in h), -1) if rows1 else -1
    if price_col >= 0:
        prices = [r[price_col] for r in data1 if r[price_col] is not None]
        # Blender should be 214.00
        has_blender_price = any(num_close(p, 214.0, 1.0) for p in prices)
        check("Blender price ~214.00", has_blender_price, f"Prices: {prices}")

    # Recipe_Matches
    rm_idx = next((i for i, s in enumerate(sheets_lower) if "recipe" in s and "match" in s), 1)
    if rm_idx < len(sheets):
        ws2 = wb[sheets[rm_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Recipe_Matches has 10+ pairings", len(data2) >= 10, f"Found {len(data2)}")
        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has matched_appliance column",
                  any("appliance" in h or "matched" in h for h in headers2),
                  f"Headers: {rows2[0]}")
            check("Has difficulty column", any("difficult" in h for h in headers2),
                  f"Headers: {rows2[0]}")

    # Survey_Results
    sr_idx = next((i for i, s in enumerate(sheets_lower) if "survey" in s and "result" in s), 2)
    if sr_idx < len(sheets):
        ws3 = wb[sheets[sr_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Survey_Results has 6 rows", len(data3) >= 6, f"Found {len(data3)}")
        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            check("Has top_answer column", any("top" in h or "answer" in h for h in headers3),
                  f"Headers: {rows3[0]}")
            check("Has response_count column", any("count" in h or "response" in h for h in headers3),
                  f"Headers: {rows3[0]}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Survey mentions cooking frequency", "cook" in all_text3 or "often" in all_text3,
              f"Text: {all_text3[:100]}")
        # Top answer for cooking frequency should be "Several times a week" (12 responses)
        check("Top cooking frequency is 'Several times a week'",
              "several times a week" in all_text3,
              f"Text: {all_text3[:200]}")

    # Product_Roadmap
    pr_idx = next((i for i, s in enumerate(sheets_lower) if "roadmap" in s or "product" in s), 3)
    if pr_idx < len(sheets):
        ws4 = wb[sheets[pr_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Product_Roadmap has 5 kits", len(data4) >= 5, f"Found {len(data4)}")
        if rows4:
            headers4 = [str(c).lower() if c else "" for c in rows4[0]]
            check("Has priority column", any("priority" in h for h in headers4),
                  f"Headers: {rows4[0]}")
            check("Has estimated_price column", any("price" in h or "estimated" in h for h in headers4),
                  f"Headers: {rows4[0]}")

        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has High priority kits", "high" in all_text4, f"Text: {all_text4[:200]}")
        check("Has kit names", "kit" in all_text4, f"Text: {all_text4[:200]}")


def check_gform():
    print("\n=== Check 2: Google Form Survey ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        meal_form = None
        for form_id, title in forms:
            if title and ("meal kit" in title.lower() or "meal_kit" in title.lower()):
                meal_form = (form_id, title)
                break
        check("Meal Kit Interest Survey form exists", meal_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        if meal_form:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (meal_form[0],))
            q_count = cur.fetchone()[0]
            check("Survey has 6 questions", q_count >= 6, f"Found {q_count}")

            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position",
                        (meal_form[0],))
            questions = cur.fetchall()
            q_text = " ".join(str(q[0]) for q in questions).lower()
            check("Has cooking frequency question", "cook" in q_text or "often" in q_text,
                  f"Questions: {q_text[:150]}")
            check("Has cuisine question", "cuisine" in q_text, f"Questions: {q_text[:150]}")
            check("Has appliance ownership question", "appliance" in q_text or "own" in q_text,
                  f"Questions: {q_text[:150]}")
            check("Has budget question", "budget" in q_text, f"Questions: {q_text[:150]}")
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_notion():
    print("\n=== Check 3: Notion Meal Kit Development Tracker ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title, properties FROM notion.databases")
        dbs = cur.fetchall()
        tracker_db = None
        for db_id, title, props in dbs:
            title_str = ""
            if isinstance(title, list):
                title_str = " ".join(
                    item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
            elif isinstance(title, str):
                try:
                    parsed = json.loads(title)
                    if isinstance(parsed, list):
                        title_str = " ".join(
                            item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                    else:
                        title_str = str(title)
                except Exception:
                    title_str = str(title)
            else:
                title_str = str(title) if title else ""
            if "meal kit" in title_str.lower() and ("tracker" in title_str.lower() or "development" in title_str.lower()):
                tracker_db = (db_id, title_str, props)
                break
        check("Meal Kit Development Tracker database exists", tracker_db is not None,
              f"Databases: {[d[1] for d in dbs]}")

        if tracker_db:
            # Check properties
            props = tracker_db[2]
            if isinstance(props, str):
                try:
                    props = json.loads(props)
                except:
                    props = {}
            props_lower = {k.lower(): v for k, v in (props or {}).items()}
            check("Has Status property",
                  any("status" in k for k in props_lower),
                  f"Properties: {list(props_lower.keys())}")
            check("Has Priority property",
                  any("priority" in k for k in props_lower),
                  f"Properties: {list(props_lower.keys())}")
            check("Has Estimated_Revenue property",
                  any("revenue" in k or "estimated" in k for k in props_lower),
                  f"Properties: {list(props_lower.keys())}")

            # Check pages
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (tracker_db[0],))
            count = cur.fetchone()[0]
            check("Tracker has 5 kit pages", count >= 5, f"Found {count}")

            # Check page content
            cur.execute("""
                SELECT properties FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (tracker_db[0],))
            pages = cur.fetchall()
            all_page_text = ""
            has_planning = False
            has_high = False
            for (page_props,) in pages:
                if isinstance(page_props, str):
                    try:
                        page_props = json.loads(page_props)
                    except:
                        page_props = {}
                page_str = json.dumps(page_props).lower()
                all_page_text += page_str + " "
                if "planning" in page_str:
                    has_planning = True
                if '"high"' in page_str:
                    has_high = True
            check("Pages have Planning status", has_planning, f"Text: {all_page_text[:200]}")
            check("Pages have High priority", has_high, f"Text: {all_page_text[:200]}")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Scripts ===")
    matcher = os.path.join(workspace, "appliance_recipe_matcher.py")
    check("appliance_recipe_matcher.py exists", os.path.exists(matcher))

    matches_json = os.path.join(workspace, "appliance_recipe_matches.json")
    check("appliance_recipe_matches.json exists", os.path.exists(matches_json))
    if os.path.exists(matches_json):
        with open(matches_json) as f:
            data = json.load(f)
        check("JSON has appliance keys", len(data) >= 3, f"Keys: {list(data.keys())[:5]}")
        total_matches = sum(len(v) if isinstance(v, list) else 0 for v in data.values())
        check("JSON has 10+ total matches", total_matches >= 10, f"Total: {total_matches}")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Meal_Kit_Analysis.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        check("No negative values in Excel", False,
                              f"Found {cell} in sheet {sheet_name}")
                        return
        check("No negative values in Excel", True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_notion()
    check_scripts(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

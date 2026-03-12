"""Evaluation for terminal-yf-wc-excel-word-notion.
Checks:
1. Commodity_Impact_Analysis.xlsx with 4 sheets
2. Pricing_Strategy_Memo.docx
3. Notion "Market Research Dashboard" database with 2 entries
4. correlation_analysis.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: Commodity_Impact_Analysis.xlsx ===")
    path = os.path.join(workspace, "Commodity_Impact_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Stock_Trends
    st_idx = next((i for i, s in enumerate(sheets_lower) if "stock" in s or "trend" in s), 0)
    ws1 = wb[sheets[st_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Stock_Trends has 2 rows", len(data1) >= 2, f"Found {len(data1)}")
    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains GC=F", "gc=f" in all_text1 or "gold" in all_text1, f"Text: {all_text1[:100]}")
    check("Contains AMZN", "amzn" in all_text1 or "amazon" in all_text1, f"Text: {all_text1[:100]}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has volatility column", any("volatil" in h for h in headers) or any("std" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Product_Margins
    pm_idx = next((i for i, s in enumerate(sheets_lower) if "product" in s or "margin" in s), 1)
    if pm_idx < len(sheets):
        ws2 = wb[sheets[pm_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Product_Margins has category rows", len(data2) >= 3, f"Found {len(data2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Has Electronics category", "electronics" in all_text2)

    # Correlation_Analysis
    ca_idx = next((i for i, s in enumerate(sheets_lower) if "correlation" in s), 2)
    if ca_idx < len(sheets):
        ws3 = wb[sheets[ca_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Correlation_Analysis has entries", len(data3) >= 2, f"Found {len(data3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Mentions gold in correlation", "gold" in all_text3, f"Text: {all_text3[:100]}")

    # Strategic_Recommendations
    sr_idx = next((i for i, s in enumerate(sheets_lower) if "strategic" in s or "recommend" in s), 3)
    if sr_idx < len(sheets):
        ws4 = wb[sheets[sr_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Strategic_Recommendations has entries", len(data4) >= 3, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has margin targets", "35" in all_text4 or "40" in all_text4,
              f"Text: {all_text4[:100]}")


def check_word(workspace):
    print("\n=== Check 2: Pricing_Strategy_Memo.docx ===")
    path = os.path.join(workspace, "Pricing_Strategy_Memo.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    check("Mentions gold or commodity", "gold" in full_text or "commodity" in full_text)
    check("Mentions AMZN or Amazon", "amzn" in full_text or "amazon" in full_text)
    check("Mentions margin", "margin" in full_text)
    check("Has substantial content", len(full_text) > 200, f"Length: {len(full_text)}")


def check_notion():
    print("\n=== Check 3: Notion Market Research Dashboard ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        dashboard_db = None
        for db_id, title in dbs:
            title_str = ""
            if isinstance(title, list):
                title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
            elif isinstance(title, str):
                try:
                    parsed = json.loads(title)
                    if isinstance(parsed, list):
                        title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                    else:
                        title_str = str(title)
                except Exception:
                    title_str = str(title)
            else:
                title_str = str(title) if title else ""
            if "market" in title_str.lower() and ("research" in title_str.lower() or "dashboard" in title_str.lower()):
                dashboard_db = (db_id, title_str)
                break
        check("Market Research Dashboard exists", dashboard_db is not None,
              f"Databases: {[d[1] for d in dbs]}")

        if dashboard_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (dashboard_db[0],))
            count = cur.fetchone()[0]
            check("Dashboard has 2 entries", count >= 2, f"Found {count}")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 4: correlation_analysis.py ===")
    path = os.path.join(workspace, "correlation_analysis.py")
    check("correlation_analysis.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Commodity_Impact_Analysis.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"stock", "trend", "product", "margin", "correlation", "strategic", "recommend"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

    # Notion: no duplicate Market Research Dashboard databases
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        market_dbs = []
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "market" in title_str and "research" in title_str:
                market_dbs.append(db_id)
        check("No duplicate Market Research databases", len(market_dbs) <= 1,
              f"Found {len(market_dbs)} matching databases")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

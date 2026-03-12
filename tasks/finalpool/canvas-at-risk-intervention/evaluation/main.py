"""Evaluation for canvas-at-risk-intervention."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "At_Risk_Report.xlsx")
    if not os.path.isfile(xlsx_path):
        check("At_Risk_Report.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("At_Risk_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # Check At Risk Students sheet
    ar_rows = load_sheet_rows(wb, "At Risk Students")
    if ar_rows is None:
        check("Sheet 'At Risk Students' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'At Risk Students' exists", True)
        data_rows = ar_rows[1:] if len(ar_rows) > 1 else []
        # Should have many at-risk students
        check("At Risk Students has data (>100 rows)", len(data_rows) > 100,
              f"Found {len(data_rows)}")

        # Check header
        header = ar_rows[0] if ar_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["student_name", "course", "current_score", "risk_level", "recommended_support"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        # Verify risk levels
        if data_rows:
            risk_levels = set(str(r[3]).strip() if len(r) > 3 and r[3] else "" for r in data_rows)
            check("Risk levels include 'Critical' and 'Warning'",
                  "Critical" in risk_levels and "Warning" in risk_levels,
                  f"Found: {risk_levels}")

            # Verify scores are below 50
            all_below_50 = all(
                float(r[2]) < 50 for r in data_rows
                if len(r) > 2 and r[2] is not None
            )
            check("All scores below 50", all_below_50)

    # Check Course Summary sheet
    cs_rows = load_sheet_rows(wb, "Course Summary")
    if cs_rows is None:
        check("Sheet 'Course Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Course Summary' exists", True)
        data_rows = cs_rows[1:] if len(cs_rows) > 1 else []
        check("Course Summary has 22 rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        # Verify known data point
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM canvas.enrollments e
            WHERE e.type = 'StudentEnrollment'
              AND (e.grades->>'current_score')::numeric < 50
              AND e.grades->>'current_score' IS NOT NULL
              AND e.course_id = (SELECT id FROM canvas.courses WHERE name LIKE 'Creative Computing%%Spring%%' LIMIT 1)
        """)
        expected_ccc_spring = cur.fetchone()[0]
        cur.close()
        conn.close()

        found_ccc = False
        for row in data_rows:
            if row and row[0] and "creative computing" in str(row[0]).lower() and "spring" in str(row[0]).lower():
                found_ccc = True
                check(f"CCC Spring at-risk count ~{expected_ccc_spring}",
                      num_close(row[1], expected_ccc_spring, 5),
                      f"Got {row[1]}")
        check("Creative Computing (Spring) row found", found_ccc)

    # Check Action Plan sheet
    ap_rows = load_sheet_rows(wb, "Action Plan")
    if ap_rows is None:
        check("Sheet 'Action Plan' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Action Plan' exists", True)
        data_rows = ap_rows[1:] if len(ap_rows) > 1 else []
        check("Action Plan has at least 5 actions", len(data_rows) >= 5,
              f"Found {len(data_rows)}")


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%academic-support%%'
               OR to_addr::text ILIKE '%%academic_support%%'
               OR subject ILIKE '%%at-risk%%'
               OR subject ILIKE '%%at_risk%%'
               OR subject ILIKE '%%intervention%%'
        """)
        emails = cur.fetchall()
        check("Email sent about at-risk students", len(emails) >= 1,
              "No matching email found")
        if emails:
            body = str(emails[0][3]) if emails[0][3] else ""
            check("Email body has content", len(body) > 20,
                  f"Body length: {len(body)}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

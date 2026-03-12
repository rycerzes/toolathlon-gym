"""Evaluation for terminal-sf-hr-diversity-gform-excel-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Diversity_Metrics_Report.xlsx ===")
    agent_file = os.path.join(agent_ws, "Diversity_Metrics_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Diversity_Metrics_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
        gwb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Department_Breakdown
    print("  Checking Department_Breakdown...")
    ws = get_sheet(awb, "Department_Breakdown")
    gws = get_sheet(gwb, "Department_Breakdown")
    check("Sheet Department_Breakdown exists", ws is not None, f"Sheets: {awb.sheetnames}")
    if ws and gws:
        a_rows = list(ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(gws.iter_rows(min_row=2, values_only=True))
        check("Department_Breakdown has 35 rows", len(a_rows) == 35, f"Got {len(a_rows)}")

        a_lookup = {}
        for row in a_rows:
            if row and row[0] and row[1]:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[key] = row

        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Row {g_row[0]}/{g_row[1]} exists", False, "Missing")
                continue
            check(f"{g_row[0]}/{g_row[1]} Employee_Count",
                  num_close(a_row[2], g_row[2], 1),
                  f"Expected {g_row[2]}, got {a_row[2]}")
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"{g_row[0]}/{g_row[1]} Diversity_Index",
                      num_close(a_row[4], g_row[4], 0.01),
                      f"Expected {g_row[4]}, got {a_row[4]}")

    # Sheet 2: Education_Analysis
    print("  Checking Education_Analysis...")
    ws2 = get_sheet(awb, "Education_Analysis")
    gws2 = get_sheet(gwb, "Education_Analysis")
    check("Sheet Education_Analysis exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2 and gws2:
        a_rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        g_rows2 = list(gws2.iter_rows(min_row=2, values_only=True))
        check("Education_Analysis has 5 rows", len(a_rows2) == 5, f"Got {len(a_rows2)}")

        a_lookup2 = {str(r[0]).strip().lower(): r for r in a_rows2 if r and r[0]}
        for g_row in g_rows2:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup2.get(key)
            if a_row is None:
                check(f"Education '{g_row[0]}' exists", False, "Missing")
                continue
            check(f"'{g_row[0]}' Total_Count",
                  num_close(a_row[1], g_row[1], 1),
                  f"Expected {g_row[1]}, got {a_row[1]}")

    # Sheet 3: Survey_Config
    print("  Checking Survey_Config...")
    ws3 = get_sheet(awb, "Survey_Config")
    check("Sheet Survey_Config exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        a_rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        check("Survey_Config has 5 questions", len(a_rows3) == 5, f"Got {len(a_rows3)}")

    # Sheet 4: Summary
    print("  Checking Summary...")
    ws4 = get_sheet(awb, "Summary")
    gws4 = get_sheet(gwb, "Summary")
    check("Sheet Summary exists", ws4 is not None, f"Sheets: {awb.sheetnames}")
    if ws4 and gws4:
        a_summary = {}
        for row in ws4.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        g_summary = {}
        for row in gws4.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                g_summary[str(row[0]).strip().lower()] = row[1]

        # Dynamic DB queries for expected values
        expected_total = 50000
        expected_num_depts = 7
        try:
            conn2 = psycopg2.connect(**DB)
            cur2 = conn2.cursor()
            cur2.execute("SELECT COUNT(*) FROM sf_data.employees")
            result = cur2.fetchone()
            if result and result[0]:
                expected_total = result[0]
            cur2.execute("SELECT COUNT(DISTINCT department) FROM sf_data.employees")
            result = cur2.fetchone()
            if result and result[0]:
                expected_num_depts = result[0]
            cur2.close()
            conn2.close()
        except Exception:
            pass

        check("Total_Employees correct",
              num_close(a_summary.get("total_employees"), expected_total, 5),
              f"Got {a_summary.get('total_employees')}, expected {expected_total}")
        check("Num_Departments correct",
              num_close(a_summary.get("num_departments"), expected_num_depts, 0),
              f"Got {a_summary.get('num_departments')}, expected {expected_num_depts}")
        check("Avg_Diversity_Index close to groundtruth",
              num_close(a_summary.get("avg_diversity_index"),
                        g_summary.get("avg_diversity_index", 1.3481), 0.02),
              f"Got {a_summary.get('avg_diversity_index')}")
        check("Departments_Meeting_Target correct",
              num_close(a_summary.get("departments_meeting_target"),
                        g_summary.get("departments_meeting_target", 7), 0),
              f"Got {a_summary.get('departments_meeting_target')}")
        hdv = a_summary.get("highest_diversity_department")
        g_hdv = g_summary.get("highest_diversity_department", "HR")
        check("Highest_Diversity_Department matches groundtruth",
              hdv is not None and str(hdv).strip().lower() == str(g_hdv).strip().lower(),
              f"Got {hdv}, expected {g_hdv}")


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gform.forms WHERE LOWER(title) LIKE '%diversity%'")
        count = cur.fetchone()[0]
        check("Diversity survey form exists", count >= 1, f"Found {count}")
        if count >= 1:
            cur.execute("""
                SELECT f.id FROM gform.forms f
                WHERE LOWER(f.title) LIKE '%diversity%' LIMIT 1
            """)
            form_id = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
            qcount = cur.fetchone()[0]
            check("Form has 5 questions", qcount == 5, f"Got {qcount}")
        cur.close()
        conn.close()
    except Exception as e:
        check("GForm check", False, str(e))


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        found_db = None
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "diversity" in title_str and ("metric" in title_str or "dashboard" in title_str):
                found_db = db_id
                break
        check("Notion 'Diversity Metrics Dashboard' exists",
              found_db is not None, f"Found {len(dbs)} dbs")

        if found_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent::text LIKE %s AND archived = false AND in_trash = false
            """, (f'%{found_db}%',))
            page_count = cur.fetchone()[0]
            check("Notion DB has 7 department pages", page_count == 7,
                  f"Got {page_count}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no unexpected sheets beyond the 4 required
    path = os.path.join(workspace, "Diversity_Metrics_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        check("Excel has no more than 6 sheets", len(wb.sheetnames) <= 6,
              f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    # Notion: no duplicate diversity dashboard databases
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE archived = false AND LOWER(title::text) LIKE '%%diversity%%'
        """)
        db_count = cur.fetchone()[0]
        check("No duplicate Diversity Metrics databases", db_count <= 1,
              f"Found {db_count} diversity databases")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

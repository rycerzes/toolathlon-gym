"""Evaluation for canvas-course-comparison-excel."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

EXPECTED_COURSES = {
    "Applied Analytics & Algorithms (Fall 2013)": {"code": "AAA-2013J", "students": 383, "avg_score": 68.7, "assignments": 6, "quizzes": 0},
    "Biochemistry & Bioinformatics (Fall 2013)":  {"code": "BBB-2013J", "students": 2237, "avg_score": 77.0, "assignments": 12, "quizzes": 5},
    "Data-Driven Design (Fall 2013)":             {"code": "DDD-2013J", "students": 1938, "avg_score": 67.2, "assignments": 7, "quizzes": 0},
    "Environmental Economics & Ethics (Fall 2013)": {"code": "EEE-2013J", "students": 1052, "avg_score": 79.6, "assignments": 5, "quizzes": 0},
    "Foundations of Finance (Fall 2013)":          {"code": "FFF-2013J", "students": 2283, "avg_score": 74.9, "assignments": 13, "quizzes": 7},
    "Global Governance & Geopolitics (Fall 2013)": {"code": "GGG-2013J", "students": 952, "avg_score": 78.1, "assignments": 10, "quizzes": 6},
}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.lower() in str(haystack).lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_dir = args.agent_workspace or os.path.join(task_root, "initial_workspace")

    agent_file = os.path.join(agent_dir, "Course_Comparison_Fall2013.xlsx")
    gt_file = os.path.join(gt_dir, "Course_Comparison_Fall2013.xlsx")

    file_errors = []
    db_errors = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    if not file_errors:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Course Stats sheet
        print("  Checking Course Stats...")
        a_rows = load_sheet_rows(agent_wb, "Course Stats")
        g_rows = load_sheet_rows(gt_wb, "Course Stats")
        if a_rows is None:
            file_errors.append("Sheet 'Course Stats' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Course Stats' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            if len(a_data) != len(g_data):
                file_errors.append(f"Course Stats row count: agent {len(a_data)} vs gt {len(g_data)}")

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing course: {g_row[0]}")
                    continue

                # Course_Code (col 1)
                if len(a_row) > 1 and len(g_row) > 1:
                    if not str_match(a_row[1], g_row[1]):
                        file_errors.append(f"{g_row[0]}: code {a_row[1]} vs {g_row[1]}")

                # Student_Count (col 2)
                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 5):
                        file_errors.append(f"{g_row[0]}: students {a_row[2]} vs {g_row[2]}")

                # Avg_Final_Score (col 3)
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 1.0):
                        file_errors.append(f"{g_row[0]}: avg_score {a_row[3]} vs {g_row[3]}")

                # Assignment_Count (col 4)
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 1):
                        file_errors.append(f"{g_row[0]}: assignments {a_row[4]} vs {g_row[4]}")

                # Quiz_Count (col 5)
                if len(a_row) > 5 and len(g_row) > 5:
                    if not num_close(a_row[5], g_row[5], 1):
                        file_errors.append(f"{g_row[0]}: quizzes {a_row[5]} vs {g_row[5]}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary metric: {g_row[0]}")
                    continue

                if len(a_row) > 1 and len(g_row) > 1:
                    g_val = g_row[1]
                    a_val = a_row[1]
                    try:
                        fa, fb = float(a_val), float(g_val)
                        if abs(fa - fb) > 5:
                            file_errors.append(f"Summary {key}: {a_val} vs {g_val}")
                    except (TypeError, ValueError):
                        if not str_contains(str(a_val), str(g_val)[:20]):
                            file_errors.append(f"Summary {key}: '{a_val}' vs '{g_val}'")

    # Check email sent (DB check)
    print("  Checking email...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%academic-affairs@university.edu%'
               OR subject ILIKE '%Fall 2013%'
            LIMIT 5
        """)
        email_rows = cur.fetchall()
        if not email_rows:
            cur.execute("SELECT COUNT(*) FROM email.messages")
            total = cur.fetchone()[0]
            db_errors.append(f"No email to academic-affairs@university.edu found (total: {total})")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Email check error: {e}")

    # Final result
    print(f"\n=== SUMMARY ===")
    print(f"  File errors: {len(file_errors)}")
    print(f"  DB errors:   {len(db_errors)} (not blocking)")
    if db_errors:
        for e in db_errors[:15]:
            print(f"    [DB] {e}")
    if file_errors:
        for e in file_errors[:15]:
            print(f"    [FILE] {e}")
        print(f"  Overall: FAIL")
        sys.exit(1)
    else:
        print(f"  Overall: PASS")
        sys.exit(0)


if __name__ == "__main__":
    main()

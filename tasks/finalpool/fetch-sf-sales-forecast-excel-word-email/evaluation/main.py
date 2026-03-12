"""Evaluation script for fetch-sf-sales-forecast-excel-word-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Sales_Forecast_Q2_2026.xlsx")
    check("Sales_Forecast_Q2_2026.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Sales_Forecast_Q2_2026.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        check("Regional_Forecast sheet exists", "Regional_Forecast" in wb.sheetnames)
        if "Regional_Forecast" in wb.sheetnames:
            ws = wb["Regional_Forecast"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Regional_Forecast has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Region', 'Current_Revenue', 'Forecasted_Revenue']:
                check(f"Regional_Forecast has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Forecast_Summary sheet exists", "Forecast_Summary" in wb.sheetnames)
        if "Forecast_Summary" in wb.sheetnames:
            ws = wb["Forecast_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Forecast_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Forecast_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Growth_Ranking sheet exists", "Growth_Ranking" in wb.sheetnames)
        if "Growth_Ranking" in wb.sheetnames:
            ws = wb["Growth_Ranking"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Growth_Ranking has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Rank', 'Region', 'Growth_Rate_Pct']:
                check(f"Growth_Ranking has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # Check Word doc
        word_path = os.path.join(agent_workspace, "Q2_Forecast_Report.docx")
        check("Forecast Word report exists", os.path.exists(word_path))
        if os.path.exists(word_path):
            from docx import Document
            doc = Document(word_path)
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word mentions forecast", "forecast" in text)
            check("Word mentions growth", "growth" in text)

        # Check email
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT subject, to_addr FROM email.messages WHERE subject ILIKE %s", ('%forecast%',))
            emails = cur.fetchall()
            check("Forecast email sent", len(emails) >= 1)
            if emails:
                check("Email to sales team", "sales-team" in str(emails[0][1]).lower())
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        check("forecast_builder.py exists", os.path.exists(os.path.join(agent_workspace, "forecast_builder.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

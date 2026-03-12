"""
Evaluation script for gform-canvas-peer-review task.

Checks:
1. Excel file (Peer_Review_Analysis.xlsx) with 3 sheets and correct data
2. Google Sheet exists with "peer review" in title and has data
3. Notion page exists with "peer review" in properties and has content blocks
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=0.5):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected():
    """
    Compute expected values from the gform responses in the database.
    Returns raw_scores list, individual_summary dict, and flagged list.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get the form
    cur.execute(
        "SELECT id FROM gform.forms WHERE LOWER(title) LIKE '%peer review%' LIMIT 1"
    )
    form_row = cur.fetchone()
    if not form_row:
        cur.close()
        conn.close()
        return None, None, None

    form_id = form_row[0]

    # Get question IDs by title pattern
    cur.execute(
        "SELECT id, title FROM gform.questions WHERE form_id = %s ORDER BY position",
        (form_id,),
    )
    questions = cur.fetchall()
    q_map = {}
    for qid, qtitle in questions:
        title_lower = qtitle.lower()
        if "your name" in title_lower:
            q_map["reviewer"] = qid
        elif "person" in title_lower or "reviewed" in title_lower:
            q_map["reviewee"] = qid
        elif "contribution" in title_lower:
            q_map["contribution"] = qid
        elif "communication" in title_lower:
            q_map["communication"] = qid
        elif "quality" in title_lower:
            q_map["quality"] = qid
        elif "comment" in title_lower:
            q_map["comments"] = qid

    # Get all responses
    cur.execute(
        "SELECT answers FROM gform.responses WHERE form_id = %s",
        (form_id,),
    )
    response_rows = cur.fetchall()

    raw_scores = []
    for (answers_json,) in response_rows:
        answers = answers_json if isinstance(answers_json, dict) else json.loads(answers_json)
        reviewer = answers.get(q_map.get("reviewer", ""), "")
        reviewee = answers.get(q_map.get("reviewee", ""), "")
        contrib = int(answers.get(q_map.get("contribution", ""), 0))
        comm = int(answers.get(q_map.get("communication", ""), 0))
        quality = int(answers.get(q_map.get("quality", ""), 0))
        avg_score = round((contrib + comm + quality) / 3.0, 2)
        raw_scores.append({
            "Reviewer": reviewer,
            "Reviewee": reviewee,
            "Contribution": contrib,
            "Communication": comm,
            "Quality": quality,
            "Average_Score": avg_score,
        })

    # Compute individual summary
    from collections import defaultdict
    student_data = defaultdict(lambda: {"contrib": [], "comm": [], "quality": []})
    for r in raw_scores:
        name = r["Reviewee"]
        student_data[name]["contrib"].append(r["Contribution"])
        student_data[name]["comm"].append(r["Communication"])
        student_data[name]["quality"].append(r["Quality"])

    individual_summary = {}
    for name, data in student_data.items():
        avg_c = round(sum(data["contrib"]) / len(data["contrib"]), 2)
        avg_m = round(sum(data["comm"]) / len(data["comm"]), 2)
        avg_q = round(sum(data["quality"]) / len(data["quality"]), 2)
        overall = round((avg_c + avg_m + avg_q) / 3.0, 2)
        individual_summary[name] = {
            "Avg_Contribution": avg_c,
            "Avg_Communication": avg_m,
            "Avg_Quality": avg_q,
            "Overall_Avg": overall,
            "Review_Count": len(data["contrib"]),
        }

    # Flagged students
    flagged = []
    for name, stats in individual_summary.items():
        if stats["Overall_Avg"] < 3.0:
            flagged.append(name)

    cur.close()
    conn.close()

    return raw_scores, individual_summary, flagged


def get_sheet(wb, name):
    """Find sheet case-insensitively."""
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, raw_scores, individual_summary, flagged):
    """Check the Excel output file."""
    print("\n=== Checking Excel Output ===")

    excel_path = os.path.join(agent_workspace, "Peer_Review_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(excel_path),
          f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Excel file readable", True)

    # Check 3 sheets exist
    check("Has 'Raw Scores' sheet",
          any(str_match(s, "Raw Scores") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")
    check("Has 'Individual Summary' sheet",
          any(str_match(s, "Individual Summary") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")
    check("Has 'Flagged' sheet",
          any(str_match(s, "Flagged") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")

    # --- Raw Scores sheet ---
    print("\n--- Raw Scores ---")
    ws = get_sheet(wb, "Raw Scores")
    if ws:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Filter out empty rows
        data_rows = [r for r in data_rows if r and r[0] is not None]
        expected_count = len(raw_scores)
        check(f"Raw Scores has ~{expected_count} rows",
              abs(len(data_rows) - expected_count) <= 2,
              f"Expected ~{expected_count}, got {len(data_rows)}")

        # Check headers
        headers = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().replace("_", "").replace(" ", "") if h else "" for h in headers]
        check("Raw Scores has Reviewer column",
              any("reviewer" in h for h in header_lower),
              f"Headers: {headers}")
        check("Raw Scores has Reviewee column",
              any("reviewee" in h for h in header_lower),
              f"Headers: {headers}")
        check("Raw Scores has Contribution column",
              any("contribution" in h for h in header_lower),
              f"Headers: {headers}")
        check("Raw Scores has Average_Score column",
              any("average" in h or "avg" in h for h in header_lower),
              f"Headers: {headers}")
    else:
        check("Raw Scores sheet accessible", False, "Sheet not found")

    # --- Individual Summary sheet ---
    print("\n--- Individual Summary ---")
    ws = get_sheet(wb, "Individual Summary")
    if ws:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]
        check("Individual Summary has 6 rows (one per student)",
              len(data_rows) == 6,
              f"Expected 6, got {len(data_rows)}")

        # Check each student's data
        headers = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().replace("_", "").replace(" ", "") if h else "" for h in headers]

        # Find column indices
        name_col = None
        overall_col = None
        count_col = None
        for i, h in enumerate(header_lower):
            if "student" in h or "name" in h:
                name_col = i
            if "overall" in h:
                overall_col = i
            if "count" in h or "reviewcount" in h:
                count_col = i

        if name_col is not None and overall_col is not None:
            for row in data_rows:
                student_name = str(row[name_col]).strip() if row[name_col] else ""
                # Find matching expected student
                matched_exp = None
                for exp_name, exp_stats in individual_summary.items():
                    if exp_name.lower() in student_name.lower() or student_name.lower() in exp_name.lower():
                        matched_exp = (exp_name, exp_stats)
                        break

                if matched_exp:
                    exp_name, exp_stats = matched_exp
                    check(f"Student '{exp_name}' Overall_Avg",
                          num_close(row[overall_col], exp_stats["Overall_Avg"], 0.5),
                          f"Expected ~{exp_stats['Overall_Avg']}, got {row[overall_col]}")
                    if count_col is not None and row[count_col] is not None:
                        check(f"Student '{exp_name}' Review_Count",
                              int(row[count_col]) == exp_stats["Review_Count"],
                              f"Expected {exp_stats['Review_Count']}, got {row[count_col]}")
        else:
            check("Individual Summary has name and overall columns", False,
                  f"Could not identify name_col or overall_col in headers: {headers}")
    else:
        check("Individual Summary sheet accessible", False, "Sheet not found")

    # --- Flagged sheet ---
    print("\n--- Flagged ---")
    ws = get_sheet(wb, "Flagged")
    if ws:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]
        check("Flagged sheet has at least 1 student",
              len(data_rows) >= 1,
              f"Got {len(data_rows)} rows")

        # Check that flagged students are listed
        all_names = " ".join(str(r[0]).lower() for r in data_rows if r[0])
        for fname in flagged:
            check(f"Flagged: '{fname}' is listed",
                  fname.lower() in all_names,
                  f"Names in flagged sheet: {all_names}")

        # Check that flagged students have overall avg below 3.0
        headers = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().replace("_", "").replace(" ", "") if h else "" for h in headers]
        overall_col = None
        for i, h in enumerate(header_lower):
            if "overall" in h or "avg" in h:
                overall_col = i
                break
        if overall_col is not None and data_rows:
            for row in data_rows:
                if row[overall_col] is not None:
                    check(f"Flagged student avg < 3.0 ({row[0]})",
                          float(row[overall_col]) < 3.0,
                          f"Overall_Avg = {row[overall_col]}")
                    break  # Just check first flagged student
    else:
        check("Flagged sheet accessible", False, "Sheet not found")


def check_google_sheet():
    """Check that a Google Sheet was created with peer review data."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute(
        """SELECT id, title FROM gsheet.spreadsheets
           WHERE LOWER(title) LIKE '%peer review%'
           LIMIT 1"""
    )
    ss_row = cur.fetchone()
    check("Google Sheet with 'peer review' in title exists",
          ss_row is not None,
          "No spreadsheet found with 'peer review' in title")

    if ss_row:
        ss_id = ss_row[0]
        ss_title = ss_row[1]
        print(f"  Found spreadsheet: '{ss_title}' (id={ss_id})")

        # Check it has data (cells)
        cur.execute(
            "SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s",
            (ss_id,),
        )
        cell_count = cur.fetchone()[0]
        check("Google Sheet has data (cells > 0)",
              cell_count > 0,
              f"Found {cell_count} cells")

        # Check it has at least 7 rows (header + 6 students)
        cur.execute(
            """SELECT COUNT(DISTINCT row_index) FROM gsheet.cells
               WHERE spreadsheet_id = %s""",
            (ss_id,),
        )
        row_count = cur.fetchone()[0]
        check("Google Sheet has at least 7 rows (header + 6 students)",
              row_count >= 7,
              f"Found {row_count} distinct rows")

    cur.close()
    conn.close()


def check_notion(flagged):
    """Check that a Notion page was created with peer review summary."""
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Search for the page
    cur.execute(
        """SELECT id, properties FROM notion.pages
           WHERE LOWER(properties::text) LIKE '%peer review%'
           ORDER BY created_time DESC LIMIT 1"""
    )
    page_row = cur.fetchone()
    check("Notion page with 'peer review' in properties exists",
          page_row is not None,
          "No Notion page found with 'peer review' in properties")

    if page_row:
        page_id = page_row[0]
        properties = page_row[1] if isinstance(page_row[1], dict) else json.loads(page_row[1]) if page_row[1] else {}
        print(f"  Found page id={page_id}")

        # Check title contains biochemistry or project
        props_text = json.dumps(properties).lower()
        check("Notion page title references Biochemistry or project",
              "biochemistry" in props_text or "project" in props_text or "peer review" in props_text,
              f"Properties: {props_text[:200]}")

        # Check the page has content blocks
        cur.execute(
            "SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s",
            (page_id,),
        )
        block_count = cur.fetchone()[0]
        check("Notion page has content blocks",
              block_count > 0,
              f"Found {block_count} blocks")

        # Check content mentions flagged students
        cur.execute(
            """SELECT block_data::text FROM notion.blocks
               WHERE parent_id = %s""",
            (page_id,),
        )
        blocks = cur.fetchall()
        all_block_text = " ".join(str(b[0]).lower() for b in blocks if b[0])

        for fname in flagged:
            check(f"Notion page mentions flagged student '{fname}'",
                  fname.lower() in all_block_text,
                  f"Student name not found in block content")

        # Check for recommendation/follow-up content
        check("Notion page has recommendation or follow-up content",
              "recommend" in all_block_text or "follow" in all_block_text or "action" in all_block_text or "flag" in all_block_text,
              "No recommendation/follow-up keywords found in blocks")

    cur.close()
    conn.close()


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    print("=== Computing Expected Values from Database ===")
    try:
        raw_scores, individual_summary, flagged = compute_expected()
        if raw_scores is None:
            print("  ERROR: Could not find peer review form in database.")
            return False, "Form not found in database"
        print(f"  Found {len(raw_scores)} raw scores")
        print(f"  Found {len(individual_summary)} students in summary")
        print(f"  Flagged students: {flagged}")
    except Exception as e:
        print(f"  ERROR computing expected values: {e}")
        import traceback
        traceback.print_exc()
        return False, f"Failed to compute expected values: {e}"

    # Run all checks
    check_excel(agent_workspace, raw_scores, individual_summary, flagged)
    check_google_sheet()
    check_notion(flagged)

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": FAIL_COUNT == 0,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return FAIL_COUNT == 0, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Rate: {pass_rate:.1%}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

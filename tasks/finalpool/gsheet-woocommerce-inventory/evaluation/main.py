"""
Evaluation script for gsheet-woocommerce-inventory task.

Check 1: Inventory_Restock_Report.xlsx
  - "Category Inventory" sheet: 8 rows, correct values per category
  - "Restock Alerts" sheet: products with stock < 10
  - "Gold Trend" sheet: 5 rows of gold closing prices
Check 2: Google Sheet "Inventory Dashboard" exists with category data
Check 3: Email to warehouse@company.com with low stock alert
"""

import argparse
import os
import sys
import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_expected_category_inventory():
    """Query WooCommerce data for expected category inventory."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT
                jsonb_array_elements(categories)->>'name' AS cat,
                COUNT(*) AS product_count,
                SUM(stock_quantity) AS total_stock,
                ROUND(AVG(regular_price::float)::numeric, 2) AS avg_price
            FROM wc.products
            GROUP BY cat
            ORDER BY cat
        """)
        rows = cur.fetchall()
        return rows
    finally:
        cur.close()
        conn.close()


def get_expected_low_stock():
    """Query WooCommerce for products with stock < 10."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT
                name,
                sku,
                stock_quantity,
                categories->0->>'name' AS category,
                regular_price
            FROM wc.products
            WHERE stock_quantity < 10
            ORDER BY stock_quantity, name
        """)
        rows = cur.fetchall()
        return rows
    finally:
        cur.close()
        conn.close()


def get_expected_gold_trend():
    """Query Yahoo Finance for latest 5 gold closing prices."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT date, close
            FROM yf.stock_prices
            WHERE symbol = 'GC=F'
            ORDER BY date DESC
            LIMIT 5
        """)
        rows = cur.fetchall()
        return rows
    finally:
        cur.close()
        conn.close()


def check_excel(agent_workspace):
    """Check the Excel report against expected data."""
    print("[eval] Checking Excel file...")
    errors = []

    excel_path = os.path.join(agent_workspace, "Inventory_Restock_Report.xlsx")
    if not os.path.exists(excel_path):
        errors.append(f"Excel file not found: {excel_path}")
        return False, errors

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        errors.append(f"Cannot open Excel file: {e}")
        return False, errors

    # --- Check Category Inventory sheet ---
    if "Category Inventory" not in wb.sheetnames:
        errors.append(f"Sheet 'Category Inventory' not found. Available: {wb.sheetnames}")
    else:
        ws = wb["Category Inventory"]
        headers = [cell.value for cell in ws[1]]

        expected_cols = ["Category", "Product_Count", "Total_Stock", "Avg_Price"]
        col_map = {}
        for ec in expected_cols:
            for idx, h in enumerate(headers):
                if h and ec.lower().replace("_", "") == str(h).lower().replace("_", "").replace(" ", ""):
                    col_map[ec] = idx
                    break
            if ec not in col_map:
                for idx, h in enumerate(headers):
                    if h and ec.lower().replace("_", " ") in str(h).lower().replace("_", " "):
                        col_map[ec] = idx
                        break
            if ec not in col_map:
                errors.append(f"Column '{ec}' not found in Category Inventory headers: {headers}")

        if not errors:
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_map["Category"]] is not None:
                    data_rows.append(row)

            if len(data_rows) != 8:
                errors.append(f"Category Inventory: expected 8 rows, found {len(data_rows)}")

            expected_cats = get_expected_category_inventory()
            for exp_cat, exp_count, exp_stock, exp_price in expected_cats:
                found = False
                for row in data_rows:
                    row_cat = str(row[col_map["Category"]]).strip()
                    if row_cat.lower() == exp_cat.lower():
                        found = True
                        # Check Product_Count (tolerance 1)
                        pc = row[col_map["Product_Count"]]
                        if pc is not None and abs(int(pc) - exp_count) > 1:
                            errors.append(
                                f"Category '{exp_cat}': Product_Count {pc} != expected {exp_count}"
                            )
                        # Check Total_Stock (tolerance 5)
                        ts = row[col_map["Total_Stock"]]
                        if ts is not None and abs(int(ts) - exp_stock) > 5:
                            errors.append(
                                f"Category '{exp_cat}': Total_Stock {ts} != expected {exp_stock}"
                            )
                        # Check Avg_Price (tolerance 1.0)
                        ap = row[col_map["Avg_Price"]]
                        if ap is not None and abs(float(ap) - float(exp_price)) > 1.0:
                            errors.append(
                                f"Category '{exp_cat}': Avg_Price {ap} != expected {exp_price}"
                            )
                        break
                if not found:
                    errors.append(f"Category '{exp_cat}' not found in Category Inventory sheet")

    # --- Check Restock Alerts sheet ---
    if "Restock Alerts" not in wb.sheetnames:
        errors.append(f"Sheet 'Restock Alerts' not found. Available: {wb.sheetnames}")
    else:
        ws = wb["Restock Alerts"]
        headers = [cell.value for cell in ws[1]]

        expected_cols_ra = ["Product_Name", "SKU", "Current_Stock", "Category", "Regular_Price"]
        col_map_ra = {}
        for ec in expected_cols_ra:
            for idx, h in enumerate(headers):
                if h and ec.lower().replace("_", "") == str(h).lower().replace("_", "").replace(" ", ""):
                    col_map_ra[ec] = idx
                    break
            if ec not in col_map_ra:
                for idx, h in enumerate(headers):
                    if h and ec.lower().replace("_", " ") in str(h).lower().replace("_", " "):
                        col_map_ra[ec] = idx
                        break
            if ec not in col_map_ra:
                errors.append(f"Column '{ec}' not found in Restock Alerts headers: {headers}")

        if all(ec in col_map_ra for ec in expected_cols_ra):
            data_rows_ra = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_map_ra["Product_Name"]] is not None:
                    data_rows_ra.append(row)

            expected_low = get_expected_low_stock()
            expected_count = len(expected_low)

            if abs(len(data_rows_ra) - expected_count) > 2:
                errors.append(
                    f"Restock Alerts: expected ~{expected_count} rows, found {len(data_rows_ra)}"
                )

            # Verify a sample of products exist
            sample_skus = [row[1] for row in expected_low[:5]]
            found_skus = set()
            for row in data_rows_ra:
                sku_val = str(row[col_map_ra["SKU"]]).strip() if row[col_map_ra["SKU"]] else ""
                found_skus.add(sku_val)

            for sku in sample_skus:
                if sku not in found_skus:
                    errors.append(f"Restock Alerts: expected SKU '{sku}' not found")

    # --- Check Gold Trend sheet ---
    if "Gold Trend" not in wb.sheetnames:
        errors.append(f"Sheet 'Gold Trend' not found. Available: {wb.sheetnames}")
    else:
        ws = wb["Gold Trend"]
        headers = [cell.value for cell in ws[1]]

        col_map_gt = {}
        for ec in ["Date", "Close_Price"]:
            for idx, h in enumerate(headers):
                if h and ec.lower().replace("_", "") == str(h).lower().replace("_", "").replace(" ", ""):
                    col_map_gt[ec] = idx
                    break
            if ec not in col_map_gt:
                for idx, h in enumerate(headers):
                    if h and ec.lower().replace("_", " ") in str(h).lower().replace("_", " "):
                        col_map_gt[ec] = idx
                        break
            if ec not in col_map_gt:
                errors.append(f"Column '{ec}' not found in Gold Trend headers: {headers}")

        if all(ec in col_map_gt for ec in ["Date", "Close_Price"]):
            data_rows_gt = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_map_gt["Date"]] is not None:
                    data_rows_gt.append(row)

            if len(data_rows_gt) != 5:
                errors.append(f"Gold Trend: expected 5 rows, found {len(data_rows_gt)}")

            # Check first row close price matches expected
            expected_gold = get_expected_gold_trend()
            if data_rows_gt and expected_gold:
                actual_close = float(data_rows_gt[0][col_map_gt["Close_Price"]])
                expected_close = float(expected_gold[0][1])
                if abs(actual_close - expected_close) > 5.0:
                    errors.append(
                        f"Gold Trend: first row Close_Price {actual_close} != expected {expected_close}"
                    )

    if errors:
        return False, errors

    print("  Excel check passed.")
    return True, []


def check_gsheet():
    """Check that an Inventory Dashboard spreadsheet exists in gsheet schema."""
    print("[eval] Checking Google Sheet...")
    errors = []
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        # Find spreadsheet with "inventory" in title (case-insensitive)
        cur.execute(
            "SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%inventory%'"
        )
        spreadsheets = cur.fetchall()
        if not spreadsheets:
            errors.append("No spreadsheet found with 'inventory' in the title")
            return False, errors

        ss_id = spreadsheets[0][0]
        ss_title = spreadsheets[0][1]
        print(f"  Found spreadsheet: '{ss_title}' (id={ss_id})")

        # Check that cells exist
        cur.execute(
            "SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,)
        )
        cell_count = cur.fetchone()[0]
        if cell_count < 9:  # At least header + 8 data rows
            errors.append(
                f"Spreadsheet has only {cell_count} cells, expected at least 9 rows of data"
            )
            return False, errors

        # Check that category data is present by looking for known category names
        cur.execute(
            """SELECT value FROM gsheet.cells
               WHERE spreadsheet_id = %s AND LOWER(value) IN (
                   'audio', 'cameras', 'electronics', 'headphones',
                   'home appliances', 'speakers', 'tv & home theater', 'watches'
               )""",
            (ss_id,),
        )
        found_cats = cur.fetchall()
        if len(found_cats) < 5:
            errors.append(
                f"Spreadsheet has only {len(found_cats)} recognizable category names, expected at least 5"
            )
            return False, errors

        print(f"  Found {len(found_cats)} category entries in spreadsheet.")

    finally:
        cur.close()
        conn.close()

    if errors:
        return False, errors

    print("  Google Sheet check passed.")
    return True, []


def check_email():
    """Check that a low stock alert email was sent to warehouse@company.com."""
    print("[eval] Checking email...")
    errors = []
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        # Look for email with relevant subject
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE (LOWER(subject) LIKE '%low stock%' OR LOWER(subject) LIKE '%alert%' OR LOWER(subject) LIKE '%restock%')
        """)
        emails = cur.fetchall()

        if not emails:
            errors.append("No email found with 'low stock', 'alert', or 'restock' in subject")
            return False, errors

        # Check at least one email goes to warehouse@company.com
        found_target = False
        matched_email = None
        for subject, to_addr, body_text in emails:
            to_str = str(to_addr).lower() if to_addr else ""
            if "warehouse@company.com" in to_str:
                found_target = True
                matched_email = (subject, body_text)
                break

        if not found_target:
            errors.append(
                "No email to warehouse@company.com found among alert emails"
            )
            return False, errors

        subject, body = matched_email
        print(f"  Found email: subject='{subject}'")

        # Check body contains some product info
        if body:
            body_lower = body.lower()
            # Check that at least a few known low-stock SKUs or product names appear
            expected_low = get_expected_low_stock()
            matches = 0
            for name, sku, stock, cat, price in expected_low[:10]:
                if sku.lower() in body_lower or name[:30].lower() in body_lower:
                    matches += 1
            if matches < 3:
                errors.append(
                    f"Email body mentions only {matches} of the first 10 low-stock products (expected >= 3)"
                )
                return False, errors
        else:
            errors.append("Email body is empty")
            return False, errors

    finally:
        cur.close()
        conn.close()

    if errors:
        return False, errors

    print("  Email check passed.")
    return True, []


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    all_passed = True

    # Check 1: Excel file
    excel_pass, excel_errors = check_excel(args.agent_workspace)
    if not excel_pass:
        all_passed = False
        print("[FAIL] Excel check failed:")
        for e in excel_errors:
            print(f"  - {e}")
    else:
        print("[PASS] Excel check passed.")

    # Check 2: Google Sheet
    gsheet_pass, gsheet_errors = check_gsheet()
    if not gsheet_pass:
        all_passed = False
        print("[FAIL] Google Sheet check failed:")
        for e in gsheet_errors:
            print(f"  - {e}")
    else:
        print("[PASS] Google Sheet check passed.")

    # Check 3: Email
    email_pass, email_errors = check_email()
    if not email_pass:
        all_passed = False
        print("[FAIL] Email check failed:")
        for e in email_errors:
            print(f"  - {e}")
    else:
        print("[PASS] Email check passed.")

    if all_passed:
        print("\nAll checks passed!")
        sys.exit(0)
    else:
        print("\nSome checks failed.")
        sys.exit(1)


if __name__ == "__main__":
    main()

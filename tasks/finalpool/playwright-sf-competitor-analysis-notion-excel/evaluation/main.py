"""Evaluation for playwright-sf-competitor-analysis-notion-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Competitive_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Competitive_Analysis.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Competitive_Analysis.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Competitor Profiles sheet
        print("  Checking Competitor Profiles sheet...")
        a_rows = load_sheet_rows(agent_wb, "Competitor Profiles")
        g_rows = load_sheet_rows(gt_wb, "Competitor Profiles")
        if a_rows is None:
            all_errors.append("Sheet 'Competitor Profiles' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            if len(a_data) < 5:
                all_errors.append(f"Competitor Profiles row count: {len(a_data)}, expected 5")
            else:
                a_lookup = {}
                for row in a_data:
                    if row and row[0]:
                        a_lookup[str(row[0]).strip().lower()] = row
                for g_row in g_rows[1:]:
                    if not g_row or not g_row[0]:
                        continue
                    key = str(g_row[0]).strip().lower()
                    a_row = a_lookup.get(key)
                    if a_row is None:
                        all_errors.append(f"Missing competitor: {g_row[0]}")
                        continue
                    # Revenue
                    if not num_close(a_row[1], g_row[1], 50):
                        all_errors.append(f"{g_row[0]} Revenue: {a_row[1]} vs {g_row[1]}")
                    # Market Share
                    if not num_close(a_row[2], g_row[2], 0.5):
                        all_errors.append(f"{g_row[0]} Market_Share: {a_row[2]} vs {g_row[2]}")
                    # Position
                    if a_row[5] and g_row[5]:
                        if str(a_row[5]).strip().lower() != str(g_row[5]).strip().lower():
                            all_errors.append(f"{g_row[0]} Position: {a_row[5]} vs {g_row[5]}")
            print("    Done.")

        # Check Internal Performance sheet
        print("  Checking Internal Performance sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Internal Performance")
        if a_rows2 is None:
            all_errors.append("Sheet 'Internal Performance' not found in agent output")
        else:
            # Check for region data
            region_names = ["europe", "asia pacific", "north america", "middle east", "latin america"]
            found_regions = 0
            for row in a_rows2:
                if row and row[0] and str(row[0]).strip().lower() in region_names:
                    found_regions += 1
            if found_regions < 5:
                all_errors.append(f"Internal Performance: found {found_regions}/5 regions")

            # Check for segment data
            segment_names = ["consumer", "enterprise", "government", "smb"]
            found_segments = 0
            for row in a_rows2:
                if row and row[0] and str(row[0]).strip().lower() in segment_names:
                    found_segments += 1
            if found_segments < 4:
                all_errors.append(f"Internal Performance: found {found_segments}/4 segments")
            print("    Done.")

        # Check SWOT Summary sheet
        print("  Checking SWOT Summary sheet...")
        a_rows3 = load_sheet_rows(agent_wb, "SWOT Summary")
        if a_rows3 is None:
            all_errors.append("Sheet 'SWOT Summary' not found in agent output")
        else:
            a_data3 = a_rows3[1:] if len(a_rows3) > 1 else []
            if len(a_data3) < 4:
                all_errors.append(f"SWOT Summary: {len(a_data3)} rows, expected at least 4")
            print("    Done.")

    # --- Check 2: Notion database ---
    print("Checking Notion database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE title::text ILIKE '%competitive%intelligence%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            all_errors.append("Notion database 'Competitive Intelligence Tracker' not found")
        else:
            # Check for competitor pages
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages p
                JOIN notion.databases d ON p.parent::jsonb->>'database_id' = d.id
                WHERE d.title::text ILIKE '%competitive%intelligence%'
                AND p.archived = false
            """)
            page_count = cur.fetchone()[0]
            if page_count < 5:
                all_errors.append(f"Notion: found {page_count}/5 competitor pages")
            else:
                print(f"    Found {page_count} competitor pages")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking Notion: {e}")

    # --- Check 3: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%board@company.com%'
            AND subject ILIKE '%competitive%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            all_errors.append("No email sent to board@company.com about competitive analysis")
        else:
            print(f"    Email found ({count} messages)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

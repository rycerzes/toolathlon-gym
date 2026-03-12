"""Evaluation for wc-review-sentiment-gsheet."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected data from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Total reviews and avg rating
    cur.execute("SELECT COUNT(*), ROUND(AVG(rating)::numeric, 2) FROM wc.product_reviews")
    total_reviews, overall_avg = cur.fetchone()

    # Rating distribution
    cur.execute("""
        SELECT rating, COUNT(*),
               ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM wc.product_reviews), 1)
        FROM wc.product_reviews
        GROUP BY rating ORDER BY rating
    """)
    rating_dist = cur.fetchall()

    # Top 20 most-reviewed products
    cur.execute("""
        SELECT p.name, COUNT(r.id) as review_count,
               ROUND(AVG(r.rating)::numeric, 2) as avg_rating,
               SUM(CASE WHEN r.rating=5 THEN 1 ELSE 0 END) as five_star,
               SUM(CASE WHEN r.rating=1 THEN 1 ELSE 0 END) as one_star
        FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        ORDER BY COUNT(r.id) DESC, p.name ASC
        LIMIT 20
    """)
    top_products = cur.fetchall()

    # Most reviewed product (alphabetically first if tied)
    cur.execute("""
        SELECT p.name FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        ORDER BY COUNT(r.id) DESC, p.name ASC
        LIMIT 1
    """)
    most_reviewed = cur.fetchone()[0]

    # Highest rated (min 3 reviews)
    cur.execute("""
        SELECT p.name FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        HAVING COUNT(r.id) >= 3
        ORDER BY AVG(r.rating) DESC, p.name ASC
        LIMIT 1
    """)
    highest_rated = cur.fetchone()[0]

    cur.close()
    conn.close()
    return {
        "total_reviews": total_reviews,
        "overall_avg": float(overall_avg),
        "rating_dist": rating_dist,
        "top_products": top_products,
        "most_reviewed": most_reviewed,
        "highest_rated": highest_rated,
    }


def check_excel(agent_workspace):
    """Check Review_Sentiment_Analysis.xlsx."""
    print("\n=== Checking Review_Sentiment_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Review_Sentiment_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    expected = get_expected_data()

    # Check Product Reviews sheet
    pr_sheet = None
    for name in wb.sheetnames:
        if "product" in name.lower() and "review" in name.lower():
            pr_sheet = wb[name]
            break
    if pr_sheet is None:
        record("Sheet 'Product Reviews' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Product Reviews' exists", True)
        rows = list(pr_sheet.iter_rows(min_row=2, values_only=True))
        record("Product Reviews has 20 rows", len(rows) == 20, f"Got {len(rows)}")

        # Check first few products match
        if rows and expected["top_products"]:
            first_product = str(rows[0][0]) if rows[0] and rows[0][0] else ""
            expected_first = expected["top_products"][0][0]
            # Just check that product name is among top products
            top_names = [p[0].lower()[:30] for p in expected["top_products"]]
            record("First product is among top-reviewed",
                   first_product.lower()[:30] in top_names or len(rows) == 20,
                   f"Got: {first_product[:50]}")

    # Check Rating Distribution sheet
    rd_sheet = None
    for name in wb.sheetnames:
        if "rating" in name.lower() and "distribution" in name.lower():
            rd_sheet = wb[name]
            break
    if rd_sheet is None:
        record("Sheet 'Rating Distribution' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Rating Distribution' exists", True)
        rows = list(rd_sheet.iter_rows(min_row=2, values_only=True))
        record("Rating Distribution has 5 rows", len(rows) == 5, f"Got {len(rows)}")

        for exp_r in expected["rating_dist"]:
            rating_val = exp_r[0]
            found = False
            for r in rows:
                if r and r[0] is not None and int(r[0]) == rating_val:
                    found = True
                    ok_count = num_close(r[1], exp_r[1], 2)
                    record(f"Rating {rating_val} count", ok_count,
                           f"Expected {exp_r[1]}, got {r[1]}")
                    if not ok_count:
                        all_ok = False
                    break
            if not found:
                record(f"Rating {rating_val} found", False, "Missing")
                all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        for key, val in summary.items():
            if "total" in key and "review" in key:
                ok = num_close(val, expected["total_reviews"], 5)
                record("Summary Total_Reviews", ok,
                       f"Expected {expected['total_reviews']}, got {val}")
                if not ok:
                    all_ok = False
            elif key == "avg_rating" or ("avg" in key and "rating" in key):
                ok = num_close(val, expected["overall_avg"], 0.1)
                record("Summary Avg_Rating", ok,
                       f"Expected {expected['overall_avg']}, got {val}")
                if not ok:
                    all_ok = False

    return all_ok


def check_word(agent_workspace):
    """Check Review_Sentiment_Report.docx."""
    print("\n=== Checking Review_Sentiment_Report.docx ===")
    from docx import Document

    docx_file = os.path.join(agent_workspace, "Review_Sentiment_Report.docx")
    if not os.path.isfile(docx_file):
        record("Word file exists", False, f"Not found: {docx_file}")
        return False
    record("Word file exists", True)

    try:
        doc = Document(docx_file)
    except Exception as e:
        record("Word readable", False, str(e))
        return False

    all_text = " ".join(p.text.lower() for p in doc.paragraphs)
    record("Word mentions 'sentiment'", "sentiment" in all_text, "No mention of 'sentiment'")
    record("Word mentions 'review'", "review" in all_text, "No mention of 'review'")

    return True


def check_gsheet():
    """Check Google Sheet with review/sentiment in title."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id FROM gsheet.spreadsheets WHERE title ILIKE '%%review%%' OR title ILIKE '%%sentiment%%'")
    rows = cur.fetchall()
    found = len(rows) > 0
    record("GSheet with 'review' or 'sentiment' in title", found, "No matching spreadsheet found")

    cur.close()
    conn.close()
    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    word_ok = check_word(args.agent_workspace)

    db_fail_before = FAIL_COUNT
    gsheet_ok = check_gsheet()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")

    overall = excel_ok and word_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

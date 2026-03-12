"""Evaluation for canvas-curriculum-audit."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_compliance():
    """Get expected compliance data from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name,
               (SELECT COUNT(*) FROM canvas.assignments a WHERE a.course_id = c.id) as asgn,
               (SELECT COUNT(*) FROM canvas.quizzes q WHERE q.course_id = c.id) as quiz,
               (SELECT COUNT(*) FROM canvas.modules m WHERE m.course_id = c.id) as mods,
               CASE WHEN c.syllabus_body IS NOT NULL AND c.syllabus_body != '' THEN true ELSE false END as has_syl
        FROM canvas.courses c ORDER BY c.name
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    result = {}
    for name, asgn, quiz, mods, syl in rows:
        compliant = asgn >= 8 and quiz >= 3 and mods >= 4 and syl
        result[name.lower()] = {
            "assignments": asgn, "quizzes": quiz, "modules": mods,
            "syllabus": syl, "compliant": compliant,
        }
    return result


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Curriculum_Audit.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Curriculum_Audit.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Curriculum_Audit.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_compliance()

    # Check Compliance Matrix sheet
    cm_rows = load_sheet_rows(wb, "Compliance Matrix")
    if cm_rows is None:
        check("Sheet 'Compliance Matrix' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Compliance Matrix' exists", True)
        data_rows = cm_rows[1:] if len(cm_rows) > 1 else []
        check("Compliance Matrix has 22 rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        # Check header
        header = cm_rows[0] if cm_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "assignments_count", "quizzes_count", "modules_count", "has_syllabus", "compliant_yn"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        # Spot-check some courses
        for row in data_rows:
            if not row or not row[0]:
                continue
            course_key = str(row[0]).strip().lower()
            if course_key in expected:
                exp = expected[course_key]
                if "applied analytics" in course_key and "fall 2013" in course_key:
                    # AAA Fall 2013: 6 assignments, 0 quizzes, 3 modules
                    check("AAA-2013 assignments=6", num_close(row[1], 6), f"Got {row[1]}")
                    check("AAA-2013 quizzes=0", num_close(row[2], 0), f"Got {row[2]}")
                    check("AAA-2013 compliant=No",
                          str(row[5]).strip().lower() in ("no", "n", "false"),
                          f"Got {row[5]}")
                elif "foundations of finance" in course_key and "fall 2013" in course_key:
                    check("FFF-2013 assignments=13", num_close(row[1], 13), f"Got {row[1]}")
                    check("FFF-2013 quizzes=7", num_close(row[2], 7), f"Got {row[2]}")

    # Check Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        check("Total_Courses = 22", num_close(lookup.get("total_courses"), 22),
              f"Got {lookup.get('total_courses')}")

        compliant_expected = sum(1 for v in expected.values() if v["compliant"])
        check(f"Compliant_Courses = {compliant_expected}",
              num_close(lookup.get("compliant_courses"), compliant_expected),
              f"Got {lookup.get('compliant_courses')}")


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Audit_Report.docx")
    if not os.path.isfile(docx_path):
        check("Audit_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Audit_Report.docx exists", True)
    check("Word doc has content (> 1KB)", os.path.getsize(docx_path) > 1000,
          f"Size: {os.path.getsize(docx_path)}")

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Report mentions compliance", "compli" in all_text, f"Sample: {all_text[:200]}")
        check("Report mentions recommendations", "recommend" in all_text, f"Sample: {all_text[:200]}")
    except ImportError:
        check("python-docx available", False, "pip install python-docx")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

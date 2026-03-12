"""
Evaluation for wc-sales-category-excel-email task.

Checks:
1. Excel file Category_Performance.xlsx with correct data
2. Email sent to sales-team@shop.com mentioning highest/lowest revenue categories
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

FILE_PASS = 0
FILE_FAIL = 0
DB_PASS = 0
DB_FAIL = 0


def check(name, condition, detail="", db=False):
    global FILE_PASS, FILE_FAIL, DB_PASS, DB_FAIL
    if condition:
        if db:
            DB_PASS += 1
        else:
            FILE_PASS += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            DB_FAIL += 1
        else:
            FILE_FAIL += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel File ===")
    agent_file = os.path.join(agent_workspace, "Category_Performance.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Category_Performance.xlsx")

    if not os.path.exists(agent_file):
        check("Excel file exists", False, f"Not found: {agent_file}")
        return
    check("Excel file exists", True)

    if not os.path.exists(gt_file):
        check("Groundtruth file exists", False, f"Not found: {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Category Summary sheet
    print("  Checking Category Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Category Summary")
    g_rows = load_sheet_rows(gt_wb, "Category Summary")

    if a_rows is None:
        check("Sheet 'Category Summary' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth sheet exists", False, "Not found")
    else:
        check("Sheet 'Category Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Category row count matches", len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Row '{g_row[0]}'", False, "Missing")
                continue

            # Product_Count (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"{key}.Product_Count",
                      num_close(a_row[1], g_row[1], 2),
                      f"{a_row[1]} vs {g_row[1]}")

            # Total_Units_Sold (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Total_Units_Sold",
                      num_close(a_row[2], g_row[2], 50),
                      f"{a_row[2]} vs {g_row[2]}")

            # Total_Revenue (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Total_Revenue",
                      num_close(a_row[3], g_row[3], 500),
                      f"{a_row[3]} vs {g_row[3]}")

            # Avg_Price (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"{key}.Avg_Price",
                      num_close(a_row[4], g_row[4], 5.0),
                      f"{a_row[4]} vs {g_row[4]}")

    # Check Top Products sheet
    print("  Checking Top Products sheet...")
    a_rows = load_sheet_rows(agent_wb, "Top Products")
    g_rows = load_sheet_rows(gt_wb, "Top Products")

    if a_rows is None:
        check("Sheet 'Top Products' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth Top Products sheet exists", False, "Not found")
    else:
        check("Sheet 'Top Products' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Top Products row count", len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        # Check that top product per category matches
        g_by_cat = {}
        for row in g_data:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat not in g_by_cat:
                    g_by_cat[cat] = row

        a_by_cat = {}
        for row in a_data:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat not in a_by_cat:
                    a_by_cat[cat] = row

        for cat, g_row in g_by_cat.items():
            a_row = a_by_cat.get(cat)
            if a_row is None:
                check(f"Top product for '{cat}'", False, "Category missing")
            else:
                check(f"Top product for '{cat}' matches",
                      str(a_row[1] or "").strip()[:30].lower() == str(g_row[1] or "").strip()[:30].lower(),
                      f"Got '{str(a_row[1])[:50]}' vs '{str(g_row[1])[:50]}'")


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for email check", False, str(e), db=True)
        return

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    # Find email to sales-team@shop.com
    target = "sales-team@shop.com"
    found = None
    for subj, from_addr, to_addr, body in all_emails:
        to_str = str(to_addr or "").lower()
        if target in to_str:
            found = (subj, from_addr, to_addr, body)
            break

    check(f"Email sent to {target}", found is not None,
          f"Found {len(all_emails)} total emails", db=True)

    if found:
        subj, from_addr, to_addr, body = found
        body_lower = (body or "").lower()
        subj_lower = (subj or "").lower()

        check("Email subject mentions category/performance/report",
              "category" in subj_lower or "performance" in subj_lower or "report" in subj_lower,
              f"Subject: {(subj or '')[:100]}", db=True)

        check("Email body mentions highest revenue category 'Headphones'",
              "headphones" in body_lower,
              "Headphones not found in body", db=True)

        check("Email body mentions lowest category 'TV & Home Theater' or similar",
              "tv" in body_lower and "theater" in body_lower or "tv & home theater" in body_lower,
              "TV & Home Theater not found in body", db=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_email()

    total_pass = FILE_PASS + DB_PASS
    total_fail = FILE_FAIL + DB_FAIL
    file_ok = FILE_FAIL == 0

    print(f"\n=== SUMMARY ===")
    print(f"  File checks - Passed: {FILE_PASS}, Failed: {FILE_FAIL}")
    print(f"  DB checks   - Passed: {DB_PASS}, Failed: {DB_FAIL}")
    if DB_FAIL > 0:
        print(f"  WARNING: {DB_FAIL} DB checks failed (not blocking)")
    print(f"  Overall: {'PASS' if file_ok else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": total_pass, "failed": total_fail, "success": file_ok}, f, indent=2)

    sys.exit(0 if file_ok else 1)


if __name__ == "__main__":
    main()

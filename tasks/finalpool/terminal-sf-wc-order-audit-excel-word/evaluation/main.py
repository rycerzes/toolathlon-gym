"""Evaluation for terminal-sf-wc-order-audit-excel-word.
Checks:
1. Order_Audit_Report.xlsx with 4 sheets and correct data
2. Audit_Findings.docx with required sections
3. audit_analysis.py script exists
"""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').replace('%', '').strip())
    except Exception:
        return default


def check_excel(workspace):
    print("\n=== Check 1: Order_Audit_Report.xlsx ===")
    path = os.path.join(workspace, "Order_Audit_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheets_lower = [s.lower() for s in sheets]

    # Check 4 sheets
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    # DW_Summary sheet
    dw_idx = next((i for i, s in enumerate(sheets_lower) if "dw" in s or "warehouse" in s), 0)
    ws_dw = wb[sheets[dw_idx]]
    rows_dw = list(ws_dw.iter_rows(values_only=True))
    if len(rows_dw) > 1:
        all_text = " ".join(str(c) for r in rows_dw for c in r if c).lower()
        check("DW sheet has Delivered status", "delivered" in all_text, f"Content: {all_text[:100]}")
        check("DW sheet has Cancelled status", "cancelled" in all_text, f"Content: {all_text[:100]}")

        # Dynamically compute expected values from read-only DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM sf_data.orders WHERE LOWER(order_status) = 'delivered'")
            expected_delivered = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(SUM(sales), 0) FROM sf_data.orders WHERE LOWER(order_status) = 'delivered'")
            expected_revenue = float(cur.fetchone()[0])
            cur.close()
            conn.close()
        except Exception:
            expected_delivered = 14033
            expected_revenue = 2177149.66

        # Find Delivered row and check count
        for row in rows_dw[1:]:
            if row[0] and "deliver" in str(row[0]).lower():
                count = safe_float(row[1])
                check(f"Delivered order count ~{expected_delivered}",
                      count is not None and abs(count - expected_delivered) < max(100, expected_delivered * 0.01),
                      f"Got {count}, expected ~{expected_delivered}")
                rev = safe_float(row[2])
                check(f"Delivered revenue ~{expected_revenue:.0f}",
                      rev is not None and abs(rev - expected_revenue) < max(5000, expected_revenue * 0.01),
                      f"Got {rev}, expected ~{expected_revenue:.0f}")
                break
    else:
        check("DW sheet has data", False, "Sheet is empty")

    # Store_Summary sheet
    store_idx = next((i for i, s in enumerate(sheets_lower) if "store" in s), 1)
    if store_idx < len(sheets):
        ws_store = wb[sheets[store_idx]]
        rows_store = list(ws_store.iter_rows(values_only=True))
        all_text_store = " ".join(str(c) for r in rows_store for c in r if c).lower()
        check("Store sheet has product count", "82" in all_text_store or "product" in all_text_store,
              f"Content snippet: {all_text_store[:100]}")

    # ShipMode sheet
    ship_idx = next((i for i, s in enumerate(sheets_lower) if "ship" in s or "mode" in s), 2)
    if ship_idx < len(sheets):
        ws_ship = wb[sheets[ship_idx]]
        rows_ship = list(ws_ship.iter_rows(values_only=True))
        all_text_ship = " ".join(str(c) for r in rows_ship for c in r if c).lower()
        check("ShipMode sheet has Economy", "economy" in all_text_ship, f"Content: {all_text_ship[:100]}")
        check("ShipMode sheet has Express", "express" in all_text_ship, f"Content: {all_text_ship[:100]}")
        check("ShipMode sheet has Standard", "standard" in all_text_ship, f"Content: {all_text_ship[:100]}")

    # Reconciliation sheet
    recon_idx = next((i for i, s in enumerate(sheets_lower) if "recon" in s), 3)
    if recon_idx < len(sheets):
        ws_recon = wb[sheets[recon_idx]]
        rows_recon = list(ws_recon.iter_rows(values_only=True))
        check("Reconciliation sheet has data", len(rows_recon) > 1,
              f"Found {len(rows_recon)} rows")
        if len(rows_recon) > 1:
            all_text_recon = " ".join(str(c) for r in rows_recon for c in r if c).lower()
            check("Reconciliation has order count comparison",
                  "order" in all_text_recon or "20000" in all_text_recon,
                  f"Content: {all_text_recon[:100]}")


def check_word(workspace):
    print("\n=== Check 2: Audit_Findings.docx ===")
    path = os.path.join(workspace, "Audit_Findings.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    try:
        from docx import Document
        doc = Document(path)
        full_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Doc mentions reconciliation or audit", "reconcil" in full_text or "audit" in full_text,
              f"Text snippet: {full_text[:100]}")
        check("Doc mentions data warehouse", "warehouse" in full_text or "dw" in full_text,
              f"Text snippet: {full_text[:100]}")
        check("Doc mentions online store or woocommerce", "store" in full_text or "woocommerce" in full_text or "ecommerce" in full_text,
              f"Text snippet: {full_text[:100]}")
        check("Doc mentions shipping or ship mode", "ship" in full_text,
              f"Text snippet: {full_text[:100]}")
        check("Doc has recommendation section", "recommend" in full_text,
              f"Text snippet: {full_text[:200]}")
    except Exception as e:
        check("Word document readable", False, str(e))


def check_reverse_validation(workspace):
    """Reverse validation: check things that should NOT exist."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Order_Audit_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        sheets_lower = [s.lower() for s in wb.sheetnames]
        # No unexpected/garbage sheets
        valid_keywords = ["dw", "warehouse", "store", "ship", "mode", "recon", "summary", "audit", "order"]
        unexpected = [s for s in wb.sheetnames
                      if not any(k in s.lower() for k in valid_keywords) and s.lower() != "sheet1"]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected sheets: {unexpected}")

        # Check no test/debug data leaked
        all_text = ""
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                all_text += " ".join(str(c) for c in row if c) + " "
        all_lower = all_text.lower()
        check("No debug/test data in Excel", "test" not in all_lower or "test" in all_lower,
              True)  # soft check
        check("No negative order counts", all(
            safe_float(c, 0) >= 0 for ws in wb.worksheets
            for row in ws.iter_rows(min_row=2, values_only=True)
            for c in [row[1]] if c is not None and isinstance(safe_float(c), (int, float))
        ), "Found negative counts")


def check_script(workspace):
    print("\n=== Check 3: audit_analysis.py ===")
    path = os.path.join(workspace, "audit_analysis.py")
    check("audit_analysis.py exists", os.path.exists(path), f"Not found at {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

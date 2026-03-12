"""
Evaluation for arxiv-method-benchmark-tracker task.
Checks Excel and Notion output.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

LEADERBOARD = [
    {"task": "Image Classification", "method": "ViT-Large", "score": 91.2, "paper_id": "2402.10001"},
    {"task": "Image Classification", "method": "ConvNeXt-XL", "score": 89.5, "paper_id": "2402.10002"},
    {"task": "Text Generation", "method": "GPT-4", "score": 95.0, "paper_id": ""},
    {"task": "Text Generation", "method": "LLaMA-3", "score": 92.3, "paper_id": "2402.10003"},
    {"task": "Image Generation", "method": "DiffusionXL", "score": 2.1, "paper_id": "2402.10004"},
]

PAPERS_WITH_ID = {"2402.10001", "2402.10002", "2402.10003", "2402.10004"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Method_Benchmark.xlsx")
    if not os.path.isfile(path):
        record("Excel file exists", False, f"Not found: {path}")
        return False
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Leaderboard sheet
    lb_rows = load_sheet_rows(wb, "Leaderboard")
    if lb_rows is None:
        record("Sheet 'Leaderboard' exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Sheet 'Leaderboard' exists", True)

    header = lb_rows[0] if lb_rows else []
    data = lb_rows[1:]
    record("Leaderboard has 5 rows", len(data) == 5, f"Found {len(data)}")

    method_col = find_col(header, ["Method", "method"])
    score_col = find_col(header, ["Score", "score", "Accuracy"])
    task_col = find_col(header, ["Task", "task"])

    if method_col is not None:
        found_methods = {str(r[method_col]).strip().lower() for r in data if method_col < len(r) and r[method_col]}
        for entry in LEADERBOARD:
            present = entry["method"].lower() in found_methods
            record(f"Method '{entry['method']}' present", present, f"Found: {found_methods}")

    # Method Details sheet
    md_rows = load_sheet_rows(wb, "Method Details") or load_sheet_rows(wb, "Method_Details")
    if md_rows is None:
        record("Sheet 'Method Details' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Method Details' exists", True)
        data2 = md_rows[1:]
        record("Method Details has 4 rows", len(data2) == 4, f"Found {len(data2)}")

        id_col = find_col(md_rows[0], ["Paper_ID", "Paper ID", "paper_id"])
        if id_col is not None:
            found_ids = {str(r[id_col]).strip() for r in data2 if id_col < len(r) and r[id_col]}
            for pid in PAPERS_WITH_ID:
                record(f"Paper {pid} in Method Details", pid in found_ids, f"Found: {found_ids}")

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        tm_key = next((k for k in metrics if "total" in k and "method" in k), None)
        if tm_key:
            record("Total_Methods = 5", num_close(metrics[tm_key], 5, tol=0), f"Got {metrics[tm_key]}")

        mwp_key = next((k for k in metrics if "with" in k and "paper" in k), None)
        if mwp_key:
            record("Methods_With_Papers = 4", num_close(metrics[mwp_key], 4, tol=0), f"Got {metrics[mwp_key]}")

        tt_key = next((k for k in metrics if "total" in k and "task" in k), None)
        if tt_key:
            record("Total_Tasks = 3", num_close(metrics[tt_key], 3, tol=0), f"Got {metrics[tt_key]}")

        ts_key = next((k for k in metrics if "top" in k and "score" in k), None)
        if ts_key:
            record("Top_Score = 95.0", num_close(metrics[ts_key], 95.0, tol=1), f"Got {metrics[ts_key]}")

    return True


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE '%%benchmark%%'
               OR properties::text ILIKE '%%method%%tracker%%'
               OR properties::text ILIKE '%%leaderboard%%'
        """)
        pages = cur.fetchall()

        if not pages:
            cur.execute("SELECT id, properties FROM notion.pages")
            all_pages = cur.fetchall()
            record("Notion page with benchmark/method content", False,
                   f"Found {len(all_pages)} pages but none matching")
            return False

        record("Notion page exists", True)

        page_ids = [p[0] for p in pages]
        cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = ANY(%s)", (page_ids,))
        count = cur.fetchone()[0]
        if count == 0:
            cur.execute("SELECT COUNT(*) FROM notion.blocks")
            count = cur.fetchone()[0]

        record("Notion page has content blocks", count >= 3, f"Found {count}")

        cur.execute("SELECT block_data FROM notion.blocks")
        blocks = cur.fetchall()
        text = " ".join(str(b[0]).lower() for b in blocks if b[0])
        props = " ".join(str(p[1]).lower() for p in pages if p[1])
        combined = text + " " + props

        has_content = any(kw in combined for kw in ["vit", "convnext", "llama", "diffusion", "benchmark", "leaderboard"])
        record("Notion mentions methods/benchmarks", has_content)

        cur.close()
        conn.close()
        return True
    except Exception as e:
        record("Notion accessible", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

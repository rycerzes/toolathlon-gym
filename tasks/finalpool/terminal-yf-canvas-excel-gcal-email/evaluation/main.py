"""Evaluation for terminal-yf-canvas-excel-gcal-email.

Checks:
1. Financial_Literacy_Workshops.xlsx with 3 sheets (Student_Tiers, Market_Events, Workshop_Schedule)
2. Google Calendar events for 3 workshops
3. Emails sent to finance_students and department_head
4. workshop_materials.txt exists
5. market_events.json exists
6. categorize_students.py and find_market_events.py scripts exist
"""
import argparse
import json
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_expected_tiers():
    """Query Canvas DB for expected tier counts."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT sub.course_id,
              SUM(CASE WHEN avg_pct < 60 THEN 1 ELSE 0 END) as needs_support,
              SUM(CASE WHEN avg_pct >= 60 AND avg_pct < 75 THEN 1 ELSE 0 END) as developing,
              SUM(CASE WHEN avg_pct >= 75 THEN 1 ELSE 0 END) as proficient,
              COUNT(*) as total
            FROM (
              SELECT a.course_id, s.user_id,
                AVG(CASE WHEN a.points_possible > 0 THEN s.score / a.points_possible * 100 ELSE NULL END) as avg_pct
              FROM canvas.submissions s
              JOIN canvas.assignments a ON s.assignment_id = a.id
              WHERE a.course_id IN (16, 17) AND s.score IS NOT NULL AND a.points_possible > 0
              GROUP BY a.course_id, s.user_id
            ) sub
            GROUP BY sub.course_id
            ORDER BY sub.course_id
        """)
        result = {}
        for row in cur.fetchall():
            result[int(row[0])] = {
                'needs_support': int(row[1]),
                'developing': int(row[2]),
                'proficient': int(row[3]),
                'total': int(row[4])
            }
        cur.close()
        conn.close()
        return result
    except Exception as e:
        print(f"  [WARN] Could not query Canvas: {e}")
        return {}


def get_expected_market_events():
    """Query YF DB for expected top 3 market events."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT symbol, date, open, close,
              ROUND(((close - open) / open * 100)::numeric, 2) as change_pct
            FROM yf.stock_prices
            WHERE symbol IN ('GOOGL', 'AMZN', 'JPM')
              AND date >= (SELECT MAX(date) - INTERVAL '30 days' FROM yf.stock_prices WHERE symbol = 'GOOGL')
            ORDER BY ABS((close - open) / open) DESC
            LIMIT 3
        """)
        events = []
        for row in cur.fetchall():
            events.append({
                'symbol': row[0],
                'date': str(row[1]),
                'change_pct': float(row[4])
            })
        cur.close()
        conn.close()
        return events
    except Exception as e:
        print(f"  [WARN] Could not query YF: {e}")
        return []


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Financial_Literacy_Workshops.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Financial_Literacy_Workshops.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Financial_Literacy_Workshops.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    check("Has 3 sheets", len(agent_wb.sheetnames) >= 3, f"Got {agent_wb.sheetnames}")

    expected_tiers = get_expected_tiers()
    expected_events = get_expected_market_events()

    # Student_Tiers sheet
    print("  Checking Student_Tiers...")
    st_sheet = get_sheet(agent_wb, "Student_Tiers")
    check("Sheet 'Student_Tiers' exists", st_sheet is not None, f"Sheets: {agent_wb.sheetnames}")
    if st_sheet:
        rows = list(st_sheet.iter_rows(min_row=2, values_only=True))
        check("Student_Tiers has 2 rows", len(rows) == 2, f"Got {len(rows)}")
        for row in rows:
            if not row or row[0] is None:
                continue
            cid = int(row[0])
            if cid in expected_tiers:
                exp = expected_tiers[cid]
                check(f"Course {cid} Needs_Support",
                      num_close(row[2], exp['needs_support'], 1),
                      f"Expected {exp['needs_support']}, got {row[2]}")
                check(f"Course {cid} Developing",
                      num_close(row[3], exp['developing'], 1),
                      f"Expected {exp['developing']}, got {row[3]}")
                check(f"Course {cid} Proficient",
                      num_close(row[4], exp['proficient'], 10),
                      f"Expected {exp['proficient']}, got {row[4]}")
                check(f"Course {cid} Total",
                      num_close(row[5], exp['total'], 10),
                      f"Expected {exp['total']}, got {row[5]}")

    # Market_Events sheet
    print("  Checking Market_Events...")
    me_sheet = get_sheet(agent_wb, "Market_Events")
    check("Sheet 'Market_Events' exists", me_sheet is not None, f"Sheets: {agent_wb.sheetnames}")
    if me_sheet:
        rows = list(me_sheet.iter_rows(min_row=2, values_only=True))
        check("Market_Events has 3 rows", len(rows) == 3, f"Got {len(rows)}")
        if expected_events and rows:
            # Check top event matches
            for i, exp_event in enumerate(expected_events):
                if i < len(rows) and rows[i]:
                    row = rows[i]
                    # Check symbol
                    row_sym = str(row[1]).strip().upper() if row[1] else ""
                    check(f"Event {i+1} symbol is {exp_event['symbol']}",
                          row_sym == exp_event['symbol'],
                          f"Got {row_sym}")
                    # Check change_pct
                    if row[2] is not None:
                        check(f"Event {i+1} change_pct",
                              num_close(row[2], exp_event['change_pct'], 0.5),
                              f"Expected {exp_event['change_pct']}, got {row[2]}")

    # Workshop_Schedule sheet
    print("  Checking Workshop_Schedule...")
    ws_sheet = get_sheet(agent_wb, "Workshop_Schedule")
    check("Sheet 'Workshop_Schedule' exists", ws_sheet is not None, f"Sheets: {agent_wb.sheetnames}")
    if ws_sheet:
        rows = list(ws_sheet.iter_rows(min_row=2, values_only=True))
        check("Workshop_Schedule has 3 rows", len(rows) == 3, f"Got {len(rows)}")

        topics_found = set()
        for row in rows:
            if row and row[1]:
                topics_found.add(str(row[1]).strip().lower())

        check("Has 'Intro to Markets' workshop",
              any("intro" in t and "market" in t for t in topics_found),
              f"Topics: {topics_found}")
        check("Has 'Portfolio Basics' workshop",
              any("portfolio" in t and "basic" in t for t in topics_found),
              f"Topics: {topics_found}")
        check("Has 'Risk Management' workshop",
              any("risk" in t and "manage" in t for t in topics_found),
              f"Topics: {topics_found}")

        # Check expected attendance for Needs Support tier
        if expected_tiers and rows:
            ns_total = sum(t.get('needs_support', 0) for t in expected_tiers.values())
            dev_total = sum(t.get('developing', 0) for t in expected_tiers.values())
            prof_total = sum(t.get('proficient', 0) for t in expected_tiers.values())
            for row in rows:
                if row and row[2]:
                    tier = str(row[2]).strip().lower()
                    if "needs" in tier or "support" in tier:
                        check("Needs Support attendance",
                              num_close(row[3], ns_total, 2),
                              f"Expected {ns_total}, got {row[3]}")
                    elif "develop" in tier:
                        check("Developing attendance",
                              num_close(row[3], dev_total, 2),
                              f"Expected {dev_total}, got {row[3]}")
                    elif "proficient" in tier:
                        check("Proficient attendance",
                              num_close(row[3], prof_total, 50),
                              f"Expected {prof_total}, got {row[3]}")


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE lower(summary) LIKE '%%intro%%market%%'
               OR lower(summary) LIKE '%%portfolio%%basic%%'
               OR lower(summary) LIKE '%%risk%%manage%%'
            ORDER BY start_datetime
        """)
        workshops = cur.fetchall()
        check("3 workshop calendar events created", len(workshops) >= 3,
              f"Found {len(workshops)} workshop events")

        if workshops:
            topics = [w[0].lower() for w in workshops]
            check("Calendar has Intro to Markets",
                  any("intro" in t and "market" in t for t in topics))
            check("Calendar has Portfolio Basics",
                  any("portfolio" in t and "basic" in t for t in topics))
            check("Calendar has Risk Management",
                  any("risk" in t and "manage" in t for t in topics))

            # Check workshops are on weekdays
            for w in workshops:
                if w[2]:
                    dt = w[2]
                    check(f"'{w[0]}' on weekday", dt.weekday() < 5,
                          f"Day: {dt.strftime('%A')}")

            # Check no two workshops on same day
            dates = set()
            for w in workshops:
                if w[2]:
                    d = w[2].date()
                    check(f"'{w[0]}' unique date", d not in dates,
                          f"Duplicate: {d}")
                    dates.add(d)

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check student announcement email
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE lower(subject) LIKE '%%workshop%%announcement%%'
               OR lower(subject) LIKE '%%financial literacy%%workshop%%'
        """)
        student_emails = cur.fetchall()
        check("Workshop announcement email sent", len(student_emails) > 0,
              f"Found {len(student_emails)}")
        if student_emails:
            to_str = str(student_emails[0][1]).lower()
            check("Announcement to finance_students",
                  "finance_students" in to_str,
                  f"To: {student_emails[0][1]}")
            body = (student_emails[0][2] or "").lower()
            check("Announcement mentions workshops",
                  "intro" in body or "portfolio" in body or "risk" in body or "workshop" in body,
                  f"Body length: {len(body)}")

        # Check department head summary email
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE lower(subject) LIKE '%%workshop%%planning%%summary%%'
               OR lower(subject) LIKE '%%workshop%%summary%%'
        """)
        head_emails = cur.fetchall()
        check("Department head summary email sent", len(head_emails) > 0,
              f"Found {len(head_emails)}")
        if head_emails:
            to_str = str(head_emails[0][1]).lower()
            check("Summary to department_head",
                  "department_head" in to_str,
                  f"To: {head_emails[0][1]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_scripts_and_outputs(agent_workspace):
    print("\n=== Checking Scripts and Output Files ===")

    # Check scripts exist
    check("categorize_students.py exists",
          os.path.isfile(os.path.join(agent_workspace, "categorize_students.py")),
          agent_workspace)
    check("find_market_events.py exists",
          os.path.isfile(os.path.join(agent_workspace, "find_market_events.py")),
          agent_workspace)
    check("generate_outline.py exists",
          os.path.isfile(os.path.join(agent_workspace, "generate_outline.py")),
          agent_workspace)

    # Check market_events.json
    mej = os.path.join(agent_workspace, "market_events.json")
    check("market_events.json exists", os.path.isfile(mej), agent_workspace)
    if os.path.isfile(mej):
        try:
            with open(mej) as f:
                events = json.load(f)
            if isinstance(events, list):
                check("market_events.json has 3 events", len(events) >= 3,
                      f"Got {len(events)}")
            elif isinstance(events, dict) and "events" in events:
                check("market_events.json has 3 events", len(events["events"]) >= 3,
                      f"Got {len(events['events'])}")
            else:
                check("market_events.json is valid list/dict", False, f"Type: {type(events)}")
        except Exception as e:
            check("market_events.json parseable", False, str(e))

    # Check workshop_materials.txt
    wmt = os.path.join(agent_workspace, "workshop_materials.txt")
    check("workshop_materials.txt exists", os.path.isfile(wmt), agent_workspace)
    if os.path.isfile(wmt):
        with open(wmt) as f:
            content = f.read().lower()
        check("workshop_materials.txt has content", len(content) > 200,
              f"Length: {len(content)}")
        check("Materials mentions Intro to Markets",
              "intro" in content and "market" in content)
        check("Materials mentions Portfolio",
              "portfolio" in content)
        check("Materials mentions Risk",
              "risk" in content)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_calendar()
    check_emails()
    check_scripts_and_outputs(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()

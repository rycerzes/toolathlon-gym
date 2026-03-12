"""
Evaluation for howtocook-event-menu-planner task.
Checks Excel and email.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Event_Menu.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Menu Plan
    mp_rows = load_sheet_rows(wb, "Menu Plan") or load_sheet_rows(wb, "Menu_Plan")
    if mp_rows is None:
        record("Sheet 'Menu Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Menu Plan' exists", True)
        data = [r for r in mp_rows[1:] if r and r[0] is not None]
        record("Menu Plan has >= 6 rows (2 per course)", len(data) >= 6, f"Found {len(data)}")

        course_col = find_col(mp_rows[0], ["Course", "course"])
        if course_col is not None:
            courses = {str(r[course_col]).strip().lower() for r in data if course_col < len(r) and r[course_col]}
            for c in ["appetizer", "main", "dessert"]:
                record(f"Course '{c}' present", c in courses, f"Found: {courses}")

        diet_col = find_col(mp_rows[0], ["Dietary_Tags", "Dietary Tags", "Tags"])
        record("Dietary_Tags column exists", diet_col is not None, f"Header: {mp_rows[0]}")

    # Ingredient List
    il_rows = load_sheet_rows(wb, "Ingredient List") or load_sheet_rows(wb, "Ingredient_List")
    if il_rows is None:
        record("Sheet 'Ingredient List' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Ingredient List' exists", True)
        data = [r for r in il_rows[1:] if r and r[0] is not None]
        record("Ingredient List has >= 5 items", len(data) >= 5, f"Found {len(data)}")

    # Cost Summary
    cs_rows = load_sheet_rows(wb, "Cost Summary") or load_sheet_rows(wb, "Cost_Summary")
    if cs_rows is None:
        record("Sheet 'Cost Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Cost Summary' exists", True)
        metrics = {}
        for row in cs_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        cpp_key = next((k for k in metrics if "cost" in k and "per" in k and "person" in k), None)
        if cpp_key:
            try:
                cpp = float(metrics[cpp_key])
                record("Cost_Per_Person <= 30", cpp <= 30, f"Got ${cpp}")
            except (TypeError, ValueError):
                record("Cost_Per_Person numeric", False)

        bv_key = next((k for k in metrics if "budget" in k and "var" in k), None)
        if bv_key:
            try:
                bv = float(metrics[bv_key])
                record("Budget_Variance >= 0 (under budget)", bv >= 0, f"Got ${bv}")
            except (TypeError, ValueError):
                pass

    # Dietary Accommodations
    da_rows = load_sheet_rows(wb, "Dietary Accommodations") or load_sheet_rows(wb, "Dietary_Accommodations")
    if da_rows is None:
        record("Sheet 'Dietary Accommodations' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Dietary Accommodations' exists", True)
        data = [r for r in da_rows[1:] if r and r[0] is not None]
        record("Dietary Accommodations has >= 3 rows", len(data) >= 3, f"Found {len(data)}")

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Event_Menu.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()

    return True


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%%menu%%' OR subject ILIKE '%%dinner%%' OR subject ILIKE '%%catering%%'
    """)
    emails = cur.fetchall()

    record("Menu/dinner email sent", len(emails) >= 1, f"Found {len(emails)}")

    if emails:
        e = emails[0]
        to = e[3]
        if isinstance(to, str):
            try:
                to = json.loads(to)
            except Exception:
                pass
        to_str = str(to).lower()
        record("Email to catering@vendor.com", "catering@vendor.com" in to_str, f"To: {to}")

        body = str(e[4]).lower() if e[4] else ""
        record("Email body mentions guests/menu", "guest" in body or "menu" in body or "dinner" in body,
               f"Body preview: {body[:200]}")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

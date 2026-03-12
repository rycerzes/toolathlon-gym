"""
Evaluation for q4-sales-reconciliation task.
Compares agent's Q4_2025_Sales_Report.xlsx and Executive_Summary.docx against groundtruth.
"""
import argparse
import os
import sys

import openpyxl


def load_sheet_data(wb, sheet_name):
    """Load all rows from a sheet as list of lists. Case-insensitive sheet lookup."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_regional_performance(agent_rows, gt_rows):
    """Check Sheet 1: Regional Performance."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Regional Performance' not found"]

    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    # Build lookup by region name (case-insensitive)
    agent_by_region = {}
    for row in agent_data:
        if row[0]:
            agent_by_region[str(row[0]).strip().lower()] = row

    for g_row in gt_data:
        region = str(g_row[0]).strip()
        region_key = region.lower()

        if region_key not in agent_by_region:
            errors.append(f"Missing region: {region}")
            continue

        a_row = agent_by_region[region_key]

        # Column mapping: Region, Target, Actual, Variance, Variance_Pct, Order_Count, Customer_Count
        field_names = ["Target", "Actual", "Variance", "Variance_Pct", "Order_Count", "Customer_Count"]
        tolerances = [1.0, 1.0, 1.0, 0.5, 0.5, 0.5]

        for j, (fname, tol) in enumerate(zip(field_names, tolerances), start=1):
            try:
                a_val = float(a_row[j]) if a_row[j] is not None else None
                g_val = float(g_row[j])
                if a_val is None:
                    errors.append(f"{region}.{fname}: missing value")
                elif abs(a_val - g_val) > tol:
                    errors.append(f"{region}.{fname}: {a_val} vs expected {g_val} (tol={tol})")
            except (ValueError, TypeError):
                errors.append(f"{region}.{fname}: invalid value '{a_row[j]}'")

    return len(errors) == 0, errors


def check_segment_breakdown(agent_rows, gt_rows):
    """Check Sheet 2: Segment Breakdown."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Segment Breakdown' not found"]

    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    # Build lookup by (region, segment) - case insensitive
    agent_lookup = {}
    for row in agent_data:
        if row[0] and row[1]:
            key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
            agent_lookup[key] = row

    for g_row in gt_data:
        region = str(g_row[0]).strip()
        segment = str(g_row[1]).strip()
        key = (region.lower(), segment.lower())

        if key not in agent_lookup:
            errors.append(f"Missing region-segment: {region}/{segment}")
            continue

        a_row = agent_lookup[key]

        # Check Revenue (col 2) with tolerance 5.0
        try:
            a_rev = float(a_row[2]) if a_row[2] is not None else None
            g_rev = float(g_row[2])
            if a_rev is None:
                errors.append(f"{region}/{segment}.Revenue: missing")
            elif abs(a_rev - g_rev) > 5.0:
                errors.append(f"{region}/{segment}.Revenue: {a_rev} vs expected {g_rev}")
        except (ValueError, TypeError):
            errors.append(f"{region}/{segment}.Revenue: invalid '{a_row[2]}'")

        # Check Orders (col 3) with tolerance 2
        try:
            a_ord = float(a_row[3]) if a_row[3] is not None else None
            g_ord = float(g_row[3])
            if a_ord is None:
                errors.append(f"{region}/{segment}.Orders: missing")
            elif abs(a_ord - g_ord) > 2:
                errors.append(f"{region}/{segment}.Orders: {a_ord} vs expected {g_ord}")
        except (ValueError, TypeError):
            errors.append(f"{region}/{segment}.Orders: invalid '{a_row[3]}'")

    return len(errors) == 0, errors


def check_executive_summary(agent_workspace):
    """Check Executive_Summary.docx exists and contains key terms."""
    errors = []
    docx_path = os.path.join(agent_workspace, "Executive_Summary.docx")

    if not os.path.exists(docx_path):
        return False, ["Executive_Summary.docx not found"]

    try:
        from docx import Document
        doc = Document(docx_path)
        full_text = " ".join([p.text for p in doc.paragraphs]).lower()

        if "asia pacific" not in full_text:
            errors.append("Executive summary missing 'Asia Pacific'")

        # Check that some revenue figure is mentioned (total ~291425)
        # Just check the doc isn't empty
        if len(full_text.strip()) < 50:
            errors.append("Executive summary is too short (less than 50 chars)")

    except ImportError:
        # If python-docx not available, just check file exists and has size
        file_size = os.path.getsize(docx_path)
        if file_size < 100:
            errors.append("Executive_Summary.docx is suspiciously small")
    except Exception as e:
        errors.append(f"Error reading Executive_Summary.docx: {e}")

    return len(errors) == 0, errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    agent_excel = os.path.join(args.agent_workspace, "Q4_2025_Sales_Report.xlsx") if args.agent_workspace else None
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_excel = os.path.join(gt_dir, "Q4_2025_Sales_Report.xlsx")

    if not agent_excel or not os.path.exists(agent_excel):
        print(f"FAIL: Agent output file not found: {agent_excel}")
        sys.exit(1)

    if not os.path.exists(gt_excel):
        print(f"FAIL: Groundtruth file not found: {gt_excel}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_excel, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_excel, data_only=True)

    # Check sheet count
    print(f"Agent sheets: {agent_wb.sheetnames}")
    print(f"Groundtruth sheets: {gt_wb.sheetnames}")

    # Check 1: Regional Performance
    print("\n[1/3] Checking Regional Performance ...")
    a_rp = load_sheet_data(agent_wb, "Regional Performance")
    g_rp = load_sheet_data(gt_wb, "Regional Performance")
    rp_ok, rp_errors = check_regional_performance(a_rp, g_rp)
    if rp_ok:
        print("  PASS")
    else:
        for e in rp_errors:
            print(f"  ERROR: {e}")

    # Check 2: Segment Breakdown
    print("[2/3] Checking Segment Breakdown ...")
    a_sb = load_sheet_data(agent_wb, "Segment Breakdown")
    g_sb = load_sheet_data(gt_wb, "Segment Breakdown")
    sb_ok, sb_errors = check_segment_breakdown(a_sb, g_sb)
    if sb_ok:
        print("  PASS")
    else:
        for e in sb_errors:
            print(f"  ERROR: {e}")

    # Check 3: Executive Summary
    print("[3/3] Checking Executive Summary ...")
    es_ok, es_errors = check_executive_summary(args.agent_workspace)
    if es_ok:
        print("  PASS")
    else:
        for e in es_errors:
            print(f"  ERROR: {e}")

    overall = rp_ok and sb_ok and es_ok
    print(f"\n=== RESULT: {'PASS' if overall else 'FAIL'} ===")
    print(f"Regional Performance: {'PASS' if rp_ok else 'FAIL'}")
    print(f"Segment Breakdown: {'PASS' if sb_ok else 'FAIL'}")
    print(f"Executive Summary: {'PASS' if es_ok else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

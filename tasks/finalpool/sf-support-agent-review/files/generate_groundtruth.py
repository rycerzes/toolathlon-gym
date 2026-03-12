"""Generate groundtruth Excel file for sf-support-agent-review task."""
import os
import psycopg2
from openpyxl import Workbook

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(TASK_ROOT, "groundtruth_workspace")

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Agent roster
    cur.execute("""
        SELECT "AGENT_NAME", "TEAM", "SKILL_LEVEL"
        FROM sf_data."SUPPORT_CENTER__PUBLIC__AGENTS"
        ORDER BY "AGENT_NAME"
    """)
    agent_rows = cur.fetchall()

    # Priority analysis with SLA compliance
    cur.execute("""
        SELECT t."PRIORITY",
               COUNT(*) as ticket_count,
               ROUND(AVG(t."RESPONSE_TIME_HOURS")::numeric, 2) as avg_response,
               ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 2) as avg_csat,
               ROUND(100.0 * SUM(CASE WHEN t."RESPONSE_TIME_HOURS" <= s."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END) / COUNT(*)::numeric, 2) as sla_compliance
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" s ON t."PRIORITY" = s."PRIORITY"
        GROUP BY t."PRIORITY"
        ORDER BY t."PRIORITY"
    """)
    priority_rows = cur.fetchall()

    # Issue type analysis
    cur.execute("""
        SELECT "ISSUE_TYPE",
               COUNT(*) as ticket_count,
               ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_response,
               ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as avg_csat
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE"
        ORDER BY COUNT(*) DESC
    """)
    issue_rows = cur.fetchall()

    # Overall summary
    cur.execute("""
        SELECT COUNT(*) as total_tickets,
               ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as overall_avg_response,
               ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as overall_avg_csat
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
    """)
    summary_row = cur.fetchone()
    total_agents = len(agent_rows)

    cur.close()
    conn.close()

    wb = Workbook()

    # Sheet 1: Agent Roster
    ws1 = wb.active
    ws1.title = "Agent Roster"
    ws1.append(["Agent_Name", "Team", "Skill_Level"])
    for row in agent_rows:
        ws1.append(list(row))

    # Sheet 2: Priority Analysis
    ws2 = wb.create_sheet("Priority Analysis")
    ws2.append(["Priority", "Ticket_Count", "Avg_Response_Hours", "Avg_Satisfaction", "SLA_Compliance_Pct"])
    for row in priority_rows:
        ws2.append([row[0], int(row[1]), float(row[2]), float(row[3]), float(row[4])])

    # Sheet 3: Issue Type Analysis
    ws3 = wb.create_sheet("Issue Type Analysis")
    ws3.append(["Issue_Type", "Ticket_Count", "Avg_Response_Hours", "Avg_Satisfaction"])
    for row in issue_rows:
        ws3.append([row[0], int(row[1]), float(row[2]), float(row[3])])

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Metric", "Value"])
    ws4.append(["Total_Agents", total_agents])
    ws4.append(["Total_Tickets", int(summary_row[0])])
    ws4.append(["Overall_Avg_Response_Hours", float(summary_row[1])])
    ws4.append(["Overall_Avg_Satisfaction", float(summary_row[2])])

    out_path = os.path.join(OUTPUT_DIR, "Support_Analysis.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")


if __name__ == "__main__":
    main()

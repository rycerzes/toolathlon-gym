"""Evaluation script for fetch-howtocook-catering-excel-gcal-email."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # --- Check Excel file ---
    excel_path = os.path.join(agent_workspace, "Wellness_Menu_Plan.xlsx")
    check("Wellness_Menu_Plan.xlsx exists", os.path.exists(excel_path))

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # Sheet 1: Daily Menu
        check("Daily Menu sheet exists", "Daily Menu" in wb.sheetnames)
        if "Daily Menu" in wb.sheetnames:
            ws = wb["Daily Menu"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))

            check("Daily Menu has 10 rows", len(data_rows) == 10, f"got {len(data_rows)}")

            for col in ["Day", "Meal_Type", "Recipe_Name", "Servings", "Ingredient_Count",
                        "Estimated_Cost", "Is_Vegetarian"]:
                check(f"Daily Menu has {col} column",
                      col in headers, f"headers: {headers}")

            # Verify days are sorted Monday-Friday
            if len(data_rows) >= 10:
                day_col = headers.index("Day") if "Day" in headers else 0
                days = [str(r[day_col]) for r in data_rows]
                expected_days = ["Monday", "Monday", "Tuesday", "Tuesday", "Wednesday",
                                "Wednesday", "Thursday", "Thursday", "Friday", "Friday"]
                check("Days sorted Monday-Friday", days == expected_days,
                      f"got {days}")

            # Verify servings are all 50
            if "Servings" in headers:
                serv_col = headers.index("Servings")
                servings = [safe_float(r[serv_col]) for r in data_rows]
                check("All servings are 50",
                      all(s == 50 for s in servings if s is not None),
                      f"servings: {servings}")

            # Verify Is_Vegetarian has at least 2 Yes values
            if "Is_Vegetarian" in headers:
                veg_col = headers.index("Is_Vegetarian")
                veg_vals = [str(r[veg_col]).strip() for r in data_rows if r[veg_col]]
                yes_count = sum(1 for v in veg_vals if v.lower() == "yes")
                check("At least 2 vegetarian options", yes_count >= 2,
                      f"found {yes_count} vegetarian")

            # Verify cost formula: Ingredient_Count * 0.50 * 50
            if "Ingredient_Count" in headers and "Estimated_Cost" in headers:
                ic_col = headers.index("Ingredient_Count")
                ec_col = headers.index("Estimated_Cost")
                cost_ok = True
                for row in data_rows:
                    ic = safe_float(row[ic_col])
                    ec = safe_float(row[ec_col])
                    if ic is not None and ec is not None:
                        expected = round(ic * 0.50 * 50, 2)
                        if abs(ec - expected) > 0.01:
                            cost_ok = False
                            break
                check("Estimated_Cost = Ingredient_Count * 0.50 * 50", cost_ok)

        # Sheet 2: Ingredient Summary
        check("Ingredient Summary sheet exists", "Ingredient Summary" in wb.sheetnames)
        if "Ingredient Summary" in wb.sheetnames:
            ws = wb["Ingredient Summary"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))

            check("Ingredient Summary has >= 20 rows", len(data_rows) >= 20,
                  f"got {len(data_rows)}")

            for col in ["Ingredient_Name", "Total_Quantity", "Unit", "Times_Used", "Total_Cost"]:
                check(f"Ingredient Summary has {col} column",
                      col in headers, f"headers: {headers}")

            # Verify sorted alphabetically
            if "Ingredient_Name" in headers:
                name_col = headers.index("Ingredient_Name")
                names = [str(r[name_col]) for r in data_rows if r[name_col]]
                check("Ingredients sorted alphabetically",
                      names == sorted(names), f"first few: {names[:5]}")

            # Verify cost formula: Times_Used * 0.50 * 50
            if "Times_Used" in headers and "Total_Cost" in headers:
                tu_col = headers.index("Times_Used")
                tc_col = headers.index("Total_Cost")
                cost_ok = True
                for row in data_rows:
                    tu = safe_float(row[tu_col])
                    tc = safe_float(row[tc_col])
                    if tu is not None and tc is not None:
                        expected = round(tu * 0.50 * 50, 2)
                        if abs(tc - expected) > 0.01:
                            cost_ok = False
                            break
                check("Ingredient Total_Cost = Times_Used * 0.50 * 50", cost_ok)

        # Sheet 3: Budget Overview
        check("Budget Overview sheet exists", "Budget Overview" in wb.sheetnames)
        if "Budget Overview" in wb.sheetnames:
            ws = wb["Budget Overview"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))

            check("Budget Overview has 6 rows", len(data_rows) == 6,
                  f"got {len(data_rows)}")

            for col in ["Label", "Value"]:
                check(f"Budget Overview has {col} column",
                      col in headers, f"headers: {headers}")

            # Build label->value map
            label_col = headers.index("Label") if "Label" in headers else 0
            value_col = headers.index("Value") if "Value" in headers else 1
            budget_map = {}
            for row in data_rows:
                if row[label_col]:
                    budget_map[str(row[label_col]).strip()] = safe_float(row[value_col])

            check("Has Total_Budget label", "Total_Budget" in budget_map)
            check("Total_Budget is 2000",
                  budget_map.get("Total_Budget") == 2000.0,
                  f"got {budget_map.get('Total_Budget')}")

            # Verify budget consistency
            tc = budget_map.get("Total_Estimated_Cost")
            br = budget_map.get("Budget_Remaining")
            if tc is not None and br is not None:
                check("Budget_Remaining = 2000 - Total_Estimated_Cost",
                      abs(br - (2000.0 - tc)) < 0.01,
                      f"remaining={br}, expected={2000.0 - tc}")

            acd = budget_map.get("Avg_Cost_Per_Day")
            if tc is not None and acd is not None:
                check("Avg_Cost_Per_Day = Total / 5",
                      abs(acd - tc / 5) < 0.01,
                      f"avg={acd}, expected={tc / 5}")

            acm = budget_map.get("Avg_Cost_Per_Meal")
            if tc is not None and acm is not None:
                check("Avg_Cost_Per_Meal = Total / 10",
                      abs(acm - tc / 10) < 0.01,
                      f"avg={acm}, expected={tc / 10}")

    # --- Check Calendar Events ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary LIKE '%Wellness Week Meal Prep%'
            ORDER BY start_datetime
        """)
        cal_rows = cur.fetchall()
        cur.close()
        conn.close()

        check("5 meal prep calendar events", len(cal_rows) == 5,
              f"found {len(cal_rows)}")

        if len(cal_rows) >= 5:
            # Check dates are March 16-20
            dates = [str(r[1])[:10] for r in cal_rows]
            expected_dates = ["2026-03-16", "2026-03-17", "2026-03-18",
                            "2026-03-19", "2026-03-20"]
            check("Calendar events on March 16-20",
                  dates == expected_dates, f"dates: {dates}")

            # Check times are 7:00-8:00
            for row in cal_rows:
                start_h = str(row[1])[11:16] if len(str(row[1])) > 15 else ""
                end_h = str(row[2])[11:16] if len(str(row[2])) > 15 else ""
                if start_h:
                    check(f"Event {str(row[1])[:10]} starts at 07:00",
                          start_h == "07:00", f"got {start_h}")
                    break  # Just check one for brevity
    except Exception as e:
        check("Calendar events query", False, str(e))

    # --- Check Emails ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr, subject, body_text
            FROM email.messages
            WHERE subject LIKE '%Wellness Week%'
            ORDER BY subject
        """)
        email_rows = cur.fetchall()
        cur.close()
        conn.close()

        check("At least 2 wellness emails sent", len(email_rows) >= 2,
              f"found {len(email_rows)}")

        # to_addr is jsonb, convert to string for matching
        recipients = [str(r[0]) for r in email_rows]
        subjects = [r[1] for r in email_rows]

        # Check vendor email
        vendor_found = any("catering_vendor" in str(r) for r in recipients)
        check("Email sent to catering_vendor@company.com", vendor_found,
              f"recipients: {recipients}")

        # Check committee email
        committee_found = any("wellness_committee" in str(r) for r in recipients)
        check("Email sent to wellness_committee@company.com", committee_found,
              f"recipients: {recipients}")

        # Check subjects
        ingredient_subj = any("ingredient" in str(s).lower() for s in subjects)
        check("Vendor email has ingredient-related subject", ingredient_subj,
              f"subjects: {subjects}")

        menu_subj = any("menu" in str(s).lower() for s in subjects)
        check("Committee email has menu-related subject", menu_subj,
              f"subjects: {subjects}")

    except Exception as e:
        check("Email query", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

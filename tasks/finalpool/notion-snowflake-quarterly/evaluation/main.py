"""
Evaluation script for notion-snowflake-quarterly task.

Checks:
1. Excel Q4_Performance.xlsx - sheet "Regional Comparison" with correct data
2. Google Sheets - "Q4 Dashboard" spreadsheet with "Summary" sheet matching Excel data

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Known targets
TARGETS = {
    "Asia Pacific": 80000,
    "Europe": 85000,
    "Latin America": 75000,
    "Middle East": 85000,
    "North America": 80000,
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def pct_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_actuals_from_snowflake():
    """Query actual Q4 2024 revenue by region from the Snowflake proxy tables."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c."REGION",
               ROUND(SUM(o."TOTAL_AMOUNT")::numeric, 2) as total_revenue
        FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
        JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c
          ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
        WHERE o."ORDER_DATE" >= '2024-10-01'
          AND o."ORDER_DATE" <= '2024-12-31'
        GROUP BY c."REGION"
        ORDER BY c."REGION"
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {row[0]: float(row[1]) for row in rows}


def compute_expected_data():
    """Compute expected rows: Region, Target, Actual, Variance, Achievement_Pct."""
    actuals = get_actuals_from_snowflake()
    rows = []
    for region in sorted(TARGETS.keys()):
        target = TARGETS[region]
        actual = actuals.get(region, 0.0)
        variance = round(actual - target, 2)
        achievement = round(actual / target * 100, 1)
        rows.append((region, target, actual, variance, achievement))
    return rows


def load_sheet_rows(wb, sheet_name):
    """Load all rows from a sheet (case-insensitive name lookup)."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_excel(agent_workspace, expected_data):
    """Check Q4_Performance.xlsx content."""
    print("\n=== Checking Excel Output ===")

    xlsx_path = os.path.join(agent_workspace, "Q4_Performance.xlsx")

    if not os.path.isfile(xlsx_path):
        record("Excel file exists", False, f"Not found: {xlsx_path}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    agent_rows = load_sheet_rows(wb, "Regional Comparison")
    if agent_rows is None:
        record("Sheet 'Regional Comparison' exists", False, "Not found")
        return False

    record("Sheet 'Regional Comparison' exists", True)

    # Skip header row
    data_rows = agent_rows[1:] if len(agent_rows) > 1 else []

    if len(data_rows) != 5:
        record("Row count is 5", False, f"Got {len(data_rows)}")
        return False

    record("Row count is 5", True)

    all_ok = True

    # Build lookup by region name
    agent_by_region = {}
    for row in data_rows:
        if row[0]:
            agent_by_region[str(row[0]).strip().lower()] = row

    for region, target, actual, variance, achievement in expected_data:
        region_key = region.lower()
        if region_key not in agent_by_region:
            record(f"Region '{region}' present", False, "Missing")
            all_ok = False
            continue

        a_row = agent_by_region[region_key]
        record(f"Region '{region}' present", True)

        # Check Target
        if not num_close(a_row[1], target, 1.0):
            record(f"{region}.Target", False, f"{a_row[1]} vs expected {target}")
            all_ok = False
        else:
            record(f"{region}.Target", True)

        # Check Actual
        if not num_close(a_row[2], actual, 1.0):
            record(f"{region}.Actual", False, f"{a_row[2]} vs expected {actual}")
            all_ok = False
        else:
            record(f"{region}.Actual", True)

        # Check Variance
        if not num_close(a_row[3], variance, 1.0):
            record(f"{region}.Variance", False, f"{a_row[3]} vs expected {variance}")
            all_ok = False
        else:
            record(f"{region}.Variance", True)

        # Check Achievement_Pct
        if not pct_close(a_row[4], achievement, 0.5):
            record(f"{region}.Achievement_Pct", False, f"{a_row[4]} vs expected {achievement}")
            all_ok = False
        else:
            record(f"{region}.Achievement_Pct", True)

    return all_ok


def check_gsheet(expected_data):
    """Check Google Sheets Q4 Dashboard with Summary sheet."""
    print("\n=== Checking Google Sheet ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Find spreadsheet titled "Q4 Dashboard"
    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE LOWER(title) LIKE '%%q4 dashboard%%'
        LIMIT 1
    """)
    row = cur.fetchone()

    if not row:
        record("Spreadsheet 'Q4 Dashboard' exists", False, "Not found in gsheet.spreadsheets")
        cur.close()
        conn.close()
        return False

    record("Spreadsheet 'Q4 Dashboard' exists", True)
    spreadsheet_id = row[0]

    # Find "Summary" sheet
    cur.execute("""
        SELECT id, title FROM gsheet.sheets
        WHERE spreadsheet_id = %s AND LOWER(title) LIKE '%%summary%%'
        LIMIT 1
    """, (spreadsheet_id,))
    sheet_row = cur.fetchone()

    if not sheet_row:
        record("Sheet 'Summary' exists", False, "Not found")
        cur.close()
        conn.close()
        return False

    record("Sheet 'Summary' exists", True)
    sheet_id = sheet_row[0]

    # Read all cells
    cur.execute("""
        SELECT row_index, col_index, value
        FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index, col_index
    """, (spreadsheet_id, sheet_id))
    cells = cur.fetchall()
    cur.close()
    conn.close()

    if not cells:
        record("Summary sheet has data", False, "No cells found")
        return False

    record("Summary sheet has data", True)

    # Build grid from cells
    grid = {}
    for row_idx, col_idx, value in cells:
        grid[(row_idx, col_idx)] = value

    all_ok = True

    # Check that we have region data rows
    # Find which rows contain region names
    region_cells = {}
    for (r, c), v in grid.items():
        if v and str(v).strip().lower() in [reg.lower() for reg in TARGETS.keys()]:
            region_cells[str(v).strip()] = r

    if len(region_cells) < 5:
        # Try case-insensitive matching
        for (r, c), v in grid.items():
            if v:
                for reg in TARGETS.keys():
                    if str(v).strip().lower() == reg.lower() and reg not in region_cells:
                        region_cells[reg] = r

    for region, target, actual, variance, achievement in expected_data:
        # Find the row for this region
        matched_row = None
        for (r, c), v in grid.items():
            if v and str(v).strip().lower() == region.lower():
                matched_row = r
                break

        if matched_row is None:
            record(f"GSheet: Region '{region}' present", False, "Not found")
            all_ok = False
            continue

        record(f"GSheet: Region '{region}' present", True)

        # Get all values in this row
        row_vals = []
        for ci in range(10):  # check up to 10 columns
            row_vals.append(grid.get((matched_row, ci)))

        # Find numeric values in this row
        numeric_vals = []
        for v in row_vals:
            if v is not None:
                try:
                    numeric_vals.append(float(v))
                except (ValueError, TypeError):
                    pass

        # Check that target value appears somewhere in the row
        target_found = any(num_close(nv, target, 5.0) for nv in numeric_vals)
        record(f"GSheet: {region}.Target", target_found,
               f"Expected ~{target} in {numeric_vals}")
        if not target_found:
            all_ok = False

        # Check that actual value appears
        actual_found = any(num_close(nv, actual, 5.0) for nv in numeric_vals)
        record(f"GSheet: {region}.Actual", actual_found,
               f"Expected ~{actual} in {numeric_vals}")
        if not actual_found:
            all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Compute expected data dynamically from Snowflake
    expected_data = compute_expected_data()
    print("Expected data:")
    for row in expected_data:
        print(f"  {row}")

    excel_ok = check_excel(args.agent_workspace, expected_data)
    gsheet_ok = check_gsheet(expected_data)

    overall = excel_ok and gsheet_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:        {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Google Sheet: {'PASS' if gsheet_ok else 'FAIL'}")
    print(f"  Overall:      {'PASS' if overall else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

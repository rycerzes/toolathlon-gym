"""Evaluation for yf-sector-rotation-excel-notion-email.

Checks:
1. Excel file (Sector_Rotation.xlsx) with 3 sheets, correct structure and values
2. Notion database "Sector Research" with 5 stock entries
3. Two emails sent to investment_team@firm.com and trading_desk@firm.com
"""
import argparse
import os
import sys
import psycopg2
import openpyxl

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

SYMBOLS = ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]

# Tolerances
PRICE_TOL = 0.5
RETURN_TOL = 0.5
RS_TOL = 0.05


def load_sheet_rows(wb, sheet_name):
    """Load rows from a sheet, case-insensitive match."""
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol):
    """Check if two numbers are close within tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison."""
    try:
        return str(a).strip().lower() == str(b).strip().lower()
    except (TypeError, AttributeError):
        return False


def check_excel(agent_ws, gt_ws):
    errors = []
    agent_path = os.path.join(agent_ws, "Sector_Rotation.xlsx")
    gt_path = os.path.join(gt_ws, "Sector_Rotation.xlsx")

    if not os.path.exists(agent_path):
        return ["Sector_Rotation.xlsx not found in agent workspace"]
    if not os.path.exists(gt_path):
        return ["Sector_Rotation.xlsx not found in groundtruth workspace"]

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        return [f"Error loading Excel files: {e}"]

    # --- Sheet 1: Momentum Analysis ---
    agent_rows = load_sheet_rows(wb_agent, "Momentum Analysis")
    gt_rows = load_sheet_rows(wb_gt, "Momentum Analysis")

    if agent_rows is None:
        errors.append("Sheet 'Momentum Analysis' not found")
    elif gt_rows is None:
        errors.append("Groundtruth 'Momentum Analysis' sheet missing")
    else:
        agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
        gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]

        if len(agent_data) < 5:
            errors.append(f"Momentum Analysis has {len(agent_data)} rows, expected 5")
        else:
            agent_by_sym = {str(r[0]).strip().upper(): r for r in agent_data}
            gt_by_sym = {str(r[0]).strip().upper(): r for r in gt_data}

            for sym in SYMBOLS:
                if sym not in agent_by_sym:
                    errors.append(f"{sym} missing from Momentum Analysis")
                    continue
                a = agent_by_sym[sym]
                g = gt_by_sym.get(sym)
                if g is None:
                    continue

                # Check Latest_Price (col index 3)
                if not num_close(a[3], g[3], PRICE_TOL):
                    errors.append(f"{sym}: Latest_Price {a[3]} != {g[3]} (tol {PRICE_TOL})")

                # Check returns (cols 4,5,6)
                for idx, name in [(4, "Return_1M_Pct"), (5, "Return_3M_Pct"), (6, "Return_6M_Pct")]:
                    if not num_close(a[idx], g[idx], RETURN_TOL):
                        errors.append(f"{sym}: {name} {a[idx]} != {g[idx]} (tol {RETURN_TOL})")

                # Check Composite_Momentum (col 7)
                if not num_close(a[7], g[7], RETURN_TOL):
                    errors.append(f"{sym}: Composite_Momentum {a[7]} != {g[7]} (tol {RETURN_TOL})")

                # Check Benchmark_Momentum (col 8)
                if not num_close(a[8], g[8], RETURN_TOL):
                    errors.append(f"{sym}: Benchmark_Momentum {a[8]} != {g[8]} (tol {RETURN_TOL})")

                # Check Signal (col 9)
                if not str_match(a[9], g[9]):
                    errors.append(f"{sym}: Signal '{a[9]}' != '{g[9]}'")

    # --- Sheet 2: Relative Strength ---
    agent_rs = load_sheet_rows(wb_agent, "Relative Strength")
    gt_rs = load_sheet_rows(wb_gt, "Relative Strength")

    if agent_rs is None:
        errors.append("Sheet 'Relative Strength' not found")
    elif gt_rs is None:
        errors.append("Groundtruth 'Relative Strength' sheet missing")
    else:
        agent_data = [r for r in agent_rs[1:] if r and r[0] is not None]
        gt_data = [r for r in gt_rs[1:] if r and r[0] is not None]

        if len(agent_data) < 5:
            errors.append(f"Relative Strength has {len(agent_data)} rows, expected 5")
        else:
            agent_by_sym = {str(r[0]).strip().upper(): r for r in agent_data}
            gt_by_sym = {str(r[0]).strip().upper(): r for r in gt_data}

            for sym in SYMBOLS:
                if sym not in agent_by_sym:
                    errors.append(f"{sym} missing from Relative Strength")
                    continue
                a = agent_by_sym[sym]
                g = gt_by_sym.get(sym)
                if g is None:
                    continue

                for idx, name in [(1, "RS_1M"), (2, "RS_3M"), (3, "RS_6M"), (4, "Avg_RS")]:
                    if not num_close(a[idx], g[idx], RS_TOL):
                        errors.append(f"{sym}: {name} {a[idx]} != {g[idx]} (tol {RS_TOL})")

                # RS_Rank
                if a[5] is not None and g[5] is not None:
                    if int(a[5]) != int(g[5]):
                        errors.append(f"{sym}: RS_Rank {a[5]} != {g[5]}")

    # --- Sheet 3: Strategy Summary ---
    agent_sum = load_sheet_rows(wb_agent, "Strategy Summary")
    gt_sum = load_sheet_rows(wb_gt, "Strategy Summary")

    if agent_sum is None:
        errors.append("Sheet 'Strategy Summary' not found")
    elif gt_sum is None:
        errors.append("Groundtruth 'Strategy Summary' sheet missing")
    else:
        def build_summary_dict(rows):
            d = {}
            for r in rows:
                if r and r[0] is not None:
                    d[str(r[0]).strip()] = r[1]
            return d

        a_sum = build_summary_dict(agent_sum[1:])
        g_sum = build_summary_dict(gt_sum[1:])

        for label in ["Overweight_Count", "Neutral_Count", "Underweight_Count"]:
            if label in g_sum:
                a_val = a_sum.get(label)
                g_val = g_sum[label]
                if a_val is None:
                    errors.append(f"Summary missing: {label}")
                elif int(a_val) != int(g_val):
                    errors.append(f"Summary {label}: {a_val} != {g_val}")

        for label in ["Top_Momentum_Stock", "Bottom_Momentum_Stock", "Portfolio_Signal"]:
            if label in g_sum:
                if not str_match(a_sum.get(label, ""), g_sum[label]):
                    errors.append(f"Summary {label}: '{a_sum.get(label)}' != '{g_sum[label]}'")

        if "Avg_Composite_Momentum" in g_sum:
            if not num_close(a_sum.get("Avg_Composite_Momentum", 0), g_sum["Avg_Composite_Momentum"], RETURN_TOL):
                errors.append(f"Summary Avg_Composite_Momentum: {a_sum.get('Avg_Composite_Momentum')} != {g_sum['Avg_Composite_Momentum']}")

    return errors


def check_notion():
    """Check for Sector Research database with 5 stock entries."""
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check for database with title containing "Sector Research"
        cur.execute("""
            SELECT id FROM notion.databases
            WHERE title::text ILIKE '%sector research%'
        """)
        db_rows = cur.fetchall()
        if not db_rows:
            # Fallback: check pages directly
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE properties::text ILIKE '%sector%'
                  AND (properties::text ILIKE '%momentum%' OR properties::text ILIKE '%signal%')
            """)
            count = cur.fetchone()[0]
            if count < 5:
                errors.append("No Notion database 'Sector Research' found and fewer than 5 sector pages")
        else:
            db_id = db_rows[0][0]
            # Check for pages linked to this database
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent::text ILIKE %s
            """, (f'%{db_id}%',))
            count = cur.fetchone()[0]
            if count < 5:
                errors.append(f"Sector Research database has {count} pages, expected 5")

        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def check_emails():
    """Check for 2 emails: one to investment_team, one to trading_desk."""
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        cur.execute("""
            SELECT subject FROM email.messages
            WHERE to_addr::text ILIKE '%investment_team@firm.com%'
            ORDER BY id DESC LIMIT 1
        """)
        rows = cur.fetchall()
        if not rows:
            errors.append("No email found to investment_team@firm.com")

        cur.execute("""
            SELECT subject FROM email.messages
            WHERE to_addr::text ILIKE '%trading_desk@firm.com%'
            ORDER BY id DESC LIMIT 1
        """)
        rows = cur.fetchall()
        if not rows:
            errors.append("No email found to trading_desk@firm.com")

        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking emails: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion database...")
    errs = check_notion()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking emails...")
    errs = check_emails()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

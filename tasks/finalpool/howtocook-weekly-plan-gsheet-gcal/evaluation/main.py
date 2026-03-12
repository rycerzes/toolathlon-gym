"""
Evaluation for howtocook-weekly-plan-gsheet-gcal task.
Checks: GSheet weekly meal plan, Excel file with 2 sheets, GCal events, email.
"""
import argparse
import os
import sys

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()

        target_ss = None
        for sid, title in spreadsheets:
            if title and "meal plan" in title.lower() and ("march" in title.lower() or "2026" in title.lower()):
                target_ss = sid
                break

        record("GSheet 'Weekly Meal Plan March 9-15 2026' exists",
               target_ss is not None,
               f"Found sheets: {[t for _, t in spreadsheets]}")

        if target_ss is None:
            conn.close()
            return

        cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s", (target_ss,))
        sheets = cur.fetchall()
        record("GSheet has at least one sheet", len(sheets) > 0)

        if not sheets:
            conn.close()
            return

        sheet_id = sheets[0][0]
        cur.execute("""
            SELECT COUNT(DISTINCT row_index) FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s AND row_index > 0
        """, (target_ss, sheet_id))
        data_rows = cur.fetchone()[0]
        record("GSheet has at least 21 meal data rows", data_rows >= 21,
               f"Found {data_rows} data rows")

        # Check meal types are present
        cur.execute("""
            SELECT LOWER(value) FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s
        """, (target_ss, sheet_id))
        cell_values = [row[0] for row in cur.fetchall() if row[0]]
        all_text = " ".join(cell_values)

        has_breakfast = "breakfast" in all_text
        has_lunch = "lunch" in all_text
        has_dinner = "dinner" in all_text
        record("GSheet contains Breakfast entries", has_breakfast)
        record("GSheet contains Lunch entries", has_lunch)
        record("GSheet contains Dinner entries", has_dinner)

        conn.close()
    except Exception as e:
        record("GSheet connection", False, str(e))


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel File ===")
    xl_path = os.path.join(agent_workspace, "Meal_Plan_Summary.xlsx")
    if not os.path.isfile(xl_path):
        record("Excel file Meal_Plan_Summary.xlsx exists", False, f"Not found at: {xl_path}")
        return
    record("Excel file Meal_Plan_Summary.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xl_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names = [s.lower() for s in wb.sheetnames]
    has_daily = any("daily" in s for s in sheet_names)
    has_category = any("category" in s for s in sheet_names)
    record("Excel has 'Daily Plan' sheet", has_daily, f"Found sheets: {wb.sheetnames}")
    record("Excel has 'Category Distribution' sheet", has_category, f"Found sheets: {wb.sheetnames}")

    # Check Daily Plan has 21 rows of data
    daily_sheet = None
    for sname in wb.sheetnames:
        if "daily" in sname.lower():
            daily_sheet = wb[sname]
            break

    if daily_sheet:
        max_row = daily_sheet.max_row
        # Count non-empty rows after header
        data_rows = sum(1 for row in daily_sheet.iter_rows(min_row=2, values_only=True)
                        if any(cell is not None and str(cell).strip() != "" for cell in row))
        record("Excel Daily Plan sheet has at least 21 data rows", data_rows >= 21,
               f"Found {data_rows} data rows")

    # Check Category Distribution has at least 3 categories
    cat_sheet = None
    for sname in wb.sheetnames:
        if "category" in sname.lower():
            cat_sheet = wb[sname]
            break

    if cat_sheet:
        cat_rows = sum(1 for row in cat_sheet.iter_rows(min_row=2, values_only=True)
                       if any(cell is not None and str(cell).strip() != "" for cell in row))
        record("Excel Category Distribution has at least 3 categories", cat_rows >= 3,
               f"Found {cat_rows} category rows")

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Meal_Plan_Summary.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Check prep event on March 8
        cur.execute("""
            SELECT summary FROM gcal.events
            WHERE LOWER(summary) LIKE '%meal prep%' OR LOWER(summary) LIKE '%prep planning%'
        """)
        prep_events = cur.fetchall()
        record("GCal has weekly meal prep planning event", len(prep_events) > 0,
               f"Found: {prep_events}")

        # Check daily reminder events (7 for March 9-15)
        cur.execute("""
            SELECT summary FROM gcal.events
            WHERE LOWER(summary) LIKE '%daily meal reminder%' OR LOWER(summary) LIKE '%meal reminder%'
        """)
        daily_events = cur.fetchall()
        record("GCal has at least 7 daily meal reminder events", len(daily_events) >= 7,
               f"Found {len(daily_events)} daily reminder events")

        # Total events check
        cur.execute("SELECT COUNT(*) FROM gcal.events")
        total_events = cur.fetchone()[0]
        record("GCal has at least 8 events total (1 prep + 7 daily)", total_events >= 8,
               f"Found {total_events} total events")

        conn.close()
    except Exception as e:
        record("GCal connection", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%meal plan%'
        """)
        emails = cur.fetchall()
        record("Email with 'Meal Plan' in subject sent", len(emails) > 0,
               f"Found {len(emails)} matching emails")

        if emails:
            # Check at least one email goes to family@home.com
            target_found = False
            for subject, to_addr in emails:
                to_str = str(to_addr).lower() if to_addr else ""
                if "family@home.com" in to_str:
                    target_found = True
                    break
            record("Email sent to family@home.com", target_found,
                   f"Recipients: {[e[1] for e in emails]}")

        conn.close()
    except Exception as e:
        record("Email connection", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_gsheet()
    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

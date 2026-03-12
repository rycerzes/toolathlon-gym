"""Evaluation for wc-inventory-reorder-forecast."""
import argparse
import json
import os
import sys

import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_data):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Inventory_Forecast.xlsx")
    if not os.path.exists(path):
        return ["Inventory_Forecast.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Stock Analysis sheet
        rows = load_sheet_rows(wb, "Stock Analysis")
        if rows is None:
            errors.append("Sheet 'Stock Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_products"]
            if abs(len(data_rows) - expected) > 5:
                errors.append(f"Stock Analysis has {len(data_rows)} rows, expected ~{expected}")

            # Check some products marked as needing reorder
            reorder_yes = [r for r in data_rows if r and len(r) > 7 and str(r[7]).strip().lower() == "yes"]
            expected_reorder = gt_data["reorder_count"]
            if abs(len(reorder_yes) - expected_reorder) > 3:
                errors.append(f"Products marked Needs_Reorder=Yes: {len(reorder_yes)}, expected ~{expected_reorder}")

        # Check Reorder Schedule sheet
        rows2 = load_sheet_rows(wb, "Reorder Schedule")
        if rows2 is None:
            errors.append("Sheet 'Reorder Schedule' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if abs(len(data_rows2) - gt_data["reorder_count"]) > 3:
                errors.append(f"Reorder Schedule has {len(data_rows2)} rows, expected ~{gt_data['reorder_count']}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal(gt_data):
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime FROM gcal.events
            WHERE summary ILIKE '%reorder%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        expected_events = gt_data["reorder_count"]
        if abs(len(rows) - expected_events) > 3:
            errors.append(f"Found {len(rows)} reorder calendar events, expected ~{expected_events}")

        if len(rows) == 0:
            errors.append("No reorder calendar events found")

    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal(gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

"""Evaluation for yf-options-expiry-monitor."""
import os
import argparse, os, sys
import psycopg2


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Options_Monitor.xlsx")
    if not os.path.exists(path):
        return ["Options_Monitor.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Position Analysis
        rows = load_sheet_rows(wb, "Position Analysis")
        if rows is None:
            errors.append("Sheet 'Position Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 20:
                errors.append(f"Position Analysis has {len(data_rows)} rows, expected at least 20")
            # Check all 5 symbols present
            symbols = {str(r[0]).strip().upper() for r in data_rows if r[0]}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols:
                    errors.append(f"Symbol {sym} missing from Position Analysis")

        # Expiry Alerts
        rows2 = load_sheet_rows(wb, "Expiry Alerts")
        if rows2 is None:
            errors.append("Sheet 'Expiry Alerts' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(f"Expiry Alerts has {len(data_rows2)} rows, expected at least 5")

        # Summary
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows3 if r[0]}
            if "total_positions" in lookup:
                if not num_close(lookup["total_positions"], 50, abs_tol=10):
                    errors.append(f"Total_Positions={lookup['total_positions']}, expected ~50")
            if "near_expiry_count" in lookup:
                if not num_close(lookup["near_expiry_count"], 16, abs_tol=10):
                    errors.append(f"Near_Expiry_Count={lookup['near_expiry_count']}, expected ~16")
            if "stocks_with_near_expiry" in lookup:
                val = str(lookup["stocks_with_near_expiry"]).upper()
                for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                    if sym not in val:
                        errors.append(f"{sym} missing from Stocks_With_Near_Expiry")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE summary ILIKE '%options expiry%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if len(rows) < 3:
            errors.append(f"Only {len(rows)} options expiry calendar events, expected at least 3")
        else:
            summaries = " ".join(r[0].upper() for r in rows)
            for sym in ["AMZN", "GOOGL", "JPM"]:
                if sym not in summaries:
                    errors.append(f"No calendar event for {sym}")
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

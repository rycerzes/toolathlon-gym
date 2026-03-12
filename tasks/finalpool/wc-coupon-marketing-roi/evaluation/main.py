"""Evaluation for wc-coupon-marketing-roi."""
import argparse
import json
import os
import sys

import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_data):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Campaign_ROI.xlsx")
    if not os.path.exists(path):
        return ["Campaign_ROI.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Campaign Performance sheet
        rows = load_sheet_rows(wb, "Campaign Performance")
        if rows is None:
            errors.append("Sheet 'Campaign Performance' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_campaigns"]
            if abs(len(data_rows) - expected) > 2:
                errors.append(f"Campaign Performance has {len(data_rows)} rows, expected {expected}")

            # Check specific campaign values
            for gc in gt_data["campaigns"][:3]:
                camp_rows = [r for r in data_rows if r[0] and gc["campaign_name"].lower() in str(r[0]).lower()]
                if camp_rows:
                    row = camp_rows[0]
                    # Check ROI with tolerance
                    if len(row) > 6 and row[6] is not None:
                        try:
                            roi_val = float(row[6])
                            if abs(roi_val - gc["roi_pct"]) > 15:
                                errors.append(f"{gc['campaign_name']} ROI={roi_val}, expected ~{gc['roi_pct']}")
                        except (ValueError, TypeError):
                            pass

        # Check Channel Summary sheet
        rows2 = load_sheet_rows(wb, "Channel Summary")
        if rows2 is None:
            errors.append("Sheet 'Channel Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 3:
                errors.append(f"Channel Summary has too few rows: {len(data_rows2)}")

        # Check Recommendations sheet
        rows3 = load_sheet_rows(wb, "Recommendations")
        if rows3 is None:
            errors.append("Sheet 'Recommendations' not found")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%campaign%' OR title ILIKE '%effectiveness%' OR title ILIKE '%feedback%'
            ORDER BY created_at DESC LIMIT 5
        """)
        rows = cur.fetchall()

        if not rows:
            errors.append("No Google Form found matching 'Campaign Effectiveness Feedback'")
        else:
            form_id = rows[0][0]
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
            q_count = cur.fetchone()[0]
            if q_count < 2:
                errors.append(f"Form has {q_count} questions, expected at least 3")

        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Google Form...")
    errs = check_gform()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

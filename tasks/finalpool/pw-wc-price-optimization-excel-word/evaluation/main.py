"""Evaluation script for pw-wc-price-optimization-excel-word."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Price_Optimization_Report.xlsx")
    check("Price_Optimization_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Price_Optimization_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        check("Price_Comparison sheet exists", "Price_Comparison" in wb.sheetnames)
        if "Price_Comparison" in wb.sheetnames:
            ws = wb["Price_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Price_Comparison has >= 10 rows", len(data_rows) >= 10, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Product_Name', 'Our_Price', 'Competitor_Price', 'Price_Difference', 'Difference_Pct', 'Recommendation']:
                check(f"Price_Comparison has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Category_Summary sheet exists", "Category_Summary" in wb.sheetnames)
        if "Category_Summary" in wb.sheetnames:
            ws = wb["Category_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Category_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Category', 'Product_Count', 'Avg_Our_Price', 'Avg_Stock']:
                check(f"Category_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Executive_Summary sheet exists", "Executive_Summary" in wb.sheetnames)
        if "Executive_Summary" in wb.sheetnames:
            ws = wb["Executive_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Executive_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Executive_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        word_path = os.path.join(agent_workspace, "Pricing_Strategy.docx")
        check("Pricing strategy Word exists", os.path.exists(word_path))
        if os.path.exists(word_path):
            from docx import Document
            doc = Document(word_path)
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word mentions pricing", "pric" in text)
            check("Word mentions recommendation", "recommend" in text)
        check("price_optimizer.py exists", os.path.exists(os.path.join(agent_workspace, "price_optimizer.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

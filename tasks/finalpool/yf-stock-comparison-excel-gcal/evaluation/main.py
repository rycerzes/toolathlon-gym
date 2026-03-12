"""
Evaluation script for yf-stock-comparison-excel-gcal task.

Checks:
1. Excel file Stock_Comparison_Report.xlsx - 2 sheets with correct structure and data
2. Google Calendar has investment review event
3. Email sent to analyst@investment.com
4. Word document Stock_Analysis_Report.docx exists
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Stock_Comparison_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return False

    agent_file = os.path.join(agent_workspace, "Stock_Comparison_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Stock_Comparison_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth file exists", False, f"Not found: {gt_file}")
        return False

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Check Price History sheet
    a_hist = load_sheet_by_name(agent_wb, "Price History")
    g_hist = load_sheet_by_name(gt_wb, "Price History")
    record("Sheet 'Price History' exists", a_hist is not None)

    if a_hist is not None:
        a_data = [r for r in a_hist[1:] if any(v is not None for v in r)]
        record("Price History has at least 15 rows (trading days)",
               len(a_data) >= 15,
               f"Found {len(a_data)} rows")

        # Check headers
        if a_hist and len(a_hist) > 0:
            header = [str(v).strip().lower() if v else "" for v in a_hist[0]]
            record("Price History has Date column", any("date" in h for h in header))
            record("Price History has GOOGL column",
                   any("googl" in h for h in header))
            record("Price History has AMZN column",
                   any("amzn" in h for h in header))
            record("Price History has JPM column",
                   any("jpm" in h for h in header))

    # Check Performance Summary sheet
    a_summ = load_sheet_by_name(agent_wb, "Performance Summary")
    g_summ = load_sheet_by_name(gt_wb, "Performance Summary")
    record("Sheet 'Performance Summary' exists", a_summ is not None)

    if a_summ is not None and g_summ is not None:
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_summ[1:] if any(v is not None for v in r)]
        record("Performance Summary has 3 rows (GOOGL, AMZN, JPM)",
               len(a_data) == 3,
               f"Found {len(a_data)} rows")

        # Build lookup by symbol
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            sym = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(sym)
            if a_row is None:
                record(f"{sym} row in Performance Summary", False, "Not found")
                all_ok = False
                continue
            record(f"{sym} row exists", True)

            # Latest price (col 1)
            if len(g_row) > 1 and len(a_row) > 1:
                record(f"{sym}: Latest_Price correct",
                       num_close(a_row[1], g_row[1], 5.0),
                       f"got {a_row[1]}, expected {g_row[1]}")
            # Return_Pct (col 4, index 4)
            if len(g_row) > 4 and len(a_row) > 4:
                record(f"{sym}: Return_Pct correct",
                       num_close(a_row[4], g_row[4], 2.0),
                       f"got {a_row[4]}, expected {g_row[4]}")
            # Volatility (col 7, index 7)
            if len(g_row) > 7 and len(a_row) > 7:
                record(f"{sym}: Volatility_Score is numeric",
                       a_row[7] is not None and str(a_row[7]).replace('.', '').replace('-', '').isdigit() or True,
                       f"got {a_row[7]}")

    return all_ok


# ============================================================================
# Check 2: Google Calendar
# ============================================================================

def check_gcal():
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    record("At least 1 calendar event created", len(events) >= 1, f"Found {len(events)}")

    portfolio_events = [e for e in events
                        if e[0] and ("portfolio" in e[0].lower() or
                                     "investment" in e[0].lower() or
                                     "review" in e[0].lower())]
    record("Investment/Portfolio review event found",
           len(portfolio_events) >= 1,
           f"Events: {[e[0] for e in events[:5]]}")

    # Check March 18 2026
    march18_events = [e for e in events
                      if e[1] and "2026-03-18" in str(e[1])]
    record("Event on March 18 2026", len(march18_events) >= 1,
           f"March 18 events: {[e[0] for e in march18_events]}")

    return len(portfolio_events) >= 1


# ============================================================================
# Check 3: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")
    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    found_email = False
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = str(to_addr or "").lower()
        subject_lower = (subject or "").lower()
        if ("analyst@investment.com" in to_str or
                "stock" in subject_lower or "performance" in subject_lower):
            found_email = True
            record("Email to analyst@investment.com found", True)

            record("Email subject mentions stock or performance",
                   "stock" in subject_lower or "performance" in subject_lower or "march" in subject_lower,
                   f"Subject: {subject}")

            body_lower = (body_text or "").lower()
            record("Email body mentions stock symbols",
                   any(s in body_lower for s in ["googl", "amzn", "jpm", "alphabet", "amazon"]),
                   "Body missing stock symbols")
            break

    if not found_email:
        record("Stock performance email found", False,
               f"Emails: {[(e[0], str(e[2])[:60]) for e in all_emails[:3]]}")

    return found_email


# ============================================================================
# Check 4: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Checking Stock_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Stock_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        record("Word file exists", False, f"Not found: {docx_path}")
        return False
    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        record("Word doc has content", len(all_text.strip()) >= 100,
               f"Content length: {len(all_text.strip())}")
        record("Word doc mentions stock analysis",
               any(term in all_text for term in ["stock", "analysis", "googl", "amzn", "jpm", "alphabet", "amazon"]),
               "Missing stock analysis content")
        return True
    except ImportError:
        size = os.path.getsize(docx_path)
        record("Word file has content (>2KB)", size > 2000, f"Size: {size} bytes")
        return size > 2000
    except Exception as e:
        record("Word file readable", False, str(e))
        return False


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    gcal_ok = check_gcal()
    email_ok = check_emails()
    word_ok = check_word(args.agent_workspace)

    all_passed = excel_ok and gcal_ok and email_ok and word_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

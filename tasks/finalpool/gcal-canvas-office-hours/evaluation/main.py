"""
Evaluation script for gcal-canvas-office-hours task.

Checks:
1. Excel file Office_Hours_Schedule.xlsx - Bookings sheet (6 rows) and Summary sheet
2. Google Calendar events - office hour events on March 9-13, 2026
3. Emails sent - confirmation emails to each student who booked
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def get_form_responses():
    """Read all form responses from the database."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get question IDs and titles
    cur.execute("""
        SELECT q.id, q.title
        FROM gform.questions q
        JOIN gform.forms f ON q.form_id = f.id
        WHERE LOWER(f.title) LIKE '%office hours%booking%'
        ORDER BY q.position
    """)
    question_map = {row[0]: row[1] for row in cur.fetchall()}

    # Get responses
    cur.execute("""
        SELECT r.answers
        FROM gform.responses r
        JOIN gform.forms f ON r.form_id = f.id
        WHERE LOWER(f.title) LIKE '%office hours%booking%'
    """)
    raw_responses = cur.fetchall()

    cur.close()
    conn.close()

    # Find question IDs by title pattern
    q_name_id = q_email_id = q_date_id = q_time_id = q_topic_id = None
    for qid, qtitle in question_map.items():
        tl = qtitle.lower()
        if "name" in tl:
            q_name_id = qid
        elif "email" in tl:
            q_email_id = qid
        elif "date" in tl:
            q_date_id = qid
        elif "time" in tl or "slot" in tl:
            q_time_id = qid
        elif "topic" in tl:
            q_topic_id = qid

    responses = []
    for (answers_raw,) in raw_responses:
        if isinstance(answers_raw, str):
            answers = json.loads(answers_raw)
        else:
            answers = answers_raw

        responses.append({
            "name": answers.get(q_name_id, ""),
            "email": answers.get(q_email_id, ""),
            "date": answers.get(q_date_id, ""),
            "time_slot": answers.get(q_time_id, ""),
            "topic": answers.get(q_topic_id, ""),
        })

    return responses


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace):
    """Check Office_Hours_Schedule.xlsx content."""
    print("\n=== Checking Excel Output ===")

    responses = get_form_responses()
    if not responses:
        check("Form responses available", False, "No form responses found in DB")
        return False

    agent_file = os.path.join(agent_workspace, "Office_Hours_Schedule.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    check("Excel file readable", True)

    all_ok = True

    # Check sheet names
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_bookings = any("booking" in s for s in sheet_names_lower)
    has_summary = any("summary" in s for s in sheet_names_lower)
    check("Sheet 'Bookings' exists", has_bookings, f"Found: {wb.sheetnames}")
    check("Sheet 'Summary' exists", has_summary, f"Found: {wb.sheetnames}")

    if not has_bookings:
        all_ok = False
    if not has_summary:
        all_ok = False

    # --- Bookings sheet ---
    print("\n--- Bookings Sheet ---")
    ws_bookings = None
    for s in wb.sheetnames:
        if "booking" in s.lower():
            ws_bookings = wb[s]
            break

    if ws_bookings:
        data_rows = list(ws_bookings.iter_rows(min_row=2, values_only=True))
        # Filter out completely empty rows
        data_rows = [r for r in data_rows if any(c is not None for c in r)]
        check("Bookings has 6 data rows", len(data_rows) == 6,
              f"Got {len(data_rows)}")
        if len(data_rows) != 6:
            all_ok = False

        # Check headers
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws_bookings[1]]
        expected_headers = ["student_name", "student_email", "preferred_date",
                            "preferred_time_slot", "topic"]
        for eh in expected_headers:
            found = any(eh.replace("_", "") in h.replace("_", "").replace(" ", "")
                        for h in headers)
            check(f"Header '{eh}' present", found, f"Headers: {headers}")
            if not found:
                all_ok = False

        # Check each response is present by student name
        agent_names = set()
        for row in data_rows:
            if row and row[0]:
                agent_names.add(str(row[0]).strip().lower())

        for resp in responses:
            name_lower = resp["name"].strip().lower()
            found = name_lower in agent_names
            check(f"Booking for '{resp['name']}' present", found)
            if not found:
                all_ok = False

    # --- Summary sheet ---
    print("\n--- Summary Sheet ---")
    ws_summary = None
    for s in wb.sheetnames:
        if "summary" in s.lower():
            ws_summary = wb[s]
            break

    if ws_summary:
        data_rows = list(ws_summary.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if any(c is not None for c in r)]

        # Count unique dates from responses
        unique_dates = set(r["date"] for r in responses)
        check("Summary has correct number of date rows",
              len(data_rows) == len(unique_dates),
              f"Expected {len(unique_dates)}, got {len(data_rows)}")

        # Check headers include Date, Total_Bookings, Time_Slots
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws_summary[1]]
        for eh in ["date", "total"]:
            found = any(eh in h for h in headers)
            check(f"Summary header contains '{eh}'", found, f"Headers: {headers}")
            if not found:
                all_ok = False

        # Check total bookings per date
        date_counts = {}
        for r in responses:
            date_counts[r["date"]] = date_counts.get(r["date"], 0) + 1

        for row in data_rows:
            if row and row[0]:
                date_val = str(row[0]).strip()
                total_val = row[1] if len(row) > 1 else None
                # Find matching expected date
                matched_date = None
                for d in date_counts:
                    if d.lower() in date_val.lower() or date_val.lower() in d.lower():
                        matched_date = d
                        break
                    # Also try matching just the day number
                    try:
                        if str(int(float(str(row[0])))) in d:
                            matched_date = d
                            break
                    except (ValueError, TypeError):
                        pass

                if matched_date and total_val is not None:
                    try:
                        check(f"Summary total for '{date_val}'",
                              int(float(str(total_val))) == date_counts[matched_date],
                              f"Expected {date_counts[matched_date]}, got {total_val}")
                    except (ValueError, TypeError):
                        check(f"Summary total for '{date_val}' is numeric", False,
                              f"Got {total_val}")

    return all_ok


# ============================================================================
# Check 2: Google Calendar events
# ============================================================================

def check_gcal():
    """Verify office hour events were created."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    for ev in events:
        print(f"  Event: {ev[0]} | {ev[2]} - {ev[3]}")

    responses = get_form_responses()

    # Count unique (date, time_slot) combos
    unique_slots = set()
    for r in responses:
        unique_slots.add((r["date"], r["time_slot"]))

    # At least 4 events (some tolerance)
    check("At least 4 calendar events created",
          len(events) >= 4,
          f"Found {len(events)}, expected at least 4")

    # Ideally should match unique slots
    check(f"Calendar events match unique slots ({len(unique_slots)})",
          len(events) >= len(unique_slots),
          f"Found {len(events)}, expected at least {len(unique_slots)}")

    # Check events have "office hours" in summary
    oh_events = [e for e in events if "office hours" in (e[0] or "").lower()]
    check("Events have 'Office Hours' in title",
          len(oh_events) >= 4,
          f"Found {len(oh_events)} with 'Office Hours' in title")

    # Check events fall in March 9-13, 2026
    valid_dates = {"2026-03-09", "2026-03-10", "2026-03-11", "2026-03-12", "2026-03-13"}

    events_in_range = 0
    for summary, description, start_dt, end_dt in events:
        if start_dt:
            date_str = start_dt.strftime("%Y-%m-%d")
            if date_str in valid_dates:
                events_in_range += 1

    check("Events on dates March 9-13, 2026",
          events_in_range >= 4,
          f"Found {events_in_range} events in valid date range")

    # Check duration is 30 minutes
    duration_ok = 0
    for summary, description, start_dt, end_dt in events:
        if start_dt and end_dt:
            duration_min = (end_dt - start_dt).total_seconds() / 60
            if 25 <= duration_min <= 35:
                duration_ok += 1

    check("Events have ~30 minute duration",
          duration_ok >= 4,
          f"{duration_ok} events have ~30min duration")


# ============================================================================
# Check 3: Emails
# ============================================================================

def check_emails():
    """Verify confirmation emails were sent to each student."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    responses = get_form_responses()

    print(f"[check_emails] Found {len(all_emails)} total emails.")

    # At least 4 emails (some tolerance)
    check("At least 4 emails sent",
          len(all_emails) >= 4,
          f"Found {len(all_emails)}, expected at least 4")

    # Check emails with "confirmation" in subject
    confirmation_emails = [
        e for e in all_emails
        if "confirmation" in (e[0] or "").lower()
    ]

    check("Emails have 'Confirmation' in subject",
          len(confirmation_emails) >= 4,
          f"Found {len(confirmation_emails)} with confirmation subject")

    # Check sender is ta@university.edu
    ta_emails = [
        e for e in all_emails
        if "ta@university.edu" in (str(e[1]) or "").lower()
    ]
    check("Emails sent from ta@university.edu",
          len(ta_emails) >= 4,
          f"Found {len(ta_emails)} from ta@university.edu")

    # Check each student received an email
    all_to_addrs = []
    for subject, from_addr, to_addr, body_text in all_emails:
        if isinstance(to_addr, list):
            for addr in to_addr:
                all_to_addrs.append(str(addr).lower())
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                if isinstance(parsed, list):
                    for addr in parsed:
                        all_to_addrs.append(str(addr).lower())
                else:
                    all_to_addrs.append(str(to_addr).lower())
            except (json.JSONDecodeError, TypeError):
                all_to_addrs.append(str(to_addr).lower())

    all_to_str = " ".join(all_to_addrs)

    students_emailed = 0
    for resp in responses:
        found = resp["email"].lower() in all_to_str
        check(f"Confirmation email sent to {resp['name']}",
              found,
              f"Email to {resp['email']} not found")
        if found:
            students_emailed += 1

    # Check email body mentions date, time, and topic for at least some emails
    if all_emails:
        bodies_checked = 0
        for subject, from_addr, to_addr, body_text in all_emails:
            body_lower = (body_text or "").lower()
            has_date = any(
                d.lower() in body_lower
                for d in ["march 9", "march 10", "march 11", "march 12", "march 13",
                           "2026-03-09", "2026-03-10", "2026-03-11", "2026-03-12", "2026-03-13"]
            )
            has_time = any(
                t.lower() in body_lower
                for t in ["9:00", "10:00", "11:00", "2:00", "14:00"]
            )
            if has_date and has_time:
                bodies_checked += 1

        check("Email bodies include date and time info",
              bodies_checked >= 4,
              f"{bodies_checked} emails mention date and time")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if FAIL_COUNT == 0 else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": FAIL_COUNT == 0,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

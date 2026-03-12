"""
Evaluation for yt-top-videos-excel-gcal task.

Checks:
1. Top_Videos_Watchlist.xlsx exists with Watchlist and Summary sheets
2. Watchlist has 10 rows, correct columns, correct top video
3. Summary has correct combined view count
4. GCal has 10 learning sessions on Tuesdays April-June 2026
5. Email sent to team@company.com with correct subject
"""
import json
import os
import sys
from argparse import ArgumentParser
from datetime import date

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Top_Videos_Watchlist.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Top_Videos_Watchlist.xlsx")
    if not os.path.exists(xlsx_path):
        record("Top_Videos_Watchlist.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Top_Videos_Watchlist.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Watchlist sheet
    if "watchlist" not in sheet_names_lower:
        record("Watchlist sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Watchlist sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("watchlist")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Watchlist has 10 rows", len(data_rows) == 10, f"Found {len(data_rows)}")

        # Check columns
        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_rank = any("rank" in h for h in headers)
            has_dur = any("duration" in h or "dur" in h for h in headers)
            has_like_rate = any("like_rate" in h or "rate" in h for h in headers)
            record("Has Rank, Duration_Min, and Like_Rate columns",
                   has_rank and has_dur and has_like_rate, f"Headers: {rows[0]}")

        # Spot check: first video should be the most viewed (DeepSeek)
        if data_rows:
            first_row = data_rows[0]
            all_text = " ".join(str(c) for r in rows for c in r if c).lower()
            has_top_video = "deepseek" in all_text or "nl7acuswykg" in all_text.lower()
            record("Top-ranked video is most-viewed (DeepSeek/Nl7aCUsWykg)",
                   has_top_video, f"First row: {first_row[:5]}")

    # Summary sheet
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        # Combined views should be ~25315767
        has_combined = "25315767" in all_text2 or "combined_views" in all_text2
        record("Summary has Combined_Views", has_combined, "Missing combined views data")
        has_10 = "10" in all_text2
        record("Summary shows Total_Selected = 10", has_10, "Missing 10 in summary")


def check_gcal():
    print("\n=== Check 2: GCal learning sessions ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE summary ILIKE '%team learning%'
        AND start_datetime >= '2026-04-07' AND start_datetime < '2026-07-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    record("At least 10 'Team Learning' events scheduled", len(events) >= 10,
           f"Found {len(events)} events")

    if events:
        # Check first event is on Tuesday April 7, 2026
        first_start = events[0][1]
        if first_start:
            first_date = first_start.date() if hasattr(first_start, 'date') else first_start
            is_april7 = str(first_date) == "2026-04-07"
            record("First session on 2026-04-07 (Tuesday)", is_april7,
                   f"First event date: {first_date}")

        # Check sessions are 1 hour
        summary, start_dt, end_dt = events[0]
        if start_dt and end_dt:
            duration_hours = (end_dt - start_dt).total_seconds() / 3600
            record("Sessions are 1 hour long", abs(duration_hours - 1.0) < 0.1,
                   f"Duration: {duration_hours:.2f} hours")

        # Check all on Tuesdays
        all_tuesdays = all(
            e[1].weekday() == 1 for e in events if e[1]
        )
        record("All sessions scheduled on Tuesdays", all_tuesdays,
               f"Weekdays: {[e[1].weekday() for e in events if e[1]]}")


def check_email():
    print("\n=== Check 3: Email sent ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT m.to_addr, m.subject FROM email.messages m
        JOIN email.sent_log sl ON sl.message_id = m.id
        WHERE m.to_addr::text ILIKE %s
        ORDER BY sl.sent_at DESC LIMIT 5
    """, ("%team%",))
    emails = cur.fetchall()

    if not emails:
        cur.execute("""
            SELECT to_addr, subject FROM email.messages
            WHERE to_addr::text ILIKE %s
            ORDER BY date DESC LIMIT 5
        """, ("%team%",))
        emails = cur.fetchall()

    cur.close()
    conn.close()

    record("Email sent to team@company.com", len(emails) >= 1,
           f"Found: {emails}")

    if emails:
        subject = str(emails[0][1]).lower() if emails[0][1] else ""
        record("Email subject mentions 'Learning Sessions'",
               "learning" in subject or "scheduled" in subject,
               f"Subject: {emails[0][1]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""Evaluation for sf-hr-mentorship-excel-gcal-email.

Checks:
1. Mentorship_Pairs.xlsx has correct sheets and data
2. Google Calendar event "Mentorship Kickoff Meeting" scheduled 7 days from launch
3. Email sent to program@hr.example.com with subject containing "Mentorship"
"""
import argparse
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Mentorship_Pairs.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Mentorship_Pairs.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Mentorship_Pairs.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # Check Pairs sheet
    agent_pairs = get_sheet(agent_wb, "Pairs")
    gt_pairs = get_sheet(gt_wb, "Pairs")
    check("Sheet 'Pairs' exists", agent_pairs is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_pairs is None:
        all_ok = False
    else:
        a_rows = list(agent_pairs.iter_rows(min_row=2, values_only=True))
        g_rows = list(gt_pairs.iter_rows(min_row=2, values_only=True))
        check("Pairs has 10 rows", len(a_rows) == 10, f"Got {len(a_rows)}")
        if len(a_rows) != 10:
            all_ok = False

        # Build lookup by mentor name
        a_lookup = {}
        for r in a_rows:
            if r and r[0]:
                a_lookup[str(r[0]).strip().lower()] = r

        mentor_names_present = 0
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().lower()
            if key in a_lookup:
                mentor_names_present += 1
        check("At least 8 mentor names match", mentor_names_present >= 8,
              f"Matched {mentor_names_present}/10")
        if mentor_names_present < 8:
            all_ok = False

        # Check column count
        if a_rows:
            check("Pairs has 6 columns", len([v for v in a_rows[0] if v is not None]) >= 5,
                  f"Got {len([v for v in a_rows[0] if v is not None])}")

    # Check Program_Summary sheet
    agent_summary = get_sheet(agent_wb, "Program_Summary")
    gt_summary = get_sheet(gt_wb, "Program_Summary")
    check("Sheet 'Program_Summary' exists", agent_summary is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if agent_summary is None:
        all_ok = False
    else:
        a_summary = {}
        for row in agent_summary.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        # Total_Pairs
        tp = a_summary.get("total_pairs")
        check("Total_Pairs = 10", num_close(tp, 10, 0), f"Got {tp}")
        if not num_close(tp, 10, 0):
            all_ok = False

        # Avg_Mentor_Rating
        amr = a_summary.get("avg_mentor_rating")
        check("Avg_Mentor_Rating close to 5.0", num_close(amr, 5.0, 0.1), f"Got {amr}")

        # Avg_Mentee_Experience
        ame = a_summary.get("avg_mentee_experience")
        check("Avg_Mentee_Experience close to 0.1", num_close(ame, 0.1, 0.5), f"Got {ame}")

    return all_ok


def check_gcal(launch_time_str):
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")
    check("At least 1 calendar event created", len(events) >= 1, f"Found {len(events)}")

    # Check for kickoff meeting
    kickoff_events = [e for e in events if "mentorship" in (e[0] or "").lower() and "kickoff" in (e[0] or "").lower()]
    check("Mentorship Kickoff Meeting event exists", len(kickoff_events) >= 1,
          f"Events: {[e[0] for e in events]}")

    if launch_time_str and kickoff_events:
        try:
            launch_dt = datetime.fromisoformat(launch_time_str)
            expected_dt = launch_dt + timedelta(days=7)
            for ev in kickoff_events:
                if ev[2]:
                    ev_dt = ev[2]
                    if hasattr(ev_dt, 'date'):
                        diff = abs((ev_dt.replace(tzinfo=None) - expected_dt).total_seconds())
                        check("Kickoff meeting 7 days from launch", diff <= 86400 * 2,
                              f"Expected around {expected_dt}, got {ev_dt}")
                        break
        except Exception as e:
            print(f"  [INFO] Could not verify date: {e}")

    return len(kickoff_events) >= 1


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr
        FROM email.messages
        WHERE subject ILIKE '%mentorship%'
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    check("Email with 'mentorship' in subject exists", len(emails) >= 1,
          f"Found {len(emails)} matching emails")

    if emails:
        e = emails[0]
        # Check recipient
        to_str = str(e[2])
        check("Email sent to program@hr.example.com",
              "program@hr.example.com" in to_str.lower(),
              f"to_addr: {to_str}")
        # Check sender
        check("Email from hr@company.example.com",
              "hr@company.example.com" in (e[1] or "").lower(),
              f"from_addr: {e[1]}")

    return len(emails) >= 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    check_gcal(args.launch_time)
    check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

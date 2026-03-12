"""
Evaluation script for playwright-wc-competitor-pricing-excel-gcal task.

Checks:
1. Competitive_Pricing_Analysis.xlsx with 4 sheets and correct data
2. Calendar events for 3 category price review meetings
3. Email sent with pricing analysis results
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=5.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


COMPETITOR_AVGS = {"electronics": 59.06, "headphones": 74.98, "speakers": 138.73}
OUR_AVG_REGULAR = {"electronics": 97.21, "headphones": 66.00, "speakers": 150.59}


def check_excel(agent_workspace):
    """Check Competitive_Pricing_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: Competitor Products ---
    comp_sheet = None
    for name in wb.sheetnames:
        if "competitor" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Competitor Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Competitor Products sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 15
        record("Competitor Products has 15 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 2: Our Products ---
    our_sheet = None
    for name in wb.sheetnames:
        if "our" in name.lower():
            our_sheet = name
            break
    if not our_sheet:
        record("Our Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Our Products sheet exists", True)
        ws = wb[our_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        # Should have products from 3 categories (30+10+5=45)
        ok = len(data_rows) >= 30
        record("Our Products has >= 30 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 3: Category Comparison ---
    cat_sheet = None
    for name in wb.sheetnames:
        if "category" in name.lower() or "comparison" in name.lower():
            cat_sheet = name
            break
    if not cat_sheet:
        record("Category Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Category Comparison sheet exists", True)
        ws = wb[cat_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 3
        record("Category Comparison has 3 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

        for row in data_rows:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat in COMPETITOR_AVGS:
                    # Check competitor avg
                    found = False
                    for cell in row[1:]:
                        if num_close(cell, COMPETITOR_AVGS[cat], tol=5.0):
                            found = True
                            break
                    record(f"Competitor avg for {cat}", found,
                           f"Expected ~{COMPETITOR_AVGS[cat]}, row: {str(row)[:200]}")
                    if not found:
                        all_ok = False

    # --- Sheet 4: Summary ---
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = name
            break
    if not sum_sheet:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Summary sheet exists", True)
        ws = wb[sum_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        for row in data_rows:
            if row and row[0]:
                metric = str(row[0]).strip().lower()
                val = row[1]
                if "total_competitor" in metric:
                    ok = num_close(val, 15, tol=0)
                    record("Total competitor products = 15", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "cheapest" in metric:
                    ok = str_contains(val, "headphone")
                    record("Cheapest category is Headphones", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "expensive" in metric:
                    ok = str_contains(val, "electronics")
                    record("Most expensive category is Electronics", ok, f"Got {val}")
                    if not ok:
                        all_ok = False

    wb.close()
    return all_ok


def check_calendar():
    """Check calendar events for 3 category price review meetings."""
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    all_ok = True
    categories_found = set()

    for summary, description, start_dt, end_dt in events:
        summary_lower = (summary or "").lower()
        if "price review" in summary_lower:
            for cat in ["electronics", "headphones", "speakers"]:
                if cat in summary_lower:
                    categories_found.add(cat)

    for cat in ["electronics", "headphones", "speakers"]:
        ok = cat in categories_found
        record(f"Price Review event for {cat}", ok,
               f"Found events for: {categories_found}")
        if not ok:
            all_ok = False

    return all_ok


def check_email():
    """Check email was sent with pricing analysis."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    all_ok = True
    found_email = False

    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if "pricing" in subj_lower or "competitive" in subj_lower or "price" in subj_lower:
            found_email = True
            record("Pricing analysis email exists", True)

            from_ok = str_contains(from_addr, "pricing") or str_contains(from_addr, "ecommerce")
            record("Email from pricing/ecommerce address", from_ok, f"From: {from_addr}")
            if not from_ok:
                all_ok = False

            body_lower = (body_text or "").lower()
            body_ok = ("electronics" in body_lower or "headphones" in body_lower or
                       "speakers" in body_lower)
            record("Email body mentions categories", body_ok,
                   f"Body preview: {(body_text or '')[:200]}")
            if not body_ok:
                all_ok = False
            break

    if not found_email:
        record("Pricing analysis email exists", False,
               f"Found {len(emails)} emails, none with pricing/competitive in subject")
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    cal_ok = check_calendar()
    email_ok = check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Calendar: {'PASS' if cal_ok else 'FAIL'}")
    print(f"  Email:    {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    overall = excel_ok and cal_ok and email_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

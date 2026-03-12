"""Evaluation for terminal-howtocook-gform-excel-notion-email.
Checks:
1. Meal_Program_Plan.xlsx with 4 sheets
2. Google Form survey with 5 questions
3. Notion "Recipe Knowledge Base" with menu entries
4. Email sent to all_staff
5. menu_planner.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: Meal_Program_Plan.xlsx ===")
    path = os.path.join(workspace, "Meal_Program_Plan.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Survey_Questions
    sq_idx = next((i for i, s in enumerate(sheets_lower) if "survey" in s or "question" in s), 0)
    ws1 = wb[sheets[sq_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Survey_Questions has 5 rows", len(data1) >= 5, f"Found {len(data1)}")
    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Has cuisine preference question", "cuisine" in all_text1 or "preferred" in all_text1,
          f"Text: {all_text1[:100]}")
    check("Has dietary restriction question", "dietary" in all_text1 or "restriction" in all_text1,
          f"Text: {all_text1[:100]}")

    # Recipe_Selection
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "recipe" in s and "select" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Recipe_Selection has 7+ recipes", len(data2) >= 7, f"Found {len(data2)}")
        if rows2:
            headers = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has difficulty column", any("difficult" in h for h in headers),
                  f"Headers: {rows2[0]}")

    # Weekly_Menu
    wm_idx = next((i for i, s in enumerate(sheets_lower) if "weekly" in s or "menu" in s), 2)
    if wm_idx < len(sheets):
        ws3 = wb[sheets[wm_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Weekly_Menu has 5 days", len(data3) >= 5, f"Found {len(data3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Menu includes Monday", "monday" in all_text3)
        check("Menu includes Friday", "friday" in all_text3)
        # Check cost column has values
        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            cost_idx = next((i for i, h in enumerate(headers3) if "cost" in h), -1)
            if cost_idx >= 0 and data3:
                costs = [r[cost_idx] for r in data3 if r[cost_idx] is not None]
                check("Cost values present", len(costs) >= 3, f"Found {len(costs)} costs")

    # Program_Summary
    ps_idx = next((i for i, s in enumerate(sheets_lower) if "program" in s or "summary" in s), 3)
    if ps_idx < len(sheets):
        ws4 = wb[sheets[ps_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Program_Summary has 4+ metrics", len(data4) >= 4, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has total weekly cost", "total" in all_text4 and ("weekly" in all_text4 or "cost" in all_text4))


def check_gform():
    print("\n=== Check 2: Google Form Survey ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        lunch_form = None
        for form_id, title in forms:
            if title and ("lunch" in title.lower() or "preference" in title.lower() or "employee" in title.lower()):
                lunch_form = (form_id, title)
                break
        check("Lunch preference survey form exists", lunch_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        if lunch_form:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (lunch_form[0],))
            q_count = cur.fetchone()[0]
            check("Survey has 5 questions", q_count >= 5, f"Found {q_count}")

            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position", (lunch_form[0],))
            questions = cur.fetchall()
            q_text = " ".join(str(q[0]) for q in questions).lower()
            check("Has cuisine question", "cuisine" in q_text or "prefer" in q_text,
                  f"Questions: {q_text[:100]}")
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_notion():
    print("\n=== Check 3: Notion Recipe Knowledge Base ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        recipe_db = None
        for db_id, title in dbs:
            title_str = ""
            if isinstance(title, list):
                title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
            elif isinstance(title, str):
                try:
                    parsed = json.loads(title)
                    if isinstance(parsed, list):
                        title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                    else:
                        title_str = str(title)
                except Exception:
                    title_str = str(title)
            else:
                title_str = str(title) if title else ""
            if "recipe" in title_str.lower() and ("knowledge" in title_str.lower() or "base" in title_str.lower()):
                recipe_db = (db_id, title_str)
                break
        check("Recipe Knowledge Base database exists", recipe_db is not None,
              f"Databases: {[d[1] for d in dbs]}")

        if recipe_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (recipe_db[0],))
            count = cur.fetchone()[0]
            check("Knowledge base has 5+ recipe entries", count >= 5, f"Found {count}")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Check 4: Email to All Staff ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%all_staff%%'
               OR subject ILIKE '%%lunch program%%'
               OR subject ILIKE '%%lunch%%survey%%'
               OR subject ILIKE '%%employee%%lunch%%'
        """)
        emails = cur.fetchall()
        if not emails:
            cur.execute("""
                SELECT id, subject, to_addr, body_text
                FROM email.drafts
                WHERE to_addr::text ILIKE '%%all_staff%%'
                   OR subject ILIKE '%%lunch%%'
            """)
            emails = cur.fetchall()
        check("Email about lunch program sent", len(emails) >= 1, "No matching email found")
        if emails:
            subject = str(emails[0][1]).lower() if emails[0][1] else ""
            check("Email subject mentions lunch or survey",
                  "lunch" in subject or "survey" in subject or "meal" in subject,
                  f"Subject: {emails[0][1]}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: menu_planner.py ===")
    path = os.path.join(workspace, "menu_planner.py")
    check("menu_planner.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative cost values
    path = os.path.join(workspace, "Meal_Program_Plan.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        check("No negative values in Excel", False,
                              f"Found {cell} in sheet {sheet_name}")
                        return
        check("No negative values in Excel", True)

    # Email: no emails to unrelated recipients
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE subject ILIKE '%%lunch%%' AND to_addr::text ILIKE '%%competitor%%'
        """)
        bad_emails = cur.fetchone()[0]
        check("No lunch emails to competitor addresses", bad_emails == 0,
              f"Found {bad_emails}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_notion()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

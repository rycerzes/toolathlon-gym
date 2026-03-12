"""Evaluation for sf-ticket-escalation-report."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Escalation_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Escalation_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check Priority Summary sheet
    print("  Checking Priority Summary...")
    a_rows = load_sheet_rows(agent_wb, "Priority Summary")
    g_rows = load_sheet_rows(gt_wb, "Priority Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Priority Summary' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Priority Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing priority: {g_row[0]}")
                continue
            # Total_Tickets
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 50):
                    errors.append(f"{key}.Total_Tickets: {a_row[1]} vs {g_row[1]}")
            # Low_CSAT_Tickets
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 50):
                    errors.append(f"{key}.Low_CSAT: {a_row[2]} vs {g_row[2]}")
            # Avg_Response_Hours
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1.0):
                    errors.append(f"{key}.Avg_Response: {a_row[4]} vs {g_row[4]}")
            # Avg_CSAT
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.1):
                    errors.append(f"{key}.Avg_CSAT: {a_row[5]} vs {g_row[5]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Summary sheet
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if "issue_type" in key or "priority" in key:
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif "pct" in key:
                    if not num_close(a_row[1], g_row[1], 1.0):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif "hours" in key or "satisfaction" in key or "csat" in key:
                    if not num_close(a_row[1], g_row[1], 0.5):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                else:
                    if not num_close(a_row[1], g_row[1], 100):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Reporter Analysis sheet
    print("  Checking Reporter Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Reporter Analysis")
    g_rows = load_sheet_rows(gt_wb, "Reporter Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Reporter Analysis' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Reporter Analysis' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing reporter: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 50):
                    errors.append(f"{key}.Total: {a_row[1]} vs {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check email was sent
    print("  Checking email sent...")
    try:
        db_config = {
            "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
            "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
            "user": "eigent", "password": "camel",
        }
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject LIKE '%Escalation%' OR subject LIKE '%escalation%'
        """)
        emails = cur.fetchall()

        # Also check sent_log (join with messages to get subject)
        cur.execute("""
            SELECT m.subject FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE m.subject LIKE '%Escalation%' OR m.subject LIKE '%escalation%'
        """)
        sent = cur.fetchall()
        cur.close()
        conn.close()

        if len(emails) + len(sent) < 1:
            all_errors.append("No escalation email found")
            print("    FAIL: no escalation email found")
        else:
            print("    PASS")
    except Exception as e:
        all_errors.append(f"Email check error: {e}")
        print(f"    ERROR: {e}")

    # Check Notion page
    print("  Checking Notion page...")
    try:
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE properties::text LIKE '%Escalation%' OR properties::text LIKE '%escalation%'")
        pages = cur.fetchall()
        cur.close()
        conn.close()

        if len(pages) < 1:
            all_errors.append("No Notion escalation page found")
            print("    FAIL: no Notion escalation page")
        else:
            print("    PASS")
    except Exception as e:
        all_errors.append(f"Notion check error: {e}")
        print(f"    ERROR: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

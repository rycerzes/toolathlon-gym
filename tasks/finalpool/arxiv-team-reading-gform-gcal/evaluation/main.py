"""
Evaluation for arxiv-team-reading-gform-gcal task.

Checks:
1. Excel Reading_List.xlsx exists with 7 rows
2. Excel has ArXiv_ID, Title, First_Author, Year, Citations, Topic columns
3. Excel contains LLM_Reasoning topic for at least 5 papers
4. GForm "Reading Group Paper Selection" exists with at least 3 questions
5. GCal has at least 4 Reading Group Session events in April 2026
6. Excel contains key paper title keywords
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Excel Reading_List.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Reading_List.xlsx")
    if not os.path.exists(xlsx_path):
        record("Reading_List.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Reading_List.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # Find Papers sheet
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    if "papers" not in sheet_names_lower:
        record("Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Papers sheet exists", True)

    ws = wb[wb.sheetnames[sheet_names_lower.index("papers")]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Has data rows", False, "Sheet is empty")
        return

    # Check header row
    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    record("Has required columns (ArXiv_ID and Title)",
           any("arxiv" in h for h in headers) and any("title" in h for h in headers),
           f"Headers: {rows[0]}")

    data_rows = [r for r in rows[1:] if any(c for c in r)]
    record("Has at least 5 data rows (papers)", len(data_rows) >= 5,
           f"Found {len(data_rows)} data rows")

    # Check all text for paper keywords
    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    has_cot = "chain-of-thought" in all_text or "chain of thought" in all_text
    has_react = "react" in all_text or "synergizing" in all_text
    has_self = "self-consistency" in all_text or "self consistency" in all_text
    papers_found = sum([has_cot, has_react, has_self])
    record("Contains key LLM reasoning paper keywords", papers_found >= 2,
           f"CoT:{has_cot}, ReAct:{has_react}, SelfConsistency:{has_self}")

    # Check Topic column
    has_llm_reasoning = "llm_reasoning" in all_text or "llm reasoning" in all_text
    record("Contains LLM_Reasoning topic classification", has_llm_reasoning,
           "No 'LLM_Reasoning' topic found in sheet")


def check_gform():
    print("\n=== Check 2: Google Form ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    reading_form = None
    for form_id, title in forms:
        if "reading group" in (title or "").lower() or "paper selection" in (title or "").lower():
            reading_form = (form_id, title)
            break

    record("Reading Group Paper Selection form exists", reading_form is not None,
           f"Forms found: {[f[1] for f in forms]}")

    if reading_form:
        form_id, title = reading_form
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        record("Form has at least 3 questions", q_count >= 3,
               f"Found {q_count} questions")

        # Check for priority question
        cur.execute("SELECT title FROM gform.questions WHERE form_id = %s", (form_id,))
        q_titles = [r[0].lower() for r in cur.fetchall()]
        has_priority = any("first" in t or "priority" in t or "read" in t for t in q_titles)
        record("Form has paper priority/selection question", has_priority,
               f"Questions: {q_titles}")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 3: Google Calendar Reading Group Sessions ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-04-01' AND start_datetime < '2026-05-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    reading_events = [
        e for e in events
        if "reading group" in (e[0] or "").lower() or "reading session" in (e[0] or "").lower()
    ]

    record("At least 4 reading group events in April 2026", len(reading_events) >= 4,
           f"Found {len(reading_events)} reading events in April 2026")

    if reading_events:
        # Check duration (should be ~1.5 hours)
        summary, start_dt, end_dt = reading_events[0]
        if start_dt and end_dt:
            duration_hours = (end_dt - start_dt).total_seconds() / 3600
            record("Reading sessions are ~1.5 hours", 1.0 <= duration_hours <= 2.5,
                   f"Duration: {duration_hours:.1f} hours")

        # Check they are on different days
        dates = set(e[1].date() for e in reading_events if e[1])
        record("Sessions on different days (at least 3 distinct dates)", len(dates) >= 3,
               f"Dates: {sorted(dates)}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

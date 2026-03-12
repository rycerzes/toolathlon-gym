"""Evaluation for terminal-canvas-arxiv-word-excel-notion."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


EXPECTED_SUBJECTS = [
    "applied analytics", "biochemistry", "creative computing",
    "data-driven design", "environmental economics", "foundations of finance",
    "global governance"
]


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    fpath = os.path.join(agent_workspace, "Curriculum_Gap_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(fpath, data_only=True)

    # Sheet 1: Course_Topics
    ct_sheet = None
    for name in wb.sheetnames:
        if "course" in name.lower() and "topic" in name.lower():
            ct_sheet = name
            break
    if not ct_sheet:
        record("Course_Topics sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Course_Topics sheet exists", True)
        ws = wb[ct_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Query dynamic unique subject count from Canvas DB
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(DISTINCT regexp_replace(name, ' \\(.*\\)$', '')) FROM canvas.courses")
            expected_course_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_course_count = 7
        record(f"Course_Topics has {expected_course_count} rows",
               len(rows) == expected_course_count, f"Got {len(rows)}")

    # Sheet 2: Research_Frontiers
    rf_sheet = None
    for name in wb.sheetnames:
        if "research" in name.lower() or "frontier" in name.lower():
            rf_sheet = name
            break
    if not rf_sheet:
        record("Research_Frontiers sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Research_Frontiers sheet exists", True)
        ws = wb[rf_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Research_Frontiers has >= 5 rows", len(rows) >= 5, f"Got {len(rows)}")

    # Sheet 3: Gap_Matrix
    gm_sheet = None
    for name in wb.sheetnames:
        if "gap" in name.lower() and "matrix" in name.lower():
            gm_sheet = name
            break
    if not gm_sheet:
        record("Gap_Matrix sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Gap_Matrix sheet exists", True)
        ws = wb[gm_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Gap_Matrix has >= 20 rows", len(rows) >= 20, f"Got {len(rows)}")

        # Check gap types exist
        gap_types = set()
        for row in rows:
            if row and len(row) > 3 and row[3]:
                gap_types.add(str(row[3]).strip().lower())
        record("Gap_Matrix has Covered type", "covered" in gap_types, f"Found types: {gap_types}")
        record("Gap_Matrix has Gap type", "gap" in gap_types, f"Found types: {gap_types}")

    wb.close()
    return True


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    fpath = os.path.join(agent_workspace, "Curriculum_Enhancement_Proposal.docx")
    if not os.path.isfile(fpath):
        record("Word document exists", False, f"Not found: {fpath}")
        return False
    record("Word document exists", True)

    from docx import Document
    doc = Document(fpath)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    record("Document mentions curriculum", "curriculum" in full_text)
    record("Document mentions gap", "gap" in full_text)
    # Check at least 3 course subjects mentioned
    mentioned = sum(1 for s in EXPECTED_SUBJECTS if s in full_text)
    record("Document mentions >= 3 course subjects", mentioned >= 3, f"Found {mentioned}")
    return True


def check_terminal_output(agent_workspace):
    print("\n=== Checking Terminal Output ===")
    fpath = os.path.join(agent_workspace, "curriculum_gap_output.txt")
    if not os.path.isfile(fpath):
        record("curriculum_gap_output.txt exists", False)
        return False
    record("curriculum_gap_output.txt exists", True)
    with open(fpath) as f:
        content = f.read().lower()
    record("Output mentions relevance or gap", "relevance" in content or "gap" in content or "score" in content)
    return True


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE title::text ILIKE '%curriculum%review%'
        """)
        count = cur.fetchone()[0]
        record("Notion Curriculum Review Tracker database exists", count >= 1, f"Found {count}")

        # Check pages (entries) exist
        cur.execute("SELECT COUNT(*) FROM notion.pages")
        page_count = cur.fetchone()[0]
        record("Notion has course subject entries", page_count >= 5, f"Found {page_count}")

        cur.close()
        conn.close()
        return count >= 1
    except Exception as e:
        record("Notion check", False, str(e))
        return False


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check no duplicate pages in Notion (each course should appear only once)
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE title::text ILIKE '%curriculum%review%'
        """)
        count = cur.fetchone()[0]
        if count >= 1:
            cur.execute("""
                SELECT p.properties::text FROM notion.pages p
                JOIN notion.databases d ON p.parent->>'database_id' = d.id
                WHERE d.title::text ILIKE '%%curriculum%%review%%'
            """)
            pages = cur.fetchall()
            page_texts = [str(p[0]).lower() for p in pages]
            # Check for duplicate entries (same text appearing twice)
            if len(page_texts) > len(set(page_texts)):
                record("No duplicate pages in Notion database", False,
                       f"Found {len(page_texts)} pages but only {len(set(page_texts))} unique")
            else:
                record("No duplicate pages in Notion database", True)
        cur.close(); conn.close()
    except Exception as e:
        record("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_terminal_output(args.agent_workspace)
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

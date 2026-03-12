"""
Evaluation script for howtocook-survey-analysis task.

Checks:
1. Excel Preference_Analysis.xlsx:
   - "Survey Results" sheet has 12 rows with correct data
   - "Preference Stats" sheet has 5 category rows with correct counts and percentages
2. Google Sheet "Recommended Menu":
   - Spreadsheet with "menu" or "recommend" in title exists
   - Has at least 5 recipe data rows
3. Cross-checks:
   - 荤菜 count=6, percentage=50.0
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


# Expected data
EXPECTED_RESPONSES = [
    {"name": "Alice", "cuisine": "荤菜", "spice": "Medium", "allergy": "None", "difficulty": "3"},
    {"name": "Bob", "cuisine": "素菜", "spice": "Mild", "allergy": "Peanut", "difficulty": "2"},
    {"name": "Carol", "cuisine": "荤菜", "spice": "Spicy", "allergy": "None", "difficulty": "4"},
    {"name": "David", "cuisine": "主食", "spice": "Medium", "allergy": "None", "difficulty": "2"},
    {"name": "Eva", "cuisine": "荤菜", "spice": "Medium", "allergy": "Shellfish", "difficulty": "3"},
    {"name": "Frank", "cuisine": "水产", "spice": "Mild", "allergy": "None", "difficulty": "2"},
    {"name": "Grace", "cuisine": "素菜", "spice": "Mild", "allergy": "None", "difficulty": "1"},
    {"name": "Henry", "cuisine": "荤菜", "spice": "Spicy", "allergy": "None", "difficulty": "5"},
    {"name": "Iris", "cuisine": "甜品", "spice": "Mild", "allergy": "Dairy", "difficulty": "2"},
    {"name": "Jack", "cuisine": "荤菜", "spice": "Medium", "allergy": "None", "difficulty": "3"},
    {"name": "Karen", "cuisine": "主食", "spice": "Medium", "allergy": "Gluten", "difficulty": "3"},
    {"name": "Leo", "cuisine": "荤菜", "spice": "Spicy", "allergy": "None", "difficulty": "4"},
]

EXPECTED_STATS = {
    "荤菜": {"count": 6, "pct": 50.0},
    "素菜": {"count": 2, "pct": 16.7},
    "主食": {"count": 2, "pct": 16.7},
    "水产": {"count": 1, "pct": 8.3},
    "甜品": {"count": 1, "pct": 8.3},
}


def check_excel(agent_workspace):
    """Check Preference_Analysis.xlsx content."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Preference_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Excel file readable", True)

    # Check sheet names
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_survey = any("survey" in s and "result" in s for s in sheet_names_lower)
    has_stats = any("preference" in s or "stat" in s for s in sheet_names_lower)
    check("Sheet 'Survey Results' exists", has_survey, f"Found: {wb.sheetnames}")
    check("Sheet 'Preference Stats' exists", has_stats, f"Found: {wb.sheetnames}")

    # --- Survey Results sheet ---
    print("\n--- Survey Results Sheet ---")
    ws_survey = None
    for s in wb.sheetnames:
        if "survey" in s.lower() and "result" in s.lower():
            ws_survey = wb[s]
            break
    if ws_survey is None:
        for s in wb.sheetnames:
            if "survey" in s.lower() or "result" in s.lower():
                ws_survey = wb[s]
                break

    if ws_survey:
        data_rows = list(ws_survey.iter_rows(min_row=2, values_only=True))
        check("Survey Results has 12 data rows", len(data_rows) == 12,
              f"Got {len(data_rows)}")

        # Check each respondent is present by name
        agent_names = set()
        for row in data_rows:
            if row and row[0]:
                agent_names.add(str(row[0]).strip().lower())

        for resp in EXPECTED_RESPONSES:
            name_lower = resp["name"].strip().lower()
            found = name_lower in agent_names
            check(f"Response for '{resp['name']}' present", found)

        # Verify specific field values
        for row in data_rows:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            for resp in EXPECTED_RESPONSES:
                if str_match(name, resp["name"]):
                    check(f"'{name}' Preferred_Type correct",
                          str_match(row[1], resp["cuisine"]),
                          f"Got '{row[1]}', expected '{resp['cuisine']}'")
                    check(f"'{name}' Spice_Level correct",
                          str_match(row[2], resp["spice"]),
                          f"Got '{row[2]}', expected '{resp['spice']}'")
                    check(f"'{name}' Allergies correct",
                          str_match(row[3], resp["allergy"]),
                          f"Got '{row[3]}', expected '{resp['allergy']}'")
                    check(f"'{name}' Max_Difficulty correct",
                          str_match(str(row[4]).rstrip('.0') if row[4] is not None else None,
                                    resp["difficulty"]),
                          f"Got '{row[4]}', expected '{resp['difficulty']}'")
                    break
    else:
        check("Survey Results sheet found", False, "Could not locate sheet")

    # --- Preference Stats sheet ---
    print("\n--- Preference Stats Sheet ---")
    ws_stats = None
    for s in wb.sheetnames:
        sl = s.lower()
        if "preference" in sl or "stat" in sl:
            ws_stats = wb[s]
            break

    if ws_stats:
        stats_rows = list(ws_stats.iter_rows(min_row=2, values_only=True))
        # Filter out empty rows
        stats_rows = [r for r in stats_rows if r and r[0] is not None and str(r[0]).strip()]
        check("Preference Stats has 5 category rows", len(stats_rows) == 5,
              f"Got {len(stats_rows)}")

        # Build dict of category -> (count, pct)
        agent_stats = {}
        for row in stats_rows:
            if row and row[0]:
                cat = str(row[0]).strip()
                count = row[1] if len(row) > 1 else None
                pct = row[2] if len(row) > 2 else None
                agent_stats[cat] = (count, pct)

        for cat, expected in EXPECTED_STATS.items():
            if cat in agent_stats:
                count, pct = agent_stats[cat]
                check(f"'{cat}' Vote_Count is {expected['count']}",
                      num_close(count, expected["count"]),
                      f"Got {count}")
                check(f"'{cat}' Percentage is {expected['pct']}",
                      num_close(pct, expected["pct"], tol=0.2),
                      f"Got {pct}")
            else:
                check(f"'{cat}' found in stats", False,
                      f"Available categories: {list(agent_stats.keys())}")

        # Check sorted by Vote_Count descending
        if len(stats_rows) >= 2:
            counts = []
            for row in stats_rows:
                try:
                    counts.append(float(row[1]))
                except (TypeError, ValueError, IndexError):
                    counts.append(0)
            is_sorted_desc = all(counts[i] >= counts[i + 1] for i in range(len(counts) - 1))
            check("Stats sorted by Vote_Count descending", is_sorted_desc,
                  f"Counts order: {counts}")
    else:
        check("Preference Stats sheet found", False, "Could not locate sheet")


def check_gsheet():
    """Check Google Sheet 'Recommended Menu' via psycopg2."""
    print("\n=== Checking Google Sheet ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Find spreadsheet with menu/recommend in title
    cur.execute("""
        SELECT id, title
        FROM gsheet.spreadsheets
        WHERE LOWER(title) LIKE '%menu%'
           OR LOWER(title) LIKE '%recommend%'
        ORDER BY created_at DESC
        LIMIT 5
    """)
    spreadsheets = cur.fetchall()

    check("Google Sheet with 'menu' or 'recommend' in title exists",
          len(spreadsheets) > 0,
          "No matching spreadsheets found")

    if not spreadsheets:
        cur.close()
        conn.close()
        return

    ss_id = spreadsheets[0][0]
    ss_title = spreadsheets[0][1]
    print(f"  [INFO] Found spreadsheet: '{ss_title}' (id={ss_id})")

    # Get sheets
    cur.execute("""
        SELECT id, title
        FROM gsheet.sheets
        WHERE spreadsheet_id = %s
    """, (ss_id,))
    sheets = cur.fetchall()
    check("Spreadsheet has at least one sheet", len(sheets) > 0)

    if not sheets:
        cur.close()
        conn.close()
        return

    sheet_id = sheets[0][0]

    # Count data rows (exclude header row at row_index=0 or 1)
    cur.execute("""
        SELECT DISTINCT row_index
        FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index
    """, (ss_id, sheet_id))
    row_indices = [r[0] for r in cur.fetchall()]

    # First row is header, rest are data
    data_row_count = max(0, len(row_indices) - 1) if row_indices else 0
    check("Recommended Menu has at least 5 recipe rows",
          data_row_count >= 5,
          f"Found {data_row_count} data rows (total rows including header: {len(row_indices)})")

    # Check that cells contain recipe names (non-empty first column values)
    cur.execute("""
        SELECT row_index, col_index, value
        FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index, col_index
    """, (ss_id, sheet_id))
    all_cells = cur.fetchall()

    # Check header row has Recipe_Name or similar
    header_cells = [c for c in all_cells if c[0] == min(r[0] for r in all_cells)]
    header_values = [str(c[2]).lower() for c in header_cells if c[2]]
    has_recipe_header = any("recipe" in h or "name" in h for h in header_values)
    has_category_header = any("category" in h or "type" in h for h in header_values)
    check("Header contains recipe name column", has_recipe_header,
          f"Headers: {header_values}")
    check("Header contains category column", has_category_header,
          f"Headers: {header_values}")

    # Check that recipe data cells are non-empty
    min_row = min(r[0] for r in all_cells) if all_cells else 0
    data_cells = [c for c in all_cells if c[0] > min_row and c[1] == 0]
    non_empty = [c for c in data_cells if c[2] and str(c[2]).strip()]
    check("Recipe name cells are non-empty",
          len(non_empty) >= 5,
          f"Found {len(non_empty)} non-empty recipe names")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gsheet()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if pass_rate >= 0.8 else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": pass_rate >= 0.8,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if pass_rate >= 0.8 else 1)


if __name__ == "__main__":
    main()

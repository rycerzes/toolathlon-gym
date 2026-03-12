"""Evaluation script for fetch-wc-inventory-forecast-excel-gcal-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Inventory_Forecast_Report.xlsx")
    check("Inventory_Forecast_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Inventory_Forecast_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        check("Stock_Status sheet exists", "Stock_Status" in wb.sheetnames)
        if "Stock_Status" in wb.sheetnames:
            ws = wb["Stock_Status"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Stock_Status has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Product', 'Current_Stock', 'Total_Sales', 'Daily_Rate', 'Days_Remaining', 'Needs_Restock']:
                check(f"Stock_Status has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Supplier_Info sheet exists", "Supplier_Info" in wb.sheetnames)
        if "Supplier_Info" in wb.sheetnames:
            ws = wb["Supplier_Info"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Supplier_Info has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Supplier', 'Lead_Time_Days', 'Min_Order_Qty', 'Reliability_Score']:
                check(f"Supplier_Info has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Restock_Summary sheet exists", "Restock_Summary" in wb.sheetnames)
        if "Restock_Summary" in wb.sheetnames:
            ws = wb["Restock_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Restock_Summary has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Restock_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT summary FROM gcal.events WHERE summary ILIKE %s", ('%inventory%',))
            check("Inventory review event created", cur.fetchone() is not None)
            cur.execute("SELECT subject FROM email.messages WHERE subject ILIKE %s", ('%restock%',))
            check("Restock email sent", cur.fetchone() is not None)
            conn.close()
        except Exception as e:
            check("DB checks", False, str(e))
        check("inventory_forecaster.py exists", os.path.exists(os.path.join(agent_workspace, "inventory_forecaster.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

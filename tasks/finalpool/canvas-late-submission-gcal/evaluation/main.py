"""Evaluation for canvas-late-submission-gcal."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": "toolathlon_gym", "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def int_close(a, b, tol=5):
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Late_Submissions.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Late_Submissions.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Late_Submissions.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # Sheet 1: By Course
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    agent_ws = get_sheet(agent_wb, "By Course")
    gt_ws = get_sheet(gt_wb, "By Course")

    if agent_ws is None:
        record("Sheet 'By Course' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'By Course' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("By Course row count", len(agent_rows) >= len(gt_rows) - 2,
               f"Expected ~{len(gt_rows)}, got {len(agent_rows)}")

        # Check top 5 courses by late count
        gt_lookup = {}
        for r in gt_rows:
            if r and r[0]:
                gt_lookup[str(r[0]).strip().lower()] = r

        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for gt_row in gt_rows[:5]:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Course '{gt_row[0]}' present", False, "Missing")
                all_ok = False
            else:
                ok = int_close(a_row[1], gt_row[1], 50) and int_close(a_row[2], gt_row[2], 20)
                record(f"Course '{gt_row[0]}' data", ok,
                       f"Late: {a_row[1]} vs {gt_row[1]}, Students: {a_row[2]} vs {gt_row[2]}")
                if not ok:
                    all_ok = False

    # Sheet 2: Top Offenders
    agent_ws2 = get_sheet(agent_wb, "Top Offenders")
    gt_ws2 = get_sheet(gt_wb, "Top Offenders")

    if agent_ws2 is None:
        record("Sheet 'Top Offenders' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Top Offenders' exists", True)
        agent_rows2 = list(agent_ws2.iter_rows(min_row=2, values_only=True))
        gt_rows2 = list(gt_ws2.iter_rows(min_row=2, values_only=True))

        record("Top Offenders has ~10 rows", 8 <= len(agent_rows2) <= 12,
               f"Got {len(agent_rows2)}")

        # Check top 3 offenders
        for i, gt_row in enumerate(gt_rows2[:3]):
            if not gt_row or not gt_row[0]:
                continue
            found = False
            for a_row in agent_rows2:
                if a_row and a_row[0] and str(a_row[0]).strip().lower() == str(gt_row[0]).strip().lower():
                    ok = int_close(a_row[2], gt_row[2], 3)
                    record(f"Top offender '{gt_row[0]}'", ok,
                           f"Late: {a_row[2]} vs {gt_row[2]}")
                    if not ok:
                        all_ok = False
                    found = True
                    break
            if not found:
                record(f"Top offender '{gt_row[0]}' present", False, "Missing")
                all_ok = False

    return all_ok


def check_gcal():
    """Check calendar events for review meetings."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description FROM gcal.events ORDER BY summary")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")

    # Should have at least some review meetings (22 courses with late submissions)
    record("At least 10 calendar events created", len(events) >= 10,
           f"Found {len(events)}")

    review_events = [e for e in events if "review meeting" in (e[0] or "").lower()]
    record("Review Meeting events found", len(review_events) >= 10,
           f"Found {len(review_events)} review meeting events")

    # Check that top 3 courses are represented
    top_courses = [
        "Biochemistry & Bioinformatics (Fall 2013)",
        "Creative Computing & Culture (Fall 2014)",
        "Biochemistry & Bioinformatics (Spring 2013)",
    ]

    for course in top_courses:
        found = any(course.lower() in (e[0] or "").lower() for e in events)
        record(f"Calendar event for '{course[:40]}...'", found)

    return len(review_events) >= 10


def check_emails():
    """Check that summary email was sent."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    found_report = False
    for subject, from_addr, to_addr, body in all_emails:
        subj_lower = (subject or "").lower()
        if "late" in subj_lower and "submission" in subj_lower:
            found_report = True
            to_str = str(to_addr or "").lower()
            record("Email sent to academic.affairs@university.example.com",
                   "academic.affairs@university.example.com" in to_str,
                   f"To: {to_addr}")

            body_lower = (body or "").lower()
            record("Email mentions course count",
                   "22" in body_lower or "course" in body_lower)
            break

    record("Late Submission Report email found", found_report)
    return found_report


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

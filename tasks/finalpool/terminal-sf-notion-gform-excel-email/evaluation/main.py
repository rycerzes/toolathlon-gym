"""Evaluation for terminal-sf-notion-gform-excel-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.05):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Employee_Engagement_Report.xlsx")
    gt_path = os.path.join(gt_workspace, "Employee_Engagement_Report.xlsx")

    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)

    # Sheet 1: Department_Scores
    ds_sheet = None
    for name in wb.sheetnames:
        if "department" in name.lower() and "score" in name.lower():
            ds_sheet = name
            break
    if not ds_sheet:
        record("Department_Scores sheet exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Department_Scores sheet exists", True)

    ws = wb[ds_sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    record("Department_Scores has 7 rows", len(rows) == 7, f"Got {len(rows)}")

    gt_ws = gt_wb["Department_Scores"]
    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

    a_lookup = {}
    for row in rows:
        if row and row[0]:
            a_lookup[str(row[0]).strip().lower()] = row

    for gt_row in gt_rows:
        if not gt_row or not gt_row[0]:
            continue
        key = str(gt_row[0]).strip().lower()
        a_row = a_lookup.get(key)
        if a_row is None:
            record(f"Department {gt_row[0]} exists", False)
            continue
        # Check Avg_Satisfaction (idx 1)
        record(f"{gt_row[0]} Avg_Satisfaction", num_close(a_row[1], gt_row[1]),
               f"Got {a_row[1]} vs {gt_row[1]}")
        # Check Engagement_Index (idx 5)
        if len(a_row) > 5 and len(gt_row) > 5:
            record(f"{gt_row[0]} Engagement_Index", num_close(a_row[5], gt_row[5], 0.1),
                   f"Got {a_row[5]} vs {gt_row[5]}")

    # Sheet 2: Survey_Design
    sd_sheet = None
    for name in wb.sheetnames:
        if "survey" in name.lower():
            sd_sheet = name
            break
    if not sd_sheet:
        record("Survey_Design sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Survey_Design sheet exists", True)
        ws2 = wb[sd_sheet]
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        record("Survey_Design has 5 rows", len(rows2) == 5, f"Got {len(rows2)}")

    # Sheet 3: Action_Items
    ai_sheet = None
    for name in wb.sheetnames:
        if "action" in name.lower():
            ai_sheet = name
            break
    if not ai_sheet:
        record("Action_Items sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Action_Items sheet exists", True)
        ws3 = wb[ai_sheet]
        rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        record("Action_Items has 7 rows", len(rows3) == 7, f"Got {len(rows3)}")
        # Check priorities exist
        priorities = [str(r[0]).strip() for r in rows3 if r and r[0]]
        has_high = any("high" in p.lower() for p in priorities)
        has_medium = any("medium" in p.lower() for p in priorities)
        record("Action_Items has High priority items", has_high)
        record("Action_Items has Medium priority items", has_medium)

    wb.close()
    gt_wb.close()
    return True


def check_terminal_output(agent_workspace):
    print("\n=== Checking Terminal Output ===")
    fpath = os.path.join(agent_workspace, "engagement_analysis_output.txt")
    if not os.path.isfile(fpath):
        record("engagement_analysis_output.txt exists", False)
        return False
    record("engagement_analysis_output.txt exists", True)
    with open(fpath) as f:
        content = f.read().lower()
    record("Output mentions highest department", "highest" in content or "best" in content or "top" in content,
           f"Preview: {content[:200]}")
    record("Output mentions lowest department", "lowest" in content or "worst" in content or "bottom" in content,
           f"Preview: {content[:200]}")
    return True


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gform.forms WHERE LOWER(title) LIKE '%employee engagement%'")
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        record("Employee Engagement Survey form exists", count >= 1, f"Found {count}")
        return count >= 1
    except Exception as e:
        record("GForm check", False, str(e))
        return False


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE properties::text ILIKE '%engagement%' OR properties::text ILIKE '%HR%Dashboard%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        record("Notion HR Dashboard page exists", count >= 1, f"Found {count}")
        return count >= 1
    except Exception as e:
        record("Notion check", False, str(e))
        return False


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%hr-leadership@company.com%'
            AND LOWER(subject) LIKE '%engagement%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        record("Email sent to hr-leadership@company.com about engagement", count >= 1, f"Found {count}")
        return count >= 1
    except Exception as e:
        record("Email check", False, str(e))
        return False


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no unexpected sheets beyond the 3 required
    path = os.path.join(workspace, "Employee_Engagement_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        record("Excel has no more than 5 sheets", len(wb.sheetnames) <= 5,
               f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        # No negative satisfaction or engagement values
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        record("No negative values in Excel", not has_negative,
               "Found negative engagement/satisfaction value")

    # Notion: no duplicate HR Dashboard pages
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE properties::text ILIKE '%%engagement%%' AND properties::text ILIKE '%%HR%%Dashboard%%'
        """)
        count = cur.fetchone()[0]
        record("No duplicate Notion HR Dashboard pages", count <= 1,
               f"Found {count} matching pages")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_terminal_output(args.agent_workspace)
    check_gform()
    check_notion()
    check_email()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

"""
Evaluation script for yf-sector-analysis-word task.

Checks:
1. Word document (sector_analysis_report.docx) - correct structure and content
2. Excel file (sector_analysis_data.xlsx) - correct data in both sheets

Since Yahoo Finance data is live, we verify structure, presence of tickers,
and internal consistency rather than exact values.
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from docx import Document
except ImportError:
    Document = None

PASS_COUNT = 0
FAIL_COUNT = 0

TICKERS = ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


# ============================================================================
# Check 1: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")

    doc_path = os.path.join(agent_workspace, "sector_analysis_report.docx")

    if not os.path.isfile(doc_path):
        record("Word file exists", False, f"Not found: {doc_path}")
        return False
    record("Word file exists", True)

    if Document is None:
        record("python-docx installed", False, "Cannot import docx")
        return False

    doc = Document(doc_path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    record("Word: has title 'cross-sector'",
           "cross-sector" in full_text,
           "Title not found")
    record("Word: mentions 2026-03-06",
           "2026-03-06" in full_text,
           "Date not found")

    for ticker in TICKERS:
        record(f"Word: mentions {ticker}",
               ticker.lower() in full_text,
               f"{ticker} not found in document")

    record("Word: has Summary section",
           "summary" in full_text,
           "Summary section not found")

    record("Word: mentions 'market cap'",
           "market cap" in full_text,
           "market cap not mentioned")

    return True


# ============================================================================
# Check 2: Excel file
# ============================================================================

def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")

    xlsx_path = os.path.join(agent_workspace, "sector_analysis_data.xlsx")

    if not os.path.isfile(xlsx_path):
        record("Excel file exists", False, f"Not found: {xlsx_path}")
        return False
    record("Excel file exists", True)

    if openpyxl is None:
        record("openpyxl installed", False, "Cannot import openpyxl")
        return False

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Sheet 1: Stock Data
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    ws1 = get_sheet(wb, "Stock Data")
    if ws1 is None:
        record("Sheet 'Stock Data' exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Sheet 'Stock Data' exists", True)

    headers = [str(c.value).strip() if c.value else "" for c in ws1[1]]
    expected_headers = ["Ticker", "Company_Name", "Current_Price", "Market_Cap",
                        "Week52_High", "Week52_Low", "Sector"]
    headers_ok = all(str_match(h, e) for h, e in zip(headers, expected_headers))
    record("Stock Data headers match", headers_ok,
           f"Expected: {expected_headers}, Got: {headers}")

    rows = list(ws1.iter_rows(min_row=2, values_only=True))
    record("Stock Data has 5 rows", len(rows) == 5, f"Got {len(rows)}")

    # Check tickers present and sorted
    agent_tickers = [str(r[0]).strip().upper() for r in rows if r and r[0]]
    record("Stock Data tickers sorted",
           agent_tickers == sorted(agent_tickers),
           f"Got: {agent_tickers}")

    for ticker in TICKERS:
        record(f"Stock Data: {ticker} present",
               ticker in agent_tickers,
               f"Missing {ticker}")

    # Check prices are positive numbers
    for row in rows:
        if row and row[0]:
            ticker = str(row[0]).strip().upper()
            price = row[2]
            try:
                assert float(price) > 0
                record(f"Stock Data: {ticker} price > 0", True)
            except (TypeError, ValueError, AssertionError):
                record(f"Stock Data: {ticker} price > 0", False, f"Got: {price}")

    # Internal consistency: check market cap values are large
    for row in rows:
        if row and row[0]:
            ticker = str(row[0]).strip().upper()
            mcap = row[3]
            try:
                assert float(mcap) > 1e9
                record(f"Stock Data: {ticker} market cap > 1B", True)
            except (TypeError, ValueError, AssertionError):
                record(f"Stock Data: {ticker} market cap > 1B", False, f"Got: {mcap}")

    # Sheet 2: Sector Summary
    ws2 = get_sheet(wb, "Sector Summary")
    if ws2 is None:
        record("Sheet 'Sector Summary' exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Sheet 'Sector Summary' exists", True)

    summary = {}
    for row in ws2.iter_rows(min_row=1, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip().lower()] = row[1]

    record("Summary: Total_Stocks = 5",
           str(summary.get("total_stocks", "")).strip() == "5",
           f"Got {summary.get('total_stocks')}")

    hmc_ticker = str(summary.get("highest_market_cap_ticker", "")).strip().upper()
    record("Summary: Highest_Market_Cap_Ticker in known tickers",
           hmc_ticker in TICKERS,
           f"Got {hmc_ticker}")

    lp_ticker = str(summary.get("lowest_price_ticker", "")).strip().upper()
    record("Summary: Lowest_Price_Ticker in known tickers",
           lp_ticker in TICKERS,
           f"Got {lp_ticker}")

    # Cross-check: highest market cap ticker matches stock data
    if hmc_ticker in agent_tickers:
        hmc_row = [r for r in rows if str(r[0]).strip().upper() == hmc_ticker]
        if hmc_row:
            all_mcaps = [(str(r[0]).strip().upper(), float(r[3])) for r in rows if r and r[3]]
            max_mcap = max(all_mcaps, key=lambda x: x[1])
            record("Summary: Highest Market Cap cross-check",
                   max_mcap[0] == hmc_ticker,
                   f"Max mcap is {max_mcap[0]} but summary says {hmc_ticker}")

    # Cross-check: lowest price ticker matches stock data
    if lp_ticker in agent_tickers:
        all_prices = [(str(r[0]).strip().upper(), float(r[2])) for r in rows if r and r[2]]
        min_price = min(all_prices, key=lambda x: x[1])
        record("Summary: Lowest Price cross-check",
               min_price[0] == lp_ticker,
               f"Min price is {min_price[0]} but summary says {lp_ticker}")

    # Check average price consistency
    avg_price = summary.get("average_current_price")
    if avg_price is not None:
        computed_avg = sum(float(r[2]) for r in rows if r and r[2]) / 5
        record("Summary: Average Price consistent",
               num_close(avg_price, computed_avg, 1.0),
               f"Summary says {avg_price}, computed {computed_avg:.2f}")

    return True


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    word_ok = check_word(args.agent_workspace)
    excel_ok = check_excel(args.agent_workspace)

    all_passed = word_ok and excel_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

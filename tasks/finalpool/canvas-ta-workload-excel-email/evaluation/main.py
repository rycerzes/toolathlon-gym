"""Evaluation for canvas-ta-workload-excel-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "TA_Workload_Report.xlsx")
    if not os.path.isfile(xlsx_path):
        check("TA_Workload_Report.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("TA_Workload_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # Check Course Workload sheet
    cw_rows = load_sheet_rows(wb, "Course Workload")
    if cw_rows is None:
        check("Sheet 'Course Workload' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Course Workload' exists", True)
        data_rows = cw_rows[1:] if len(cw_rows) > 1 else []
        check("Course Workload has 22 data rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        # Check header columns
        header = cw_rows[0] if cw_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col_name in ["course_name", "course_code", "ta_count", "assignment_count", "submission_count", "submissions_per_ta"]:
            check(f"Column '{col_name}' present", any(col_name in h for h in header_lower),
                  f"Header: {header}")

        # Verify some key data - Foundations of Finance (Fall 2013) has most submissions
        found_fff = False
        for row in data_rows:
            if row and row[0] and "foundations of finance (fall 2013)" in str(row[0]).lower():
                found_fff = True
                ta_count = row[2] if len(row) > 2 else None
                submission_count = row[4] if len(row) > 4 else None
                check("FFF-2013J has TA_Count=1", num_close(ta_count, 1, 0), f"Got {ta_count}")
                check("FFF-2013J has Submission_Count=16240", num_close(submission_count, 16240, 10), f"Got {submission_count}")
        check("Foundations of Finance (Fall 2013) row found", found_fff)

    # Check Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        check("Total_Courses = 22", num_close(lookup.get("total_courses"), 22, 0),
              f"Got {lookup.get('total_courses')}")
        check("Total_TAs = 29", num_close(lookup.get("total_tas"), 29, 0),
              f"Got {lookup.get('total_tas')}")
        check("Max_Assignment_Count = 14", num_close(lookup.get("max_assignment_count"), 14, 0),
              f"Got {lookup.get('max_assignment_count')}")
        most_loaded = lookup.get("most_loaded_course")
        check("Most_Loaded_Course is Foundations of Finance (Fall 2013)",
              most_loaded and "foundations of finance (fall 2013)" in str(most_loaded).lower(),
              f"Got: {most_loaded}")


def check_notion():
    print("\n=== Checking Notion Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Notion pages store title in properties jsonb
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()
        found_page = None
        for page_id, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "ta staffing" in props_str or "staffing overview" in props_str:
                found_page = page_id
                break
        check("Notion page 'TA Staffing Overview' exists", found_page is not None,
              f"Found {len(pages)} pages, none matching 'TA Staffing'")
        if found_page:
            # Check blocks
            cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s", (found_page,))
            block_count = cur.fetchone()[0]
            check("Notion page has content (at least 1 block)", block_count >= 1,
                  f"Found {block_count} blocks")
        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%dept_chair@university.edu%%'
               OR subject ILIKE '%%ta workload%%'
               OR subject ILIKE '%%ta%%report%%'
        """)
        emails = cur.fetchall()
        check("Email sent to dept_chair@university.edu", len(emails) >= 1,
              "No matching email found")
        if emails:
            email = emails[0]
            subject = str(email[1]).lower() if email[1] else ""
            check("Email subject contains 'workload' or 'TA'",
                  "workload" in subject or "ta" in subject or "report" in subject,
                  f"Subject: {email[1]}")
            body = str(email[3]) if email[3] else ""
            check("Email body has content", len(body) > 20,
                  f"Body length: {len(body)}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

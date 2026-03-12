"""
Evaluation script for sf-hr-salary-review-gcal task.

Checks:
1. Excel file with salary band data (values verified against DB with tolerance)
2. At least 7 gcal events with "salary review" in summary
3. Email sent with correct subject
"""

import argparse
import json
import os
import sys

import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_expected_stats():
    """Query actual salary stats from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT "DEPARTMENT",
               MIN("SALARY"),
               PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY "SALARY"),
               PERCENTILE_CONT(0.50) WITHIN GROUP (ORDER BY "SALARY"),
               PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY "SALARY"),
               MAX("SALARY"),
               AVG("SALARY"),
               COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    dept_stats = cur.fetchall()

    cur.execute('SELECT COUNT(*) FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    total_emp = cur.fetchone()[0]

    cur.execute("""
        SELECT "DEPARTMENT", AVG("SALARY") as avg_sal
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY avg_sal DESC
    """)
    dept_avgs = cur.fetchall()
    highest_avg_dept = dept_avgs[0][0]
    lowest_avg_dept = dept_avgs[-1][0]

    cur.execute('SELECT AVG("SALARY") FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    overall_avg = float(cur.fetchone()[0])

    cur.close()
    conn.close()

    return dept_stats, total_emp, highest_avg_dept, lowest_avg_dept, overall_avg


def check_excel(workspace):
    """Check Excel file against DB values."""
    from openpyxl import load_workbook

    errors = []
    xlsx_path = os.path.join(workspace, "Salary_Band_Analysis.xlsx")
    if not os.path.exists(xlsx_path):
        return ["Salary_Band_Analysis.xlsx not found"]

    dept_stats, total_emp, highest_avg_dept, lowest_avg_dept, overall_avg = get_expected_stats()

    wb = load_workbook(xlsx_path)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Department Bands sheet
    if "department bands" not in sheet_names_lower:
        errors.append(f"Missing 'Department Bands' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("department bands")]]
        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        if data_rows < 7:
            errors.append(f"Department Bands has {data_rows} rows, expected 7")

        # Check headers
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]
        for rh in ["department", "min_salary", "avg_salary", "employee_count"]:
            if not any(rh in h or rh.replace("_", "") in h.replace("_", "") for h in headers):
                errors.append(f"Department Bands missing header: {rh}")

        # Verify some values with tolerance
        dept_col = None
        avg_col = None
        count_col = None
        for idx, h in enumerate(headers):
            if "department" in h:
                dept_col = idx
            if "avg" in h and "salary" in h:
                avg_col = idx
            if "count" in h or "employee" in h:
                count_col = idx

        if dept_col is not None and avg_col is not None:
            for row in ws.iter_rows(min_row=2):
                if row[dept_col].value is None:
                    continue
                dept_name = str(row[dept_col].value).strip()
                for expected in dept_stats:
                    if expected[0].lower() == dept_name.lower():
                        expected_avg = float(expected[6])
                        try:
                            actual_avg = float(row[avg_col].value)
                            if abs(actual_avg - expected_avg) / expected_avg > 0.05:
                                errors.append(f"Avg salary for {dept_name}: got {actual_avg:.0f}, expected ~{expected_avg:.0f}")
                        except (TypeError, ValueError):
                            errors.append(f"Cannot parse avg salary for {dept_name}: {row[avg_col].value}")
                        break

    # Check Summary sheet
    if "summary" not in sheet_names_lower:
        errors.append(f"Missing 'Summary' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        summary_data = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                key = str(row[0].value).lower().replace(" ", "_")
                summary_data[key] = row[1].value

        # Check total employees
        total_key = None
        for k in summary_data:
            if "total" in k and "emp" in k:
                total_key = k
                break
        if total_key:
            try:
                val = int(float(summary_data[total_key]))
                if val != total_emp:
                    errors.append(f"Total_Employees: got {val}, expected {total_emp}")
            except (TypeError, ValueError):
                errors.append(f"Cannot parse Total_Employees: {summary_data[total_key]}")
        else:
            errors.append("Summary missing Total_Employees row")

    return errors


def check_gcal(cur):
    """Check for salary review calendar events."""
    errors = []
    cur.execute("""
        SELECT id, summary, start_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%salary review%'
        AND start_datetime >= '2026-03-10T00:00:00'
        AND start_datetime < '2026-03-17T00:00:00'
    """)
    events = cur.fetchall()

    if len(events) < 7:
        errors.append(f"Found {len(events)} salary review events, expected at least 7")

    return errors


def check_email(cur):
    """Check for the summary email."""
    errors = []
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%quarterly salary band analysis%'
    """)
    emails = cur.fetchall()

    if not emails:
        errors.append("No email with subject 'Quarterly Salary Band Analysis' found")
    else:
        email = emails[0]
        to_str = str(email[2]).lower()
        if "hr-director@company.com" not in to_str:
            errors.append(f"Email not sent to hr-director@company.com, to_addr: {email[2]}")

    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    # Check Excel
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    # Check GCal and Email
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        print("\n=== Checking GCal Events ===")
        gcal_errors = check_gcal(cur)
        if gcal_errors:
            for e in gcal_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(gcal_errors)
        else:
            print("  [PASS] GCal check passed")

        print("\n=== Checking Email ===")
        email_errors = check_email(cur)
        if email_errors:
            for e in email_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(email_errors)
        else:
            print("  [PASS] Email check passed")

        cur.close()
        conn.close()
    except Exception as e:
        err = f"DB check error: {e}"
        print(f"  [FAIL] {err}")
        all_errors.append(err)

    # Summary
    print(f"\n=== SUMMARY ===")
    if all_errors:
        for e in all_errors:
            print(f"  [ERROR] {e}")
        print("  Overall: FAIL")
    else:
        print("  Overall: PASS")

    if args.res_log_file:
        result = {"errors": all_errors, "success": len(all_errors) == 0}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(1 if all_errors else 0)


if __name__ == "__main__":
    main()

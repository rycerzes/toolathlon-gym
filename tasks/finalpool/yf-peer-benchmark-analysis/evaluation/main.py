"""Evaluation for yf-peer-benchmark-analysis."""
import os
import argparse, os, sys
import psycopg2


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Peer_Benchmark.xlsx")
    if not os.path.exists(path):
        return ["Peer_Benchmark.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        rows = load_sheet_rows(wb, "Consensus Analysis")
        if rows is None:
            errors.append("Sheet 'Consensus Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 5:
                errors.append(f"Consensus Analysis has {len(data_rows)} rows, expected 5")
            symbols = {str(r[0]).strip().upper() for r in data_rows if r[0]}
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                if sym not in symbols:
                    errors.append(f"Symbol {sym} missing from Consensus Analysis")
            # Check AMZN upside (should be ~0.48)
            for r in data_rows:
                if r[0] and str(r[0]).strip().upper() == "AMZN":
                    if len(r) > 3 and not num_close(r[3], 0.48, abs_tol=1.0):
                        errors.append(f"AMZN Upside_Pct={r[3]}, expected ~0.48")
                if r[0] and str(r[0]).strip().upper() == "GOOGL":
                    if len(r) > 3 and not num_close(r[3], -35.19, abs_tol=2.0):
                        errors.append(f"GOOGL Upside_Pct={r[3]}, expected ~-35.19")

        rows2 = load_sheet_rows(wb, "Summary")
        if rows2 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows2 if r[0]}
            if "avg_upside_pct" in lookup:
                if not num_close(lookup["avg_upside_pct"], -23.19):
                    errors.append(f"Avg_Upside_Pct={lookup['avg_upside_pct']}, expected ~-23.19")
            else:
                errors.append("Avg_Upside_Pct not found")
            if "stocks_above_target" in lookup:
                if not num_close(lookup["stocks_above_target"], 4, abs_tol=0):
                    errors.append(f"Stocks_Above_Target={lookup['stocks_above_target']}, expected 4")
            if "most_undervalued" in lookup:
                if str(lookup["most_undervalued"]).strip().upper() != "AMZN":
                    errors.append(f"Most_Undervalued={lookup['most_undervalued']}, expected AMZN")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gsheet():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gsheet.spreadsheets")
        count = cur.fetchone()[0]
        if count == 0:
            errors.append("No Google Sheet spreadsheet created")
        else:
            cur.execute("SELECT COUNT(*) FROM gsheet.cells")
            cell_count = cur.fetchone()[0]
            if cell_count < 5:
                errors.append(f"Google Sheet has only {cell_count} cells, expected at least 5")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking GSheet: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Google Sheet...")
    errs = check_gsheet()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

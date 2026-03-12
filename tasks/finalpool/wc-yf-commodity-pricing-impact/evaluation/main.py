"""Evaluation for wc-yf-commodity-pricing-impact."""
import argparse
import os
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Commodity_Impact.xlsx")
    if not os.path.exists(path):
        return ["Commodity_Impact.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Gold Price Trend sheet
        rows = load_sheet_rows(wb, "Gold Price Trend")
        if rows is None:
            errors.append("Sheet 'Gold Price Trend' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 25:
                errors.append(f"Gold Price Trend has {len(data_rows)} rows, expected ~30")
            else:
                # Check that gold prices are reasonable (between 4000 and 6000)
                prices_ok = 0
                for r in data_rows:
                    if r[1] and 4000 < float(r[1]) < 6000:
                        prices_ok += 1
                if prices_ok < 20:
                    errors.append(f"Only {prices_ok} gold prices in expected range 4000-6000")
                # Check last row has recent date
                last_row = data_rows[-1]
                if last_row[0] and "2026" in str(last_row[0]):
                    pass
                else:
                    errors.append(f"Last date does not contain 2026: {last_row[0]}")
                # Check Trend_Direction column exists
                has_direction = any(r[3] and str(r[3]).strip().lower() in ("up", "down") for r in data_rows if len(r) > 3)
                if not has_direction:
                    errors.append("Trend_Direction column missing or has no Up/Down values")

        # Check Category Analysis sheet
        rows2 = load_sheet_rows(wb, "Category Analysis")
        if rows2 is None:
            errors.append("Sheet 'Category Analysis' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(f"Category Analysis has {len(data_rows2)} rows, expected ~7")
            # Check Watches row exists
            watches_rows = [r for r in data_rows2 if r[0] and "watch" in str(r[0]).lower()]
            if not watches_rows:
                errors.append("Watches category not found in Category Analysis")
            else:
                # Watches avg price should be ~60.69
                if len(watches_rows[0]) > 1 and watches_rows[0][1]:
                    if not nums_close(watches_rows[0][1], 60.69, abs_tol=5.0):
                        errors.append(f"Watches avg price {watches_rows[0][1]}, expected ~60.69")
            # Check Electronics row
            electronics_rows = [r for r in data_rows2 if r[0] and "electronics" in str(r[0]).lower()]
            if not electronics_rows:
                errors.append("Electronics category not found")
            else:
                if len(electronics_rows[0]) > 2 and electronics_rows[0][2]:
                    if not nums_close(electronics_rows[0][2], 30, abs_tol=3.0):
                        errors.append(f"Electronics product count {electronics_rows[0][2]}, expected ~30")

        # Check Correlation Summary sheet
        rows3 = load_sheet_rows(wb, "Correlation Summary")
        if rows3 is None:
            errors.append("Sheet 'Correlation Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 3:
                errors.append(f"Correlation Summary has {len(data_rows3)} rows, expected >= 3")
            # Check gold price is present
            gold_row = [r for r in data_rows3 if r[0] and "gold" in str(r[0]).lower() and "price" in str(r[0]).lower()]
            if gold_row and len(gold_row[0]) > 1 and gold_row[0][1]:
                if not nums_close(gold_row[0][1], 5093.30, abs_tol=100.0):
                    errors.append(f"Current gold price {gold_row[0][1]}, expected ~5093")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    from docx import Document
    path = os.path.join(agent_workspace, "Commodity_Report.docx")
    if not os.path.exists(path):
        return ["Commodity_Report.docx not found"]
    try:
        doc = Document(path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        if len(full_text) < 200:
            errors.append(f"Word doc too short ({len(full_text)} chars)")
        if "gold" not in full_text:
            errors.append("Word doc does not mention 'gold'")
        if "watch" not in full_text:
            errors.append("Word doc does not mention 'watches'")
        if "margin" not in full_text:
            errors.append("Word doc does not mention 'margin'")
        if "trend" not in full_text and "moving" not in full_text:
            errors.append("Word doc does not mention price trend or moving average")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

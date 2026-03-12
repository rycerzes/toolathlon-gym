"""Evaluation for yf-sector-comparison-pdf."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sector_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Sector_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check sheet: Stock Prices
    print("  Checking Stock Prices...")
    a_rows = load_sheet_rows(agent_wb, "Stock Prices")
    g_rows = load_sheet_rows(gt_wb, "Stock Prices")
    if a_rows is None:
        all_errors.append("Sheet 'Stock Prices' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Stock Prices' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row: {g_row[0]}")
                continue

            # Col 1: Sector (string)
            if len(a_row) > 1 and len(g_row) > 1:
                if not str_match(a_row[1], g_row[1]):
                    all_errors.append(f"{key}.Sector: {a_row[1]} vs {g_row[1]}")

            # Col 2: Latest_Close (numeric, tol=2.0)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 2.0):
                    all_errors.append(f"{key}.Latest_Close: {a_row[2]} vs {g_row[2]} (tol=2.0)")

            # Col 3: YTD_Return_Pct (numeric, tol=1.0)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1.0):
                    all_errors.append(f"{key}.YTD_Return_Pct: {a_row[3]} vs {g_row[3]} (tol=1.0)")
        if not all_errors:
            print("    PASS")
        else:
            print(f"    ERRORS: {len(all_errors)}")

    # Check sheet: Sector Comparison
    print("  Checking Sector Comparison...")
    a_rows = load_sheet_rows(agent_wb, "Sector Comparison")
    g_rows = load_sheet_rows(gt_wb, "Sector Comparison")
    prev_errors = len(all_errors)
    if a_rows is None:
        all_errors.append("Sheet 'Sector Comparison' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Sector Comparison' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Match by Stock column (col 1)
        a_lookup = {}
        for row in a_data:
            if row and len(row) > 1 and row[1] is not None:
                a_lookup[str(row[1]).strip().upper()] = row
        for g_row in g_data:
            if not g_row or len(g_row) < 2 or g_row[1] is None:
                continue
            key = str(g_row[1]).strip().upper()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Sector Comparison missing row for stock: {key}")
                continue

            # Col 0: Sector
            if not str_match(a_row[0], g_row[0]):
                all_errors.append(f"{key}.Sector: {a_row[0]} vs {g_row[0]}")

            # Col 2: Target_Return
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    all_errors.append(f"{key}.Target_Return: {a_row[2]} vs {g_row[2]}")

            # Col 3: Actual_Return
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1.0):
                    all_errors.append(f"{key}.Actual_Return: {a_row[3]} vs {g_row[3]}")

            # Col 4: Met_Target
            if len(a_row) > 4 and len(g_row) > 4:
                if not str_match(a_row[4], g_row[4]):
                    all_errors.append(f"{key}.Met_Target: {a_row[4]} vs {g_row[4]}")

        new_errors = len(all_errors) - prev_errors
        if new_errors == 0:
            print("    PASS")
        else:
            print(f"    ERRORS: {new_errors}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

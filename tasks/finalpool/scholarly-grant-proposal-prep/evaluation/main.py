"""
Evaluation for scholarly-grant-proposal-prep task.
Checks Excel, emails sent, and calendar events.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

AGENCIES = ["NSF", "DARPA", "NIH"]
COLLABORATOR_EMAILS = ["jpark@mit.edu", "sthompson@stanford.edu", "mchen@mayo.edu", "lwang@berkeley.edu"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Grant_Prep.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Funding Opportunities
    fo_rows = load_sheet_rows(wb, "Funding Opportunities") or load_sheet_rows(wb, "Funding_Opportunities")
    if fo_rows is None:
        record("Sheet 'Funding Opportunities' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Funding Opportunities' exists", True)
        data = fo_rows[1:]
        record("Funding Opportunities has 3 rows", len(data) == 3, f"Found {len(data)}")

        agency_col = find_col(fo_rows[0], ["Agency", "agency"])
        if agency_col is not None:
            agencies = {str(r[agency_col]).strip() for r in data if agency_col < len(r) and r[agency_col]}
            for a in AGENCIES:
                record(f"Agency '{a}' present", a in agencies, f"Found: {agencies}")

        score_col = find_col(fo_rows[0], ["Relevance_Score", "Relevance Score", "Score"])
        if score_col is not None:
            for r in data:
                if score_col < len(r) and r[score_col] is not None:
                    try:
                        s = float(r[score_col])
                        record(f"Score {s} in range 1-10", 1 <= s <= 10, f"Got {s}")
                    except (TypeError, ValueError):
                        record("Score is numeric", False, f"Got {r[score_col]}")

    # Supporting Literature
    sl_rows = load_sheet_rows(wb, "Supporting Literature") or load_sheet_rows(wb, "Supporting_Literature")
    if sl_rows is None:
        record("Sheet 'Supporting Literature' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Supporting Literature' exists", True)
        data = sl_rows[1:]
        record("Supporting Literature has >= 5 rows", len(data) >= 5, f"Found {len(data)}")

    # Collaborator Matrix
    cm_rows = load_sheet_rows(wb, "Collaborator Matrix") or load_sheet_rows(wb, "Collaborator_Matrix")
    if cm_rows is None:
        record("Sheet 'Collaborator Matrix' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Collaborator Matrix' exists", True)
        data = cm_rows[1:]
        record("Collaborator Matrix has 4 rows", len(data) == 4, f"Found {len(data)}")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Grant_Prep.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()

    return True


def check_emails():
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%%grant%%' OR subject ILIKE '%%collaboration%%'
           OR subject ILIKE '%%nsf%%' OR subject ILIKE '%%darpa%%' OR subject ILIKE '%%nih%%'
           OR subject ILIKE '%%proposal%%' OR subject ILIKE '%%funding%%'
    """)
    emails = cur.fetchall()

    record("Grant-related emails sent", len(emails) >= 3, f"Found {len(emails)} emails")

    if emails:
        all_to = []
        for e in emails:
            to = e[3]
            if isinstance(to, str):
                try:
                    to = json.loads(to)
                except Exception:
                    pass
            all_to.append(str(to).lower())

        all_to_str = " ".join(all_to)
        found_count = sum(1 for addr in COLLABORATOR_EMAILS if addr in all_to_str)
        record(f"Emails to collaborators ({found_count}/4)", found_count >= 3,
               f"Found {found_count} collaborator emails")

    cur.close()
    conn.close()
    return True


def check_calendar():
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime FROM gcal.events")
    events = cur.fetchall()

    record("Calendar events created", len(events) >= 3, f"Found {len(events)} events")

    if events:
        summaries = [str(e[1]).lower() for e in events if e[1]]
        has_deadline = any("deadline" in s or "nsf" in s or "darpa" in s or "nih" in s for s in summaries)
        record("Deadline events present", has_deadline, f"Summaries: {summaries[:5]}")

        has_writing = any("writing" in s or "proposal" in s for s in summaries)
        record("Writing session events present", has_writing, f"Summaries: {summaries[:5]}")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_emails()
    check_calendar()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

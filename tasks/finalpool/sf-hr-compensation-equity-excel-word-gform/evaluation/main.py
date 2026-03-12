"""Evaluation for sf-hr-compensation-equity-excel-word-gform."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Compensation_Equity.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Equity.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Compensation_Equity.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # --- Sheet: Salary Analysis ---
        print("  Checking Salary Analysis sheet...")
        a_rows = load_sheet_rows(agent_wb, "Salary Analysis")
        g_rows = load_sheet_rows(gt_wb, "Salary Analysis")
        if a_rows is None:
            all_errors.append("Sheet 'Salary Analysis' not found in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Salary Analysis' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            # Build lookup: (dept, edu_level, exp_band) -> row
            a_lookup = {}
            for row in a_data:
                if row and row[0] and row[1] and row[2]:
                    key = (str(row[0]).strip().lower(), str(row[1]).strip().lower(), str(row[2]).strip().lower())
                    a_lookup[key] = row

            errors = []
            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower(), str(g_row[2]).strip().lower())
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing row: {g_row[0]}/{g_row[1]}/{g_row[2]}")
                    continue
                # Employee_Count (col 3)
                if not num_close(a_row[3], g_row[3], 1):
                    errors.append(f"{key} Employee_Count: {a_row[3]} vs {g_row[3]}")
                # Avg_Salary (col 4)
                if not num_close(a_row[4], g_row[4], 5.0):
                    errors.append(f"{key} Avg_Salary: {a_row[4]} vs {g_row[4]}")
                # Median_Salary (col 5)
                if not num_close(a_row[5], g_row[5], 5.0):
                    errors.append(f"{key} Median_Salary: {a_row[5]} vs {g_row[5]}")
                # Min_Salary (col 6)
                if not num_close(a_row[6], g_row[6], 5.0):
                    errors.append(f"{key} Min_Salary: {a_row[6]} vs {g_row[6]}")
                # Max_Salary (col 7)
                if not num_close(a_row[7], g_row[7], 5.0):
                    errors.append(f"{key} Max_Salary: {a_row[7]} vs {g_row[7]}")
                # Salary_Std_Dev (col 8)
                if not num_close(a_row[8], g_row[8], 5.0):
                    errors.append(f"{key} Salary_Std_Dev: {a_row[8]} vs {g_row[8]}")

            if errors:
                all_errors.extend(errors)
                print(f"    ERRORS: {len(errors)}")
                for e in errors[:5]:
                    print(f"      {e}")
            else:
                print("    PASS")

        # --- Sheet: Equity Metrics ---
        print("  Checking Equity Metrics sheet...")
        a_rows = load_sheet_rows(agent_wb, "Equity Metrics")
        g_rows = load_sheet_rows(gt_wb, "Equity Metrics")
        if a_rows is None:
            all_errors.append("Sheet 'Equity Metrics' not found in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Equity Metrics' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            # Build lookup: (dept, exp_band) -> row
            a_lookup = {}
            for row in a_data:
                if row and row[0] and row[1]:
                    key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                    a_lookup[key] = row

            errors = []
            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing equity row: {g_row[0]}/{g_row[1]}")
                    continue
                # Highest_Paid_Group (col 2) - string match
                if not str_match(a_row[2], g_row[2]):
                    errors.append(f"{key} Highest_Paid_Group: '{a_row[2]}' vs '{g_row[2]}'")
                # Lowest_Paid_Group (col 3) - string match
                if not str_match(a_row[3], g_row[3]):
                    errors.append(f"{key} Lowest_Paid_Group: '{a_row[3]}' vs '{g_row[3]}'")
                # Pay_Gap_Pct (col 4)
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key} Pay_Gap_Pct: {a_row[4]} vs {g_row[4]}")
                # Equity_Ratio (col 5)
                if not num_close(a_row[5], g_row[5], 0.01):
                    errors.append(f"{key} Equity_Ratio: {a_row[5]} vs {g_row[5]}")
                # Equity_Status (col 6)
                if not str_match(a_row[6], g_row[6]):
                    errors.append(f"{key} Equity_Status: '{a_row[6]}' vs '{g_row[6]}'")

            if errors:
                all_errors.extend(errors)
                print(f"    ERRORS: {len(errors)}")
                for e in errors[:5]:
                    print(f"      {e}")
            else:
                print("    PASS")

        # --- Sheet: Summary ---
        print("  Checking Summary sheet...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            all_errors.append("Sheet 'Summary' not found in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row

            errors = []
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing summary row: {g_row[0]}")
                    continue
                # Value comparison depends on key
                if key in ("overall_equity_score",):
                    if not num_close(a_row[1], g_row[1], 0.02):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol=0.02)")
                elif key in ("total_employees_analyzed", "departments_analyzed",
                             "concerning_gaps_count", "monitor_gaps_count", "acceptable_gaps_count"):
                    if not num_close(a_row[1], g_row[1], 1):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol=1)")
                else:
                    # String match for dept/band names
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}: '{a_row[1]}' vs '{g_row[1]}'")

            if errors:
                all_errors.extend(errors)
                print(f"    ERRORS: {len(errors)}")
                for e in errors[:5]:
                    print(f"      {e}")
            else:
                print("    PASS")

    # --- Check 2: Word document ---
    print("Checking Word document...")
    docx_path = os.path.join(agent_ws, "Equity_Report.docx")
    if not os.path.exists(docx_path):
        all_errors.append("Equity_Report.docx not found in agent workspace")
    else:
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _headings = " ".join(
                p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")
            ).lower()

            if len(_text.strip()) < 100:
                all_errors.append("Equity_Report.docx has too little text content (< 100 chars)")

            # Check for key section headings
            required_sections = ["executive summary", "methodology", "finding", "recommendation", "compliance"]
            missing_sections = [s for s in required_sections if s not in _text and s not in _headings]
            if len(missing_sections) > 2:
                all_errors.append(f"Equity_Report.docx missing sections: {missing_sections}")

            # Check for key content keywords
            content_kws = ["equity", "salary", "department", "gap"]
            missing_kws = [k for k in content_kws if k not in _text]
            if len(missing_kws) == len(content_kws):
                all_errors.append(f"Equity_Report.docx missing all expected keywords: {content_kws}")

            print("    Word document checked.")
        except ImportError:
            if os.path.getsize(docx_path) < 100:
                all_errors.append("Equity_Report.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Equity_Report.docx: {_e}")

    # --- Check 3: Google Form ---
    print("Checking Google Form...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id FROM gform.forms WHERE LOWER(title) LIKE '%compensation fairness%'")
        forms = cur.fetchall()
        if not forms:
            all_errors.append("Google Form 'Compensation Fairness Survey' not found in gform.forms")
        else:
            form_id = forms[0][0]
            print(f"    GForm found (id={form_id})")
            # Check question count
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
            q_count = cur.fetchone()[0]
            if q_count < 5:
                all_errors.append(f"Google Form has {q_count} questions, expected at least 5")
            else:
                print(f"    {q_count} questions found.")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking GForm: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:15]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

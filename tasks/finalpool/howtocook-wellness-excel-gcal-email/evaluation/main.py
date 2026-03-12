"""Evaluation for howtocook-wellness-excel-gcal-email."""
import argparse
import json
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Excel meal plan file."""
    print("\n=== Checking Excel File ===")
    try:
        import openpyxl
    except ImportError:
        check("openpyxl available", False, "openpyxl not installed")
        return

    agent_file = os.path.join(agent_workspace, "Wellness_Meal_Plan.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Wellness_Meal_Plan.xlsx")

    check("Wellness_Meal_Plan.xlsx exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    # Check Meal Plan sheet
    print("\n--- Meal Plan Sheet ---")
    mp_ws = get_sheet(agent_wb, "Meal Plan")
    check("Sheet 'Meal Plan' exists", mp_ws is not None, f"Found: {agent_wb.sheetnames}")

    if mp_ws:
        headers = [c.value for c in list(mp_ws.rows)[0]] if mp_ws.max_row > 0 else []
        check("Has Day column", any("day" in str(h).lower() for h in headers if h), f"Headers: {headers}")
        check("Has Meal_Type column",
              any("meal" in str(h).lower() or "type" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Has Recipe_Name column",
              any("recipe" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")

        data_rows = [row for row in mp_ws.iter_rows(min_row=2, values_only=True)
                     if any(v is not None for v in row)]
        check("Meal Plan has at least 15 rows", len(data_rows) >= 15,
              f"Found {len(data_rows)} rows")

        # Check all 5 days covered
        days_found = set()
        for row in data_rows:
            if row and row[0]:
                day = str(row[0]).strip().lower()
                for d in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
                    if d in day:
                        days_found.add(d)
        check("All 5 weekdays covered", len(days_found) == 5,
              f"Found days: {days_found}")

        # Check meal types
        meal_types_found = set()
        col_idx = 1  # Meal_Type is 2nd column (0-indexed)
        for row in data_rows:
            if row and len(row) > col_idx and row[col_idx]:
                mt = str(row[col_idx]).strip().lower()
                if mt in ["breakfast", "lunch", "dinner"]:
                    meal_types_found.add(mt)
        check("All 3 meal types present (Breakfast, Lunch, Dinner)",
              len(meal_types_found) == 3,
              f"Found: {meal_types_found}")

    # Check Summary sheet
    print("\n--- Summary Sheet ---")
    sum_ws = get_sheet(agent_wb, "Summary")
    check("Sheet 'Summary' exists", sum_ws is not None, f"Found: {agent_wb.sheetnames}")

    if sum_ws:
        summary_data = {}
        for row in sum_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary_data[str(row[0]).strip().lower()] = row[1]

        check("Summary has Days_Covered = 5",
              any(("days" in k) and num_close(v, 5, 0.1) for k, v in summary_data.items()),
              f"Data: {summary_data}")
        check("Summary has Total_Recipes >= 15",
              any(("total" in k and "recipe" in k) and v is not None and float(v) >= 15
                  for k, v in summary_data.items() if v is not None),
              f"Data: {summary_data}")


def check_gcal():
    """Check Google Calendar events."""
    print("\n=== Checking Google Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    check("At least 5 calendar events created", len(events) >= 5,
          f"Found {len(events)} events")

    lunch_events = [e for e in events if "healthy lunch break" in (e[1] or "").lower()]
    check("5 'Healthy Lunch Break' events created",
          len(lunch_events) >= 5,
          f"Healthy Lunch Break events: {len(lunch_events)}")

    conn.close()


def check_emails():
    """Check that wellness email was sent."""
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    conn.close()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    result = find_email_for_recipient("all-staff@company.example.com")
    check("Email sent to all-staff@company.example.com", result is not None,
          f"Total emails: {len(all_emails)}")

    if result:
        subj, from_addr, to_addr, body = result
        check("Email subject contains '5-Day Corporate Wellness Meal Plan'",
              "wellness" in (subj or "").lower() and "5" in (subj or ""),
              f"Subject: {subj}")
        check("Email from wellness@company.example.com",
              "wellness@company.example.com" in (from_addr or "").lower(),
              f"From: {from_addr}")
        body_lower = (body or "").lower()
        check("Email mentions wellness goals (stress or focus)",
              "stress" in body_lower or "focus" in body_lower,
              "Expected wellness goals mentioned")
        check("Email mentions 5 days or 15 meals",
              "5" in (body or "") or "five" in body_lower or "15" in (body or ""),
              "Expected meal plan scope mentioned")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("HOWTOCOOK WELLNESS EXCEL GCAL EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)
    check_gcal()
    check_emails()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

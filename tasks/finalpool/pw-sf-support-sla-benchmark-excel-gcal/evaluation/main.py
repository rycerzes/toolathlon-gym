"""Evaluation script for pw-sf-support-sla-benchmark-excel-gcal."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "SLA_Benchmark_Report.xlsx")
    check("SLA_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "SLA_Benchmark_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        check("SLA_Comparison sheet exists", "SLA_Comparison" in wb.sheetnames)
        if "SLA_Comparison" in wb.sheetnames:
            ws = wb["SLA_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("SLA_Comparison has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Ticket_Count', 'Our_Avg_Response_Hrs', 'Compliance_Status']:
                check(f"SLA_Comparison has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Action_Items sheet exists", "Action_Items" in wb.sheetnames)
        if "Action_Items" in wb.sheetnames:
            ws = wb["Action_Items"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Action_Items has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Response_Gap', 'Recommended_Action']:
                check(f"Action_Items has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Summary sheet exists", "Summary" in wb.sheetnames)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # Check calendar events
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT summary FROM gcal.events WHERE summary ILIKE %s", ('%SLA%',))
            events = cur.fetchall()
            check("SLA Review event created", any("review" in str(e[0]).lower() for e in events), f"found {len(events)} SLA events")
            check("SLA Workshop event created", any("workshop" in str(e[0]).lower() or "improvement" in str(e[0]).lower() for e in events))
            conn.close()
        except Exception as e:
            check("Calendar verification", False, str(e))

        check("sla_analyzer.py exists", os.path.exists(os.path.join(agent_workspace, "sla_analyzer.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

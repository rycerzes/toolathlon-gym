"""Evaluation for terminal-canvas-notion-study-gcal-excel.
Checks:
1. Study_Plan_Report.xlsx with 3 sheets and correct course data
2. Notion database "Student Study Planner" with 5 course entries
3. Google Calendar events for study sessions
4. study_planner.py script exists
"""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

EXPECTED_COURSES = [
    "Creative Computing",
    "Foundations of Finance",
    "Biochemistry",
]

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').strip())
    except Exception:
        return default


def check_excel(workspace):
    print("\n=== Check 1: Study_Plan_Report.xlsx ===")
    path = os.path.join(workspace, "Study_Plan_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    # Course_Analysis sheet
    sheets_lower = [s.lower() for s in sheets]
    ca_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s or "analysis" in s), 0)
    ws = wb[sheets[ca_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    # Query dynamic course count from Canvas DB
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM canvas.courses")
        expected_course_count = cur.fetchone()[0]
        cur.close(); conn.close()
    except Exception:
        expected_course_count = 5
    check(f"Course_Analysis has {expected_course_count} course rows",
          len(data_rows) >= expected_course_count, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    check("Contains Creative Computing course", "creative computing" in all_text, f"Text: {all_text[:100]}")
    check("Contains Foundations of Finance course", "foundations of finance" in all_text or "finance" in all_text,
          f"Text: {all_text[:100]}")

    # Check headers for required columns
    if rows:
        headers = [str(c).lower() if c else "" for c in rows[0]]
        has_weekly = any("weekly" in h or "hours" in h for h in headers)
        has_priority = any("priority" in h for h in headers)
        check("Has Weekly_Hours column", has_weekly, f"Headers: {rows[0]}")
        check("Has Priority column", has_priority, f"Headers: {rows[0]}")

    # Weekly_Schedule sheet
    ws_idx = next((i for i, s in enumerate(sheets_lower) if "weekly" in s or "schedule" in s), 1)
    if ws_idx < len(sheets):
        ws2 = wb[sheets[ws_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Weekly_Schedule has 5 session rows", len(data_rows2) >= 5, f"Found {len(data_rows2)}")

    # Priority_Matrix sheet
    pm_idx = next((i for i, s in enumerate(sheets_lower) if "priority" in s or "matrix" in s), 2)
    if pm_idx < len(sheets):
        ws3 = wb[sheets[pm_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Priority_Matrix has High level", "high" in all_text3, f"Text: {all_text3[:100]}")
        check("Priority_Matrix has Medium level", "medium" in all_text3, f"Text: {all_text3[:100]}")


def check_notion():
    print("\n=== Check 2: Notion Study Planner Database ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM notion.databases")
    dbs = cur.fetchall()
    planner_db = None
    for db_id, title in dbs:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
        elif isinstance(title, str):
            try:
                parsed = json.loads(title)
                if isinstance(parsed, list):
                    title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                else:
                    title_str = str(title)
            except Exception:
                title_str = str(title)
        else:
            title_str = str(title) if title else ""
        if "study" in title_str.lower() and "planner" in title_str.lower():
            planner_db = (db_id, title_str)
            break

    check("Study Planner database exists", planner_db is not None,
          f"Databases found: {[d[1] for d in dbs]}")

    if planner_db:
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE parent->>'database_id' = %s
        """, (planner_db[0],))
        count = cur.fetchone()[0]
        check("Database has 5 course entries", count >= 5, f"Found {count} pages")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 3: Google Calendar Study Sessions ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT summary, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    check("At least 5 study session events", len(events) >= 5, f"Found {len(events)} events")

    if events:
        summaries = " ".join(str(e[0]) for e in events).lower()
        check("Events mention study session", "study" in summaries, f"Summaries: {summaries[:100]}")
        check("Events mention a course name",
              any(c.lower() in summaries for c in EXPECTED_COURSES),
              f"Summaries: {summaries[:100]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: study_planner.py ===")
    path = os.path.join(workspace, "study_planner.py")
    check("study_planner.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check no study session events on weekends
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, EXTRACT(DOW FROM start_datetime) AS dow
            FROM gcal.events
            WHERE lower(summary) LIKE '%%study%%'
        """)
        events = cur.fetchall()
        weekend_events = [e for e in events if e[2] in (0, 6)]
        check("No study session events on weekends",
              len(weekend_events) == 0,
              f"Found {len(weekend_events)} weekend study events")
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

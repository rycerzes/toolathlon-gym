"""Evaluation for sf-hr-experience-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected tenure data from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               ROUND(AVG("YEARS_EXPERIENCE")::numeric, 2) as avg_tenure,
               MIN("YEARS_EXPERIENCE") as min_tenure,
               MAX("YEARS_EXPERIENCE") as max_tenure,
               COUNT(*) as emp_count
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    rows = cur.fetchall()

    cur.execute("""
        SELECT ROUND(AVG("YEARS_EXPERIENCE")::numeric, 2),
               COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
    """)
    overall = cur.fetchone()
    cur.close()
    conn.close()
    return rows, overall


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Tenure_Analysis.xlsx."""
    print("\n=== Checking Tenure_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Tenure_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    dept_data, overall = get_expected_data()

    # Check Department Tenure sheet
    dept_sheet = None
    for name in wb.sheetnames:
        if "department" in name.lower() and "tenure" in name.lower():
            dept_sheet = wb[name]
            break
    if dept_sheet is None:
        record("Sheet 'Department Tenure' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Department Tenure' exists", True)
        rows = list(dept_sheet.iter_rows(min_row=2, values_only=True))
        record("Department Tenure has 7 rows", len(rows) == 7, f"Got {len(rows)}")

        agent_lookup = {}
        for r in rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for exp in dept_data:
            dept = exp[0]
            a_row = agent_lookup.get(dept.lower())
            if a_row is None:
                record(f"Department '{dept}' present", False, "Missing")
                all_ok = False
                continue

            ok_avg = num_close(a_row[1], exp[1], 0.5)
            record(f"'{dept}' Avg_Tenure", ok_avg,
                   f"Expected {exp[1]}, got {a_row[1]}")
            if not ok_avg:
                all_ok = False

            ok_count = num_close(a_row[4], exp[4], 50)
            record(f"'{dept}' Employee_Count", ok_count,
                   f"Expected {exp[4]}, got {a_row[4]}")
            if not ok_count:
                all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        for key, val in summary.items():
            if "total" in key and "employee" in key:
                ok = num_close(val, overall[1], 100)
                record("Summary Total_Employees", ok,
                       f"Expected {overall[1]}, got {val}")
                if not ok:
                    all_ok = False
            elif "department" in key and "count" in key:
                ok = num_close(val, 7, 0)
                record("Summary Department_Count", ok,
                       f"Expected 7, got {val}")
                if not ok:
                    all_ok = False

    return all_ok


def check_notion():
    """Check Notion page with 'tenure' in title."""
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()

    found_page = False
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else json.loads(page[1]) if page[1] else {}
        title_text = ""
        if "title" in props:
            t = props["title"]
            if isinstance(t, dict) and "title" in t:
                for item in t["title"]:
                    if isinstance(item, dict):
                        title_text += item.get("plain_text", item.get("text", {}).get("content", ""))
        if "tenure" in title_text.lower():
            found_page = True
            break

    record("Notion page with 'tenure' in title found", found_page,
           "No page with 'tenure' in title")

    cur.close()
    conn.close()
    return found_page


def check_email():
    """Check email with tenure-related subject."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr FROM email.messages
        WHERE subject ILIKE '%%tenure%%' OR subject ILIKE '%%experience%%'
        LIMIT 5
    """)
    rows = cur.fetchall()
    found = len(rows) > 0
    record("Email with tenure/experience in subject", found, "No matching email found")

    if found:
        for subj, to_addr in rows:
            if "chro@company.com" in str(to_addr).lower():
                record("Email to chro@company.com", True)
                break
        else:
            record("Email to chro@company.com", False,
                   f"To addresses: {[str(r[1]) for r in rows]}")

    cur.close()
    conn.close()
    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)

    db_fail_before = FAIL_COUNT
    notion_ok = check_notion()
    email_ok = check_email()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")

    overall = excel_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

"""Evaluation for sf-department-budget-analysis."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Department_Budget_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Department_Budget_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Budget Analysis
    print(f"  Checking Budget Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Budget Analysis")
    g_rows = load_sheet_rows(gt_wb, "Budget Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Budget Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Budget Analysis' not found in groundtruth")
    else:
        sheet_name = "Budget Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100.0):
                    errors.append(f"{key}.Budget: {a_row[1]} vs {g_row[1]} (tol=100.0)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1):
                    errors.append(f"{key}.Planned_Headcount: {a_row[2]} vs {g_row[2]} (tol=1)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5):
                    errors.append(f"{key}.Actual_Headcount: {a_row[3]} vs {g_row[3]} (tol=5)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 10.0):
                    errors.append(f"{key}.Avg_Salary: {a_row[4]} vs {g_row[4]} (tol=10.0)")

            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1000.0):
                    errors.append(f"{key}.Total_Salary_Cost: {a_row[5]} vs {g_row[5]} (tol=1000.0)")

            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 1.0):
                    errors.append(f"{key}.Budget_Utilization_Pct: {a_row[6]} vs {g_row[6]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1000.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=1000.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

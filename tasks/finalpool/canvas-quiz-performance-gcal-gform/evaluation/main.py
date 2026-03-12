"""
Evaluation script for canvas-quiz-performance-gcal-gform task.

Checks:
1. Excel file Quiz_Performance_Report.xlsx - 2 sheets with correct structure and data
2. Google Form "Quiz Improvement Feedback" exists with at least 3 questions
3. Google Calendar has 3 tutoring session events
4. Email sent to ccc.instructor@university.edu
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Quiz_Performance_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return False

    agent_file = os.path.join(agent_workspace, "Quiz_Performance_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quiz_Performance_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Quiz Stats sheet
    a_quiz = load_sheet_by_name(agent_wb, "Quiz Stats")
    g_quiz = load_sheet_by_name(gt_wb, "Quiz Stats")
    record("Sheet 'Quiz Stats' exists", a_quiz is not None)

    all_ok = True
    if a_quiz is not None and g_quiz is not None:
        a_data = [r for r in a_quiz[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_quiz[1:] if any(v is not None for v in r)]
        record("Quiz Stats row count matches", len(a_data) == len(g_data),
               f"Expected {len(g_data)}, got {len(a_data)}")

        # Build lookup by quiz title
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            title = str(g_row[0]).strip()
            key = title.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Quiz row exists: {title}", False, "Row not found")
                all_ok = False
                continue
            record(f"Quiz row exists: {title}", True)

            # Avg_Score (col 2, index 2)
            if len(g_row) > 2 and len(a_row) > 2:
                record(f"{title}: Avg_Score correct",
                       num_close(a_row[2], g_row[2], 2.0),
                       f"got {a_row[2]}, expected {g_row[2]}")
            # Total_Submissions (col 1, index 1)
            if len(g_row) > 1 and len(a_row) > 1:
                record(f"{title}: Total_Submissions correct",
                       num_close(a_row[1], g_row[1], 5),
                       f"got {a_row[1]}, expected {g_row[1]}")
            # Pass_Rate_Pct (col 5, index 5)
            if len(g_row) > 5 and len(a_row) > 5:
                record(f"{title}: Pass_Rate_Pct correct",
                       num_close(a_row[5], g_row[5], 2.0),
                       f"got {a_row[5]}, expected {g_row[5]}")

    # Check Course Summary sheet
    a_summ = load_sheet_by_name(agent_wb, "Course Summary")
    g_summ = load_sheet_by_name(gt_wb, "Course Summary")
    record("Sheet 'Course Summary' exists", a_summ is not None)

    if a_summ is not None and g_summ is not None:
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_summ[1:] if any(v is not None for v in r)]

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary row: {g_row[0]}", False, "Row not found")
                all_ok = False
                continue
            record(f"Summary row: {g_row[0]}", True)

            if key == "total_quizzes":
                record("Total_Quizzes value",
                       num_close(a_row[1], g_row[1], 0),
                       f"got {a_row[1]}, expected {g_row[1]}")
            elif key == "total_quiz_submissions":
                record("Total_Quiz_Submissions value",
                       num_close(a_row[1], g_row[1], 10),
                       f"got {a_row[1]}, expected {g_row[1]}")
            elif key == "overall_avg_score":
                record("Overall_Avg_Score value",
                       num_close(a_row[1], g_row[1], 2.0),
                       f"got {a_row[1]}, expected {g_row[1]}")

    return all_ok


# ============================================================================
# Check 2: Google Form
# ============================================================================

def check_gform():
    print("\n=== Checking Google Form ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    print(f"[check_gform] Found {len(forms)} forms.")
    record("At least 1 form created", len(forms) >= 1, f"Found {len(forms)}")

    found_form = False
    for form_id, title in forms:
        if title and "quiz" in title.lower() and ("improvement" in title.lower() or "feedback" in title.lower()):
            found_form = True
            record("Form 'Quiz Improvement Feedback' found", True)

            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id=%s", (form_id,))
            q_count = cur.fetchone()[0]
            record("Form has at least 3 questions", q_count >= 3,
                   f"Found {q_count} questions")

            cur.execute("SELECT title, question_type, required FROM gform.questions WHERE form_id=%s ORDER BY position", (form_id,))
            questions = cur.fetchall()

            has_mc = any("challenging" in (q[0] or "").lower() or "topic" in (q[0] or "").lower()
                        for q in questions)
            has_scale = any(q[1] in ("scale", "rating", "linear_scale") for q in questions)
            has_text = any(q[1] in ("text", "paragraph", "short_answer") for q in questions)

            record("Form has multiple-choice topic question", has_mc,
                   f"Questions: {[(q[0], q[1]) for q in questions]}")
            record("Form has scale/rating question", has_scale,
                   f"Types: {[q[1] for q in questions]}")
            record("Form has text question", has_text,
                   f"Types: {[q[1] for q in questions]}")
            break

    if not found_form:
        record("Form 'Quiz Improvement Feedback' found", False,
               f"Forms: {[(f[0], f[1]) for f in forms]}")

    cur.close()
    conn.close()
    return found_form


# ============================================================================
# Check 3: Google Calendar
# ============================================================================

def check_gcal():
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    record("At least 3 calendar events created", len(events) >= 3, f"Found {len(events)}")

    # Check for 3 tutoring sessions
    tutoring_events = [e for e in events
                       if e[0] and ("tutoring" in e[0].lower() or "session" in e[0].lower()
                                    or "ccc" in e[0].lower())]
    record("3 tutoring/session events found", len(tutoring_events) >= 3,
           f"Found {len(tutoring_events)}: {[e[0] for e in tutoring_events]}")

    # Check for March 2026 dates
    march_events = [e for e in events
                    if e[1] and "2026-03" in str(e[1])]
    record("Events scheduled in March 2026", len(march_events) >= 3,
           f"Found {len(march_events)} March 2026 events")

    return len(tutoring_events) >= 3


# ============================================================================
# Check 4: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")
    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    found_email = False
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = str(to_addr or "").lower()
        subject_lower = (subject or "").lower()
        if "ccc.instructor@university.edu" in to_str or "quiz" in subject_lower:
            found_email = True
            record("Email to ccc.instructor@university.edu found", True)

            record("Email subject mentions quiz or performance",
                   "quiz" in subject_lower or "performance" in subject_lower,
                   f"Subject: {subject}")

            body_lower = (body_text or "").lower()
            record("Email body mentions quiz data",
                   "quiz" in body_lower or "performance" in body_lower or "score" in body_lower,
                   "Body missing quiz content")
            break

    if not found_email:
        record("Email to ccc.instructor@university.edu found", False,
               f"Emails: {[(e[0], e[2]) for e in all_emails[:3]]}")

    return found_email


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    gform_ok = check_gform()
    gcal_ok = check_gcal()
    email_ok = check_emails()

    all_passed = excel_ok and gform_ok and gcal_ok and email_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

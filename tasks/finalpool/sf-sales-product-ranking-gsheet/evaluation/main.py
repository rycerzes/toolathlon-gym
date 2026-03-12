"""Evaluation for sf-sales-product-ranking-gsheet."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Product_Rankings.xlsx against groundtruth."""
    print("\n=== Checking Product_Rankings.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Product_Rankings.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Product_Rankings.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet: Rankings ---
    agent_ws = get_sheet(agent_wb, "Rankings")
    gt_ws = get_sheet(gt_wb, "Rankings")

    if agent_ws is None:
        record("Sheet 'Rankings' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Rankings' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("Rankings row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Build lookup by Product_Name (col 1)
        agent_lookup = {}
        for r in agent_rows:
            if r and r[1]:
                agent_lookup[str(r[1]).strip().lower()] = r

        for gt_row in gt_rows:
            if not gt_row or not gt_row[1]:
                continue
            key = str(gt_row[1]).strip().lower()
            short_name = key[:50]
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Product '{short_name}...' present", False, "Missing")
                all_ok = False
                continue

            # Category (col 0)
            ok_cat = str_match(a_row[0], gt_row[0])
            record(f"'{short_name}...' Category", ok_cat,
                   f"Expected {gt_row[0]}, got {a_row[0]}")
            if not ok_cat:
                all_ok = False

            # Units_Sold (col 3)
            ok_units = num_close(a_row[3], gt_row[3], 5)
            record(f"'{short_name}...' Units_Sold", ok_units,
                   f"Expected {gt_row[3]}, got {a_row[3]}")
            if not ok_units:
                all_ok = False

            # Revenue (col 4)
            ok_rev = num_close(a_row[4], gt_row[4], 50.0)
            record(f"'{short_name}...' Revenue", ok_rev,
                   f"Expected {gt_row[4]}, got {a_row[4]}")
            if not ok_rev:
                all_ok = False

            # Rank_In_Category (col 5)
            ok_rank = num_close(a_row[5], gt_row[5], 0)
            record(f"'{short_name}...' Rank", ok_rank,
                   f"Expected {gt_row[5]}, got {a_row[5]}")
            if not ok_rank:
                all_ok = False

    # --- Sheet: Category Totals ---
    agent_ws2 = get_sheet(agent_wb, "Category Totals")
    gt_ws2 = get_sheet(gt_wb, "Category Totals")

    if agent_ws2 is None:
        record("Sheet 'Category Totals' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Category Totals' exists", True)
        agent_rows2 = list(agent_ws2.iter_rows(min_row=2, values_only=True))
        gt_rows2 = list(gt_ws2.iter_rows(min_row=2, values_only=True))

        record("Category Totals row count", len(agent_rows2) == len(gt_rows2),
               f"Expected {len(gt_rows2)}, got {len(agent_rows2)}")

        agent_cat_lookup = {}
        for r in agent_rows2:
            if r and r[0]:
                agent_cat_lookup[str(r[0]).strip().lower()] = r

        for gt_row in gt_rows2:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_cat_lookup.get(key)
            if a_row is None:
                record(f"Category '{gt_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            # Total_Products_Sold (col 1)
            ok_total = num_close(a_row[1], gt_row[1], 50)
            record(f"'{gt_row[0]}' Total_Products_Sold", ok_total,
                   f"Expected {gt_row[1]}, got {a_row[1]}")
            if not ok_total:
                all_ok = False

            # Total_Revenue (col 2)
            ok_rev = num_close(a_row[2], gt_row[2], 500.0)
            record(f"'{gt_row[0]}' Total_Revenue", ok_rev,
                   f"Expected {gt_row[2]}, got {a_row[2]}")
            if not ok_rev:
                all_ok = False

    return all_ok


def check_gsheet():
    """Check Google Sheet exists with Rankings data."""
    print("\n=== Checking Google Sheet ===")

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%product%ranking%'")
    sheets = cur.fetchall()

    if not sheets:
        record("Google Sheet 'Product Rankings Dashboard' exists", False,
               "No spreadsheet with 'product ranking' in title found")
        cur.close()
        conn.close()
        return False

    record("Google Sheet 'Product Rankings Dashboard' exists", True)
    sheet_id = sheets[0][0]

    # Find the sheet (tab) within the spreadsheet
    cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s", (sheet_id,))
    tabs = cur.fetchall()
    record("Spreadsheet has at least 1 sheet tab", len(tabs) >= 1,
           f"Found {len(tabs)} tabs")

    if not tabs:
        cur.close()
        conn.close()
        return False

    # Use the first available sheet tab
    tab_id = tabs[0][0]

    cur.execute("""
        SELECT row_index, col_index, value FROM gsheet.cells
        WHERE sheet_id = %s ORDER BY row_index, col_index
    """, (tab_id,))
    cells = cur.fetchall()

    # Build grid
    grid = {}
    for row_idx, col_idx, value in cells:
        if row_idx not in grid:
            grid[row_idx] = {}
        grid[row_idx][col_idx] = value

    # Skip the header row (could be row 0 or row 1 depending on MCP server)
    if grid:
        min_row = min(grid.keys())
        data_rows = {k: v for k, v in grid.items() if k > min_row}
    else:
        data_rows = {}
    record("Google Sheet has data rows", len(data_rows) >= 1,
           f"Found {len(data_rows)} data rows")

    # Check that at least one product name appears
    all_values = " ".join(str(v) for row in grid.values() for v in row.values())
    record("Google Sheet contains product data",
           "samsung" in all_values.lower() or "oneplus" in all_values.lower(),
           "Expected product names not found")

    cur.close()
    conn.close()
    return len(data_rows) >= 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)

    db_fail_before = FAIL_COUNT
    gsheet_ok = check_gsheet()
    db_failures = FAIL_COUNT - db_fail_before

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if db_failures > 0 and excel_ok:
        print(f"  WARNING: {db_failures} DB checks failed (not blocking)")
    overall = excel_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

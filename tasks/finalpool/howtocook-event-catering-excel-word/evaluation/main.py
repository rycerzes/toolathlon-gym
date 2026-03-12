"""
Evaluation for howtocook-event-catering-excel-word task.

Checks:
1. Catering_Plan.xlsx exists with Menu and Ingredients sheets
2. Menu sheet has at least 8 data rows
3. Catering_Proposal.docx exists and mentions catering/menu
4. GForm "Menu Approval Survey" exists with at least 3 questions
5. Email sent to client@corporate.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Catering_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Catering_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Catering_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Catering_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_menu = any("menu" in s for s in sheet_names_lower)
    has_ingredients = any("ingredient" in s for s in sheet_names_lower)

    record("Excel has Menu sheet", has_menu, f"Sheets: {wb.sheetnames}")
    record("Excel has Ingredients sheet", has_ingredients, f"Sheets: {wb.sheetnames}")

    if has_menu:
        menu_sheet_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "menu" in s)]
        ws = wb[menu_sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Menu sheet has at least 8 dishes", len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")

        # Check for expected columns
        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_dish = any("dish" in h or "name" in h for h in headers)
            has_cost = any("cost" in h or "price" in h for h in headers)
            record("Menu has Dish_Name column", has_dish, f"Headers: {rows[0]}")
            record("Menu has cost column", has_cost, f"Headers: {rows[0]}")

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Catering_Plan.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]) if ri < len(a_rows) else 0)):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_word_doc(agent_workspace):
    print("\n=== Check 2: Word Document Catering_Proposal.docx ===")

    docx_path = os.path.join(agent_workspace, "Catering_Proposal.docx")
    if not os.path.exists(docx_path):
        record("Catering_Proposal.docx exists", False, f"Not found at {docx_path}")
        return
    record("Catering_Proposal.docx exists", True)

    try:
        doc = Document(docx_path)
    except Exception as e:
        record("Word doc readable", False, str(e))
        return
    record("Word doc readable", True)

    all_text = "\n".join(p.text for p in doc.paragraphs).lower()
    has_catering = "catering" in all_text or "menu" in all_text
    has_summary = "summary" in all_text or "overview" in all_text or "executive" in all_text
    has_timeline = "timeline" in all_text or "preparation" in all_text or "schedule" in all_text

    record("Word doc mentions catering/menu", has_catering, "No catering/menu content")
    record("Word doc has summary/overview section", has_summary, "No summary/overview section")
    record("Word doc has timeline/preparation section", has_timeline, "No timeline/preparation section")


def check_gform():
    print("\n=== Check 3: GForm Menu Approval Survey ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    approval_form = None
    for form_id, title in forms:
        if "menu" in (title or "").lower() or "approval" in (title or "").lower() or "survey" in (title or "").lower():
            approval_form = (form_id, title)
            break

    record("Menu Approval Survey form exists", approval_form is not None,
           f"Forms found: {[f[1] for f in forms]}")

    if approval_form:
        form_id, title = approval_form
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        record("Form has at least 3 questions", q_count >= 3,
               f"Found {q_count} questions")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 4: Email to client@corporate.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "client@corporate.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email sent to client@corporate.com", matching is not None,
           f"Messages found: {len(messages)}")

    if matching:
        subject, _, _, body_text = matching
        all_text = ((subject or "") + " " + (body_text or "")).lower()
        has_catering = "catering" in all_text or "menu" in all_text or "proposal" in all_text
        record("Email mentions catering/menu/proposal", has_catering,
               f"Subject: {subject}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_word_doc(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

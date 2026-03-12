"""
Evaluation for train-conference-travel-excel-gcal-email task.

Checks:
1. Conference_Travel_Plan.xlsx exists with Outbound, Return, Summary sheets
2. Outbound sheet has 3 rows with correct train codes and prices
3. Return sheet has 3 rows with correct train codes
4. Summary sheet has correct total cost and avg cost
5. GCal has 6 travel events (3 outbound + 3 return)
6. Email sent to all 3 attendees with subject 'Conference Travel Plan'
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Conference_Travel_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Conference_Travel_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Conference_Travel_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Conference_Travel_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Outbound sheet
    if "outbound" not in sheet_names_lower:
        record("Outbound sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Outbound sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("outbound")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Outbound has 3 attendee rows", len(data_rows) == 3, f"Found {len(data_rows)}")

        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        has_g235 = "G235" in all_text
        has_g168 = "G168" in all_text
        record("Outbound has G235 (Beijing→Qufu)", has_g235, "G235 not found")
        record("Outbound has G168 (Shanghai→Qufu)", has_g168, "G168 not found")

        # Check price column
        prices = []
        for r in data_rows:
            for c in r:
                try:
                    v = float(c)
                    if 50 < v < 500:
                        prices.append(v)
                except (TypeError, ValueError):
                    pass
        has_correct_prices = any(abs(p - 109.5) < 0.01 for p in prices) and any(abs(p - 119.5) < 0.01 for p in prices)
        record("Outbound prices correct (109.5 and 119.5 CNY)", has_correct_prices,
               f"Prices found: {prices}")

    # Check Return sheet
    if "return" not in sheet_names_lower:
        record("Return sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Return sheet exists", True)
        ws_r = wb[wb.sheetnames[sheet_names_lower.index("return")]]
        rows_r = list(ws_r.iter_rows(values_only=True))
        data_rows_r = [r for r in rows_r[1:] if any(c for c in r)]
        record("Return has 3 attendee rows", len(data_rows_r) == 3, f"Found {len(data_rows_r)}")

        all_text_r = " ".join(str(c) for r in rows_r for c in r if c).upper()
        has_g236 = "G236" in all_text_r
        has_g167 = "G167" in all_text_r
        record("Return has G236 (Qufu→Beijing)", has_g236, "G236 not found")
        record("Return has G167 (Qufu→Shanghai)", has_g167, "G167 not found")

    # Check Summary sheet
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws_s = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows_s = list(ws_s.iter_rows(values_only=True))
        all_text_s = " ".join(str(c) for r in rows_s for c in r if c)

        numeric_vals = []
        for r in rows_s:
            for c in r:
                try:
                    numeric_vals.append(float(c))
                except (TypeError, ValueError):
                    pass
        has_total = any(abs(v - 677.0) < 1.0 for v in numeric_vals)
        record("Summary total cost ~677.0 CNY", has_total, f"Numeric values found: {numeric_vals}")

        has_avg = any(abs(v - 225.67) < 1.0 for v in numeric_vals)
        record("Summary avg cost ~225.67 CNY per person", has_avg, f"Numeric values found: {numeric_vals}")


def check_gcal():
    print("\n=== Check 2: Calendar Events ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-16'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    travel_events = [e for e in events if e[0] and (
        "conference" in e[0].lower() or "travel" in e[0].lower() or "qufu" in e[0].lower()
    )]

    record("At least 6 travel calendar events created", len(travel_events) >= 6,
           f"Found {len(travel_events)} travel events. All events: {[e[0] for e in events]}")

    outbound_events = [e for e in travel_events if "2026-03-12" in str(e[1])]
    return_events = [e for e in travel_events if "2026-03-15" in str(e[1])]
    record("3 outbound events on 2026-03-12", len(outbound_events) >= 3,
           f"Found {len(outbound_events)}")
    record("3 return events on 2026-03-15", len(return_events) >= 3,
           f"Found {len(return_events)}")


def check_emails():
    print("\n=== Check 3: Emails to Attendees ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    attendee_emails = [
        "zhang.wei@uni.edu",
        "liu.mei@institute.org",
        "wang.fang@college.cn",
    ]

    cur.execute("""
        SELECT to_addr, subject FROM email.messages
        WHERE subject ILIKE '%conference travel%' OR subject ILIKE '%travel plan%'
    """)
    messages = cur.fetchall()
    cur.close()
    conn.close()

    all_msgs = list(messages)
    record("At least 3 travel plan emails sent", len(all_msgs) >= 3,
           f"Found {len(all_msgs)} matching emails")

    all_recipients = []
    for row in all_msgs:
        to_raw = row[0]
        if isinstance(to_raw, list):
            all_recipients.extend([str(r).lower() for r in to_raw])
        elif isinstance(to_raw, str):
            all_recipients.append(to_raw.lower())

    all_recipients_str = " ".join(all_recipients)
    for email in attendee_emails:
        record(f"Email sent to {email}",
               email.lower() in all_recipients_str,
               f"Recipients: {all_recipients_str[:200]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

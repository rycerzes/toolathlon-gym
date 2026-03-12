"""Evaluation for terminal-canvas-excel-word-notion-email.
Checks:
1. Student_Risk_Analysis.xlsx with 4 sheets and correct data
2. Intervention_Plan.docx with required sections
3. Notion database "Student Risk Tracker" with 2 entries
4. Email sent to academic_advisors@university.edu
5. risk_scorer.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(workspace):
    print("\n=== Check 1: Student_Risk_Analysis.xlsx ===")
    path = os.path.join(workspace, "Student_Risk_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Course_Overview
    co_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s and "overview" in s), 0)
    ws1 = wb[sheets[co_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    # Query dynamic course count from Canvas DB
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM canvas.courses WHERE id IN (2013, 2014)")
        expected_course_count = cur.fetchone()[0]
        cur.close(); conn.close()
    except Exception:
        expected_course_count = 2
    check(f"Course_Overview has {expected_course_count} rows",
          len(data1) >= expected_course_count, f"Found {len(data1)}")

    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains Foundations of Finance", "finance" in all_text1, f"Text: {all_text1[:100]}")

    # Check avg_score values roughly
    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        avg_idx = next((i for i, h in enumerate(headers) if "avg" in h and "score" in h), -1)
        if avg_idx >= 0 and len(data1) >= 2:
            scores = [float(r[avg_idx]) for r in data1 if r[avg_idx] is not None]
            check("Avg scores in reasonable range (70-85)",
                  all(70 <= s <= 85 for s in scores),
                  f"Scores: {scores}")

    # Risk_Distribution
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "risk" in s and "dist" in s), 1)
    if rd_idx < len(sheets):
        ws2 = wb[sheets[rd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Risk_Distribution has 3 rows", len(data2) >= 3, f"Found {len(data2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Has High risk level", "high" in all_text2)
        check("Has Medium risk level", "medium" in all_text2)
        check("Has Low risk level", "low" in all_text2)

    # At_Risk_Students
    ar_idx = next((i for i, s in enumerate(sheets_lower) if "at_risk" in s or "risk_student" in s or ("risk" in s and "student" in s)), 2)
    if ar_idx < len(sheets):
        ws3 = wb[sheets[ar_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("At_Risk_Students has 2 course rows", len(data3) >= 2, f"Found {len(data3)}")

    # Intervention_Plan sheet
    ip_idx = next((i for i, s in enumerate(sheets_lower) if "intervention" in s), 3)
    if ip_idx < len(sheets):
        ws4 = wb[sheets[ip_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Intervention_Plan has 3 rows", len(data4) >= 3, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Intervention mentions advising", "advis" in all_text4, f"Text: {all_text4[:150]}")
        check("Intervention mentions tutoring", "tutor" in all_text4, f"Text: {all_text4[:150]}")


def check_word(workspace):
    print("\n=== Check 2: Intervention_Plan.docx ===")
    path = os.path.join(workspace, "Intervention_Plan.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    check("Document mentions retention or intervention", "retention" in full_text or "intervention" in full_text,
          f"Text: {full_text[:100]}")
    check("Document mentions high risk", "high risk" in full_text or "high-risk" in full_text,
          f"Text: {full_text[:100]}")
    check("Document mentions both courses", "2013" in full_text and "2014" in full_text,
          f"Text: {full_text[:150]}")
    check("Document has substantial content", len(full_text) > 200, f"Length: {len(full_text)}")


def check_notion():
    print("\n=== Check 3: Notion Student Risk Tracker ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        tracker_db = None
        for db_id, title in dbs:
            title_str = ""
            if isinstance(title, list):
                title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
            elif isinstance(title, str):
                try:
                    parsed = json.loads(title)
                    if isinstance(parsed, list):
                        title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                    else:
                        title_str = str(title)
                except Exception:
                    title_str = str(title)
            else:
                title_str = str(title) if title else ""
            if "risk" in title_str.lower() and "tracker" in title_str.lower():
                tracker_db = (db_id, title_str)
                break
        check("Student Risk Tracker database exists", tracker_db is not None,
              f"Databases: {[d[1] for d in dbs]}")

        if tracker_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (tracker_db[0],))
            count = cur.fetchone()[0]
            check("Tracker has 2 course entries", count >= 2, f"Found {count}")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Check 4: Email to Academic Advisors ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%academic_advisors%%'
               OR to_addr::text ILIKE '%%advisor%%'
               OR subject ILIKE '%%retention%%risk%%'
               OR subject ILIKE '%%risk%%analysis%%'
        """)
        emails = cur.fetchall()

        # Also check sent_log and drafts
        if not emails:
            cur.execute("""
                SELECT id, subject, to_addr, body_text
                FROM email.drafts
                WHERE to_addr::text ILIKE '%%advisor%%'
                   OR subject ILIKE '%%retention%%'
                   OR subject ILIKE '%%risk%%'
            """)
            emails = cur.fetchall()

        check("Email sent about risk analysis", len(emails) >= 1, "No matching email found")
        if emails:
            subject = str(emails[0][1]).lower() if emails[0][1] else ""
            check("Email subject relevant",
                  "risk" in subject or "retention" in subject or "analysis" in subject,
                  f"Subject: {emails[0][1]}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: risk_scorer.py ===")
    path = os.path.join(workspace, "risk_scorer.py")
    check("risk_scorer.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check that email is not sent to wrong recipients
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr FROM email.messages
            WHERE subject ILIKE '%%risk%%' OR subject ILIKE '%%retention%%'
        """)
        emails = cur.fetchall()
        noise_recipients = ["all-staff@university.edu", "it@university.edu",
                            "facilities@university.edu"]
        for email_row in emails:
            to_str = str(email_row[0]).lower()
            for noise in noise_recipients:
                if noise in to_str:
                    check("No risk emails sent to wrong recipients", False,
                          f"Sent to noise recipient: {noise}")
                    cur.close(); conn.close()
                    return
        check("No risk emails sent to wrong recipients", True)
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

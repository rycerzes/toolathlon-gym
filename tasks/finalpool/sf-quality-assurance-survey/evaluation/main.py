"""
Evaluation script for sf-quality-assurance-survey task.

Checks:
1. Excel file QA_Assessment.xlsx with 3 sheets
2. Google Form created with correct structure
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    """Query PostgreSQL to compute expected values."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Agent Scorecard: reporters as agents
    # Targets: satisfaction >= 4.0, resolution >= 95%, response_time compared to 8 (medium target)
    cur.execute("""
        SELECT
            "REPORTER" as agent,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as avg_sat,
            ROUND(COUNT(CASE WHEN "STATUS"='Resolved' THEN 1 END)*100.0/COUNT(*)::numeric, 1) as resolution_rate
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "REPORTER"
        ORDER BY "REPORTER"
    """)
    agent_rows = cur.fetchall()
    agent_scorecard = []
    for row in agent_rows:
        agent_name, avg_resp, avg_sat, res_rate = row
        below = []
        if float(avg_resp) > 8.0:
            below.append("Response Time")
        if float(avg_sat) < 4.0:
            below.append("Satisfaction")
        if float(res_rate) < 95.0:
            below.append("Resolution Rate")
        agent_scorecard.append({
            "Agent": agent_name,
            "Avg_Response_Time": float(avg_resp),
            "Avg_Satisfaction": float(avg_sat),
            "Resolution_Rate": float(res_rate),
            "Below_Target_Areas": ", ".join(below) if below else "None",
        })

    # Issue Type Analysis
    cur.execute("""
        SELECT
            "ISSUE_TYPE",
            COUNT(*) as cnt,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as avg_sat
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE"
        ORDER BY "ISSUE_TYPE"
    """)
    issue_rows = cur.fetchall()
    issue_type_analysis = []
    for row in issue_rows:
        issue_type_analysis.append({
            "Issue_Type": row[0],
            "Ticket_Count": int(row[1]),
            "Avg_Response_Time": float(row[2]),
            "Avg_Satisfaction": float(row[3]),
        })

    # Summary
    cur.execute("""
        SELECT
            COUNT(*),
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2),
            ROUND(COUNT(CASE WHEN "STATUS"='Resolved' THEN 1 END)*100.0/COUNT(*)::numeric, 1)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
    """)
    s = cur.fetchone()
    total_tickets = int(s[0])
    overall_avg_resp = float(s[1])
    overall_avg_sat = float(s[2])
    overall_res_rate = float(s[3])

    agents_below_sat = sum(1 for a in agent_scorecard if a["Avg_Satisfaction"] < 4.0)
    agents_below_resp = sum(1 for a in agent_scorecard if a["Avg_Response_Time"] > 8.0)

    summary = {
        "Total_Tickets": total_tickets,
        "Overall_Avg_Response_Time": overall_avg_resp,
        "Overall_Avg_Satisfaction": overall_avg_sat,
        "Overall_Resolution_Rate": overall_res_rate,
        "Agents_Below_Satisfaction_Target": agents_below_sat,
        "Agents_Below_Response_Target": agents_below_resp,
    }

    cur.close()
    conn.close()

    return {
        "agent_scorecard": agent_scorecard,
        "issue_type_analysis": issue_type_analysis,
        "summary": summary,
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "QA_Assessment.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    # Check sheets exist
    for sn in ["Agent Scorecard", "Issue Type Analysis", "Summary"]:
        found = any(str_match(s, sn) for s in wb.sheetnames)
        check(f"Sheet '{sn}' exists", found, f"Found: {wb.sheetnames}")

    # --- Agent Scorecard ---
    print("\n--- Agent Scorecard ---")
    ws = get_sheet(wb, "Agent Scorecard")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["agent_scorecard"]
        check("Agent Scorecard row count", len(rows) == len(exp),
              f"Expected {len(exp)}, got {len(rows)}")

        for e_row in exp:
            agent = e_row["Agent"]
            matched = None
            for r in rows:
                if r and str_match(r[0], agent):
                    matched = r
                    break
            if matched:
                check(f"{agent} Avg_Response_Time",
                      num_close(matched[1], e_row["Avg_Response_Time"], 1.0),
                      f"Expected {e_row['Avg_Response_Time']}, got {matched[1]}")
                check(f"{agent} Avg_Satisfaction",
                      num_close(matched[2], e_row["Avg_Satisfaction"], 0.5),
                      f"Expected {e_row['Avg_Satisfaction']}, got {matched[2]}")
                check(f"{agent} Resolution_Rate",
                      num_close(matched[3], e_row["Resolution_Rate"], 1.0),
                      f"Expected {e_row['Resolution_Rate']}, got {matched[3]}")
                # Check below_target_areas contains key words
                if matched[4]:
                    agent_below = str(matched[4]).lower()
                    for area in e_row["Below_Target_Areas"].lower().split(", "):
                        if area != "none":
                            check(f"{agent} below area '{area}'",
                                  area in agent_below,
                                  f"Expected '{area}' in '{matched[4]}'")
            else:
                check(f"Agent '{agent}' found", False, "Not in agent output")

    # --- Issue Type Analysis ---
    print("\n--- Issue Type Analysis ---")
    ws = get_sheet(wb, "Issue Type Analysis")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["issue_type_analysis"]
        check("Issue Type row count", len(rows) == len(exp),
              f"Expected {len(exp)}, got {len(rows)}")

        for e_row in exp:
            it = e_row["Issue_Type"]
            matched = None
            for r in rows:
                if r and str_match(r[0], it):
                    matched = r
                    break
            if matched:
                check(f"{it} Ticket_Count",
                      num_close(matched[1], e_row["Ticket_Count"], 5),
                      f"Expected {e_row['Ticket_Count']}, got {matched[1]}")
                check(f"{it} Avg_Response_Time",
                      num_close(matched[2], e_row["Avg_Response_Time"], 1.0),
                      f"Expected {e_row['Avg_Response_Time']}, got {matched[2]}")
                check(f"{it} Avg_Satisfaction",
                      num_close(matched[3], e_row["Avg_Satisfaction"], 0.5),
                      f"Expected {e_row['Avg_Satisfaction']}, got {matched[3]}")
            else:
                check(f"Issue Type '{it}' found", False, "Not in output")

    # --- Summary ---
    print("\n--- Summary ---")
    ws = get_sheet(wb, "Summary")
    if ws:
        data = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                data[str(row[0]).strip().lower().replace(" ", "_")] = row[1]

        for key, gt_val in expected["summary"].items():
            key_lower = key.lower()
            agent_val = data.get(key_lower)
            if agent_val is None:
                for ak, av in data.items():
                    if key_lower.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                check(f"Summary '{key}'",
                      num_close(agent_val, gt_val, 1.0),
                      f"Expected {gt_val}, got {agent_val}")
            else:
                check(f"Summary '{key}'",
                      str_match(agent_val, gt_val),
                      f"Expected '{gt_val}', got '{agent_val}'")


def check_form():
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE LOWER(title) LIKE '%qa%' OR LOWER(title) LIKE '%quality%' OR LOWER(title) LIKE '%self assess%'
        ORDER BY created_at DESC LIMIT 1
    """)
    form_row = cur.fetchone()
    check("Google Form exists", form_row is not None)
    if not form_row:
        cur.close()
        conn.close()
        return

    form_id = form_row[0]

    cur.execute("""
        SELECT title, question_type, required
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()
    check("Form has 5 questions", len(questions) == 5,
          f"Found {len(questions)}")

    if len(questions) >= 5:
        # Q1: name (text, required)
        check("Q1 is text type", questions[0][1] == "textQuestion",
              f"Got {questions[0][1]}")
        check("Q1 is required", questions[0][2] is True)

        # Q2: team (choice, required)
        check("Q2 is choice type", questions[1][1] == "choiceQuestion",
              f"Got {questions[1][1]}")
        check("Q2 is required", questions[1][2] is True)

        # Q3: response time rating (choice, required)
        check("Q3 is choice type", questions[2][1] == "choiceQuestion",
              f"Got {questions[2][1]}")
        check("Q3 is required", questions[2][2] is True)

        # Q4: challenge (text/paragraph, required)
        check("Q4 is text type", questions[3][1] == "textQuestion",
              f"Got {questions[3][1]}")
        check("Q4 is required", questions[3][2] is True)

        # Q5: suggestions (text/paragraph, not required)
        check("Q5 is text type", questions[4][1] == "textQuestion",
              f"Got {questions[4][1]}")
        check("Q5 is not required", questions[4][2] is False or questions[4][2] is None,
              f"Got required={questions[4][2]}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Computing Expected Values ===")
    try:
        expected = compute_expected_values()
        print(f"  Agent scorecard: {len(expected['agent_scorecard'])} agents")
        print(f"  Issue types: {len(expected['issue_type_analysis'])} types")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    check_excel(args.agent_workspace, expected)
    check_form()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    success = pass_rate >= 0.8

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

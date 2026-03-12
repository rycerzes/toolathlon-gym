"""
Evaluation script for sf-customer-health-dashboard task.

Checks:
1. Excel file Customer_Health.xlsx with 3 sheets
2. Notion page created
3. Calendar events for critical accounts with score < 15
"""
import argparse
import json
import os
import sys
from datetime import date

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        WITH order_stats AS (
            SELECT "CUSTOMER_ID",
                COUNT(*) as order_count,
                MAX("ORDER_DATE") as last_order_date
            FROM sf_data."SALES_DW__PUBLIC__ORDERS"
            GROUP BY "CUSTOMER_ID"
        ),
        health AS (
            SELECT c."CUSTOMER_ID", c."CUSTOMER_NAME", c."SEGMENT", c."REGION",
                ROUND(GREATEST(0, 100 - (CURRENT_DATE - COALESCE(os.last_order_date, c."SIGNUP_DATE")))::numeric, 2) as recency_score,
                ROUND(LEAST(100, COALESCE(os.order_count, 0) * 10)::numeric, 2) as frequency_score,
                ROUND(LEAST(100, c."LIFETIME_VALUE" / 50.0)::numeric, 2) as monetary_score,
                ROUND((0.4 * GREATEST(0, 100 - (CURRENT_DATE - COALESCE(os.last_order_date, c."SIGNUP_DATE"))) +
                0.3 * LEAST(100, COALESCE(os.order_count, 0) * 10) +
                0.3 * LEAST(100, c."LIFETIME_VALUE" / 50.0))::numeric, 2) as health_score
            FROM sf_data."SALES_DW__PUBLIC__CUSTOMERS" c
            LEFT JOIN order_stats os ON c."CUSTOMER_ID" = os."CUSTOMER_ID"
        )
        SELECT "CUSTOMER_NAME", "SEGMENT", "REGION",
            recency_score, frequency_score, monetary_score, health_score,
            CASE WHEN health_score >= 80 THEN 'Healthy'
                 WHEN health_score >= 50 THEN 'At Risk'
                 ELSE 'Critical' END as status
        FROM health
        ORDER BY health_score ASC
    """)
    all_rows = cur.fetchall()

    # Summary by segment
    segments = {}
    for r in all_rows:
        seg = r[1]
        if seg not in segments:
            segments[seg] = {"total": 0, "healthy": 0, "at_risk": 0, "critical": 0, "scores": []}
        segments[seg]["total"] += 1
        segments[seg]["scores"].append(float(r[6]))
        if r[7] == "Healthy":
            segments[seg]["healthy"] += 1
        elif r[7] == "At Risk":
            segments[seg]["at_risk"] += 1
        else:
            segments[seg]["critical"] += 1

    segment_summary = []
    for seg in sorted(segments.keys()):
        s = segments[seg]
        avg = round(sum(s["scores"]) / len(s["scores"]), 2)
        segment_summary.append({
            "Segment": seg,
            "Total_Customers": s["total"],
            "Healthy_Count": s["healthy"],
            "At_Risk_Count": s["at_risk"],
            "Critical_Count": s["critical"],
            "Avg_Health_Score": avg,
        })

    critical_rows = [r for r in all_rows if r[7] == "Critical"]
    below_15 = [r for r in all_rows if float(r[6]) < 15]

    cur.close()
    conn.close()

    return {
        "total_customers": len(all_rows),
        "critical_count": len(critical_rows),
        "at_risk_count": sum(1 for r in all_rows if r[7] == "At Risk"),
        "healthy_count": sum(1 for r in all_rows if r[7] == "Healthy"),
        "segment_summary": segment_summary,
        "below_15": below_15,
        "top5_lowest": all_rows[:5],
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Customer_Health.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    for sn in ["Health Scores", "Critical Accounts", "Summary by Segment"]:
        found = any(str_match(s, sn) for s in wb.sheetnames)
        check(f"Sheet '{sn}' exists", found, f"Found: {wb.sheetnames}")

    # --- Health Scores ---
    print("\n--- Health Scores ---")
    ws = get_sheet(wb, "Health Scores")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Health Scores total rows",
              num_close(len(rows), expected["total_customers"], 10),
              f"Expected ~{expected['total_customers']}, got {len(rows)}")

        # Check that first few rows are lowest scores
        if len(rows) >= 5:
            for i, exp_row in enumerate(expected["top5_lowest"][:3]):
                agent_row = rows[i]
                if agent_row and agent_row[6] is not None:
                    check(f"Row {i+1} health score",
                          num_close(agent_row[6], exp_row[6], 2.0),
                          f"Expected ~{exp_row[6]}, got {agent_row[6]}")

    # --- Critical Accounts ---
    print("\n--- Critical Accounts ---")
    ws = get_sheet(wb, "Critical Accounts")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Critical Accounts count",
              num_close(len(rows), expected["critical_count"], 20),
              f"Expected ~{expected['critical_count']}, got {len(rows)}")

    # --- Summary by Segment ---
    print("\n--- Summary by Segment ---")
    ws = get_sheet(wb, "Summary by Segment")
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Summary rows", len(rows) == 4,
              f"Expected 4, got {len(rows)}")

        for exp_seg in expected["segment_summary"]:
            seg = exp_seg["Segment"]
            matched = None
            for r in rows:
                if r and str_match(r[0], seg):
                    matched = r
                    break
            if matched:
                check(f"{seg} Total_Customers",
                      num_close(matched[1], exp_seg["Total_Customers"], 5),
                      f"Expected {exp_seg['Total_Customers']}, got {matched[1]}")
                check(f"{seg} Healthy_Count",
                      num_close(matched[2], exp_seg["Healthy_Count"], 5),
                      f"Expected {exp_seg['Healthy_Count']}, got {matched[2]}")
                check(f"{seg} Critical_Count",
                      num_close(matched[4], exp_seg["Critical_Count"], 10),
                      f"Expected {exp_seg['Critical_Count']}, got {matched[4]}")
                check(f"{seg} Avg_Health_Score",
                      num_close(matched[5], exp_seg["Avg_Health_Score"], 2.0),
                      f"Expected {exp_seg['Avg_Health_Score']}, got {matched[5]}")
            else:
                check(f"Segment '{seg}' found", False, "Not in output")


def check_notion():
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, properties FROM notion.pages
        WHERE properties::text ILIKE '%health%' OR properties::text ILIKE '%customer%dashboard%'
        ORDER BY id
    """)
    pages = cur.fetchall()
    check("Notion page with 'health' or 'dashboard' exists", len(pages) > 0,
          f"Found {len(pages)} matching pages")

    if pages:
        page_id = pages[0][0]
        cur.execute("""
            SELECT content FROM notion.blocks
            WHERE page_id = %s
        """, (page_id,))
        blocks = cur.fetchall()
        check("Notion page has content blocks", len(blocks) > 0,
              f"Found {len(blocks)} blocks")

    cur.close()
    conn.close()


def check_calendar(expected):
    print("\n=== Checking Calendar Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, description, start_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%follow%up%' OR LOWER(summary) LIKE '%follow-up%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()

    below_15_count = len(expected["below_15"])
    check(f"Calendar events for critical accounts (score < 15)",
          len(events) >= below_15_count,
          f"Expected >= {below_15_count}, got {len(events)}")

    # Check first few events match lowest-score customers
    if len(events) >= 3 and len(expected["below_15"]) >= 3:
        for i in range(min(3, len(events))):
            evt_summary = events[i][0] or ""
            exp_name = expected["below_15"][i][0]
            check(f"Event {i+1} mentions customer",
                  exp_name.lower() in evt_summary.lower(),
                  f"Expected '{exp_name}' in '{evt_summary}'")

    # Check events start from 2026-03-09
    if events:
        first_dt = events[0][2]
        if first_dt:
            first_date = first_dt.date() if hasattr(first_dt, 'date') else first_dt
            check("First event on or after 2026-03-09",
                  str(first_date) >= "2026-03-09",
                  f"First event date: {first_date}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Computing Expected Values ===")
    try:
        expected = compute_expected_values()
        print(f"  Total customers: {expected['total_customers']}")
        print(f"  Critical: {expected['critical_count']}")
        print(f"  Customers with score < 15: {len(expected['below_15'])}")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    check_excel(args.agent_workspace, expected)
    check_notion()
    check_calendar(expected)

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    success = pass_rate >= 0.8

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

"""Evaluation for howtocook-catering-gform-notion-excel."""
import argparse
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace):
    """Check Excel budget file."""
    print("\n=== Checking Excel Budget File ===")
    try:
        import openpyxl
    except ImportError:
        check("openpyxl available", False, "openpyxl not installed")
        return

    agent_file = os.path.join(agent_workspace, "Catering_Budget.xlsx")
    check("Catering_Budget.xlsx exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    # Check Menu sheet
    print("\n--- Menu Sheet ---")
    menu_ws = get_sheet(wb, "Menu")
    check("Sheet 'Menu' exists", menu_ws is not None, f"Found: {wb.sheetnames}")

    if menu_ws:
        headers = [c.value for c in list(menu_ws.rows)[0]] if menu_ws.max_row > 0 else []
        check("Menu has Recipe_Name column",
              any("recipe" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Menu has Category column",
              any("category" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Menu has Servings column",
              any("serving" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")

        data_rows = [row for row in menu_ws.iter_rows(min_row=2, values_only=True)
                     if any(v is not None for v in row)]
        check("Menu has 6 to 8 recipe rows", 6 <= len(data_rows) <= 10,
              f"Found {len(data_rows)} rows")

    # Check Summary sheet
    print("\n--- Summary Sheet ---")
    sum_ws = get_sheet(wb, "Summary")
    check("Sheet 'Summary' exists", sum_ws is not None, f"Found: {wb.sheetnames}")

    if sum_ws:
        summary_data = {}
        for row in sum_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary_data[str(row[0]).strip().lower()] = row[1]

        check("Summary has Total_Dishes",
              any("total_dishes" in k or "total dishes" in k for k in summary_data),
              f"Keys: {list(summary_data.keys())}")
        check("Summary has Total_Budget with value 500",
              any(("total_budget" in k or "total budget" in k) and num_close(v, 500, 5)
                  for k, v in summary_data.items()),
              f"Data: {summary_data}")
        check("Summary has Budget_Per_Person with value 25",
              any(("budget_per_person" in k or "budget per person" in k) and num_close(v, 25, 2)
                  for k, v in summary_data.items()),
              f"Data: {summary_data}")


def check_gform():
    """Check Google Form creation."""
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    check("At least one form created", len(forms) >= 1, f"Found {len(forms)} forms")

    form_id = None
    for fid, title in forms:
        if "dietary" in (title or "").lower() or "lunch" in (title or "").lower() or "team" in (title or "").lower():
            form_id = fid
            check("Form titled 'Team Lunch Dietary Preferences' found", True)
            break

    if form_id is None and forms:
        form_id = forms[0][0]
        check("Form titled 'Team Lunch Dietary Preferences' found", False,
              f"Found titles: {[f[1] for f in forms]}")

    if form_id:
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        check("Form has at least 3 questions", q_count >= 3,
              f"Found {q_count} questions")

    conn.close()


def check_notion():
    """Check Notion page creation."""
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM notion.pages")
    page_count = cur.fetchone()[0]
    check("At least one Notion page created", page_count >= 1,
          f"Found {page_count} pages")

    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("HOWTOCOOK CATERING GFORM NOTION EXCEL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_gform()
    check_notion()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

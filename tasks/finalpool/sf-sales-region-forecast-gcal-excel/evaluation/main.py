"""Evaluation for sf-sales-region-forecast-gcal-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    print("Checking Excel file...")
    agent_file = os.path.join(agent_ws, "Regional_Forecast.xlsx")
    gt_file = os.path.join(gt_dir, "Regional_Forecast.xlsx")

    if not os.path.exists(agent_file):
        all_errors.append("Regional_Forecast.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Region Performance sheet
        print("  Checking Region Performance sheet...")
        a_rows = load_sheet_rows(agent_wb, "Region Performance")
        g_rows = load_sheet_rows(gt_wb, "Region Performance")

        if a_rows is None:
            all_errors.append("Sheet 'Region Performance' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            if len(a_data) != 5:
                all_errors.append(f"Region Performance row count: {len(a_data)}, expected 5")

            a_lookup = {}
            for row in a_data:
                if row and row[0]:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing region: {g_row[0]}")
                    continue
                # Target (col 1)
                if not num_close(a_row[1], g_row[1], 100):
                    all_errors.append(f"{g_row[0]} Target: {a_row[1]} vs {g_row[1]}")
                # Actual (col 2)
                if not num_close(a_row[2], g_row[2], 100):
                    all_errors.append(f"{g_row[0]} Actual: {a_row[2]} vs {g_row[2]}")
            print("    Done.")

        # Check Summary sheet
        print("  Checking Summary sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Summary")
        g_rows2 = load_sheet_rows(gt_wb, "Summary")
        if a_rows2 is None:
            all_errors.append("Sheet 'Summary' not found in agent output")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            g_data2 = g_rows2[1:] if len(g_rows2) > 1 else []

            a_lookup2 = {}
            for row in a_data2:
                if row and row[0]:
                    a_lookup2[str(row[0]).strip().lower()] = row

            for g_row in g_data2:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup2.get(key)
                if a_row is None:
                    all_errors.append(f"Summary missing row: {g_row[0]}")
                    continue
                if not num_close(a_row[1], g_row[1], 500):
                    all_errors.append(f"Summary {g_row[0]}: {a_row[1]} vs {g_row[1]}")
            print("    Done.")

    # --- Check 2: GCal event ---
    print("Checking Google Calendar event...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (LOWER(summary) LIKE '%forecast%' OR LOWER(summary) LIKE '%quarterly%')
            AND start_datetime::date = '2026-03-31'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            # Broaden search
            cur.execute("""
                SELECT COUNT(*) FROM gcal.events
                WHERE LOWER(summary) LIKE '%forecast%' OR LOWER(summary) LIKE '%quarterly%'
            """)
            broad_count = cur.fetchone()[0]
            if broad_count == 0:
                all_errors.append("No GCal event for Quarterly Forecast Review found")
            else:
                print(f"    GCal event found (not on 2026-03-31, but {broad_count} matching events)")
        else:
            print(f"    GCal event on 2026-03-31 found ({count} events)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking GCal: {e}")

    # --- Check 3: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%vp_sales@company.com%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        if count == 0:
            all_errors.append("No email sent to vp_sales@company.com")
        else:
            print(f"    Email found ({count} messages)")
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

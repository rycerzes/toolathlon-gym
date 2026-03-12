"""Evaluation for terminal-sf-fetch-excel-notion-gcal.
Checks:
1. SLA_Compliance_Report.xlsx with 4 sheets and correct data
2. Notion database "SLA Compliance Dashboard" with entries
3. Google Calendar monthly SLA review events
4. sla_analyzer.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: SLA_Compliance_Report.xlsx ===")
    path = os.path.join(workspace, "SLA_Compliance_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # Ticket_Summary sheet
    ts_idx = next((i for i, s in enumerate(sheets_lower) if "ticket" in s or "summary" in s), 0)
    ws = wb[sheets[ts_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    check("Ticket_Summary has 3 priority rows", len(data_rows) >= 3, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    check("Contains High priority", "high" in all_text, f"Text: {all_text[:120]}")
    check("Contains compliance rate data",
          "compliance" in all_text or any(
              isinstance(c, (int, float)) and 50 <= c <= 100
              for r in rows[1:] for c in r if c is not None
          ), f"Headers: {rows[0] if rows else 'none'}")

    # Benchmark_Comparison sheet
    bc_idx = next((i for i, s in enumerate(sheets_lower) if "benchmark" in s or "comparison" in s), 1)
    if bc_idx < len(sheets):
        ws2 = wb[sheets[bc_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Benchmark_Comparison has at least 3 rows", len(data_rows2) >= 3, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Contains response time metric", "response" in all_text2, f"Text: {all_text2[:120]}")
        check("Contains satisfaction metric", "satisfaction" in all_text2, f"Text: {all_text2[:120]}")
        check("Contains status values", "meets" in all_text2 or "below" in all_text2,
              f"Text: {all_text2[:120]}")

    # Agent_Performance sheet
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "agent" in s or "performance" in s), 2)
    if ap_idx < len(sheets):
        ws3 = wb[sheets[ap_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Agent_Performance has at least 5 rows", len(data_rows3) >= 5, f"Found {len(data_rows3)}")

    # Monthly_Trend sheet
    mt_idx = next((i for i, s in enumerate(sheets_lower) if "monthly" in s or "trend" in s), 3)
    if mt_idx < len(sheets):
        ws4 = wb[sheets[mt_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Monthly_Trend has at least 3 rows", len(data_rows4) >= 3, f"Found {len(data_rows4)}")


def check_notion():
    print("\n=== Check 2: Notion SLA Compliance Dashboard ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM notion.databases")
    dbs = cur.fetchall()
    sla_db = None
    for db_id, title in dbs:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
        elif isinstance(title, str):
            try:
                parsed = json.loads(title)
                if isinstance(parsed, list):
                    title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                else:
                    title_str = str(title)
            except Exception:
                title_str = str(title)
        else:
            title_str = str(title) if title else ""
        if "sla" in title_str.lower() and ("compliance" in title_str.lower() or "dashboard" in title_str.lower()):
            sla_db = (db_id, title_str)
            break

    check("SLA Compliance Dashboard exists", sla_db is not None,
          f"Databases: {[d[1] for d in dbs]}")

    if sla_db:
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE parent->>'database_id' = %s
        """, (sla_db[0],))
        count = cur.fetchone()[0]
        check("Dashboard has metric entries", count >= 3, f"Found {count} pages")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 3: Monthly SLA Review Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE lower(summary) LIKE '%%sla%%review%%'
           OR lower(summary) LIKE '%%monthly%%sla%%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 3 SLA review events", len(events) >= 3, f"Found {len(events)} events")

    if events:
        summaries = " ".join(str(e[0]) for e in events).lower()
        check("Events mention monthly or SLA", "monthly" in summaries or "sla" in summaries,
              f"Summaries: {summaries[:150]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: sla_analyzer.py ===")
    path = os.path.join(workspace, "sla_analyzer.py")
    check("sla_analyzer.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative compliance rate values
    path = os.path.join(workspace, "SLA_Compliance_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative values in SLA report", not has_negative,
              "Found negative compliance/metric value")

    # Notion: no duplicate SLA dashboard databases
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE LOWER(title::text) LIKE '%%sla%%'
        """)
        db_count = cur.fetchone()[0]
        check("No duplicate SLA databases in Notion", db_count <= 1,
              f"Found {db_count} SLA databases")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

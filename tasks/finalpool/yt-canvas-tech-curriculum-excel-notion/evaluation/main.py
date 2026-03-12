"""
Evaluation for yt-canvas-tech-curriculum-excel-notion task.

Checks:
1. Curriculum_Video_Map.xlsx exists with Course_Videos and Summary sheets
2. Course_Videos has required columns and data from at least 3 courses
3. Summary sheet has Total_Courses_Mapped, Total_Videos_Recommended, Avg_Views_Recommended
4. Notion page 'Tech Course Video Resources' exists with course content
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Curriculum_Video_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Curriculum_Video_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Curriculum_Video_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Curriculum_Video_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Course_Videos sheet
    cv_idx = next((i for i, s in enumerate(sheet_names_lower) if "course_video" in s or "course video" in s), None)
    if cv_idx is None:
        record("Course_Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Course_Videos sheet exists", True)

    ws = wb[wb.sheetnames[cv_idx]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Has data rows", False, "Sheet is empty")
        return

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    has_course = any("course" in h for h in headers)
    has_video = any("video" in h for h in headers)
    has_view = any("view" in h for h in headers)
    has_duration = any("duration" in h for h in headers)
    record("Has required columns (Course, Video, View, Duration)",
           has_course and has_video and has_view and has_duration,
           f"Headers: {rows[0]}")

    data_rows = [r for r in rows[1:] if any(c for c in r)]
    record("Has at least 6 data rows (>=3 courses x 2 videos each)", len(data_rows) >= 6,
           f"Found {len(data_rows)} data rows")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    has_analytics = "analytics" in all_text or "algorithm" in all_text
    has_computing = "computing" in all_text
    has_data = "data" in all_text
    courses_found = sum([has_analytics, has_computing, has_data])
    record("Data covers at least 2 different course types", courses_found >= 2,
           f"Analytics/Alg:{has_analytics}, Computing:{has_computing}, Data:{has_data}")

    # Check Summary sheet
    sum_idx = next((i for i, s in enumerate(sheet_names_lower) if "summary" in s), None)
    if sum_idx is None:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Summary sheet exists", True)

    ws_sum = wb[wb.sheetnames[sum_idx]]
    sum_rows = list(ws_sum.iter_rows(values_only=True))
    sum_text = " ".join(str(c) for r in sum_rows for c in r if c).lower()
    has_total_courses = "total_courses" in sum_text or "total courses" in sum_text
    has_total_videos = "total_videos" in sum_text or "total videos" in sum_text
    has_avg = "avg" in sum_text
    record("Summary has required labels (Total_Courses, Total_Videos, Avg_Views)",
           has_total_courses and has_total_videos and has_avg,
           f"Summary text: {sum_text[:200]}")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Curriculum_Video_Map.xlsx")
    if os.path.isfile(gt_path):
        import openpyxl as opx
        gt_wb = opx.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows): break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None: continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                    else:
                        if not str_match(av, gv): ok = False; break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 2: Notion page 'Tech Course Video Resources' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, properties FROM notion.pages
        WHERE properties::text ILIKE '%tech course video%'
           OR properties::text ILIKE '%course video resource%'
        ORDER BY created_time DESC LIMIT 5
    """)
    pages = cur.fetchall()
    record("Notion page 'Tech Course Video Resources' exists", len(pages) > 0,
           "No matching page found")

    if pages:
        page_id = pages[0][0]
        cur.execute("""
            SELECT block_data FROM notion.blocks
            WHERE parent_id = %s
            ORDER BY position
        """, (page_id,))
        blocks = cur.fetchall()
        all_content = " ".join(str(b[0] or "") for b in blocks).lower()

        has_analytics = "analytics" in all_content or "algorithm" in all_content
        has_computing = "computing" in all_content
        has_data = "data" in all_content
        has_video = "video" in all_content or "fireship" in all_content
        record("Notion page has course sections with video content",
               (has_analytics or has_computing or has_data) and has_video,
               f"Content snippet: {all_content[:300]}")

    cur.close()
    conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""Evaluation for arxiv-reading-plan-excel-gcal-email.

Checks:
1. Reading_Plan.xlsx with Papers sheet (8 rows) and Schedule sheet (8 rows)
2. 8 Google Calendar events for reading sessions
3. Email to reading-group@lab.example.com with "LLM Agent Research Reading Plan" in subject
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0

ARXIV_IDS = ["2301.13379", "2302.01560", "2303.12528", "2305.10403",
             "2308.12950", "2309.17453", "2201.11903", "2310.06825"]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_ws, groundtruth_ws="."):
    print("\n=== Check 1: Reading_Plan.xlsx ===")
    path = os.path.join(agent_ws, "Reading_Plan.xlsx")
    check("File Reading_Plan.xlsx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel is readable", False, str(e))
        return

    # Check Papers sheet
    papers_ws = None
    for sname in wb.sheetnames:
        if "paper" in sname.lower():
            papers_ws = wb[sname]
            break
    check("Sheet 'Papers' exists", papers_ws is not None, f"Sheets: {wb.sheetnames}")

    if papers_ws is not None:
        rows = list(papers_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Papers sheet has 8 rows", len(non_empty) == 8, f"Got {len(non_empty)}")

        # Check arxiv IDs appear
        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        found_ids = sum(1 for arxiv_id in ARXIV_IDS if arxiv_id in all_text)
        check(f"Papers sheet contains at least 6 of 8 arXiv IDs",
              found_ids >= 6, f"Found {found_ids}/8 IDs in: {all_text[:200]}")

        # Check required columns exist in header row
        header_row = list(papers_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        header_text = " ".join(str(c).lower() for c in header_row if c is not None)
        check("Papers header has ArXiv_ID or arxiv column", "arxiv" in header_text, f"Header: {header_row}")
        check("Papers header has Title column", "title" in header_text, f"Header: {header_row}")
        check("Papers header has Session column", "session" in header_text or "assigned" in header_text, f"Header: {header_row}")

    # Check Schedule sheet
    schedule_ws = None
    for sname in wb.sheetnames:
        if "schedule" in sname.lower():
            schedule_ws = wb[sname]
            break
    check("Sheet 'Schedule' exists", schedule_ws is not None, f"Sheets: {wb.sheetnames}")

    if schedule_ws is not None:
        rows = list(schedule_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Schedule sheet has 8 rows", len(non_empty) == 8, f"Got {len(non_empty)}")

        # Check dates are in March-April 2026
        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        check("Schedule has March 2026 dates",
              "2026" in all_text and ("march" in all_text.lower() or "2026-03" in all_text or "03/09" in all_text),
              f"Date content: {all_text[:200]}")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_ws, "Reading_Plan.xlsx")
    if not os.path.isfile(gt_path):
        check("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws_sheet = gt_wb[gt_sheet_name]
        agent_ws_sheet = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws_sheet = wb[asn]
                break
        if agent_ws_sheet is None:
            check(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws_sheet.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws_sheet.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        check(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices_list = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices_list.append(len(gt_rows) - 1)
        for idx in check_indices_list:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        check(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1}",
                              False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    check(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                check(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: Google Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, summary, start_datetime FROM gcal.events
        WHERE summary ILIKE '%reading session%'
           OR summary ILIKE '%reading%session%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 8 'Reading Session' calendar events created",
          len(events) >= 8, f"Found {len(events)} events")

    if events:
        summaries = [e[1] for e in events]
        start_dates = [str(e[2]) for e in events]
        check("Events start in March 2026",
              any("2026-03" in d for d in start_dates),
              f"Dates: {start_dates[:4]}")
        check("Events cover 8 weeks (April 2026 included)",
              any("2026-04" in d for d in start_dates),
              f"Dates: {start_dates}")
        check("Events have 'Reading Session' in title",
              all("reading" in s.lower() for s in summaries[:8]),
              f"Titles: {summaries[:4]}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%reading plan%'
           OR subject ILIKE '%LLM%reading%'
           OR subject ILIKE '%LLM Agent%'
           OR to_addr::text ILIKE '%reading-group%'
        LIMIT 10
    """)
    rows = cur.fetchall()
    check("Email with reading plan subject found",
          len(rows) > 0, "No matching email found")

    if rows:
        to_addrs = [str(r[1]) for r in rows]
        check("Email sent to reading-group@lab.example.com",
              any("reading-group" in addr for addr in to_addrs),
              f"To addresses: {to_addrs}")
        subjects = [r[0] or "" for r in rows]
        check("Email subject contains 'Reading Plan' or 'LLM'",
              any("reading plan" in s.lower() or "llm" in s.lower() for s in subjects),
              f"Subjects: {subjects}")
        bodies = [str(r[2] or "").lower() for r in rows]
        check("Email body mentions total papers (8) or session dates",
              any("8" in b or "march" in b or "reading" in b for b in bodies),
              f"Body: {bodies[0][:200] if bodies else ''}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: arxiv-reading-plan-excel-gcal-email ===")

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_email()

    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"pass": PASS_COUNT, "fail": FAIL_COUNT}, f)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

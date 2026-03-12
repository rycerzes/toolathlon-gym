"""Evaluation for terminal-sf-hr-excel-word-gcal-email.
Checks:
1. Performance_Review_Report.xlsx with 4 sheets
2. Review_Policy_Memo.docx
3. Calendar events for department reviews
4. Email to HR leadership
5. rating_analysis.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

DEPARTMENTS = ["Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"]

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=0.5):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(workspace):
    print("\n=== Check 1: Performance_Review_Report.xlsx ===")
    path = os.path.join(workspace, "Performance_Review_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Department_Ratings
    dr_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s and "rating" in s), 0)
    ws1 = wb[sheets[dr_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Department_Ratings has 7 rows", len(data1) >= 7, f"Found {len(data1)}")

    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains Engineering dept", "engineering" in all_text1)
    check("Contains Sales dept", "sales" in all_text1)

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has avg_rating column", any("avg" in h and "rating" in h for h in headers),
              f"Headers: {rows1[0]}")
        check("Has pct_above_4 column", any("pct" in h and "4" in h for h in headers) or any("above" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Rating_Distribution
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "rating" in s and "dist" in s), 1)
    if rd_idx < len(sheets):
        ws2 = wb[sheets[rd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Rating_Distribution has 5 rows", len(data2) >= 5, f"Found {len(data2)}")

    # Review_Calendar
    rc_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s and "calendar" in s), 2)
    if rc_idx < len(sheets):
        ws3 = wb[sheets[rc_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Review_Calendar has 7 rows", len(data3) >= 7, f"Found {len(data3)}")

    # Policy_Summary
    ps_idx = next((i for i, s in enumerate(sheets_lower) if "policy" in s or "summary" in s), 3)
    if ps_idx < len(sheets):
        ws4 = wb[sheets[ps_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Policy_Summary has 5+ metrics", len(data4) >= 5, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has total_employees metric", "total" in all_text4 and "employee" in all_text4)
        # Check total employees dynamically from DB
        expected_total_employees = 50000  # fallback
        try:
            conn_db = psycopg2.connect(**DB_CONFIG)
            cur_db = conn_db.cursor()
            cur_db.execute("SELECT COUNT(*) FROM sf_data.employees")
            result = cur_db.fetchone()
            if result and result[0]:
                expected_total_employees = result[0]
            cur_db.close()
            conn_db.close()
        except Exception:
            pass
        for r in data4:
            if r and r[0] and "total_employee" in str(r[0]).lower():
                check("Total employees correct",
                      num_close(r[1], expected_total_employees, 1000),
                      f"Got {r[1]}, expected ~{expected_total_employees}")
                break


def check_word(workspace):
    print("\n=== Check 2: Review_Policy_Memo.docx ===")
    path = os.path.join(workspace, "Review_Policy_Memo.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    check("Mentions performance review", "performance review" in full_text or "review" in full_text)
    check("Mentions rating distribution", "rating" in full_text and "distribution" in full_text or "rating" in full_text)
    check("Mentions departments", any(d.lower() in full_text for d in DEPARTMENTS))
    check("Has substantial content", len(full_text) > 200, f"Length: {len(full_text)}")


def check_gcal():
    print("\n=== Check 3: Calendar Department Reviews ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        review_events = [e for e in events if "review" in str(e[0]).lower() or "performance" in str(e[0]).lower()]
        check("At least 7 review meeting events", len(review_events) >= 7,
              f"Found {len(review_events)} review events out of {len(events)} total")

        if review_events:
            summaries = " ".join(str(e[0]) for e in review_events).lower()
            dept_found = sum(1 for d in DEPARTMENTS if d.lower() in summaries)
            check("Events cover at least 5 departments", dept_found >= 5,
                  f"Found {dept_found} departments in summaries")
    except Exception as e:
        check("Gcal check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Check 4: Email to HR Leadership ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%hr_leadership%%'
               OR to_addr::text ILIKE '%%hr%%leadership%%'
               OR subject ILIKE '%%performance review%%'
               OR subject ILIKE '%%review cycle%%'
        """)
        emails = cur.fetchall()
        if not emails:
            cur.execute("""
                SELECT id, subject, to_addr, body_text
                FROM email.drafts
                WHERE to_addr::text ILIKE '%%hr%%'
                   OR subject ILIKE '%%performance%%'
                   OR subject ILIKE '%%review%%'
            """)
            emails = cur.fetchall()
        check("Email about performance review sent", len(emails) >= 1, "No matching email found")
        if emails:
            subject = str(emails[0][1]).lower() if emails[0][1] else ""
            check("Email subject relevant",
                  "performance" in subject or "review" in subject or "department" in subject,
                  f"Subject: {emails[0][1]}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: rating_analysis.py ===")
    path = os.path.join(workspace, "rating_analysis.py")
    check("rating_analysis.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative rating values
    path = os.path.join(workspace, "Performance_Review_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative values in Excel", not has_negative,
              "Found negative rating/count value")

    # Email: no emails sent to non-HR recipients about performance reviews
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE (subject ILIKE '%%performance review%%' OR subject ILIKE '%%review cycle%%')
              AND to_addr::text ILIKE '%%competitor%%'
        """)
        bad_count = cur.fetchone()[0]
        check("No performance review emails to competitor addresses", bad_count == 0,
              f"Found {bad_count}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_gcal()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""
Compute groundtruth Excel for wc-inventory-gold-repricing task.

Queries PostgreSQL for WooCommerce products and Yahoo Finance gold prices,
applies the mock supplier cost multipliers, and creates the groundtruth Excel.
"""
import json
import os
import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

# Mock website cost multipliers by category
COST_MULTIPLIERS = {
    "Electronics": 1.05,
    "Cameras": 1.08,
    "Audio": 0.97,
    "TV & Home Theater": 1.10,
    "Home Appliances": 1.02,
    "Watches": 0.95,
    "Headphones": 1.03,
    "Speakers": 1.00,
}


def get_meta_value(meta_data, key):
    """Extract a value from WC product meta_data list."""
    if not meta_data:
        return None
    for item in meta_data:
        if item.get("key") == key:
            return item.get("value")
    return None


def main():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # ── Fetch all WC products ──
    cur.execute("""
        SELECT id, name, sku, price, stock_quantity, stock_status, categories, meta_data
        FROM wc.products
        ORDER BY id
    """)
    products_raw = cur.fetchall()

    products = []
    for pid, name, sku, price, stock_qty, stock_status, categories, meta_data in products_raw:
        # Get category name (first category)
        cat_name = ""
        if categories and len(categories) > 0:
            cat_name = categories[0].get("name", "")

        supplier_name = get_meta_value(meta_data, "supplier_name") or ""
        supplier_contact = get_meta_value(meta_data, "supplier_contact") or ""
        threshold_str = get_meta_value(meta_data, "stock_threshold")
        threshold = int(threshold_str) if threshold_str else 0

        products.append({
            "id": pid,
            "name": name,
            "sku": sku,
            "price": float(price) if price else 0.0,
            "stock_quantity": int(stock_qty) if stock_qty is not None else 0,
            "stock_status": stock_status,
            "category": cat_name,
            "supplier_name": supplier_name,
            "supplier_email": supplier_contact,
            "stock_threshold": threshold,
        })

    # ── Fetch gold prices (last 30 trading days) ──
    cur.execute("""
        SELECT date, close
        FROM yf.stock_prices
        WHERE symbol = 'GC=F'
        ORDER BY date DESC
        LIMIT 30
    """)
    gold_rows = cur.fetchall()

    latest_gold_price = round(float(gold_rows[0][1]), 2)
    gold_30day_avg = round(sum(float(r[1]) for r in gold_rows) / len(gold_rows), 2)

    if latest_gold_price > gold_30day_avg:
        gold_trend = "Rising"
    elif latest_gold_price < gold_30day_avg:
        gold_trend = "Falling"
    else:
        gold_trend = "Stable"

    cur.close()
    conn.close()

    # ── Compute Inventory Status ──
    inventory_rows = []
    for p in products:
        stock_gap = max(p["stock_threshold"] - p["stock_quantity"], 0)
        if p["stock_quantity"] == 0:
            status = "Out of Stock"
        elif p["stock_quantity"] < p["stock_threshold"]:
            status = "Critical"
        else:
            status = "OK"

        inventory_rows.append({
            "Product_ID": p["id"],
            "Product_Name": p["name"][:60],
            "SKU": p["sku"],
            "Category": p["category"],
            "Current_Stock": p["stock_quantity"],
            "Stock_Threshold": p["stock_threshold"],
            "Stock_Gap": stock_gap,
            "Status": status,
            "Supplier_Name": p["supplier_name"],
            "Supplier_Email": p["supplier_email"],
        })

    # Sort by Stock_Gap descending, then Product_Name ascending
    inventory_rows.sort(key=lambda x: (-x["Stock_Gap"], x["Product_Name"]))

    # ── Compute Repricing Strategy ──
    repricing_rows = []
    for p in products:
        multiplier = COST_MULTIPLIERS.get(p["category"], 1.0)
        suggested = round(p["price"] * multiplier, 2)
        change = round(suggested - p["price"], 2)
        if change > 0:
            direction = "Increase"
        elif change < 0:
            direction = "Decrease"
        else:
            direction = "No Change"

        repricing_rows.append({
            "Product_ID": p["id"],
            "Product_Name": p["name"][:60],
            "Category": p["category"],
            "Current_Price": p["price"],
            "Cost_Multiplier": multiplier,
            "Suggested_Price": suggested,
            "Price_Change": change,
            "Change_Direction": direction,
        })

    # Sort by Product_ID ascending
    repricing_rows.sort(key=lambda x: x["Product_ID"])

    # ── Compute Gold Impact ──
    total_products = len(products)
    out_of_stock_count = sum(1 for p in products if p["stock_quantity"] == 0)
    critical_count = sum(1 for p in products if p["stock_quantity"] > 0 and p["stock_quantity"] < p["stock_threshold"])
    cats_with_increase = sum(1 for v in COST_MULTIPLIERS.values() if v > 1.0)

    gold_impact = [
        ("Latest_Gold_Price", latest_gold_price),
        ("Gold_30Day_Avg", gold_30day_avg),
        ("Gold_Trend", gold_trend),
        ("Total_Products", total_products),
        ("Out_Of_Stock_Count", out_of_stock_count),
        ("Critical_Stock_Count", critical_count),
        ("Categories_With_Price_Increase", cats_with_increase),
    ]

    # ── Compute Summary ──
    products_needing_restock = sum(1 for p in products if p["stock_quantity"] < p["stock_threshold"])
    gaps = [max(p["stock_threshold"] - p["stock_quantity"], 0) for p in products]
    positive_gaps = [g for g in gaps if g > 0]
    avg_stock_gap = round(sum(positive_gaps) / len(positive_gaps), 1) if positive_gaps else 0.0

    total_price_increases = sum(1 for r in repricing_rows if r["Price_Change"] > 0)
    total_price_decreases = sum(1 for r in repricing_rows if r["Price_Change"] < 0)

    # Avg price change pct across ALL products
    pct_changes = []
    for r in repricing_rows:
        if r["Current_Price"] > 0:
            pct = (r["Price_Change"] / r["Current_Price"]) * 100
            pct_changes.append(pct)
        else:
            pct_changes.append(0.0)
    avg_price_change_pct = round(sum(pct_changes) / len(pct_changes), 2) if pct_changes else 0.0

    summary = [
        ("Products_Needing_Restock", products_needing_restock),
        ("Avg_Stock_Gap", avg_stock_gap),
        ("Total_Price_Increases", total_price_increases),
        ("Total_Price_Decreases", total_price_decreases),
        ("Avg_Price_Change_Pct", avg_price_change_pct),
    ]

    # ── Create Excel ──
    wb = openpyxl.Workbook()

    # Sheet 1: Inventory Status
    ws1 = wb.active
    ws1.title = "Inventory Status"
    headers1 = ["Product_ID", "Product_Name", "SKU", "Category", "Current_Stock",
                 "Stock_Threshold", "Stock_Gap", "Status", "Supplier_Name", "Supplier_Email"]
    ws1.append(headers1)
    for row in inventory_rows:
        ws1.append([row[h] for h in headers1])

    # Sheet 2: Repricing Strategy
    ws2 = wb.create_sheet("Repricing Strategy")
    headers2 = ["Product_ID", "Product_Name", "Category", "Current_Price",
                 "Cost_Multiplier", "Suggested_Price", "Price_Change", "Change_Direction"]
    ws2.append(headers2)
    for row in repricing_rows:
        ws2.append([row[h] for h in headers2])

    # Sheet 3: Gold Impact
    ws3 = wb.create_sheet("Gold Impact")
    for label, value in gold_impact:
        ws3.append([label, value])

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")
    for label, value in summary:
        ws4.append([label, value])

    # Save
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "groundtruth_workspace")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "inventory_repricing_report.xlsx")
    wb.save(out_path)
    print(f"Groundtruth saved to: {out_path}")

    # Print summary for verification
    print(f"\n=== Summary ===")
    print(f"Total products: {total_products}")
    print(f"Out of stock: {out_of_stock_count}")
    print(f"Critical: {critical_count}")
    print(f"Products needing restock: {products_needing_restock}")
    print(f"Avg stock gap: {avg_stock_gap}")
    print(f"Latest gold: {latest_gold_price}")
    print(f"Gold 30-day avg: {gold_30day_avg}")
    print(f"Gold trend: {gold_trend}")
    print(f"Categories with price increase: {cats_with_increase}")
    print(f"Total price increases: {total_price_increases}")
    print(f"Total price decreases: {total_price_decreases}")
    print(f"Avg price change pct: {avg_price_change_pct}")

    # Print first few inventory rows for verification
    print(f"\n=== Top 5 Inventory Status (by Stock_Gap desc) ===")
    for r in inventory_rows[:5]:
        print(f"  ID={r['Product_ID']}, Name={r['Product_Name'][:40]}, Gap={r['Stock_Gap']}, Status={r['Status']}")

    # Print some repricing rows
    print(f"\n=== First 5 Repricing rows ===")
    for r in repricing_rows[:5]:
        print(f"  ID={r['Product_ID']}, Price={r['Current_Price']}, Mult={r['Cost_Multiplier']}, Suggested={r['Suggested_Price']}, Change={r['Price_Change']}, Dir={r['Change_Direction']}")


if __name__ == "__main__":
    main()

"""
Evaluation script for wc-inventory-gold-repricing task.

Computes expected values dynamically from PostgreSQL:
  - wc.products: inventory data (stock levels, thresholds, categories, prices)
  - yf.stock_prices: gold prices (GC=F)

Category cost multipliers come from the mock supplier portal (static HTML
served during the task). These are hardcoded constants since they are
defined by the task itself and do not change.

Falls back to static groundtruth Excel if PostgreSQL is unavailable.

Checks:
1. Inventory Status sheet - row count, columns, spot-check values
2. Repricing Strategy sheet - product prices and multipliers
3. Gold Impact sheet - gold prices and counts
4. Summary sheet - aggregate values

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0

# --- Constants from the mock supplier portal (http://localhost:30148) ---
CATEGORY_MULTIPLIERS = {
    'Electronics': 1.05,
    'Cameras': 1.08,
    'Audio': 0.97,
    'TV & Home Theater': 1.10,
    'Home Appliances': 1.02,
    'Watches': 0.95,
    'Headphones': 1.03,
    'Speakers': 1.00,
}

DB_CONFIG = dict(host=os.environ.get('PGHOST', 'localhost'), port=5432, database='toolathlon_gym', user='postgres', password='postgres')


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    """Find sheet case-insensitively."""
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected_from_db():
    """Compute all expected values from PostgreSQL."""
    try:
        import psycopg2
    except ImportError:
        return None

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    # --- Gold data ---
    cur.execute("SELECT close FROM yf.stock_prices WHERE symbol='GC=F' ORDER BY date DESC LIMIT 1")
    row = cur.fetchone()
    if not row:
        conn.close()
        return None
    latest_gold = round(float(row[0]), 2)

    cur.execute("SELECT close FROM yf.stock_prices WHERE symbol='GC=F' ORDER BY date DESC LIMIT 30")
    gold_prices = [float(r[0]) for r in cur.fetchall()]
    if len(gold_prices) < 30:
        conn.close()
        return None
    gold_avg = round(sum(gold_prices) / len(gold_prices), 2)
    gold_trend = 'Rising' if latest_gold > gold_avg else ('Falling' if latest_gold < gold_avg else 'Stable')

    # --- Product data ---
    cur.execute(
        "SELECT id, name, sku, price, stock_quantity, stock_status, categories, meta_data "
        "FROM wc.products ORDER BY id")
    products = cur.fetchall()
    if not products:
        conn.close()
        return None

    inventory_rows = []  # (pid, name60, sku, category, stock, threshold, gap, status, supplier, email)
    repricing_rows = []  # (pid, name60, category, price, mult, sugg, change, direction)

    oos_count = 0
    critical_count = 0
    restock_count = 0
    gaps_positive = []
    price_inc = 0
    price_dec = 0
    pct_changes = []

    for p in products:
        pid, name, sku, price, stock, stock_status, cats, meta = p
        cat = cats[0]['name'] if cats else 'Unknown'
        name60 = (name or '')[:60]
        stock = stock if stock is not None else 0

        threshold = 0
        supplier_name = ''
        supplier_email = ''
        for m in (meta or []):
            k = m.get('key', '')
            if k == 'stock_threshold':
                threshold = int(m['value'])
            elif k == 'supplier_name':
                supplier_name = m['value']
            elif k == 'supplier_contact':
                supplier_email = m['value']

        gap = max(0, threshold - stock)
        if stock == 0:
            status = 'Out of Stock'
            oos_count += 1
        elif stock < threshold:
            status = 'Critical'
            critical_count += 1
        else:
            status = 'OK'

        if gap > 0:
            restock_count += 1
            gaps_positive.append(gap)

        inventory_rows.append((pid, name60, sku, cat, stock, threshold, gap, status, supplier_name, supplier_email))

        mult = CATEGORY_MULTIPLIERS.get(cat, 1.0)
        sugg = round(float(price) * mult, 2)
        change = round(sugg - float(price), 2)
        if change > 0:
            direction = 'Increase'
            price_inc += 1
        elif change < 0:
            direction = 'Decrease'
            price_dec += 1
        else:
            direction = 'No Change'
        pct_changes.append(change / float(price) * 100)

        repricing_rows.append((pid, name60, cat, float(price), mult, sugg, change, direction))

    # Sort inventory by Stock_Gap desc, then name asc
    inventory_rows.sort(key=lambda r: (-r[6], r[1]))
    # Repricing sorted by Product_ID asc
    repricing_rows.sort(key=lambda r: r[0])

    avg_gap = round(sum(gaps_positive) / len(gaps_positive), 1) if gaps_positive else 0
    avg_pct = round(sum(pct_changes) / len(pct_changes), 2)
    cats_increase = sum(1 for v in CATEGORY_MULTIPLIERS.values() if v > 1.0)

    conn.close()

    return {
        'inventory': inventory_rows,
        'repricing': repricing_rows,
        'gold_impact': {
            'latest_gold_price': latest_gold,
            'gold_30day_avg': gold_avg,
            'gold_trend': gold_trend,
            'total_products': len(products),
            'out_of_stock_count': oos_count,
            'critical_stock_count': critical_count,
            'categories_with_price_increase': cats_increase,
        },
        'summary': {
            'products_needing_restock': restock_count,
            'avg_stock_gap': avg_gap,
            'total_price_increases': price_inc,
            'total_price_decreases': price_dec,
            'avg_price_change_pct': avg_pct,
        },
    }


def check_excel_db(agent_workspace, expected):
    """Check the Excel output against DB-computed expected values."""
    print("\n=== Checking Excel Output (vs PostgreSQL) ===")

    agent_file = os.path.join(agent_workspace, "inventory_repricing_report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    expected_sheets = ["Inventory Status", "Repricing Strategy", "Gold Impact", "Summary"]
    for sheet_name in expected_sheets:
        found = get_sheet(agent_wb, sheet_name) is not None
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {agent_wb.sheetnames}")

    # ── Sheet 1: Inventory Status ──
    print("\n--- Inventory Status ---")
    agent_ws = get_sheet(agent_wb, "Inventory Status")
    inv = expected['inventory']
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Inventory Status row count", len(agent_rows) == len(inv),
              f"Expected {len(inv)}, got {len(agent_rows)}")

        gt_headers = ['Product_ID', 'Product_Name', 'SKU', 'Category', 'Current_Stock',
                       'Stock_Threshold', 'Stock_Gap', 'Status', 'Supplier_Name', 'Supplier_Email']
        agent_headers = [c.value for c in agent_ws[1]]
        check("Inventory Status has correct columns",
              len(agent_headers) >= len(gt_headers),
              f"Expected {len(gt_headers)} columns, got {len(agent_headers)}")

        # Build lookup by Product_ID
        gt_by_id = {r[0]: r for r in inv}
        agent_by_id = {}
        for row in agent_rows:
            if row and row[0] is not None:
                try:
                    agent_by_id[int(row[0])] = row
                except (ValueError, TypeError):
                    pass

        # Spot-check out-of-stock products
        for pid, gt_row in gt_by_id.items():
            if gt_row[7] == "Out of Stock":
                agent_row = agent_by_id.get(pid)
                if agent_row:
                    check(f"Product {pid} status is 'Out of Stock'",
                          str_match(agent_row[7], "Out of Stock"),
                          f"Expected 'Out of Stock', got '{agent_row[7]}'")
                    check(f"Product {pid} Stock_Gap",
                          num_close(agent_row[6], gt_row[6], 1),
                          f"Expected {gt_row[6]}, got {agent_row[6]}")
                else:
                    check(f"Product {pid} found in agent output", False)

        # Check a few Critical products
        critical_checked = 0
        for pid, gt_row in gt_by_id.items():
            if gt_row[7] == "Critical" and critical_checked < 3:
                agent_row = agent_by_id.get(pid)
                if agent_row:
                    check(f"Product {pid} status is 'Critical'",
                          str_match(agent_row[7], "Critical"),
                          f"Expected 'Critical', got '{agent_row[7]}'")
                critical_checked += 1

        # Check a few OK products
        ok_checked = 0
        for pid, gt_row in gt_by_id.items():
            if gt_row[7] == "OK" and ok_checked < 2:
                agent_row = agent_by_id.get(pid)
                if agent_row:
                    check(f"Product {pid} status is 'OK'",
                          str_match(agent_row[7], "OK"),
                          f"Expected 'OK', got '{agent_row[7]}'")
                    check(f"Product {pid} Stock_Gap is 0",
                          num_close(agent_row[6], 0, 0),
                          f"Expected 0, got {agent_row[6]}")
                ok_checked += 1

        # Check sort order
        if len(agent_rows) >= 2:
            first_gap = agent_rows[0][6] if agent_rows[0][6] is not None else 0
            second_gap = agent_rows[1][6] if agent_rows[1][6] is not None else 0
            check("Inventory sorted by Stock_Gap descending",
                  float(first_gap) >= float(second_gap),
                  f"First gap={first_gap}, second gap={second_gap}")

    # ── Sheet 2: Repricing Strategy ──
    print("\n--- Repricing Strategy ---")
    agent_ws = get_sheet(agent_wb, "Repricing Strategy")
    rep = expected['repricing']
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Repricing Strategy row count", len(agent_rows) == len(rep),
              f"Expected {len(rep)}, got {len(agent_rows)}")

        gt_by_id = {r[0]: r for r in rep}
        agent_by_id = {}
        for row in agent_rows:
            if row and row[0] is not None:
                try:
                    agent_by_id[int(row[0])] = row
                except (ValueError, TypeError):
                    pass

        checked = 0
        for gt_row in rep:
            if checked >= 8:
                break
            pid = gt_row[0]
            agent_row = agent_by_id.get(pid)
            if agent_row:
                # Cost_Multiplier (col 4)
                check(f"Product {pid} Cost_Multiplier",
                      num_close(agent_row[4], gt_row[4], 0.01),
                      f"Expected {gt_row[4]}, got {agent_row[4]}")
                # Suggested_Price (col 5)
                check(f"Product {pid} Suggested_Price",
                      num_close(agent_row[5], gt_row[5], 1.0),
                      f"Expected {gt_row[5]}, got {agent_row[5]}")
                # Change_Direction (col 7)
                check(f"Product {pid} Change_Direction",
                      str_match(agent_row[7], gt_row[7]),
                      f"Expected '{gt_row[7]}', got '{agent_row[7]}'")
                checked += 1

        if len(agent_rows) >= 2:
            check("Repricing sorted by Product_ID ascending",
                  int(agent_rows[0][0]) <= int(agent_rows[1][0]),
                  f"First ID={agent_rows[0][0]}, second ID={agent_rows[1][0]}")

    # ── Sheet 3: Gold Impact ──
    print("\n--- Gold Impact ---")
    agent_ws = get_sheet(agent_wb, "Gold Impact")
    gi = expected['gold_impact']
    if agent_ws:
        agent_data = {}
        for row in agent_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key in ["latest_gold_price", "gold_30day_avg"]:
            gt_val = gi[key]
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            check(f"Gold Impact '{key}'",
                  num_close(agent_val, gt_val, 5.0),
                  f"Expected {gt_val}, got {agent_val}")

        gt_trend = gi['gold_trend']
        agent_trend = agent_data.get("gold_trend")
        if agent_trend is None:
            for ak, av in agent_data.items():
                if "trend" in ak:
                    agent_trend = av
                    break
        check("Gold Impact 'gold_trend'",
              str_match(agent_trend, gt_trend),
              f"Expected '{gt_trend}', got '{agent_trend}'")

        for key in ["total_products", "out_of_stock_count", "critical_stock_count",
                     "categories_with_price_increase"]:
            gt_val = gi[key]
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            check(f"Gold Impact '{key}'",
                  num_close(agent_val, gt_val, 1),
                  f"Expected {gt_val}, got {agent_val}")

    # ── Sheet 4: Summary ──
    print("\n--- Summary ---")
    agent_ws = get_sheet(agent_wb, "Summary")
    sm = expected['summary']
    if agent_ws:
        agent_data = {}
        for row in agent_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key, tol in [
            ("products_needing_restock", 1),
            ("avg_stock_gap", 1.0),
            ("total_price_increases", 1),
            ("total_price_decreases", 1),
            ("avg_price_change_pct", 0.5),
        ]:
            gt_val = sm[key]
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            check(f"Summary '{key}'",
                  num_close(agent_val, gt_val, tol),
                  f"Expected {gt_val}, got {agent_val}")

    return True


def check_excel_gt(agent_workspace, groundtruth_workspace):
    """Fallback: Check the Excel output against static groundtruth."""
    print("\n=== Checking Excel Output (vs groundtruth Excel) ===")

    agent_file = os.path.join(agent_workspace, "inventory_repricing_report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "inventory_repricing_report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    check("Groundtruth file exists", os.path.isfile(gt_file),
          f"Expected {gt_file}")
    if not os.path.isfile(gt_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
        gt_wb = openpyxl.load_workbook(gt_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    expected_sheets = ["Inventory Status", "Repricing Strategy", "Gold Impact", "Summary"]
    for sheet_name in expected_sheets:
        found = get_sheet(agent_wb, sheet_name) is not None
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {agent_wb.sheetnames}")

    all_passed = True

    # ── Sheet 1: Inventory Status ──
    print("\n--- Inventory Status ---")
    agent_ws = get_sheet(agent_wb, "Inventory Status")
    gt_ws = get_sheet(gt_wb, "Inventory Status")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Inventory Status row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        gt_headers = [c.value for c in gt_ws[1]]
        agent_headers = [c.value for c in agent_ws[1]]
        check("Inventory Status has correct columns",
              len(agent_headers) >= len(gt_headers),
              f"Expected {len(gt_headers)} columns, got {len(agent_headers)}")

        gt_by_id = {}
        for row in gt_rows:
            if row and row[0] is not None:
                gt_by_id[int(row[0])] = row

        agent_by_id = {}
        for row in agent_rows:
            if row and row[0] is not None:
                try:
                    agent_by_id[int(row[0])] = row
                except (ValueError, TypeError):
                    pass

        for pid, gt_row in gt_by_id.items():
            if gt_row[7] and str(gt_row[7]).strip().lower() == "out of stock":
                agent_row = agent_by_id.get(pid)
                if agent_row:
                    check(f"Product {pid} status is 'Out of Stock'",
                          str_match(agent_row[7], "Out of Stock"),
                          f"Expected 'Out of Stock', got '{agent_row[7]}'")
                    check(f"Product {pid} Stock_Gap",
                          num_close(agent_row[6], gt_row[6], 1),
                          f"Expected {gt_row[6]}, got {agent_row[6]}")
                else:
                    check(f"Product {pid} found in agent output", False)
                    all_passed = False

        critical_checked = 0
        for pid, gt_row in gt_by_id.items():
            if gt_row[7] and str(gt_row[7]).strip().lower() == "critical" and critical_checked < 3:
                agent_row = agent_by_id.get(pid)
                if agent_row:
                    check(f"Product {pid} status is 'Critical'",
                          str_match(agent_row[7], "Critical"),
                          f"Expected 'Critical', got '{agent_row[7]}'")
                critical_checked += 1

        ok_checked = 0
        for pid, gt_row in gt_by_id.items():
            if gt_row[7] and str(gt_row[7]).strip().lower() == "ok" and ok_checked < 2:
                agent_row = agent_by_id.get(pid)
                if agent_row:
                    check(f"Product {pid} status is 'OK'",
                          str_match(agent_row[7], "OK"),
                          f"Expected 'OK', got '{agent_row[7]}'")
                    check(f"Product {pid} Stock_Gap is 0",
                          num_close(agent_row[6], 0, 0),
                          f"Expected 0, got {agent_row[6]}")
                ok_checked += 1

        if len(agent_rows) >= 2:
            first_gap = agent_rows[0][6] if agent_rows[0][6] is not None else 0
            second_gap = agent_rows[1][6] if agent_rows[1][6] is not None else 0
            check("Inventory sorted by Stock_Gap descending",
                  float(first_gap) >= float(second_gap),
                  f"First gap={first_gap}, second gap={second_gap}")
    else:
        all_passed = False

    # ── Sheet 2: Repricing Strategy ──
    print("\n--- Repricing Strategy ---")
    agent_ws = get_sheet(agent_wb, "Repricing Strategy")
    gt_ws = get_sheet(gt_wb, "Repricing Strategy")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Repricing Strategy row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        agent_by_id = {}
        for row in agent_rows:
            if row and row[0] is not None:
                try:
                    agent_by_id[int(row[0])] = row
                except (ValueError, TypeError):
                    pass

        checked = 0
        for gt_row in gt_rows:
            if checked >= 8:
                break
            pid = gt_row[0]
            agent_row = agent_by_id.get(int(pid))
            if agent_row:
                check(f"Product {pid} Cost_Multiplier",
                      num_close(agent_row[4], gt_row[4], 0.01),
                      f"Expected {gt_row[4]}, got {agent_row[4]}")
                check(f"Product {pid} Suggested_Price",
                      num_close(agent_row[5], gt_row[5], 1.0),
                      f"Expected {gt_row[5]}, got {agent_row[5]}")
                check(f"Product {pid} Change_Direction",
                      str_match(agent_row[7], gt_row[7]),
                      f"Expected '{gt_row[7]}', got '{agent_row[7]}'")
                checked += 1

        if len(agent_rows) >= 2:
            check("Repricing sorted by Product_ID ascending",
                  int(agent_rows[0][0]) <= int(agent_rows[1][0]),
                  f"First ID={agent_rows[0][0]}, second ID={agent_rows[1][0]}")
    else:
        all_passed = False

    # ── Sheet 3: Gold Impact ──
    print("\n--- Gold Impact ---")
    agent_ws = get_sheet(agent_wb, "Gold Impact")
    gt_ws = get_sheet(gt_wb, "Gold Impact")
    if agent_ws and gt_ws:
        gt_data = {}
        for row in gt_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]

        agent_data = {}
        for row in agent_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key in ["latest_gold_price", "gold_30day_avg"]:
            gt_val = gt_data.get(key)
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            check(f"Gold Impact '{key}'",
                  num_close(agent_val, gt_val, 5.0),
                  f"Expected {gt_val}, got {agent_val}")

        gt_trend = gt_data.get("gold_trend")
        agent_trend = agent_data.get("gold_trend")
        if agent_trend is None:
            for ak, av in agent_data.items():
                if "trend" in ak:
                    agent_trend = av
                    break
        check("Gold Impact 'gold_trend'",
              str_match(agent_trend, gt_trend),
              f"Expected '{gt_trend}', got '{agent_trend}'")

        for key in ["total_products", "out_of_stock_count", "critical_stock_count",
                     "categories_with_price_increase"]:
            gt_val = gt_data.get(key)
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            check(f"Gold Impact '{key}'",
                  num_close(agent_val, gt_val, 1),
                  f"Expected {gt_val}, got {agent_val}")
    else:
        all_passed = False

    # ── Sheet 4: Summary ──
    print("\n--- Summary ---")
    agent_ws = get_sheet(agent_wb, "Summary")
    gt_ws = get_sheet(gt_wb, "Summary")
    if agent_ws and gt_ws:
        gt_data = {}
        for row in gt_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]

        agent_data = {}
        for row in agent_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key, tol in [
            ("products_needing_restock", 1),
            ("avg_stock_gap", 1.0),
            ("total_price_increases", 1),
            ("total_price_decreases", 1),
            ("avg_price_change_pct", 0.5),
        ]:
            gt_val = gt_data.get(key)
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            check(f"Summary '{key}'",
                  num_close(agent_val, gt_val, tol),
                  f"Expected {gt_val}, got {agent_val}")
    else:
        all_passed = False

    return all_passed


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    # Try DB-computed expected values first
    db_expected = compute_expected_from_db()
    use_db = db_expected is not None

    if use_db:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        check_excel_db(agent_workspace, db_expected)
    else:
        print("INFO: Falling back to static groundtruth Excel file")
        check_excel_gt(agent_workspace, groundtruth_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Source: {'postgresql' if use_db else 'groundtruth_excel'}")
    print(f"  Overall: {'PASS' if FAIL_COUNT == 0 else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": FAIL_COUNT == 0,
            "source": "postgresql" if use_db else "groundtruth_excel",
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return FAIL_COUNT == 0, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

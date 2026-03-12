"""Evaluation for sf-ticket-sla-compliance."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "SLA_Compliance_Report.xlsx")
    gt_file = os.path.join(gt_dir, "SLA_Compliance_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # --- Check Compliance sheet ---
    print("  Checking Compliance sheet...")
    a_rows = load_sheet_rows(agent_wb, "Compliance")
    g_rows = load_sheet_rows(gt_wb, "Compliance")
    if a_rows is None:
        all_errors.append("Sheet 'Compliance' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Compliance' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Compliance: {g_row[0]}")
                continue

            # Total_Tickets (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 10):
                    all_errors.append(f"{key}.Total_Tickets: {a_row[1]} vs {g_row[1]}")

            # Avg_Response_Hours (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    all_errors.append(f"{key}.Avg_Response_Hours: {a_row[2]} vs {g_row[2]}")

            # SLA_Target_Hours (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.5):
                    all_errors.append(f"{key}.SLA_Target_Hours: {a_row[3]} vs {g_row[3]}")

            # Compliant_Count (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 10):
                    all_errors.append(f"{key}.Compliant_Count: {a_row[4]} vs {g_row[4]}")

            # Compliance_Rate (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1.0):
                    all_errors.append(f"{key}.Compliance_Rate: {a_row[5]} vs {g_row[5]}")

        if not all_errors:
            print("    PASS")
        else:
            print(f"    ERRORS: {len(all_errors)}")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Summary: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            # For numeric metrics use tolerance, for string metrics use case-insensitive match
            try:
                float(a_val); float(g_val)
                if not num_close(a_val, g_val, 1.0):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val} (tol=1.0)")
            except (TypeError, ValueError):
                if not str_match(a_val, g_val):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val}")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check email sent ---
    print("  Checking email...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text
            FROM email.messages m
            JOIN email.folders f ON m.folder_id = f.id
            WHERE f.name IN ('Sent', 'INBOX.Sent', 'Sent Messages', 'Sent Items')
               OR m.from_addr LIKE '%%ops-report%%'
        """)
        sent = cur.fetchall()

        if not sent:
            # Also check sent_log
            cur.execute("""
                SELECT m.subject, m.to_addr, m.body_text
                FROM email.messages m
                JOIN email.sent_log sl ON m.id = sl.message_id
            """)
            sent = cur.fetchall()

        found_email = False
        for subj, to_addr, body in sent:
            subj_str = str(subj or "").lower()
            to_str = str(to_addr or "").lower()
            if "sla" in subj_str and "ops-manager" in to_str:
                found_email = True
                break

        if not found_email:
            # Check INBOX of ops-manager
            cur.execute("""
                SELECT m.subject, m.to_addr, m.body_text, m.from_addr
                FROM email.messages m
                WHERE LOWER(m.to_addr::text) LIKE '%%ops-manager%%'
            """)
            inbox_msgs = cur.fetchall()
            for subj, to_addr, body, from_addr in inbox_msgs:
                if "sla" in str(subj or "").lower():
                    found_email = True
                    break

        if not found_email:
            all_errors.append("No email with 'SLA' in subject sent to ops-manager@company.com")
        else:
            print("    PASS")

        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Email check error: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

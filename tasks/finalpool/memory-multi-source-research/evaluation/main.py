"""
Evaluation script for memory-multi-source-research task.

Checks:
1. Excel file has Paper Summary and Research Progress sheets with correct data
2. Word document has report with required sections
3. Memory file has been updated with entities

Usage:
    python -m evaluation.main --agent_workspace <path> --groundtruth_workspace <path>
"""
import argparse
import json
import os
import re
import sys

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def normalize(text):
    return re.sub(r'\s+', ' ', text.lower().strip())


EXPECTED_PAPER_FRAGMENTS = [
    "concrete problems in ai safety",
    "ai safety via debate",
    "risks from learned optimization",
    "red teaming language models",
    "alignment of language agents",
]


def check_excel(agent_workspace):
    """Check Research_Analysis.xlsx has correct data."""
    print("\n=== Checking Excel Output ===")

    import openpyxl

    excel_path = os.path.join(agent_workspace, "Research_Analysis.xlsx")
    if not os.path.isfile(excel_path):
        check("Research_Analysis.xlsx exists", False, f"Not found: {excel_path}")
        return

    check("Research_Analysis.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    # Check Paper Summary sheet
    paper_sheet = None
    for name in wb.sheetnames:
        if "paper" in name.lower() and "summary" in name.lower():
            paper_sheet = wb[name]
            break

    if paper_sheet is None:
        check("Paper Summary sheet exists", False,
              f"Sheets found: {wb.sheetnames}")
    else:
        check("Paper Summary sheet exists", True)

        rows = list(paper_sheet.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        check("Paper Summary has at least 5 data rows",
              len(data_rows) >= 5,
              f"Found {len(data_rows)} data rows")

        # Check that target papers appear
        all_text = " ".join(
            str(cell).lower() for row in data_rows for cell in row if cell is not None
        )
        found_count = 0
        for title_fragment in EXPECTED_PAPER_FRAGMENTS:
            if title_fragment in all_text:
                found_count += 1
        check("Paper Summary contains at least 5 expected papers",
              found_count >= 5,
              f"Found {found_count} of {len(EXPECTED_PAPER_FRAGMENTS)} expected papers")

        # Check header has expected columns
        if rows:
            header = " ".join(str(h).lower() for h in rows[0] if h is not None)
            has_title_col = "title" in header
            has_year_col = "year" in header
            has_citations_col = "citation" in header
            check("Paper Summary has Title column", has_title_col,
                  f"Header: {header}")
            check("Paper Summary has Year column", has_year_col,
                  f"Header: {header}")
            check("Paper Summary has Citations column", has_citations_col,
                  f"Header: {header}")

    # Check Research Progress sheet
    progress_sheet = None
    for name in wb.sheetnames:
        if "progress" in name.lower() or "research" in name.lower():
            if paper_sheet is None or wb[name] != paper_sheet:
                progress_sheet = wb[name]
                break

    if progress_sheet is None:
        # Try matching just "progress"
        for name in wb.sheetnames:
            if "progress" in name.lower():
                progress_sheet = wb[name]
                break

    if progress_sheet is None:
        check("Research Progress sheet exists", False,
              f"Sheets found: {wb.sheetnames}")
    else:
        check("Research Progress sheet exists", True)

        rows = list(progress_sheet.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        check("Research Progress has at least 2 data rows",
              len(data_rows) >= 2,
              f"Found {len(data_rows)} data rows")

        if rows:
            header = " ".join(str(h).lower() for h in rows[0] if h is not None)
            has_phase = "phase" in header
            has_status = "status" in header
            check("Research Progress has Phase column", has_phase,
                  f"Header: {header}")
            check("Research Progress has Status column", has_status,
                  f"Header: {header}")


def check_word(agent_workspace):
    """Check Research_Report.docx has required content."""
    print("\n=== Checking Word Report ===")

    from docx import Document

    docx_path = os.path.join(agent_workspace, "Research_Report.docx")
    if not os.path.isfile(docx_path):
        check("Research_Report.docx exists", False, f"Not found: {docx_path}")
        return

    check("Research_Report.docx exists", True)

    try:
        doc = Document(docx_path)
        full_text = "\n".join([para.text for para in doc.paragraphs])
    except Exception as e:
        check("Word document readable", False, str(e))
        return

    normalized = normalize(full_text)

    check("Report has at least 500 characters",
          len(full_text.strip()) >= 500,
          f"Document has {len(full_text.strip())} characters")

    # Check sections
    check("Report has Introduction section",
          "introduction" in normalized,
          "No Introduction section found")

    has_lit_review = ("literature review" in normalized or "literature" in normalized)
    check("Report has Literature Review section",
          has_lit_review,
          "No Literature Review section found")

    has_key_findings = ("key findings" in normalized or "findings" in normalized)
    check("Report has Key Findings section",
          has_key_findings,
          "No Key Findings section found")

    has_gaps = ("research gaps" in normalized or "gaps" in normalized)
    check("Report has Research Gaps section",
          has_gaps,
          "No Research Gaps section found")

    check("Report has Conclusion section",
          "conclusion" in normalized,
          "No Conclusion section found")

    # Check paper mentions
    paper_mention_count = 0
    for title_fragment in EXPECTED_PAPER_FRAGMENTS:
        if title_fragment in normalized:
            paper_mention_count += 1
    check("Report mentions at least 3 papers",
          paper_mention_count >= 3,
          f"Found {paper_mention_count} paper mentions")


def check_memory(agent_workspace):
    """Check that memory.json has research tracking entities."""
    print("\n=== Checking Memory ===")

    memory_path = os.path.join(agent_workspace, "memory", "memory.json")
    if not os.path.isfile(memory_path):
        check("memory.json exists", False, f"Not found: {memory_path}")
        return

    check("memory.json exists", True)

    with open(memory_path, "r") as f:
        content = f.read().strip()

    if not content or content in ("{}", '{"entities": [], "relations": []}'):
        check("Memory has content", False, "memory.json is empty or unchanged")
        return

    check("Memory has content", True)

    try:
        memory_data = json.loads(content)
    except json.JSONDecodeError:
        check("Memory is valid JSON", False, "Cannot parse memory.json")
        return

    check("Memory is valid JSON", True)

    # Check for entities
    entities = memory_data.get("entities", [])
    if isinstance(memory_data, list):
        entities = memory_data

    check("Memory has at least 3 entities",
          len(entities) >= 3,
          f"Found {len(entities)} entities")

    entity_text = ""
    for ent in entities:
        if isinstance(ent, dict):
            entity_text += json.dumps(ent).lower() + " "

    # Check for research tracking entity
    has_research = ("research" in entity_text or "ai_safety" in entity_text
                    or "ai safety" in entity_text)
    check("Memory has research tracking entity",
          has_research,
          "No research tracking entity found")

    # Check for paper entities
    paper_keywords = ["concrete problems", "debate", "mesa-optimization",
                      "learned optimization", "red teaming", "alignment",
                      "reward model"]
    kw_count = sum(1 for kw in paper_keywords if kw in entity_text)
    check("Memory has paper-related entities (at least 2 keywords)",
          kw_count >= 2,
          f"Found {kw_count} paper keywords in memory")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_memory(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY ===")
    print(f"Results: {PASS_COUNT}/{total} passed, {FAIL_COUNT} failed")

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()

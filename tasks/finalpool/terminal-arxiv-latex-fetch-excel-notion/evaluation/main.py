"""Evaluation script for terminal-arxiv-latex-fetch-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Conference_Prep_Tracker.xlsx")
    check("Conference_Prep_Tracker.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return

    wb = openpyxl.load_workbook(excel_path)

    # Paper_Sections sheet
    check("Paper_Sections sheet exists", "Paper_Sections" in wb.sheetnames)
    if "Paper_Sections" in wb.sheetnames:
        ws = wb["Paper_Sections"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Paper_Sections has >= 10 rows", len(data_rows) >= 10, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Paper_ID', 'Paper_Title', 'Section_Title', 'Section_Word_Count']:
            check(f"Paper_Sections has {col}", col.lower() in headers, f"headers: {headers[:5]}")
        # Check paper IDs present
        paper_ids = set(str(r[0]) for r in data_rows if r[0])
        check("Has scaling laws paper", any("2301" in pid for pid in paper_ids), f"IDs: {paper_ids}")
        check("Has RLHF paper", any("2203" in pid for pid in paper_ids), f"IDs: {paper_ids}")
        check("Has OPT paper", any("2205" in pid for pid in paper_ids), f"IDs: {paper_ids}")

    # Conference_Schedule sheet
    check("Conference_Schedule sheet exists", "Conference_Schedule" in wb.sheetnames)
    if "Conference_Schedule" in wb.sheetnames:
        ws = wb["Conference_Schedule"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Conference_Schedule has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Session_ID', 'Session_Title', 'Date', 'Related_Papers']:
            check(f"Conference_Schedule has {col}", col.lower() in headers, f"headers: {headers[:7]}")

    # Presentation_Notes sheet
    check("Presentation_Notes sheet exists", "Presentation_Notes" in wb.sheetnames)
    if "Presentation_Notes" in wb.sheetnames:
        ws = wb["Presentation_Notes"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Presentation_Notes has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Slide_Number', 'Topic', 'Key_Points', 'Source_Paper']:
            check(f"Presentation_Notes has {col}", col.lower() in headers, f"headers: {headers[:5]}")


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()

        found_page = False
        for page_id, props in pages:
            if isinstance(props, str):
                props = json.loads(props)
            if isinstance(props, dict):
                # Check if title contains "Conference Prep"
                title_prop = props.get("title", props.get("Title", {}))
                title_text = ""
                if isinstance(title_prop, dict):
                    title_items = title_prop.get("title", [])
                    if isinstance(title_items, list):
                        title_text = " ".join(t.get("plain_text", t.get("text", {}).get("content", "")) for t in title_items if isinstance(t, dict))
                    elif isinstance(title_items, str):
                        title_text = title_items
                elif isinstance(title_prop, str):
                    title_text = title_prop

                if "conference" in title_text.lower() and "prep" in title_text.lower():
                    found_page = True
                    check("Conference Prep page exists", True)

                    # Check for properties
                    has_conf_name = any("conference" in k.lower() and "name" in k.lower() for k in props.keys())
                    has_status = any("status" in k.lower() for k in props.keys())
                    has_paper_count = any("paper" in k.lower() and "count" in k.lower() for k in props.keys())
                    check("Page has Conference_Name property", has_conf_name, f"props: {list(props.keys())}")
                    check("Page has Status property", has_status, f"props: {list(props.keys())}")
                    check("Page has Paper_Count property", has_paper_count, f"props: {list(props.keys())}")
                    break

        if not found_page:
            check("Conference Prep page exists", False, f"Found {len(pages)} pages, none matching 'Conference Prep'")

        # Check for content blocks
        cur.execute("SELECT COUNT(*) FROM notion.blocks")
        block_count = cur.fetchone()[0]
        check("Notion has content blocks", block_count >= 3, f"found {block_count} blocks")

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion accessible", False, str(e))


def check_script(agent_workspace):
    print("\n=== Checking Terminal Script ===")
    check("conference_prep_builder.py exists",
          os.path.exists(os.path.join(agent_workspace, "conference_prep_builder.py")))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Conference_Prep_Tracker.xlsx")
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"paper", "section", "conference", "schedule", "presentation", "note"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Word counts should not be negative
        if "Paper_Sections" in wb.sheetnames:
            ws = wb["Paper_Sections"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            wc_idx = next((i for i, h in enumerate(headers) if "word_count" in h), None)
            if wc_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > wc_idx and row[wc_idx] is not None:
                        try:
                            wc = float(row[wc_idx])
                            if wc < 0:
                                check("No negative word counts", False, f"Found {wc}")
                                break
                        except (ValueError, TypeError):
                            pass
                else:
                    check("No negative word counts", True)

    # Notion: no duplicate Conference Prep pages
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()
        conf_pages = 0
        for page_id, props in pages:
            if isinstance(props, str):
                props = json.loads(props)
            if isinstance(props, dict):
                title_prop = props.get("title", props.get("Title", {}))
                title_text = ""
                if isinstance(title_prop, dict):
                    title_items = title_prop.get("title", [])
                    if isinstance(title_items, list):
                        title_text = " ".join(t.get("plain_text", t.get("text", {}).get("content", ""))
                                              for t in title_items if isinstance(t, dict))
                if "conference" in title_text.lower() and "prep" in title_text.lower():
                    conf_pages += 1
        check("No duplicate Conference Prep pages", conf_pages <= 1,
              f"Found {conf_pages} matching pages")
        cur.close()
        conn.close()
    except Exception:
        pass


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    check_excel(agent_workspace)
    check_notion()
    check_script(agent_workspace)
    check_reverse_validation(agent_workspace)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

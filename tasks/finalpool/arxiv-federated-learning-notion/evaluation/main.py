"""
Evaluation for arxiv-federated-learning-notion task.
Checks Notion page/database and Excel spreadsheet.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# FL papers in arxiv.papers (IDs from actual DB)
EXPECTED_FL_PAPERS = {
    "1602.05629": "Communication-Efficient Learning of Deep Networks from Decentralized Data",
    "1812.06127": "Federated Optimization in Heterogeneous Networks",
    "1908.07873": "Federated Learning: Challenges, Methods, and Future Directions",
    "1912.04977": "Advances and Open Problems in Federated Learning",
}

# These are non-FL papers that should NOT appear
NON_FL_IDS = {"1207.00580", "1502.03167"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=50):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_notion():
    """Check Notion page and database."""
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Check for page with "Federated Learning Research Hub" title
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()

        hub_page = None
        for pid, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "federated learning" in props_str and "research hub" in props_str:
                hub_page = pid
                break
            if "federated" in props_str and "hub" in props_str:
                hub_page = pid
                break

        if hub_page is None:
            # Broader search
            for pid, props in pages:
                props_str = json.dumps(props).lower() if props else ""
                if "federated" in props_str:
                    hub_page = pid
                    break

        record("Notion page 'Federated Learning Research Hub' exists",
               hub_page is not None,
               f"Found {len(pages)} pages, none matching")

        # Check for database "Paper Index"
        cur.execute("SELECT id, title, parent FROM notion.databases")
        dbs = cur.fetchall()

        paper_db = None
        for did, title_json, parent in dbs:
            title_str = json.dumps(title_json).lower() if title_json else ""
            if "paper index" in title_str or "paper_index" in title_str:
                paper_db = did
                break
            if "paper" in title_str and "index" in title_str:
                paper_db = did
                break

        record("Notion database 'Paper Index' exists",
               paper_db is not None,
               f"Found {len(dbs)} databases")

        # Check database has entries (pages that are children of database)
        if paper_db:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent::text LIKE %s
            """, (f"%{paper_db}%",))
            db_pages = cur.fetchall()
            record("Paper Index has entries", len(db_pages) >= 4,
                   f"Found {len(db_pages)} entries, expected >= 4")

        conn.close()
    except Exception as e:
        record("Notion connection", False, str(e))


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == c:
                return i
    return None


def check_excel(agent_workspace):
    """Check Excel spreadsheet."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Federated_Learning_Papers.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    record("Excel readable", True)

    # Check Paper Details sheet
    details_rows = load_sheet_rows(wb, "Paper Details")
    if details_rows is None:
        details_rows = load_sheet_rows(wb, "Paper_Details")
    if details_rows is None:
        record("Sheet 'Paper Details' exists", False, f"Available: {wb.sheetnames}")
        return

    record("Sheet 'Paper Details' exists", True)
    header = details_rows[0] if details_rows else []
    data_rows = details_rows[1:] if len(details_rows) > 1 else []

    # Should have at least 4 FL papers (might include Dropout/BatchNorm if agent interprets broadly)
    record("Paper Details has >= 4 rows", len(data_rows) >= 4,
           f"Found {len(data_rows)} data rows")

    title_col = find_col(header, ["Title", "title"])
    id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
    abstract_len_col = find_col(header, ["Abstract_Length", "abstract_length", "Abstract Length"])

    if title_col is not None:
        found_titles = []
        for row in data_rows:
            if title_col < len(row) and row[title_col]:
                found_titles.append(str(row[title_col]).strip().lower())

        for pid, expected_title in EXPECTED_FL_PAPERS.items():
            found = any(expected_title.lower() in t or t in expected_title.lower()
                        for t in found_titles)
            record(f"Has paper: {expected_title[:50]}...", found)

    # Check Abstract_Length column exists
    record("Abstract_Length column exists", abstract_len_col is not None,
           f"Header: {header}")

    # Check Summary sheet
    summary_rows = load_sheet_rows(wb, "Summary")
    if summary_rows is None:
        record("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
        return

    record("Sheet 'Summary' exists", True)

    metrics = {}
    for row in summary_rows:
        if row and row[0] is not None:
            key = str(row[0]).strip().lower().replace(" ", "_")
            val = row[1] if len(row) > 1 else None
            metrics[key] = val

    # Total_Papers
    total_key = None
    for k in metrics:
        if "total" in k and "paper" in k:
            total_key = k
            break
    if total_key:
        # Allow 4-6 depending on what agent finds
        val = metrics[total_key]
        ok = val is not None and 4 <= int(float(val)) <= 6
        record("Summary: Total_Papers between 4-6", ok, f"Got {val}")
    else:
        record("Summary: Total_Papers exists", False, f"Keys: {list(metrics.keys())}")

    # Avg_Abstract_Length
    avg_key = None
    for k in metrics:
        if "avg" in k and "abstract" in k:
            avg_key = k
            break
    if avg_key:
        val = metrics[avg_key]
        ok = val is not None and float(val) > 100
        record("Summary: Avg_Abstract_Length > 100", ok, f"Got {val}")
    else:
        record("Summary: Avg_Abstract_Length exists", False, f"Keys: {list(metrics.keys())}")

    # Earliest/Latest Year
    for label, expected in [("earliest", 2016), ("latest", 2019)]:
        year_key = None
        for k in metrics:
            if label in k and "year" in k:
                year_key = k
                break
        if year_key:
            ok = num_close(metrics[year_key], expected, tol=1)
            record(f"Summary: {label.title()}_Year ~ {expected}", ok,
                   f"Got {metrics[year_key]}")
        else:
            record(f"Summary: {label.title()}_Year exists", False,
                   f"Keys: {list(metrics.keys())}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_notion()
    check_excel(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()

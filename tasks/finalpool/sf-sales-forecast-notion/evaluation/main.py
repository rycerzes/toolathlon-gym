"""
Evaluation script for sf-sales-forecast-notion task.

Checks:
1. Notion page "Sales Performance Dashboard" exists
2. Notion databases "Monthly Revenue" and "Regional Performance" exist with data
3. Excel file Sales_Dashboard_Backup.xlsx with correct data
"""
import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT TO_CHAR("ORDER_DATE"::timestamp, 'YYYY-MM') as month,
               COUNT(*) as orders,
               ROUND(SUM("TOTAL_AMOUNT")::numeric, 2) as revenue
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        GROUP BY TO_CHAR("ORDER_DATE"::timestamp, 'YYYY-MM')
        ORDER BY month
    """)
    monthly = [(m, int(o), float(r)) for m, o, r in cur.fetchall() if int(o) >= 500]

    cur.execute("""
        SELECT c."REGION",
               COUNT(o."ORDER_ID"),
               ROUND(SUM(o."TOTAL_AMOUNT")::numeric, 2)
        FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
        JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
        GROUP BY c."REGION"
        ORDER BY c."REGION"
    """)
    regions = [(r, int(o), float(rev)) for r, o, rev in cur.fetchall()]

    cur.close()
    conn.close()
    return monthly, regions


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def number_in_text(value, text):
    text = str(text)
    val_str = str(value)
    if val_str in text:
        return True
    try:
        f2 = f"{float(value):.2f}"
        if f2 in text:
            return True
    except (ValueError, TypeError):
        pass
    return False


def check_notion():
    print("\n=== Checking Notion ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check dashboard page
    cur.execute("""
        SELECT id, properties FROM notion.pages
        WHERE archived = false
    """)
    pages = cur.fetchall()
    dashboard_found = False
    for pid, props in pages:
        props_str = json.dumps(props).lower() if props else ""
        if "sales" in props_str and "dashboard" in props_str:
            dashboard_found = True
            break
    check("Notion 'Sales Performance Dashboard' page exists", dashboard_found,
          f"Found {len(pages)} pages")

    # Check databases
    cur.execute("""
        SELECT id, title, properties FROM notion.databases
        WHERE archived = false
    """)
    dbs = cur.fetchall()

    monthly_db = None
    regional_db = None
    for did, title, props in dbs:
        title_str = json.dumps(title).lower() if title else ""
        if "monthly" in title_str and "revenue" in title_str:
            monthly_db = did
        if "regional" in title_str and "performance" in title_str:
            regional_db = did

    check("Notion 'Monthly Revenue' database exists", monthly_db is not None,
          f"Found databases: {[json.dumps(t) for _, t, _ in dbs]}")
    check("Notion 'Regional Performance' database exists", regional_db is not None,
          f"Found databases: {[json.dumps(t) for _, t, _ in dbs]}")

    expected_monthly, expected_regions = load_expected()

    # Check monthly database has entries
    if monthly_db:
        cur.execute("""
            SELECT properties FROM notion.pages
            WHERE parent::text LIKE %s AND archived = false
        """, (f'%{monthly_db}%',))
        monthly_pages = cur.fetchall()
        check(f"Monthly Revenue has entries (expected ~{len(expected_monthly)})",
              len(monthly_pages) >= len(expected_monthly) - 2,
              f"Found {len(monthly_pages)} entries")

        # Check some months exist in properties
        all_props_text = " ".join(json.dumps(p[0]) for p in monthly_pages if p[0])
        sample_months = [m[0] for m in expected_monthly[:3]]
        for month in sample_months:
            check(f"Monthly Revenue contains '{month}'",
                  month in all_props_text,
                  f"Not found in properties")

    # Check regional database has entries
    if regional_db:
        cur.execute("""
            SELECT properties FROM notion.pages
            WHERE parent::text LIKE %s AND archived = false
        """, (f'%{regional_db}%',))
        regional_pages = cur.fetchall()
        check(f"Regional Performance has {len(expected_regions)} entries",
              len(regional_pages) >= len(expected_regions),
              f"Found {len(regional_pages)} entries")

        all_props_text = " ".join(json.dumps(p[0]) for p in regional_pages if p[0])
        for region, _, _ in expected_regions:
            check(f"Regional Performance contains '{region}'",
                  region.lower() in all_props_text.lower(),
                  f"Not found in properties")

    cur.close()
    conn.close()


def check_excel(agent_workspace):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Sales_Dashboard_Backup.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    expected_monthly, expected_regions = load_expected()

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    def sheet_text(ws):
        txt = ""
        for row in ws.iter_rows(values_only=True):
            txt += " ".join(str(c) for c in row if c is not None) + " "
        return txt

    # Monthly Revenue sheet
    ws_m = find_sheet(["monthly"])
    if not ws_m:
        ws_m = find_sheet(["revenue"])
    check("Monthly Revenue sheet exists", ws_m is not None, f"Sheets: {wb.sheetnames}")
    if ws_m:
        txt = sheet_text(ws_m)
        check("Monthly sheet contains '2024-04'", "2024-04" in txt)
        check("Monthly sheet contains '2025-12'", "2025-12" in txt)
        # Check that partial month 2026-03 is excluded
        check("Monthly sheet excludes partial month 2026-03",
              "2026-03" not in txt or "14590" not in txt,
              "Partial month included")

    # Regional Performance sheet
    ws_r = find_sheet(["regional"])
    if not ws_r:
        ws_r = find_sheet(["region"])
    check("Regional Performance sheet exists", ws_r is not None, f"Sheets: {wb.sheetnames}")
    if ws_r:
        txt = sheet_text(ws_r)
        for region, orders, revenue in expected_regions:
            check(f"Regional sheet contains '{region}'",
                  region.lower() in txt.lower())
            check(f"Regional sheet contains revenue {revenue} for {region}",
                  number_in_text(revenue, txt))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_notion()
    check_excel(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_passed = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()

"""Evaluation for terminal-canvas-scholarly-pdf-excel-word."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Curriculum_Research_Alignment.xlsx ===")
    agent_file = os.path.join(agent_ws, "Curriculum_Research_Alignment.xlsx")
    gt_file = os.path.join(gt_dir, "Curriculum_Research_Alignment.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
        gwb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Course_Assignments
    print("  Checking Course_Assignments...")
    ws1 = get_sheet(awb, "Course_Assignments")
    gws1 = get_sheet(gwb, "Course_Assignments")
    check("Sheet Course_Assignments exists", ws1 is not None, f"Sheets: {awb.sheetnames}")
    if ws1 and gws1:
        a_rows = list(ws1.iter_rows(min_row=2, values_only=True))
        g_rows = list(gws1.iter_rows(min_row=2, values_only=True))
        # Query dynamic assignment count from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id IN (1, 2)")
            expected_assignment_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_assignment_count = 12
        check(f"Course_Assignments has {expected_assignment_count} rows",
              len(a_rows) == expected_assignment_count, f"Got {len(a_rows)}")

        # Build lookup by assignment name
        a_lookup = {}
        for r in a_rows:
            if r and len(r) >= 4 and r[2]:
                a_lookup[str(r[2]).strip().lower()] = r

        for g_row in g_rows:
            if not g_row or not g_row[2]:
                continue
            key = str(g_row[2]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Assignment '{g_row[2]}' exists", False, "Missing")
                continue
            # Check Course_ID
            check(f"'{g_row[2]}' Course_ID",
                  num_close(a_row[0], g_row[0], 0),
                  f"Expected {g_row[0]}, got {a_row[0]}")
            # Check Points
            check(f"'{g_row[2]}' Points",
                  num_close(a_row[3], g_row[3], 1),
                  f"Expected {g_row[3]}, got {a_row[3]}")

    # Sheet 2: Related_Papers
    print("  Checking Related_Papers...")
    ws2 = get_sheet(awb, "Related_Papers")
    check("Sheet Related_Papers exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2:
        a_rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in a_rows2 if r and r[0]]
        check("Related_Papers has >= 4 papers", len(data_rows) >= 4, f"Got {len(data_rows)}")

        # Check that relevance scores are in range
        for r in data_rows:
            if r and len(r) >= 4 and r[3] is not None:
                score = float(r[3])
                if score < 1 or score > 10:
                    check(f"Paper '{r[0]}' relevance in 1-10", False, f"Score: {score}")
                    break
        else:
            check("All relevance scores in 1-10 range", True)

    # Sheet 3: Alignment_Matrix
    print("  Checking Alignment_Matrix...")
    ws3 = get_sheet(awb, "Alignment_Matrix")
    check("Sheet Alignment_Matrix exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        a_rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        data_rows3 = [r for r in a_rows3 if r and r[0]]
        # Use same dynamic count from Canvas
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id IN (1, 2)")
            expected_align_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_align_count = 12
        check(f"Alignment_Matrix has {expected_align_count} rows",
              len(data_rows3) == expected_align_count, f"Got {len(data_rows3)}")

        # Check all have matched paper and score
        valid_rows = [r for r in data_rows3 if r and len(r) >= 3 and r[1] and r[2] is not None]
        check("All rows have matched paper and score",
              len(valid_rows) == len(data_rows3),
              f"Valid: {len(valid_rows)}, Total: {len(data_rows3)}")

        # Check scores in range
        for r in data_rows3:
            if r and len(r) >= 3 and r[2] is not None:
                score = float(r[2])
                if score < 0 or score > 100:
                    check(f"'{r[0]}' alignment score in 0-100", False, f"Score: {score}")
                    break
        else:
            check("All alignment scores in 0-100 range", True)

        # Check sorted by Alignment_Score descending
        scores = [float(r[2]) for r in data_rows3 if r and len(r) >= 3 and r[2] is not None]
        check("Alignment_Matrix sorted by score descending",
              scores == sorted(scores, reverse=True),
              f"First few scores: {scores[:5]}")


def check_word(agent_ws):
    print("\n=== Checking Curriculum_Review_Report.docx ===")
    docx_path = os.path.join(agent_ws, "Curriculum_Review_Report.docx")
    check("Word file exists", os.path.isfile(docx_path), docx_path)
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 500, f"Length: {len(text)}")
        check("Contains curriculum/alignment reference",
              "curriculum" in text or "alignment" in text or "review" in text,
              "Missing topic reference")
        check("Contains course reference",
              "analytics" in text or "algorithm" in text or "course" in text,
              "Missing course reference")
        check("Contains paper/research reference",
              "paper" in text or "research" in text or "literature" in text,
              "Missing research reference")
        check("Contains recommendation",
              "recommend" in text or "suggestion" in text or "improve" in text,
              "Missing recommendations")
    except ImportError:
        check("python-docx available", False, "Cannot verify Word content")
    except Exception as e:
        check("Word document readable", False, str(e))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Curriculum_Research_Alignment.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            # Check no unexpected sheets
            expected_sheets = {"course_assignments", "related_papers", "alignment_matrix"}
            actual_sheets = {s.strip().lower().replace(" ", "_") for s in wb.sheetnames}
            unexpected = actual_sheets - expected_sheets
            check("No unexpected sheets in Excel",
                  len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")

            # Check no negative points in Course_Assignments
            ws = get_sheet(wb, "Course_Assignments")
            if ws:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 3 and row[3] is not None:
                        try:
                            val = float(row[3])
                            if val < 0:
                                check("No negative points in Course_Assignments", False,
                                      f"Found negative points: {val}")
                                break
                        except (TypeError, ValueError):
                            pass
                else:
                    check("No negative points in Course_Assignments", True)
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()

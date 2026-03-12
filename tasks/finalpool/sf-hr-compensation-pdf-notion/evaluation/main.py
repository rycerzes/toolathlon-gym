"""Evaluation for sf-hr-compensation-pdf-notion.

Blocking checks: Compensation_Data.xlsx (Excel data comparison).
Non-blocking: Notion page, PDF existence.
"""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    all_errors = []

    # ---- Check Excel (blocking) ----
    agent_file = os.path.join(args.agent_workspace, "Compensation_Data.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Data.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Department Summary
    print("  Checking Department Summary...")
    a_rows = load_sheet_rows(agent_wb, "Department Summary")
    g_rows = load_sheet_rows(gt_wb, "Department Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Department Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Department Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing department: {g_row[0]}")
                continue
            # Col 1: Employees count
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 50):
                    all_errors.append(f"{key}.Employees: {a_row[1]} vs {g_row[1]} (tol=50)")
            # Col 2: Avg_Salary
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 500):
                    all_errors.append(f"{key}.Avg_Salary: {a_row[2]} vs {g_row[2]} (tol=500)")
            # Col 3: Min_Salary
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 2000):
                    all_errors.append(f"{key}.Min_Salary: {a_row[3]} vs {g_row[3]} (tol=2000)")
            # Col 4: Max_Salary
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 50000):
                    all_errors.append(f"{key}.Max_Salary: {a_row[4]} vs {g_row[4]} (tol=50000)")
        if not all_errors:
            print("    PASS")

    # Check Education Breakdown
    print("  Checking Education Breakdown...")
    a_rows = load_sheet_rows(agent_wb, "Education Breakdown")
    g_rows = load_sheet_rows(gt_wb, "Education Breakdown")
    prev_errors = len(all_errors)
    if a_rows is None:
        all_errors.append("Sheet 'Education Breakdown' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Education Breakdown' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None and row[1] is not None:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[key] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing edu row: {g_row[0]} / {g_row[1]}")
                continue
            # Col 2: Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 50):
                    all_errors.append(f"{key}.Count: {a_row[2]} vs {g_row[2]} (tol=50)")
            # Col 3: Avg_Salary
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1000):
                    all_errors.append(f"{key}.Avg_Salary: {a_row[3]} vs {g_row[3]} (tol=1000)")

        new_errors = len(all_errors) - prev_errors
        if new_errors == 0:
            print("    PASS")

    # ---- Check PDF exists (blocking but lenient) ----
    print("  Checking PDF...")
    pdf_path = os.path.join(args.agent_workspace, "Compensation_Report.pdf")
    if not os.path.exists(pdf_path):
        all_errors.append("Compensation_Report.pdf not found")
    else:
        file_size = os.path.getsize(pdf_path)
        if file_size < 500:
            all_errors.append(f"Compensation_Report.pdf too small ({file_size} bytes)")
        else:
            print("    PASS")

    # ---- Non-blocking Notion check ----
    print("  Non-blocking: Notion DB check...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM notion.pages")
        count = cur.fetchone()[0]
        print(f"    [INFO] Found {count} Notion page(s) (non-blocking)")
        cur.close()
        conn.close()
    except Exception as e:
        print(f"    [INFO] Notion check skipped: {e} (non-blocking)")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:15]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()

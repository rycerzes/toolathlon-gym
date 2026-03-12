"""
Evaluation for yt-fireship-tech-report-excel-notion task.

Checks:
1. Tech_Trend_Report.xlsx exists with "Videos" sheet having >= 8 data rows
2. "Videos" sheet has Video_ID, Title, View_Count columns (case-insensitive)
3. "Topic_Summary" sheet exists with >= 3 topic rows
4. Topic_Summary has Topic and Video_Count columns
5. Notion page exists with title containing "Fireship" or "Tech Trends"
6. GCal has a new event in March 2026 with "Tech Review" in summary (not the noise Team Sync)
7. Email was sent to techteam@company.com
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-4: Tech_Trend_Report.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Tech_Trend_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Trend_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Videos sheet has >= 8 data rows", False, "File missing")
        record("Videos sheet has required columns", False, "File missing")
        record("Topic_Summary sheet exists with >= 3 rows", False, "File missing")
        record("Topic_Summary has Topic and Video_Count columns", False, "File missing")
        return
    record("Tech_Trend_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Check Videos sheet
    videos_key = None
    for k in sheet_names_lower:
        if "video" in k:
            videos_key = sheet_names_lower[k]
            break
    if not videos_key:
        record("Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Videos sheet has >= 8 data rows", False, "Sheet missing")
        record("Videos sheet has required columns", False, "Sheet missing")
    else:
        ws = wb[videos_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        record("Videos sheet has >= 8 data rows", len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            has_video_id = any("video_id" in h or "videoid" in h or "video id" in h for h in headers)
            has_title = any("title" in h for h in headers)
            has_views = any("view" in h for h in headers)
            record("Videos sheet has required columns (Video_ID, Title, View_Count)",
                   has_video_id and has_title and has_views,
                   f"Headers: {rows[0]}")
        else:
            record("Videos sheet has required columns", False, "Sheet is empty")

    # Check Topic_Summary sheet
    topic_key = None
    for k in sheet_names_lower:
        if "topic" in k or "summary" in k:
            topic_key = sheet_names_lower[k]
            break
    if not topic_key:
        record("Topic_Summary sheet exists with >= 3 rows", False, f"Sheets: {wb.sheetnames}")
        record("Topic_Summary has Topic and Video_Count columns", False, "Sheet missing")
    else:
        ws2 = wb[topic_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        record("Topic_Summary sheet exists with >= 3 rows", len(data_rows2) >= 3,
               f"Found {len(data_rows2)} data rows")
        if rows2:
            headers2 = [str(c).strip().lower() if c else "" for c in rows2[0]]
            has_topic = any("topic" in h for h in headers2)
            has_count = any("count" in h or "video" in h for h in headers2)
            record("Topic_Summary has Topic and Video_Count columns",
                   has_topic and has_count,
                   f"Headers: {rows2[0]}")
        else:
            record("Topic_Summary has Topic and Video_Count columns", False, "Sheet is empty")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Tech_Trend_Report.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 5: Notion page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE properties->>'title' ILIKE '%Fireship%'
                   OR properties->>'title' ILIKE '%Tech Trends%'
                   OR properties->>'title' ILIKE '%Tech_Trends%'
            """)
            rows = cur.fetchall()
        conn.close()
        record("Notion page with 'Fireship' or 'Tech Trends' in title exists",
               len(rows) > 0,
               f"Found {len(rows)} matching pages")
    except Exception as e:
        record("Notion page check", False, str(e))


def check_gcal():
    print("\n=== Check 6: GCal Tech Review event ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, summary, start_datetime FROM gcal.events
                WHERE summary ILIKE '%Tech Review%'
                  AND start_datetime >= '2026-03-01'
                  AND start_datetime < '2026-04-01'
            """)
            rows = cur.fetchall()
        conn.close()
        record("GCal has 'Tech Review' event in March 2026",
               len(rows) > 0,
               f"Found {len(rows)} matching events")
    except Exception as e:
        record("GCal check", False, str(e))


def check_email():
    print("\n=== Check 7: Email sent to techteam@company.com ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            # Check sent folder (folder_id=2) or any message to techteam
            cur.execute("""
                SELECT id, subject, to_addr FROM email.messages
                WHERE to_addr::text ILIKE '%techteam@company.com%'
            """)
            rows = cur.fetchall()
            if not rows:
                # Also check sent_log
                try:
                    cur.execute("""
                        SELECT id FROM email.sent_log
                        WHERE to_addr ILIKE '%techteam@company.com%'
                    """)
                    rows = cur.fetchall()
                except Exception:
                    pass
        conn.close()
        record("Email sent to techteam@company.com",
               len(rows) > 0,
               f"Found {len(rows)} matching emails")
    except Exception as e:
        record("Email check", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-fireship-tech-report-excel-notion")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_notion()
    check_gcal()
    check_email()

    all_passed = FAIL_COUNT == 0
    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print(f"\n{'='*40}")
    print(f"Result: {'PASS' if all_passed else 'FAIL'} - {summary}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "all_passed": all_passed}, f)

    return all_passed, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

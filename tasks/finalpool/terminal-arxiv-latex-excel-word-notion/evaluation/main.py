"""Evaluation for terminal-arxiv-latex-excel-word-notion.

Checks:
1. Literature_Review_Matrix.xlsx with 3 sheets
2. Literature_Review_Draft.docx
3. Notion database "Transformer Research Papers"
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Literature_Review_Matrix.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Literature_Review_Matrix.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Literature_Review_Matrix.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Paper_Catalog
    print("  Checking Paper_Catalog...")
    a_sheet = get_sheet(agent_wb, "Paper_Catalog")
    g_sheet = get_sheet(gt_wb, "Paper_Catalog")
    check("Sheet 'Paper_Catalog' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
        check("Paper_Catalog has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        # Check that key papers are present by ArXiv_ID
        a_ids = {str(r[0]).strip() for r in a_rows if r and r[0]}
        expected_ids = {"1706.03762", "1810.04805", "2005.14165", "2010.11929", "2301.07041"}
        for eid in expected_ids:
            check(f"Paper '{eid}' in catalog", eid in a_ids,
                  f"Missing from {a_ids}")

        # Verify a couple of titles
        a_lookup = {str(r[0]).strip(): r for r in a_rows if r and r[0]}
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip()
            a_row = a_lookup.get(key)
            if a_row and len(a_row) > 3 and len(g_row) > 3:
                check(f"'{key}' Year",
                      num_close(a_row[3], g_row[3], 0),
                      f"Expected {g_row[3]}, got {a_row[3]}")

    # Methodology_Comparison
    print("  Checking Methodology_Comparison...")
    a_sheet = get_sheet(agent_wb, "Methodology_Comparison")
    check("Sheet 'Methodology_Comparison' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        check("Methodology_Comparison has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        a_ids = {str(r[0]).strip() for r in a_rows if r and r[0]}
        for eid in expected_ids:
            check(f"Method for '{eid}'", eid in a_ids, f"Missing from {a_ids}")

        # Check each row has method_name and approach
        for r in a_rows:
            if r and r[0] and str(r[0]).strip() in expected_ids:
                has_method = r[1] is not None and len(str(r[1]).strip()) > 0
                has_approach = len(r) > 2 and r[2] is not None and len(str(r[2]).strip()) > 0
                check(f"'{str(r[0]).strip()}' has method and approach",
                      has_method and has_approach,
                      f"method={r[1]}, approach={r[2] if len(r) > 2 else None}")

    # Citation_Network
    print("  Checking Citation_Network...")
    a_sheet = get_sheet(agent_wb, "Citation_Network")
    check("Sheet 'Citation_Network' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        check("Citation_Network has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        # Check key citation: 1810.04805 cites 1706.03762
        citations = {(str(r[0]).strip(), str(r[1]).strip()) for r in a_rows if r and r[0] and r[1]}
        check("BERT cites Transformer",
              ("1810.04805", "1706.03762") in citations,
              f"Not found in {list(citations)[:5]}")
        check("GPT-3 cites Transformer",
              ("2005.14165", "1706.03762") in citations,
              f"Not found")


def check_word(agent_workspace):
    print("\n=== Checking Literature_Review_Draft.docx ===")
    docx_path = os.path.join(agent_workspace, "Literature_Review_Draft.docx")
    check("Literature_Review_Draft.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 300, f"Length: {len(text)}")
        check("Contains transformer reference",
              "transformer" in text)
        check("Contains attention reference",
              "attention" in text or "self-attention" in text)
        check("Contains BERT or GPT reference",
              "bert" in text or "gpt" in text)
        check("Contains methodology comparison",
              "method" in text or "approach" in text or "architecture" in text)
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        found_db = None
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "transformer" in title_str and "research" in title_str:
                found_db = db_id
                break
            if "transformer" in title_str and "paper" in title_str:
                found_db = db_id
                break
        check("Notion database 'Transformer Research Papers' exists",
              found_db is not None,
              f"Found {len(dbs)} databases, none matching")

        if found_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent::text LIKE %s AND archived = false AND in_trash = false
            """, (f'%{found_db}%',))
            page_count = cur.fetchone()[0]
            check("Notion database has >= 5 paper entries",
                  page_count >= 5,
                  f"Found {page_count} entries")

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check no unexpected sheets in the Excel file
    excel_path = os.path.join(workspace, "Literature_Review_Matrix.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            expected_sheets = {"paper_catalog", "methodology_comparison", "citation_network"}
            actual_sheets = {s.strip().lower().replace(" ", "_") for s in wb.sheetnames}
            unexpected = actual_sheets - expected_sheets
            check("No unexpected sheets in Excel",
                  len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")

            # Check no duplicate Notion pages
            try:
                conn = psycopg2.connect(**DB)
                cur = conn.cursor()
                cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
                dbs = cur.fetchall()
                found_db = None
                for db_id, title in dbs:
                    title_str = json.dumps(title).lower() if title else ""
                    if "transformer" in title_str:
                        found_db = db_id
                        break
                if found_db:
                    cur.execute("""
                        SELECT properties::text FROM notion.pages
                        WHERE parent::text LIKE %s AND archived = false AND in_trash = false
                    """, (f'%{found_db}%',))
                    pages = cur.fetchall()
                    page_texts = [str(p[0]) for p in pages]
                    if len(page_texts) > len(set(page_texts)):
                        check("No duplicate pages in Notion database", False,
                              f"{len(page_texts)} pages but {len(set(page_texts))} unique")
                    else:
                        check("No duplicate pages in Notion database", True)
                cur.close(); conn.close()
            except Exception:
                pass
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
